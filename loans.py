"""
loans.py  v1.1

Loans & Borrowings Monitor (Per Third Party Enhancement)

New in v1.1:
  * Added Third Party filter (pulls from third_parties.name).
  * Aggregation now keyed by (code, third_party) so balances are per third party + code.
  * Added 'Third Party' column (detailed mode) placed AFTER 'Description' as requested.
  * Context menu 'Turn to Donation' now converts only the rows for the selected (code, third_party) pair.
  * Excel export updated to include Third Party column in detailed mode.
  * Filter summary includes Third Party.
  * If multiple third parties exist for same code, simple mode will show several rows (same code) with different third-party groupings (third party not shown in simple mode to match original request).
    NOTE: This can look like duplicates; consider adding Third Party to simple mode later if needed.

Carry‑over (v1.0):
  * Tracks Quantity Given (Loan + Return of Borrowing OUT) and Quantity Received (In Borrowing + In Return of Loan IN).
  * Balance & Status (to receive / to give / settled).
  * Turn to Donation converts involved loan / borrowing transaction types to donation ones and appends a tag to document numbers.
  * Multi-format date parsing & optional tkcalendar.
  * Simple / Detailed mode toggle, scrolling, export, keyboard shortcuts.

"""

import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import re
from datetime import date, datetime
from calendar import monthrange
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Optional calendar
try:
    from tkcalendar import DateEntry
    TKCAL_AVAILABLE = True
except Exception:
    TKCAL_AVAILABLE = False

from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang
from popup_utils import custom_popup, custom_askyesno

# ---------------- Theme ----------------
BG_MAIN        = "#F0F4F8"
BG_PANEL       = "#FFFFFF"
COLOR_PRIMARY  = "#2C3E50"
COLOR_BORDER   = "#D0D7DE"
ROW_ALT_COLOR  = "#F7FAFC"
ROW_NORM_COLOR = "#FFFFFF"
BTN_EXPORT     = "#2980B9"
BTN_REFRESH    = "#2563EB"
BTN_CLEAR      = "#7F8C8D"
BTN_TOGGLE     = "#8E44AD"

KIT_FILL_COLOR    = "228B22"
MODULE_FILL_COLOR = "ADD8E6"

OUT_TYPES_GIVEN = {"Loan","Return of Borrowing"}
IN_TYPES_RECEIVED = {"In Borrowing","In Return of Loan"}

DONATION_OUT = "Out Donation"
DONATION_IN  = "In Donation"
DONATION_TAG = " (turned to donation)"

# ---------------- Date Parsing ----------------
DATE_REGEXES = [
    ("%Y-%m-%d", re.compile(r"^\d{4}-\d{2}-\d{2}$")),
    ("%d/%m/%Y", re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")),
    ("%d-%m-%Y", re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$")),
    ("%Y/%m/%d", re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")),
    ("%d %b %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$")),
    ("%d %B %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]+\s+\d{4}$")),
]
MONTH_ONLY_REGEXES = [
    ("%Y-%m", re.compile(r"^\d{4}-\d{1,2}$")),
    ("%m/%Y", re.compile(r"^\d{1,2}/\d{4}$")),
    ("%b-%Y", re.compile(r"^[A-Za-z]{3}-\d{4}$")),
    ("%B-%Y", re.compile(r"^[A-Za-z]+-\d{4}$")),
]

def parse_user_date(text: str, bound: str):
    if not text:
        return None
    raw = text.strip()
    for fmt, rx in DATE_REGEXES:
        if rx.match(raw):
            try: return datetime.strptime(raw, fmt).date()
            except: pass
    for fmt, rx in MONTH_ONLY_REGEXES:
        if rx.match(raw):
            try:
                tmp = datetime.strptime(raw, fmt)
                y, m = tmp.year, tmp.month
                return date(y, m, 1) if bound=="from" else date(y, m, monthrange(y,m)[1])
            except: pass
    if re.match(r"^\d{4}$", raw):
        y = int(raw)
        return date(y,1,1) if bound=="from" else date(y,12,31)
    return None

# ---------------- DB Helpers ----------------
def fetch_scenarios():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM scenarios ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except:
        return []
    finally:
        cur.close(); conn.close()

def fetch_kit_numbers():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT Kit FROM stock_transactions
            WHERE Kit IS NOT NULL AND Kit!='None' AND Kit!=''
            ORDER BY Kit
        """)
        return [r[0] for r in cur.fetchall()]
    except:
        return []
    finally:
        cur.close(); conn.close()

def fetch_module_numbers():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT Module FROM stock_transactions
            WHERE Module IS NOT NULL AND Module!='None' AND Module!=''
            ORDER BY Module
        """)
        return [r[0] for r in cur.fetchall()]
    except:
        return []
    finally:
        cur.close(); conn.close()

def fetch_third_parties():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM third_parties ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except:
        return []
    finally:
        cur.close(); conn.close()

# ---------------- Aggregation (per code + third_party) ----------------
def aggregate_loans(filters):
    scenario = filters.get("scenario")
    kit_number = filters.get("kit")
    module_number = filters.get("module")
    type_filter = filters.get("type")
    item_search = filters.get("item_search")
    doc_search = filters.get("doc_number")
    third_party_filter = filters.get("third_party")
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")

    conn = connect_db()
    if conn is None:
        return []

    cur = conn.cursor()
    where = []
    params = []

    if scenario and scenario.lower() != "all":
        where.append("Scenario = ?"); params.append(scenario)
    if kit_number and kit_number.lower() != "all":
        where.append("Kit = ?"); params.append(kit_number)
    if module_number and module_number.lower() != "all":
        where.append("Module = ?"); params.append(module_number)
    if third_party_filter and third_party_filter.lower() != "all":
        where.append("Third_Party = ?"); params.append(third_party_filter)
    if doc_search:
        where.append("document_number LIKE ?"); params.append(f"%{doc_search}%")
    if date_from:
        where.append("Date >= ?"); params.append(date_from.strftime("%Y-%m-%d"))
    if date_to:
        where.append("Date <= ?"); params.append(date_to.strftime("%Y-%m-%d"))

    loan_clause = """
       (Out_Type IN ('Loan','Return of Borrowing')
        OR IN_Type IN ('In Borrowing','In Return of Loan'))
    """
    base_where = " AND ".join(where) if where else "1=1"

    sql = f"""
      SELECT code, Scenario, Kit, Module,
             Qty_IN, IN_Type, Qty_Out, Out_Type,
             document_number, Third_Party
      FROM stock_transactions
      WHERE {base_where} AND {loan_clause}
    """
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()
    except:
        cur.close(); conn.close()
        return []

    grouped = {}
    scen_map = defaultdict(set)
    kit_map = defaultdict(set)
    mod_map = defaultdict(set)
    doc_map = defaultdict(set)

    for (code, scen, kit, module,
         qty_in, in_type, qty_out, out_type,
         doc, third_party) in rows:
        if not code:
            continue
        desc = get_item_description(code)
        dtype = detect_type(code, desc)

        if type_filter and type_filter.lower() != "all":
            if dtype.lower() != type_filter.lower():
                continue
        if item_search:
            if dtype.lower() != "item":
                continue
            if (item_search.lower() not in code.lower()
                and item_search.lower() not in desc.lower()):
                continue

        tp = third_party or ""  # normalize None -> ""
        key = (code, tp)
        rec = grouped.setdefault(key, {
            "code": code,
            "third_party": tp,
            "description": desc,
            "type": dtype,
            "qty_given": 0,
            "qty_received": 0
        })

        if out_type in OUT_TYPES_GIVEN and qty_out:
            try: rec["qty_given"] += int(qty_out)
            except: pass
        if in_type in IN_TYPES_RECEIVED and qty_in:
            try: rec["qty_received"] += int(qty_in)
            except: pass

        if scen: scen_map[key].add(scen)
        if kit: kit_map[key].add(kit)
        if module: mod_map[key].add(module)
        if doc: doc_map[key].add(doc)

    result = []
    for key, rec in grouped.items():
        balance = rec["qty_given"] - rec["qty_received"]
        if balance > 0:
            status = f"{balance} " + lang.t("loans.status_to_receive","to receive")
        elif balance < 0:
            status = f"{abs(balance)} " + lang.t("loans.status_to_give","to give")
        else:
            status = lang.t("loans.status_settled","Settled")
        rec["balance"] = balance
        rec["status"] = status
        rec["scenarios"] = ", ".join(sorted(scen_map[key])) if scen_map[key] else ""
        rec["kits"] = ", ".join(sorted(kit_map[key])) if kit_map[key] else ""
        rec["modules"] = ", ".join(sorted(mod_map[key])) if mod_map[key] else ""
        rec["documents"] = ", ".join(sorted(doc_map[key])) if doc_map[key] else ""
        result.append(rec)

    result.sort(key=lambda r: (r["type"], r["code"], r["third_party"]))
    cur.close(); conn.close()
    return result

# ---------------- UI Class ----------------
class Loans(tk.Frame):
    SIMPLE_COLS = ["code","description","qty_given","qty_received","balance","status"]
    DETAIL_COLS = [
        "scenarios","kits","modules","type","code","description",
        "third_party",  # inserted after description (as per request)
        "qty_given","qty_received","balance","status","documents"
    ]

    def __init__(self, parent, app, *args, **kwargs):
        super().__init__(parent, bg=BG_MAIN, *args, **kwargs)
        self.app = app
        self.rows = []
        self.simple_mode = False

        # Filters
        self.scenario_var = tk.StringVar(value="All")
        self.kit_var = tk.StringVar(value="All")
        self.module_var = tk.StringVar(value="All")
        self.type_var = tk.StringVar(value="All")
        self.third_party_var = tk.StringVar(value="All")
        self.item_search_var = tk.StringVar()
        self.doc_var = tk.StringVar()
        self.from_var = tk.StringVar()
        self.to_var = tk.StringVar()

        self.status_var = tk.StringVar(value=lang.t("loans.ready","Ready"))
        self.tree = None

        self._build_ui()
        self.populate_filters()
        self._set_default_dates()
        self.refresh()

    # ---------- UI ----------
    def _build_ui(self):
        header = tk.Frame(self, bg=BG_MAIN)
        header.pack(fill="x", padx=12, pady=(12,6))
        tk.Label(header, text=lang.t("menu.reports.loans","Loans"),
                 font=("Helvetica",20,"bold"),
                 bg=BG_MAIN, fg=COLOR_PRIMARY).pack(side="left")
        self.toggle_btn = tk.Button(header,
                                    text=lang.t("loans.detailed","Detailed"),
                                    bg=BTN_TOGGLE, fg="#FFFFFF", relief="flat",
                                    padx=14, pady=6, command=self.toggle_mode)
        self.toggle_btn.pack(side="right", padx=(6,0))
        tk.Button(header, text=lang.t("generic.export","Export"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.export_excel)\
            .pack(side="right", padx=(6,0))
        tk.Button(header, text=lang.t("generic.clear","Clear"),
                  bg=BTN_CLEAR, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.clear_filters)\
            .pack(side="right", padx=(6,0))
        tk.Button(header, text=lang.t("generic.refresh","Refresh"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.refresh)\
            .pack(side="right", padx=(6,0))

        filters = tk.Frame(self, bg=BG_MAIN)
        filters.pack(fill="x", padx=12, pady=(0,8))

        # Row 1
        r1 = tk.Frame(filters, bg=BG_MAIN); r1.pack(fill="x", pady=2)
        tk.Label(r1, text=lang.t("generic.scenario","Scenario"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.scenario_cb = ttk.Combobox(r1, textvariable=self.scenario_var, state="readonly", width=20)
        self.scenario_cb.grid(row=0, column=1, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.kit_number","Kit Number"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.kit_cb = ttk.Combobox(r1, textvariable=self.kit_var, state="readonly", width=16)
        self.kit_cb.grid(row=0, column=3, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.module_number","Module Number"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.module_cb = ttk.Combobox(r1, textvariable=self.module_var, state="readonly", width=16)
        self.module_cb.grid(row=0, column=5, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.type","Type"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.type_cb = ttk.Combobox(r1, textvariable=self.type_var, state="readonly",
                                    width=12, values=["All","Kit","Module","Item"])
        self.type_cb.grid(row=0, column=7, padx=(0,12))

        tk.Label(r1, text=lang.t("loans.third_party","Third Party"), bg=BG_MAIN)\
            .grid(row=0, column=8, sticky="w", padx=(0,4))
        self.third_party_cb = ttk.Combobox(r1, textvariable=self.third_party_var,
                                           state="readonly", width=18)
        self.third_party_cb.grid(row=0, column=9, padx=(0,4))

        # Row 2
        r2 = tk.Frame(filters, bg=BG_MAIN); r2.pack(fill="x", pady=2)
        tk.Label(r2, text=lang.t("loans.item_search","Item Search"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.item_entry = tk.Entry(r2, textvariable=self.item_search_var, width=22)
        self.item_entry.grid(row=0, column=1, padx=(0,12))
        self.item_entry.bind("<Return>", lambda e: self.refresh())
        self.item_entry.bind("<Escape>", lambda e: self._clear_field(self.item_search_var))

        tk.Label(r2, text=lang.t("generic.document_number","Document Number"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.doc_entry = tk.Entry(r2, textvariable=self.doc_var, width=22)
        self.doc_entry.grid(row=0, column=3, padx=(0,12))
        self.doc_entry.bind("<Return>", lambda e: self.refresh())
        self.doc_entry.bind("<Escape>", lambda e: self._clear_field(self.doc_var))

        tk.Label(r2, text=lang.t("generic.from_date","From Date"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.from_entry = self._date_widget(r2, self.from_var)
        self.from_entry.grid(row=0, column=5, padx=(0,12))

        tk.Label(r2, text=lang.t("generic.to_date","To Date"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.to_entry = self._date_widget(r2, self.to_var)
        self.to_entry.grid(row=0, column=7, padx=(0,12))

        # Status
        tk.Label(self, textvariable=self.status_var, anchor="w",
                 bg=BG_MAIN, fg=COLOR_PRIMARY, relief="sunken")\
            .pack(fill="x", padx=12, pady=(0,8))

        # Table
        table_frame = tk.Frame(self, bg=COLOR_BORDER, bd=1, relief="solid")
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self.tree = ttk.Treeview(table_frame, columns=(), show="headings", height=20)
        self.tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        hsb.pack(fill="x", padx=12, pady=(0,12))

        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        style = ttk.Style()
        try: style.theme_use("clam")
        except: pass
        style.configure("Treeview",
                        background=BG_PANEL,
                        fieldbackground=BG_PANEL,
                        foreground=COLOR_PRIMARY,
                        rowheight=24,
                        font=("Helvetica",10))
        style.configure("Treeview.Heading",
                        background="#E5E8EB", foreground=COLOR_PRIMARY,
                        font=("Helvetica",11,"bold"))
        self.tree.tag_configure("alt", background=ROW_ALT_COLOR)
        self.tree.tag_configure("kitrow", background="#228B22", foreground="#FFFFFF")
        self.tree.tag_configure("modrow", background="#ADD8E6")

        # Context menu / key
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Return>", lambda e: self._turn_to_donation_prompt() if self.tree.selection() else None)

        # ESC global (clear all)
        self.bind_all("<Escape>", self._esc_global_handler)

    def _esc_global_handler(self, event):
        # If inside entry, let field-level handle
        if isinstance(event.widget, tk.Entry):
            return
        self.clear_filters()

    def _date_widget(self, parent, var):
        if TKCAL_AVAILABLE:
            w = DateEntry(parent, textvariable=var, width=12,
                          date_pattern="yyyy-mm-dd",
                          showweeknumbers=False,
                          background="#2563EB", foreground="white", borderwidth=1)
            w.bind("<Return>", lambda e: self.refresh())
            w.bind("<Escape>", lambda e: self._clear_field(var))
            return w
        else:
            e = tk.Entry(parent, textvariable=var, width=12)
            e.bind("<Return>", lambda ev: self.refresh())
            e.bind("<Escape>", lambda ev: self._clear_field(var))
            return e

    def _clear_field(self, tk_var):
        tk_var.set("")
        self.refresh()

    def _set_default_dates(self):
        today = date.today()
        try:
            past = date(today.year - 1, today.month, today.day)
        except:
            # leap day safeguard
            past = today
        self.from_var.set(past.strftime("%Y-%m-%d"))
        self.to_var.set(today.strftime("%Y-%m-%d"))

    # ---------- Filters ----------
    def populate_filters(self):
        scenarios = fetch_scenarios()
        kits = fetch_kit_numbers()
        modules = fetch_module_numbers()
        tps = fetch_third_parties()

        self.scenario_cb['values'] = ["All"] + scenarios
        self.kit_cb['values'] = ["All"] + kits
        self.module_cb['values'] = ["All"] + modules
        self.third_party_cb['values'] = ["All"] + tps

        for var, cb in [
            (self.scenario_var, self.scenario_cb),
            (self.kit_var, self.kit_cb),
            (self.module_var, self.module_cb),
            (self.third_party_var, self.third_party_cb)
        ]:
            if var.get() not in cb['values']:
                var.set("All")

    # ---------- Mode ----------
    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        self.toggle_btn.config(text=lang.t("loans.simple","Simple") if self.simple_mode
                               else lang.t("loans.detailed","Detailed"))
        self._populate_tree()

    # ---------- Refresh ----------
    def refresh(self):
        filters = {
            "scenario": self.scenario_var.get(),
            "kit": self.kit_var.get(),
            "module": self.module_var.get(),
            "type": self.type_var.get(),
            "third_party": self.third_party_var.get(),
            "item_search": self.item_search_var.get().strip(),
            "doc_number": self.doc_var.get().strip(),
            "date_from": parse_user_date(self.from_var.get().strip(), "from") if self.from_var.get().strip() else None,
            "date_to": parse_user_date(self.to_var.get().strip(), "to") if self.to_var.get().strip() else None
        }
        if filters["date_to"] and filters["date_to"] > date.today():
            filters["date_to"] = date.today()
        self.rows = aggregate_loans(filters)
        self._populate_tree()
        self.status_var.set(lang.t("loans.loaded","Loaded {n} rows").format(n=len(self.rows)))

    def clear_filters(self):
        self.scenario_var.set("All")
        self.kit_var.set("All")
        self.module_var.set("All")
        self.type_var.set("All")
        self.third_party_var.set("All")
        self.item_search_var.set("")
        self.doc_var.set("")
        self._set_default_dates()
        self.refresh()

    # ---------- Columns ----------
    def _current_columns(self):
        return self.SIMPLE_COLS if self.simple_mode else self.DETAIL_COLS

    def _populate_tree(self):
        cols = self._current_columns()
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols

        headings = {
            "scenarios": lang.t("generic.scenario","Scenario(s)"),
            "kits": lang.t("generic.kit_number","Kit Number(s)"),
            "modules": lang.t("generic.module_number","Module Number(s)"),
            "type": lang.t("generic.type","Type"),
            "code": lang.t("generic.code","Code"),
            "description": lang.t("generic.description","Description"),
            "third_party": lang.t("loans.third_party","Third Party"),
            "qty_given": lang.t("loans.qty_given","Quantity Given"),
            "qty_received": lang.t("loans.qty_received","Quantity Received"),
            "balance": lang.t("loans.balance","Balance"),
            "status": lang.t("loans.status","Status"),
            "documents": lang.t("generic.document_number","Document Number(s)")
        }

        for c in cols:
            width = 120
            if c == "description": width = 340
            elif c == "code": width = 160
            elif c in ("scenarios","kits","modules","documents"): width = 200
            elif c == "third_party": width = 180
            elif c in ("qty_given","qty_received","balance"): width = 130
            elif c == "status": width = 140
            self.tree.heading(c, text=headings.get(c, c))
            self.tree.column(c, width=width, anchor="w", stretch=True)

        for idx, r in enumerate(self.rows):
            if self.simple_mode:
                vals = [
                    r["code"], r["description"],
                    r["qty_given"], r["qty_received"],
                    r["balance"], r["status"]
                ]
            else:
                vals = [
                    r["scenarios"], r["kits"], r["modules"], r["type"],
                    r["code"], r["description"], r["third_party"],
                    r["qty_given"], r["qty_received"],
                    r["balance"], r["status"], r["documents"]
                ]
            tag = "alt" if idx % 2 else ""
            if r["type"].upper() == "KIT": tag = "kitrow"
            elif r["type"].upper() == "MODULE": tag = "modrow"
            self.tree.insert("", "end", values=vals, tags=(tag,))

    # ---------- Context Menu ----------
    def _show_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid: return
        self.tree.selection_set(iid)
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(label=lang.t("loans.turn_to_donation","Turn to Donation"),
                         command=self._turn_to_donation_prompt)
        menu.post(event.x_root, event.y_root)

    def _selected_code_and_third_party(self):
        sel = self.tree.selection()
        if not sel:
            return None, None
        vals = self.tree.item(sel[0],"values")
        if not vals:
            return None, None
        if self.simple_mode:
            # simple mode has no third_party column
            code_idx = self._current_columns().index("code")
            return vals[code_idx], None
        else:
            code_idx = self._current_columns().index("code")
            tp_idx   = self._current_columns().index("third_party")
            return vals[code_idx], vals[tp_idx]

    def _turn_to_donation_prompt(self):
        code, tp = self._selected_code_and_third_party()
        if not code:
            return
        confirm = custom_askyesno(self,
                                  lang.t("loans.turn_to_donation","Turn to Donation"),
                                  lang.t("loans.turn_confirm_tp",
                                         "Convert loan/borrowing rows for code {code}{tp_part} to Donation?",
                                         code=code,
                                         tp_part=(f' / {tp}' if tp else '')))
        if confirm != "yes":
            return
        changed = self._convert_code_third_party_to_donation(code, tp)
        if changed:
            custom_popup(self,
                         lang.t("generic.success","Success"),
                         lang.t("loans.turn_success_tp","Updated {n} rows to Donation for {code}{tp_part}")
                         .format(n=changed, code=code, tp_part=(f' / {tp}' if tp else '')),
                         "info")
            self.refresh()
        else:
            custom_popup(self,
                         lang.t("generic.info","Info"),
                         lang.t("loans.turn_none","No loan/borrowing rows found to update."),
                         "info")

    def _convert_code_third_party_to_donation(self, code, third_party):
        conn = connect_db()
        if conn is None: return 0
        cur = conn.cursor()
        count = 0
        try:
            if third_party:
                cur.execute("""
                  SELECT rowid, IN_Type, Out_Type, document_number
                  FROM stock_transactions
                  WHERE code=? AND Third_Party=?
                    AND (Out_Type IN ('Loan','Return of Borrowing')
                         OR IN_Type IN ('In Borrowing','In Return of Loan'))
                """,(code, third_party))
            else:
                cur.execute("""
                  SELECT rowid, IN_Type, Out_Type, document_number
                  FROM stock_transactions
                  WHERE code=?
                    AND (Out_Type IN ('Loan','Return of Borrowing')
                         OR IN_Type IN ('In Borrowing','In Return of Loan'))
                """,(code,))
            rows = cur.fetchall()
            for rowid, in_type, out_type, doc in rows:
                new_in = in_type
                new_out = out_type
                modified = False
                if out_type in OUT_TYPES_GIVEN:
                    new_out = DONATION_OUT; modified = True
                if in_type in IN_TYPES_RECEIVED:
                    new_in = DONATION_IN; modified = True
                if modified:
                    new_doc = doc or ""
                    if DONATION_TAG not in new_doc:
                        new_doc = (new_doc + DONATION_TAG) if new_doc else DONATION_TAG.strip()
                    cur.execute("""
                      UPDATE stock_transactions
                      SET IN_Type=?, Out_Type=?, document_number=?
                      WHERE rowid=?""",(new_in,new_out,new_doc,rowid))
                    count += 1
            conn.commit()
        except:
            conn.rollback()
            count = 0
        finally:
            cur.close(); conn.close()
        return count

    # ---------- Export ----------
    def export_excel(self):
        if not self.rows:
            custom_popup(self, lang.t("generic.info","Info"),
                         lang.t("loans.no_data_export","Nothing to export."),"warning")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            title=lang.t("loans.export_title","Save Loans Report"),
            initialfile=f"Loans_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not path: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Loans"

            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([lang.t("generic.generated","Generated"), now_str])
            ws.append([lang.t("generic.filters_used","Filters Used")])
            ws.append(["Scenario", self.scenario_var.get(),
                       "Kit", self.kit_var.get(),
                       "Module", self.module_var.get()])
            ws.append(["Type", self.type_var.get(),
                       "Third Party", self.third_party_var.get(),
                       "Document", self.doc_var.get()])
            ws.append(["From", self.from_var.get(),
                       "To", self.to_var.get(),
                       "Mode", "Simple" if self.simple_mode else "Detailed"])
            ws.append([])

            cols = self._current_columns()
            headers = [c.replace("_"," ").title() for c in cols]
            ws.append(headers)

            kit_fill = PatternFill(start_color=KIT_FILL_COLOR,end_color=KIT_FILL_COLOR,fill_type="solid")
            module_fill = PatternFill(start_color=MODULE_FILL_COLOR,end_color=MODULE_FILL_COLOR,fill_type="solid")

            for r in self.rows:
                if self.simple_mode:
                    line = [
                        r["code"], r["description"],
                        r["qty_given"], r["qty_received"],
                        r["balance"], r["status"]
                    ]
                else:
                    line = [
                        r["scenarios"], r["kits"], r["modules"], r["type"],
                        r["code"], r["description"], r["third_party"],
                        r["qty_given"], r["qty_received"],
                        r["balance"], r["status"], r["documents"]
                    ]
                ws.append(line)
                dtype = r["type"].upper()
                if dtype == "KIT":
                    for c in ws[ws.max_row]: c.fill = kit_fill
                elif dtype == "MODULE":
                    for c in ws[ws.max_row]: c.fill = module_fill

            for col in ws.columns:
                max_len=0
                letter=get_column_letter(col[0].column)
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(v))
                ws.column_dimensions[letter].width = min(max_len+2, 60)

            wb.save(path)
            custom_popup(self, lang.t("generic.success","Success"),
                         lang.t("loans.export_success","Export completed: {f}").format(f=path),
                         "info")
        except Exception as e:
            custom_popup(self, lang.t("generic.error","Error"),
                         lang.t("loans.export_fail","Export failed: {err}").format(err=str(e)),
                         "error")

# ---------------- Module Export ----------------
__all__ = ["Loans"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Loans Monitor v1.1")
    Loans(root, None).pack(fill="both", expand=True)
    root.geometry("1580x820")
    root.mainloop()
