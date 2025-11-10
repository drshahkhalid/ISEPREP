import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import sqlite3
import os
import re
import logging
from db import connect_db
from language_manager import lang
from stock_data import parse_expiry   # only need parse_expiry now
from transaction_utils import log_transaction
from popup_utils import custom_popup, custom_askyesno, custom_dialog
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# ---------------- Helper lookups ---------------- #

def get_active_designation(code):
    conn = connect_db()
    if conn is None:
        return code
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT designation, designation_en, designation_fr, designation_sp
            FROM items_list WHERE code = ?
        """, (code,))
        row = cur.fetchone()
        if not row:
            return code
        lang_code = (lang.lang_code or "en").lower()
        pref_map = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp"
        }
        order = [pref_map.get(lang_code, "designation_en"),
                 "designation_en", "designation_fr", "designation_sp", "designation"]
        for col in order:
            if col in row.keys() and row[col]:
                return row[col]
        return code
    finally:
        cur.close()
        conn.close()


def get_active_item_type(code):
    conn = connect_db()
    if conn is None:
        return "ITEM"
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT type FROM items_list WHERE code = ?", (code,))
        row = cur.fetchone()
        if not row:
            return "ITEM"
        t = row["type"] if "type" in row.keys() else None
        return (t or "ITEM").upper()
    finally:
        cur.close()
        conn.close()


def check_expiry_required(code):
    conn = connect_db()
    if conn is None:
        return False
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT remarks FROM items_list WHERE code = ?", (code,))
        r = cur.fetchone()
        if not r:
            return False
        remarks = r["remarks"] if "remarks" in r.keys() else ""
        return bool(remarks and 'exp' in remarks.lower())
    finally:
        cur.close()
        conn.close()


def fetch_project_details():
    conn = connect_db()
    if conn is None:
        return ("Unknown Project", "Unknown Code")
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        if not row:
            return ("Unknown Project", "Unknown Code")
        return (row[0] or "Unknown Project", row[1] or "Unknown Code")
    finally:
        cur.close()
        conn.close()


# ---------------- Main Class ---------------- #

class StockOut(tk.Frame):
    def __init__(self, parent, root, role="supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.root = root
        self.role = role.lower()
        self.scenario_map = self._load_scenarios()
        self.reverse_scenario_map = {v: k for k, v in self.scenario_map.items()}
        self.user_inputs = {}
        self.row_data = {}
        self.current_document_number = None
        self.pack(fill="both", expand=True)
        self._build_ui()
        self.populate_table()

    # ---- Popup helpers ----
    def _info(self, key, default_text, **fmt):
        custom_popup(self, lang.t("dialog_titles.success", "Success"),
                     lang.t(key, default_text, **fmt), "info")

    def _err(self, key, default_text, **fmt):
        custom_popup(self, lang.t("dialog_titles.error", "Error"),
                     lang.t(key, default_text, **fmt), "error")

    def _ask(self, key, default_text, **fmt):
        return custom_askyesno(self, lang.t("dialog_titles.confirm", "Confirm"),
                               lang.t(key, default_text, **fmt)) == "yes"

    # ---- Scenario / schema ----
    def _load_scenarios(self):
        conn = connect_db()
        if conn is None:
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios")
            return {str(r['scenario_id']): r['name'] for r in cur.fetchall()}
        finally:
            cur.close()
            conn.close()

    # ---- Document Number ----
    def generate_document_number(self, out_type_text: str) -> str:
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").upper()
        base_map = {
            "Issue to End User": "OEU",
            "Expired Items": "OEXP",
            "Damaged Items": "ODMG",
            "Cold Chain Break": "OCCB",
            "Batch Recall": "OBRC",
            "Theft": "OTHF",
            "Other Losses": "OLS",
            "Out Donation": "ODN",
            "Loan": "OLOAN",
            "Return of Borrowing": "OROB",
            "Quarantine": "OQRT"
        }
        raw = (out_type_text or "").strip()
        norm = re.sub(r'[^a-z0-9]+', '', raw.lower())
        abbr = None
        for k, v in base_map.items():
            if re.sub(r'[^a-z0-9]+', '', k.lower()) == norm:
                abbr = v
                break
        if not abbr:
            tokens = re.split(r'\s+', raw.upper())
            stop = {"OF", "FROM", "THE", "AND", "DE", "DU", "DES", "LA", "LE", "LES", "TO"}
            parts = []
            for t in tokens:
                if t and t not in stop:
                    parts.append("MSF" if t == "MSF" else t[0])
            abbr = "".join(parts) if parts else (raw[:4].upper() or "OUT")
            abbr = abbr[:8]
        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"
        serial = 1
        conn = connect_db()
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT document_number FROM stock_transactions
                    WHERE document_number LIKE ?
                    ORDER BY document_number DESC LIMIT 1
                """, (prefix + "/%",))
                r = cur.fetchone()
                if r and r[0]:
                    tail = r[0].rsplit('/', 1)[-1]
                    if tail.isdigit():
                        serial = int(tail) + 1
            except:
                pass
            finally:
                cur.close(); conn.close()
        doc = f"{prefix}/{serial:04d}"
        self.current_document_number = doc
        return doc

    # ---- Data Fetch (final_qty & management_mode filters) ----
    def _fetch_rows(self, code_filter=None, scenario_filter=None):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_data)")
            cols = [c[1] for c in cur.fetchall()]
            if "final_qty" not in cols or "management_mode" not in cols:
                logging.warning("Required columns (final_qty / management_mode) missing.")
                return []
            sql = """
                SELECT unique_id, scenario, final_qty, exp_date, management_mode
                FROM stock_data
                WHERE final_qty > 0
                  AND REPLACE(REPLACE(LOWER(TRIM(management_mode)),'-',' '),' ','_') = 'on_shelf'
            """
            params = []
            if code_filter:
                sql += " AND unique_id LIKE ?"
                params.append(f"%/%/%/{code_filter}/%")
            if scenario_filter and scenario_filter != lang.t("stock_out.all_scenarios", "All Scenarios"):
                scen_id = self.reverse_scenario_map.get(scenario_filter)
                if scen_id:
                    sql += " AND (scenario = ? OR scenario = ?)"
                    params.extend([scenario_filter, scen_id])
                else:
                    sql += " AND scenario = ?"
                    params.append(scenario_filter)
            sql += " ORDER BY scenario, unique_id"
            cur.execute(sql, params)
            rows = cur.fetchall()
            processed = []
            for r in rows:
                unique_id = r["unique_id"]
                parts = unique_id.split("/")
                scenario_token = parts[0] if len(parts) > 0 else (r["scenario"] if "scenario" in r.keys() else "")
                item_code = parts[3] if len(parts) > 3 else ""
                exp_segment = parts[5] if len(parts) > 5 else (r["exp_date"] if "exp_date" in r.keys() else "")
                final_qty = r["final_qty"] if "final_qty" in r.keys() else 0
                scenario_display = self.scenario_map.get(str(scenario_token), scenario_token)
                processed.append({
                    "unique_id": unique_id,
                    "scenario_name": scenario_display,
                    "code": item_code,
                    "description": get_active_designation(item_code),
                    "type": get_active_item_type(item_code),
                    "exp_date": exp_segment if exp_segment not in ("None", None) else "",
                    "current_stock": final_qty,
                })
            return processed
        finally:
            cur.close()
            conn.close()

    # ---- UI construction (unchanged layout) ----
    def _build_ui(self):
        bg = "#F5F7FA"
        self.configure(bg=bg)
        tk.Label(self, text=lang.t("stock_out.title", "Stock Out"),
                 font=("Helvetica", 20, "bold"), bg=bg).pack(pady=10)

        top = tk.Frame(self, bg=bg); top.pack(fill="x", pady=5)
        can_edit = self.role in ["admin", "manager"]
        tk.Button(top, text=lang.t("stock_out.add_button", "Process Stock Out"),
                  bg="#27AE60", fg="white", command=self.save_all,
                  state="normal" if can_edit else "disabled").pack(side="left", padx=5)
        tk.Button(top, text=lang.t("stock_out.clear_all", "Clear All"),
                  bg="#7F8C8D", fg="white", command=self.clear_form).pack(side="left", padx=5)
        tk.Button(top, text=lang.t("stock_out.export", "Export"),
                  bg="#2980B9", fg="white", command=self.export_data).pack(side="left", padx=5)

        filt = tk.Frame(self, bg=bg); filt.pack(fill="x", pady=5)
        tk.Label(filt, text=lang.t("stock_out.scenario", "Scenario:"), bg=bg).grid(row=0, column=0, padx=5, sticky="w")
        self.scenario_var = tk.StringVar(value=lang.t("stock_out.all_scenarios", "All Scenarios"))
        scenario_values = list(self.scenario_map.values()) + [lang.t("stock_out.all_scenarios", "All Scenarios")]
        self.scenario_cb = ttk.Combobox(filt, textvariable=self.scenario_var,
                                        values=scenario_values, state="readonly", width=35)
        self.scenario_cb.grid(row=0, column=1, padx=5, pady=3)
        self.scenario_cb.bind("<<ComboboxSelected>>", lambda e: self.populate_table())

        type_frame = tk.Frame(self, bg=bg); type_frame.pack(fill="x", pady=5)
        tk.Label(type_frame, text=lang.t("stock_out.out_type", "OUT Type:"), bg=bg).grid(row=0, column=0, padx=5, sticky="w")
        self.trans_type_var = tk.StringVar()
        self.trans_type_cb = ttk.Combobox(type_frame, textvariable=self.trans_type_var,
                                          values=[
                                              lang.t("stock_out.issue_to_end_user", "Issue to End User"),
                                              lang.t("stock_out.expired_items", "Expired Items"),
                                              lang.t("stock_out.damaged_items", "Damaged Items"),
                                              lang.t("stock_out.cold_chain_break", "Cold Chain Break"),
                                              lang.t("stock_out.batch_recall", "Batch Recall"),
                                              lang.t("stock_out.theft", "Theft"),
                                              lang.t("stock_out.other_losses", "Other Losses"),
                                              lang.t("stock_out.out_donation", "Out Donation"),
                                              lang.t("stock_out.loan", "Loan"),
                                              lang.t("stock_out.return_of_borrowing", "Return of Borrowing"),
                                              lang.t("stock_out.quarantine", "Quarantine"),
                                          ], state="readonly", width=30)
        self.trans_type_cb.grid(row=0, column=1, padx=5, pady=3)
        self.trans_type_cb.bind("<<ComboboxSelected>>", self._update_party_enable)

        tk.Label(type_frame, text=lang.t("stock_out.end_user", "End User:"), bg=bg).grid(row=0, column=2, padx=5, sticky="w")
        self.end_user_var = tk.StringVar()
        self.end_user_cb = ttk.Combobox(type_frame, textvariable=self.end_user_var, state="disabled", width=30)
        self.end_user_cb['values'] = self._fetch_end_users()
        self.end_user_cb.grid(row=0, column=3, padx=5, pady=3)

        tk.Label(type_frame, text=lang.t("stock_out.third_party", "Third Party:"), bg=bg).grid(row=0, column=4, padx=5, sticky="w")
        self.third_party_var = tk.StringVar()
        self.third_party_cb = ttk.Combobox(type_frame, textvariable=self.third_party_var,
                                           state="disabled", width=30)
        self.third_party_cb['values'] = self._fetch_third_parties()
        self.third_party_cb.grid(row=0, column=5, padx=5, pady=3)

        tk.Label(type_frame, text=lang.t("stock_out.remarks", "Remarks:"), bg=bg).grid(row=0, column=6, padx=5, sticky="w")
        self.remarks_entry = tk.Entry(type_frame, width=50)
        self.remarks_entry.grid(row=0, column=7, padx=5, pady=3)

        search_frame = tk.Frame(self, bg=bg); search_frame.pack(fill="x", pady=8)
        tk.Label(search_frame, text=lang.t("stock_out.item_code", "Item Code"), bg=bg).grid(row=0, column=0, padx=5, sticky="w")
        self.code_entry = tk.Entry(search_frame, width=25)
        self.code_entry.grid(row=0, column=1, padx=5)
        self.code_entry.bind("<KeyRelease>", self._live_search)
        self.code_entry.bind("<Return>", lambda e: self._select_first_search())
        tk.Button(search_frame, text=lang.t("stock_out.clear_search", "Clear Search"),
                  bg="#7F8C8D", fg="white", command=self._clear_search).grid(row=0, column=2, padx=5)
        self.search_listbox = tk.Listbox(search_frame, height=5, width=80)
        self.search_listbox.grid(row=0, column=3, padx=5, pady=5, sticky="we")
        self.search_listbox.bind("<<ListboxSelect>>", self._apply_search_selection)

        self.cols = ("unique_id", "code", "description", "type",
                     "scenario_name", "current_stock", "exp_date",
                     "qty_out", "remaining_quantity")
        tree_frame = tk.Frame(self, bg=bg); tree_frame.pack(expand=True, fill="both", pady=5)
        self.tree = ttk.Treeview(tree_frame, columns=self.cols, show="headings", height=20)
        self.tree.tag_configure("light_red", background="#FF9999")
        headers = {
            "code": lang.t("stock_out.code", "Code"),
            "description": lang.t("stock_out.description", "Description"),
            "type": lang.t("stock_out.type", "Type"),
            "scenario_name": lang.t("stock_out.scenario_name", "Scenario"),
            "current_stock": lang.t("stock_out.current_stock", "Current Stock"),
            "exp_date": lang.t("stock_out.expiry_date", "Expiry Date"),
            "qty_out": lang.t("stock_out.qty_out", "Qty Out"),
            "remaining_quantity": lang.t("stock_out.remaining_quantity", "Remaining Qty")
        }
        widths = {
            "code": 120, "description": 420, "type": 110, "scenario_name": 170,
            "current_stock": 110, "exp_date": 120, "qty_out": 90, "remaining_quantity": 130
        }
        for c in self.cols:
            if c == "unique_id":
                self.tree.column(c, width=0, stretch=False)
            else:
                self.tree.heading(c, text=headers.get(c, c))
                self.tree.column(c, width=widths.get(c, 120), stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview); vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview); hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        self.tree.bind("<Double-1>", self._start_edit_qty_out)

        self.status_var = tk.StringVar(value=lang.t("stock_out.ready", "Ready"))
        tk.Label(self, textvariable=self.status_var, anchor="w", relief="sunken",
                 bg=bg).pack(fill="x", pady=(4, 0))

    # ---- Search & populate ----
    def _live_search(self, event=None):
        q = self.code_entry.get().strip()
        self.search_listbox.delete(0, tk.END)
        if not q:
            self.populate_table()
            return
        rows = self._fetch_rows(None, self.scenario_var.get())
        seen = set()
        for r in rows:
            if q.lower() in r["code"].lower() and r["code"] not in seen:
                self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
                seen.add(r["code"])
        self.status_var.set(lang.t("stock_out.found_items", "Found {size} items").format(size=self.search_listbox.size()))

    def _select_first_search(self):
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self._apply_search_selection()

    def _apply_search_selection(self, event=None):
        sel = self.search_listbox.curselection()
        if not sel:
            return
        code = self.search_listbox.get(sel[0]).split(" - ")[0]
        self.search_listbox.delete(0, tk.END)
        self._show_only_code(code)

    def _clear_search(self):
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.populate_table()

    def _show_only_code(self, code):
        self._preserve_user_inputs()
        self.tree.delete(*self.tree.get_children())
        rows = self._fetch_rows(code, self.scenario_var.get())
        if not rows:
            self.status_var.set(lang.t("stock_out.no_items", "No items found for code {code}").format(code=code))
            return
        for r in rows:
            self._insert_row(r, restore=True)
        self.status_var.set(lang.t("stock_out.loaded", "Loaded {count} records for code {code}")
                            .format(count=len(self.tree.get_children()), code=code))

    def populate_table(self):
        self._preserve_user_inputs()
        self.tree.delete(*self.tree.get_children())
        rows = self._fetch_rows(None, self.scenario_var.get())
        if not rows:
            self.status_var.set(lang.t("stock_out.no_items", "No stock data available"))
            return
        for r in rows:
            self._insert_row(r, restore=True)
        self.status_var.set(lang.t("stock_out.loaded", "Loaded {count} records")
                            .format(count=len(self.tree.get_children())))

    def _insert_row(self, row_dict, restore=False):
        uid = row_dict["unique_id"]
        current_stock = row_dict["current_stock"]
        if restore and uid in self.user_inputs:
            qty_out = self.user_inputs[uid].get("qty_out", "")
        else:
            qty_out = ""
        remaining = self._remaining(current_stock, qty_out)
        iid = self.tree.insert("", "end", values=(
            uid, row_dict["code"], row_dict["description"], row_dict["type"],
            row_dict["scenario_name"], current_stock, row_dict["exp_date"],
            qty_out, remaining
        ))
        self.row_data[iid] = row_dict

    def _preserve_user_inputs(self):
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 9:
                continue
            uid = vals[0]
            self.user_inputs[uid] = {"qty_out": vals[7], "remaining": vals[8]}

    # ---- Editing qty_out ----
    def _start_edit_qty_out(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not row_id or not col:
            return
        col_index = int(col.replace("#", "")) - 1
        if col_index != 7:
            return
        bbox = self.tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox
        current_val = self.tree.set(row_id, "qty_out")
        current_stock = int(self.tree.set(row_id, "current_stock"))
        code = self.tree.set(row_id, "code")
        expiry_required = check_expiry_required(code)
        entry = tk.Entry(self.tree, font=("Helvetica", 10))
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, current_val)
        entry.focus()

        def finish(_=None):
            new_val = entry.get().strip()
            if new_val and not new_val.isdigit():
                self._err("stock_out.invalid_qty", "Qty Out must be an integer")
                return
            if new_val and int(new_val) > current_stock:
                self._err("stock_out.qty_exceeds_stock",
                          "Qty Out ({new_val}) cannot exceed Current Stock ({current_stock})",
                          new_val=new_val, current_stock=current_stock)
                return
            self.tree.set(row_id, "qty_out", new_val)
            exp_date = self.tree.set(row_id, "exp_date")
            remaining = self._remaining(current_stock, new_val)
            self.tree.set(row_id, "remaining_quantity", remaining)
            if new_val and new_val.isdigit() and expiry_required and not self.validate_expiry_for_save(code, new_val, exp_date):
                self.tree.item(row_id, tags=("light_red",))
            else:
                self.tree.item(row_id, tags=())
            uid = self.tree.set(row_id, "unique_id")
            self.user_inputs[uid] = {"qty_out": new_val, "remaining": remaining}
            entry.destroy()

        entry.bind("<Return>", finish)
        entry.bind("<Tab>", finish)
        entry.bind("<FocusOut>", finish)

    # ---- Calculations / validation ----
    def _remaining(self, current_stock, qty_out):
        if qty_out and qty_out.isdigit():
            return max(int(current_stock) - int(qty_out), 0)
        return int(current_stock)

    def validate_expiry_for_save(self, code, qty_out, exp_date):
        if not qty_out or not qty_out.isdigit():
            return True
        if not check_expiry_required(code):
            return True
        parsed = parse_expiry(exp_date) if exp_date else None
        today = datetime.now().date()
        return bool(parsed and parsed > today)

    # ---- Direct delta application (NO final_qty touch) ----
    def _apply_stock_out_delta(self, unique_id, delta_qty_out, expiry_date=None, scenario_name=None):
        """
        Increment qty_out by delta_qty_out exactly once.
        Relies on DB triggers to recompute final_qty.
        """
        if delta_qty_out <= 0:
            return
        conn = connect_db()
        if conn is None:
            raise RuntimeError("DB connection failed")
        cur = conn.cursor()
        try:
            # Debug before
            cur.execute("SELECT qty_in, qty_out, final_qty FROM stock_data WHERE unique_id=?", (unique_id,))
            before = cur.fetchone()
            logging.info(f"[STOCK_OUT][BEFORE] {unique_id} -> {before}")

            cur.execute("""
                UPDATE stock_data
                   SET qty_out = COALESCE(qty_out,0) + ?,
                       exp_date = COALESCE(?, exp_date),
                       scenario = COALESCE(?, scenario)
                 WHERE unique_id = ?
            """, (delta_qty_out, expiry_date, scenario_name, unique_id))
            if cur.rowcount == 0:
                # Optionally insert if missing (rare)
                cur.execute("""
                    INSERT INTO stock_data (unique_id, scenario, qty_in, qty_out, exp_date)
                    VALUES (?, ?, 0, ?, ?)
                """, (unique_id, scenario_name, delta_qty_out, expiry_date))
            conn.commit()

            # Debug after
            cur.execute("SELECT qty_in, qty_out, final_qty FROM stock_data WHERE unique_id=?", (unique_id,))
            after = cur.fetchone()
            logging.info(f"[STOCK_OUT][AFTER] {unique_id} -> {after}")

        except Exception:
            conn.rollback()
            raise
        finally:
            cur.close()
            conn.close()

    # ---- Current final (read only) ----
    def _current_final(self, unique_id):
        conn = connect_db()
        if conn is None:
            return 0
        cur = conn.cursor()
        try:
            cur.execute("SELECT final_qty FROM stock_data WHERE unique_id = ?", (unique_id,))
            row = cur.fetchone()
            if not row or row[0] is None:
                return 0
            return max(row[0], 0)
        finally:
            cur.close()
            conn.close()

    def _has_stock(self, unique_id, qty_out_int):
        # Use current final_qty
        return self._current_final(unique_id) >= qty_out_int

    # ---- SAVE ----
    def save_all(self):
        if self.role not in ["admin", "manager"]:
            self._err("stock_out.no_permission", "Only admin or manager roles can save changes.")
            return
        rows = self.tree.get_children()
        if not rows:
            self._err("stock_out.no_rows", "No rows to save.")
            return
        out_type = self.trans_type_var.get()
        if not out_type:
            self._err("stock_out.no_out_type", "OUT Type is required.")
            return

        end_user = self.end_user_var.get().strip()
        third_party = self.third_party_var.get().strip()
        remarks = self.remarks_entry.get().strip()

        if out_type == lang.t("stock_out.issue_to_end_user", "Issue to End User"):
            if not end_user or end_user not in self.end_user_cb['values']:
                self._err("stock_out.invalid_end_user", "A valid End User must be selected.")
                return
        if out_type in [
            lang.t("stock_out.out_donation", "Out Donation"),
            lang.t("stock_out.loan", "Loan"),
            lang.t("stock_out.return_of_borrowing", "Return of Borrowing")
        ]:
            if not third_party or third_party not in self.third_party_cb['values']:
                self._err("stock_out.invalid_third_party",
                          "A valid Third Party must be selected for {ttype}.",
                          ttype=out_type)
                return

        doc_number = self.generate_document_number(out_type)
        invalid = []
        export_rows = []

        # Validate
        for iid in rows:
            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 9:
                continue
            code = vals[1]
            qty_out = vals[7]
            exp_date = vals[6]
            current_stock = int(vals[5]) if str(vals[5]).isdigit() else 0
            if qty_out and qty_out.isdigit():
                if int(qty_out) > current_stock or not self.validate_expiry_for_save(code, qty_out, exp_date):
                    invalid.append(code)
                    self.tree.item(iid, tags=("light_red",))
                else:
                    self.tree.item(iid, tags=())
        if invalid:
            self._err("stock_out.invalid_qty_or_expiry",
                      "Valid Qty Out (<= Current Stock) and future expiry dates are required for items: {items}",
                      items=", ".join(invalid))
            return

        # Apply each delta
        for iid in rows:
            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 9:
                continue
            unique_id, code, description, item_type, scenario_name, current_stock, exp_date, qty_out_txt, remaining = vals
            if not qty_out_txt or not qty_out_txt.isdigit():
                continue
            delta_qty_out = int(qty_out_txt)
            if delta_qty_out <= 0:
                continue

            # Guard final stock before applying
            if not self._has_stock(unique_id, delta_qty_out):
                self._err("stock_out.qty_exceeds_stock",
                          "Qty Out ({new_val}) cannot exceed Current Stock ({current_stock})",
                          new_val=delta_qty_out, current_stock=self._current_final(unique_id))
                continue

            parsed_exp = parse_expiry(exp_date)
            expiry_date = parsed_exp.strftime("%Y-%m-%d") if parsed_exp else None

            parts = unique_id.split("/")
            kit_number = parts[6] if len(parts) > 6 and parts[6] != "None" else ""
            module_number = parts[7] if len(parts) > 7 and parts[7] != "None" else ""

            try:
                self._apply_stock_out_delta(unique_id, delta_qty_out,
                                            expiry_date=expiry_date,
                                            scenario_name=scenario_name)
            except Exception as e:
                self._err("stock_out.save_failed", "Failed updating stock_data: {error}", error=str(e))
                continue

            try:
                log_transaction(
                    unique_id=unique_id,
                    code=code,
                    Description=description,
                    Expiry_date=expiry_date,
                    Batch_Number=None,
                    Scenario=scenario_name,
                    Kit=kit_number or None,
                    Module=module_number or None,
                    Qty_IN=None,
                    IN_Type=None,
                    Qty_Out=delta_qty_out,
                    Out_Type=out_type,
                    Third_Party=third_party if third_party else None,
                    End_User=end_user if end_user else None,
                    Remarks=remarks,
                    Movement_Type="stock_out",
                    document_number=doc_number
                )
                export_rows.append({
                    "code": code,
                    "description": description,
                    "type": item_type,
                    "kit_number": kit_number,
                    "module_number": module_number,
                    "current_stock": int(current_stock) if str(current_stock).isdigit() else 0,
                    "expiry_date": exp_date or "",
                    "batch_number": "",
                    "qty_issued": delta_qty_out
                })
            except Exception as e:
                self._err("stock_out.save_failed", "Failed to log transaction: {error}", error=str(e))
                continue

        self._info("stock_out.saved", "Stock OUT saved successfully.")
        self.status_var.set(lang.t("stock_out.document_number_saved",
                                   "Saved. Document Number: {doc}").format(doc=doc_number))

        if export_rows and self._ask("stock_out.save_excel_prompt",
                                     "Do you want to save the stock issuance to Excel?"):
            self.export_data(export_rows)

        self.clear_form()

    # ---- Clear ----
    def clear_form(self):
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.user_inputs.clear()
        self.trans_type_var.set("")
        self.end_user_var.set("")
        self.third_party_var.set("")
        self.remarks_entry.delete(0, tk.END)
        self.current_document_number = None
        self.populate_table()
        self.status_var.set(lang.t("stock_out.ready", "Ready"))

    # ---- Parties ----
    def _fetch_third_parties(self):
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM third_parties ORDER BY name")
            return [r[0] for r in cur.fetchall()]
        finally:
            cur.close(); conn.close()

    def _fetch_end_users(self):
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM end_users ORDER BY name")
            return [r[0] for r in cur.fetchall()]
        finally:
            cur.close(); conn.close()

    # ---- Export ----
    def export_data(self, rows_to_issue=None):
        logging.info("[STOCK_OUT] export_data invoked")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("[STOCK_OUT] parent window invalid; abort export")
            return
        try:
            if rows_to_issue is None:
                rows = []
                for iid in self.tree.get_children():
                    vals = self.tree.item(iid, "values")
                    if not vals or len(vals) < 9:
                        continue
                    unique_id, code, desc, tfield, scenario_name, current_stock, exp_date, qty_out, remaining = vals
                    if not qty_out or not qty_out.isdigit() or int(qty_out) <= 0:
                        continue
                    parts = unique_id.split("/")
                    kit_number = parts[6] if len(parts) > 6 and parts[6] != "None" else ""
                    module_number = parts[7] if len(parts) > 7 and parts[7] != "None" else ""
                    rows.append({
                        "code": code,
                        "description": desc,
                        "type": tfield,
                        "kit_number": kit_number,
                        "module_number": module_number,
                        "current_stock": int(current_stock) if str(current_stock).isdigit() else 0,
                        "expiry_date": exp_date or "",
                        "batch_number": "",
                        "qty_issued": int(qty_out)
                    })
            else:
                rows = [r for r in rows_to_issue if r.get("qty_issued", 0) > 0]

            if not rows:
                custom_popup(self.parent, lang.t("dialog_titles.error", "Error"),
                             lang.t("stock_out.no_issue_qty", "No quantities entered to issue."), "error")
                return

            project_name, project_code = fetch_project_details()
            out_type_raw = self.trans_type_var.get() or lang.t("stock_out.unknown", "Unknown")
            scenario_name = self.scenario_var.get()
            doc_number = getattr(self, "current_document_number", None)
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            def sanitize(s: str) -> str:
                s = re.sub(r'[^A-Za-z0-9]+', '_', s)
                s = re.sub(r'_+', '_', s)
                return s.strip('_') or "Unknown"

            out_slug = sanitize(out_type_raw)
            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)
            filename = f"StockOut_{out_slug}_{current_time.replace(':','-')}.xlsx"

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=filename,
                initialdir=default_dir
            )
            if not path:
                self.status_var.set(lang.t("stock_out.export_cancelled", "Export cancelled"))
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "StockOut"[:31]

            if doc_number:
                ws['A1'] = f"Date: {current_time}{' ' * 8}Document Number: {doc_number}"
            else:
                ws['A1'] = f"Date: {current_time}"
            ws['A1'].font = Font(name="Calibri", size=11)
            ws['A1'].alignment = Alignment(horizontal="left")

            ws['A2'] = "Stock Out – Movement: Stock Out"
            ws['A2'].font = Font(name="Tahoma", size=14, bold=True)
            ws['A2'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A2:I2')

            ws['A3'] = f"{project_name} - {project_code}"
            ws['A3'].font = Font(name="Tahoma", size=14, bold=True)
            ws['A3'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A3:I3')

            ws['A4'] = f"{lang.t('stock_out.out_type','OUT Type')}: {out_type_raw}"
            ws['A4'].font = Font(name="Tahoma", size=12, bold=True)
            ws['A4'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A4:I4')

            ws['A5'] = f"Scenario: {scenario_name}"
            ws['A5'].font = Font(name="Tahoma", size=12, bold=True)
            ws['A5'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A5:I5')

            ws.append([])

            headers = [
                lang.t("stock_out.code", "Code"),
                lang.t("stock_out.description", "Description"),
                lang.t("stock_out.type", "Type"),
                lang.t("stock_out.kit_number", "Kit Number"),
                lang.t("stock_out.module_number", "Module Number"),
                lang.t("stock_out.current_stock", "Current Stock"),
                lang.t("stock_out.expiry_date", "Expiry Date"),
                lang.t("stock_out.batch_number", "Batch Number"),
                lang.t("stock_out.qty_out", "Qty Issued")
            ]
            ws.append(headers)
            header_row = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=c)
                cell.font = Font(name="Tahoma", size=11, bold=True)

            for r in rows:
                r_type = (r.get("type") or "").lower()
                ws.append([
                    r.get("code",""),
                    r.get("description",""),
                    r.get("type",""),
                    r.get("kit_number",""),
                    r.get("module_number",""),
                    r.get("current_stock",0),
                    r.get("expiry_date",""),
                    r.get("batch_number",""),
                    r.get("qty_issued",0)
                ])
                data_idx = ws.max_row
                fill = None
                bold_flag = r_type in ("kit", "module")
                if r_type == "kit":
                    fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif r_type == "module":
                    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=data_idx, column=col)
                    cell.font = Font(name="Calibri", size=11, bold=bold_flag)
                    if fill:
                        cell.fill = fill

            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col:
                    val = cell.value
                    l = len(str(val)) if val is not None else 0
                    if l > max_len:
                        max_len = l
                ws.column_dimensions[letter].width = min(max_len + 2, 50)

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = '1:7'
            ws.oddFooter.center.text = "&P of &N"
            ws.evenFooter.center.text = "&P of &N"

            wb.save(path)
            custom_popup(self.parent, lang.t("dialog_titles.success", "Success"),
                         lang.t("stock_out.export_success", f"Exported to {path}"), "info")
            self.status_var.set(lang.t("stock_out.export_success", f"Export successful: {path}"))
        except Exception as e:
            logging.error(f"[STOCK_OUT] Export failed: {e}")
            custom_popup(self.parent, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_out.export_failed", f"Export failed: {str(e)}"), "error")

    # ---- Party enabling ----
    def _update_party_enable(self, event=None):
        t = self.trans_type_var.get()
        self.end_user_cb.config(state="disabled")
        self.third_party_cb.config(state="disabled")
        if t == lang.t("stock_out.issue_to_end_user", "Issue to End User"):
            self.end_user_cb.config(state="readonly")
        elif t in [
            lang.t("stock_out.out_donation", "Out Donation"),
            lang.t("stock_out.loan", "Loan"),
            lang.t("stock_out.return_of_borrowing", "Return of Borrowing")
        ]:
            self.third_party_cb.config(state="readonly")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Out")
    shell = tk.Tk()
    shell.role = "admin"
    try:
        StockOut(root, shell, role="admin")
    except Exception as e:
        logging.error(f"Error launching StockOut: {e}")
    root.mainloop()
