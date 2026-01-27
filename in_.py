import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import sqlite3
import re
from calendar import monthrange
from language_manager import lang
from db import connect_db
from transaction_utils import log_transaction
from stock_data import StockData
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from popup_utils import custom_popup, custom_askyesno, custom_dialog
import os

# Unified popup utilities (consistent with receive_kit / dispatch_kit)
from popup_utils import custom_popup, custom_askyesno

# ---------------------- Utility / Parsing ---------------------- #
def parse_expiry(date_str):
    """
    Parse various date formats; for MM/YYYY or YYYY/MM return last day of month.
    Returns datetime.date or None.
    """
    if not date_str or str(date_str).lower() == 'none':
        return None
    date_str = date_str.strip()
    date_str = re.sub(r'[-/\s]+', '/', date_str)
    formats = [
        r'(\d{1,2})/(\d{1,2})/(\d{4})',  # DD/MM/YYYY
        r'(\d{4})/(\d{1,2})/(\d{1,2})',  # YYYY/MM/DD
        r'(\d{1,2})/(\d{4})',            # MM/YYYY
        r'(\d{4})/(\d{1,2})'             # YYYY/MM
    ]
    from calendar import monthrange as mr
    try:
        for fmt in formats:
            m = re.match(fmt, date_str)
            if not m:
                continue
            g = m.groups()
            if len(g) == 3:
                if fmt.startswith(r'(\d{1,2})'):  # DD/MM/YYYY
                    day, month, year = map(int, g)
                else:  # YYYY/MM/DD
                    year, month, day = map(int, g)
                return datetime(year, month, day).date()
            elif len(g) == 2:
                if fmt.startswith(r'(\d{1,2})'):
                    month, year = map(int, g)
                else:
                    year, month = map(int, g)
                _, last_day = mr(year, month)
                return datetime(year, month, last_day).date()
        return None
    except ValueError:
        return None

def get_active_designation(code):
    """
    Returns description in active language (fallback chain).
    """
    conn = connect_db()
    if conn is None:
        return lang.t("stock_in.no_description", "No Description")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT designation, designation_en, designation_fr, designation_sp
            FROM items_list WHERE code=?
        """, (code,))
        row = cur.fetchone()
        if not row:
            return lang.t("stock_in.no_description", "No Description")
        lang_code = (lang.lang_code or "en").lower()
        mapping = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp"
        }
        active_col = mapping.get(lang_code, "designation_en")
        if row[active_col]:
            return row[active_col]
        if row["designation_en"]:
            return row["designation_en"]
        return row["designation"] if row["designation"] else lang.t("stock_in.no_description", "No Description")
    finally:
        cur.close()
        conn.close()

def check_expiry_required(code):
    """
    Returns True if 'exp' found in items_list.remarks.
    """
    conn = connect_db()
    if conn is None:
        return False
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT remarks FROM items_list WHERE code=?", (code,))
        row = cur.fetchone()
        return bool(row and row['remarks'] and 'exp' in row['remarks'].lower())
    finally:
        cur.close()
        conn.close()



def fetch_project_details():
    """
    Returns (project_name, project_code).
    """
    conn = connect_db()
    if conn is None:
        return lang.t("stock_in.unknown_project", "Unknown Project"), lang.t("stock_in.unknown_code", "Unknown Code")
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        return (row[0] if row and row[0] else lang.t("stock_in.unknown_project", "Unknown Project"),
                row[1] if row and row[1] else lang.t("stock_in.unknown_code", "Unknown Code"))
    finally:
        cur.close()
        conn.close()

def fetch_qty_needed(self, code, scenario_name, scenario_id, std_qty):
    """
    Compute qty_needed = qty_to_order_per_scenario - (session qty_in).
    Fallback to std_qty if no stored target.
    """
    conn = connect_db()
    if conn is None:
        return 0
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT COALESCE(qty_to_order_per_scenario, 0)
            FROM stock_data
            WHERE item=? AND scenario=?
        """, (code, scenario_name))
        row = cur.fetchone()
        base_needed = int(row[0]) if row else None
        if base_needed is None:
            base_needed = int(std_qty) if str(std_qty).isdigit() else 0
        session_in = 0
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if vals[0] == code and vals[2] == scenario_name:
                qin = vals[7]
                if qin and str(qin).isdigit():
                    session_in += int(qin)
        return max(base_needed - session_in, 0)
    finally:
        cur.close()
        conn.close()



# ---------------------- Main Class ---------------------- #
class StockIn(tk.Frame):


# ---------- Canonical IN Types (class-level constant) ----------
    IN_TYPE_CANONICAL = [
        "In MSF",
        "In Local Purchase",
        "In from Quarantine",
        "In Donation",
        "Return from End User",
        "In Supply Non-MSF",
        "In Borrowing",
        "In Return of Loan",
        "In Correction of Previous Transaction",
    ]
        
    def __init__(self, parent: tk.Widget, root: tk.Tk, role: str = "supervisor"):
        super().__init__(parent)
        self.root = root
        self.role = role.lower()
        self.editing_cell = None
        self.scenario_map = self.fetch_scenario_map()
        self.reverse_scenario_map = {v: k for k, v in self.scenario_map.items()}
        self.user_inputs = {}
        self.row_data = {}             # item_id -> {'unique_id','unique_id_2'}
        self.context_menu_row = None
        self.current_document_number = None  # store doc number for export
        self.pack(fill="both", expand=True)
        self.render_ui()

    # ---------- Unified popup wrappers ---------- #
    def show_error(self, msg_key_default, default_text, **fmt):
        custom_popup(self, lang.t("dialog_titles.error", "Error"),
                     lang.t(msg_key_default, default_text, **fmt), "error")

    def show_info(self, msg_key_default, default_text, **fmt):
        custom_popup(self, lang.t("dialog_titles.success", "Success"),
                     lang.t(msg_key_default, default_text, **fmt), "info")

    def ask_yes_no(self, msg_key_default, default_text, **fmt):
        return custom_askyesno(self, lang.t("dialog_titles.confirm", "Confirm"),
                               lang.t(msg_key_default, default_text, **fmt)) == "yes"


    #-----------Helpers for Expiry date column and popups---------#

    def _validate_date_format(self, date_str):
        """Validate date format (YYYY-MM-DD)."""
        if not date_str:
            return True
        import re
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return False
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    
    def _is_date_in_future(self, date_str):
        """Check if date is in the future (or today)."""
        if not date_str:
            return True
        try:
            exp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            today = datetime.today().date()
            return exp_date >= today
        except ValueError:
            return False

    def _last_day_of_month(self, year, month):
        """Get last day of month in YYYY-MM-DD format."""
        from calendar import monthrange
        if not (1 <= month <= 12):
            raise ValueError(f"Invalid month: {month}")
        last_day = monthrange(year, month)[1]
        return f"{year:04d}-{month:02d}-{last_day:02d}"
    
    def _parse_ambiguous_slash_date(self, groups):
        """Parse DD/MM/YYYY vs MM/DD/YYYY."""
        first = int(groups[0])
        second = int(groups[1])
        year = groups[2]
        if first > 12:
            return f"{year}-{second:02d}-{first:02d}"
        if second > 12:
            return f"{year}-{first:02d}-{second:02d}"
        return f"{year}-{second:02d}-{first:02d}"

    def _parse_flexible_date(self, date_input):
        """
        Parse flexible date formats and convert to YYYY-MM-DD.
        Returns: (success: bool, parsed_date: str or None, error_msg: str)
        """
        if not date_input:
            return True, None, ""
        
        import re
        date_input = date_input.strip()
        
        patterns = [
            (r'^(\d{4})-(\d{2})-(\d{2})$', 'YYYY-MM-DD', 
             lambda m: f"{m[0]}-{m[1]}-{m[2]}"),
            (r'^(\d{4})-(\d{2})$', 'YYYY-MM', 
             lambda m: self._last_day_of_month(int(m[0]), int(m[1]))),
            (r'^(\d{4})$', 'YYYY', 
             lambda m: f"{m[0]}-12-31"),
            (r'^(\d{1,2})/(\d{4})$', 'MM/YYYY', 
             lambda m: self._last_day_of_month(int(m[1]), int(m[0]))),
            (r'^(\d{2})-(\d{2})-(\d{4})$', 'DD-MM-YYYY', 
             lambda m: f"{m[2]}-{m[1]}-{m[0]}"),
            (r'^(\d{1,2})/(\d{1,2})/(\d{4})$', 'DD/MM/YYYY or MM/DD/YYYY', 
             self._parse_ambiguous_slash_date),
        ]
        
        for pattern, format_name, converter in patterns:
            match = re.match(pattern, date_input)
            if match:
                try:
                    groups = match.groups()
                    result = converter(groups)
                    if result is None:
                        continue
                    datetime.strptime(result, '%Y-%m-%d')
                    return True, result, ""
                except (ValueError, IndexError):
                    continue
        
        error_msg = (
            f"Cannot parse date: '{date_input}'\n\n"
            f"Supported formats:\n"
            f"  • YYYY-MM-DD    (e.g., 2027-12-31)\n"
            f"  • YYYY-MM       (e.g., 2027-12 → 2027-12-31)\n"
            f"  • YYYY          (e.g., 2027 → 2027-12-31)\n"
            f"  • MM/YYYY       (e.g., 12/2029 → 2029-12-31)\n"
            f"  • DD-MM-YYYY    (e.g., 31-12-2027)\n"
            f"  • DD/MM/YYYY    (e.g., 31/12/2027)\n"
            f"  • MM/DD/YYYY    (e.g., 12/31/2027)"
        )
        return False, None, error_msg    


    # ---------- Helpers for canonical IN types ----------
    def get_canonical_in_type(self) -> str:
        """
        Convert the currently selected display label to canonical English
        using translations section 'stock_in.in_types_map'.
        Returns a canonical English string (e.g., 'In MSF').
        """
        display_val = (self.trans_type_var.get() or "").strip()
        return lang.enum_to_canonical("stock_in.in_types_map", display_val).strip()

    def get_display_in_type_list(self):
        """
        Return localized display list for IN types from the canonical list.
        These are shown in the combobox, but will be converted back to
        canonical English for DB writes.
        """
        return lang.enum_to_display_list("stock_in.in_types_map", self.IN_TYPE_CANONICAL)

    # ---------- Document Number Generation ----------
    def generate_document_number(self, in_type_text: str) -> str:
        """
        Format: YYYY/MM/<PROJECT_CODE>/<ABBR>/<SERIAL>
        ABBR from canonical English; in_type_text MUST be canonical English.
        """
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").strip().upper()

        base_map = {
            "In MSF": "IMSF",
            "In Local Purchase": "ILP",
            "In from Quarantine": "IFQ",
            "In Donation": "IDN",
            "Return from End User": "IREU",
            "In Supply Non-MSF": "ISNM",
            "In Borrowing": "IBR",
            "In Return of Loan": "IRL",
            "In Correction of Previous Transaction": "ICOR"
        }
        raw = (in_type_text or "").strip()
        norm = re.sub(r'[^a-z0-9]+', '', raw.lower())
        abbr = None
        for k, v in base_map.items():
            if re.sub(r'[^a-z0-9]+', '', k.lower()) == norm:
                abbr = v
                break
        if not abbr:
            tokens = re.split(r'\s+', raw.upper())
            stop = {"OF", "FROM", "THE", "AND", "DE", "DU", "DES", "LA", "LE", "LES"}
            parts = []
            for t in tokens:
                if not t or t in stop:
                    continue
                parts.append("MSF" if t == "MSF" else t[0])
            abbr = "".join(parts) if parts else (raw[:4].upper() or "DOC").replace(" ", "")
            if len(abbr) > 8:
                abbr = abbr[:8]

        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"
        serial = 1
        conn = connect_db()
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT document_number
                      FROM stock_transactions
                     WHERE document_number LIKE ?
                     ORDER BY document_number DESC
                     LIMIT 1
                """, (prefix + "/%",))
                row = cur.fetchone()
                if row and row[0]:
                    tail = row[0].rsplit("/", 1)[-1]
                    if tail.isdigit():
                        serial = int(tail) + 1
            except Exception:
                pass
            finally:
                cur.close()
                conn.close()
        doc = f"{prefix}/{serial:04d}"
        self.current_document_number = doc
        return doc
    
    # -------- Data Fetch Helpers -------- #
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None:
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            return {str(r['scenario_id']): r['name'] for r in cur.fetchall()}
        finally:
            cur.close()
            conn.close()

    def fetch_search_results(self, query):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        lang_code = (lang.lang_code or "en").lower()
        mapping = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp"
        }
        active_des = mapping.get(lang_code, "designation_en")
        sql = f"""
            SELECT DISTINCT c.code
              FROM compositions c
              LEFT JOIN items_list i ON c.code = i.code
             WHERE c.quantity > 0
               AND (
                    LOWER(c.code) LIKE ?
                 OR LOWER(i.{active_des}) LIKE ?
                 OR (i.{active_des} IS NULL AND LOWER(i.designation_en) LIKE ?)
                 OR (i.{active_des} IS NULL AND i.designation_en IS NULL AND LOWER(i.designation) LIKE ?)
               )
        """
        params = [f"%{query.lower()}%"] * 4
        if self.scenario_var.get() != lang.t("stock_in.all_scenarios", "All Scenarios"):
            sql += " AND c.scenario_id = ?"
            params.append(self.reverse_scenario_map.get(self.scenario_var.get()))
        sql += " ORDER BY c.code"
        try:
            cur.execute(sql, params)
            rows = cur.fetchall()
            return [{'code': r['code'], 'description': get_active_designation(r['code'])} for r in rows]
        finally:
            cur.close()
            conn.close()

    def fetch_all_compositions(self):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        base = """
            SELECT c.code, c.quantity, s.name AS scenario_name, c.scenario_id
              FROM compositions c
              LEFT JOIN scenarios s ON c.scenario_id = s.scenario_id
             WHERE c.quantity > 0
        """
        params = []
        if self.scenario_var.get() != lang.t("stock_in.all_scenarios", "All Scenarios"):
            base += " AND c.scenario_id = ?"
            params.append(self.reverse_scenario_map.get(self.scenario_var.get()))
        base += " ORDER BY c.code, s.name"
        try:
            cur.execute(base, params)
            rows = cur.fetchall()
            out = []
            for r in rows:
                code = r['code']
                scenario_name = r['scenario_name']
                out.append({
                    'unique_id': f"{r['scenario_id']}/None/-----/{code}",
                    'unique_id_2': f"{r['scenario_id']}/None/-----/{code}",
                    'scenario_id': r['scenario_id'],
                    'scenario_name': scenario_name,
                    'kit_code': '-----',
                    'module_code': '-----',
                    'code': code,
                    'quantity': r['quantity'],
                    'description': get_active_designation(code)
                })
            return out
        finally:
            cur.close()
            conn.close()

    def fetch_compositions_for_code(self, code):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        sql = """
            SELECT c.code, c.quantity, s.name AS scenario_name, c.scenario_id
              FROM compositions c
              LEFT JOIN scenarios s ON c.scenario_id = s.scenario_id
             WHERE c.quantity > 0 AND c.code = ?
        """
        params = [code]
        if self.scenario_var.get() != lang.t("stock_in.all_scenarios", "All Scenarios"):
            sql += " AND c.scenario_id = ?"
            params.append(self.reverse_scenario_map.get(self.scenario_var.get()))
        sql += " ORDER BY c.code, s.name"
        try:
            cur.execute(sql, params)
            rows = cur.fetchall()
            out = []
            for r in rows:
                out.append({
                    'unique_id': f"{r['scenario_id']}/None/-----/{r['code']}",
                    'unique_id_2': f"{r['scenario_id']}/None/-----/{r['code']}",
                    'scenario_id': r['scenario_id'],
                    'scenario_name': r['scenario_name'],
                    'kit_code': '-----',
                    'module_code': '-----',
                    'code': r['code'],
                    'quantity': r['quantity'],
                    'description': get_active_designation(r['code'])
                })
            return out
        finally:
            cur.close()
            conn.close()

    def fetch_third_parties(self):
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM third_parties ORDER BY name")
            return [r[0] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()

    def fetch_end_users(self):
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM end_users ORDER BY name")
            return [r[0] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()

    # -------- UI Rendering -------- #
    def render_ui(self):
        # Clear existing widgets
        for w in self.winfo_children():
            try:
                w.destroy()
            except:
                pass

        bg = "#F0F4F8"
        self.configure(bg=bg)

        tk.Label(self, text=lang.t("stock_in.title", "Stock In"),
                 font=("Helvetica", 20, "bold"), bg=bg).pack(pady=10)

        # Top buttons
        # Top buttons (REMOVE role restriction)
        btn_frame = tk.Frame(self, bg=bg)
        btn_frame.pack(pady=5, fill="x")
    
        tk.Button(btn_frame, text=lang.t("stock_in.add_button", "Add Stock"),
                bg="#27AE60", fg="white", activebackground="#1E874B",
                command=self.save_all).pack(side="left", padx=5)
    
        tk.Button(btn_frame, text=lang.t("stock_in.clear_all", "Clear All"),
                bg="#7F8C8D", fg="white", activebackground="#666E70",
                command=self. clear_form).pack(side="left", padx=5)
    
        tk.Button(btn_frame, text=lang.t("stock_in.export", "Export"),
                bg="#2980B9", fg="white", activebackground="#1F6390",
                command=self.export_data).pack(side="left", padx=5)

        # Scenario filter
        filter_frame = tk.Frame(self, bg=bg); filter_frame.pack(pady=5, fill="x")
        tk.Label(filter_frame, text=lang.t("stock_in.scenario", "Scenario:"),
                 bg=bg).grid(row=0, column=0, padx=5, sticky="w")
        self.scenario_var = tk.StringVar(value=lang.t("stock_in.all_scenarios", "All Scenarios"))
        scenarios = list(self.scenario_map.values()) + [lang.t("stock_in.all_scenarios", "All Scenarios")]
        self.scenario_cb = ttk.Combobox(filter_frame, textvariable=self.scenario_var,
                                        values=scenarios, state="readonly", width=30)
        self.scenario_cb.grid(row=0, column=1, padx=5, pady=5)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)

        # IN Type + parties
        type_frame = tk.Frame(self, bg=bg); type_frame.pack(pady=5, fill="x")
        tk.Label(type_frame, text=lang.t("stock_in.in_type", "IN Type:"), bg=bg)\
            .grid(row=0, column=0, padx=5, sticky="w")
        self.trans_type_var = tk.StringVar()
        # DISPLAY LIST IS LOCALIZED; we convert back to canonical on save
        self.trans_type_cb = ttk.Combobox(
            type_frame,
            textvariable=self.trans_type_var,
            values=self.get_display_in_type_list(),
            state="readonly", width=30
        )
        self.trans_type_cb.grid(row=0, column=1, padx=5, pady=5)
        self.trans_type_cb.bind("<<ComboboxSelected>>", self.update_dropdown_visibility)
        
        
        tk.Label(type_frame, text=lang.t("stock_in.end_user", "End User:"), bg=bg)\
            .grid(row=0, column=2, padx=5, sticky="w")
        self.end_user_var = tk.StringVar()
        self.end_user_cb = ttk.Combobox(type_frame, textvariable=self.end_user_var, state="disabled", width=30)
        self.end_user_cb['values'] = self.fetch_end_users()
        self.end_user_cb.grid(row=0, column=3, padx=5, pady=5)

        tk.Label(type_frame, text=lang.t("stock_in.third_party", "Third Party:"), bg=bg)\
            .grid(row=0, column=4, padx=5, sticky="w")
        self.third_party_var = tk.StringVar()
        self.third_party_cb = ttk.Combobox(type_frame, textvariable=self.third_party_var, state="disabled", width=30)
        self.third_party_cb['values'] = self.fetch_third_parties()
        self.third_party_cb.grid(row=0, column=5, padx=5, pady=5)

        tk.Label(type_frame, text=lang.t("stock_in.remarks", "Remarks:"), bg=bg)\
            .grid(row=0, column=6, padx=5, sticky="w")
        self.remarks_entry = tk.Entry(type_frame, width=40, state="disabled")
        self.remarks_entry.grid(row=0, column=7, padx=5, pady=5)

        # Search
        search_frame = tk.Frame(self, bg=bg); search_frame.pack(pady=10, fill="x")
        tk.Label(search_frame, text=lang.t("stock_in.item_code", "Item Code"), bg=bg)\
            .grid(row=0, column=0, padx=5, sticky="w")
        self.code_entry = tk.Entry(search_frame)
        self.code_entry.grid(row=0, column=1, padx=5, pady=5)
        self.code_entry.bind("<KeyRelease>", self.search_items)
        self.code_entry.bind("<Return>", self.select_first_result)
        tk.Button(search_frame, text=lang.t("stock_in.clear_search", "Clear Search"),
                  bg="#7F8C8D", fg="white", activebackground="#666E70",
                  command=self.clear_search).grid(row=0, column=2, padx=5, pady=5)
        self.search_listbox = tk.Listbox(search_frame, height=5, width=80)
        self.search_listbox.grid(row=0, column=3, padx=5, pady=5, sticky="we")
        self.search_listbox.bind("<<ListboxSelect>>", self.fill_code_from_search)

        # Table
        tree_frame = tk.Frame(self, bg=bg); tree_frame.pack(expand=True, fill="both", pady=10)
        self.cols = ("code", "description", "scenario_name", "kit_code", "module_code",
                     "std_qty", "qty_needed", "qty_in", "expiry_date", "batch_no")
        self.tree = ttk.Treeview(tree_frame, columns=self.cols, show="headings", height=18)
        self.tree.tag_configure("light_red", background="#FF9999")

        self.headers = {
            "code": lang.t("stock_in.code", "Code"),
            "description": lang.t("stock_in.description", "Description"),
            "scenario_name": lang.t("stock_in.scenario_name", "Scenario Name"),
            "kit_code": lang.t("stock_in.kit_code", "Kit"),
            "module_code": lang.t("stock_in.module_code", "Module"),
            "std_qty": lang.t("stock_in.std_qty", "Std Qty"),
            "qty_needed": lang.t("stock_in.qty_needed", "Qty Needed"),
            "qty_in": lang.t("stock_in.qty_in", "Qty In"),
            "expiry_date": lang.t("stock_in.expiry_date", "Expiry Date"),
            "batch_no": lang.t("stock_in.batch_no", "Batch No")
        }
        self.widths = {
            "code": 100, "description": 335, "scenario_name": 120, "kit_code": 120,
            "module_code": 120, "std_qty": 80, "qty_needed": 90, "qty_in": 90,
            "expiry_date": 110, "batch_no": 120
        }
        for c in self.cols:
            self.tree.heading(c, text=self.headers[c])
            self.tree.column(c, width=self.widths[c], stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview); vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview); hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.context_menu = tk.Menu(self.tree, tearoff=0)
        self.context_menu.add_command(
            label=lang.t("stock_in.add_another_line", "Add another line for this item"),
            command=self.add_another_line
        )
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.start_edit)

        self.status_var = tk.StringVar(value=lang.t("stock_in.ready", "Ready"))
        tk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w",
                 bg=bg).pack(fill="x", pady=(5, 0))

        self.populate_table()

    # -------- Context Menu -------- #
    def show_context_menu(self, event):
        self.context_menu_row = self.tree.identify_row(event.y)
        if self.context_menu_row:
            self.tree.selection_set(self.context_menu_row)
            self.context_menu.post(event.x_root, event.y_root)

    def add_another_line(self):
        if not self.context_menu_row:
            return
        vals = self.tree.item(self.context_menu_row, "values")
        item_id = self.context_menu_row
        unique_id_2 = self.row_data[item_id]['unique_id_2']
        code = vals[0]
        description = vals[1]
        scenario_name = vals[2]
        kit_code = vals[3]
        module_code = vals[4]
        std_qty = int(vals[5]) if vals[5] and str(vals[5]).isdigit() else 0
    
        # ✅ Get scenario_id
        scenario_id = self.reverse_scenario_map.get(scenario_name)
    
        qty_needed = fetch_qty_needed(self, code, scenario_name, scenario_id, std_qty)
    
        # ✅ Use scenario_id in unique_id (correct!)
        new_unique_id = f"{scenario_id}/None/-----/{code}/None"
    
        # ✅ User input key uses scenario_name (for UI consistency)
        input_key = f"{scenario_name}/{code}/{std_qty}"
        user_input = self.user_inputs.get(input_key, {"qty_in": "", "expiry_date": "", "batch_no": ""})
    
        index = self.tree.index(self.context_menu_row) + 1
        new_item_id = self.tree.insert("", index, values=(
            code, description, scenario_name, kit_code, module_code,
            std_qty, qty_needed, user_input["qty_in"], user_input["expiry_date"], user_input["batch_no"]
        ))
    
        self.row_data[new_item_id] = {'unique_id': new_unique_id, 'unique_id_2': unique_id_2}
        self.status_var.set(lang.t("stock_in.added_line", "Added new line for {code}").format(code=code))
    
        # Update qty_needed for all matching items
        for item_id in self.tree.get_children():
            item_vals = self.tree.item(item_id, "values")
            if item_vals[0] == code and item_vals[2] == scenario_name:
                item_std_qty = int(item_vals[5]) if item_vals[5] and str(item_vals[5]).isdigit() else 0
                new_qty_needed = fetch_qty_needed(self, code, scenario_name, scenario_id, item_std_qty)
                self.tree.set(item_id, "qty_needed", new_qty_needed)


            
    def update_tree_columns(self, event=None):
        for c in self.cols:
            self.tree.heading(c, text=self.headers[c])
            self.tree.column(c, width=self.widths[c], stretch=True)
        self.tree.bind("<Double-1>", self.start_edit)

    def on_scenario_selected(self, event=None):
        self.populate_table()

    # ---------- Dropdown visibility based on canonical ----------
    def update_dropdown_visibility(self, event=None):
        """Enable/disable dropdowns based on canonical IN Type selection."""
        canonical_type = self.get_canonical_in_type()
        
        # Reset all to disabled
        self.end_user_cb. config(state="disabled")
        self.third_party_cb. config(state="disabled")
        self.remarks_entry.config(state="disabled")
        
        # Enable based on canonical English type
        if canonical_type == "Return from End User":
            self.end_user_cb.config(state="readonly")
        elif canonical_type == "In Donation":
            self.third_party_cb.config(state="readonly")
        elif canonical_type == "In Correction of Previous Transaction":
            self.remarks_entry.config(state="normal")

    def search_items(self, event=None):
        query = self.code_entry.get().strip()
        self.search_listbox.delete(0, tk.END)
        if not query:
            self.populate_table()
            return
        results = self.fetch_search_results(query)
        for r in results:
            display = f"{r['code']} - {r['description'] or lang.t('stock_in.no_description', 'No Description')}"
            self.search_listbox.insert(tk.END, display)
        self.status_var.set(
            lang.t("stock_in.found_items", "Found {count} items", count=self.search_listbox.size())
        )

    def clear_search(self):
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.populate_table()

    def select_first_result(self, event=None):
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self.fill_code_from_search()

    def fill_code_from_search(self, event=None):
        sel = self.search_listbox.curselection()
        if not sel:
            return
        code = self.search_listbox.get(sel[0]).split(" - ")[0]
        self.code_entry.delete(0, tk.END)
        self.code_entry.insert(0, code)
        self.search_listbox.delete(0, tk.END)
        self.save_user_inputs()
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.status_var.set(lang.t("stock_in.loading", "Loading..."))
        self.update_idletasks()
        data = self.fetch_compositions_for_code(code)
        if not data:
            self.status_var.set(
                lang.t("stock_in.no_items", "No items found for code {code}").format(code=code)
            )
        else:
            for row in data:
                unique_id = row['unique_id']
                unique_id_2 = row['unique_id_2']
                scenario_name = row['scenario_name']
                kit_code = row['kit_code']
                module_code = row['module_code']
                code_ = row['code']
                description = row['description']
                std_qty = int(row['quantity']) if row['quantity'] and str(row['quantity']).isdigit() else 0
                qty_needed = fetch_qty_needed(self, code_, scenario_name, row['scenario_id'], std_qty)
                input_key = f"{scenario_name}/{code_}/{std_qty}"
                user_input = self.user_inputs.get(input_key, {"qty_in": "", "expiry_date": "", "batch_no": ""})
                values = (
                    code_, description, scenario_name, kit_code, module_code,
                    std_qty, qty_needed, user_input["qty_in"], user_input["expiry_date"], user_input["batch_no"]
                )
                item_id = self.tree.insert("", "end", values=values)
                self.row_data[item_id] = {'unique_id': unique_id, 'unique_id_2': unique_id_2}
                if user_input["qty_in"] and user_input["qty_in"].isdigit() and check_expiry_required(code_):
                    if not self.validate_expiry_for_save(code_, user_input["qty_in"], user_input["expiry_date"]):
                        self.tree.item(item_id, tags=("light_red",))
            self.status_var.set(
                lang.t("stock_in.loaded", "Loaded {count} records for code {code}")
                .format(count=len(self.tree.get_children()), code=code)
            )

    def save_user_inputs(self):
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            scenario_name = vals[2]
            code = vals[0]
            std_qty = vals[5]
            key = f"{scenario_name}/{code}/{std_qty}"
            self.user_inputs[key] = {
                "qty_in": vals[7],
                "expiry_date": vals[8],
                "batch_no": vals[9]
            }

    def populate_table(self, event=None):
        self.save_user_inputs()
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        data = self.fetch_all_compositions()
        for row in data:
            unique_id = row['unique_id']
            unique_id_2 = row['unique_id_2']
            scenario_name = row['scenario_name']
            kit_code = row['kit_code']
            module_code = row['module_code']
            code = row['code']
            description = row['description']
            std_qty = int(row['quantity']) if row['quantity'] and str(row['quantity']).isdigit() else 0
            qty_needed = fetch_qty_needed(self, code, scenario_name, row['scenario_id'], std_qty)
            key = f"{scenario_name}/{code}/{std_qty}"
            user_input = self.user_inputs.get(key, {"qty_in": "", "expiry_date": "", "batch_no": ""})
            values = (
                code, description, scenario_name, kit_code, module_code,
                std_qty, qty_needed, user_input["qty_in"], user_input["expiry_date"], user_input["batch_no"]
            )
            item_id = self.tree.insert("", "end", values=values)
            self.row_data[item_id] = {'unique_id': unique_id, 'unique_id_2': unique_id_2}
            if user_input["qty_in"] and user_input["qty_in"].isdigit() and check_expiry_required(code):
                if not self.validate_expiry_for_save(code, user_input["qty_in"], user_input["expiry_date"]):
                    self.tree.item(item_id, tags=("light_red",))
        self.status_var.set(
            lang.t("stock_in.loaded", "Loaded {count} records").format(count=len(self.tree.get_children()))
        )

    def clear_form(self):
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.trans_type_var.set("")
        self.end_user_var.set("")
        self.third_party_var.set("")
        self.remarks_entry.delete(0, tk.END)
        self.end_user_cb.config(state="disabled")
        self.third_party_cb.config(state="disabled")
        self.remarks_entry.config(state="disabled")
        self.user_inputs = {}
        self.current_document_number = None
        self.populate_table()
        self.status_var.set(lang.t("stock_in.ready", "Ready"))

    def validate_expiry_for_save(self, code, qty_in, expiry_date):
        if not qty_in:
            return True
        if isinstance(qty_in, str) and not qty_in.isdigit():
            return True
        if not check_expiry_required(code):
            return True
        parsed_expiry = parse_expiry(expiry_date) if expiry_date else None
        current_date = datetime.now().date()
        if not parsed_expiry or parsed_expiry <= current_date:
            return False
        return True

    # -------- Editing (Cell) -------- #
    def start_edit(self, event):
        """
        Edit a cell inline.
        ✅ Single popup on error, clear cell, keep editor open
        ✅ Flexible date parsing for expiry column
        """
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not row_id or not col:
            return
        
        col_index = int(col.replace("#", "")) - 1
        
        # Only allow editing columns 7 (qty_in), 8 (expiry_date), 9 (batch_no)
        if col_index not in [7, 8, 9]:
            return
        
        # Destroy existing editor
        if self.editing_cell:
            try:
                self.editing_cell.destroy()
            except:
                pass
            self.editing_cell = None
        
        bbox = self.tree.bbox(row_id, col)
        if not bbox:
            return
        
        x, y, w, h = bbox
        value = self.tree.set(row_id, self.cols[col_index])
        code = self.tree.set(row_id, "code")
        scenario_name = self.tree.set(row_id, "scenario_name")
        scenario_id = self.reverse_scenario_map.get(scenario_name)
        
        try:
            std_qty = int(self.tree.set(row_id, "std_qty"))
        except:
            std_qty = 0
        
        expiry_required = check_expiry_required(code)

        # Create entry widget
        entry = tk.Entry(
            self.tree,
            font=("Helvetica", 10),
            background="#FFFBE0"  # Light yellow
        )
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, value)
        entry.focus()
        self.editing_cell = entry

        def cleanup():
            """Safely cleanup editor widget."""
            try:
                if entry.winfo_exists():
                    entry.destroy()
            except:
                pass
            self.editing_cell = None

        def save_edit(event=None):
            """Save edited value with validation."""
            if row_id not in self.tree.get_children():
                cleanup()
                return
            
            new_val = entry.get().strip()
            
            # ===== COLUMN 7: QTY IN =====
            if col_index == 7:
                if new_val and not new_val.isdigit():
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("stock_in.invalid_qty_detail",
                               "Invalid quantity for item: {code}\n\n"
                               "Quantity must be a positive integer.\n\n"
                               "Examples: 1, 10, 100",
                               code=code),
                        "error"
                    )
                    entry.focus()
                    return
            
            # ===== COLUMN 8: EXPIRY DATE =====
            elif col_index == 8:
                if new_val:
                    # ✅ Parse flexible date format
                    success, parsed_date, parse_error = self._parse_flexible_date(new_val)
                    
                    if not success:
                        # ✅ SINGLE POPUP: Clear cell, show error, keep editor open
                        entry.delete(0, tk.END)
                        entry.configure(background="#FFCCCC")
                        custom_popup(
                            self,
                            lang.t("dialog_titles.error", "Error"),
                            lang.t("stock_in.invalid_expiry_format_detail",
                                   "Invalid expiry date format for item: {code}\n\n"
                                   "Entered: '{date}'\n\n"
                                   "Supported formats:\n"
                                   "  • YYYY-MM-DD    (e.g., 2027-12-31)\n"
                                   "  • YYYY-MM       (e.g., 2027-12 → 2027-12-31)\n"
                                   "  • YYYY          (e.g., 2027 → 2027-12-31)\n"
                                   "  • MM/YYYY       (e.g., 12/2029 → 2029-12-31)\n"
                                   "  • DD-MM-YYYY    (e.g., 31-12-2027)\n"
                                   "  • DD/MM/YYYY    (e.g., 31/12/2027)",
                                   code=code,
                                   date=new_val),
                            "error"
                        )
                        entry.focus()
                        return
                    
                    # ✅ Validate date is in future
                    if not self._is_date_in_future(parsed_date):
                        entry.delete(0, tk.END)
                        entry.configure(background="#FFCCCC")
                        custom_popup(
                            self,
                            lang.t("dialog_titles.error", "Error"),
                            lang.t("stock_in.expiry_future_detail",
                                   "Expiry date must be in the future.\n\n"
                                   "Item: {code}\n"
                                   "Entered date: {date}\n"
                                   "Today: {today}\n\n"
                                   "Please enter a future date.",
                                   code=code,
                                   date=parsed_date,
                                   today=datetime.now().strftime("%Y-%m-%d")),
                            "error"
                        )
                        entry.focus()
                        return
                    
                    # Format as YYYY-MM-DD
                    new_val = parsed_date
            
            # ===== COLUMN 9: BATCH NO =====
            elif col_index == 9:
                if new_val and len(new_val) > 30:
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("stock_in.batch_no_length_detail",
                               "Batch number is too long.\n\n"
                               "Maximum length: 30 characters\n"
                               "Entered length: {length} characters\n\n"
                               "Please shorten the batch number.",
                               length=len(new_val)),
                        "error"
                    )
                    entry.focus()
                    return
            
            # ===== ALL VALIDATIONS PASSED - SAVE =====
            self.tree.set(row_id, self.cols[col_index], new_val)
            
            # Update unique_id_2 if expiry date changed
            if col_index == 8:
                unique_id_2_base = self.row_data[row_id]['unique_id_2'].rsplit('/', 1)[0]
                new_unique_id_2 = f"{unique_id_2_base}/{new_val or 'None'}"
                self.row_data[row_id]['unique_id_2'] = new_unique_id_2
                self.row_data[row_id]['unique_id'] = new_unique_id_2

            # Validate expiry requirement and apply tags
            qty_in = self.tree.set(row_id, "qty_in")
            expiry_date = self.tree.set(row_id, "expiry_date")
            
            if qty_in and qty_in.isdigit() and expiry_required and not self.validate_expiry_for_save(code, qty_in, expiry_date):
                self.tree.item(row_id, tags=("light_red",))
            else:
                self.tree.item(row_id, tags=())

            # Save user inputs
            input_key = f"{scenario_name}/{code}/{std_qty}"
            self.user_inputs[input_key] = {
                "qty_in": self.tree.set(row_id, "qty_in"),
                "expiry_date": self.tree.set(row_id, "expiry_date"),
                "batch_no": self.tree.set(row_id, "batch_no")
            }
            
            # Update qty_needed for all matching items
            for item_id in self.tree.get_children():
                item_vals = self.tree.item(item_id, "values")
                if item_vals[0] == code and item_vals[2] == scenario_name:
                    item_std_qty = int(item_vals[5]) if item_vals[5] and str(item_vals[5]).isdigit() else 0
                    new_qty_needed = fetch_qty_needed(self, code, scenario_name, scenario_id, item_std_qty)
                    self.tree.set(item_id, "qty_needed", new_qty_needed)
            
            cleanup()

        # Bind events
        entry.bind("<Return>", save_edit)
        entry.bind("<KP_Enter>", save_edit)
        entry.bind("<Tab>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", lambda e: cleanup())

    def start_edit_cell(self, row_id, col_index):
        """
        Edit a cell inline.
        ✅ Single popup on error, clear cell, keep editor open
        ✅ Detailed error messages like out_kit.py
        """
        col = f"#{col_index + 1}"
        bbox = self.tree.bbox(row_id, col)
        if not bbox:
            return False
        
        x, y, width, height = bbox
        value = self.tree.set(row_id, self.cols[col_index])
        code = self.tree.set(row_id, "code")
        scenario_name = self.tree.set(row_id, "scenario_name")
        scenario_id = self.reverse_scenario_map.get(scenario_name)
        
        try:
            std_qty = int(self.tree.set(row_id, "std_qty"))
        except:
            std_qty = 0
        
        expiry_required = check_expiry_required(code)

        # Destroy existing editor
        if self.editing_cell:
            self.editing_cell.destroy()
            self.editing_cell = None

        # Create entry widget
        entry = tk.Entry(
            self.tree,
            font=("Helvetica", 10),
            background="#FFFBE0"  # Light yellow background
        )
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus()
        self.editing_cell = entry

        def cleanup():
            """Safely cleanup editor widget."""
            try:
                if entry.winfo_exists():
                    entry.destroy()
            except:
                pass
            self.editing_cell = None

        def save_edit(event=None, row_id=row_id, col_index=col_index, code=code,
                      scenario_name=scenario_name, scenario_id=scenario_id,
                      std_qty=std_qty, expiry_required=expiry_required):
            """Save edited value with validation."""
            
            # Check if row still exists
            if row_id not in self.tree.get_children():
                cleanup()
                return
            
            new_val = entry.get().strip()
            
            # ===== COLUMN 7: QTY IN (Integer Validation) =====
            if col_index == 7:
                if new_val and not new_val.isdigit():
                    # ✅ Clear cell, red background, single popup
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("stock_in.invalid_qty_detail",
                               "Invalid quantity for item: {code}\n\n"
                               "Quantity must be a positive integer.\n\n"
                               "Examples: 1, 10, 100",
                               code=code),
                        "error"
                    )
                    entry.focus_set()
                    return
                
                if new_val and int(new_val) <= 0:
                    # ✅ Clear cell, red background, single popup
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("stock_in.qty_must_be_positive",
                               "Quantity must be greater than zero.\n\n"
                               "Entered: {qty}",
                               qty=new_val),
                        "error"
                    )
                    entry.focus_set()
                    return
            
            # ===== COLUMN 8: EXPIRY DATE (Date Validation) =====
            if col_index == 8:
                if new_val:
                    # Parse the date
                    parsed = parse_expiry(new_val)
                    
                    if not parsed:
                        # ✅ Clear cell, red background, single popup
                        entry.delete(0, tk.END)
                        entry.configure(background="#FFCCCC")
                        custom_popup(
                            self,
                            lang.t("dialog_titles.error", "Error"),
                            lang.t("stock_in.invalid_expiry_format_detail",
                                   "Invalid expiry date format for item: {code}\n\n"
                                   "Entered: '{date}'\n\n"
                                   "Supported formats:\n"
                                   "  • YYYY-MM-DD    (e.g., 2027-12-31)\n"
                                   "  • YYYY-MM       (e.g., 2027-12 → 2027-12-31)\n"
                                   "  • YYYY          (e.g., 2027 → 2027-12-31)\n"
                                   "  • MM/YYYY       (e.g., 12/2029 → 2029-12-31)\n"
                                   "  • DD-MM-YYYY    (e.g., 31-12-2027)\n"
                                   "  • DD/MM/YYYY    (e.g., 31/12/2027)",
                                   code=code,
                                   date=new_val),
                            "error"
                        )
                        entry.focus_set()
                        return
                    
                    # Check if in future
                    if parsed <= datetime.now().date():
                        # ✅ Clear cell, red background, single popup
                        entry.delete(0, tk.END)
                        entry.configure(background="#FFCCCC")
                        custom_popup(
                            self,
                            lang.t("dialog_titles.error", "Error"),
                            lang.t("stock_in.expiry_future_detail",
                                   "Expiry date must be in the future.\n\n"
                                   "Item: {code}\n"
                                   "Entered date: {date}\n"
                                   "Today: {today}\n\n"
                                   "Please enter a future date.",
                                   code=code,
                                   date=parsed.strftime("%Y-%m-%d"),
                                   today=datetime.now().strftime("%Y-%m-%d")),
                            "error"
                        )
                        entry.focus_set()
                        return
                    
                    # Format as YYYY-MM-DD
                    new_val = parsed.strftime("%Y-%m-%d")
            
            # ===== COLUMN 9: BATCH NO (Length Validation) =====
            if col_index == 9:
                if new_val and len(new_val) > 30:
                    # ✅ Clear cell, red background, single popup
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("stock_in.batch_no_length_detail",
                               "Batch number is too long.\n\n"
                               "Maximum length: 30 characters\n"
                               "Entered length: {length} characters\n\n"
                               "Please shorten the batch number.",
                               length=len(new_val)),
                        "error"
                    )
                    entry.focus_set()
                    return
            
            # ===== ALL VALIDATIONS PASSED - SAVE =====
            self.tree.set(row_id, self.cols[col_index], new_val)
            
            # Update unique_id_2 if expiry date changed
            if col_index == 8:
                unique_id_2_base = self.row_data[row_id]['unique_id_2'].rsplit('/', 1)[0]
                new_unique_id_2 = f"{unique_id_2_base}/{new_val or 'None'}"
                self.row_data[row_id]['unique_id_2'] = new_unique_id_2
                self.row_data[row_id]['unique_id'] = new_unique_id_2

            # Validate expiry requirement and apply tags
            qty_in = self.tree.set(row_id, "qty_in")
            expiry_date = self.tree.set(row_id, "expiry_date")
            
            if qty_in and qty_in.isdigit() and expiry_required and not self.validate_expiry_for_save(code, qty_in, expiry_date):
                self.tree.item(row_id, tags=("light_red",))
            else:
                self.tree.item(row_id, tags=())

            # Save user inputs
            scenario_name2 = self.tree.set(row_id, "scenario_name")
            code2 = self.tree.set(row_id, "code")
            std_qty2 = int(self.tree.set(row_id, "std_qty")) if self.tree.set(row_id, "std_qty") and self.tree.set(row_id, "std_qty").isdigit() else 0
            input_key = f"{scenario_name2}/{code2}/{std_qty2}"
            
            self.user_inputs[input_key] = {
                "qty_in": self.tree.set(row_id, "qty_in"),
                "expiry_date": self.tree.set(row_id, "expiry_date"),
                "batch_no": self.tree.set(row_id, "batch_no")
            }
            
            # Update qty_needed for all matching items
            for item_id in self.tree.get_children():
                item_vals = self.tree.item(item_id, "values")
                if item_vals[0] == code2 and item_vals[2] == scenario_name2:
                    item_std_qty = int(item_vals[5]) if item_vals[5] and str(item_vals[5]).isdigit() else 0
                    new_qty_needed = fetch_qty_needed(self, code2, scenario_name2, scenario_id, item_std_qty)
                    self.tree.set(item_id, "qty_needed", new_qty_needed)
            
            cleanup()
            
            # Move to next cell if Tab was pressed
            if event and event.keysym == "Tab":
                self.move_to_next_editable_cell(row_id, col_index)
        
        # Bind events
        entry.bind("<Return>", save_edit)
        entry.bind("<Tab>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", lambda e: cleanup())
        
        return True

        
    # -------- Save / Persist -------- #
    def save_all(self):
        if self.role.lower() not in ["admin", "manager"]:
            self.show_error("stock_in.no_permission", "Only admin or manager roles can save changes.")
            return
        rows = self.tree.get_children()
        if not rows:
            self.show_error("stock_in.no_rows", "No rows to save.")
            return

        # ALWAYS use canonical English for DB/logging
        ttype_canonical = self.get_canonical_in_type()
    
        # Guard: ensure the canonical value is one of the supported enums
        if not ttype_canonical or ttype_canonical not in self.IN_TYPE_CANONICAL:
            self.show_error("stock_in.no_in_type", "IN Type is required.")
            return

        remarks = self.remarks_entry.get().strip()
        end_user = self.end_user_var.get().strip()
        third_party = self.third_party_var.get().strip()

        # Validate specific IN types
        if ttype_canonical == "In Donation":
            if not third_party or third_party not in self.third_party_cb['values']:
                self.show_error("stock_in.invalid_third_party", "A valid Third Party must be selected for In Donation.")
                return
        if ttype_canonical == "Return from End User":
            if not end_user or end_user not in self.end_user_cb['values']:
                self.show_error("stock_in.invalid_end_user", "A valid End User must be selected for Return from End User.")
                return
        if ttype_canonical == "In Correction of Previous Transaction":
            if len(remarks) < 10 or len(remarks) > 300:
                self.show_error("stock_in.remarks_length", "Remarks must be between 10 and 300 characters for In Correction of Previous Transaction.")
                return

        # Generate Document Number from CANONICAL type
        doc_number = self.generate_document_number(ttype_canonical)
        self.status_var.set(
            lang.t("stock_in.generating_document", "Generated Document Number: {doc}").format(doc=doc_number)
        )

        invalid_items = []
        exported_rows = []
    
        for iid in rows:
            vals = self.tree.item(iid, "values")
            code = vals[0]
            description = vals[1]
            scenario_name = vals[2]
            kit_code = vals[3]
            module_code = vals[4]
            std_qty = vals[5]
            qty_needed = vals[6]
            qty_in = vals[7]
            expiry_date = vals[8]
            batch_no = vals[9]

            if not qty_in or not qty_in.isdigit():
                continue
        
            qty_in_int = int(qty_in)
            parsed_exp = parse_expiry(expiry_date)
            expiry_fmt = parsed_exp.strftime("%Y-%m-%d") if parsed_exp else None

            if not self.validate_expiry_for_save(code, qty_in_int, expiry_fmt):
                invalid_items.append(code)
                self.tree.item(iid, tags=("light_red",))
                continue

            # ✅ Convert scenario_name to scenario_id
            scenario_id = self.reverse_scenario_map.get(scenario_name)
        
            if not scenario_id:
                custom_popup(
                    self,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("stock_in.invalid_scenario", "Invalid scenario: {scenario}").format(scenario=scenario_name),
                    "error"
                )
                continue

            exp_part = expiry_fmt or "None"

            # ✅ Use scenario_id in unique_id (6-layer format)
            six_layer_unique_id = f"{scenario_id}/{kit_code if kit_code != '-----' else 'None'}/{module_code if module_code != '-----' else 'None'}/{code}/{std_qty}/{exp_part}"

            try:
                # ✅ Store canonical English + scenario_id in DB
                log_transaction(
                    unique_id=six_layer_unique_id,
                    code=code,
                    Description=description,
                    Expiry_date=expiry_fmt,
                    Batch_Number=batch_no,
                    Scenario=str(scenario_id),          # ✅ Use scenario_id (as string)
                    Kit=kit_code if kit_code != "-----" else None,
                    Module=module_code if module_code != "-----" else None,
                    Qty_IN=qty_in_int,
                    IN_Type=ttype_canonical,            # ✅ Canonical English
                    Third_Party=third_party if third_party else None,
                    End_User=end_user if end_user else None,
                    Remarks=remarks,
                    Movement_Type="stock_in",
                    document_number=doc_number
                )

                # ✅ Pass scenario_id to StockData
                StockData.add_or_update(
                    unique_id=six_layer_unique_id,
                    qty_in=qty_in_int,
                    qty_out=0,
                    exp_date=expiry_fmt
                )
                
            
                exported_rows.append({
                    'code': code,
                    'description': description,
                    'scenario_name': scenario_name,     # Keep for Excel export
                    'kit_code': kit_code,
                    'module_code': module_code,
                    'std_qty': std_qty,
                    'qty_needed': qty_needed,
                    'qty_in': qty_in_int,
                    'expiry_date': expiry_fmt,
                    'batch_no': batch_no
                })

            except Exception as e:
                custom_popup(
                    self,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("stock_in.save_failed", "Failed to save row: {error}").format(error=str(e)),
                    "error"
                )
                continue

        if invalid_items:
            self.show_error(
                "stock_in.invalid_expiry",
                "Valid future expiry dates are required for items: {items}",
                items=', '.join(invalid_items)
            )
            return

        self.show_info("stock_in.saved", "Stock IN saved successfully.")
    
        # Offer export
        if exported_rows and self.ask_yes_no("stock_in.save_excel_prompt", "Do you want to save the stock issuance to Excel?"):
            self.export_data(exported_rows)

        self.clear_form()

    # -------- Export -------- #
    def export_data(self, rows_to_export=None):
        try:
            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Show display (localized) in Excel for user friendliness
            ttype_display = self.trans_type_var.get() or lang.t("stock_in.unknown", "Unknown")
            safe_time = current_time.replace(":", "-").replace(" ", "_")
            file_name = f"IsEPREP_Stock-In_{ttype_display.replace(' ', '_')}_{safe_time}.xlsx"

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                title=lang.t("stock_in.save_excel", "Save Excel"),
                initialfile=file_name,
                initialdir=default_dir
            )
            if not path:
                self.status_var.set(lang.t("stock_in.export_cancelled", "Export cancelled"))
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = lang.t("stock_in.stock_in", "Stock In")

            project_name, project_code = fetch_project_details()
            doc_number = getattr(self, "current_document_number", None)

            # A1: Date + Doc number
            if doc_number:
                ws['A1'] = f"Date: {current_time}{' ' * 8}Document Number: {doc_number}"
            else:
                ws['A1'] = f"Date: {current_time}"
            ws['A1'].font = Font(name="Helvetica", size=10)
            ws['A1'].alignment = Alignment(horizontal="left")

            # A2: Title
            ws['A2'] = lang.t("stock_in.stock_in", "Stock In")
            ws['A2'].font = Font(name="Tahoma", size=14)
            ws['A2'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A2:J2')

            # A3: Project
            ws['A3'] = f"{project_name} - {project_code}"
            ws['A3'].font = Font(name="Tahoma", size=14)
            ws['A3'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A3:J3')

            # A4: IN Type (display)
            ws['A4'] = f"{lang.t('stock_in.in_type', 'In Type')}: {ttype_display}"
            ws['A4'].font = Font(name="Tahoma")
            ws['A4'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A4:J4')

            ws.append([])  # Blank row

            headers = [
                lang.t("stock_in.code", "Code"),
                lang.t("stock_in.description", "Description"),
                lang.t("stock_in.scenario_name", "Scenario Name"),
                lang.t("stock_in.kit_code", "Kit"),
                lang.t("stock_in.module_code", "Module"),
                lang.t("stock_in.std_qty", "Std Qty"),
                lang.t("stock_in.qty_needed", "Qty Needed"),
                lang.t("stock_in.qty_in", "Qty In"),
                lang.t("stock_in.expiry_date", "Expiry Date"),
                lang.t("stock_in.batch_no", "Batch No")
            ]
            ws.append(headers)

            kit_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            module_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

            rows_data = rows_to_export or [
                {k: v for k, v in zip(
                    ['code', 'description', 'scenario_name', 'kit_code', 'module_code', 'std_qty', 'qty_needed', 'qty_in', 'expiry_date', 'batch_no'],
                    self.tree.item(i)["values"]
                )}
                for i in self.tree.get_children()
                if self.tree.item(i)["values"][7]
            ]

            start_row = ws.max_row + 1
            for offset, r in enumerate(rows_data):
                row_idx = start_row + offset
                ws.append([
                    r['code'], r['description'], r['scenario_name'], r['kit_code'], r['module_code'],
                    r['std_qty'], r['qty_needed'], r['qty_in'], r['expiry_date'], r['batch_no']
                ])
                # Type highlight
                conn = connect_db()
                cur = conn.cursor()
                cur.execute("SELECT type FROM items_list WHERE code=?", (r['code'],))
                type_row = cur.fetchone()
                cur.close()
                if conn:
                    conn.close()
                item_type = type_row[0].upper() if type_row and type_row[0] else None
                if r['kit_code'] != '-----' and item_type == "KIT":
                    for col_cells in ws[f"A{row_idx}:J{row_idx}"]:
                        for cell in col_cells:
                            cell.fill = kit_fill
                elif r['module_code'] != '-----' and item_type == "MODULE":
                    for col_cells in ws[f"A{row_idx}:J{row_idx}"]:
                        for cell in col_cells:
                            cell.fill = module_fill

            ws.column_dimensions['A'].width = 100 / 7
            ws.column_dimensions['B'].width = 335 / 7
            ws.column_dimensions['C'].width = 120 / 7
            ws.column_dimensions['D'].width = 120 / 7
            ws.column_dimensions['E'].width = 120 / 7
            ws.column_dimensions['F'].width = 80 / 7
            ws.column_dimensions['G'].width = 90 / 7
            ws.column_dimensions['H'].width = 90 / 7
            ws.column_dimensions['I'].width = 110 / 7
            ws.column_dimensions['J'].width = 120 / 7

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1

            wb.save(path)
            wb.close()
            msg = lang.t("stock_in.export_success", "Export successful: {path}").format(path=path)
            custom_popup(self, lang.t("dialog_titles.success", "Success"), msg, "info")
            self.status_var.set(msg)
        except Exception as e:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_in.export_failed", "Export failed: {error}").format(error=str(e)), "error")
            self.status_var.set(lang.t("stock_in.export_error", "Export error: {error}").format(error=str(e)))

# ---------------------- Runner ---------------------- #
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock In")
    app = tk.Tk()
    app.role = "admin"
    StockIn(root, app, role="admin")
    root.mainloop()