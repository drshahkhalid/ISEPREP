"""
expiry_data.py  (Stock Expiry / Expiry Projection)  v1.5

Change Log v1.5:
  * All quantity figures (Expired, This Month, each projected month column, Total)
    are now strictly integers (rounded) — no decimals.
  * Projection logic now produces integer quantities: projected_qty =
        int(round(max(original_qty - (AMC * months_inclusive), 0)))
  * Row total and numeric totals are recomputed after integer conversion to stay consistent.
  * AMC values remain in decimal form (two decimals as before, e.g. 0.0 / 3.27).
  * Simple / Detailed toggle & Enter-to-refresh behavior retained from v1.4.

Notes:
  - Original qty sources (qty_in - qty_out) are assumed integers; intermediary
    projection math can create floats before rounding.
  - If AMC is large enough to exceed lot quantity, projected quantity becomes 0.
  - Rounding uses round() then int() to follow normal .5-up rounding.

Previous v1.4 features retained:
  * Simple / Detailed mode (Simple: Code, Description, Already Expired, This Month, Total)
  * Export always uses full detailed set
  * Dropdown kit/module filters, AMC=0 override
  * Robust schema fallbacks & scenario mapping
"""

import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
from datetime import date, datetime
from calendar import monthrange
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from db import connect_db
from language_manager import lang
from popup_utils import custom_popup
from manage_items import get_item_description, detect_type

# ---------------- Theme Constants ----------------
BG_MAIN        = "#F0F4F8"
BG_PANEL       = "#FFFFFF"
COLOR_PRIMARY  = "#2C3E50"
COLOR_ACCENT   = "#2563EB"
COLOR_BORDER   = "#D0D7DE"
ROW_ALT_COLOR  = "#F7FAFC"
ROW_NORM_COLOR = "#FFFFFF"
BTN_EXPORT     = "#2980B9"
BTN_REFRESH    = "#2563EB"
BTN_CLEAR      = "#7F8C8D"
BTN_TOGGLE     = "#8E44AD"

# ---------------- Helper Functions ----------------
def months_between_inclusive(start_ym, end_ym):
    (y1, m1), (y2, m2) = start_ym, end_ym
    return (y2 - y1) * 12 + (m2 - m1) + 1

def add_months(y, m, n):
    total = y * 12 + (m - 1) + n
    ny = total // 12
    nm = (total % 12) + 1
    return ny, nm

def ym_label(y, m):
    return date(y, m, 1).strftime("%b-%Y")


# ---------------- Core Calculator ----------------
class ExpiryDataCalculator:
    def __init__(self,
                 scenario_name_map,
                 management_mode_filter="All",
                 scenario_filter="All",
                 kit_number_filter="All",
                 module_number_filter="All",
                 item_search="",
                 type_filter="All",
                 expiry_period_months=12,
                 amc_months=6):
        self.scenario_name_map = scenario_name_map
        self.mgmt_mode = (management_mode_filter or "All").strip()
        self.scenario_filter = (scenario_filter or "All").strip()
        self.kit_number_filter = (kit_number_filter or "All").strip()
        self.module_number_filter = (module_number_filter or "All").strip()
        self.item_search = (item_search or "").strip()
        self.type_filter = (type_filter or "All").strip()
        self.expiry_period = max(1, min(int(expiry_period_months or 1), 99))
        self.amc_months = max(0, min(int(amc_months or 0), 99))  # allow 0
        today = date.today()
        self.current_year = today.year
        self.current_month = today.month
        self.today = today

    def _table_columns_lower(self, cursor, table):
        try:
            cursor.execute(f"PRAGMA table_info({table})")
            return [r[1].lower() for r in cursor.fetchall()]
        except sqlite3.Error:
            return []

    def _stock_data_rows(self, conn):
        cur = conn.cursor()
        cols = self._table_columns_lower(cur, "stock_data")
        has_code = "code" in cols
        has_mgmt = "management_mode" in cols
        has_scenario = "scenario" in cols
        has_comments = "comments" in cols

        base_cols = ["unique_id", "qty_in", "qty_out", "exp_date", "kit_number", "module_number", "updated_at"]
        if has_code: base_cols.append("code")
        if has_mgmt: base_cols.append("management_mode")
        if has_scenario: base_cols.append("scenario")
        if has_comments: base_cols.append("comments")

        select_sql = ", ".join(base_cols)
        data = []
        try:
            cur.execute(f"SELECT {select_sql} FROM stock_data")
            for row in cur.fetchall():
                rec = dict(zip(base_cols, row))
                rec.setdefault("code", None)
                rec.setdefault("management_mode", None)
                rec.setdefault("scenario", None)
                rec.setdefault("comments", None)
                if not rec["code"]:
                    rec["code"] = self._derive_code_from_unique_id(rec.get("unique_id"))
                raw_scen = rec.get("scenario")
                if not raw_scen:
                    raw_scen = self._derive_scenario_from_unique_id(rec.get("unique_id"))
                if raw_scen and raw_scen.isdigit() and raw_scen in self.scenario_name_map:
                    rec["scenario"] = self.scenario_name_map[raw_scen]
                else:
                    rec["scenario"] = raw_scen
                data.append(rec)
        except sqlite3.Error:
            pass
        finally:
            cur.close()
        return data

    def _derive_code_from_unique_id(self, unique_id):
        if not unique_id or "/" not in unique_id:
            return None
        parts = unique_id.split("/")
        if len(parts) < 4:
            return None
        kit_part = parts[1] if len(parts) > 1 else None
        module_part = parts[2] if len(parts) > 2 else None
        item_part = parts[3] if len(parts) > 3 else None
        for c in (item_part, module_part, kit_part):
            if c and c.lower() != "none":
                return c
        return None

    def _derive_scenario_from_unique_id(self, unique_id):
        if not unique_id or "/" not in unique_id:
            return None
        scen = unique_id.split("/")[0]
        if scen and scen.isdigit() and scen in self.scenario_name_map:
            return self.scenario_name_map[scen]
        return scen

    def _items_type_map(self, conn):
        cur = conn.cursor()
        mapping = {}
        try:
            cur.execute("SELECT code, type FROM items_list")
            for c, t in cur.fetchall():
                mapping[c] = t
        except sqlite3.Error:
            pass
        finally:
            cur.close()
        return mapping

    def _amc_map(self, conn):
        if self.amc_months == 0:
            return {}
        cur = conn.cursor()
        mapping = {}
        start_year, start_month = add_months(self.current_year, self.current_month, -(self.amc_months - 1))
        start_date = date(start_year, start_month, 1)
        end_year, end_month = self.current_year, self.current_month
        end_day = monthrange(end_year, end_month)[1]
        end_date = date(end_year, end_month, end_day)
        try:
            tx_cols = self._table_columns_lower(cur, "stock_transactions")
            needed = {"date", "qty_out", "out_type", "code"}
            if not needed.issubset(set(tx_cols)):
                return {}
            sql = """
                SELECT code, SUM(COALESCE(Qty_Out,0))
                FROM stock_transactions
                WHERE Out_Type IS NOT NULL
                  AND LOWER(Out_Type) = LOWER('Out MSF')
                  AND Date >= ? AND Date <= ?
                GROUP BY code
            """
            cur.execute(sql, (start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))
            for code, total_out in cur.fetchall():
                if not code:
                    continue
                mapping[code] = (total_out or 0) / self.amc_months
        except sqlite3.Error:
            pass
        finally:
            cur.close()
        return mapping

    def compute(self):
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed.")
        try:
            stock_rows = self._stock_data_rows(conn)
            type_map = self._items_type_map(conn)
            amc_map = self._amc_map(conn)

            per_code_expiry = {}
            code_comments = {}

            def norm(v):
                return (v or "").strip().lower()

            for r in stock_rows:
                code = (r.get("code") or "").strip()
                if not code:
                    continue
                qty_in = r.get("qty_in") or 0
                qty_out = r.get("qty_out") or 0
                final_qty = qty_in - qty_out
                if final_qty <= 0:
                    continue

                mgmt_mode = r.get("management_mode")
                scenario_val = r.get("scenario")
                kit_num = (r.get("kit_number") or "")
                module_num = (r.get("module_number") or "")
                comments_val = (r.get("comments") or "").strip()
                exp_date = r.get("exp_date")

                if self.mgmt_mode.lower() != "all":
                    if norm(mgmt_mode) != self.mgmt_mode.lower():
                        continue
                if self.scenario_filter.lower() != "all":
                    if not scenario_val or scenario_val.lower() != self.scenario_filter.lower():
                        continue
                if self.kit_number_filter.lower() != "all":
                    if kit_num.lower() != self.kit_number_filter.lower():
                        continue
                if self.module_number_filter.lower() != "all":
                    if module_num.lower() != self.module_number_filter.lower():
                        continue

                code_type = type_map.get(code)
                if not code_type:
                    code_type = detect_type(code, "")

                if self.type_filter.lower() != "all":
                    if not code_type or code_type.lower() != self.type_filter.lower():
                        continue

                if self.item_search:
                    if (code_type or "").lower() != "item":
                        continue
                    desc = get_item_description(code)
                    if self.item_search.lower() not in code.lower() and self.item_search.lower() not in desc.lower():
                        continue

                exp_year = exp_month = None
                if exp_date and isinstance(exp_date, str) and len(exp_date) >= 7:
                    parts = exp_date.split("-")
                    try:
                        if len(parts) >= 2:
                            ey = int(parts[0]); em = int(parts[1])
                            if 1 <= em <= 12:
                                exp_year, exp_month = ey, em
                    except ValueError:
                        exp_year = exp_month = None

                lot_key = (code, exp_year, exp_month)
                per_code_expiry.setdefault(lot_key, 0)
                per_code_expiry[lot_key] += int(final_qty)  # ensure integer accumulation

                if comments_val:
                    code_comments.setdefault(code, set()).add(comments_val)

            current_ym = (self.current_year, self.current_month)
            horizon_end = add_months(self.current_year, self.current_month, self.expiry_period - 1)

            def within_horizon(y, m):
                if y is None or m is None:
                    return False
                if (y < self.current_year) or (y == self.current_year and m < self.current_month):
                    return False
                hy, hm = horizon_end
                if (y > hy) or (y == hy and m > hm):
                    return False
                return True

            horizon_months = set()
            for (_code, y, m), qty in per_code_expiry.items():
                if y is None or m is None:
                    continue
                if within_horizon(y, m):
                    horizon_months.add((y, m))
            horizon_months_sorted = sorted(horizon_months, key=lambda x: (x[0], x[1]))

            static_cols = [
                ("code",          lang.t("expiry_data.code", "Code")),
                ("description",   lang.t("expiry_data.description", "Description")),
                ("comments",      lang.t("expiry_data.comments", "Comments")),
                ("amc",           lang.t("expiry_data.amc", "AMC")),
                ("expired_qty",   lang.t("expiry_data.expired_qty", "Already Expired")),
                ("this_month_qty",lang.t("expiry_data.this_month", "This Month"))
            ]
            dynamic_cols = []
            for (y, m) in horizon_months_sorted:
                if (y, m) == current_ym:
                    continue
                key = f"proj_{y}_{m:02d}"
                dynamic_cols.append((key, ym_label(y, m)))

            per_code_rows = {}
            for (code, y, m), qty in per_code_expiry.items():
                row = per_code_rows.setdefault(code, {
                    "code": code,
                    "description": get_item_description(code),
                    "comments": ", ".join(sorted(code_comments.get(code, []))) if code in code_comments else "",
                    "amc": 0.0,
                    "expired_qty": 0,
                    "this_month_qty": 0
                })
                if y is None or m is None:
                    continue
                if (y, m) < current_ym:
                    row["expired_qty"] = int(row.get("expired_qty", 0)) + int(qty)
                elif (y, m) == current_ym:
                    row["this_month_qty"] = int(row.get("this_month_qty", 0)) + int(qty)
                else:
                    k = f"proj_{y}_{m:02d}"
                    row[k] = int(row.get(k, 0)) + int(qty)

            for code, row in per_code_rows.items():
                # Keep AMC as decimal (two decimals)
                row["amc"] = round(amc_map.get(code, 0.0), 2)

            # Projection -> integer quantities
            for code, row in per_code_rows.items():
                amc_val = row["amc"] if self.amc_months > 0 else 0.0
                for (y, m) in horizon_months_sorted:
                    if (y, m) == current_ym:
                        continue
                    k = f"proj_{y}_{m:02d}"
                    base_qty = float(row.get(k, 0))
                    if base_qty <= 0:
                        row[k] = 0
                        continue
                    months_inc = months_between_inclusive(current_ym, (y, m))
                    projected = max(base_qty - (amc_val * months_inc), 0)
                    row[k] = int(round(projected))

            dynamic_keys = [k for k, _ in dynamic_cols]
            for row in per_code_rows.values():
                total_val = int(row.get("expired_qty", 0)) + int(row.get("this_month_qty", 0)) + sum(int(row.get(k, 0)) for k in dynamic_keys)
                row["row_total"] = int(total_val)

            all_columns = static_cols + dynamic_cols + [("row_total", lang.t("expiry_data.total", "Total"))]

            rows_list = list(per_code_rows.values())
            rows_list.sort(key=lambda r: r["code"])

            totals = {}
            numeric_keys = {k for k, _ in all_columns if k not in ("code", "description", "comments", "amc")}
            for k in numeric_keys:
                totals[k] = int(sum(int(r.get(k, 0)) for r in rows_list))
            # AMC total (optional aggregate, keep decimal if needed)
            totals["amc"] = round(sum(r.get("amc", 0.0) for r in rows_list), 2)

            return all_columns, rows_list, totals, [c[0] for c in dynamic_cols]
        finally:
            try:
                conn.close()
            except Exception:
                pass


# ---------------- UI View ----------------
class ExpiryDataView(tk.Frame):
    SIMPLE_BASE = ["code", "description", "expired_qty", "this_month_qty", "row_total"]

    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_MAIN)
        self.app = app
        self.tree = None
        self.columns_meta = []
        self.rows_cache = []
        self.totals_cache = {}
        self.future_month_keys = []
        self.scenario_map = self._load_scenario_map()
        self.simple_mode = False  # start Detailed
        self._build_ui()
        self.populate_kit_module_lists()
        self.refresh()

    def t(self, key, fallback=None, **kwargs):
        return lang.t(f"expiry_data.{key}", fallback=fallback if fallback else key, **kwargs)

    def _load_scenario_map(self):
        conn = connect_db()
        if conn is None:
            return {}
        cur = conn.cursor()
        mapping = {}
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios")
            for sid, name in cur.fetchall():
                mapping[str(sid)] = name
        except sqlite3.Error:
            pass
        finally:
            cur.close(); conn.close()
        return mapping

    def _build_ui(self):
        header_frame = tk.Frame(self, bg=BG_MAIN)
        header_frame.pack(fill="x", padx=12, pady=(12,4))
        tk.Label(header_frame,
                 text=self.t("title", "Stock Expiry / Projection"),
                 font=("Helvetica", 20, "bold"),
                 bg=BG_MAIN, fg=COLOR_PRIMARY,
                 anchor="w").pack(side="left", fill="x", expand=True)

        self.toggle_btn = tk.Button(header_frame,
                                    text=self.t("toggle_detailed", "Detailed"),
                                    bg=BTN_TOGGLE, fg="#FFFFFF",
                                    padx=14, pady=6, relief="flat",
                                    command=self.toggle_mode)
        self.toggle_btn.pack(side="right")

        filters = tk.Frame(self, bg=BG_MAIN)
        filters.pack(fill="x", padx=12, pady=(0, 10))

        r1 = tk.Frame(filters, bg=BG_MAIN); r1.pack(fill="x", pady=2)

        tk.Label(r1, text=self.t("management_mode","Management Mode"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.mgmt_mode_var = tk.StringVar(value="All")
        self.mgmt_mode_cb = ttk.Combobox(r1, textvariable=self.mgmt_mode_var, state="readonly", width=14,
                                         values=["All","on-shelf","in-box"])
        self.mgmt_mode_cb.grid(row=0, column=1, padx=(0,14))

        tk.Label(r1, text=self.t("scenario","Scenario"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.scenario_var = tk.StringVar(value="All")
        self.scenario_cb = ttk.Combobox(r1, textvariable=self.scenario_var, state="readonly", width=20)
        self.scenario_cb.grid(row=0, column=3, padx=(0,14))

        tk.Label(r1, text=self.t("kit_number","Kit Number"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.kit_var = tk.StringVar(value="All")
        self.kit_cb = ttk.Combobox(r1, textvariable=self.kit_var, state="readonly", width=16, values=["All"])
        self.kit_cb.grid(row=0, column=5, padx=(0,14))

        tk.Label(r1, text=self.t("module_number","Module Number"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.module_var = tk.StringVar(value="All")
        self.module_cb = ttk.Combobox(r1, textvariable=self.module_var, state="readonly", width=16, values=["All"])
        self.module_cb.grid(row=0, column=7, padx=(0,14))

        r2 = tk.Frame(filters, bg=BG_MAIN); r2.pack(fill="x", pady=2)

        tk.Label(r2, text=self.t("item_search","Item Search"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.item_search_var = tk.StringVar()
        item_entry = tk.Entry(r2, textvariable=self.item_search_var, width=22)
        item_entry.grid(row=0, column=1, padx=(0,14))
        item_entry.bind("<Return>", lambda e: self.refresh())

        tk.Label(r2, text=self.t("type","Type"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.type_var = tk.StringVar(value="All")
        self.type_cb = ttk.Combobox(r2, textvariable=self.type_var, state="readonly", width=12,
                                    values=["All","Kit","Module","Item"])
        self.type_cb.grid(row=0, column=3, padx=(0,14))

        tk.Label(r2, text=self.t("expiry_period","Expiry Period (Months)"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.expiry_period_var = tk.StringVar(value="12")
        tk.Entry(r2, textvariable=self.expiry_period_var, width=6,
                 validate="key", validatecommand=(self.register(self._val_1_99), "%P"))\
            .grid(row=0, column=5, padx=(0,14))

        tk.Label(r2, text=self.t("amc_months","AMC Months (0=No Consumption)"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.amc_months_var = tk.StringVar(value="6")
        tk.Entry(r2, textvariable=self.amc_months_var, width=6,
                 validate="key", validatecommand=(self.register(self._val_0_99), "%P"))\
            .grid(row=0, column=7, padx=(0,14))

        btn_row = tk.Frame(filters, bg=BG_MAIN); btn_row.pack(fill="x", pady=(6,4))
        tk.Button(btn_row, text=self.t("refresh","Refresh"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.refresh).pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("clear","Clear"),
                  bg=BTN_CLEAR, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.clear_filters).pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("export","Export"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.export_excel).pack(side="left", padx=(0,6))

        self.status_var = tk.StringVar(value=self.t("ready","Ready"))
        tk.Label(self, textvariable=self.status_var, anchor="w",
                 bg=BG_MAIN, fg=COLOR_PRIMARY, relief="sunken")\
            .pack(fill="x", padx=12, pady=(0,8))

        outer = tk.Frame(self, bg=COLOR_BORDER, bd=1, relief="solid")
        outer.pack(fill="both", expand=True, padx=12, pady=(0,12))
        self.tree = ttk.Treeview(outer, columns=(), show="headings", height=22)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vsb.set)

        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("Treeview",
                        background=BG_PANEL,
                        fieldbackground=BG_PANEL,
                        foreground=COLOR_PRIMARY,
                        rowheight=24,
                        font=("Helvetica",10))
        style.configure("Treeview.Heading",
                        background="#E5E8EB",
                        foreground=COLOR_PRIMARY,
                        font=("Helvetica",11,"bold"))
        self.tree.tag_configure("norm", background=ROW_NORM_COLOR)
        self.tree.tag_configure("alt", background=ROW_ALT_COLOR)

    def _val_1_99(self, P):
        if P == "": return True
        return P.isdigit() and 1 <= int(P) <= 99

    def _val_0_99(self, P):
        if P == "": return True
        return P.isdigit() and 0 <= int(P) <= 99

    def load_scenarios(self):
        conn = connect_db()
        if conn is None:
            self.scenario_cb['values'] = ["All"]
            return
        cur = conn.cursor()
        names = []
        try:
            cur.execute("SELECT name FROM scenarios ORDER BY name")
            names = [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            names = []
        finally:
            cur.close(); conn.close()
        self.scenario_cb['values'] = ["All"] + names
        if self.scenario_var.get() not in self.scenario_cb['values']:
            self.scenario_var.set("All")

    def populate_kit_module_lists(self):
        conn = connect_db()
        kits, modules = set(), set()
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("PRAGMA table_info(stock_data)")
                cols = [r[1].lower() for r in cur.fetchall()]
                if "kit_number" in cols:
                    cur.execute("""SELECT DISTINCT kit_number FROM stock_data
                                   WHERE kit_number IS NOT NULL AND kit_number!='None'
                                   ORDER BY kit_number""")
                    for (k,) in cur.fetchall():
                        if k: kits.add(str(k))
                if "module_number" in cols:
                    cur.execute("""SELECT DISTINCT module_number FROM stock_data
                                   WHERE module_number IS NOT NULL AND module_number!='None'
                                   ORDER BY module_number""")
                    for (m,) in cur.fetchall():
                        if m: modules.add(str(m))
            finally:
                cur.close(); conn.close()
        kit_vals = ["All"] + sorted(kits, key=lambda x: (len(x), x))
        mod_vals = ["All"] + sorted(modules, key=lambda x: (len(x), x))
        self.kit_cb['values'] = kit_vals
        self.module_cb['values'] = mod_vals
        if self.kit_var.get() not in kit_vals:
            self.kit_var.set("All")
        if self.module_var.get() not in mod_vals:
            self.module_var.set("All")

    def refresh(self):
        self.load_scenarios()
        self.populate_kit_module_lists()
        try:
            expiry_period = int(self.expiry_period_var.get() or "12")
        except ValueError:
            expiry_period = 12; self.expiry_period_var.set("12")
        try:
            amc_months = int(self.amc_months_var.get() or "6")
        except ValueError:
            amc_months = 6; self.amc_months_var.set("6")

        calc = ExpiryDataCalculator(
            scenario_name_map=self.scenario_map,
            management_mode_filter=self.mgmt_mode_var.get(),
            scenario_filter=self.scenario_var.get(),
            kit_number_filter=self.kit_var.get(),
            module_number_filter=self.module_var.get(),
            item_search=self.item_search_var.get(),
            type_filter=self.type_var.get(),
            expiry_period_months=expiry_period,
            amc_months=amc_months
        )
        try:
            cols, rows, totals, future_keys = calc.compute()
        except ValueError as e:
            custom_popup(self, self.t("error","Error"), str(e), "error")
            return
        self.columns_meta = cols
        self.rows_cache = rows
        self.totals_cache = totals
        self.future_month_keys = future_keys
        self._populate_tree()
        self.status_var.set(self.t("loaded","Loaded {n} codes").format(n=len(rows)))

    def current_display_columns(self):
        full_order = [cid for cid, _ in self.columns_meta]
        if not self.simple_mode:
            return full_order
        base_set = set(self.SIMPLE_BASE)
        return [c for c in full_order if c in base_set]

    def _populate_tree(self):
        display_cols = self.current_display_columns()
        label_map = {cid: lbl for cid, lbl in self.columns_meta}

        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = display_cols

        for cid in display_cols:
            width = 140
            if cid == "description": width = 340
            elif cid == "comments": width = 220
            elif cid == "code": width = 160
            elif cid == "amc": width = 80
            elif cid.endswith("_qty") or cid.startswith("proj_") or cid == "row_total": width = 110
            self.tree.heading(cid, text=label_map.get(cid, cid))
            self.tree.column(cid, width=width, anchor="w")

        for idx, r in enumerate(self.rows_cache):
            # Ensure integer display for quantity columns (already int but enforce)
            row_display = {}
            for k, v in r.items():
                if k == "amc":
                    row_display[k] = v  # keep decimal
                elif isinstance(v, (int, float)) and (k.endswith("_qty") or k.startswith("proj_") or k == "row_total"):
                    row_display[k] = int(v)
                else:
                    row_display[k] = v
            vals = [row_display.get(c, "") for c in display_cols]
            tag = "alt" if idx % 2 else "norm"
            self.tree.insert("", "end", values=vals, tags=(tag,))

    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        if self.simple_mode:
            self.toggle_btn.config(text=self.t("toggle_simple","Simple"))
            self.status_var.set(self.t("mode_simple_status","Simple mode"))
        else:
            self.toggle_btn.config(text=self.t("toggle_detailed","Detailed"))
            self.status_var.set(self.t("mode_detailed_status","Detailed mode"))
        self._populate_tree()

    def clear_filters(self):
        self.mgmt_mode_var.set("All")
        self.scenario_var.set("All")
        self.kit_var.set("All")
        self.module_var.set("All")
        self.item_search_var.set("")
        self.type_var.set("All")
        self.expiry_period_var.set("12")
        self.amc_months_var.set("6")
        self.refresh()

    def export_excel(self):
        if not self.rows_cache:
            custom_popup(self, self.t("no_data","No Data"),
                         self.t("nothing_export","Nothing to export."), "warning")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files","*.xlsx")],
            title=self.t("export_dialog","Save Expiry Projection"),
            initialfile=f"Expiry_Projection_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ExpiryProjection"
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([self.t("generated","Generated"), now_str])
            ws.append([self.t("filters","Filters Used")])
            ws.append(["Management Mode", self.mgmt_mode_var.get(),
                       "Scenario", self.scenario_var.get(),
                       "Type", self.type_var.get()])
            ws.append(["Expiry Period", self.expiry_period_var.get(),
                       "AMC Months", self.amc_months_var.get(),
                       "Mode", "Simple" if self.simple_mode else "Detailed"])
            ws.append(["Kit Number", self.kit_var.get(),
                       "Module Number", self.module_var.get(),
                       "Item Search", self.item_search_var.get()])
            ws.append([])

            header = [lbl for _cid, lbl in self.columns_meta]
            ws.append(header)

            kit_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            module_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            for r in self.rows_cache:
                out_row = []
                for cid, _lbl in self.columns_meta:
                    val = r.get(cid, "")
                    if cid != "amc" and isinstance(val, (int, float)) and (cid.endswith("_qty") or cid.startswith("proj_") or cid == "row_total"):
                        val = int(val)
                    out_row.append(val)
                ws.append(out_row)
                code = r.get("code","")
                dtype = detect_type(code, r.get("description",""))
                if dtype.upper() == "KIT":
                    for c in ws[ws.max_row]: c.fill = kit_fill
                elif dtype.upper() == "MODULE":
                    for c in ws[ws.max_row]: c.fill = module_fill

            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(v))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            wb.save(file_path)
            custom_popup(self, self.t("success","Success"),
                         self.t("export_ok","Export completed: {f}").format(f=file_path),
                         "info")
        except Exception as e:
            custom_popup(self, self.t("error","Error"),
                         self.t("export_fail","Export failed: {err}").format(err=str(e)),
                         "error")


class StockExpiry(ExpiryDataView):
    pass

__all__ = ["ExpiryDataView", "StockExpiry"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Expiry / Projection v1.5")
    class Dummy: pass
    d = Dummy()
    StockExpiry(root, d).pack(fill="both", expand=True)
    root.geometry("1400x760")
    root.mainloop()