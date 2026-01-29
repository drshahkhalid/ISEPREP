from py_compile import main
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog
from datetime import datetime
import sqlite3
import re
from calendar import monthrange
from collections import defaultdict
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import messagebox as mb
from popup_utils import custom_popup, custom_askyesno, custom_dialog
import os

from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang
from popup_utils import custom_popup

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")

# ---------------------------------------------------------------
# Dialog centering helper
# ---------------------------------------------------------------
def _center_child(win: tk.Toplevel, parent: tk.Widget):
    win.update_idletasks()
    try:
        if parent and parent.winfo_exists():
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            w, h = win.winfo_width(), win.winfo_height()
            x = px + (pw // 2) - (w // 2)
            y = py + (ph // 2) - (h // 2)
        else:
            sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
            w, h = win.winfo_width(), win.winfo_height()
            x = (sw // 2) - (w // 2)
            y = (sh // 2) - (h // 2)
        x = max(x, 0)
        y = max(y, 0)
        win.geometry(f"+{x}+{y}")
    except Exception:
        pass

# Center tkinter.simpledialog (monkey patch once)
import tkinter.simpledialog as _sd
if not hasattr(_sd, "_CenteredQueryString"):
    _OrigQS = _sd._QueryString
    class _CenteredQueryString(_OrigQS):
        def body(self, master):
            res = super().body(master)
            _center_child(self, self.parent)
            return res
    _sd._QueryString = _CenteredQueryString

# ---------------------------------------------------------------
# Parsing / DB helper functions
# ---------------------------------------------------------------
def parse_expiry(date_str: str) -> str:
    if not date_str:
        return None
    ds = str(date_str).strip()
    if ds.lower() == "none":
        return None
    ds = re.sub(r'[-.\s]+', '/', ds)
    patterns = [
        r'^(\d{1,2})/(\d{1,2})/(\d{4})$',
        r'^(\d{4})/(\d{1,2})/(\d{1,2})$',
        r'^(\d{1,2})/(\d{4})$',
        r'^(\d{4})/(\d{1,2})$'
    ]
    for idx, pat in enumerate(patterns):
        m = re.match(pat, ds)
        if not m:
            continue
        g = m.groups()
        try:
            if idx == 0:
                d, mth, y = map(int, g)
                return datetime(y, mth, d).strftime("%Y-%m-%d")
            elif idx == 1:
                y, mth, d = map(int, g)
                return datetime(y, mth, d).strftime("%Y-%m-%d")
            elif idx == 2:
                mth, y = map(int, g)
                _, last = monthrange(y, mth)
                return datetime(y, mth, last).strftime("%Y-%m-%d")
            elif idx == 3:
                y, mth = map(int, g)
                _, last = monthrange(y, mth)
                return datetime(y, mth, last).strftime("%Y-%m-%d")
        except ValueError as e:
            logging.error(f"[parse_expiry] invalid date: {ds} -> {e}")
            return None
    return None

def check_expiry_required(code: str) -> bool:
    if not code:
        return False
    conn = connect_db()
    if conn is None:
        return False
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("PRAGMA table_info(items_management)")
        cols = {r[1].lower() for r in cur.fetchall()}
        if "remarks" not in cols or "code" not in cols:
            return False
        cur.execute("SELECT remarks FROM items_management WHERE code=?",(code,))
        row = cur.fetchone()
        if not row or not row["remarks"]:
            return False
        return "exp" in row["remarks"].lower()
    except sqlite3.Error as e:
        logging.error(f"[check_expiry_required] {e}")
        return False
    finally:
        cur.close()
        conn.close()

def fetch_project_details():
    conn = connect_db()
    if conn is None:
        return ("Unknown Project","PRJ")
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        r = cur.fetchone()
        return (r[0] if r and r[0] else "Unknown Project",
                r[1] if r and r[1] else "PRJ")
    except sqlite3.Error as e:
        logging.error(f"[fetch_project_details] {e}")
        return ("Unknown Project","PRJ")
    finally:
        cur.close()
        conn.close()

def fetch_third_parties():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM third_parties ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error as e:
        logging.error(f"[fetch_third_parties] {e}")
        return []
    finally:
        cur.close()
        conn.close()

def fetch_end_users():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM end_users ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error as e:
        logging.error(f"[fetch_end_users] {e}")
        return []
    finally:
        cur.close()
        conn.close()

# ---------------------------------------------------------------
# stock_data manipulator
# ---------------------------------------------------------------
class StockData:
    @staticmethod
    def add_or_update(unique_id, scenario, qty_in=0, qty_out=0,
                      exp_date=None, kit_number=None, module_number=None):
        conn = connect_db()
        if conn is None:
            raise ValueError("DB connection failed")
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_data)")
            cols = {row[1].lower() for row in cur.fetchall()}
            has_scenario = "scenario" in cols
            cur.execute("SELECT qty_in, qty_out FROM stock_data WHERE unique_id=?",(unique_id,))
            row = cur.fetchone()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if row:
                new_in = (row[0] or 0) + qty_in
                new_out = (row[1] or 0) + qty_out
                if has_scenario:
                    cur.execute("""
                        UPDATE stock_data
                           SET qty_in=?, qty_out=?, exp_date=?, kit_number=?, module_number=?, scenario=?, updated_at=?
                         WHERE unique_id=?""",
                        (new_in, new_out, exp_date, kit_number, module_number, scenario, now, unique_id))
                else:
                    cur.execute("""
                        UPDATE stock_data
                           SET qty_in=?, qty_out=?, exp_date=?, kit_number=?, module_number=?, updated_at=?
                         WHERE unique_id=?""",
                        (new_in, new_out, exp_date, kit_number, module_number, now, unique_id))
            else:
                if has_scenario:
                    cur.execute("""
                        INSERT INTO stock_data
                            (unique_id, qty_in, qty_out, exp_date, kit_number, module_number, scenario, updated_at)
                        VALUES (?,?,?,?,?,?,?,?)""",
                        (unique_id, qty_in, qty_out, exp_date, kit_number, module_number, scenario, now))
                else:
                    cur.execute("""
                        INSERT INTO stock_data
                            (unique_id, qty_in, qty_out, exp_date, kit_number, module_number, updated_at)
                        VALUES (?,?,?,?,?,?,?)""",
                        (unique_id, qty_in, qty_out, exp_date, kit_number, module_number, now))
            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"[StockData.add_or_update] {e}")
            raise
        finally:
            cur.close()
            conn.close()

    @staticmethod
    def consume_by_line_id(line_id: int, qty_out_add: int):
        """
        Increase qty_out for existing stock_data line identified by line_id.
        """
        if qty_out_add <= 0:
            return
        conn = connect_db()
        if conn is None:
            raise ValueError("DB connection failed")
        cur = conn.cursor()
        try:
            cur.execute("SELECT qty_in, qty_out FROM stock_data WHERE line_id=?", (line_id,))
            row = cur.fetchone()
            if not row:
                logging.warning(f"[consume_by_line_id] line_id {line_id} not found")
                return
            new_out = (row[1] or 0) + qty_out_add
            cur.execute("""
                UPDATE stock_data
                   SET qty_out = ?, updated_at = CURRENT_TIMESTAMP
                 WHERE line_id = ?
            """, (new_out, line_id))
            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"[consume_by_line_id] {e}")
        finally:
            cur.close()
            conn.close()

# ---------------------------------------------------------------
# Main UI class
# ---------------------------------------------------------------
class StockInKit(tk.Frame):
    FIXED_IN_TYPE = "Internal move from on-shelf items"

    def __init__(self, parent, app, role: str = "supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        self.role = role.lower()

        def _sv(default=""):
            try:
                return tk.StringVar(value=default)
            except Exception:
                class DummyVar:
                    def __init__(self, v): self._v = v
                    def set(self, v): self._v = v
                    def get(self): return self._v
                return DummyVar(default)

        self.trans_type_var    = _sv(self.FIXED_IN_TYPE)
        self.scenario_var      = _sv("")
        self.mode_var          = _sv("")
        self.kit_var           = _sv("")
        self.kit_number_var    = _sv("")
        self.module_var        = _sv("")
        self.module_number_var = _sv("")
        self.end_user_var      = _sv("")
        self.third_party_var   = _sv("")
        self.status_var        = _sv(lang.t("receive_kit.ready","Ready"))

        self.tree = None
        self.search_var = None
        self.search_listbox = None
        self.scenario_cb = None
        self.mode_cb = None
        self.kit_cb = None
        self.kit_number_cb = None
        self.module_cb = None
        self.module_number_cb = None
        self.end_user_cb = None
        self.third_party_cb = None
        self.remarks_entry = None

        self.row_data = {}
        self.code_to_iid = {}
        self.mode_definitions = []
        self.mode_label_to_key = {}
        self.selected_scenario_id = None
        self.selected_scenario_name = None
        self.suggested_usage = defaultdict(int)
        self.editing_cell = None

        self.scenario_map = self.fetch_scenario_map()
        self.reverse_scenario_map = {v: k for k, v in self.scenario_map.items()}

        logging.info("StockInKit initialized")

        if self.parent is not None and self.parent.winfo_exists():
            self.pack(fill="both", expand=True)
            self.after(50, self.initialize_ui)
        else:
            logging.error("Parent window missing at initialization")


    # ---------------------------------------------------------
    # Localization & Code Extraction Helpers
    # ---------------------------------------------------------
    def _all_label(self):
        """Return localized 'All' label for dropdowns."""
        return lang.t("in_kit.all", "All")
    
    def _norm_all(self, val):
        """Normalize 'All' variants to English 'All' for internal use."""
        all_lbl = self._all_label()
        return "All" if (val is None or val == "" or val == all_lbl) else val
    
    def _extract_code_from_display(self, display_string: str) -> str:
        """
        Extract code from "CODE - Description" format.
        Handles prefixes like "● CODE - Description".
        
        Args:
            display_string: Either "CODE" or "CODE - Description"
        
        Returns:
            Just the code part, or None if empty/invalid
        """
        if not display_string:
            return None
        
        display_string = display_string.strip()
        
        if display_string == "-----":
            return None
        
        # Strip visual indicators
        prefixes = ["●", "■", "◆", "►", "[S]", "[Primary]", "[Standalone]"]
        for prefix in prefixes:
            if display_string.startswith(prefix):
                display_string = display_string[len(prefix):].strip()
                break
        
        # Extract code from "CODE - Description"
        if " - " in display_string:
            code = display_string.split(" - ", 1)[0].strip()
            return code if code else None
        
        return display_string

    def _canon_movement_type(self, display_label: str) -> str:
        """
        Convert localized movement type to canonical English for database storage.
        
        Args:
            display_label: Localized label from dropdown
        
        Returns:
            Canonical English movement type
        """
        internal_key = self.mode_label_to_key.get(display_label)
        
        if not internal_key:
            logging.warning(f"[IN_KIT] Unknown movement type label: {display_label}")
            return display_label
        
        # Map internal keys to canonical English
        canon_map = {
            "receive_kit": "Receive Kit",
            "receive_standalone": "Receive standalone items",
            "receive_module_scenario": "Receive module from scenario",
            "receive_module_kit": "Receive module from Kit",
            "receive_items_kit": "Receive items from Kit",
            "receive_items_module": "Receive items from module"
        }
        
        canonical = canon_map.get(internal_key, internal_key)
        logging.debug(f"[IN_KIT] Movement type: '{display_label}' → '{canonical}'")
        return canonical

    def _display_for_movement_type(self, canonical_value: str) -> str:
        """
        Convert canonical English movement type to localized display label.
        
        Args:
            canonical_value: English movement type from database
        
        Returns:
            Localized display label
        """
        reverse_canon_map = {
            "Receive Kit": "receive_kit",
            "Receive standalone items": "receive_standalone",
            "Receive module from scenario": "receive_module_scenario",
            "Receive module from Kit": "receive_module_kit",
            "Receive items from Kit": "receive_items_kit",
            "Receive items from module": "receive_items_module"
        }
        
        internal_key = reverse_canon_map.get(canonical_value, "receive_kit")
        
        for label, key in self.mode_label_to_key.items():
            if key == internal_key:
                return label
        
        return canonical_value


    # ------------- Ensure Vars ----------------
    def ensure_vars_ready(self):
        needed = {
            "trans_type_var": self.FIXED_IN_TYPE,
            "scenario_var": "",
            "mode_var": "",
            "kit_var": "",
            "kit_number_var": "",
            "module_var": "",
            "module_number_var": "",
            "end_user_var": "",
            "third_party_var": "",
            "status_var": lang.t("receive_kit.ready","Ready")
        }
        for name, default in needed.items():
            cur = getattr(self, name, None)
            if cur is None or not hasattr(cur, "set"):
                try:
                    setattr(self, name, tk.StringVar(value=default))
                except Exception:
                    class DummyVar:
                        def __init__(self,v): self._v=v
                        def set(self,v): self._v=v
                        def get(self): return self._v
                    setattr(self, name, DummyVar(default))

    # ------------- Scenario helpers ------------
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None:
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            return {str(r['scenario_id']): r['name'] for r in cur.fetchall()}
        except sqlite3.Error as e:
            logging.error(f"[fetch_scenario_map] {e}")
            return {}
        finally:
            cur.close()
            conn.close()

    def fetch_scenarios(self):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            return [{"id": r['scenario_id'], "name": r['name']} for r in cur.fetchall()]
        except sqlite3.Error as e:
            logging.error(f"[fetch_scenarios] {e}")
            return []
        finally:
            cur.close()
            conn.close()

    # ------------- Modes -----------------------
    def build_mode_definitions(self):
        scenario = self.selected_scenario_name or ""
        self.mode_definitions = [
            ("receive_kit",        lang.t("receive_kit.mode_generate_kit", "Generate Kit")),
            ("add_standalone",     lang.t("receive_kit.mode_add_standalone",     "Add standalone item/s in {scenario}", scenario=scenario)),
            ("add_module_scenario",lang.t("receive_kit.mode_add_module_scenario","Add module to {scenario}", scenario=scenario)),
            ("add_module_kit",     lang.t("receive_kit.mode_add_module_kit",     "Add module to a kit")),
            ("add_items_kit",      lang.t("receive_kit.mode_add_items_kit",      "Add items to a kit")),
            ("add_items_module",   lang.t("receive_kit.mode_add_items_module",   "Add items to a module"))
        ]
        self.mode_label_to_key = {lbl: key for key, lbl in self.mode_definitions}

    def current_mode_key(self):
        label = self.mode_var.get()
        return self.mode_label_to_key.get(label)

    def ensure_mode_ready(self):
        if not self.mode_definitions:
            self.build_mode_definitions()
        mk = self.current_mode_key()
        if not mk and self.mode_definitions:
            self.mode_var.set(self.mode_definitions[0][1])
            mk = self.mode_definitions[0][0]
        return mk

    # ------------- UI Init --------------------
    def initialize_ui(self):
        if not (self.parent and self.parent.winfo_exists()):
            return
        try:
            self.render_ui()
        except tk.TclError as e:
            logging.error(f"[initialize_ui] {e}")
            custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                         f"Failed to render UI: {e}", "error")

    def render_ui(self):
        self.ensure_vars_ready()
        for w in list(self.parent.winfo_children()):
            try: w.destroy()
            except: pass

        title_lbl = tk.Label(self.parent,
                             text=lang.t("receive_kit.title", "Receive Kit-Module"),
                             font=("Helvetica", 20, "bold"),
                             bg="#F0F4F8")
        title_lbl.pack(pady=10, fill="x")

        main = tk.Frame(self.parent, bg="#F0F4F8")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(main, text=lang.t("receive_kit.scenario","Scenario:"), bg="#F0F4F8")\
            .grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.scenario_cb = ttk.Combobox(main, textvariable=self.scenario_var,
                                        state="readonly", width=40)
        self.scenario_cb.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)

        tk.Label(main, text=lang.t("receive_kit.movement_type","Movement Type:"), bg="#F0F4F8")\
            .grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.mode_cb = ttk.Combobox(main, textvariable=self.mode_var,
                                    state="readonly", width=40)
        self.mode_cb.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.mode_cb.bind("<<ComboboxSelected>>", self.update_mode)

        tk.Label(main, text=lang.t("receive_kit.select_kit","Select Kit:"), bg="#F0F4F8")\
            .grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.kit_cb = ttk.Combobox(main, textvariable=self.kit_var, state="disabled", width=80)
        self.kit_cb.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.kit_cb.bind("<<ComboboxSelected>>", self.on_kit_selected)

        # ✅ Kit Number - Combobox (enabled conditionally for selecting EXISTING kits)
        tk.Label(main, text=lang.t("in_kit.kit_number","Kit Number:"), bg="#F0F4F8")\
            .grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.kit_number_cb = ttk.Combobox(main, textvariable=self.kit_number_var, 
                                          state="disabled", width=22)
        self.kit_number_cb.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.kit_number_cb.bind("<<ComboboxSelected>>", self.on_kit_number_selected)

        tk.Label(main, text=lang.t("receive_kit.select_module","Select Module:"), bg="#F0F4F8")\
            .grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.module_cb = ttk.Combobox(main, textvariable=self.module_var, state="disabled", width=80)
        self.module_cb.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.module_cb.bind("<<ComboboxSelected>>", self.on_module_selected)

        # ✅ Module Number - Combobox (enabled conditionally for selecting EXISTING modules)
        tk.Label(main, text=lang.t("in_kit.module_number","Module Number:"), bg="#F0F4F8")\
            .grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.module_number_cb = ttk.Combobox(main, textvariable=self.module_number_var, 
                                             state="disabled", width=22)
        self.module_number_cb.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        self.module_number_cb.bind("<<ComboboxSelected>>", self.on_module_number_selected)
        
        type_frame = tk.Frame(main, bg="#F0F4F8")
        type_frame.grid(row=4, column=0, columnspan=4, pady=5, sticky="w")
        tk.Label(type_frame, text=lang.t("receive_kit.in_type","IN Type:"), bg="#F0F4F8")\
            .grid(row=0, column=0, padx=5, sticky="w")
        tk.Label(type_frame, textvariable=self.trans_type_var,
                 bg="#E0E0E0", fg="#000", relief="sunken", width=30, anchor="w")\
            .grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(type_frame, text=lang.t("receive_kit.end_user","End User:"), bg="#F0F4F8")\
            .grid(row=0, column=2, padx=5, sticky="w")
        self.end_user_cb = ttk.Combobox(type_frame, textvariable=self.end_user_var,
                                        state="disabled", width=30)
        self.end_user_cb['values'] = fetch_end_users()
        self.end_user_cb.grid(row=0, column=3, padx=5, pady=5)

        tk.Label(type_frame, text=lang.t("receive_kit.third_party","Third Party:"), bg="#F0F4F8")\
            .grid(row=0, column=4, padx=5, sticky="w")
        self.third_party_cb = ttk.Combobox(type_frame, textvariable=self.third_party_var,
                                           state="disabled", width=30)
        self.third_party_cb['values'] = fetch_third_parties()
        self.third_party_cb.grid(row=0, column=5, padx=5, pady=5)

        tk.Label(type_frame, text=lang.t("receive_kit.remarks","Remarks:"), bg="#F0F4F8")\
            .grid(row=0, column=6, padx=5, sticky="w")
        self.remarks_entry = tk.Entry(type_frame, width=40, state="disabled")
        self.remarks_entry.grid(row=0, column=7, padx=5, pady=5)

        tk.Label(main, text=lang.t("receive_kit.search_item","Search Kit/Module/Item:"), bg="#F0F4F8")\
            .grid(row=5, column=0, padx=5, pady=(10,0), sticky="w")
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(main, textvariable=self.search_var, width=50)
        search_entry.grid(row=5, column=1, columnspan=3, padx=5, pady=(10,0), sticky="w")
        search_entry.bind("<KeyRelease>", self.search_items)
        search_entry.bind("<Return>", self.select_first_result)

        self.search_listbox = tk.Listbox(main, height=5, width=60)
        self.search_listbox.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky="we")
        self.search_listbox.bind("<<ListboxSelect>>", self.fill_from_search)

        # Columns + hidden line_id & qty_out (last two)
        cols = ("code","description","type","kit","module",
                "std_qty","qty_to_receive","expiry_date","batch_no","line_id","qty_out")
        self.tree = ttk.Treeview(main, columns=cols, show="headings", height=22)
        self.tree.tag_configure("light_red", background="#FF9999")
        self.tree.tag_configure("kit", background="#228B22", foreground="white")
        self.tree.tag_configure("module", background="#ADD8E6")

        headers = {
            "code": lang.t("receive_kit.code","Code"),
            "description": lang.t("receive_kit.description","Description"),
            "type": lang.t("receive_kit.type","Type"),
            "kit": lang.t("receive_kit.kit","Kit"),
            "module": lang.t("receive_kit.module","Module"),
            "std_qty": lang.t("receive_kit.std_qty","Std Qty"),
            "qty_to_receive": lang.t("receive_kit.qty_to_receive","Qty to Receive"),
            "expiry_date": lang.t("receive_kit.expiry_date","Expiry Date"),
            "batch_no": lang.t("receive_kit.batch_no","Batch No"),
            "line_id": "line_id (hidden)",
            "qty_out": "qty_out (hidden)"
        }
        widths = {
            "code":130,"description":380,"type":110,"kit":120,
            "module":120,"std_qty":90,"qty_to_receive":120,
            "expiry_date":120,"batch_no":120,"line_id":1,"qty_out":1
        }
        aligns = {
            "code":"w","description":"w","type":"w","kit":"w","module":"w",
            "std_qty":"e","qty_to_receive":"e","expiry_date":"w",
            "batch_no":"w","line_id":"w","qty_out":"e"
        }
        for c in cols:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths[c], anchor=aligns[c], stretch=False if c in ("line_id","qty_out") else True)
        vsb = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(main, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")
        vsb.grid(row=7, column=4, sticky="ns")
        hsb.grid(row=8, column=0, columnspan=4, sticky="ew")
        main.grid_rowconfigure(7, weight=1)
        main.grid_columnconfigure(1, weight=1)

        for ev in ("<Double-1>", "<KeyPress-Return>", "<KeyPress-Tab>"):
            self.tree.bind(ev, self.start_edit)
        for ev in ("<KeyPress-Up>", "<KeyPress-Down>", "<KeyPress-Left>", "<KeyPress-Right>"):
            self.tree.bind(ev, self.navigate_tree)
        self.tree.bind("<Button-3>", self.show_context_menu)

        btn_frame = tk.Frame(main, bg="#F0F4F8")
        btn_frame.grid(row=9, column=0, columnspan=4, pady=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.add_missing","Add Missing Item"),
                  bg="#FFA500", fg="white", command=self.add_missing_item).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.save","Save"),
                  bg="#27AE60", fg="white",
                  state="normal" if self.role in ["admin","manager"] else "disabled",
                  command=self.save_all).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.clear","Clear"),
                  bg="#7F8C8D", fg="white", command=self.clear_form).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.export","Export"),
                  bg="#2980B9", fg="white", command=self.export_data).pack(side="left", padx=5)

        tk.Label(main, textvariable=self.status_var, relief="sunken",
                 anchor="w", bg="#F0F4F8").grid(row=10, column=0, columnspan=4, sticky="ew")

        self.load_scenarios()

    # ------------- Scenario & mode logic -------------
    def load_scenarios(self):
        self.ensure_vars_ready()
        scenarios = self.fetch_scenarios()
        names = [f"{s['id']} - {s['name']}" for s in scenarios]
        self.scenario_cb['values'] = names
        if names:
            self.scenario_cb.current(0)
            self.on_scenario_selected()

    def on_scenario_selected(self, event=None):
        scen = self.scenario_var.get()
        if not scen:
            self.selected_scenario_id = None
            self.selected_scenario_name = None
            self.clear_search()
            return
        self.selected_scenario_id, self.selected_scenario_name = scen.split(" - ", 1)
        self.build_mode_definitions()
        self.mode_cb['values'] = [lbl for _, lbl in self.mode_definitions]
        if self.mode_definitions:
            self.mode_var.set(self.mode_definitions[0][1])
        self.update_mode()
        self.clear_search()

    def update_mode(self, event=None):
        """
        Called when movement type changes.
        Enables/disables appropriate selectors based on mode.
        
        ✅ Workflows:
        1. receive_kit: Kit CODE → Popup for NEW kit number
        2. add_standalone: Search only (no kit/module)
        3. add_module_scenario: Module CODE → Popup for NEW module number (NO KIT)
        4. add_module_kit: Kit NUMBER (existing) → Module CODE → Popup for module number
        5. add_items_kit: Kit NUMBER (existing) → Search
        6. add_items_module: Kit NUMBER (existing) → Module NUMBER (existing) → Search
        """
        self.ensure_vars_ready()
        mode_key = self.current_mode_key()
        
        logging.debug(f"[IN_KIT] Mode changed to: {mode_key}")

        # ===== DISABLE ALL SELECTORS =====
        for cb in [self.kit_cb, self.kit_number_cb, self.module_cb, self.module_number_cb]:
            if cb:
                cb.config(state="disabled")

        # ===== CLEAR ALL VALUES =====
        self.kit_var.set("")
        self.kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        
        # Clear dropdown values
        if self.kit_cb:
            self.kit_cb['values'] = []
        if self.kit_number_cb:
            self.kit_number_cb['values'] = []
        if self.module_cb:
            self.module_cb['values'] = []
        if self.module_number_cb:
            self.module_number_cb['values'] = []

        if not self.selected_scenario_id:
            return

        # ===== MODE-SPECIFIC LOGIC =====
        
        if mode_key == "receive_kit":
            # ✅ Workflow: Kit CODE → Popup for kit number
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.status_var.set(
                lang.t("in_kit.select_kit_to_generate", "Select a kit to generate")
            )
            logging.debug(f"[IN_KIT] receive_kit: {len(self.kit_cb['values'])} kits available")
        
        elif mode_key == "add_standalone":
            # ✅ Workflow: Search only (no kit/module) - AUTO-POPULATE
            self.status_var.set(
                lang.t("in_kit.search_standalone", "Search for standalone items")
            )
            logging.debug(f"[IN_KIT] add_standalone: Search enabled")
            
            # ✅ AUTO-POPULATE search results
            self.after(100, self.auto_populate_search)
        
        elif mode_key == "add_module_scenario":
            # ✅ Workflow: Module CODE → Popup for module number (NO KIT NEEDED)
            self.module_cb.config(state="readonly")
            modules = self.fetch_all_modules(self.selected_scenario_id)
            self.module_cb['values'] = modules
            self.status_var.set(
                lang.t("in_kit.select_primary_module", 
                       "Select a primary module to add (no kit needed)")
            )
            logging.debug(f"[IN_KIT] add_module_scenario: {len(modules)} primary modules")
        
        elif mode_key == "add_module_kit":
            # ✅ Workflow: Kit CODE → Kit NUMBER → Module CODE → Popup for module number
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.status_var.set(
                lang.t("in_kit.select_kit_code_first",
                       "Step 1: Select kit type ({count} available)",
                       count=len(self.kit_cb['values']))
            )
            logging.debug(f"[IN_KIT] add_module_kit: {len(self.kit_cb['values'])} kit types")
            

        
        elif mode_key == "add_items_kit":
            # ✅ Workflow: Kit CODE → Kit NUMBER → Search
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.status_var.set(
                lang.t("in_kit.select_kit_code_first",
                       "Step 1: Select kit type ({count} available)",
                       count=len(self.kit_cb['values']))
            )
            logging.debug(f"[IN_KIT] add_items_kit: {len(self.kit_cb['values'])} kit types")
        
        elif mode_key == "add_items_module":
            # ✅ Workflow: Kit CODE → Kit NUMBER → Module CODE → Module NUMBER → Search
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.status_var.set(
                lang.t("in_kit.select_kit_code_for_module",
                       "Step 1: Select kit type ({count} available)",
                       count=len(self.kit_cb['values']))
            )
            logging.debug(f"[IN_KIT] add_items_module: {len(self.kit_cb['values'])} kit types")
            
    def on_kit_selected(self, event=None):
        """
        Handle kit selection.
        Flow:
        1. Extract kit code from display
        2. If mode is receive_kit → load_hierarchy() will handle popup for kit number
        3. If mode is add_module_kit → enable module dropdown (popup happens when module selected)
        4. Other modes → ready for next action
        """
        self.ensure_vars_ready()
        kit_display = self.kit_var.get()
    
        if not kit_display:
            return
    
        # Extract code from "CODE - Description"
        kit_code = self._extract_code_from_display(kit_display)
    
        logging.debug(f"[IN_KIT] Kit selected: '{kit_display}' -> Code: '{kit_code}'")
    
        # Clear dependent fields (but don't clear kit_number if already set)
        self.module_var.set("")
        self.module_number_var.set("")
    
        if not kit_code:
            return
    
        mode_key = self.current_mode_key()
    
        if mode_key == "receive_kit":
            # ✅ Load hierarchy will trigger popup for kit number
            self.load_hierarchy(kit_code)
    
        elif mode_key == "add_module_kit":
            # ✅ Step 2: Enable kit NUMBER dropdown (filtered by kit code)
            kit_numbers = self.fetch_existing_kit_numbers_for_code(
                self.selected_scenario_id,
                kit_code
            )
            
            if kit_numbers:
                self.kit_number_cb.config(state="readonly")
                self.kit_number_cb['values'] = kit_numbers
                self.status_var.set(
                    lang.t("in_kit.select_kit_number_for_module",
                           "Step 2: Select kit number ({count} of this kit type)",
                           count=len(kit_numbers))
                )
                logging.debug(f"[IN_KIT] Enabled {len(kit_numbers)} kit numbers for {kit_code}")
            else:
                self.status_var.set(
                    lang.t("in_kit.no_kit_numbers_for_type",
                           "No kit numbers found for this kit type")
                )
    
        elif mode_key == "add_items_kit":
            # ✅ Step 2: Enable kit NUMBER dropdown (filtered by selected kit CODE)
            kit_numbers = self.fetch_existing_kit_numbers_for_code(
                self.selected_scenario_id, 
                kit_code
            )
            
            if kit_numbers:
                self.kit_number_cb.config(state="readonly")
                self.kit_number_cb['values'] = kit_numbers
                self.status_var.set(
                    lang.t("in_kit.select_kit_number_then_search",
                           "Step 2: Select kit number ({count} of this kit type)",
                           count=len(kit_numbers))
                )
                logging.debug(f"[IN_KIT] Enabled {len(kit_numbers)} kit numbers for {kit_code}")
            else:
                self.status_var.set(
                    lang.t("in_kit.no_kit_numbers_for_type", 
                           "No kit numbers found for this kit type")
                )
        
        elif mode_key == "add_items_module":
            # ✅ Step 2: Enable kit NUMBER dropdown
            kit_numbers = self.fetch_existing_kit_numbers_for_code(
                self.selected_scenario_id,
                kit_code
            )
            
            if kit_numbers:
                self.kit_number_cb.config(state="readonly")
                self.kit_number_cb['values'] = kit_numbers
                self.status_var.set(
                    lang.t("in_kit.select_kit_number_then_module",
                           "Step 2: Select kit number ({count} available)",
                           count=len(kit_numbers))
                )
                logging.debug(f"[IN_KIT] Enabled {len(kit_numbers)} kit numbers for {kit_code}")
            else:
                self.status_var.set(
                    lang.t("in_kit.no_kit_numbers_for_type",
                           "No kit numbers found for this kit type")
                )
    
        else:
            # Default fallback
            self.status_var.set(
                lang.t("in_kit.kit_selected", "Kit selected: {code}", code=kit_code)
            )

    def on_kit_number_selected(self, event=None):
        """
        Handle selection of existing kit NUMBER.
        Used in modes: add_module_kit, add_items_kit, add_items_module
        
        ✅ Workflow:
        - add_module_kit: Kit NUMBER selected → Enable module CODE dropdown
        - add_items_kit: Kit NUMBER selected → AUTO-POPULATE search
        - add_items_module: Kit NUMBER selected → Enable module CODE dropdown
        """
        self.ensure_vars_ready()
        kit_number = self.kit_number_var.get().strip()
        
        if not kit_number:
            return
        
        mode_key = self.current_mode_key()
        
        logging.debug(f"[IN_KIT] Kit number selected: {kit_number} (mode: {mode_key})")
        
        if mode_key == "add_module_kit":
            # ✅ Step 3: Enable module CODE dropdown
            kit_code = self._get_kit_code_for_number(kit_number)
            
            if kit_code:
                modules = self.fetch_modules_for_kit(self.selected_scenario_id, kit_code)
                
                if modules:
                    self.module_cb.config(state="readonly")
                    self.module_cb['values'] = modules
                    self.status_var.set(
                        lang.t("in_kit.select_module_to_add_to_kit",
                               "Step 3: Select module to add ({count} available)",
                               count=len(modules))
                    )
                    logging.debug(f"[IN_KIT] Enabled {len(modules)} modules for kit {kit_code}")
                else:
                    self.status_var.set(
                        lang.t("in_kit.no_modules_for_kit", "No modules available for this kit type")
                    )
            else:
                self.status_var.set(
                    lang.t("in_kit.kit_code_not_found", "Could not determine kit type")
                )
            
        elif mode_key == "add_items_kit":
            # ✅ AUTO-POPULATE search results for items
            self.status_var.set(
                lang.t("in_kit.search_items_for_kit",
                       "Kit number: {num}. Items loaded below.",
                       num=kit_number)
            )
            # ✅ Trigger auto-populate
            self.after(100, self.auto_populate_search)
        
        elif mode_key == "add_items_module":
            # ✅ Step 3: Enable module CODE dropdown
            kit_code = self._extract_code_from_display(self.kit_var.get())
            
            if kit_code:
                modules = self.fetch_modules_for_kit(self.selected_scenario_id, kit_code)
                
                if modules:
                    self.module_cb.config(state="readonly")
                    self.module_cb['values'] = modules
                    self.status_var.set(
                        lang.t("in_kit.select_module_code_then_number",
                               "Step 3: Select module type ({count} available)",
                               count=len(modules))
                    )
                    logging.debug(f"[IN_KIT] Enabled {len(modules)} module types")
                else:
                    self.status_var.set(
                        lang.t("in_kit.no_modules_for_kit", "No modules in this kit type")
                    )
                
    def on_module_selected(self, event=None):
        """
        Handle module selection.
        
        ✅ Workflows:
        - add_module_scenario: Module CODE selected → Popup for NEW module number
        - add_module_kit: Module CODE selected → Popup for NEW module number
        - add_items_module: Module CODE selected → Enable module NUMBER dropdown (existing)
        """
        self.ensure_vars_ready()
        module_display = self.module_var.get()
    
        if not module_display:
            return
    
        module_code = self._extract_code_from_display(module_display)
    
        logging.debug(f"[IN_KIT] Module selected: '{module_display}' -> Code: '{module_code}'")
    
        if not module_code:
            return
    
        mode_key = self.current_mode_key()
        kit_number = self.kit_number_var.get().strip() or None
    
        # ===== WORKFLOW: Add Items to Module =====
        if mode_key == "add_items_module":
            # ✅ Step 4: Enable module NUMBER dropdown (existing)
            module_numbers = self.fetch_existing_module_numbers(
                self.selected_scenario_id,
                kit_number
            )
            
            if module_numbers:
                self.module_number_cb.config(state="readonly")
                self.module_number_cb['values'] = module_numbers
                self.status_var.set(
                    lang.t("in_kit.select_module_number_final",
                           "Step 4: Select module number ({count} available)",
                           count=len(module_numbers))
                )
                logging.debug(f"[IN_KIT] Enabled {len(module_numbers)} module numbers")
            else:
                self.status_var.set(
                    lang.t("in_kit.no_module_numbers", "No module numbers found")
                )
        
        # ===== WORKFLOW: Add Module (new module number via popup) =====
        else:
            # ✅ Trigger popup for NEW module number
            module_number = self.ask_module_number(kit_number, module_code)
        
            if module_number:
                self.module_number_var.set(module_number)
                self.status_var.set(
                    lang.t("in_kit.module_number_set", 
                        "Module number set: {num}. Ready to add items.", 
                        num=module_number)
                )
                logging.info(f"[IN_KIT] Module number entered: {module_number}")
            else:
                # User cancelled
                self.module_var.set("")
                self.module_number_var.set("")
                self.status_var.set(
                    lang.t("in_kit.module_cancelled", "Module selection cancelled")
                )


    def on_module_number_selected(self, event=None):
        """
        Handle selection of existing module NUMBER.
        Used in mode: add_items_module
        
        ✅ Workflow:
        - add_items_module: Module NUMBER selected → AUTO-POPULATE search
        """
        self.ensure_vars_ready()
        module_number = self.module_number_var.get().strip()
        
        if not module_number:
            return
        
        mode_key = self.current_mode_key()
        
        logging.debug(f"[IN_KIT] Module number selected: {module_number} (mode: {mode_key})")
        
        if mode_key == "add_items_module":
            # ✅ AUTO-POPULATE search results for items
            kit_number = self.kit_number_var.get().strip()
            self.status_var.set(
                lang.t("in_kit.search_items_for_module",
                       "Kit: {kit}, Module: {mod}. Items loaded below.",
                       kit=kit_number,
                       mod=module_number)
            )
            # ✅ Trigger auto-populate
            self.after(100, self.auto_populate_search)



    def auto_populate_search(self):
        """
        Automatically populate search results without user typing.
        Used when entering certain modes or after selecting kit/module numbers.
        
        ✅ Triggers automatic search with empty query to show all available items.
        """
        if not self.selected_scenario_id:
            return
        
        mode_key = self.current_mode_key()
        
        logging.debug(f"[IN_KIT] Auto-populating search for mode: {mode_key}")
        
        # Clear search input (show we're loading all items)
        if self.search_var:
            self.search_var.set("")
        
        # Clear existing results
        if self.search_listbox:
            self.search_listbox.delete(0, tk.END)
        
        # Fetch ALL items for current mode (empty query = all items)
        results = self.fetch_search_results("", self.selected_scenario_id, mode_key)
        
        # Populate search listbox
        for r in results:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        
        # Update status
        count = len(results)
        if count > 0:
            self.status_var.set(
                lang.t("in_kit.items_loaded", 
                       "{count} items available (search to filter)",
                       count=count)
            )
        else:
            self.status_var.set(
                lang.t("in_kit.no_items_available", "No items available")
            )
        
        logging.info(f"[IN_KIT] Auto-populated {count} items for mode: {mode_key}")			

                
    # ------------- Clear & Search -------------
    def clear_search(self):
        if self.search_var:
            self.search_var.set("")
        if self.search_listbox:
            self.search_listbox.delete(0, tk.END)
        if self.tree:
            self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        self.suggested_usage.clear()
        self.status_var.set(lang.t("receive_kit.ready","Ready"))

    def clear_form(self):
        self.ensure_vars_ready()
        self.clear_search()
        self.trans_type_var.set(self.FIXED_IN_TYPE)
        self.scenario_var.set("")
        self.mode_var.set("")
        self.kit_var.set("")
        self.kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        self.end_user_var.set("")
        self.third_party_var.set("")
        if self.remarks_entry:
            self.remarks_entry.config(state="normal")
            self.remarks_entry.delete(0, tk.END)
            self.remarks_entry.config(state="disabled")
        if self.end_user_cb:
            self.end_user_cb.config(state="disabled")
        if self.third_party_cb:
            self.third_party_cb.config(state="disabled")
        self.load_scenarios()

    def search_items(self, event=None):
        if not self.selected_scenario_id:
            return
        q = self.search_var.get().strip() if self.search_var else ""
        self.search_listbox.delete(0, tk.END)
        mk = self.current_mode_key()
        results = self.fetch_search_results(q, self.selected_scenario_id, mk)
        for r in results:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        self.status_var.set(lang.t("receive_kit.found_items", f"Found {self.search_listbox.size()} items"))

    def select_first_result(self, event=None):
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self.fill_from_search()

    def fill_from_search(self, event=None):
        sel = self.search_listbox.curselection()
        if not sel:
            return
        line = self.search_listbox.get(sel[0])
        code = line.split(" - ")[0]
        self.search_listbox.selection_clear(0, tk.END)
        mk = self.current_mode_key()
        if mk == "receive_kit":
            self.load_hierarchy(code)
        else:
            self.add_to_tree(code)

    # ------------- Fetch domain data -------------
    def fetch_kits(self, scenario_id):
        """
        Fetch PRIMARY kits from kit_items (level='primary').
        Only includes items with type='Kit' (language-independent).
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[IN_KIT] DB connection failed in fetch_kits")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get primary kits from kit_items
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id,))
            
            kit_codes = [r['code'] for r in cur.fetchall()]
            
            if not kit_codes:
                logging.debug(f"[IN_KIT] No primary kits found for scenario {scenario_id}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for kit_code in kit_codes:
                desc = get_item_description(kit_code)
                item_type = detect_type(kit_code, desc).upper()
                
                # Only include if type is KIT
                if item_type == "KIT":
                    display = f"{kit_code} - {desc}" if desc else kit_code
                    result.append(display)
            
            logging.info(f"[IN_KIT] Found {len(result)} primary kits for scenario {scenario_id}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_kits error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_all_modules(self, scenario_id):
        """
        Fetch PRIMARY modules from kit_items (level='primary', standalone modules).
        Only includes items with type='Module' (language-independent).
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[IN_KIT] DB connection failed in fetch_all_modules")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get primary standalone modules
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND module IS NOT NULL 
                  AND module != ''
                  AND module != 'None'
                  AND (kit IS NULL OR kit = '' OR kit = 'None')
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id,))
            
            module_codes = [r['code'] for r in cur.fetchall()]
            
            if not module_codes:
                logging.debug(f"[IN_KIT] No primary modules found for scenario {scenario_id}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()
                
                # Only include if type is MODULE
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)
            
            logging.info(f"[IN_KIT] Found {len(result)} primary modules for scenario {scenario_id}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_all_modules error: {e}")
            return []
        finally:
            cur.close()
            conn.close()


    def fetch_modules_for_kit(self, scenario_id, kit_code):
        """
        Fetch SECONDARY modules inside a specific kit from kit_items.
        Only includes items with type='Module'.
        
        Args:
            scenario_id: Scenario ID
            kit_code: Kit code (extracted from dropdown)
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[IN_KIT] DB connection failed in fetch_modules_for_kit")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get secondary modules inside this kit
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND kit=?
                  AND level='secondary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id, kit_code))
            
            module_codes = [r['code'] for r in cur.fetchall()]
            
            if not module_codes:
                logging.debug(f"[IN_KIT] No modules found in kit {kit_code}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()
                
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)
            
            logging.debug(f"[IN_KIT] Found {len(result)} modules in kit {kit_code}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_modules_for_kit error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_available_kit_numbers(self, scenario_id, kit_code=None):
        if not scenario_id: return []
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            if kit_code:
                cur.execute("""
                    SELECT DISTINCT kit_number
                      FROM stock_data
                     WHERE kit_number IS NOT NULL
                       AND kit_number!='None'
                       AND unique_id LIKE ?
                     ORDER BY kit_number
                """,(f"{scenario_id}/{kit_code}/%",))
            else:
                cur.execute("""
                    SELECT DISTINCT kit_number
                      FROM stock_data
                     WHERE kit_number IS NOT NULL
                       AND kit_number!='None'
                       AND unique_id LIKE ?
                     ORDER BY kit_number
                """,(f"{scenario_id}/%",))
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error as e:
            logging.error(f"[fetch_available_kit_numbers] {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_module_numbers(self, scenario_id, kit_code=None, module_code=None):
        if not scenario_id: return []
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            where = ["module_number IS NOT NULL","module_number!='None'","unique_id LIKE ?"]
            params = [f"{scenario_id}/%"]
            if kit_code:
                where.append("unique_id LIKE ?")
                params.append(f"{scenario_id}/{kit_code}/%")
            if module_code:
                where.append("unique_id LIKE ?")
                params.append(f"{scenario_id}/%/{module_code}/%")
            sql = f"""
                SELECT DISTINCT module_number
                  FROM stock_data
                 WHERE {' AND '.join(where)}
                 ORDER BY module_number
            """
            cur.execute(sql, params)
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error as e:
            logging.error(f"[fetch_module_numbers] {e}")
            return []
        finally:
            cur.close()
            conn.close()

		#--------------------------HELPERS for Dropd Downs-------------#

    def fetch_existing_kit_numbers(self, scenario_id):
        """
        Fetch existing kit numbers with stock > 0 from stock_data.
        Used for selecting which kit to add modules/items to.
        
        Returns:
            List of kit numbers (strings), sorted
        """
        if not scenario_id:
            return []
        
        conn = connect_db()
        if conn is None:
            return []
        
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT kit_number
                FROM stock_data
                WHERE scenario = ?
                  AND kit_number IS NOT NULL
                  AND kit_number != 'None'
                  AND kit_number != ''
                  AND (qty_in - COALESCE(qty_out, 0)) > 0
                ORDER BY kit_number
            """, (str(scenario_id),))
            
            results = [r[0] for r in cur.fetchall()]
            logging.debug(f"[IN_KIT] Found {len(results)} existing kit numbers")
            return results
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_existing_kit_numbers error: {e}")
            return []
        finally:
            cur.close()
            conn.close()
    
    def fetch_existing_module_numbers(self, scenario_id, kit_number=None):
        """
        Fetch existing module numbers with stock > 0 from stock_data.
        Used for selecting which module to add items to.
        
        Args:
            scenario_id: Scenario ID
            kit_number: Optional kit number filter
        
        Returns:
            List of module numbers (strings), sorted
        """
        if not scenario_id:
            return []
        
        conn = connect_db()
        if conn is None:
            return []
        
        cur = conn.cursor()
        try:
            sql = """
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE scenario = ?
                  AND module_number IS NOT NULL
                  AND module_number != 'None'
                  AND module_number != ''
                  AND (qty_in - COALESCE(qty_out, 0)) > 0
            """
            params = [str(scenario_id)]
            
            if kit_number:
                sql += " AND kit_number = ?"
                params.append(kit_number)
            
            sql += " ORDER BY module_number"
            
            cur.execute(sql, params)
            
            results = [r[0] for r in cur.fetchall()]
            logging.debug(f"[IN_KIT] Found {len(results)} existing module numbers")
            return results
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_existing_module_numbers error: {e}")
            return []
        finally:
            cur.close()
            conn.close()
	#-------------------Helper------------

    def fetch_existing_kit_numbers_for_code(self, scenario_id, kit_code):
        """
        Fetch existing kit numbers for a SPECIFIC kit code.
        Filters by both scenario AND kit code.
        
        Args:
            scenario_id: Scenario ID
            kit_code: Kit code (e.g., "CHOLKIT001")
        
        Returns:
            List of kit numbers for this specific kit type
        """
        if not scenario_id or not kit_code:
            return []
        
        conn = connect_db()
        if conn is None:
            return []
        
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT kit_number
                FROM stock_data
                WHERE scenario = ?
                  AND kit = ?
                  AND kit_number IS NOT NULL
                  AND kit_number != 'None'
                  AND kit_number != ''
                  AND (qty_in - COALESCE(qty_out, 0)) > 0
                ORDER BY kit_number
            """, (str(scenario_id), kit_code))
            
            results = [r[0] for r in cur.fetchall()]
            logging.debug(f"[IN_KIT] Found {len(results)} kit numbers for {kit_code}")
            return results
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_existing_kit_numbers_for_code error: {e}")
            return []
        finally:
            cur.close()
            conn.close()


    def _get_kit_code_for_number(self, kit_number: str) -> str:
        """
        Get kit CODE from kit NUMBER by querying stock_data.
        
        Args:
            kit_number: Kit number (e.g., "CHOLKIT001-K1")
        
        Returns:
            Kit code (e.g., "CHOLKIT001"), or None if not found
        """
        if not kit_number or not self.selected_scenario_id:
            return None
        
        conn = connect_db()
        if conn is None:
            return None
        
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT kit
                FROM stock_data
                WHERE scenario = ?
                  AND kit_number = ?
                  AND kit IS NOT NULL
                  AND kit != 'None'
                  AND kit != ''
                LIMIT 1
            """, (str(self.selected_scenario_id), kit_number))
            
            row = cur.fetchone()
            return row[0] if row else None
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] _get_kit_code_for_number error: {e}")
            return None
        finally:
            cur.close()
            conn.close()
            
#------------------HELPERS Drop down finished----------			



    def fetch_search_results(self, query, scenario_id, mode_key):
        """
        Search for items based on mode and query.
        Returns list of dicts with code, description, level, type.
        """
        if not scenario_id:
            return []
        if not mode_key:
            mode_key = self.ensure_mode_ready()
        if mode_key in (self.mode_label_to_key or {}):
            mode_key = self.mode_label_to_key[mode_key]
        
        mk = (mode_key or "").lower()
        conn = connect_db()
        if conn is None:
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            q = (query or "").lower()
            
            def common_params():
                return (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            
            if mk == "receive_kit":
                # Search for PRIMARY level kits
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND LOWER(ki.level)='primary'
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, *common_params())
            
            elif mk == "add_standalone":
                # ✅ FIXED: Search for PRIMARY level standalone ITEMS (not in kit/module)
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND LOWER(ki.level)='primary'
                      AND (ki.kit IS NULL OR ki.kit = '' OR ki.kit = 'None')
                      AND (ki.module IS NULL OR ki.module = '' OR ki.module = 'None')
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, *common_params())
            
            elif mk == "add_module_scenario":
                # Search for PRIMARY level modules (not in kits)
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND LOWER(ki.level)='primary'
                      AND (ki.kit IS NULL OR ki.kit = '' OR ki.kit = 'None')
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, *common_params())
            
            elif mk == "add_module_kit":
                # Search for modules in selected kit
                kit_code = self._extract_code_from_display(self.kit_var.get())
                if not kit_code:
                    return []
                
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND ki.kit=?
                      AND LOWER(ki.level)='secondary'
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, kit_code, *common_params())
            
            elif mk == "add_items_kit":
                # ✅ FIXED: Search for ITEMS in selected kit
                kit_code = self._extract_code_from_display(self.kit_var.get())
                if not kit_code:
                    return []
                
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND ki.kit=?
                      AND LOWER(ki.level)='tertiary'
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, kit_code, *common_params())
            
            elif mk == "add_items_module":
                # ✅ FIXED: Search for ITEMS in selected module
                kit_code = self._extract_code_from_display(self.kit_var.get())
                module_code = self._extract_code_from_display(self.module_var.get())
                
                if not kit_code or not module_code:
                    return []
                
                sql = """
                   SELECT DISTINCT ki.code, ki.level
                     FROM kit_items ki
                LEFT JOIN items_list il ON il.code=ki.code
                    WHERE ki.scenario_id=?
                      AND ki.kit=?
                      AND ki.module=?
                      AND LOWER(ki.level)='tertiary'
                      AND (
                        UPPER(ki.code) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?)
                        OR UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                      )
                 ORDER BY ki.code
                """
                params = (scenario_id, kit_code, module_code, *common_params())
            
            else:
                return []
            
            cur.execute(sql, params)
            rows = cur.fetchall()
            
            out = []
            for r in rows:
                code = r['code']
                desc = get_item_description(code)
                t = detect_type(code, desc)
                out.append({
                    'code': code,
                    'description': desc,
                    'level': r['level'],
                    'type': t
                })
            
            logging.debug(f"[IN_KIT] Search found {len(out)} results for mode '{mk}'")
            return out
            
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_search_results error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_kit_items(self, scenario_id, code):
        if not scenario_id or not code:
            return []
        conn = connect_db()
        if conn is None: return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT code, level, kit, module, item,
                       COALESCE(std_qty,0) AS std_qty,
                       treecode
                  FROM kit_items
                 WHERE scenario_id=?
                   AND (code=? OR kit=? OR module=? OR item=?)
                 ORDER BY treecode
            """,(scenario_id, code, code, code, code))
            comps = []
            for r in cur.fetchall():
                dsc = get_item_description(r['code'])
                comps.append({
                    'code': r['code'],
                    'description': dsc,
                    'type': detect_type(r['code'], dsc),
                    'kit': r['kit'] or "-----",
                    'module': r['module'] or "-----",
                    'item': r['item'] or "-----",
                    'std_qty': r['std_qty'],
                    'treecode': r['treecode']
                })
            return comps
        except sqlite3.Error as e:
            logging.error(f"[fetch_kit_items] {e}")
            return []
        finally:
            cur.close()
            conn.close()

    # ------------- On-shelf batch extraction (per line) -------------
    def fetch_on_shelf_batches(self, code: str):
        """
        Fetch on-shelf stock for a given item code, sorted by LONGEST expiry first.
    
        Returns list of dicts:
            - code: item code
            - expiry: parsed YYYY-MM-DD or original string
         - management_mode: 'on_shelf'
            - final_qty: available quantity
            - line_id: stock_data.line_id
    
        ✅ Filters:
            - Same scenario (using scenario_id)
            - On-shelf items only (≤6 slashes in unique_id OR management_mode='on-shelf')
            - final_qty > 0

        ✅ Sorting:
            - Longest expiry first (DESC)
            - NULL/empty expiry last
        """
        if not code:
            return []
    
        code = code.strip().upper()
        scenario_id = self.selected_scenario_id
    
        if not scenario_id:
            logging.warning("[IN_KIT] No scenario selected in fetch_on_shelf_batches")
            return []
    
        conn = connect_db()
        if conn is None:
            logging.error("[IN_KIT] DB connection failed in fetch_on_shelf_batches")
            return []
    
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
    
        try:
            # ✅ Query: Filter by scenario_id, on-shelf status, positive stock
            cur.execute("""
                SELECT 
                    unique_id,
                    exp_date,
                    qty_in,
                    qty_out,
                    management_mode,
                    line_id
                FROM stock_data
                WHERE item = ?
                AND scenario = ?
                AND (qty_in - COALESCE(qty_out, 0)) > 0
                AND (
                    management_mode = 'on-shelf' 
                    OR LENGTH(unique_id) - LENGTH(REPLACE(unique_id, '/', '')) <= 6
                )
                ORDER BY 
                CASE WHEN exp_date IS NULL OR exp_date = '' OR exp_date = 'None' THEN 1 ELSE 0 END,
                exp_date DESC
            """, (code, scenario_id))
        
            rows = cur.fetchall()
        
            if not rows:
                logging.info(f"[IN_KIT] No on-shelf stock found for {code} in scenario {scenario_id}")
                return []
        
            batches = []
        
            for r in rows:
                # Calculate final_qty
                qty_in = r['qty_in'] or 0
                qty_out = r['qty_out'] or 0
                final_qty = qty_in - qty_out
            
                if final_qty <= 0:
                    continue
            
                # Parse expiry date
                exp_date = r['exp_date']
            
                # If exp_date column is empty, try extracting from unique_id
                if not exp_date or exp_date == 'None':
                    parts = r['unique_id'].split('/')
                    if len(parts) >= 6:
                        exp_part = parts[5]
                        if exp_part and exp_part != 'None':
                            exp_date = exp_part
            
                # Parse to standard format if possible
                if exp_date and exp_date != 'None':
                    parsed = parse_expiry(exp_date)
                    if parsed:
                        exp_date = parsed
                else:
                    exp_date = None
            
                batches.append({
                    'code': code,
                    'expiry': exp_date,
                    'management_mode': 'on_shelf',
                    'final_qty': final_qty,
                    'line_id': r['line_id']
                })
        
            # ✅ Log results
            if batches:
                longest_exp = batches[0]['expiry']
                logging.info(
                    f"[IN_KIT] Found {len(batches)} on-shelf batches for {code}, "
                    f"longest expiry: {longest_exp}"
                )
        
            return batches
        
        except sqlite3.Error as e:
            logging.error(f"[IN_KIT] fetch_on_shelf_batches error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def remaining_available(self, code, expiry, management_mode, total_final):
        key = (code, expiry or "", management_mode)
        used = self.suggested_usage.get(key, 0)
        return max(total_final - used, 0)

    def record_suggested(self, code, expiry, management_mode, qty):
        key = (code, expiry or "", management_mode)
        self.suggested_usage[key] += qty

    def _remaining_std_allowance(self, code: str) -> int:
        std_val = None
        assigned = 0
        for top in self.tree.get_children():
            stack = [top]
            while stack:
                iid = stack.pop()
                vals = self.tree.item(iid, "values")
                if not vals: continue
                c, t, s, q = vals[0], vals[2], vals[5], vals[6]
                if c == code and t.upper() == "ITEM":
                    if std_val is None and str(s).isdigit():
                        std_val = int(s)
                    if q and str(q).isdigit():
                        assigned += int(q)
                stack.extend(self.tree.get_children(iid))
        if not std_val or std_val <= 0:
            return 10**12
        rem = std_val - assigned
        return rem if rem > 0 else 0

    # ------------- Module number helper -------------
    def ask_module_number(self, kit_number: str, module_code: str) -> str:
        """
        Ask user for a new module number with pre-filled suggestion.
        
        Args:
            kit_number: Kit number (for suggestion)
            module_code: Module code (for suggestion if no kit)
        
        Returns:
            Entered module number (validated as unique), or None if cancelled
        """
        # ✅ Generate smart suggestion
        if kit_number and kit_number.strip():
            suggestion = f"{kit_number.strip()}-M"
        elif module_code and module_code.strip():
            suggestion = f"{module_code.strip()}-M1"
        else:
            suggestion = "M1"
        
        while True:
            # ✅ Use custom dialog with pre-filled suggestion
            entered = self.ask_custom_text(
                title=lang.t("receive_kit.module_number", "Module Number"),
                prompt=lang.t(
                    "in_kit.enter_module_number",
                    "Enter a unique Module Number.\n\n"
                    "Suggestion: {suggestion}\n\n"
                    "You can edit the suggestion or enter a different number.",
                    suggestion=suggestion
                ),
                initial_value=suggestion  # ✅ Pre-fill with suggestion
            )
            
            # User cancelled
            if entered is None:
                return None
            
            # Validate not empty
            if not entered.strip():
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.warning", "Warning"),
                    lang.t("in_kit.module_number_empty", "Module number cannot be empty."),
                    "warning"
                )
                continue
            
            # ✅ Check uniqueness
            if self.is_module_number_unique(kit_number, entered):
                logging.info(f"[IN_KIT] Module number entered: {entered}")
                return entered
            

            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "receive_kit.duplicate_module_number",
                    "Module Number '{num}' already exists.\n\n"
                    "Please enter a different number.",
                    num=entered
                ),
                "error"
            )
        


    # ------------- Custom text input dialog -------------
    def ask_custom_text(self, title: str, prompt: str, initial_value: str = "") -> str:
        """
        Show a custom dialog to ask for text input.
        
        Args:
            title: Dialog title
            prompt: Prompt message
            initial_value: Pre-filled text (suggestion)
        
        Returns:
            Entered text (stripped), or None if cancelled
        """
        result = {"value": None}
        
        # Create dialog
        dlg = tk.Toplevel(self.parent)
        dlg.title(title)
        dlg.transient(self.parent)
        dlg.grab_set()
        dlg.resizable(False, False)
        
        # Configure style
        dlg.configure(bg="#F0F4F8", padx=20, pady=20)
        
        # Prompt label
        tk.Label(
            dlg, 
            text=prompt, 
            font=("Helvetica", 10),
            bg="#F0F4F8",
            wraplength=400,
            justify="left"
        ).pack(pady=(0, 10))
        
        # Entry field with pre-filled suggestion
        entry_var = tk.StringVar(value=initial_value)
        entry = tk.Entry(
            dlg, 
            textvariable=entry_var,
            font=("Helvetica", 11),
            width=40,
            relief="solid",
            borderwidth=1
        )
        entry.pack(pady=(0, 15), ipady=4)
        entry.focus()
        entry.select_range(0, tk.END)  # Select all text for easy editing
        
        # Button frame
        btn_frame = tk.Frame(dlg, bg="#F0F4F8")
        btn_frame.pack()
        
        def on_ok():
            entered = entry_var.get().strip()
            if entered:
                result["value"] = entered
                dlg.destroy()
        
        def on_cancel():
            result["value"] = None
            dlg.destroy()
        
        # OK button
        tk.Button(
            btn_frame,
            text=lang.t("dialog.ok", "OK"),
            command=on_ok,
            bg="#27AE60",
            fg="white",
            font=("Helvetica", 10, "bold"),
            padx=20,
            pady=5,
            relief="flat",
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        # Cancel button
        tk.Button(
            btn_frame,
            text=lang.t("dialog.cancel", "Cancel"),
            command=on_cancel,
            bg="#7F8C8D",
            fg="white",
            font=("Helvetica", 10),
            padx=20,
            pady=5,
            relief="flat",
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        # Bind keys
        entry.bind("<Return>", lambda e: on_ok())
        entry.bind("<Escape>", lambda e: on_cancel())
        dlg.bind("<Escape>", lambda e: on_cancel())
        
        # Center dialog
        dlg.update_idletasks()
        _center_child(dlg, self.parent)
        
        # Wait for dialog
        dlg.wait_window()
        
        return result["value"]



    # ------------- Hierarchy load -------------
    def load_hierarchy(self, kit_code):
        if not self.selected_scenario_id or not kit_code:
            return
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        self.suggested_usage.clear()
        comps = self.fetch_kit_items(self.selected_scenario_id, kit_code)
        if not comps:
            self.status_var.set(lang.t("receive_kit.no_items", f"No items for {kit_code}"))
            return
        kit_number = self.kit_number_var.get().strip() if self.kit_number_var.get() else None
        if not kit_number:
            # ✅ Generate smart suggestion for kit number
            suggestion = f"{kit_code}-K1" if kit_code else "K1"
            
            while True:
                # ✅ Use custom dialog with pre-filled suggestion
                kn = self.ask_custom_text(
                    title=lang.t("receive_kit.kit_number", "Kit Number"),
                    prompt=lang.t(
                        "in_kit.enter_kit_number",
                        "Enter a unique Kit Number for: {kit_code}\n\n"
                        "Suggestion: {suggestion}\n\n"
                        "You can edit the suggestion or enter a different number.",
                        kit_code=kit_code,
                        suggestion=suggestion
                    ),
                    initial_value=suggestion  # ✅ Pre-fill with suggestion
                )
                
                # User cancelled
                if kn is None:
                    self.status_var.set(
                        lang.t("receive_kit.kit_number_cancelled", "Kit number entry cancelled")
                    )
                    return
                
                # Validate not empty
                if not kn.strip():
                    custom_popup(
                        self.parent,
                        lang.t("dialog_titles.warning", "Warning"),
                        lang.t("in_kit.kit_number_empty", "Kit number cannot be empty."),
                        "warning"
                    )
                    continue
                
                # ✅ Check uniqueness
                if self.is_kit_number_unique(kn):
                    kit_number = kn
                    self.kit_number_var.set(kn)
                    logging.info(f"[IN_KIT] Kit number entered: {kit_number}")
                    break
                
        
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t(
                        "receive_kit.duplicate_kit_number",
                        "Kit Number '{num}' already exists globally.\n\n"
                        "Please enter a different number.",
                        num=kn
                    ),
                    "error"
                )
                # Loop again with same suggestion
        treecode_to_iid = {}
        module_number_map = {}
        for comp in sorted(comps, key=lambda x: x['treecode']):
            tc = comp['treecode']
            parent_tc = tc[:-3] if len(tc) >= 3 else ""
            parent_iid = treecode_to_iid.get(parent_tc, "")
            ctype = comp['type'].upper()
            std_qty = comp['std_qty']
            kit_disp = comp['kit'] or "-----"
            module_disp = comp['module'] or "-----"
            module_number = None
            if ctype == "MODULE":
                module_number = self.ask_module_number(kit_number, comp['code'])
                if module_number is None:
                    self.status_var.set(lang.t("receive_kit.module_number_cancelled","Cancelled"))
                    return
                module_number_map[comp['code']] = module_number
            else:
                if comp['module'] and comp['module'] in module_number_map:
                    module_number = module_number_map[comp['module']]
            row_iid = self.tree.insert(parent_iid, "end", values=(
                comp['code'], comp['description'], comp['type'],
                kit_disp, module_disp,
                std_qty, (std_qty if ctype!="ITEM" else 0), "", "",
                "", ""  # hidden line_id, qty_out
            ))
            treecode_to_iid[tc] = row_iid
            self.code_to_iid[comp['code']] = row_iid
            if ctype == "KIT":
                self.tree.item(row_iid, tags=("kit",))
            elif ctype == "MODULE":
                self.tree.item(row_iid, tags=("module",))
            self.row_data[row_iid] = {
                'kit_number': kit_number,
                'module_number': module_number,
                'treecode': tc
            }
            if ctype == "ITEM":
                self.tree.delete(row_iid)
                del self.row_data[row_iid]
                if comp['code'] in self.code_to_iid:
                    del self.code_to_iid[comp['code']]
                self.insert_item_batches(
                    code=comp['code'],
                    parent_iid=parent_iid,
                    kit_number=kit_number,
                    module_number=module_number,
                    structural_kit=kit_disp if kit_disp != "-----" else None,
                    structural_module=module_disp if module_disp != "-----" else None,
                    std_qty=std_qty
                )
        self.update_child_quantities()
        self.update_parent_expiry()
        self.status_var.set(lang.t("receive_kit.loaded_records", f"Loaded hierarchy for {kit_code}"))

    # ------------- Insert batch rows (with line_id & hidden qty_out) -------------
    def insert_item_batches(self, code, parent_iid, kit_number, module_number,
                            structural_kit, structural_module, std_qty):
        """
        Insert item batches into tree, allocating from longest expiry stock.
        
        ✅ Logic:
            1. Fetch all on-shelf batches (sorted by longest expiry)
            2. Allocate up to std_qty from batches
            3. If one batch insufficient, use multiple batches
            4. Display longest expiry in tree
            5. Track allocation to prevent over-allocation
        
        Args:
            code: Item code
            parent_iid: Parent tree item ID
            kit_number: Kit number (for metadata)
            module_number: Module number (for metadata)
            structural_kit: Kit code (for display)
            structural_module: Module code (for display)
            std_qty: Standard quantity needed
        """
        batches = self.fetch_on_shelf_batches(code)
        desc = get_item_description(code)
        
        # Convert std_qty to integer
        try:
            std_qty_int = int(std_qty) if std_qty else 0
        except (ValueError, TypeError):
            std_qty_int = 0
        
        # If no stock available, insert warning row
        if not batches:
            logging.warning(f"[IN_KIT] No on-shelf stock for {code}")
            iid = self.tree.insert(parent_iid, "end", values=(
                code, desc, "ITEM",
                structural_kit or "-----",
                structural_module or "-----",
                std_qty_int,
                0,  # qty_to_receive = 0
                "",  # expiry_date = empty
                "",  # batch_no = empty
                "",  # line_id (hidden)
                "0"  # qty_out (hidden)
            ), tags=("light_red",))
            
            self.row_data[iid] = {
                'kit_number': kit_number,
                'module_number': module_number,
                'max_qty': 0,
                'management_mode': 'on_shelf',
                'expiry_key': "",
                'line_id': None
            }
            return
        
        # Calculate how much we can/should allocate
        def remaining_std_allowance_for_insertion():
            """Calculate how much more of this item we can allocate (respecting std_qty)."""
            total_assigned = 0
            current_std = std_qty_int
            
            for top in self.tree.get_children():
                stack = [top]
                while stack:
                    iid = stack.pop()
                    vals = self.tree.item(iid, "values")
                    if vals and vals[0] == code and vals[2].upper() == "ITEM":
                        q = vals[6]  # qty_to_receive column
                        if q and str(q).isdigit():
                            total_assigned += int(q)
                    stack.extend(self.tree.get_children(iid))
            
            if current_std <= 0:
                return 10**12  # No limit if std_qty is 0 or invalid
            
            rem = current_std - total_assigned
            return rem if rem > 0 else 0
        
        # Track longest expiry (first batch has it due to sorting)
        longest_expiry = batches[0]['expiry'] if batches else None
        
        # Allocate from batches
        inserted_any = False
        
        for b in batches:
            expiry = b['expiry'] or ""
            final_qty = b['final_qty']
            line_id = b['line_id']
            
            # Check physical availability (after accounting for prior allocations)
            remain_physical = self.remaining_available(code, expiry, b['management_mode'], final_qty)
            
            if remain_physical <= 0:
                continue
            
            # Check standard quantity allowance
            remain_std = remaining_std_allowance_for_insertion()
            
            if remain_std <= 0:
                break  # Already allocated enough
            
            # Allocate minimum of physical and standard allowance
            allocate = min(remain_physical, remain_std)
            
            if allocate <= 0:
                continue
            
            # ✅ Insert tree row with expiry date visible
            iid = self.tree.insert(parent_iid, "end", values=(
                code,
                desc,
                "ITEM",
                structural_kit or "-----",
                structural_module or "-----",
                std_qty_int,
                allocate,  # qty_to_receive
                expiry if expiry else "",  # ✅ EXPIRY DATE VISIBLE
                "",  # batch_no (user can edit)
                str(line_id),  # line_id (hidden)
                str(allocate)  # qty_out (hidden) - mirrors qty_to_receive
            ))
            
            # Store metadata
            self.row_data[iid] = {
                'kit_number': kit_number,
                'module_number': module_number,
                'max_qty': final_qty,
                'management_mode': b['management_mode'],
                'expiry_key': expiry,
                'line_id': line_id
            }
            
            # Record allocation to prevent over-use
            self.record_suggested(code, expiry, b['management_mode'], allocate)
            
            inserted_any = True
            
            logging.debug(
                f"[IN_KIT] Allocated {allocate} of {code} from batch "
                f"(expiry: {expiry}, line_id: {line_id})"
            )
        
        # If no batches were inserted (shouldn't happen if batches exist), add warning row
        if not inserted_any:
            logging.warning(f"[IN_KIT] No batches allocated for {code} (likely allocation issue)")
            iid = self.tree.insert(parent_iid, "end", values=(
                code, desc, "ITEM",
                structural_kit or "-----",
                structural_module or "-----",
                std_qty_int,
                0,  # qty_to_receive = 0
                longest_expiry if longest_expiry else "",  # Show longest available expiry
                "",
                "",  # line_id
                "0"  # qty_out
            ), tags=("light_red",))
            
            self.row_data[iid] = {
                'kit_number': kit_number,
                'module_number': module_number,
                'max_qty': 0,
                'management_mode': 'on_shelf',
                'expiry_key': longest_expiry or "",
                'line_id': None
            }
    # ------------- Add single code -------------
    def add_to_tree(self, code):
        mk = self.current_mode_key()
        desc = get_item_description(code)
        item_type = detect_type(code, desc).upper()
        kit_code = self.kit_var.get() if mk in ["add_module_kit","add_items_kit","add_items_module"] else None
        module_code = self.module_var.get() if mk == "add_items_module" else None
        kit_number = self.kit_number_var.get().strip() if self.kit_number_var.get() else None
        module_number = self.module_number_var.get().strip() if self.module_number_var.get() else None
        if mk == "add_module_kit":
            if not kit_code or not kit_number:
                custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                             lang.t("receive_kit.no_kit_number","Please select a Kit & Kit Number"),"error")
                return
        elif mk == "add_items_kit":
            if not kit_number:
                custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                             lang.t("receive_kit.no_kit_number","Please select a Kit Number"),"error")
                return
        elif mk == "add_items_module":
            if not module_code or not module_number:
                custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                             lang.t("receive_kit.no_module_number","Please select a Module & Module Number"),"error")
                return
        scenario_module_mode = (mk == "add_module_scenario")
        if item_type in ["KIT","MODULE"]:
            if item_type == "MODULE" and mk in ["add_module_kit","add_module_scenario"]:
                mn = self.ask_module_number(kit_number, code)
                if mn is None:
                    self.status_var.set(lang.t("receive_kit.module_number_cancelled","Cancelled"))
                    return
                module_number = mn
            iid = self.tree.insert("", "end", values=(
                code, desc, item_type,
                kit_code or "-----",
                module_code or "-----",
                0, 1, "", "",
                "", ""  # hidden
            ))
            if item_type == "KIT":
                self.tree.item(iid, tags=("kit",))
            else:
                self.tree.item(iid, tags=("module",))
            self.row_data[iid] = {
                'kit_number': kit_number if not scenario_module_mode else None,
                'module_number': module_number,
                'treecode': None
            }
            self.status_var.set(lang.t("receive_kit.added_item", f"Added {item_type} {code}"))
            return
        structural_kit = None if scenario_module_mode else kit_code
        structural_module = module_code
        self.insert_item_batches(
            code=code,
            parent_iid="",
            kit_number=kit_number,
            module_number=module_number,
            structural_kit=structural_kit,
            structural_module=structural_module,
            std_qty=0
        )
        self.status_var.set(lang.t("receive_kit.added_item", f"Added ITEM {code} (batches)"))

    # ------------- Add missing item dialog -------------
    def add_missing_item(self):
        if not self.selected_scenario_id:
            custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                         lang.t("receive_kit.no_scenario","Please select a scenario"),"error")
            return
        dlg = tk.Toplevel(self.parent)
        dlg.title(lang.t("receive_kit.add_missing","Add Missing Item"))
        dlg.geometry("560x360")
        dlg.transient(self.parent)
        dlg.grab_set()
        tk.Label(dlg, text=lang.t("receive_kit.search_item","Search Kit/Module/Item"),
                 font=("Helvetica",10,"bold")).pack(pady=6)
        sv = tk.StringVar()
        entry = tk.Entry(dlg, textvariable=sv, font=("Helvetica",10), width=60)
        entry.pack(padx=10, pady=4, fill=tk.X)
        lb = tk.Listbox(dlg, font=("Helvetica",10), height=14,
                        selectbackground="#2563EB", selectforeground="white")
        lb.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        def refresh(*_):
            lb.delete(0, tk.END)
            mk = self.current_mode_key()
            res = self.fetch_search_results(sv.get().strip(), self.selected_scenario_id, mk)
            for r in res:
                lb.insert(tk.END, f"{r['code']} - {r['description']}")
            if not res:
                lb.insert(tk.END, lang.t("receive_kit.no_items_found","No items found"))
        sv.trace_add("write", refresh)
        refresh()
        def choose(_=None):
            idxs = lb.curselection()
            if idxs:
                line = lb.get(idxs[0])
                c = line.split(" - ")[0]
                dlg.destroy()
                mk = self.current_mode_key()
                if mk == "receive_kit":
                    self.load_hierarchy(c)
                else:
                    self.add_to_tree(c)
            else:
                dlg.destroy()
        lb.bind("<Double-Button-1>", choose)
        entry.bind("<Return>", choose)
        entry.focus()
        _center_child(dlg, self.parent)
        dlg.wait_window()

    # ------------- Editing -------------
    def start_edit(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_index = int(col_id.replace("#","")) - 1
        # qty_to_receive(6), expiry_date(7), batch_no(8) only
        if col_index not in [6,7,8]:
            return
        self._start_edit_cell(row_id, col_index)

    def _start_edit_cell(self, row_id, col_index):
        try:
            bbox = self.tree.bbox(row_id, f"#{col_index+1}")
            if not bbox:
                return
            x,y,w,h = bbox
            orig_val = self.tree.set(row_id, self.tree["columns"][col_index])
            entry = tk.Entry(self.tree, font=("Helvetica",10))
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, orig_val)
            entry.focus()
            self.editing_cell = entry
            def finalize(_=None):
                new_val = entry.get().strip()
                if col_index == 6:
                    new_qty = int(new_val) if new_val.isdigit() else 0
                    t = self.tree.set(row_id, "type").upper()
                    if t == "ITEM":
                        code = self.tree.set(row_id, "code")
                        expiry = self.tree.set(row_id, "expiry_date") or ""
                        meta = self.row_data.get(row_id, {})
                        mgmt_mode = meta.get('management_mode')
                        if mgmt_mode == 'on_shelf':
                            max_qty = meta.get('max_qty')
                            key = (code, expiry, mgmt_mode)
                            prev_used = self.suggested_usage.get(key, 0)
                            old_qty = int(orig_val) if orig_val.isdigit() else 0
                            used_minus_self = prev_used - old_qty
                            physical_remaining = (max_qty if max_qty is not None else new_qty) - used_minus_self
                            physical_remaining = max(physical_remaining, 0)
                            if new_qty > physical_remaining:
                                new_qty = physical_remaining
                            std_allow = self._remaining_std_allowance(code) + old_qty
                            if new_qty > std_allow:
                                new_qty = std_allow
                            if new_qty < 0: new_qty = 0
                            self.suggested_usage[key] = used_minus_self + new_qty
                    self.tree.set(row_id, "qty_to_receive", str(new_qty))
                    # Mirror into hidden qty_out
                    self.tree.set(row_id, "qty_out", str(new_qty))
                    if t in ["KIT","MODULE"]:
                        self.update_child_quantities()
                elif col_index == 7:
                    if new_val:
                        parsed = parse_expiry(new_val)
                        self.tree.set(row_id, "expiry_date", parsed if parsed else "")
                    else:
                        self.tree.set(row_id, "expiry_date", "")
                elif col_index == 8:
                    if len(new_val) > 30:
                        new_val = new_val[:30]
                    self.tree.set(row_id, "batch_no", new_val)
                entry.destroy()
                self.editing_cell = None
                self.update_parent_expiry()
            entry.bind("<Return>", finalize)
            entry.bind("<Tab>", finalize)
            entry.bind("<FocusOut>", finalize)
            entry.bind("<Escape>", lambda e: (entry.destroy(), setattr(self, "editing_cell", None)))
        except Exception as e:
            logging.error(f"[start_edit_cell] {e}")

    # ------------- Quantity propagation -------------
    def update_child_quantities(self):
        try:
            for iid in self.tree.get_children():
                self._update_subtree_quantities(iid)
        except Exception as e:
            logging.error(f"[update_child_quantities] {e}")

    def _update_subtree_quantities(self, iid):
        vals = self.tree.item(iid, "values")
        if not vals: return
        t = vals[2].upper()
        if t in ["KIT","MODULE"]:
            for child in self.tree.get_children(iid):
                self._update_subtree_quantities(child)

    # ------------- Expiry propagation -------------
    def update_parent_expiry(self):
        try:
            for iid in self.tree.get_children():
                self._update_parent_expiry_recursive(iid)
        except Exception as e:
            logging.error(f"[update_parent_expiry] {e}")

    def _update_parent_expiry_recursive(self, iid):
        vals = self.tree.item(iid, "values")
        if not vals: return
        t = vals[2].upper()
        child_exp = []
        for c in self.tree.get_children(iid):
            self._update_parent_expiry_recursive(c)
            cv = self.tree.item(c,"values")
            if not cv: continue
            if cv[2].upper()=="ITEM":
                p = parse_expiry(cv[7])
                if p: child_exp.append(p)
            else:
                cp = parse_expiry(cv[7])
                if cp: child_exp.append(cp)
        if t in ["KIT","MODULE"] and child_exp:
            self.tree.set(iid, "expiry_date", min(child_exp))
        code = vals[0]
        qty = vals[6]
        expv = self.tree.set(iid, "expiry_date")
        if qty and str(qty).isdigit() and int(qty)>0 and check_expiry_required(code) and not parse_expiry(expv):
            self.tree.item(iid, tags=("light_red",))
        else:
            if t == "KIT":
                self.tree.item(iid, tags=("kit",))
            elif t == "MODULE":
                self.tree.item(iid, tags=("module",))

    # ------------- Navigation & Context Menu -------------
    def navigate_tree(self, event):
        if self.editing_cell:
            return
        all_rows = []
        def collect(p=""):
            for r in self.tree.get_children(p):
                all_rows.append(r)
                collect(r)
        collect()
        if not all_rows:
            return
        sel = self.tree.selection()
        current = sel[0] if sel else all_rows[0]
        if current not in all_rows:
            current = all_rows[0]
        idx = all_rows.index(current)
        if event.keysym == "Up" and idx > 0:
            self.tree.selection_set(all_rows[idx-1]); self.tree.focus(all_rows[idx-1])
        elif event.keysym == "Down" and idx < len(all_rows)-1:
            self.tree.selection_set(all_rows[idx+1]); self.tree.focus(all_rows[idx+1])

    def show_context_menu(self, event):
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(label=lang.t("receive_kit.add_new_row","Add New Row"),
                         command=lambda: None)
        menu.post(event.x_root, event.y_root)

    # ------------- Uniqueness helpers -------------
    def is_kit_number_unique(self, kit_number: str) -> bool:
        if not kit_number or kit_number.lower()=="none":
            return False
        conn = connect_db()
        if conn is None: return False
        cur = conn.cursor()
        try:
            cur.execute("""SELECT COUNT(*) FROM stock_data
                           WHERE kit_number=? AND kit_number!='None'""",(kit_number.strip(),))
            return cur.fetchone()[0] == 0
        finally:
            cur.close()
            conn.close()

    def is_module_number_unique(self, kit_number: str, module_number: str) -> bool:
        if not module_number or module_number.lower()=="none":
            return False
        conn = connect_db()
        if conn is None: return False
        cur = conn.cursor()
        try:
            cur.execute("""SELECT COUNT(*) FROM stock_data
                           WHERE module_number=? AND module_number!='None'""",(module_number.strip(),))
            return cur.fetchone()[0] == 0
        finally:
            cur.close()
            conn.close()

    def generate_unique_id(self, scenario_id, kit, module, item, std_qty, exp_date, kit_number, module_number):
        return f"{scenario_id}/{kit or 'None'}/{module or 'None'}/{item or 'None'}/{std_qty}/{exp_date or 'None'}/{kit_number or 'None'}/{module_number or 'None'}"

    # ------------- Document & Logging -------------
    def generate_document_number(self, in_type_text: str) -> str:
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").upper()
        base_map = {
            "In MSF":"IMSF","In Local Purchase":"ILP","In from Quarantine":"IFQ",
            "In Donation":"IDN","Return from End User":"IREU","In Supply Non-MSF":"ISNM",
            "In Borrowing":"IBR","In Return of Loan":"IRL","In Correction of Previous Transaction":"ICOR",
            "In from on-shelf items":"INTERNAL"
        }
        raw = (in_type_text or "").strip()
        abbr = None
        norm_raw = re.sub(r'[^a-z0-9]+','', raw.lower())
        for k,v in base_map.items():
            if re.sub(r'[^a-z0-9]+','', k.lower()) == norm_raw:
                abbr = v; break
        if not abbr:
            tokens = re.split(r'\s+', raw.upper())
            stop = {"OF","FROM","THE","AND","DE","DU","DES","LA","LE","LES"}
            letters=[]
            for t in tokens:
                if not t or t in stop: continue
                if t=="MSF": letters.append("MSF")
                else: letters.append(t[0])
            abbr = "".join(letters) or (raw[:4].upper() or "DOC")
            abbr = abbr[:8]
        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"
        conn = connect_db()
        serial = 1
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT document_number FROM stock_transactions
                     WHERE document_number LIKE ?
                     ORDER BY document_number DESC LIMIT 1
                """,(prefix+"/%",))
                r = cur.fetchone()
                if r and r[0]:
                    last = r[0].rsplit('/',1)[-1]
                    if last.isdigit():
                        serial = int(last)+1
            finally:
                cur.close()
                conn.close()
        doc = f"{prefix}/{serial:04d}"
        self.current_document_number = doc
        return doc

    def log_transaction(self,
                        unique_id: str,
                        code: str,
                        description: str,
                        expiry_date: str,
                        batch_number: str,
                        scenario: str,
                        kit: str,
                        module: str,
                        qty_in,
                        in_type: str,
                        qty_out,
                        out_type: str,
                        third_party: str,
                        end_user: str,
                        remarks: str,
                        movement_type: str,
                        document_number: str):
        """
        Insert a single stock transaction row.

        Note: Per your new requirement, save_all() now calls this twice
        for each ITEM row:
          1) Incoming (qty_in > 0, qty_out = None, in_type set, out_type None)
          2) Outgoing mirror (qty_in = None, qty_out > 0, in_type None, out_type = original in_type)

        Parameters may be None; they are written as NULL in SQLite.
        """
        conn = connect_db()
        if conn is None:
            return
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO stock_transactions
                (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
                 Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
                 Third_Party, End_User, Remarks, Movement_Type, document_number)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,(
                datetime.today().strftime('%Y-%m-%d'),
                datetime.now().strftime('%H:%M:%S'),
                unique_id,
                code,
                description,
                expiry_date,
                batch_number,
                scenario,
                kit,
                module,
                qty_in,
                in_type,
                qty_out,
                out_type,
                third_party,
                end_user,
                remarks,
                movement_type,
                document_number
            ))
            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"[log_transaction] {e}")
        finally:
            cur.close()
            conn.close()


    # ------------- Uniqueness enforcement -------------
    def _enforce_unique_numbers_before_save(self):
        kits = {}
        modules = {}
        for iid, meta in self.row_data.items():
            vals = self.tree.item(iid, "values")
            if not vals: continue
            t = (vals[2] or "").upper()
            kn = meta.get('kit_number')
            mn = meta.get('module_number')
            if t == "KIT" and kn and kn.lower() != 'none':
                kits.setdefault(kn, []).append(iid)
            if t == "MODULE" and mn and mn.lower() != 'none':
                modules.setdefault(mn, []).append(iid)
        dup_kits = {k: v for k,v in kits.items() if len(v) > 1}
        dup_modules = {m: v for m,v in modules.items() if len(v) > 1}
        if not dup_kits and not dup_modules:
            return True
        for kit_number, iids in dup_kits.items():
            for iid in iids[1:]:
                while True:
                    new_kn = simpledialog.askstring(
                        lang.t("receive_kit.kit_number","Kit Number"),
                        f"Duplicate Kit Number '{kit_number}'. Enter a new Kit Number:",
                        parent=self.parent
                    )
                    if new_kn is None:
                        custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                                     f"Save aborted (duplicate Kit Number {kit_number}).","error")
                        return False
                    new_kn = new_kn.strip()
                    if not new_kn:
                        continue
                    if self.is_kit_number_unique(new_kn):
                        self.row_data[iid]['kit_number'] = new_kn
                        break
                    custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                                 f"Kit Number '{new_kn}' already exists globally.","error")
        for mod_number, iids in dup_modules.items():
            for iid in iids[1:]:
                while True:
                    new_mn = simpledialog.askstring(
                        lang.t("receive_kit.module_number","Module Number"),
                        f"Duplicate Module Number '{mod_number}'. Enter a new Module Number:",
                        parent=self.parent
                    )
                    if new_mn is None:
                        custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                                     f"Save aborted (duplicate Module Number {mod_number}).","error")
                        return False
                    new_mn = new_mn.strip()
                    if not new_mn:
                        continue
                    if self.is_module_number_unique(None, new_mn):
                        self.row_data[iid]['module_number'] = new_mn
                        break
                    custom_popup(self.parent, lang.t("dialog_titles.error","Error"),
                                 f"Module Number '{new_mn}' already exists globally.","error")
        return True

    # ------------- 70% rule -------------
    def _validate_global_70_rule(self):
        per_code = {}
        for top in self.tree.get_children():
            stack=[top]
            while stack:
                iid=stack.pop()
                vals=self.tree.item(iid,"values")
                if vals:
                    code, _, t, _, _, std_qty, qty, _, _, _, _ = vals
                    if (t or "").upper()=="ITEM":
                        std = int(std_qty) if str(std_qty).isdigit() else 0
                        issued = int(qty) if str(qty).isdigit() else 0
                        if code not in per_code:
                            per_code[code]={'std':std,'issued':issued}
                        else:
                            per_code[code]['issued'] += issued
                            if std > per_code[code]['std']:
                                per_code[code]['std'] = std
                stack.extend(self.tree.get_children(iid))
        if not per_code:
            return True
        total = len(per_code)
        covered = 0
        uncovered=[]
        for c,st in per_code.items():
            if st['std']<=0 or st['issued']>=st['std']:
                covered+=1
            else:
                uncovered.append((c, st['std'], st['issued']))
        coverage = covered/total if total else 1
        if coverage >= 0.70:
            return True
        uncovered.sort(key=lambda x: (x[1]-x[2]), reverse=True)
        sample = "\n".join([f"{c}: {issd}/{std} ({(issd/std*100):.1f}%)" for c,std,issd in uncovered[:15]])
        more = f"\n... {len(uncovered)-15} more." if len(uncovered)>15 else ""
        mode_key = self.current_mode_key()
        context = {
            "receive_kit":"kit",
            "add_module_kit":"module",
            "add_module_scenario":"module",
            "add_items_kit":"kit",
            "add_items_module":"module",
            "add_standalone":"operation"
        }.get(mode_key,"operation")
        msg = (f"At least 70% of items must have enough stock. Current: {coverage*100:.2f}% "
               f"({covered}/{total}). The {context} cannot be created.\n\nBelow standard (sample):\n"
               f"{sample}{more}")
        custom_popup(self.parent, lang.t("dialog_titles.error","Error"), msg, "error")
        return False

    # ------------- Save logic -------------
    def save_all(self):
        """
        Save current tree (kit/module/items) into stock_data and log transactions.

        ✅ Updated behavior:
          - Consumes existing on-shelf batches (increases qty_out on original lines)
          - Adds new stock_data lines for built kit/module/item composition (qty_in)
          - Uses scenario_id (not scenario_name) for database storage
          - Logs TWO transaction rows per ITEM with qty_to_receive > 0:
              1) Incoming:  Qty_IN = qty_to_receive, IN_Type = trans_type_var value
              2) Outgoing mirror: Qty_Out = hidden 'qty_out' column value,
                 Out_Type = original IN_Type value, Qty_IN = NULL
        """
        self.ensure_vars_ready()
        
        # ===== PERMISSION CHECK =====
        if self.role not in ["admin", "manager"]:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("receive_kit.no_permission", 
                       "Only admin or manager roles can save changes."),
                "error"
            )
            return
        
        # ===== VALIDATE DATA EXISTS =====
        if not self.tree.get_children():
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("receive_kit.no_rows", "No rows to save."),
                "error"
            )
            return

        # ===== GET IN TYPE =====
        in_type = self.trans_type_var.get()
        if not in_type:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                "Internal error: IN Type missing.",
                "error"
            )
            return

        # ===== VALIDATE UNIQUE NUMBERS =====
        if not self._enforce_unique_numbers_before_save():
            return

        # ===== VALIDATE 70% COVERAGE RULE =====
        if not self._validate_global_70_rule():
            return

        # ===== GENERATE DOCUMENT NUMBER =====
        doc = self.generate_document_number(in_type)
        
        # ✅ Get scenario_id (for database storage)
        scenario_id = self.selected_scenario_id
        
        if not scenario_id:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                "No scenario selected. Cannot save.",
                "error"
            )
            return
        
        # Get scenario_name (for display/logging purposes)
        scenario_name = self.scenario_map.get(scenario_id, "Unknown")

        saved = 0
        errors = []
        all_iids = []

        # ===== COLLECT ALL TREE ITEMS =====
        def collect(root=""):
            for r in self.tree.get_children(root):
                all_iids.append(r)
                collect(r)
        collect()

        # ===== STEP 1: CONSUME EXISTING STOCK LINES =====
        consumed_lines = set()
        
        for iid in all_iids:
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            
            (code, description, t, kit_col, module_col, std_qty,
             qty_to_receive, expiry_date, batch_no, line_id, qty_out_hidden) = vals
            
            if (t or "").upper() != "ITEM":
                continue
            
            if not line_id or not line_id.isdigit():
                continue
            
            if not qty_out_hidden or not qty_out_hidden.isdigit():
                continue
            
            consume_qty = int(qty_out_hidden)
            
            if consume_qty <= 0:
                continue
            
            if line_id in consumed_lines:
                continue
            
            # ✅ Update existing stock_data line's qty_out
            try:
                StockData.consume_by_line_id(int(line_id), consume_qty)
                consumed_lines.add(line_id)
                logging.info(f"[IN_KIT] Consumed {consume_qty} from line_id {line_id} for {code}")
            except Exception as e:
                logging.error(f"[IN_KIT] Failed to consume line_id {line_id}: {e}")

        # ===== STEP 2: INSERT NEW LINES & LOG TRANSACTIONS =====
        for iid in all_iids:
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            
            (code, description, t, kit_col, module_col, std_qty,
             qty_to_receive, expiry_date, batch_no, line_id, qty_out_hidden) = vals

            if not qty_to_receive or not str(qty_to_receive).isdigit():
                continue
            
            qty_to_receive = int(qty_to_receive)
            
            if qty_to_receive <= 0:
                continue

            # ===== PARSE EXPIRY DATE =====
            parsed_exp = parse_expiry(expiry_date) if expiry_date else None

            # ===== VALIDATE EXPIRY IF REQUIRED =====
            if (t or "").upper() == "ITEM" and check_expiry_required(code):
                if not parsed_exp:
                    self.tree.item(iid, tags=("light_red",))
                    errors.append(code)
                    logging.warning(f"[IN_KIT] Missing expiry for {code} (required)")
                    continue

            # ===== GET METADATA =====
            meta = self.row_data.get(iid, {})
            kit_number = meta.get('kit_number')
            module_number = meta.get('module_number')

            # ===== BUILD UNIQUE_ID COMPONENTS =====
            kit_struct = kit_col if kit_col and kit_col != "-----" else None
            module_struct = module_col if module_col and module_col != "-----" else None
            item_part = code if (t or "").upper() == "ITEM" else None
            std_numeric = int(std_qty) if str(std_qty).isdigit() else 0

            # ===== GENERATE UNIQUE_ID (8-PART FORMAT) =====
            unique_id = self.generate_unique_id(
                scenario_id,
                kit_struct,
                module_struct,
                item_part,
                std_numeric,
                parsed_exp,
                kit_number,
                module_number
            )

            # ===== ADD/UPDATE STOCK_DATA (NEW COMPOSITION LINE) =====
            try:
                StockData.add_or_update(
                    unique_id,
                    scenario=str(scenario_id),
                    qty_in=qty_to_receive,
                    exp_date=parsed_exp,
                    kit_number=kit_number,
                    module_number=module_number
                )
                logging.info(
                    f"[IN_KIT] Added stock_data: {unique_id}, "
                    f"qty_in={qty_to_receive}, scenario={scenario_id}"
                )
            except Exception as e:
                logging.error(f"[IN_KIT] Failed to add stock_data for {code}: {e}")
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    f"Failed to save {code}: {str(e)}",
                    "error"
                )
                continue

            # ===== LOG TRANSACTION 1: INCOMING (QTY_IN) =====
            try:
                self.log_transaction(
                    unique_id=unique_id,
                    code=code,
                    description=description,
                    expiry_date=parsed_exp,
                    batch_number=batch_no or None,
                    scenario=str(scenario_id),
                    kit=kit_number,
                    module=module_number,
                    qty_in=qty_to_receive,
                    in_type=in_type,
                    qty_out=None,
                    out_type=None,
                    third_party=self.third_party_var.get() or None,
                    end_user=self.end_user_var.get() or None,
                    remarks=self.remarks_entry.get().strip() if self.remarks_entry else None,
                    movement_type=self.mode_var.get() or "stock_in_kit",
                    document_number=doc
                )
                logging.debug(f"[IN_KIT] Logged incoming transaction for {code}")
            except Exception as e:
                logging.error(f"[IN_KIT] Failed to log incoming transaction for {code}: {e}")

            # ===== LOG TRANSACTION 2: OUTGOING MIRROR (QTY_OUT) =====
            if qty_out_hidden and str(qty_out_hidden).isdigit():
                out_q = int(qty_out_hidden)
                
                if out_q > 0:
                    try:
                        self.log_transaction(
                            unique_id=unique_id,
                            code=code,
                            description=description,
                            expiry_date=parsed_exp,
                            batch_number=batch_no or None,
                            scenario=str(scenario_id),
                            kit=kit_number,
                            module=module_number,
                            qty_in=None,
                            in_type=None,
                            qty_out=out_q,
                            out_type=in_type,
                            third_party=self.third_party_var.get() or None,
                            end_user=self.end_user_var.get() or None,
                            remarks=self.remarks_entry.get().strip() if self.remarks_entry else None,
                            movement_type=self.mode_var.get() or "stock_in_kit",
                            document_number=doc
                        )
                        logging.debug(f"[IN_KIT] Logged outgoing mirror transaction for {code}")
                    except Exception as e:
                        logging.error(f"[IN_KIT] Failed to log outgoing transaction for {code}: {e}")

            saved += 1

        # ===== HANDLE ERRORS =====
        if errors:
            error_list = ', '.join(set(errors))
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("receive_kit.invalid_expiry",
                       "Valid expiry dates are required for these items:\n{items}",
                       items=error_list),
                "error"
            )
            return

        # ===== SUCCESS MESSAGE =====
        custom_popup(
            self.parent,
            lang.t("dialog_titles.success", "Success"),
            lang.t("receive_kit.save_success",
                   "Kit/Module received successfully.\n\n"
                   "Items saved: {saved}\n"
                   "Transactions logged: {transactions}\n"
                   "Document Number: {doc}",
                   saved=saved,
                   transactions=saved * 2,
                   doc=doc),
            "info"
        )

        self.status_var.set(
            lang.t("receive_kit.document_number_generated",
                   "Saved {saved} items ({transactions} transactions). Doc: {doc}",
                   saved=saved,
                   transactions=saved * 2,
                   doc=doc)
        )
        
        # ===== CLEAR FORM =====
        self.clear_form()
        

    # ------------- Export -------------
    def export_data(self, rows_to_export=None):
        if not self.tree.get_children():
            self.status_var.set(lang.t("receive_kit.export_cancelled","Export cancelled"))
            return
        default_dir = "D:/ISEPREP"
        if not os.path.exists(default_dir):
            try: os.makedirs(default_dir)
            except: pass
        current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        in_type_raw = self.trans_type_var.get() or lang.t("receive_kit.unknown","Unknown")
        movement_type_raw = self.mode_var.get() or lang.t("receive_kit.unknown","Unknown")
        doc_number = getattr(self, "current_document_number", "")
        safe = lambda s: (re.sub(r'[^A-Za-z0-9]+','_', s or '') or "Unknown").strip('_')
        file_name = f"in_kit_{safe(movement_type_raw)}_{safe(in_type_raw)}_{current_time}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=file_name,
            initialdir=default_dir
        )
        if not path:
            self.status_var.set(lang.t("receive_kit.export_cancelled","Export cancelled"))
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws_title = lang.t("receive_kit.title","Receive Kit-Module")
        ws.title = ws_title[:31]
        project_name, project_code = fetch_project_details()
        ws['A1'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    Document Number: {doc_number}"
        ws['A2'] = f"{ws_title} – Movement: {movement_type_raw}"
        ws['A3'] = f"{project_name} - {project_code}"
        ws['A4'] = f"{lang.t('receive_kit.in_type','IN Type')}: {in_type_raw}"
        ws['A5'] = f"{lang.t('receive_kit.movement_type','Movement Type')}: {movement_type_raw}"
        ws.append([])
        headers = [
            lang.t("receive_kit.code","Code"),
            lang.t("receive_kit.description","Description"),
            lang.t("receive_kit.type","Type"),
            lang.t("receive_kit.kit","Kit"),
            lang.t("receive_kit.module","Module"),
            lang.t("receive_kit.std_qty","Std Qty"),
            lang.t("receive_kit.qty_to_receive","Qty to Receive"),
            lang.t("receive_kit.expiry_date","Expiry Date"),
            lang.t("receive_kit.batch_no","Batch No"),
            "line_id",
            "qty_out_consumed"
        ]
        ws.append(headers)
        def append_rows_recursive(iid):
            vals = self.tree.item(iid,"values")
            if vals:
                qty = vals[6]
                if qty and str(qty).isdigit() and int(qty)>0 and vals[2].upper()=="ITEM":
                    ws.append(list(vals))
            for c in self.tree.get_children(iid):
                append_rows_recursive(c)
        for iid in self.tree.get_children():
            append_rows_recursive(iid)
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len+2, 55)
        wb.save(path)
        custom_popup(self.parent, lang.t("dialog_titles.success","Success"),
                     lang.t("receive_kit.export_success", f"Export successful: {path}"), "info")
        self.status_var.set(lang.t("receive_kit.export_success", f"Export successful: {path}"))

# ------------- Standalone harness -------------
if __name__ == "__main__":
    root = tk.Tk()
    root.title("In Kit (On-Shelf Batches)")
    class DummyApp: pass
    DummyApp.role = "admin"
    app = DummyApp()
    StockInKit(root, app, role="admin")
    root.geometry("1400x850")
    root.mainloop()