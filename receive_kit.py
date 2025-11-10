"""
receive_kit.py (Full Consolidated) v1.74
Implements:
  - Unified expiry parsing & strict validation
  - Kit / Module / Item hierarchy receives & edits
  - Editable "Qty to Receive" for all rows
  - Group earliest expiry (exp_module / exp_kit) + auto/adopt logic
  - Comments column logic (only when expiry matches adopted group earliest)
  - Scenario-level module grouping (modules without kits)
  - Interactive uniqueness resolution (kit_number / module_number)
  - Transaction logging & stock_data persistence
  - Excel export with Comments last
  - Context menus, duplication, adoption flows
Changes from previous version:
  * Qty to Receive editable for all rows
  * "Comment" column renamed to "Comments"
  * Comments blank for manual distinct expiry; only set when expiry matches exp_module or exp_kit earliest
  * Added module_groups_nokit for scenario-level modules (no kit) to produce exp_module values
  * Recalculated comments after expiry edits / recompute
  * Export order updated (Comments last)
"""
from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog
from tkinter import messagebox as mb
import sqlite3
import re
import os
import logging
import calendar as _cal
from calendar import monthrange
from datetime import datetime as _dt, datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
# External project modules
from db import connect_db
from manage_items import get_item_description, detect_type
from kits_Composition import KitsComposition # retained (even if unused) for compatibility
from language_manager import lang
from popup_utils import custom_popup, custom_askyesno, custom_dialog
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
# ---------------------------------------------------------------------
# UI Utility: Center Toplevel
# ---------------------------------------------------------------------
def _center_child(win: tk.Toplevel, parent: tk.Widget | None):
    try:
        win.update_idletasks()
        if parent and parent.winfo_exists():
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            w, h = win.winfo_width(), win.winfo_height()
            x = px + (pw // 2) - (w // 2)
            y = py + (ph // 2) - (h // 2)
        else:
            sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
            w, h = win.winfo_width(), win.winfo_height()
            x = (sw // 2) - (w // 2)
            y = (sh // 2) - (h // 2)
        if x < 0: x = 0
        if y < 0: y = 0
        win.geometry(f"+{x}+{y}")
    except Exception:
        pass
# Center simpledialog askstring
import tkinter.simpledialog as _sd
if not hasattr(_sd, "_CenteredQueryString"):
    _OrigQ = _sd._QueryString
    class _CenteredQueryString(_OrigQ):
        def body(self, master):
            r = super().body(master)
            try: _center_child(self, self.parent)
            except Exception: pass
            return r
    _sd._QueryString = _CenteredQueryString
# Simple wrappers
def win_popup(kind: str, title: str, message: str):
    if kind == "error":
        mb.showerror(title, message)
    elif kind == "warning":
        mb.showwarning(title, message)
    else:
        mb.showinfo(title, message)
def win_confirm(title: str, message: str) -> bool:
    return mb.askyesno(title, message)
# ---------------------------------------------------------------------
# EXPIRY PARSING
# ---------------------------------------------------------------------
_PARSE_EXPIRY_PATTERNS = [
    ("dmy_slash", re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{4})$')),
    ("ymd_slash", re.compile(r'^(\d{4})/(\d{1,2})/(\d{1,2})$')),
    ("my_slash", re.compile(r'^(\d{1,2})/(\d{4})$')),
    ("ym_slash", re.compile(r'^(\d{4})/(\d{1,2})$')),
    ("iso_dash", re.compile(r'^(\d{4})-(\d{2})-(\d{2})$')),
    ("dmy_dash", re.compile(r'^(\d{1,2})-(\d{1,2})-(\d{4})$')),
]
def parse_expiry(date_str: str) -> str | None:
    if not date_str:
        return None
    raw = str(date_str).replace("(adopted)", "").strip()
    if not raw or raw.lower() == "none":
        return None
    norm_slash = re.sub(r'\s+', ' ', raw)
    for tag, rx in _PARSE_EXPIRY_PATTERNS:
        m = rx.match(raw) if 'dash' in tag else rx.match(norm_slash)
        if not m:
            continue
        try:
            if tag == "iso_dash":
                y, mo, d = map(int, m.groups()); return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag == "dmy_dash":
                d, mo, y = map(int, m.groups()); return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag == "dmy_slash":
                d, mo, y = map(int, m.groups()); return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag == "ymd_slash":
                y, mo, d = map(int, m.groups()); return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag == "my_slash":
                mo, y = map(int, m.groups()); ld = monthrange(y, mo)[1]; return f"{y:04d}-{mo:02d}-{ld:02d}"
            if tag == "ym_slash":
                y, mo = map(int, m.groups()); ld = monthrange(y, mo)[1]; return f"{y:04d}-{mo:02d}-{ld:02d}"
        except ValueError:
            return None
    # Fallback display format (DD-Mon-YYYY)
    try:
        dt = _dt.strptime(raw, "%d-%b-%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None
_EXP_PATTERNS = [
    ("dmy_slash", re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{4})$')),
    ("dmy_dash", re.compile(r'^(\d{1,2})-(\d{1,2})-(\d{4})$')),
    ("ymd_slash", re.compile(r'^(\d{4})/(\d{1,2})/(\d{1,2})$')),
    ("ymd_dash", re.compile(r'^(\d{4})-(\d{1,2})-(\d{1,2})$')),
    ("my_slash", re.compile(r'^(\d{1,2})/(\d{4})$')),
    ("ym_slash", re.compile(r'^(\d{4})/(\d{1,2})$')),
    ("my_dash", re.compile(r'^(\d{1,2})-(\d{4})$')),
    ("ym_dash", re.compile(r'^(\d{4})-(\d{1,2})$')),
    ("dmy_dot", re.compile(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$')),
    ("my_dot", re.compile(r'^(\d{1,2})\.(\d{4})$')),
    ("ym_dot", re.compile(r'^(\d{4})\.(\d{1,2})$')),
]
_DISPLAY_DATE_FORMAT = "%d-%b-%Y"
DISPLAY_DATE_FORMAT = _DISPLAY_DATE_FORMAT # legacy alias
def strict_parse_expiry(user_text: str) -> str | None:
    if user_text is None:
        return None
    raw = user_text.replace("(adopted)", "").strip()
    if raw == "":
        return None
    base = parse_expiry(raw)
    if base:
        return base
    for tag, rx in _EXP_PATTERNS:
        m = rx.match(raw)
        if not m: continue
        try:
            if tag in ("dmy_slash","dmy_dash","dmy_dot"):
                d, mo, y = map(int, m.groups()); _dt(y, mo, d)
                return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag in ("ymd_slash","ymd_dash"):
                y, mo, d = map(int, m.groups()); _dt(y, mo, d)
                return f"{y:04d}-{mo:02d}-{d:02d}"
            if tag in ("my_slash","my_dash","my_dot"):
                mo, y = map(int, m.groups()); ld = _cal.monthrange(y, mo)[1]
                return f"{y:04d}-{mo:02d}-{ld:02d}"
            if tag in ("ym_slash","ym_dash","ym_dot"):
                y, mo = map(int, m.groups()); ld = _cal.monthrange(y, mo)[1]
                return f"{y:04d}-{mo:02d}-{ld:02d}"
        except Exception:
            raise ValueError("Invalid date values.")
    raise ValueError("Unrecognized expiry date format.")
def format_expiry_display(iso: str) -> str:
    """
    Always returns a plain formatted date (DD-Mon-YYYY) without any '(adopted)' suffix.
    """
    if not iso:
        return ""
    try:
        dt = _dt.strptime(iso, "%Y-%m-%d")
        return dt.strftime(_DISPLAY_DATE_FORMAT)
    except Exception:
        return iso
# ---------------------------------------------------------------------
# DATABASE HELPERS
# ---------------------------------------------------------------------
def check_expiry_required(code: str) -> bool:
    conn = connect_db()
    if conn is None:
        return False
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT remarks FROM items_list WHERE code = ?", (code,))
        row = cur.fetchone()
        return bool(row and row['remarks'] and 'exp' in row['remarks'].lower())
    except sqlite3.Error:
        return False
    finally:
        cur.close(); conn.close()
def fetch_project_details():
    conn = connect_db()
    if conn is None:
        return ("Unknown Project","Unknown Code")
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        return (row[0] if row and row[0] else "Unknown Project",
                row[1] if row and row[1] else "Unknown Code")
    except sqlite3.Error:
        return ("Unknown Project","Unknown Code")
    finally:
        cur.close(); conn.close()
def fetch_third_parties():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM third_parties ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()
def fetch_end_users():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM end_users ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()
class StockData:
    @staticmethod
    def add_or_update(unique_id,
                      scenario,
                      qty_in=0,
                      qty_out=0,
                      exp_date=None,
                      kit_number=None,
                      module_number=None,
                      comments=None):
        """
        Insert or accumulate (qty_in, qty_out) for a unique_id, updating expiry,
        kit/module numbers and comments. Supports optional 'scenario' column
        (auto‑detected) and new 'comments' column (must exist in schema).
        """
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed")

        cur = conn.cursor()
        try:
            # Inspect table columns (case-insensitive)
            cur.execute("PRAGMA table_info(stock_data)")
            cols = {row[1].lower(): row[1] for row in cur.fetchall()}

            has_scenario = 'scenario' in cols
            has_comments = 'comments' in cols

            cur.execute("SELECT qty_in, qty_out FROM stock_data WHERE unique_id = ?", (unique_id,))
            existing = cur.fetchone()
            now_ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

            if existing:
                new_in = existing[0] + qty_in
                new_out = existing[1] + qty_out

                if has_scenario and has_comments:
                    cur.execute(f"""
                        UPDATE stock_data
                           SET qty_in=?,
                               qty_out=?,
                               exp_date=?,
                               kit_number=?,
                               module_number=?,
                               scenario=?,
                               comments=?,
                               updated_at=?
                         WHERE unique_id=?
                    """, (new_in, new_out, exp_date, kit_number, module_number,
                          scenario, comments, now_ts, unique_id))
                elif has_scenario:
                    cur.execute(f"""
                        UPDATE stock_data
                           SET qty_in=?,
                               qty_out=?,
                               exp_date=?,
                               kit_number=?,
                               module_number=?,
                               scenario=?,
                               updated_at=?
                         WHERE unique_id=?
                    """, (new_in, new_out, exp_date, kit_number, module_number,
                          scenario, now_ts, unique_id))
                elif has_comments:
                    cur.execute(f"""
                        UPDATE stock_data
                           SET qty_in=?,
                               qty_out=?,
                               exp_date=?,
                               kit_number=?,
                               module_number=?,
                               comments=?,
                               updated_at=?
                         WHERE unique_id=?
                    """, (new_in, new_out, exp_date, kit_number, module_number,
                          comments, now_ts, unique_id))
                else:
                    cur.execute(f"""
                        UPDATE stock_data
                           SET qty_in=?,
                               qty_out=?,
                               exp_date=?,
                               kit_number=?,
                               module_number=?,
                               updated_at=?
                         WHERE unique_id=?
                    """, (new_in, new_out, exp_date, kit_number, module_number,
                          now_ts, unique_id))
            else:
                # Build dynamic insert
                base_cols = ["unique_id", "qty_in", "qty_out", "exp_date",
                             "kit_number", "module_number"]
                base_vals = [unique_id, qty_in, qty_out, exp_date,
                             kit_number, module_number]
                if has_scenario:
                    base_cols.append("scenario")
                    base_vals.append(scenario)
                if has_comments:
                    base_cols.append("comments")
                    base_vals.append(comments)
                base_cols.append("updated_at")
                base_vals.append(now_ts)

                placeholders = ",".join("?" for _ in base_vals)
                col_list = ",".join(base_cols)
                cur.execute(f"""
                    INSERT INTO stock_data ({col_list})
                    VALUES ({placeholders})
                """, base_vals)

            conn.commit()

        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"StockData.add_or_update error: {e}")
            raise
        finally:
            cur.close()
            conn.close()
# ---------------------------------------------------------------------
# MAIN UI CLASS
# ---------------------------------------------------------------------
class StockReceiveKit(tk.Frame):
    def __init__(self, parent, app, role: str = "supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        self.role = role.lower()
        # Scenario mappings
        self.scenario_map = self.fetch_scenario_map()
        self.reverse_scenario_map = {v: k for k, v in self.scenario_map.items()}
        self.selected_scenario_id: str | None = None
        self.selected_scenario_name: str | None = None
        # UI variables / components
        self.tree: ttk.Treeview | None = None
        self.scenario_var = None
        self.mode_var = None
        self.kit_var = None
        self.kit_number_var = None
        self.module_var = None
        self.module_number_var = None
        self.search_var = None
        self.search_listbox = None
        self.trans_type_var = None
        self.end_user_var = None
        self.third_party_var = None
        self.remarks_entry = None
        self.editing_cell: tk.Entry | None = None
        # Data structures
        self.row_data: dict[str, dict] = {}
        self.code_to_iid: dict[str, str] = {}
        self.renamed_modules = []
        self.parent_expiry_map = {}
        self.current_document_number: str | None = None
        # State
        self.expiry_validation_active = False
        # Fonts / styling
        self.font_normal = ("Helvetica", 10)
        self.font_bold = ("Helvetica", 10, "bold")
        self.adopted_font = ("Helvetica", 10, "underline")
        if self.parent and self.parent.winfo_exists():
            self.pack(fill="both", expand=True)
            self.after(120, self.initialize_ui)
    # ----------------- Scenario / init -----------------
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None:
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            return {str(r['scenario_id']): r['name'] for r in cur.fetchall()}
        except sqlite3.Error:
            return {}
        finally:
            cur.close(); conn.close()
    def fetch_scenarios(self):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            return [{"id": r['scenario_id'], "name": r['name']} for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def initialize_ui(self):
        try:
            self.render_ui()
        except tk.TclError as e:
            logging.error(f"UI render error: {e}")
            custom_popup(self.parent, "Error", f"Failed to render UI: {e}", "error")
    def build_mode_definitions(self):
        scenario = self.selected_scenario_name or ""
        self.mode_definitions = [
            ("receive_kit", lang.t("receive_kit.mode_receive_kit", "Receive Kit")),
            ("add_standalone", lang.t("receive_kit.mode_add_standalone", "Add standalone item/s in {scenario}", scenario=scenario)),
            ("add_module_scenario",lang.t("receive_kit.mode_add_module_scenario", "Add module to {scenario}", scenario=scenario)),
            ("add_module_kit", lang.t("receive_kit.mode_add_module_kit", "Add module to a kit")),
            ("add_items_kit", lang.t("receive_kit.mode_add_items_kit", "Add items to a kit")),
            ("add_items_module", lang.t("receive_kit.mode_add_items_module", "Add items to a module"))
        ]
        self.mode_label_to_key = {lbl: key for key, lbl in self.mode_definitions}
    def current_mode_key(self):
        if not hasattr(self, "mode_label_to_key"):
            return None
        return self.mode_label_to_key.get(self.mode_var.get())
    def render_ui(self):
        # Clear existing
        for w in self.parent.winfo_children():
            try: w.destroy()
            except Exception: pass
        tk.Label(self.parent, text=lang.t("receive_kit.title","Receive Kit-Module"),
                 font=("Helvetica", 20, "bold"), bg="#F0F4F8").pack(pady=10)
        main = tk.Frame(self.parent, bg="#F0F4F8")
        main.pack(fill="both", expand=True, padx=10, pady=10)
        # Scenario
        tk.Label(main, text=lang.t("receive_kit.scenario","Scenario:"), bg="#F0F4F8")\
            .grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(main, textvariable=self.scenario_var, state="readonly", width=40)
        self.scenario_cb.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)
        # Movement Type
        tk.Label(main, text=lang.t("receive_kit.movement_type","Movement Type:"), bg="#F0F4F8")\
            .grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.mode_var = tk.StringVar()
        self.mode_cb = ttk.Combobox(main, textvariable=self.mode_var, state="readonly", width=40)
        self.mode_cb.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.mode_cb.bind("<<ComboboxSelected>>", self.update_mode)
        # Kit selectors
        self.kit_label = tk.Label(main, text=lang.t("receive_kit.select_kit","Select Kit:"), bg="#F0F4F8")
        self.kit_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.kit_var = tk.StringVar()
        self.kit_cb = ttk.Combobox(main, textvariable=self.kit_var, state="disabled", width=40)
        self.kit_cb.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.kit_cb.bind("<<ComboboxSelected>>", self.on_kit_selected)
        self.kit_number_label = tk.Label(main, text=lang.t("receive_kit.select_kit_number","Select Kit Number:"), bg="#F0F4F8")
        self.kit_number_label.grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.kit_number_var = tk.StringVar()
        self.kit_number_cb = ttk.Combobox(main, textvariable=self.kit_number_var, state="disabled", width=20)
        self.kit_number_cb.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.kit_number_cb.bind("<<ComboboxSelected>>", self.on_kit_number_selected)
        # Module selectors
        self.module_label = tk.Label(main, text=lang.t("receive_kit.select_module","Select Module:"), bg="#F0F4F8")
        self.module_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.module_var = tk.StringVar()
        self.module_cb = ttk.Combobox(main, textvariable=self.module_var, state="disabled", width=40)
        self.module_cb.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.module_cb.bind("<<ComboboxSelected>>", self.on_module_selected)
        self.module_number_label = tk.Label(main, text=lang.t("receive_kit.select_module_number","Select Module Number:"), bg="#F0F4F8")
        self.module_number_label.grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.module_number_var = tk.StringVar()
        self.module_number_cb = ttk.Combobox(main, textvariable=self.module_number_var, state="disabled", width=20)
        self.module_number_cb.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        self.module_number_cb.bind("<<ComboboxSelected>>", self.on_module_number_selected)
        # IN Type / End User / Third Party / Remarks
        type_frame = tk.Frame(main, bg="#F0F4F8")
        type_frame.grid(row=4, column=0, columnspan=4, pady=5, sticky="w")
        tk.Label(type_frame, text=lang.t("receive_kit.in_type","IN Type:"), bg="#F0F4F8")\
            .grid(row=0, column=0, padx=5, sticky="w")
        self.trans_type_var = tk.StringVar()
        self.trans_type_cb = ttk.Combobox(
            type_frame, textvariable=self.trans_type_var, state="readonly", width=30,
            values=[
                lang.t("receive_kit.in_msf","In MSF"),
                lang.t("receive_kit.in_local_purchase","In Local Purchase"),
                lang.t("receive_kit.in_from_quarantine","In from Quarantine"),
                lang.t("receive_kit.in_donation","In Donation"),
                lang.t("receive_kit.return_from_end_user","Return from End User"),
                lang.t("receive_kit.in_supply_non_msf","In Supply Non-MSF"),
                lang.t("receive_kit.in_borrowing","In Borrowing"),
                lang.t("receive_kit.in_return_loan","In Return of Loan"),
                lang.t("receive_kit.in_correction","In Correction of Previous Transaction")
            ]
        )
        self.trans_type_cb.grid(row=0, column=1, padx=5, pady=5)
        self.trans_type_cb.bind("<<ComboboxSelected>>", self.update_dropdown_visibility)
        tk.Label(type_frame, text=lang.t("receive_kit.end_user","End User:"), bg="#F0F4F8")\
            .grid(row=0, column=2, padx=5, sticky="w")
        self.end_user_var = tk.StringVar()
        self.end_user_cb = ttk.Combobox(type_frame, textvariable=self.end_user_var, state="disabled", width=30)
        self.end_user_cb['values'] = fetch_end_users()
        self.end_user_cb.grid(row=0, column=3, padx=5, pady=5)
        tk.Label(type_frame, text=lang.t("receive_kit.third_party","Third Party:"), bg="#F0F4F8")\
            .grid(row=0, column=4, padx=5, sticky="w")
        self.third_party_var = tk.StringVar()
        self.third_party_cb = ttk.Combobox(type_frame, textvariable=self.third_party_var, state="disabled", width=30)
        self.third_party_cb['values'] = fetch_third_parties()
        self.third_party_cb.grid(row=0, column=5, padx=5, pady=5)
        tk.Label(type_frame, text=lang.t("receive_kit.remarks","Remarks:"), bg="#F0F4F8")\
            .grid(row=0, column=6, padx=5, sticky="w")
        self.remarks_entry = tk.Entry(type_frame, width=40, state="disabled")
        self.remarks_entry.grid(row=0, column=7, padx=5, pady=5)
        # Search
        tk.Label(main, text=lang.t("receive_kit.item","Kit/Module/Item:"), bg="#F0F4F8")\
            .grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(main, textvariable=self.search_var, width=40)
        self.search_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", self.search_items)
        self.search_entry.bind("<Return>", self.select_first_result)
        tk.Button(main, text=lang.t("receive_kit.clear_search","Clear Search"),
                  bg="#7F8C8D", fg="white", command=self.clear_search)\
            .grid(row=5, column=2, padx=5, pady=5)
        self.search_listbox = tk.Listbox(main, height=5, width=60)
        self.search_listbox.grid(row=6, column=1, columnspan=3, padx=5, pady=5, sticky="we")
        self.search_listbox.bind("<<ListboxSelect>>", self.fill_from_search)
        # Tree columns
        cols = (
            "code","description","type","kit","module",
            "std_qty","qty_to_receive","expiry_date","batch_no",
            "exp_module","exp_kit","comments","unique_id"
        )
        self.tree = ttk.Treeview(main, columns=cols, show="headings", height=20)
        # Tag styles
        self.tree.tag_configure("light_red", background="#FF9999")
        self.tree.tag_configure("kit", background="#228B22", foreground="white", font=self.font_bold)
        self.tree.tag_configure("module", background="#ADD8E6", font=self.font_bold)
        self.tree.tag_configure("auto_expiry_gray", foreground="#666666")
        headers = {
                "code": "Code",
                "description": "Description",
                "type": "Type",
                "kit": "Kit",
                "module": "Module",
                "std_qty": "Std Qty",
                "qty_to_receive": "Qty to Receive",
                "expiry_date": "Expiry Date",
                "batch_no": "Batch No",
                "exp_module": "Exp Module",
                "exp_kit": "Exp Kit",
                "comments": "Comments",
                "unique_id": "Unique ID"
        }
        widths = {
                "code": 160,
                "description": 360,
                "type": 95,
                "kit": 110,
                "module": 110,
                "std_qty": 70,
                "qty_to_receive": 120,
                "expiry_date": 140,
                "batch_no": 120,
                "exp_module": 0,
                "exp_kit": 0,
                "comments": 170,
                "unique_id": 0
        }
        aligns = {
                "std_qty": "e",
                "qty_to_receive": "e"
        }
        # Configure columns
        for c in cols:
                self.tree.heading(c, text=headers[c])
                if c in ["exp_module", "exp_kit", "unique_id"]:
                        # Explicitly hide these columns
                        self.tree.column(c, width=0, minwidth=0, stretch=False)
                else:
                        self.tree.column(c, width=widths.get(c, 110), anchor=aligns.get(c, "w"), stretch=True)
        vsb = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(main, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")
        vsb.grid(row=7, column=4, sticky="ns")
        hsb.grid(row=8, column=0, columnspan=4, sticky="ew")
        main.grid_rowconfigure(7, weight=1)
        main.grid_columnconfigure(1, weight=1)
        # Bindings
        self.tree.bind("<Double-1>", self.start_edit)
        self.tree.bind("<KeyPress-Return>", self.start_edit)
        self.tree.bind("<KeyPress-Tab>", self.start_edit)
        self.tree.bind("<KeyPress-Up>", self.navigate_tree)
        self.tree.bind("<KeyPress-Down>", self.navigate_tree)
        self.tree.bind("<KeyPress-Left>", self.navigate_tree)
        self.tree.bind("<KeyPress-Right>", self.navigate_tree)
        self.tree.bind("<Button-3>", self.show_context_menu)
        # Buttons
        btn_frame = tk.Frame(main, bg="#F0F4F8")
        btn_frame.grid(row=9, column=0, columnspan=4, pady=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.add_missing","Add Missing Item"),
                  bg="#FFA500", fg="white", command=self.add_missing_item).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.save","Save"),
                  bg="#27AE60", fg="white", command=self.save_all,
                  state="normal" if self.role in ["admin","manager"] else "disabled").pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.clear","Clear"),
                  bg="#7F8C8D", fg="white", command=self.clear_form).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("receive_kit.export","Export"),
                  bg="#2980B9", fg="white", command=self.export_data).pack(side="left", padx=5)
        self.status_var = tk.StringVar(value=lang.t("receive_kit.ready","Ready"))
        tk.Label(main, textvariable=self.status_var, relief="sunken",
                 anchor="w", bg="#F0F4F8").grid(row=10, column=0, columnspan=4, sticky="ew")
        self.load_scenarios()
    # ----------------- Scenario Handling -----------------
    def load_scenarios(self):
        scens = self.fetch_scenarios()
        values = [f"{s['id']} - {s['name']}" for s in scens]
        self.scenario_cb['values'] = values
        if values:
            self.scenario_cb.current(0)
            self.on_scenario_selected()
    def on_scenario_selected(self, event=None):
        scen_string = self.scenario_var.get()
        if not scen_string:
            self.selected_scenario_id = None
            self.selected_scenario_name = None
            self.tree.delete(*self.tree.get_children())
            self.row_data.clear()
            self.code_to_iid.clear()
            self.status_var.set("")
            return
        parts = scen_string.split(" - ")
        self.selected_scenario_id = parts[0]
        self.selected_scenario_name = parts[1] if len(parts) > 1 else ""
        self.build_mode_definitions()
        self.mode_cb['values'] = [lbl for _, lbl in self.mode_definitions]
        self.mode_var.set(self.mode_definitions[0][1])
        # Reset tree and search
        self.kit_var.set(""); self.kit_number_var.set("")
        self.module_var.set(""); self.module_number_var.set("")
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear(); self.code_to_iid.clear()
        self.search_var.set(""); self.search_listbox.delete(0, tk.END)
        self.update_mode()
    def on_kit_selected(self, event=None):
        kit_code = self.kit_var.get()
        mode_key = self.current_mode_key()
        self.kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        if mode_key in ["add_module_kit","add_items_kit","add_items_module"]:
            if kit_code:
                self.kit_number_cb['values'] = self.fetch_available_kit_numbers(self.selected_scenario_id, kit_code)
                self.kit_number_cb.config(state="readonly" if self.kit_number_cb['values'] else "normal")
            else:
                self.kit_number_cb['values'] = []
                self.kit_number_cb.config(state="normal")
        if mode_key == "add_items_module":
            if kit_code:
                self.module_cb['values'] = self.fetch_modules_for_kit(self.selected_scenario_id, kit_code)
            else:
                self.module_cb['values'] = self.fetch_all_modules(self.selected_scenario_id)
            self.module_number_cb['values'] = self.fetch_module_numbers(self.selected_scenario_id,
                                                                        kit_code=kit_code if kit_code else None)
            self.module_number_cb.config(state="readonly" if self.module_number_cb['values'] else "normal")
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        self.search_listbox.delete(0, tk.END)
        results = self.fetch_search_results("", self.selected_scenario_id, mode_key)
        for r in results:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        self.status_var.set(f"Found {self.search_listbox.size()} items")
    def on_kit_number_selected(self, event=None):
        kit_number = self.kit_number_var.get().strip() if self.kit_number_var.get() else ""
        kit_code = self.kit_var.get()
        mode_key = self.current_mode_key()
        if mode_key == "add_items_module":
            module_code = self.module_var.get()
            if module_code:
                nums = self.fetch_module_numbers_for_module_instance(
                    self.selected_scenario_id, module_code, kit_number or None
                )
                self.module_number_cb['values'] = nums
                self.module_number_cb.config(state="readonly" if nums else "disabled")
            return
        if not kit_number or mode_key not in ["add_module_kit", "add_items_kit"]:
            # Reset tree if no kit number chosen in these modes
            self.tree.delete(*self.tree.get_children())
            self.row_data.clear()
            self.code_to_iid.clear()
            self.recompute_exp_groups()
            return
        items = self.fetch_stock_data_for_kit_number(self.selected_scenario_id, kit_number, kit_code)
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        treecode_to_iid = {}
        for item in sorted(items, key=lambda x: x['treecode']):
            parent_tc = item['treecode'][:-3] if len(item['treecode']) >= 3 else ''
            parent_iid = treecode_to_iid.get(parent_tc, '')
            iid = self.tree.insert(parent_iid, "end", values=(
                item['code'], item['description'], item['type'], item['kit'], item['module'],
                item['std_qty'], item['qty_to_receive'], item['expiry_date'] or "", "",
                "", "", "" # hidden
            ))
            treecode_to_iid[item['treecode']] = iid
            self.code_to_iid[item['code']] = iid
            if item['type'].upper() == 'KIT':
                self.tree.item(iid, tags=("kit",))
            elif item['type'].upper() == 'MODULE':
                self.tree.item(iid, tags=("module",))
            self.row_data[iid] = {
                'unique_id': item['unique_id'],
                'kit_number': item['kit_number'],
                'module_number': item['module_number'],
                'treecode': item['treecode']
            }
        self.recompute_exp_groups()
        self.status_var.set(f"Loaded {len(self.tree.get_children())} records for kit number {kit_number}")
    def on_module_selected(self, event=None):
        module_code = self.module_var.get()
        mode_key = self.current_mode_key()
        kit_code = self.kit_var.get()
        self.module_number_var.set("")
        if mode_key == "add_items_module":
            self.module_number_cb['values'] = self.fetch_module_numbers(
                self.selected_scenario_id,
                kit_code=kit_code if kit_code else None,
                module_code=module_code if module_code else None
            )
            self.module_number_cb.config(state="readonly" if self.module_number_cb['values'] else "normal")
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        self.search_listbox.delete(0, tk.END)
        results = self.fetch_search_results("", self.selected_scenario_id, mode_key)
        for r in results:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        self.status_var.set(f"Found {self.search_listbox.size()} items")
    def on_module_number_selected(self, event=None):
        module_number = self.module_number_var.get()
        mode_key = self.current_mode_key()
        if mode_key != "add_items_module":
            return
        if not module_number:
            self.tree.delete(*self.tree.get_children())
            self.row_data.clear()
            self.code_to_iid.clear()
            self.recompute_exp_groups()
            return
        kit_code = self.kit_var.get()
        module_code = self.module_var.get()
        items = self.fetch_stock_data_for_module_number(self.selected_scenario_id,
                                                        module_number, kit_code, module_code)
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        treecode_to_iid = {}
        for item in sorted(items, key=lambda x: x['treecode']):
            parent_tc = item['treecode'][:-3] if len(item['treecode']) >= 3 else ''
            parent_iid = treecode_to_iid.get(parent_tc, '')
            iid = self.tree.insert(parent_iid, "end", values=(
                item['code'], item['description'], item['type'], item['kit'], item['module'],
                item['std_qty'], item['qty_to_receive'], item['expiry_date'] or "", "",
                "", "", ""
            ))
            treecode_to_iid[item['treecode']] = iid
            self.code_to_iid[item['code']] = iid
            if item['type'].upper() == 'KIT':
                self.tree.item(iid, tags=("kit",))
            elif item['type'].upper() == 'MODULE':
                self.tree.item(iid, tags=("module",))
            self.row_data[iid] = {
                'unique_id': item['unique_id'],
                'kit_number': item['kit_number'],
                'module_number': item['module_number'],
                'treecode': item['treecode']
            }
        self.recompute_exp_groups()
        self.status_var.set(f"Loaded {len(self.tree.get_children())} records for module number {module_number}")
   
    def update_mode(self, event=None):
        mode_key = self.current_mode_key()
        scenario_module_mode = (mode_key == "add_module_scenario")
        # Disable all selectors initially
        for cb in [self.kit_cb, self.kit_number_cb, self.module_cb, self.module_number_cb]:
            cb.config(state="disabled")
        for lab in [self.kit_label, self.kit_number_label, self.module_label, self.module_number_label]:
            lab.config(state="disabled")
        self.kit_var.set(""); self.kit_number_var.set("")
        self.module_var.set(""); self.module_number_var.set("")
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear(); self.code_to_iid.clear()
        self.search_var.set(""); self.search_listbox.delete(0, tk.END)
        if mode_key == "add_module_kit":
            self.kit_label.config(state="normal")
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.kit_number_label.config(state="normal")
            self.kit_number_cb.config(state="normal")
            self.kit_number_cb['values'] = self.fetch_available_kit_numbers(self.selected_scenario_id)
            if self.kit_cb['values']:
                self.kit_cb.current(0)
                self.on_kit_selected()
        elif scenario_module_mode:
            pass
        elif mode_key == "add_items_kit":
            self.kit_label.config(state="normal")
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.kit_number_label.config(state="normal")
            self.kit_number_cb.config(state="normal")
            self.kit_number_cb['values'] = self.fetch_available_kit_numbers(self.selected_scenario_id)
            if self.kit_cb['values']:
                self.kit_cb.current(0)
                self.on_kit_selected()
        elif mode_key == "add_items_module":
            self.kit_label.config(state="normal")
            self.kit_cb.config(state="readonly")
            self.kit_cb['values'] = self.fetch_kits(self.selected_scenario_id)
            self.kit_number_label.config(state="normal")
            self.kit_number_cb.config(state="normal")
            self.kit_number_cb['values'] = self.fetch_available_kit_numbers(self.selected_scenario_id)
            self.module_label.config(state="normal")
            self.module_cb.config(state="readonly")
            self.module_cb['values'] = self.fetch_all_modules(self.selected_scenario_id)
            self.module_number_label.config(state="normal")
            self.module_number_cb.config(state="normal")
            self.module_number_cb['values'] = self.fetch_module_numbers(self.selected_scenario_id)
            self.module_number_cb.config(state="readonly" if self.module_number_cb['values'] else "normal")
        # Populate search results
        results = self.fetch_search_results("", self.selected_scenario_id, mode_key)
        for r in results:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        self.status_var.set(lang.t("receive_kit.found_items", f"Found {self.search_listbox.size()} items"))
    def ensure_mode_ready(self):
        if not hasattr(self, "mode_definitions") or not self.mode_definitions:
            self.build_mode_definitions()
        if not hasattr(self, "mode_label_to_key") or not self.mode_label_to_key:
            self.build_mode_definitions()
        label = self.mode_var.get() if self.mode_var else ""
        key = self.mode_label_to_key.get(label)
        if not key and self.mode_definitions:
            first_label = self.mode_definitions[0][1]
            self.mode_var.set(first_label)
            key = self.mode_definitions[0][0]
        return key
    # -----------------------------------------------------------------
    # Unique ID refresh
    # -----------------------------------------------------------------
    def update_unique_ids_and_column(self):
        scenario_id = self.selected_scenario_id or "None"
        if not self.tree:
            return
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = list(self.tree.item(iid, "values"))
            if not vals or len(vals) < 13:
                continue
            code = vals[0]
            type_field = (vals[2] or "").strip().upper()
            kit_col = vals[3] if vals[3] and vals[3] != "-----" else None
            module_col = vals[4] if vals[4] and vals[4] != "-----" else None
            std_qty = vals[5]
            kit_code = kit_col
            module_code = module_col
            item_code = None
            if type_field == "KIT" and not kit_code:
                kit_code = code
            if type_field == "MODULE" and not module_code:
                module_code = code
            if type_field == "ITEM":
                item_code = code
            rd = self.row_data.setdefault(iid, {})
            expiry_iso = rd.get("expiry_iso")
            if not expiry_iso:
                raw_exp = (vals[7] or "").replace("(adopted)", "").strip()
                expiry_iso = parse_expiry(raw_exp)
                if expiry_iso:
                    rd["expiry_iso"] = expiry_iso
            kit_number = rd.get("kit_number") or "None"
            module_number = rd.get("module_number") or "None"
            try:
                std_int = int(std_qty) if str(std_qty).isdigit() else 0
            except Exception:
                std_int = 0
            unique_id = self.generate_unique_id(
                scenario_id,
                kit_code if kit_code else None,
                module_code if module_code else None,
                item_code,
                std_int,
                expiry_iso,
                kit_number,
                module_number
            )
            rd["unique_id"] = unique_id
            vals[12] = unique_id
            self.tree.item(iid, values=tuple(vals))
    # -----------------------------------------------------------------
    # Fetch Helpers
    # -----------------------------------------------------------------
    def fetch_available_kit_numbers(self, scenario_id, kit_code=None):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            if kit_code:
                cur.execute("""
                    SELECT DISTINCT kit_number FROM stock_data
                    WHERE kit_number IS NOT NULL AND kit_number!='None'
                      AND unique_id LIKE ?
                """, (f"{scenario_id}/{kit_code}/%",))
            else:
                cur.execute("""
                    SELECT DISTINCT kit_number FROM stock_data
                    WHERE kit_number IS NOT NULL AND kit_number!='None'
                      AND unique_id LIKE ?
                """, (f"{scenario_id}/%",))
            return sorted([r[0] for r in cur.fetchall()])
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_available_module_numbers(self, scenario_id, kit_code=None, module_code=None):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            if kit_code and module_code:
                cur.execute("""
                    SELECT DISTINCT module_number FROM stock_data
                    WHERE module_number IS NOT NULL AND module_number!='None'
                      AND unique_id LIKE ?
                """, (f"{scenario_id}/{kit_code}/{module_code}/%",))
            elif kit_code:
                cur.execute("""
                    SELECT DISTINCT module_number FROM stock_data
                    WHERE module_number IS NOT NULL AND module_number!='None'
                      AND unique_id LIKE ?
                """, (f"{scenario_id}/{kit_code}/%",))
            else:
                cur.execute("""
                    SELECT DISTINCT module_number FROM stock_data
                    WHERE module_number IS NOT NULL AND module_number!='None'
                      AND unique_id LIKE ?
                """, (f"{scenario_id}/%",))
            return sorted([r[0] for r in cur.fetchall()])
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_module_numbers_for_module_instance(self, scenario_id, module_code, kit_number=None):
        if not module_code:
            return []
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            sql = """
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE module_number IS NOT NULL
                  AND module_number != 'None'
                  AND unique_id LIKE ?
                  AND SUBSTR(
                        unique_id,
                        INSTR(unique_id,'/',1,2)+1,
                        INSTR(unique_id,'/',1,3)-INSTR(unique_id,'/',1,2)-1
                      ) = ?
            """
            params = [f"{scenario_id}/%", module_code]
            if kit_number:
                sql += " AND kit_number=?"
                params.append(kit_number)
            sql += " ORDER BY module_number"
            cur.execute(sql, params)
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_module_numbers(self, scenario_id: str, kit_code: str = None, module_code: str = None):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            where = ["module_number IS NOT NULL", "module_number!='None'", "unique_id LIKE ?"]
            params = [f"{scenario_id}/%"]
            if kit_code:
                where.append("unique_id LIKE ?")
                params.append(f"{scenario_id}/{kit_code}/%")
            if module_code:
                where.append("unique_id LIKE ?")
                params.append(f"{scenario_id}/%/{module_code}/%")
            sql = f"""
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE {' AND '.join(where)}
                ORDER BY module_number
            """
            cur.execute(sql, params)
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_search_results(self, query, scenario_id, mode_key):
        if not mode_key:
            return []
        if hasattr(self, "mode_label_to_key") and mode_key in self.mode_label_to_key:
            mode_key = self.mode_label_to_key[mode_key]
        mode_key = mode_key.lower().replace(" ", "_")
        conn = connect_db()
        if conn is None: return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            q = (query or "").lower()
            if mode_key == "add_standalone":
                sql = """
                    SELECT DISTINCT ki.code, ki.level
                    FROM kit_items ki
                    LEFT JOIN items_list il ON ki.code=il.code
                    WHERE ki.scenario_id=? AND LOWER(ki.level)='tertiary' AND (
                        UPPER(ki.code) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                    ) ORDER BY ki.code
                """
                params = (scenario_id, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            elif mode_key == "receive_kit":
                sql = """
                    SELECT DISTINCT ki.code, ki.level
                    FROM kit_items ki
                    LEFT JOIN items_list il ON ki.code=il.code
                    WHERE ki.scenario_id=? AND LOWER(ki.level)='primary' AND (
                        UPPER(ki.code) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                    ) ORDER BY ki.code
                """
                params = (scenario_id, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            elif mode_key == "add_module_kit":
                kit_code = self.kit_var.get()
                sql = """
                    SELECT DISTINCT ki.code, ki.level
                    FROM kit_items ki LEFT JOIN items_list il ON ki.code=il.code
                    WHERE ki.scenario_id=? AND ki.kit=? AND LOWER(ki.level)='secondary' AND (
                        UPPER(ki.code) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                    ) ORDER BY ki.code
                """
                params = (scenario_id, kit_code, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            elif mode_key == "add_module_scenario":
                sql = """
                    SELECT DISTINCT ki.code, ki.level
                    FROM kit_items ki LEFT JOIN items_list il ON ki.code=il.code
                    WHERE ki.scenario_id=? AND LOWER(ki.level)='secondary' AND (
                        UPPER(ki.code) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                    ) ORDER BY ki.code
                """
                params = (scenario_id, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            elif mode_key == "add_items_kit":
                kit_code = self.kit_var.get()
                sql = """
                    SELECT DISTINCT ki.code, ki.level
                    FROM kit_items ki LEFT JOIN items_list il ON ki.code=il.code
                    WHERE ki.scenario_id=? AND ki.kit=? AND LOWER(ki.level)='tertiary' AND (
                        UPPER(ki.code) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                        UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                    ) ORDER BY ki.code
                """
                params = (scenario_id, kit_code, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
            elif mode_key == "add_items_module":
                kit_code = self.kit_var.get()
                module_code = self.module_var.get()
                if kit_code and module_code:
                    sql = """
                        SELECT DISTINCT ki.code, ki.level
                        FROM kit_items ki LEFT JOIN items_list il ON ki.code=il.code
                        WHERE ki.scenario_id=? AND ki.kit=? AND ki.module=? AND LOWER(ki.level)='tertiary' AND (
                            UPPER(ki.code) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                        ) ORDER BY ki.code
                    """
                    params = (scenario_id, kit_code, module_code, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
                elif module_code:
                    sql = """
                        SELECT DISTINCT ki.code, ki.level
                        FROM kit_items ki LEFT JOIN items_list il ON ki.code=il.code
                        WHERE ki.scenario_id=? AND ki.module=? AND LOWER(ki.level)='tertiary' AND (
                            UPPER(ki.code) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_en,'')) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_fr,'')) LIKE UPPER(?) OR
                            UPPER(COALESCE(il.designation_sp,'')) LIKE UPPER(?)
                        ) ORDER BY ki.code
                    """
                    params = (scenario_id, module_code, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")
                else:
                    sql = "SELECT code, level FROM kit_items WHERE 1=0"
                    params = ()
            else:
                sql = "SELECT code, level FROM kit_items WHERE 1=0"
                params = ()
            cur.execute(sql, params)
            rtn = []
            for row in cur.fetchall():
                code = row['code']
                desc = get_item_description(code)
                rtn.append({
                    "code": code,
                    "description": desc,
                    "level": row['level'],
                    "type": detect_type(code, desc)
                })
            return rtn
        except sqlite3.Error as e:
            logging.error(f"fetch_search_results error: {e}")
            return []
        finally:
            cur.close(); conn.close()
    def fetch_kit_items(self, scenario_id, code):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT ki.code, ki.level, ki.kit, ki.module, ki.item,
                       COALESCE(ki.std_qty,0) AS std_qty, ki.treecode
                FROM kit_items ki
                WHERE ki.scenario_id=? AND (
                      ki.code=? OR ki.kit=? OR ki.module=? OR ki.item=?)
                ORDER BY ki.treecode
            """, (scenario_id, code, code, code, code))
            out = []
            for r in cur.fetchall():
                desc = get_item_description(r['code'])
                out.append({
                    'code': r['code'],
                    'description': desc,
                    'type': detect_type(r['code'], desc),
                    'kit': r['kit'] or "-----",
                    'module': r['module'] or "-----",
                    'item': r['item'] or "-----",
                    'std_qty': r['std_qty'],
                    'qty_to_receive': r['std_qty'],
                    'expiry_date': "",
                    'batch_no': "",
                    'treecode': r['treecode']
                })
            return out
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_kits(self, scenario_id):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT code FROM kit_items
                WHERE scenario_id=? AND level='primary'
                ORDER BY code
            """, (scenario_id,))
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_modules_for_kit(self, scenario_id, kit_code):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT code FROM kit_items
                WHERE scenario_id=? AND kit=? AND level='secondary'
                ORDER BY code
            """, (scenario_id, kit_code))
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_all_modules(self, scenario_id):
        conn = connect_db()
        if conn is None: return []
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT code FROM kit_items
                WHERE scenario_id=? AND level='secondary'
                ORDER BY code
            """, (scenario_id,))
            return [r[0] for r in cur.fetchall()]
        except sqlite3.Error:
            return []
        finally:
            cur.close(); conn.close()
    def fetch_stock_data_for_kit_number(self, scenario_id, kit_number, kit_code=None):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            query = """
                SELECT sd.unique_id, sd.qty_in, sd.exp_date, sd.kit_number, sd.module_number,
                       ki.code, ki.kit, ki.module, ki.item, COALESCE(ki.std_qty,0) AS std_qty, ki.treecode
                FROM stock_data sd
                JOIN kit_items ki ON sd.unique_id LIKE ? AND ki.scenario_id=? AND (
                    ki.code = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                     INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.kit = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                       INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.module = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                          INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.item = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,3)+1,
                                        INSTR(sd.unique_id,'/',1,4)-INSTR(sd.unique_id,'/',1,3)-1)
                )
                WHERE sd.kit_number=? AND sd.qty_in>0
            """
            params = (f"{scenario_id}/%", scenario_id, kit_number)
            if kit_code:
                query += " AND ki.kit=?"
                params += (kit_code,)
            cur.execute(query, params)
            lst = []
            for row in cur.fetchall():
                desc = get_item_description(row['code'])
                lst.append({
                    'code': row['code'],
                    'description': desc,
                    'type': detect_type(row['code'], desc),
                    'kit': row['kit'] or "-----",
                    'module': row['module'] or "-----",
                    'item': row['item'] or "-----",
                    'std_qty': row['std_qty'],
                    'qty_to_receive': row['qty_in'],
                    'expiry_date': row['exp_date'] or "",
                    'batch_no': "",
                    'treecode': row['treecode'],
                    'unique_id': row['unique_id'],
                    'kit_number': row['kit_number'],
                    'module_number': row['module_number']
                })
            return lst
        except sqlite3.Error as e:
            logging.error(f"fetch_stock_data_for_kit_number: {e}")
            return []
        finally:
            cur.close(); conn.close()
    def fetch_stock_data_for_module_number(self, scenario_id, module_number, kit_code=None, module_code=None):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory=sqlite3.Row
        cur = conn.cursor()
        try:
            query = """
                SELECT sd.unique_id, sd.qty_in, sd.exp_date, sd.kit_number, sd.module_number,
                       ki.code, ki.kit, ki.module, ki.item, COALESCE(ki.std_qty,0) AS std_qty, ki.treecode
                FROM stock_data sd
                JOIN kit_items ki ON sd.unique_id LIKE ? AND ki.scenario_id=? AND (
                    ki.code = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                     INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.kit = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                       INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.module = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,2)+1,
                                          INSTR(sd.unique_id,'/',1,3)-INSTR(sd.unique_id,'/',1,2)-1)
                    OR ki.item = SUBSTR(sd.unique_id, INSTR(sd.unique_id,'/',1,3)+1,
                                        INSTR(sd.unique_id,'/',1,4)-INSTR(sd.unique_id,'/',1,3)-1)
                )
                WHERE sd.module_number=? AND sd.qty_in>0
            """
            params = (f"{scenario_id}/%", scenario_id, module_number)
            if kit_code and module_code:
                query += " AND ki.kit=? AND ki.module=?"
                params += (kit_code, module_code)
            elif kit_code:
                query += " AND ki.kit=?"
                params += (kit_code,)
            cur.execute(query, params)
            lst = []
            for row in cur.fetchall():
                desc = get_item_description(row['code'])
                lst.append({
                    'code': row['code'],
                    'description': desc,
                    'type': detect_type(row['code'], desc),
                    'kit': row['kit'] or "-----",
                    'module': row['module'] or "-----",
                    'item': row['item'] or "-----",
                    'std_qty': row['std_qty'],
                    'qty_to_receive': row['qty_in'],
                    'expiry_date': row['exp_date'] or "",
                    'batch_no': "",
                    'treecode': row['treecode'],
                    'unique_id': row['unique_id'],
                    'kit_number': row['kit_number'],
                    'module_number': row['module_number']
                })
            return lst
        except sqlite3.Error as e:
            logging.error(f"fetch_stock_data_for_module_number: {e}")
            return []
        finally:
            cur.close(); conn.close()
    # -----------------------------------------------------------------
    # Module descendant retrieval strategies
    # -----------------------------------------------------------------
    def fetch_full_module_subtree(self, scenario_id: str, module_code: str):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory=sqlite3.Row
        cur=conn.cursor()
        try:
            cur.execute("""
                SELECT code, level, kit, module, item, COALESCE(std_qty,0) AS std_qty, treecode
                FROM kit_items
                WHERE scenario_id=? AND code=? AND level='secondary'
                LIMIT 1
            """, (scenario_id, module_code))
            base = cur.fetchone()
            if not base:
                return []
            prefix = base['treecode']
            if not prefix: return []
            cur.execute("""
                SELECT code, level, kit, module, item, COALESCE(std_qty,0) AS std_qty, treecode
                FROM kit_items
                WHERE scenario_id=? AND treecode LIKE ?
                ORDER BY treecode
            """, (scenario_id, f"{prefix}%"))
            lst=[]
            for r in cur.fetchall():
                desc = get_item_description(r['code'])
                lst.append({
                    'code': r['code'],
                    'description': desc,
                    'type': detect_type(r['code'], desc),
                    'kit': r['kit'] or "-----",
                    'module': r['module'] or "-----",
                    'item': r['item'] or "-----",
                    'std_qty': r['std_qty'],
                    'qty_to_receive': r['std_qty'],
                    'expiry_date': "",
                    'batch_no': "",
                    'treecode': r['treecode']
                })
            return lst
        except sqlite3.Error as e:
            logging.error(f"fetch_full_module_subtree error: {e}")
            return []
        finally:
            cur.close(); conn.close()
    def fetch_items_by_module_column(self, scenario_id: str, module_code: str):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory=sqlite3.Row
        cur=conn.cursor()
        try:
            cur.execute("""
                SELECT code, level, kit, module, item, COALESCE(std_qty,0) AS std_qty, treecode
                FROM kit_items
                WHERE scenario_id=? AND module=?
                ORDER BY treecode
            """, (scenario_id, module_code))
            lst=[]
            for r in cur.fetchall():
                desc = get_item_description(r['code'])
                lst.append({
                    'code': r['code'],
                    'description': desc,
                    'type': detect_type(r['code'], desc),
                    'kit': r['kit'] or "-----",
                    'module': r['module'] or "-----",
                    'item': r['item'] or "-----",
                    'std_qty': r['std_qty'],
                    'qty_to_receive': r['std_qty'],
                    'expiry_date': "",
                    'batch_no': "",
                    'treecode': r['treecode']
                })
            return lst
        except sqlite3.Error as e:
            logging.error(f"fetch_items_by_module_column error: {e}")
            return []
        finally:
            cur.close(); conn.close()
    def fetch_items_by_code_prefix(self, scenario_id: str, module_code: str):
        conn = connect_db()
        if conn is None: return []
        conn.row_factory=sqlite3.Row
        cur=conn.cursor()
        try:
            like_pattern = module_code + '%'
            cur.execute("""
                SELECT code, level, kit, module, item, COALESCE(std_qty,0) AS std_qty, treecode
                FROM kit_items
                WHERE scenario_id=? AND code LIKE ?
                ORDER BY treecode
            """, (scenario_id, like_pattern))
            lst=[]
            for r in cur.fetchall():
                desc = get_item_description(r['code'])
                lst.append({
                    'code': r['code'],
                    'description': desc,
                    'type': detect_type(r['code'], desc),
                    'kit': r['kit'] or "-----",
                    'module': r['module'] or "-----",
                    'item': r['item'] or "-----",
                    'std_qty': r['std_qty'],
                    'qty_to_receive': r['std_qty'],
                    'expiry_date': "",
                    'batch_no': "",
                    'treecode': r['treecode']
                })
            return lst
        except sqlite3.Error as e:
            logging.error(f"fetch_items_by_code_prefix error: {e}")
            return []
        finally:
            cur.close(); conn.close()
    # -----------------------------------------------------------------
    # Add Missing Item Dialog
    # -----------------------------------------------------------------
    def add_missing_item(self):
        if not self.selected_scenario_id:
            custom_popup(self.parent, "Error", "Please select a scenario", "error")
            return
        dialog = tk.Toplevel(self.parent)
        dialog.title(lang.t("receive_kit.add_missing","Add Missing Item"))
        dialog.geometry("520x300")
        dialog.transient(self.parent)
        dialog.grab_set()
        tk.Label(dialog, text=lang.t("receive_kit.search_item","Search Kit/Module/Item"),
                 font=("Helvetica",10), pady=10).pack()
        search_var = tk.StringVar()
        entry = tk.Entry(dialog, textvariable=search_var, font=("Helvetica",10), width=40)
        entry.pack(fill=tk.X, padx=10, pady=5)
        listbox = tk.Listbox(dialog, font=("Helvetica",10),
                             selectbackground="#0077CC", selectforeground="white", height=10)
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        def update_list(*_):
            q = search_var.get().strip()
            listbox.delete(0, tk.END)
            mode_key = self.current_mode_key()
            res = self.fetch_search_results(q, self.selected_scenario_id, mode_key)
            for r in res:
                listbox.insert(tk.END, f"{r['code']} - {r['description']}")
            if not res:
                listbox.insert(tk.END, lang.t("receive_kit.no_items_found","No items found"))
        search_var.trace("w", update_list)
        update_list()
        def on_select(_=None):
            idx = listbox.curselection()
            if idx:
                sel = listbox.get(idx[0])
                code = sel.split(" - ")[0]
                dialog.destroy()
                if self.current_mode_key() == "receive_kit":
                    self.load_hierarchy(code)
                else:
                    self.add_to_tree(code)
            else:
                dialog.destroy()
        listbox.bind("<Double-Button-1>", on_select)
        entry.bind("<Return>", on_select)
        entry.focus()
        dialog.wait_window()
    # -----------------------------------------------------------------
    # Select Kit Popup (Receive Kit mode)
    # -----------------------------------------------------------------
    def select_kit_popup(self):
        if not self.selected_scenario_id:
            custom_popup(self.parent, "Error", "Please select a scenario", "error")
            return None, None
        dialog = tk.Toplevel(self.parent)
        dialog.withdraw()
        dialog.title(lang.t("receive_kit.select_kit","Select Kit"))
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog.configure(bg="#FFFFFF")
        PAD_X, PAD_Y = 18, 14
        outer = tk.Frame(dialog, bg="#FFFFFF", padx=PAD_X, pady=PAD_Y)
        outer.pack(fill="both", expand=True)
        tk.Label(outer, text=lang.t("receive_kit.select_kit","Select Kit"),
                 font=("Helvetica",12,"bold"), bg="#FFFFFF", anchor="w").pack(fill="x", pady=(0,6))
        search_var = tk.StringVar()
        search_entry = tk.Entry(outer, textvariable=search_var, font=("Helvetica",10), width=52)
        search_entry.pack(fill="x", pady=(0,6))
        list_frame = tk.Frame(outer, bg="#FFFFFF")
        list_frame.pack(fill="both", expand=True, pady=(0,10))
        listbox = tk.Listbox(list_frame, font=("Helvetica",10), height=6,
                             activestyle="dotbox", selectbackground="#2563EB",
                             selectforeground="white", exportselection=False)
        listbox.pack(side="left", fill="both", expand=True)
        sb = tk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        sb.pack(side="right", fill="y")
        listbox.configure(yscrollcommand=sb.set)
        tk.Label(outer, text=lang.t("receive_kit.kit_number","Kit Number"),
                 font=("Helvetica",10), bg="#FFFFFF", anchor="w").pack(fill="x", pady=(0,2))
        kit_number_var = tk.StringVar()
        kit_number_entry = tk.Entry(outer, textvariable=kit_number_var, font=("Helvetica",10), width=26)
        kit_number_entry.pack(anchor="w", pady=(0,6))
        error_label = tk.Label(outer, text="", fg="#B91C1C", font=("Helvetica",9),
                               bg="#FFFFFF", wraplength=400, justify="left", anchor="w")
        error_label.pack(fill="x", pady=(0,10))
        btn_frame = tk.Frame(outer, bg="#FFFFFF")
        btn_frame.pack(fill="x", pady=(4,0))
        selected_kit = [None]; selected_kit_number = [None]
        def update_list(*_):
            q = search_var.get().strip().lower()
            listbox.delete(0, tk.END)
            kits = self.fetch_search_results(q, self.selected_scenario_id, "receive_kit")
            for k in kits:
                listbox.insert(tk.END, f"{k['code']} - {k['description']}")
            if not kits:
                listbox.insert(tk.END, "No kits found")
        def validate_kit_number():
            kn = kit_number_var.get().strip()
            if not kn:
                error_label.config(text="Kit Number is required")
                return False
            conn = connect_db()
            if conn is None:
                error_label.config(text="DB connection failed")
                return False
            cur = conn.cursor()
            try:
                cur.execute("SELECT COUNT(*) FROM stock_data WHERE kit_number=? AND unique_id LIKE ?",
                            (kn, f"{self.selected_scenario_id}/%"))
                if cur.fetchone()[0] > 0:
                    error_label.config(text="Kit Number already exists")
                    return False
                error_label.config(text=""); return True
            finally:
                cur.close(); conn.close()
        def close_dialog():
            try: dialog.grab_release()
            except: pass
            dialog.destroy()
        def on_confirm(_=None):
            idx = listbox.curselection()
            if not idx:
                error_label.config(text="Please select a kit")
                return
            if not validate_kit_number():
                return
            line = listbox.get(idx[0])
            selected_kit[0] = line.split(" - ")[0]
            selected_kit_number[0] = kit_number_var.get().strip()
            close_dialog()
        tk.Button(btn_frame, text=lang.t("receive_kit.confirm","Confirm"),
                  bg="#2563EB", fg="white", font=("Helvetica",10,"bold"), relief="flat",
                  padx=16, pady=6, command=on_confirm).pack(side="left", padx=(0,10))
        tk.Button(btn_frame, text=lang.t("receive_kit.cancel","Cancel"),
                  bg="#E5E7EB", fg="#111827", relief="flat", padx=16, pady=6,
                  command=close_dialog).pack(side="left")
        listbox.bind("<Double-Button-1>", on_confirm)
        search_entry.bind("<Return>", on_confirm)
        kit_number_entry.bind("<Return>", on_confirm)
        dialog.bind("<Escape>", lambda e: close_dialog())
        search_var.trace_add("write", update_list)
        update_list()
        dialog.update_idletasks()
        _center_child(dialog, self.parent)
        dialog.deiconify()
        search_entry.focus()
        dialog.wait_window()
        return selected_kit[0], selected_kit_number[0]
    # -----------------------------------------------------------------
    # Select Module Popup
    # -----------------------------------------------------------------
    def select_module_popup(self, kit_number, module_code, module_description):
        dialog = tk.Toplevel(self.parent)
        dialog.withdraw()
        dialog.title(lang.t("receive_kit.select_module","Select Module Number"))
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog.configure(bg="#FFFFFF")
        PAD_X, PAD_Y = 18, 14
        outer = tk.Frame(dialog, bg="#FFFFFF", padx=PAD_X, pady=PAD_Y)
        outer.pack(fill="both", expand=True)
        tk.Label(outer, text=f"{lang.t('receive_kit.module','Module')}: {module_code}",
                 font=("Helvetica",11,"bold"), bg="#FFFFFF", anchor="w").pack(fill="x", pady=(0,4))
        if module_description:
            tk.Label(outer, text=module_description, font=("Helvetica",9),
                     fg="#374151", bg="#FFFFFF", wraplength=380,
                     justify="left", anchor="w").pack(fill="x", pady=(0,8))
        suggested = f"{kit_number}-M" if kit_number else "MOD-"
        module_number_var = tk.StringVar(value=suggested)
        tk.Label(outer, text=lang.t("receive_kit.module_number","Module Number"),
                 font=("Helvetica",10), bg="#FFFFFF", anchor="w").pack(fill="x", pady=(0,2))
        entry = tk.Entry(outer, textvariable=module_number_var, font=("Helvetica",10), width=26)
        entry.pack(anchor="w", pady=(0,6))
        error_label = tk.Label(outer, text="", fg="#B91C1C", font=("Helvetica",9),
                               bg="#FFFFFF", wraplength=380, justify="left", anchor="w")
        error_label.pack(fill="x", pady=(0,8))
        btn_bar = tk.Frame(outer, bg="#FFFFFF")
        btn_bar.pack(fill="x", pady=(4,0))
        selected_module_number = [None]
        def validate():
            val = module_number_var.get().strip()
            if not val:
                error_label.config(text="Module Number is required"); return False
            conn = connect_db()
            if conn is None:
                error_label.config(text="DB error"); return False
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT COUNT(*) FROM stock_data
                    WHERE module_number=? AND module_number!='None'
                """, (val,))
                if cur.fetchone()[0] > 0:
                    error_label.config(text="Module Number already exists")
                    return False
                error_label.config(text=""); return True
            finally:
                cur.close(); conn.close()
        def on_confirm(_=None):
            if validate():
                selected_module_number[0] = module_number_var.get().strip()
                close()
        def close():
            try: dialog.grab_release()
            except: pass
            dialog.destroy()
        tk.Button(btn_bar, text=lang.t("receive_kit.confirm","Confirm"),
                  command=on_confirm, bg="#2563EB", fg="white",
                  relief="flat", padx=14, pady=5,
                  font=("Helvetica",10,"bold")).pack(side="left", padx=(0,8))
        tk.Button(btn_bar, text=lang.t("receive_kit.cancel","Cancel"),
                  command=close, bg="#E5E7EB", fg="#111827",
                  relief="flat", padx=14, pady=5,
                  font=("Helvetica",10)).pack(side="left")
        entry.bind("<Return>", on_confirm)
        dialog.bind("<Escape>", lambda e: close())
        dialog.update_idletasks()
        _center_child(dialog, self.parent)
        dialog.deiconify()
        entry.focus()
        dialog.wait_window()
        return selected_module_number[0]
    # -----------------------------------------------------------------
    # Load hierarchy (Receive Kit)
    # -----------------------------------------------------------------
    def load_hierarchy(self, code):
        # Clear existing tree and metadata
        self.tree.delete(*self.tree.get_children())
        self.row_data.clear()
        self.code_to_iid.clear()
        comps = self.fetch_kit_items(self.selected_scenario_id, code)
        if not comps:
            self.status_var.set(f"No items found for code {code}")
            return
        kit_number_global = self.kit_number_var.get() or None
        treecode_map = {}
        module_number_map = {}
        for comp in sorted(comps, key=lambda x: x['treecode']):
            parent_tc = comp['treecode'][:-3] if len(comp['treecode']) >= 3 else ''
            parent_iid = treecode_map.get(parent_tc, '')
            qty_to_receive = comp['std_qty']
            kit_display = comp['kit'] or "-----"
            module_display = comp['module'] or "-----"
            # Inherit kit number from parent if present
            inherited_kn = kit_number_global
            if parent_iid in self.row_data:
                pn = self.row_data[parent_iid].get('kit_number')
                if pn and isinstance(pn, str) and pn.lower() == "none":
                    pn = None
                inherited_kn = pn or inherited_kn
            kit_number = inherited_kn
            module_number = None
            if comp['type'].upper() == "MODULE":
                module_number = self.select_module_popup(kit_number, comp['code'], comp['description'])
                if module_number is None:
                    self.status_var.set(f"Module number selection cancelled for {comp['code']}")
                    return
                module_number_map[comp['code']] = module_number
            else:
                # Inherit module number from parent or mapping
                if parent_iid in self.row_data:
                    pm = self.row_data[parent_iid].get('module_number')
                    if pm and isinstance(pm, str) and pm.lower() == "none":
                        pm = None
                    module_number = pm
                if not module_number and comp['module']:
                    module_number = module_number_map.get(comp['module'])
            iid = self.tree.insert(parent_iid, "end", values=(
                comp['code'], comp['description'], comp['type'], kit_display,
                module_display, comp['std_qty'], qty_to_receive, "", "",
                "", "", "", "" # hidden columns remain blank
            ))
            treecode_map[comp['treecode']] = iid
            self.code_to_iid[comp['code']] = iid
            # Tag structural rows
            if comp['type'].upper() == "KIT":
                self.tree.item(iid, tags=("kit",))
            elif comp['type'].upper() == "MODULE":
                self.tree.item(iid, tags=("module",))
            self.row_data[iid] = {
                'kit_number': kit_number,
                'module_number': module_number,
                'treecode': comp['treecode']
            }
        self.recompute_exp_groups()
        self.status_var.set(f"Loaded {len(self.tree.get_children())} top-level children for {code}")
    # -----------------------------------------------------------------
    # Expiry validation toggles
    # -----------------------------------------------------------------
    def activate_global_expiry_validation(self):
        self.expiry_validation_active = True
    def update_parent_expiry(self):
        if not self.expiry_validation_active: return
        for top in self.tree.get_children():
            self.update_subtree_expiry(top)
    def update_subtree_expiry(self, iid):
        if not self.tree.exists(iid): return
        vals = self.tree.item(iid, "values")
        if not vals: return
        self._validate_and_tag_row(iid)
        for c in self.tree.get_children(iid):
            self.update_subtree_expiry(c)
    def _validate_and_tag_row(self, iid, force=False):
        if not self.tree.exists(iid): return
        vals = self.tree.item(iid, "values")
        if not vals: return
        rtype = (vals[2] or "").strip().upper()
        is_struct = rtype in ("KIT","MODULE","PRIMARY","SECONDARY")
        rd = self.row_data.get(iid, {})
        tags = []
        if rtype in ("KIT","PRIMARY"):
            tags.append("kit")
        elif rtype in ("MODULE","SECONDARY"):
            tags.append("module")
        expiry_cell = (vals[7] or "")
        has_iso = bool(rd.get("expiry_iso"))
        if not has_iso:
            base = parse_expiry(expiry_cell.replace("(adopted)","").strip())
            if base: has_iso = True
        if (self.expiry_validation_active or force) and is_struct and not has_iso:
            tags.insert(0,"light_red")
        if rd.get("auto_expiry") and not rd.get("user_manual_expiry"):
            tags.append("auto_expiry_gray")
        self.tree.item(iid, tags=tuple(dict.fromkeys(tags)))
    # -----------------------------------------------------------------
    # Inline editing (Qty / Expiry / Batch)
    # -----------------------------------------------------------------
    def start_edit(self, event):
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_index = int(self.tree.identify_column(event.x).replace("#","")) - 1
        if col_index not in [6,7,8]:
            return
        if self.editing_cell:
            try: self.editing_cell.destroy()
            except Exception: pass
        self.start_edit_cell(row_id, col_index)
    def start_edit_cell(self, row_id, col_index):
        if not self.tree.exists(row_id): return
        if col_index not in (6,7,8): return
        bbox = self.tree.bbox(row_id, f"#{col_index+1}")
        if not bbox: return
        x,y,w,h = bbox
        original = self.tree.set(row_id, self.tree["columns"][col_index])
        if self.editing_cell:
            try: self.editing_cell.destroy()
            except: pass
        entry = tk.Entry(self.tree, font=self.font_normal)
        entry.place(x=x,y=y,width=w,height=h)
        if original:
            entry.insert(0, original)
        entry.focus()
        self.editing_cell = entry
        def finish(edited=False, edited_col=None):
            # Only expiry edits trigger comment + group recompute
            if self.editing_cell is entry:
                try: entry.destroy()
                except: pass
                self.editing_cell = None
            if not edited:
                return
            if edited_col == 7: # expiry
                self._validate_and_tag_row(row_id, force=True)
                self.recompute_exp_groups()
                self.update_row_comment(row_id, force=True, sticky_mode=True)
            elif edited_col == 6: # quantity
                # No expiry/ comment force; just keep tags (no force)
                self._validate_and_tag_row(row_id, force=False)
            elif edited_col == 8: # batch
                pass # nothing extra
        def save_and_close(event=None):
            if self.editing_cell is not entry:
                return
            new = entry.get().strip()
            edited = False
            edited_col = col_index
            try:
                vals = self.tree.item(row_id, "values")
                if not vals: return
                elif col_index == 6:
                    # Quantity (base user quantity)
                    rtype = self._get_row_type(row_id)
                    if rtype.upper() in ("KIT", "MODULE"):
                        mb.showinfo("Info", "Quantity for Kit or Module must always remain 1.")
                        return  # Prevent change
                    if new == "":
                        # Allow blank -> treat as 1
                        new_qty = 1
                    else:
                        if not new.isdigit():
                            custom_popup(self.parent, "Invalid", "Quantity must be a whole number", "error")
                            return
                        new_qty = int(new)
                        if new_qty < 0:
                            custom_popup(self.parent, "Invalid", "Quantity cannot be negative", "error")
                            return
                    self.row_data[row_id]['user_qty'] = new_qty
                    vals = list(self.tree.item(row_id, "values"))
                    vals[6] = str(new_qty)
                    self.tree.item(row_id, values=tuple(vals))
                    edited = True
                elif col_index == 7:
                    # Expiry
                    rd = self.row_data.setdefault(row_id, {})
                    if new == "":
                        self.tree.set(row_id, "expiry_date", "")
                        rd.pop("expiry_iso", None)
                        rd["user_manual_expiry"] = False
                        rd["auto_expiry"] = False
                        edited = True
                    else:
                        iso = parse_expiry(new)
                        if not iso:
                            try:
                                iso = strict_parse_expiry(new)
                            except Exception:
                                iso = None
                        if not iso:
                            custom_popup(self.parent, "Invalid Expiry",
                                         "Unrecognized date. Examples:\n2029-10-05\n05/10/2029\n10/2029\n2029-10",
                                         "error")
                            return
                        rd["expiry_iso"] = iso
                        rd["user_manual_expiry"] = True
                        rd["auto_expiry"] = False
                        self.tree.set(row_id, "expiry_date", format_expiry_display(iso))
                        tset = set(self.tree.item(row_id, "tags") or [])
                        if "auto_expiry_gray" in tset:
                            tset.remove("auto_expiry_gray")
                        self.tree.item(row_id, tags=tuple(tset))
                        edited = True
                elif col_index == 8:
                    # Batch No
                    self.tree.set(row_id, "batch_no", new[:30])
                    edited = True
            finally:
                finish(edited, edited_col)
                if edited and col_index == 6:
                    # Preserve comment explicitly (re-apply existing sticky if any) – no recalculation
                    rd = self.row_data.get(row_id, {})
                    sticky = rd.get('sticky_comment')
                    if sticky:
                        current_vals = list(self.tree.item(row_id, "values"))
                        if len(current_vals) >= 12 and current_vals[11] != sticky:
                            current_vals[11] = sticky
                            self.tree.item(row_id, values=tuple(current_vals))
                if event and event.keysym in ("Tab","Return","Up","Down","Left","Right"):
                    self.navigate_tree(event)
        def cancel(_=None):
            finish(False, None)
        entry.bind("<Return>", save_and_close)
        entry.bind("<Tab>", save_and_close)
        entry.bind("<Escape>", cancel)
        entry.bind("<FocusOut>", save_and_close)
        entry.bind("<Up>", save_and_close)
        entry.bind("<Down>", save_and_close)
        entry.bind("<Left>", save_and_close)
        entry.bind("<Right>", save_and_close)
    # -----------------------------------------------------------------
    # Quantity Propagation & Navigation
    # -----------------------------------------------------------------
    def navigate_tree(self, event):
        if self.editing_cell:
            return
        sel = self.tree.selection()
        if not sel:
            rows = self.tree.get_children()
            if rows:
                self.tree.selection_set(rows[0])
                self.tree.focus(rows[0])
                self.start_edit_cell(rows[0], 6)
            return
        current = sel[0]
        rows = list(self.tree.get_children())
        for p in rows:
            rows.extend(self.tree.get_children(p))
        if current not in rows:
            return
        idx = rows.index(current)
        if event.keysym == "Up" and idx > 0:
            nxt = rows[idx-1]
            self.tree.selection_set(nxt); self.tree.focus(nxt)
            self.start_edit_cell(nxt, 6)
        elif event.keysym == "Down" and idx < len(rows)-1:
            nxt = rows[idx+1]
            self.tree.selection_set(nxt); self.tree.focus(nxt)
            self.start_edit_cell(nxt, 6)
        elif event.keysym in ("Tab","Return"):
            nxt = rows[(idx+1) % len(rows)]
            self.tree.selection_set(nxt); self.tree.focus(nxt)
            self.start_edit_cell(nxt, 6)
    # -----------------------------------------------------------------
    # Search / Selection
    # -----------------------------------------------------------------
    def search_items(self, event=None):
        q = self.search_var.get().strip()
        self.search_listbox.delete(0, tk.END)
        if not q:
            self.tree.delete(*self.tree.get_children())
            self.row_data.clear()
            self.code_to_iid.clear()
            self.status_var.set("")
            return
        if not self.selected_scenario_id:
            self.status_var.set("Please select a scenario")
            return
        mode_key = self.current_mode_key()
        res = self.fetch_search_results(q, self.selected_scenario_id, mode_key)
        for r in res:
            self.search_listbox.insert(tk.END, f"{r['code']} - {r['description']}")
        self.status_var.set(f"Found {self.search_listbox.size()} items")
    def select_first_result(self, event=None):
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self.fill_from_search()
    def fill_from_search(self, event=None):
        sel = self.search_listbox.curselection()
        if not sel:
            return
        chosen = self.search_listbox.get(sel[0])
        code = chosen.split(" - ")[0]
        self.search_var.set(chosen)
        self.search_listbox.delete(0, tk.END)
        if self.current_mode_key() == "receive_kit":
            kit_code, kit_number = self.select_kit_popup()
            if kit_code and kit_number:
                self.kit_number_var.set(kit_number)
                self.load_hierarchy(kit_code)
        else:
            self.add_to_tree(code)
    # -----------------------------------------------------------------
    # Add single code (mode-dependent)
    # -----------------------------------------------------------------
    def add_to_tree(self, code):
        mode_key = self.current_mode_key()
        scen_module_mode = (mode_key == "add_module_scenario")
        kit_code = self.kit_var.get() if mode_key in [
            "add_module_kit", "add_items_kit", "add_items_module"
        ] else None
        module_code_selected = self.module_var.get() if mode_key == "add_items_module" else None
        kit_number = self.kit_number_var.get().strip() if self.kit_number_var.get() else None
        module_number = (self.module_number_var.get().strip()
                         if (mode_key == "add_items_module" and self.module_number_var.get())
                         else None)
        # Validation per mode
        if mode_key == "add_module_kit":
            if not kit_code:
                custom_popup(self.parent, "Error", "Select a Kit.", "error"); return
            if not kit_number:
                custom_popup(self.parent, "Error", "Select a Kit Number.", "error"); return
        elif mode_key == "add_items_kit":
            if not kit_number:
                custom_popup(self.parent, "Error", "Select a Kit Number.", "error"); return
        elif mode_key == "add_items_module":
            if not module_code_selected:
                custom_popup(self.parent, "Error", "Select a Module.", "error"); return
            if not module_number:
                custom_popup(self.parent, "Error", "Select a Module Number.", "error"); return
        desc = get_item_description(code)
        item_type = detect_type(code, desc)
        std_qty = 0
        qty_to_receive = 1 if item_type.upper() in ('KIT', 'MODULE') else 1  # Set to 1 for Kit/Module
        exp_date = ""
        # Determine parent iid depending on mode
        if mode_key == "add_module_kit":
            parent_iid = ''
        elif scen_module_mode:
            parent_iid = ''
        elif mode_key == "add_items_module":
            parent_iid = self.code_to_iid.get(module_code_selected, '')
        elif mode_key == "add_items_kit":
            parent_iid = self.code_to_iid.get(kit_code, '')
        else:
            parent_iid = ''
        kit_number_for_id = kit_number or "None"
        module_number_for_id = module_number or "None"
        kit_display = "-----" if scen_module_mode else (kit_code or "-----")
        module_display = module_code_selected or "-----"
        if code in self.code_to_iid:
            self.status_var.set(f"Item {code} already in tree")
            return
        # Adding a MODULE (with subtree strategies) in add_module_kit / add_module_scenario
        if item_type.upper() == "MODULE" and mode_key in ("add_module_kit", "add_module_scenario"):
            module_number_for_id = self.select_module_popup(kit_number, code, desc)
            if not module_number_for_id:
                self.status_var.set(f"Cancelled adding module {code}")
                return
            # Try retrieval strategies
            comps1 = self.fetch_full_module_subtree(self.selected_scenario_id, code)
            comps2 = self.fetch_items_by_module_column(self.selected_scenario_id, code)
            comps3 = self.fetch_items_by_code_prefix(self.selected_scenario_id, code)
            candidates = [("treecode", comps1), ("module_col", comps2), ("code_prefix", comps3)]
            chosen_label, components = None, []
            for lbl, lst in sorted(candidates, key=lambda x: len(x[1]), reverse=True):
                if lst:
                    chosen_label, components = lbl, lst
                    break
            if not components:
                # Insert only the module row
                iid_only = self.tree.insert(parent_iid, "end", values=(
                    code, desc, "MODULE", kit_display, code,
                    std_qty, 1, "", "",
                    "", "", "", ""
                ))
                self.tree.item(iid_only, tags=("module",))
                self.code_to_iid[code] = iid_only
                self.row_data[iid_only] = {
                    'unique_id': self.generate_unique_id(
                        self.selected_scenario_id,
                        None if scen_module_mode else (kit_code or code),
                        code, None, std_qty, "",
                        kit_number_for_id if not scen_module_mode else "None",
                        module_number_for_id
                    ),
                    'kit_number': None if scen_module_mode else kit_number_for_id,
                    'module_number': module_number_for_id,
                    'treecode': None
                }
                self.recompute_exp_groups()
                self.status_var.set(f"Added module {code} (no descendants)")
                return
            SEG = 3
            tc_map = {}
            for comp in sorted(components, key=lambda x: x['treecode'] or ""):
                tc = comp['treecode']
                parent_tc = tc[:-SEG] if tc and len(tc) > SEG else None
                parent_ref = tc_map.get(parent_tc, parent_iid)
                comp_type = comp['type'].upper()
                iid = self.tree.insert(parent_ref, "end", values=(
                    comp['code'], comp['description'], comp['type'],
                    kit_display,
                    comp['module'] if comp['module'] != "-----"
                    else (comp['code'] if comp_type == "MODULE" else comp['module']),
                    comp['std_qty'], comp['std_qty'], "", "",
                    "", "", "", ""
                ))
                if comp_type == "MODULE":
                    self.tree.item(iid, tags=("module",))
                elif comp_type == "KIT":
                    self.tree.item(iid, tags=("kit",))
                tc_map[tc] = iid
                self.code_to_iid[comp['code']] = iid
                if scen_module_mode:
                    kit_part_for_id = None
                    kit_number_part = "None"
                else:
                    kit_part_for_id = kit_code if comp_type != 'KIT' else comp['code']
                    kit_number_part = kit_number_for_id
                module_part_for_id = comp['module'] if comp_type != 'MODULE' else comp['code']
                item_part_for_id = comp['item'] if comp_type == 'ITEM' else None
                unique_id = self.generate_unique_id(
                    self.selected_scenario_id,
                    kit_part_for_id,
                    module_part_for_id,
                    item_part_for_id,
                    comp['std_qty'],
                    None,
                    kit_number_part,
                    module_number_for_id
                )
                self.row_data[iid] = {
                    'unique_id': unique_id,
                    'kit_number': None if scen_module_mode else kit_number_for_id,
                    'module_number': module_number_for_id,
                    'treecode': tc
                }
            self.recompute_exp_groups()
            self.status_var.set(f"Added module {code} with descendants (strategy {chosen_label})")
            return
        # Simple single row insertion
        kit_for_id = None if scen_module_mode else (kit_code if item_type.upper() != "KIT" else code)
        module_for_id = module_code_selected if item_type.upper() != "MODULE" else code
        item_for_id = code if item_type.upper() == "ITEM" else None
        unique_id = self.generate_unique_id(
            self.selected_scenario_id,
            kit_for_id,
            module_for_id,
            item_for_id,
            std_qty,
            exp_date,
            kit_number_for_id,
            module_number_for_id
        )
        iid = self.tree.insert(parent_iid, "end", values=(
            code, desc, item_type,
            kit_display if not scen_module_mode else "-----",
            module_display,
            std_qty, qty_to_receive, "", "",
            "", "", "", ""
        ))
        self.code_to_iid[code] = iid
        if item_type.upper() == "KIT":
            self.tree.item(iid, tags=("kit",))
        elif item_type.upper() == "MODULE":
            self.tree.item(iid, tags=("module",))
        self.row_data[iid] = {
            'unique_id': unique_id,
            'kit_number': None if scen_module_mode else kit_number_for_id,
            'module_number': module_number_for_id
        }
        self.recompute_exp_groups()
        self.status_var.set(f"Added item {code}")
    # -----------------------------------------------------------------
    # Uniqueness helpers
    # -----------------------------------------------------------------
    def is_kit_number_unique(self, kit_number: str) -> bool:
        if not kit_number or kit_number.strip().lower() == 'none':
            return False
        conn = connect_db()
        if conn is None: return False
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT COUNT(*) FROM stock_data
                WHERE kit_number=? AND kit_number!='None'
            """, (kit_number.strip(),))
            return cur.fetchone()[0] == 0
        finally:
            cur.close(); conn.close()
    def is_module_number_unique(self, kit_number: str, module_number: str) -> bool:
        if not module_number or module_number.strip().lower() == 'none':
            return False
        conn = connect_db()
        if conn is None: return False
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT COUNT(*) FROM stock_data
                WHERE module_number=? AND module_number!='None'
            """, (module_number.strip(),))
            return cur.fetchone()[0] == 0
        finally:
            cur.close(); conn.close()
    def _tc_len_type(self, treecode: str | None):
        """
        Return a synthetic structural type based strictly on treecode length:
          5 -> KIT
          8 -> MODULE
          11 -> ITEM
        Fallback: returns None if unrecognized.
        """
        if not treecode:
            return None
        L = len(treecode)
        if L == 5:
            return "KIT"
        if L == 8:
            return "MODULE"
        if L == 11:
            return "ITEM"
        return None
   
    def generate_unique_id(self, scenario_id, kit, module, item, std_qty, exp_date, kit_number, module_number):
        return f"{scenario_id}/{kit or 'None'}/{module or 'None'}/{item or 'None'}/{std_qty}/{exp_date or 'None'}/{kit_number or 'None'}/{module_number or 'None'}"
    def ensure_module_number_consistency(self):
        for top_iid in self.tree.get_children():
            self._ensure_module_subtree(top_iid)
    def _ensure_module_subtree(self, iid):
        vals = self.tree.item(iid, "values")
        if not vals: return
        code, desc, tfield = vals[0], vals[1], vals[2]
        if (tfield or "").upper() == "MODULE":
            rd = self.row_data.get(iid, {})
            mod_num = rd.get('module_number')
            kit_num = rd.get('kit_number')
            if mod_num:
                self.propagate_module_number_to_children(iid, kit_num, mod_num)
        for child in self.tree.get_children(iid):
            self._ensure_module_subtree(child)
    def propagate_module_number_to_children(self, module_iid, kit_number, module_number):
        if module_iid not in self.row_data:
            return
        stack = list(self.tree.get_children(module_iid))
        while stack:
            child_iid = stack.pop()
            vals = self.tree.item(child_iid, "values")
            if not vals: continue
            ctype = (vals[2] or "").upper()
            if ctype == "MODULE":
                continue
            rd = self.row_data.setdefault(child_iid, {})
            rd['kit_number'] = kit_number
            rd['module_number'] = module_number
            kit_for_id = None if vals[3] == "-----" else vals[3]
            module_for_id = None if vals[4] == "-----" else vals[4]
            item_for_id = vals[0] if ctype == "ITEM" else None
            parsed_exp = parse_expiry(vals[7])
            try:
                std_qty_num = int(vals[5]) if str(vals[5]).isdigit() else 0
            except Exception:
                std_qty_num = 0
            rd['unique_id'] = self.generate_unique_id(
                self.selected_scenario_id,
                kit_for_id,
                module_for_id,
                item_for_id,
                std_qty_num,
                parsed_exp,
                kit_number,
                module_number
            )
            stack.extend(self.tree.get_children(child_iid))
    def ensure_unique_numbers_interactively(self):
        """
        Enforce uniqueness for kit_number (KIT rows) and module_number (MODULE rows).
        Prompts user to rename duplicates or DB conflicts one-by-one.
        Returns True if uniqueness satisfied or user resolved all; False if user cancels.
        """
        try:
            def build_signature():
                entries = []
                for iid in self._gather_full_tree_nodes():
                    vals = self.tree.item(iid, "values")
                    if not vals: continue
                    rtype = (vals[2] or "").upper()
                    if rtype not in ("KIT","MODULE"): continue
                    rd = self.row_data.get(iid, {})
                    num = rd.get('kit_number') if rtype == "KIT" else rd.get('module_number')
                    entries.append((rtype, (num or ""), vals[0]))
                entries.sort()
                return tuple(entries)
            def collect_maps():
                kit_map = {}
                mod_map = {}
                for iid in self._gather_full_tree_nodes():
                    vals = self.tree.item(iid, "values")
                    if not vals: continue
                    rtype = (vals[2] or "").upper()
                    if rtype not in ("KIT","MODULE"): continue
                    rd = self.row_data.get(iid, {})
                    if rtype == "KIT":
                        num = rd.get('kit_number')
                        if num and str(num).lower() != "none":
                            kit_map.setdefault(num, []).append(iid)
                    else:
                        num = rd.get('module_number')
                        if num and str(num).lower() != "none":
                            mod_map.setdefault(num, []).append(iid)
                return kit_map, mod_map
            def prompt_single_rename(kind_label, iid, current_number):
                vals = self.tree.item(iid, "values")
                code = vals[0] if vals else "UNKNOWN"
                while True:
                    new_val = simpledialog.askstring(
                        f"Rename {kind_label} Number",
                        f"{kind_label} '{code}' uses duplicate {kind_label} Number '{current_number}'.\n"
                        f"Enter NEW {kind_label} Number (Cancel aborts SAVE):",
                        parent=self.parent
                    )
                    if new_val is None:
                        custom_popup(self.parent, "Cancelled",
                                     f"{kind_label} number resolution cancelled. Save aborted.",
                                     "warning")
                        return False
                    new_val = new_val.strip()
                    if not new_val:
                        continue
                    kit_map_now, mod_map_now = collect_maps()
                    present_map = kit_map_now if kind_label == "Kit" else mod_map_now
                    if new_val in present_map and iid not in present_map[new_val]:
                        custom_popup(self.parent, "Duplicate",
                                     f"{kind_label} Number '{new_val}' already used. Try another.",
                                     "error")
                        continue
                    rd = self.row_data.setdefault(iid, {})
                    if kind_label == "Kit":
                        rd['kit_number'] = new_val
                    else:
                        old_mod = rd.get('module_number')
                        if old_mod != new_val:
                            rd['module_number'] = new_val
                            module_code = vals[4] or vals[0]
                            self.renamed_modules.append({
                                'module_code': module_code,
                                'old_module_number': old_mod,
                                'new_module_number': new_val,
                                'kit_number': rd.get('kit_number'),
                                'scenario_id': self.selected_scenario_id
                            })
                            if (vals[2] or "").upper() == "MODULE":
                                self.propagate_module_number_to_children(iid, rd.get('kit_number'), new_val)
                    return True
            current_sig = build_signature()
            if getattr(self, "_last_uniqueness_signature", None) == current_sig:
                return True
            while True:
                kit_map, mod_map = collect_maps()
                duplicate_fixed = False
                for number, iids in kit_map.items():
                    if len(iids) > 1:
                        target = iids[0]
                        if not prompt_single_rename("Kit", target, number):
                            return False
                        duplicate_fixed = True
                        break
                if duplicate_fixed:
                    continue
                for number, iids in mod_map.items():
                    if len(iids) > 1:
                        target = iids[0]
                        if not prompt_single_rename("Module", target, number):
                            return False
                        duplicate_fixed = True
                        break
                if duplicate_fixed:
                    continue
                break
            conn = connect_db()
            if conn:
                cur = conn.cursor()
                def db_in_use(field, value):
                    cur.execute(f"""
                        SELECT COUNT(*) FROM stock_data
                        WHERE {field}=? AND unique_id LIKE ?
                    """, (value, f"{self.selected_scenario_id}/%"))
                    return cur.fetchone()[0] > 0
                while True:
                    conflict = False
                    kit_map, mod_map = collect_maps()
                    for number, iids in kit_map.items():
                        if db_in_use("kit_number", number):
                            if not prompt_single_rename("Kit", iids[0], number):
                                cur.close(); conn.close()
                                return False
                            conflict = True
                            break
                    if conflict: continue
                    for number, iids in mod_map.items():
                        if db_in_use("module_number", number):
                            if not prompt_single_rename("Module", iids[0], number):
                                cur.close(); conn.close()
                                return False
                            conflict = True
                            break
                    if not conflict:
                        break
                cur.close()
                conn.close()
            self._last_uniqueness_signature = build_signature()
            return True
        except Exception as e:
            logging.error("ensure_unique_numbers_interactively error", exc_info=True)
            return True # fail-open
    # -----------------------------------------------------------------
    # Document number & transaction log
    # -----------------------------------------------------------------
    def generate_document_number(self, in_type_text: str) -> str:
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").strip().upper()
        base_map = {
            "In MSF": "IMSF",
            "In Local Purchase": "ILP",
            "In from Quarantine": "IFQ",
            "In Donation": "IDN",
            "Return from End User": "IREU",
            "In Supply Non-MSF": "ISNM",
            "In Borrowing": "IBR",
            "In Return of Loan": "IRL",
            "In Correction of Previous Transaction": "ICOR"
        }
        raw = (in_type_text or "").strip()
        abbr = None
        for k, v in base_map.items():
            if re.sub(r'[^a-z0-9]+','', k.lower()) == re.sub(r'[^a-z0-9]+','', raw.lower()):
                abbr = v; break
        if not abbr:
            tokens = re.split(r'\s+', raw.upper())
            stop = {"OF","FROM","THE","AND","DE","DU","DES","LA","LE","LES"}
            letters = []
            for t in tokens:
                if not t or t in stop: continue
                if t == "MSF": letters.append("MSF")
                else: letters.append(t[0])
            if not letters:
                abbr = (raw[:4].upper() or "DOC").replace(" ","")
            else:
                abbr = "".join(letters)
            if len(abbr) > 8:
                abbr = abbr[:8]
        now = _dt.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"
        conn = connect_db()
        serial = 1
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT document_number FROM stock_transactions
                    WHERE document_number LIKE ?
                    ORDER BY document_number DESC
                    LIMIT 1
                """, (prefix + "/%",))
                row = cur.fetchone()
                if row and row[0]:
                    last_serial = row[0].rsplit("/",1)[-1]
                    if last_serial.isdigit():
                        serial = int(last_serial) + 1
            except Exception:
                pass
            finally:
                cur.close(); conn.close()
        document_number = f"{prefix}/{serial:04d}"
        self.current_document_number = document_number
        return document_number
    def log_transaction(self,
                        unique_id,
                        code,
                        description,
                        expiry_date,
                        batch_number,
                        scenario,
                        kit,
                        module,
                        qty_in,
                        in_type,
                        qty_out,
                        out_type,
                        third_party,
                        end_user,
                        remarks,
                        movement_type,
                        document_number,
                        comments=None):
        """
        Insert a transaction row. Extended to persist 'comments' (Adopted …) into
        stock_transactions.Comments (column must exist).
        """
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed")
        cur = conn.cursor()
        try:
            # Detect if Comments column exists (case-insensitive)
            cur.execute("PRAGMA table_info(stock_transactions)")
            cols = {row[1].lower(): row[1] for row in cur.fetchall()}
            has_comments = 'comments' in cols

            if has_comments:
                cur.execute(f"""
                    INSERT INTO stock_transactions
                    (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
                     Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
                     Third_Party, End_User, Remarks, Movement_Type, document_number, Comments)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    _dt.today().strftime('%Y-%m-%d'),
                    _dt.now().time().strftime('%H:%M:%S'),
                    unique_id, code, description, expiry_date, batch_number,
                    scenario, kit, module, qty_in, in_type, qty_out, out_type,
                    third_party, end_user, remarks, movement_type, document_number, comments
                ))
            else:
                # Backward compatibility if Comments column missing
                cur.execute(f"""
                    INSERT INTO stock_transactions
                    (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
                     Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
                     Third_Party, End_User, Remarks, Movement_Type, document_number)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    _dt.today().strftime('%Y-%m-%d'),
                    _dt.now().time().strftime('%H:%M:%S'),
                    unique_id, code, description, expiry_date, batch_number,
                    scenario, kit, module, qty_in, in_type, qty_out, out_type,
                    third_party, end_user, remarks, movement_type, document_number
                ))

            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"Error logging transaction: {e}")
            raise
        finally:
            cur.close()
            conn.close()
    # -----------------------------------------------------------------
    # Tree iteration helpers
    # -----------------------------------------------------------------
    def _gather_full_tree_nodes(self):
        result = []
        stack = list(self.tree.get_children(""))
        while stack:
            iid = stack.pop()
            result.append(iid)
            stack.extend(self.tree.get_children(iid))
        return result
    def _module_subtree_iids(self, module_iid):
        out = []
        stack = list(self.tree.get_children(module_iid))
        while stack:
            nid = stack.pop()
            out.append(nid)
            stack.extend(self.tree.get_children(nid))
        return out
    def _collect_structural_nodes(self):
        structural = []
        def walk(iid):
            if not self.tree.exists(iid): return
            vals = self.tree.item(iid, "values")
            if vals:
                rt = (vals[2] or "").strip().upper()
                if rt in ("KIT","MODULE","PRIMARY","SECONDARY"):
                    structural.append(iid)
            for ch in self.tree.get_children(iid):
                walk(ch)
        for top in self.tree.get_children(""):
            walk(top)
        return structural

    def _compute_row_comment_text(self, rd):
        """Utility (not used directly above but available if you refactor)."""
        if not rd or rd.get("user_manual_expiry") or not rd.get("expiry_iso"):
            return ""
        if rd.get("exp_module_iso") and rd["expiry_iso"] == rd["exp_module_iso"]:
            return "Adopted Module expiry"
        if rd.get("exp_kit_iso") and rd["expiry_iso"] == rd["exp_kit_iso"]:
            return "Adopted Kit expiry"
        return ""

    def _structural_iids(self):
        out = []
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid): continue
            vals = self.tree.item(iid, "values")
            if not vals: continue
            rt = (vals[2] or "").strip().upper()
            if rt in ("KIT","MODULE","PRIMARY","SECONDARY"):
                out.append(iid)
        return out
    def _get_row_type(self, iid):
        if not self.tree.exists(iid): return ""
        vals = self.tree.item(iid, "values")
        if not vals: return ""
        return (vals[2] or "").strip().upper()
    def _split_treecode(self, tc: str | None):
        if not tc:
            return []
        # 3-char chunks
        return [tc[i:i+3] for i in range(0, len(tc), 3)]
    def _auto_fill_expiry_with_precedence(self):
        """
        Improved precedence logic:
        - If user has NOT manually entered an expiry:
            * Prefer module earliest (exp_module_iso):
                - If no current expiry -> adopt module earliest.
                - If current expiry == kit earliest but differs from module earliest -> upgrade to module earliest.
            * If no module earliest -> adopt kit earliest only if there is no current expiry.
        - Never override a manual expiry.
        """
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = list(self.tree.item(iid, "values"))
            if len(vals) < 13:
                continue
            rd = self.row_data.setdefault(iid, {})
            # Skip if user entered manually
            if rd.get("user_manual_expiry"):
                continue
            exp_mod = rd.get("exp_module_iso")
            exp_kit = rd.get("exp_kit_iso")
            current = rd.get("expiry_iso")
            changed = False
            if exp_mod:
                # Adopt or upgrade to module earliest
                if current is None:
                    rd["expiry_iso"] = exp_mod
                    changed = True
                elif current == exp_kit and exp_mod != exp_kit:
                    rd["expiry_iso"] = exp_mod
                    changed = True
            else:
                # No module earliest: adopt kit earliest if empty
                if current is None and exp_kit:
                    rd["expiry_iso"] = exp_kit
                    changed = True
            if changed:
                vals[7] = format_expiry_display(rd["expiry_iso"])
                self.tree.item(iid, values=tuple(vals))
    def _recalc_after_quantity_edit(self, edited_iid, was_structural):
        """
        Recalculate quantities after a user edits one row's quantity.
        If structural (KIT/MODULE), recompute entire subtree (affects descendants).
        If item, only that row's display is re-derived (but still using current ancestor multipliers).
        """
        if not self.tree.exists(edited_iid):
            return
        # If structural, full recompute:
        if was_structural:
            changed = self._compute_multiplicative_quantities()
            if changed > 1:
                self.status_var.set(f"Qty recalculated for {changed - 1} descendant row(s) due to structural change.")
            else:
                self.status_var.set("Qty updated.")
        else:
            # Non-structural (ITEM) edit: only reapply formula to that row
            # Re-run compute but capture its path multipliers quickly.
            # Simpler approach: temporarily compute all (small overhead) for consistency.
            changed = self._compute_multiplicative_quantities()
            if changed > 0:
                self.status_var.set("Qty updated.")
    def update_row_comment(self, iid, force=False, sticky_mode=True):
        """
        NEW simplified comment rule set:

          Structural rows (KIT / MODULE / PRIMARY / SECONDARY): always blank.

          For ITEM (or other non-structural):
            - If no expiry_iso OR user_manual_expiry: blank.
            - else if expiry_iso == exp_module_iso: "Adopted Module expiry"
            - else if expiry_iso == exp_kit_iso: "Adopted Kit expiry"
            - else blank.

        The 'sticky_mode' parameter is kept for interface compatibility but is
        not needed now: logic always fully re-evaluates (force respected).
        """
        if not self.tree or not self.tree.exists(iid):
            return

        vals = list(self.tree.item(iid, "values"))
        if len(vals) < 13:
            return

        row_type = (vals[2] or "").strip().upper()
        rd = self.row_data.setdefault(iid, {})

        # Structural rows: blank comment
        if row_type in ("KIT", "PRIMARY", "MODULE", "SECONDARY"):
            if vals[11] != "":
                vals[11] = ""
                rd['sticky_comment'] = ""
                self.tree.item(iid, values=tuple(vals))
            return

        expiry_iso = rd.get("expiry_iso")
        exp_mod_iso = rd.get("exp_module_iso")
        exp_kit_iso = rd.get("exp_kit_iso")
        manual = rd.get("user_manual_expiry")

        if not expiry_iso or manual:
            new_comment = ""
        else:
            if exp_mod_iso and expiry_iso == exp_mod_iso:
                new_comment = "Adopted Module expiry"
            elif exp_kit_iso and expiry_iso == exp_kit_iso:
                new_comment = "Adopted Kit expiry"
            else:
                new_comment = ""

        if vals[11] != new_comment:
            vals[11] = new_comment
            rd['sticky_comment'] = new_comment
            self.tree.item(iid, values=tuple(vals))

    # -----------------------------------------------------------------
    # Group expiry computation + auto-fill + comments
    # -----------------------------------------------------------------
    def recompute_exp_groups(self):
        """
        Recompute earliest kit/module expiries (by treecode prefixes) and
        update hidden columns. Then auto fill missing expiries (no '(adopted)').
        """
        if not self.tree or not self.selected_scenario_id:
            return
        earliest_kit = {}
        earliest_module = {}
        # Gather existing expiries
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 13:
                continue
            rd = self.row_data.setdefault(iid, {})
            tc = rd.get('treecode')
            if not tc:
                continue
            iso = rd.get("expiry_iso")
            if not iso:
                cell_raw = (vals[7] or "").strip()
                iso = parse_expiry(cell_raw)
                if iso:
                    rd["expiry_iso"] = iso
            if not iso:
                continue
            if len(tc) >= 5:
                prev = earliest_kit.get(tc[:5])
                if prev is None or iso < prev:
                    earliest_kit[tc[:5]] = iso
            if len(tc) >= 8:
                prevm = earliest_module.get(tc[:8])
                if prevm is None or iso < prevm:
                    earliest_module[tc[:8]] = iso
        # Write group columns
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = list(self.tree.item(iid, "values"))
            if len(vals) < 13:
                continue
            rd = self.row_data.setdefault(iid, {})
            tc = rd.get('treecode')
            exp_mod_disp = ""
            exp_kit_disp = ""
            rd["exp_module_iso"] = None
            rd["exp_kit_iso"] = None
            if tc:
                if len(tc) >= 8 and tc[:8] in earliest_module:
                    rd["exp_module_iso"] = earliest_module[tc[:8]]
                    exp_mod_disp = format_expiry_display(rd["exp_module_iso"])
                if len(tc) >= 5 and tc[:5] in earliest_kit:
                    rd["exp_kit_iso"] = earliest_kit[tc[:5]]
                    exp_kit_disp = format_expiry_display(rd["exp_kit_iso"])
            if vals[9] != exp_mod_disp or vals[10] != exp_kit_disp:
                vals[9] = exp_mod_disp
                vals[10] = exp_kit_disp
                self.tree.item(iid, values=tuple(vals))
        # Auto fill missing
        self._auto_fill_expiry_with_precedence()
        # Force upgrade for all non-manual items if module differs from current/kit
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid): continue
            rd = self.row_data.get(iid, {})
            if rd.get("user_manual_expiry"): continue  # Skip manuals
            current = rd.get("expiry_iso")
            exp_mod = rd.get("exp_module_iso")
            exp_kit = rd.get("exp_kit_iso")
            if exp_mod and current != exp_mod:  # Always upgrade to module if available and not matching
                rd["expiry_iso"] = exp_mod
                vals = list(self.tree.item(iid, "values"))
                vals[7] = format_expiry_display(exp_mod)
                self.tree.item(iid, values=tuple(vals))
                # Clear auto tag if needed
                tags = list(self.tree.item(iid, "tags") or [])
                if "auto_expiry_gray" in tags: tags.remove("auto_expiry_gray")
                self.tree.item(iid, tags=tuple(tags))
        # Tag + comments
        for iid in self._gather_full_tree_nodes():
            self._validate_and_tag_row(iid, force=False)
        self._rederive_comments(force=False, sticky_mode=True)
        self.update_unique_ids_and_column()
    def _auto_fill_expiry_with_precedence(self):
        """
        Improved precedence logic:
        - If user has NOT manually entered an expiry:
            * If there is a module earliest (exp_module_iso):
                - If no current expiry -> adopt module earliest.
                - Else if current expiry equals kit earliest (exp_kit_iso) and differs from module earliest -> upgrade to module earliest.
            * Else (no module earliest) and no current expiry, but kit earliest exists -> adopt kit earliest.
        - Never override a manual expiry.
        This lets items first temporarily adopt kit expiry (e.g. from structural propagation)
        and later 'upgrade' to module expiry when module earliest becomes known, without needing adopt_source.
        """
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = list(self.tree.item(iid, "values"))
            if len(vals) < 13:
                continue
            rd = self.row_data.setdefault(iid, {})
            # Skip manual entries
            if rd.get("user_manual_expiry"):
                continue
            exp_mod = rd.get("exp_module_iso")
            exp_kit = rd.get("exp_kit_iso")
            current = rd.get("expiry_iso")
            changed = False
            if exp_mod:
                # Upgrade or adopt module earliest
                if current is None:
                    rd["expiry_iso"] = exp_mod
                    changed = True
                elif current != exp_mod and current == exp_kit:
                    # Current matches kit earliest but module earliest is different -> upgrade
                    rd["expiry_iso"] = exp_mod
                    changed = True

            if changed:
                # Update expiry cell (no '(adopted)' suffix)
                vals[7] = format_expiry_display(rd["expiry_iso"])
                self.tree.item(iid, values=tuple(vals))
    def _rederive_comments(self, force=False, sticky_mode=True):
        """
        Bulk comment updater applying the new adoption rules.
        'force' and 'sticky_mode' kept for signature compatibility.
        """
        if not self.tree:
            return
        for iid in self._gather_full_tree_nodes():
            self.update_row_comment(iid, force=True, sticky_mode=False)


    def _auto_adopt_mandatory_item_expiries(self):
        """
        For items requiring expiry (DB flag) that still have no expiry_iso,
        adopt earliest module then kit (same precedence). No '(adopted)' text.
        """
        adopted = 0
        for iid in self._gather_full_tree_nodes():
            if not self.tree.exists(iid):
                continue
            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 13:
                continue
            type_field = (vals[2] or "").strip().upper()
            if type_field != "ITEM":
                continue
            rd = self.row_data.setdefault(iid, {})
            if rd.get("expiry_iso"):
                continue
            code = vals[0]
            if not check_expiry_required(code):
                continue
            raw_cell = (vals[7] or "").strip()
            existing_iso = parse_expiry(raw_cell)
            if existing_iso:
                rd["expiry_iso"] = existing_iso
                continue
            iso = rd.get("exp_module_iso") or rd.get("exp_kit_iso")
            if not iso:
                continue
            rd["expiry_iso"] = iso
            new_vals = list(vals)
            new_vals[7] = format_expiry_display(iso)
            self.tree.item(iid, values=tuple(new_vals))
            adopted += 1
        self._rederive_comments()
        return adopted
    def _prompt_for_structural_expiries(self):
        structural = self._collect_structural_nodes()
        if not structural:
            return True
        title = lang.t("receive_kit.expiry_prompt_title","Expiry Required")
        for iid in structural:
            if not self.tree.exists(iid):
                continue
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            code, desc = vals[0], vals[1]
            rd = self.row_data.setdefault(iid, {})
            if rd.get("expiry_iso"):
                continue
            cell_raw = (vals[7] or "").strip()
            if cell_raw:
                parsed = parse_expiry(cell_raw)
                if parsed:
                    rd["expiry_iso"] = parsed
                    self.tree.set(iid, "expiry_date", format_expiry_display(parsed))
                    continue
            while not rd.get("expiry_iso"):
                prompt_msg = lang.t(
                    "receive_kit.expiry_prompt_message",
                    "Enter expiry (YYYY-MM-DD, DD/MM/YYYY, MM/YYYY, etc.) for {code} - {desc}:",
                    code=code, desc=desc
                )
                user_input = simpledialog.askstring(title, prompt_msg, parent=self.parent)
                if user_input is None:
                    custom_popup(
                        self.parent,
                        lang.t("receive_kit.cancelled_title","Cancelled"),
                        lang.t("receive_kit.expiry_cancelled","Save cancelled: missing structural expiry."),
                        "warning"
                    )
                    return False
                user_input = user_input.strip()
                if not user_input:
                    continue
                parsed = parse_expiry(user_input)
                if not parsed:
                    try:
                        parsed = strict_parse_expiry(user_input)
                    except Exception:
                        parsed = None
                if not parsed:
                    custom_popup(
                        self.parent,
                        lang.t("receive_kit.invalid_expiry_title","Invalid Expiry"),
                        lang.t("receive_kit.invalid_expiry_msg",
                               "Format not recognized. Examples:\n05/10/2029\n2029-10-05\n10/2029\n2029-10"),
                        "error"
                    )
                    continue
                rd["expiry_iso"] = parsed
                self.tree.set(iid, "expiry_date", format_expiry_display(parsed))
        return True
    def _propagate_structural_expiries_top_down(self):
        """
        Propagate structural (KIT / MODULE) expiries to descendant ITEM rows that
        have no manual expiry. Module precedence still applies (a module expiry
        can replace a previously propagated kit expiry).
        """
        structural = []
        for iid in self._collect_structural_nodes():
            rd = self.row_data.get(iid, {})
            tc = rd.get('treecode') or ""
            depth = len(tc) // 3 if tc else 0
            structural.append((depth, iid))
        structural.sort(reverse=True)
        def propagate_from(struct_iid, iso, is_module):
            stack = list(self.tree.get_children(struct_iid))
            while stack:
                nid = stack.pop()
                if not self.tree.exists(nid):
                    continue
                stack.extend(self.tree.get_children(nid))
                vals = self.tree.item(nid, "values")
                if not vals:
                    continue
                rtype = (vals[2] or "").strip().upper()
                if rtype in ("KIT","MODULE","PRIMARY","SECONDARY"):
                    continue
                rd = self.row_data.setdefault(nid, {})
                if rd.get("user_manual_expiry"):
                    continue
                current_iso = rd.get("expiry_iso")
                if current_iso is None:
                    rd["expiry_iso"] = iso
                    new_vals = list(vals)
                    new_vals[7] = format_expiry_display(iso)
                    self.tree.item(nid, values=tuple(new_vals))
                else:
                    if is_module and current_iso != iso and not rd.get("user_manual_expiry"):
                        rd["expiry_iso"] = iso
                        new_vals = list(vals)
                        new_vals[7] = format_expiry_display(iso)
                        self.tree.item(nid, values=tuple(new_vals))
        for _depth, iid in structural:
            rd = self.row_data.get(iid, {})
            iso = rd.get("expiry_iso")
            if not iso:
                continue
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            rtype = (vals[2] or "").strip().upper()
            if rtype in ("MODULE","SECONDARY"):
                propagate_from(iid, iso, is_module=True)
            elif rtype in ("KIT","PRIMARY"):
                propagate_from(iid, iso, is_module=False)
        self._rederive_comments()
    def _ensure_expiry_when_no_structurals(self) -> bool:
        """
        If there are no KIT/MODULE structural rows, ensure every row has an expiry.
        Prompts user for one date to apply to all if none can be inferred.
        """
        all_iids = self._gather_full_tree_nodes()
        if not all_iids:
            return True
        have_any = any(self.row_data.get(iid, {}).get("expiry_iso") for iid in all_iids)
        if have_any:
            return True
        title = lang.t("receive_kit.expiry_prompt_title","Expiry Required")
        prompt_msg = lang.t(
            "receive_kit.expiry_prompt_message",
            "Enter expiry (YYYY-MM-DD, DD/MM/YYYY, MM/YYYY, etc.) for ALL items:"
        )
        while True:
            user_input = simpledialog.askstring(title, prompt_msg, parent=self.parent)
            if user_input is None:
                custom_popup(
                    self.parent,
                    lang.t("receive_kit.cancelled_title","Cancelled"),
                    lang.t("receive_kit.expiry_cancelled","Save cancelled: missing expiry."),
                    "warning"
                )
                return False
            user_input = user_input.strip()
            if not user_input:
                continue
            parsed = parse_expiry(user_input)
            if not parsed:
                try:
                    parsed = strict_parse_expiry(user_input)
                except Exception:
                    parsed = None
            if not parsed:
                custom_popup(
                    self.parent,
                    lang.t("receive_kit.invalid_expiry_title","Invalid Expiry"),
                    lang.t("receive_kit.invalid_expiry_msg",
                           "Format not recognized. Examples:\n05/10/2029\n2029-10-05\n10/2029\n2029-10"),
                    "error"
                )
                continue
            for iid in all_iids:
                rd = self.row_data.setdefault(iid, {})
                if not rd.get("expiry_iso"):
                    rd["expiry_iso"] = parsed
                    vals = self.tree.item(iid, "values")
                    if vals:
                        new_vals = list(vals)
                        new_vals[7] = format_expiry_display(parsed)
                        self.tree.item(iid, values=tuple(new_vals))
            self._rederive_comments()
            return True
    # -----------------------------------------------------------------
    # Duplicate Row (Add New Row)
    # -----------------------------------------------------------------
    def add_new_row(self):
        selected = self.tree.selection()
        if not selected:
            custom_popup(self.parent, "Error", "Please select a row to duplicate", "error")
            return
        if not self.selected_scenario_id:
            custom_popup(self.parent, "Error", "Please select a scenario", "error")
            return
        source_iid = selected[0]
        vals = self.tree.item(source_iid, "values")
        if not vals or len(vals) < 13:
            custom_popup(self.parent, "Error", "Selected row invalid", "error")
            return
        code, description, item_type, kit_val, module_val, std_qty = (
            vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]
        )
        parent_iid = self.tree.parent(source_iid)
        siblings = list(self.tree.get_children(parent_iid))
        try:
            src_index = siblings.index(source_iid)
            insert_index = src_index + 1
        except ValueError:
            insert_index = "end"
        rd_src = self.row_data.get(source_iid, {})
        kit_number = rd_src.get("kit_number") or "None"
        module_number = rd_src.get("module_number") or "None"
        kit_for_id = None if kit_val == "-----" else (kit_val if item_type.upper() != "KIT" else code)
        module_for_id = None if module_val == "-----" else (module_val if item_type.upper() != "MODULE" else code)
        item_for_id = code if item_type.upper() == "ITEM" else None
        try:
            std_qty_int = int(std_qty) if str(std_qty).isdigit() else 0
        except Exception:
            std_qty_int = 0
        new_unique_id = self.generate_unique_id(
            self.selected_scenario_id,
            kit_for_id,
            module_for_id,
            item_for_id,
            std_qty_int,
            None,
            kit_number,
            module_number
        )
        new_iid = self.tree.insert(parent_iid, insert_index, values=(
            code, description, item_type,
            kit_val, module_val,
            std_qty, 1, "", "",
            "", "", "", new_unique_id
        ))
        if item_type.upper() == "KIT":
            self.tree.item(new_iid, tags=("kit",))
        elif item_type.upper() == "MODULE":
            self.tree.item(new_iid, tags=("module",))
        self.row_data[new_iid] = {
            "unique_id": new_unique_id,
            "kit_number": None if kit_number == "None" else kit_number,
            "module_number": None if module_number == "None" else module_number,
            "treecode": self.row_data.get(source_iid, {}).get("treecode")
        }
        self.update_unique_ids_and_column()
        self.tree.selection_set(new_iid)
        self.tree.focus(new_iid)
        self.status_var.set(f"Duplicated {code} below selected row")
    # -----------------------------------------------------------------
    # Dropdown visibility
    # -----------------------------------------------------------------
    def update_dropdown_visibility(self, event=None):
        ttype = self.trans_type_var.get()
        self.end_user_cb.config(state="disabled")
        self.third_party_cb.config(state="disabled")
        self.remarks_entry.config(state="disabled")
        self.end_user_var.set(""); self.third_party_var.set("")
        self.remarks_entry.delete(0, tk.END)
        if ttype in [
            lang.t("receive_kit.in_donation","In Donation"),
            lang.t("receive_kit.in_borrowing","In Borrowing"),
            lang.t("receive_kit.in_return_loan","In Return of Loan")
        ]:
            self.third_party_cb.config(state="readonly")
        elif ttype == lang.t("receive_kit.return_from_end_user","Return from End User"):
            self.end_user_cb.config(state="readonly")
        elif ttype == lang.t("receive_kit.in_correction","In Correction of Previous Transaction"):
            self.remarks_entry.config(state="normal")
    # -----------------------------------------------------------------
    # Rewrite module numbers in DB after renames
    # -----------------------------------------------------------------
    def rewrite_module_number_rows(self):
        if not self.renamed_modules:
            return
        compressed = {}
        for r in self.renamed_modules:
            key = (r['scenario_id'], r['kit_number'], r['module_code'])
            compressed[key] = r
        renames = list(compressed.values())
        conn = connect_db()
        if conn is None:
            return
        cur = conn.cursor()
        try:
            for entry in renames:
                scenario_id = entry['scenario_id']
                kit_number = entry['kit_number']
                module_code = entry['module_code']
                old_mod_num = entry['old_module_number']
                new_mod_num = entry['new_module_number']
                like_pattern = f"{scenario_id}/%/{module_code}/%/%/%/{kit_number}/{old_mod_num}"
                cur.execute("""
                    SELECT unique_id, qty_in, qty_out
                    FROM stock_data
                    WHERE unique_id LIKE ?
                """, (like_pattern,))
                sd_rows = cur.fetchall()
                for (unique_id, qty_in, qty_out) in sd_rows:
                    parts = unique_id.split('/')
                    if len(parts) != 8: continue
                    if parts[-1] != old_mod_num:
                        continue
                    new_unique_id = '/'.join(parts[:-1] + [new_mod_num])
                    cur.execute("SELECT qty_in, qty_out FROM stock_data WHERE unique_id=?",
                                (new_unique_id,))
                    existing = cur.fetchone()
                    if existing:
                        merged_in = existing[0] + qty_in
                        merged_out = existing[1] + qty_out
                        cur.execute("""
                            UPDATE stock_data
                            SET qty_in=?, qty_out=?, module_number=?, updated_at=?
                            WHERE unique_id=?
                        """, (merged_in, merged_out, new_mod_num,
                              _dt.now().strftime('%Y-%m-%d %H:%M:%S'), new_unique_id))
                        cur.execute("DELETE FROM stock_data WHERE unique_id=?", (unique_id,))
                    else:
                        cur.execute("""
                            UPDATE stock_data
                            SET unique_id=?, module_number=?, updated_at=?
                            WHERE unique_id=?
                        """, (new_unique_id, new_mod_num,
                              _dt.now().strftime('%Y-%m-%d %H:%M:%S'), unique_id))
                cur.execute("""
                    SELECT rowid, unique_id, Module FROM stock_transactions
                    WHERE unique_id LIKE ?
                """, (like_pattern,))
                tx_rows = cur.fetchall()
                for (rowid, tx_unique_id, tx_module_val) in tx_rows:
                    parts = tx_unique_id.split('/')
                    if len(parts) != 8: continue
                    if parts[-1] != old_mod_num: continue
                    new_tx_unique_id = '/'.join(parts[:-1] + [new_mod_num])
                    cur.execute("""
                        UPDATE stock_transactions
                        SET unique_id=?, Module=?
                        WHERE rowid=?
                    """, (new_tx_unique_id, new_mod_num, rowid))
            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            logging.error(f"rewrite_module_number_rows error: {e}")
        finally:
            cur.close(); conn.close()
        self.renamed_modules.clear()
    # -----------------------------------------------------------------
    # Recursive save subtree
    # -----------------------------------------------------------------
    def save_subtree(self, iid, invalid_items, exported_rows,
                     kit_number=None, module_number=None,
                     document_number=None, effective_expiry=None):
        """
        Modified to:
          - Capture the 'Comments' column text and persist it to stock_data (comments)
            and stock_transactions (Comments).
        """
        try:
            if not self.tree.exists(iid):
                return True

            vals = self.tree.item(iid, "values")
            if not vals or len(vals) < 13:
                return True

            (code, desc, type_field, kit_val, mod_val, std_qty, qty_str,
             display_expiry, batch_no, exp_module_col, exp_kit_col,
             comments_col, _unique_visible) = vals

            rd_local = self.row_data.get(iid, {})

            node_iso = rd_local.get('expiry_iso')
            if not node_iso:
                clean_disp = (display_expiry or "").replace("(adopted)", "").strip()
                if clean_disp:
                    try:
                        node_iso = strict_parse_expiry(clean_disp)
                        rd_local['expiry_iso'] = node_iso
                    except ValueError:
                        node_iso = None

            final_expiry = node_iso or effective_expiry
            qty_to_receive = int(qty_str) if (qty_str and qty_str.isdigit()) else 0

            parent_kn = kit_number
            parent_mn = module_number
            rd_kn = rd_local.get('kit_number')
            rd_mn = rd_local.get('module_number')

            def norm(v):
                return None if (isinstance(v, str) and v.lower() == "none") else v

            kit_number = norm(rd_kn) or norm(parent_kn)
            module_number = norm(rd_mn) or norm(parent_mn)

            # Ensure kit/module numbers for structural nodes
            if type_field.upper() == "KIT":
                while not kit_number:
                    entered = simpledialog.askstring("Kit Number",
                                                     f"Enter Kit Number for {code}",
                                                     parent=self.parent)
                    if entered is None:
                        return False
                    if entered.strip() and self.is_kit_number_unique(entered.strip()):
                        kit_number = entered.strip()
                        break
                    custom_popup(self.parent, "Error", "Kit Number exists or invalid.", "error")

            if type_field.upper() == "MODULE":
                while not module_number:
                    entered = simpledialog.askstring("Module Number",
                                                     f"Enter Module Number for {code}",
                                                     parent=self.parent)
                    if entered is None:
                        return False
                    if entered.strip() and self.is_module_number_unique(kit_number, entered.strip()):
                        module_number = entered.strip()
                        break
                    custom_popup(self.parent, "Error", "Module Number exists or invalid.", "error")

            # Track module renames (propagate)
            original_mod_num = rd_local.get('module_number')
            if type_field.upper() == "MODULE" and original_mod_num and original_mod_num != module_number:
                self.renamed_modules.append({
                    'module_code': mod_val if mod_val != "-----" else code,
                    'old_module_number': original_mod_num,
                    'new_module_number': module_number,
                    'kit_number': kit_number,
                    'scenario_id': self.selected_scenario_id
                })
                self.propagate_module_number_to_children(iid, kit_number, module_number)

            rd_local['kit_number'] = kit_number
            rd_local['module_number'] = module_number
            if final_expiry:
                rd_local['expiry_iso'] = final_expiry

            kit_for_id = None if kit_val == "-----" else kit_val
            module_for_id = None if mod_val == "-----" else mod_val
            item_part = code if type_field.upper() == "ITEM" else None

            unique_id = self.generate_unique_id(
                self.selected_scenario_id,
                kit_for_id,
                module_for_id,
                item_part,
                std_qty,
                final_expiry,
                kit_number,
                module_number
            )
            rd_local['unique_id'] = unique_id

            cur_vals = list(vals)
            cur_vals[12] = unique_id
            self.tree.item(iid, values=tuple(cur_vals))

            # Persist only if quantity > 0 (movement)
            if qty_to_receive > 0:
                scenario_name = self.scenario_map.get(self.selected_scenario_id, "Unknown")
                # Save to stock_data (comments)
                StockData.add_or_update(
                    unique_id=unique_id,
                    scenario=scenario_name,
                    qty_in=qty_to_receive,
                    exp_date=final_expiry,
                    kit_number=kit_number,
                    module_number=module_number,
                    comments=comments_col or None
                )
                # Transaction log
                self.log_transaction(
                    unique_id=unique_id,
                    code=code,
                    description=desc,
                    expiry_date=final_expiry,
                    batch_number=batch_no or None,
                    scenario=scenario_name,
                    kit=kit_number,
                    module=module_number,
                    qty_in=qty_to_receive,
                    in_type=self.trans_type_var.get() or "stock_in",
                    qty_out=None,
                    out_type=None,
                    third_party=self.third_party_var.get() or None,
                    end_user=self.end_user_var.get() or None,
                    remarks=self.remarks_entry.get().strip() or None,
                    movement_type=self.mode_var.get() or "stock_in",
                    document_number=document_number,
                    comments=comments_col or None
                )

                exported_rows.append({
                    'code': code,
                    'description': desc,
                    'type': type_field,
                    'kit': kit_val,
                    'module': mod_val,
                    'std_qty': std_qty,
                    'qty_to_receive': qty_to_receive,
                    'expiry_date': final_expiry or '',
                    'batch_no': batch_no,
                    'exp_module': exp_module_col,
                    'exp_kit': exp_kit_col,
                    'comments': comments_col,
                    'unique_id': unique_id
                })

            # Recurse
            for child in self.tree.get_children(iid):
                ok = self.save_subtree(
                    child,
                    invalid_items,
                    exported_rows,
                    kit_number,
                    module_number,
                    document_number,
                    effective_expiry=final_expiry
                )
                if not ok:
                    return False
            return True

        except Exception as e:
            logging.error(f"save_subtree error iid={iid}: {e}", exc_info=True)
            return False
       
    # -----------------------------------------------------------------
    # Export
    # -----------------------------------------------------------------
    def export_data(self, rows_to_export=None):
        if self.parent is None or not self.parent.winfo_exists():
            return
        try:
            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)
            current_time = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
            in_type_raw = self.trans_type_var.get() or "Unknown"
            movement_type_raw = self.mode_var.get() or "Unknown"
            def sanitize(s: str) -> str:
                s = re.sub(r'[^A-Za-z0-9]+','_',s)
                s = re.sub(r'_+','_',s)
                return s.strip('_') or "Unknown"
            file_name = f"IsEPREP_{sanitize(movement_type_raw)}_{sanitize(in_type_raw)}_{current_time.replace(':','-')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files","*.xlsx")],
                initialfile=file_name,
                initialdir=default_dir
            )
            if not file_path:
                self.status_var.set("Export cancelled")
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws_title_base = lang.t("receive_kit.title", "Receive Kit-Module")
            ws.title = (ws_title_base[:15] + "-" + sanitize(movement_type_raw)[:12])[:31]
            doc_number = getattr(self, "current_document_number", None)
            if doc_number:
                ws['A1'] = f"Date: {current_time} Document Number: {doc_number}"
            else:
                ws['A1'] = f"Date: {current_time}"
            ws['A1'].font = Font(name="Helvetica", size=10)
            ws['A1'].alignment = Alignment(horizontal="left")
            project_name, project_code = fetch_project_details()
            ws['A2'] = f"{ws_title_base} – Movement: {movement_type_raw}"
            ws['A2'].font = Font(name="Helvetica", size=14)
            ws['A2'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A2:L2')
            ws['A3'] = f"{project_name} - {project_code}"
            ws['A3'].font = Font(name="Helvetica", size=14)
            ws['A3'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A3:L3')
            ws['A4'] = f"IN Type: {in_type_raw}"
            ws['A4'].font = Font(name="Helvetica", size=12)
            ws['A4'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A4:L4')
            ws['A5'] = f"Movement Type: {movement_type_raw}"
            ws['A5'].font = Font(name="Helvetica", size=12)
            ws['A5'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A5:L5')
            ws.append([])
            headers = ["Code","Description","Type","Kit","Module","Std Qty","Qty to Receive",
                       "Expiry Date","Batch No","Exp Module","Exp Kit","Comments"]
            ws.append(headers)
            kit_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            module_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            if rows_to_export:
                source = rows_to_export
            else:
                source = []
                for iid in self._gather_full_tree_nodes():
                    vals = self.tree.item(iid,"values")
                    if not vals: continue
                    qty = vals[6]
                    if qty and qty.isdigit() and int(qty) > 0:
                        source.append({
                            "code":vals[0],"description":vals[1],"type":vals[2],"kit":vals[3],
                            "module":vals[4],"std_qty":vals[5],"qty_to_receive":qty,
                            "expiry_date":vals[7],"batch_no":vals[8],
                            "exp_module":vals[9],"exp_kit":vals[10],"comments":vals[11]
                        })
            for r in source:
                ws.append([
                    r['code'], r['description'], r['type'], r['kit'], r['module'],
                    r['std_qty'], r['qty_to_receive'], r.get('expiry_date',""),
                    r.get('batch_no',""), r.get('exp_module',""), r.get('exp_kit',""), r.get('comments',"")
                ])
                row_idx = ws.max_row
                if r['type'].upper()=="KIT":
                    for cell in ws[row_idx]:
                        cell.fill=kit_fill
                elif r['type'].upper()=="MODULE":
                    for cell in ws[row_idx]:
                        cell.fill=module_fill
            for col in ws.columns:
                max_len=0
                letter=get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[letter].width = min(max_len + 2, 48)
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            wb.save(file_path)
            custom_popup(self.parent, "Success", f"Export successful: {file_path}", "info")
            self.status_var.set(f"Export successful: {file_path}")
        except Exception as e:
            logging.error(f"Export failed: {e}")
            custom_popup(self.parent, "Error", f"Export failed: {e}", "error")
    # -----------------------------------------------------------------
    # Clear / Context Menu
    # -----------------------------------------------------------------
    def clear_search(self):
                self.search_var.set("")
                self.search_listbox.delete(0, tk.END)
                self.tree.delete(*self.tree.get_children())
                for rd in self.row_data.values():
                        rd.pop('sticky_comment', None)
                self.row_data.clear()
                self.code_to_iid.clear()
                self.parent_expiry_map.clear()
                self.status_var.set(lang.t("receive_kit.ready","Ready"))
                self.expiry_validation_active = False
    def clear_form(self):
                self.search_var.set("")
                self.search_listbox.delete(0, tk.END)
                self.tree.delete(*self.tree.get_children())
                for rd in self.row_data.values():
                        rd.pop('sticky_comment', None)
                self.row_data.clear()
                self.code_to_iid.clear()
                self.parent_expiry_map.clear()
                self.scenario_var.set("")
                self.mode_var.set("")
                self.kit_var.set("")
                self.kit_number_var.set("")
                self.module_var.set("")
                self.module_number_var.set("")
                self.trans_type_var.set("")
                self.end_user_var.set("")
                self.third_party_var.set("")
                self.remarks_entry.delete(0, tk.END)
                self.end_user_cb.config(state="disabled")
                self.third_party_cb.config(state="disabled")
                self.remarks_entry.config(state="disabled")
                self.status_var.set(lang.t("receive_kit.ready","Ready"))
                self.expiry_validation_active = False
                self.load_scenarios()
    def show_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        self.tree.selection_set(iid)
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(label=lang.t("receive_kit.add_new_row","Add New Row"),
                         command=self.add_new_row)
        menu.post(event.x_root, event.y_root)
    # Legacy alias
    def save(self, *args, **kwargs):
        self.save_all()
    # -----------------------------------------------------------------
    # Save Workflow
    # -----------------------------------------------------------------
    def save_all(self):
        if self.role.lower() not in ("admin","manager"):
            custom_popup(self.parent,
                         lang.t("receive_kit.perm_error_title","Permission Denied"),
                         lang.t("receive_kit.perm_error_msg","Only admin or manager can save."),
                         "error")
            return
        if not self.tree.get_children():
            custom_popup(self.parent,
                         lang.t("receive_kit.no_rows_title","Nothing to Save"),
                         lang.t("receive_kit.no_rows_msg","No rows present."),
                         "error")
            return
        if not self.trans_type_var.get():
            custom_popup(self.parent,
                         lang.t("receive_kit.in_type_missing_title","IN Type Missing"),
                         lang.t("receive_kit.in_type_missing_msg","Please select an IN Type."),
                         "error")
            return
        if not self.ensure_unique_numbers_interactively():
            return
        self.activate_global_expiry_validation()
        structural_nodes = self._collect_structural_nodes()
        if structural_nodes:
            if not self._prompt_for_structural_expiries():
                return
            still_missing = [self.tree.item(iid,"values")[0]
                             for iid in structural_nodes
                             if not self.row_data.get(iid,{}).get("expiry_iso")]
            if still_missing:
                custom_popup(self.parent,
                             lang.t("receive_kit.missing_expiry_title","Missing Expiry"),
                             "Still missing expiry for:\n" + "\n".join(still_missing),
                             "error")
                return
            self._propagate_structural_expiries_top_down()
        else:
            if not self._ensure_expiry_when_no_structurals():
                return
        for iid in self._gather_full_tree_nodes():
            self._validate_and_tag_row(iid, force=True)
        self.recompute_exp_groups()
        adopted_items = self._auto_adopt_mandatory_item_expiries()
        if adopted_items:
            self.recompute_exp_groups()
        missing_struct_final = [
            self.tree.item(iid, "values")[0]
            for iid in self._structural_iids()
            if not self.row_data.get(iid, {}).get("expiry_iso")
        ]
        if missing_struct_final:
            custom_popup(self.parent,
                         lang.t("receive_kit.missing_expiry_title","Missing Expiry"),
                         "Structural rows still missing expiry:\n" + "\n".join(missing_struct_final),
                         "error")
            return
        for iid in self._gather_full_tree_nodes():
            self._validate_and_tag_row(iid, force=True)
        self.update_unique_ids_and_column()
        review_title = lang.t("receive_kit.review_title","Review Before Saving")
        base_msg = "Structural expiries captured. Mandatory item expiries auto-adopted where possible."
        if adopted_items:
            base_msg += f"\n{adopted_items} mandatory item(s) received an adopted expiry."
        review_message = base_msg + "\nProceed with save?"
        try:
            choice = custom_dialog(
                self.parent,
                review_title,
                review_message,
                buttons=[
                    {"key":"save","text":lang.t("receive_kit.button_save","Save"),
                     "style":"Primary.Popup.TButton"},
                    {"key":"review","text":lang.t("receive_kit.button_review","Review"),
                     "style":"Secondary.Popup.TButton"}
                ]
            )
        except Exception:
            choice = "save" if win_confirm(review_title, review_message) else "review"
        if choice != "save":
            self.status_var.set(lang.t("receive_kit.review_mode_status","Review mode - not saved yet."))
            return
        document_number = self.generate_document_number(self.trans_type_var.get())
        self.ensure_module_number_consistency()
        exported_rows = []
        for root in self.tree.get_children(""):
            if not self.save_subtree(root, [], exported_rows, document_number=document_number):
                custom_popup(self.parent,
                             lang.t("receive_kit.save_error_title","Save Error"),
                             lang.t("receive_kit.save_error_msg","An error occurred during save."),
                             "error")
                return
        self.rewrite_module_number_rows()
        custom_popup(self.parent,
                     lang.t("receive_kit.save_success_title","Success"),
                     lang.t("receive_kit.save_success_msg","Data saved successfully."),
                     "success")
        self.status_var.set(
            lang.t("receive_kit.saved_doc_status","Saved. Document Number: {doc}", doc=document_number)
        )
        if exported_rows and custom_askyesno(
            self.parent,
            lang.t("receive_kit.export_confirm_title","Export"),
            lang.t("receive_kit.export_confirm_msg","Export this movement to Excel?")
        ) == "yes":
            self.export_data(exported_rows)
        self.clear_form()
# ---------------------------------------------------------------------
# VERSION
# ---------------------------------------------------------------------
__version__ = "v1.74"
if __name__ == "__main__":
    root = tk.Tk()
    root.title(f"Receive Kit-Module ({__version__})")
    app = tk.Toplevel(root)
    app.role = "admin"
    StockReceiveKit(app, app, role="admin")
    root.mainloop()