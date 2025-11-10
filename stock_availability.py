"""
stock_availability.py  v1.1

Updates (v1.1):
  * Kit & Module filters changed to dropdown (Combobox) populated from DISTINCT
    stock_data.kit_number / module_number (excluding NULL/'None'). Both include "All".
  * Quantity now strictly uses stock_data.final_qty (NOT qty_in - qty_out). Only
    rows with final_qty > 0 are displayed.
  * Expiry Date column sourced directly from stock_data.exp_date (unchanged source)
    but display format changed to "DD-Month-YYYY" (e.g. 22-July-2025). Month name
    is translation-aware via lang.t("month.<lowercase_english_name>").
    If translation key not found, English month name is used.
  * Expired lots (expiry date < today) are highlighted with a very light red
    background (#FFF5F5) tag 'expired_light'. Kit & Module coloring still applied
    (priority: expired highlighting first but subtle).
  * Internal variable names adjusted (kit_filter_var/module_filter_var) for dropdown.
  * Refresh repopulates kit/module dropdown lists based on current scenario (if filtered)
    and other filters do NOT cascade-limit the distinct lists (for performance simplicity).
  * Minor refactor of calculator to remove fallback quantity computation; final_qty
    must exist in schema (as provided). Robust if column missing: row skipped.

Original Features (v1.0 retained):
  - Management Mode, Scenario, Item search (only for Items), Type filter
  - Expiry Horizon (Months) for inclusion (expired included automatically)
  - AMC computation (Qty_Out where Out_Type='Out MSF' over last N months / N)
  - Simple / Detailed toggle
  - Export to Excel with color coding for KIT (green) and MODULE (light blue)
  - Translation system integration


"""

import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
from datetime import date, datetime
from calendar import monthrange
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from db import connect_db
from language_manager import lang
from popup_utils import custom_popup
from manage_items import get_item_description, detect_type

# ---------------- Theme Constants ----------------
BG_MAIN        = "#F0F4F8"
BG_PANEL       = "#FFFFFF"
COLOR_PRIMARY  = "#2C3E50"
COLOR_BORDER   = "#D0D7DE"
ROW_ALT_COLOR  = "#F7FAFC"
ROW_NORM_COLOR = "#FFFFFF"
BTN_EXPORT     = "#2980B9"
BTN_REFRESH    = "#2563EB"
BTN_CLEAR      = "#7F8C8D"
BTN_TOGGLE     = "#8E44AD"

EXPIRED_LIGHT_COLOR = "#FFF5F5"  # very light red

# ---------------- Helpers ----------------
def add_months(year: int, month: int, delta: int):
    total = year * 12 + (month - 1) + delta
    return total // 12, (total % 12) + 1

def months_between_inclusive(start_ym, end_ym):
    (y1, m1), (y2, m2) = start_ym, end_ym
    return (y2 - y1) * 12 + (m2 - m1) + 1

def month_name_translated(dt_obj: date):
    name_en = dt_obj.strftime("%B")  # English full month
    key = f"month.{name_en.lower()}"
    translated = lang.t(key, fallback=name_en)
    return translated

def format_expiry_display(iso_str: str):
    if not iso_str:
        return ""
    try:
        y, m, d = map(int, iso_str.split("-"))
        dt_obj = date(y, m, d)
        return f"{dt_obj.day:02d}-{month_name_translated(dt_obj)}-{dt_obj.year}"
    except Exception:
        return iso_str  # fallback raw if parse fails

# ---------------- Data Aggregator ----------------
class StockAvailabilityCalculator:
    def __init__(self,
                 scenario_name_map,
                 management_mode="All",
                 scenario_filter="All",
                 kit_filter="All",
                 module_filter="All",
                 item_search="",
                 type_filter="All",
                 expiry_period=12,
                 amc_months=6):
        self.scenario_name_map = scenario_name_map
        self.management_mode = management_mode.strip()
        self.scenario_filter = scenario_filter.strip()
        self.kit_filter = kit_filter.strip()
        self.module_filter = module_filter.strip()
        self.item_search = item_search.strip()
        self.type_filter = type_filter.strip()
        self.expiry_period = max(1, min(expiry_period, 99))
        self.amc_months = max(1, min(amc_months, 99))
        today = date.today()
        self.current_year = today.year
        self.current_month = today.month
        self.today = today

    # ---- Internals ----
    def _table_cols(self, cur, table):
        try:
            cur.execute(f"PRAGMA table_info({table})")
            return [r[1].lower() for r in cur.fetchall()]
        except sqlite3.Error:
            return []

    def _load_stock(self, conn):
        cur = conn.cursor()
        cols = self._table_cols(cur, "stock_data")
        needed = [
            "unique_id","scenario","kit_number","module_number",
            "kit","module","item",
            "qty_in","qty_out","final_qty",
            "exp_date","management_mode","discrepancy","comments"
        ]
        present = [c for c in needed if c in cols]
        if "final_qty" not in present:
            return []  # cannot produce quantity list per spec
        sql = f"SELECT {', '.join(present)} FROM stock_data"
        out = []
        try:
            cur.execute(sql)
            for row in cur.fetchall():
                rd = dict(zip(present, row))
                # Scenario normalization (id->name)
                scen = rd.get("scenario")
                if scen and str(scen).isdigit() and str(scen) in self.scenario_name_map:
                    rd["scenario"] = self.scenario_name_map[str(scen)]
                out.append(rd)
        except sqlite3.Error:
            pass
        finally:
            cur.close()
        return out

    def _amc_map(self, conn):
        cur = conn.cursor()
        mapping = {}
        try:
            tx_cols = self._table_cols(cur, "stock_transactions")
            needed = {"date","qty_out","out_type","code"}
            if not needed.issubset(set(tx_cols)):
                return mapping
            # Date window inclusive
            start_year, start_month = add_months(self.current_year, self.current_month, -(self.amc_months - 1))
            start_date = date(start_year, start_month, 1)
            end_day = monthrange(self.current_year, self.current_month)[1]
            end_date = date(self.current_year, self.current_month, end_day)
            cur.execute("""
                SELECT code, SUM(COALESCE(Qty_Out,0))
                FROM stock_transactions
                WHERE Out_Type IS NOT NULL
                  AND LOWER(Out_Type)=LOWER('Out MSF')
                  AND Date >= ? AND Date <= ?
                GROUP BY code
            """,(start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))
            for code, total_out in cur.fetchall():
                if code:
                    mapping[code] = (total_out or 0)/self.amc_months
        except sqlite3.Error:
            pass
        finally:
            cur.close()
        return mapping

    def _derive_code(self, row):
        # Priority: item > module > kit
        for k in ("item","module","kit"):
            v = row.get(k)
            if v and str(v).lower() != "none":
                return v
        uid = row.get("unique_id") or ""
        parts = uid.split("/")
        if len(parts) >= 4:
            for idx in (3,2,1):
                if idx < len(parts):
                    vv = parts[idx]
                    if vv and vv.lower() != "none":
                        return vv
        return None

    def _parse_iso(self, iso):
        if not iso:
            return None
        try:
            y,m,d = map(int, iso.split("-"))
            return date(y,m,d)
        except Exception:
            return None

    def compute(self):
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed")
        try:
            stock_rows = self._load_stock(conn)
            amc_map = self._amc_map(conn)
            current_ym = (self.current_year, self.current_month)
            horizon_end = add_months(self.current_year, self.current_month, self.expiry_period - 1)
            horizon_end_ym = horizon_end

            results = []
            for r in stock_rows:
                final_qty = r.get("final_qty")
                if final_qty is None or final_qty <= 0:
                    continue  # only positive quantities

                code = self._derive_code(r)
                if not code:
                    continue

                mgmt = (r.get("management_mode") or "").strip()
                if self.management_mode.lower() != "all":
                    if mgmt.lower() != self.management_mode.lower():
                        continue

                scen_name = (r.get("scenario") or "").strip()
                if self.scenario_filter.lower() != "all":
                    if scen_name.lower() != self.scenario_filter.lower():
                        continue

                kit_num = (r.get("kit_number") or "")
                if self.kit_filter.lower() != "all":
                    if kit_num.lower() != self.kit_filter.lower():
                        continue

                module_num = (r.get("module_number") or "")
                if self.module_filter.lower() != "all":
                    if module_num.lower() != self.module_filter.lower():
                        continue

                code_type = detect_type(code, "")
                if self.type_filter.lower() != "all":
                    if code_type.lower() != self.type_filter.lower():
                        continue

                if self.item_search:
                    if code_type.lower() != "item":
                        continue
                    desc_search = get_item_description(code)
                    if (self.item_search.lower() not in code.lower()
                        and self.item_search.lower() not in desc_search.lower()):
                        continue

                expiry_iso = r.get("exp_date")
                exp_dt = self._parse_iso(expiry_iso)
                months_to_expiry = None
                if exp_dt:
                    months_to_expiry = (exp_dt.year - self.current_year)*12 + (exp_dt.month - self.current_month)
                    # Horizon filtering (keep expired or within horizon)
                    if months_to_expiry > (self.expiry_period - 1):
                        continue

                desc = get_item_description(code)
                amc = round(amc_map.get(code, 0.0), 2)
                discrepancy = r.get("discrepancy") or 0
                comments = r.get("comments") or ""

                results.append({
                    "code": code,
                    "description": desc,
                    "amc": amc,
                    "quantity": final_qty,
                    "expiry_date": expiry_iso or "",
                    "comments": comments,
                    "scenario": scen_name,
                    "kit_number": kit_num,
                    "module_number": module_num,
                    "management_mode": mgmt,
                    "type": code_type,
                    "qty_in": r.get("qty_in") or 0,
                    "qty_out": r.get("qty_out") or 0,
                    "discrepancy": discrepancy,
                    "months_to_expiry": months_to_expiry,
                    "_expired_flag": (exp_dt is not None and exp_dt < self.today)
                })

            # Sort: expired first (optional) then months_to_expiry then code
            def sort_key(x):
                exp_sort = -1 if x["_expired_flag"] else 0
                mte = x.get("months_to_expiry")
                mte_sort = 9999 if mte is None else mte
                return (exp_sort, mte_sort, x["code"])
            results.sort(key=sort_key)
            return results
        finally:
            try:
                conn.close()
            except Exception:
                pass

# ---------------- UI ----------------
class StockAvailability(tk.Frame):
    SIMPLE_COLUMNS = ["code","description","quantity","expiry_date","amc","comments"]
    DETAILED_EXTRA = ["scenario","kit_number","module_number","management_mode",
                      "type","qty_in","qty_out","discrepancy","months_to_expiry"]
    COL_LABELS = {
        "code": "Code",
        "description": "Description",
        "amc": "AMC",
        "quantity": "Quantity",
        "expiry_date": "Expiry Date",
        "comments": "Comments",
        "scenario": "Scenario",
        "kit_number": "Kit Number",
        "module_number": "Module Number",
        "management_mode": "Management Mode",
        "type": "Type",
        "qty_in": "Qty IN",
        "qty_out": "Qty OUT",
        "discrepancy": "Discrepancy",
        "months_to_expiry": "Months To Expiry"
    }

    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_MAIN)
        self.app = app
        self.tree = None
        self.data_rows = []
        self.simple_mode = True
        self._build_ui()
        self.refresh()

    def t(self, key, fallback=None, **kwargs):
        return lang.t(f"stock_availability.{key}", fallback=fallback if fallback else key, **kwargs)

    # ---------- UI Build ----------
    def _build_ui(self):
        tk.Label(
            self,
            text=self.t("title","Stock Availability"),
            font=("Helvetica", 20, "bold"),
            bg=BG_MAIN,
            fg=COLOR_PRIMARY,
            anchor="w"
        ).pack(fill="x", padx=12, pady=(12, 4))

        filt = tk.Frame(self, bg=BG_MAIN)
        filt.pack(fill="x", padx=12, pady=(0,10))

        # Row 1
        r1 = tk.Frame(filt, bg=BG_MAIN)
        r1.pack(fill="x", pady=2)

        tk.Label(r1, text=self.t("management_mode","Management Mode"), bg=BG_MAIN).grid(row=0, column=0, sticky="w", padx=(0,4))
        self.mgmt_mode_var = tk.StringVar(value="All")
        ttk.Combobox(r1, textvariable=self.mgmt_mode_var, state="readonly", width=14,
                     values=["All","on-shelf","in-box"]).grid(row=0, column=1, padx=(0,14))

        tk.Label(r1, text=self.t("scenario","Scenario"), bg=BG_MAIN).grid(row=0, column=2, sticky="w", padx=(0,4))
        self.scenario_var = tk.StringVar(value="All")
        self.scenario_cb = ttk.Combobox(r1, textvariable=self.scenario_var, state="readonly", width=20)
        self.scenario_cb.grid(row=0, column=3, padx=(0,14))

        tk.Label(r1, text=self.t("kit_number","Kit"), bg=BG_MAIN).grid(row=0, column=4, sticky="w", padx=(0,4))
        self.kit_filter_var = tk.StringVar(value="All")
        self.kit_cb = ttk.Combobox(r1, textvariable=self.kit_filter_var, state="readonly", width=16, values=["All"])
        self.kit_cb.grid(row=0, column=5, padx=(0,14))

        tk.Label(r1, text=self.t("module_number","Module"), bg=BG_MAIN).grid(row=0, column=6, sticky="w", padx=(0,4))
        self.module_filter_var = tk.StringVar(value="All")
        self.module_cb = ttk.Combobox(r1, textvariable=self.module_filter_var, state="readonly", width=16, values=["All"])
        self.module_cb.grid(row=0, column=7, padx=(0,14))

        # Row 2
        r2 = tk.Frame(filt, bg=BG_MAIN)
        r2.pack(fill="x", pady=2)

        tk.Label(r2, text=self.t("item_search","Item Search"), bg=BG_MAIN).grid(row=0, column=0, sticky="w", padx=(0,4))
        self.item_search_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.item_search_var, width=22).grid(row=0, column=1, padx=(0,14))

        tk.Label(r2, text=self.t("type","Type"), bg=BG_MAIN).grid(row=0, column=2, sticky="w", padx=(0,4))
        self.type_var = tk.StringVar(value="All")
        ttk.Combobox(r2, textvariable=self.type_var, state="readonly", width=12,
                     values=["All","Kit","Module","Item"]).grid(row=0, column=3, padx=(0,14))

        tk.Label(r2, text=self.t("expiry_period","Expiry Period (Months)"), bg=BG_MAIN).grid(row=0, column=4, sticky="w", padx=(0,4))
        self.expiry_period_var = tk.StringVar(value="12")
        tk.Entry(r2, textvariable=self.expiry_period_var, width=6,
                 validate="key",
                 validatecommand=(self.register(self._val_1_99), "%P")).grid(row=0, column=5, padx=(0,14))

        tk.Label(r2, text=self.t("amc_months","AMC Months"), bg=BG_MAIN).grid(row=0, column=6, sticky="w", padx=(0,4))
        self.amc_months_var = tk.StringVar(value="6")
        tk.Entry(r2, textvariable=self.amc_months_var, width=6,
                 validate="key",
                 validatecommand=(self.register(self._val_1_99), "%P")).grid(row=0, column=7, padx=(0,14))

        # Buttons
        btn_row = tk.Frame(filt, bg=BG_MAIN)
        btn_row.pack(fill="x", pady=(6,4))
        tk.Button(btn_row, text=self.t("refresh","Refresh"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.refresh).pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("clear","Clear"),
                  bg=BTN_CLEAR, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.clear_filters).pack(side="left", padx=(0,6))
        self.toggle_btn = tk.Button(btn_row, text=self.t("toggle_detailed","Detailed"),
                                    bg=BTN_TOGGLE, fg="#FFFFFF", relief="flat",
                                    padx=14, pady=6, command=self.toggle_mode)
        self.toggle_btn.pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("export","Export"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.export_excel).pack(side="left", padx=(0,6))

        self.status_var = tk.StringVar(value=self.t("ready","Ready"))
        tk.Label(self, textvariable=self.status_var, anchor="w",
                 bg=BG_MAIN, fg=COLOR_PRIMARY, relief="sunken").pack(fill="x", padx=12, pady=(0,8))

        # Tree
        outer = tk.Frame(self, bg=COLOR_BORDER, bd=1, relief="solid")
        outer.pack(fill="both", expand=True, padx=12, pady=(0,12))
        self.tree = ttk.Treeview(outer, columns=(), show="headings", height=24)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vsb.set)

        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("Treeview",
                        background=BG_PANEL,
                        fieldbackground=BG_PANEL,
                        foreground=COLOR_PRIMARY,
                        rowheight=24,
                        font=("Helvetica",10))
        style.configure("Treeview.Heading",
                        background="#E5E8EB",
                        foreground=COLOR_PRIMARY,
                        font=("Helvetica",11,"bold"))
        self.tree.tag_configure("norm", background=ROW_NORM_COLOR)
        self.tree.tag_configure("alt", background=ROW_ALT_COLOR)
        self.tree.tag_configure("kitrow", background="#228B22", foreground="#FFFFFF")
        self.tree.tag_configure("modrow", background="#ADD8E6")
        self.tree.tag_configure("expired_light", background=EXPIRED_LIGHT_COLOR)

    # ---------- Validation ----------
    def _val_1_99(self, P):
        if P == "":
            return True
        return P.isdigit() and 1 <= int(P) <= 99

    # ---------- Scenario / dropdown loading ----------
    def _scenario_map(self):
        conn = connect_db()
        mapping = {}
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("SELECT scenario_id, name FROM scenarios")
                for sid, nm in cur.fetchall():
                    mapping[str(sid)] = nm
            except sqlite3.Error:
                pass
            finally:
                cur.close(); conn.close()
        return mapping

    def load_scenarios(self):
        conn = connect_db()
        names = []
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("SELECT name FROM scenarios ORDER BY name")
                names = [r[0] for r in cur.fetchall()]
            except sqlite3.Error:
                names = []
            finally:
                cur.close(); conn.close()
        self.scenario_cb['values'] = ["All"] + names
        if self.scenario_var.get() not in self.scenario_cb['values']:
            self.scenario_var.set("All")

    def populate_kit_module_lists(self):
        conn = connect_db()
        kits, modules = set(), set()
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("PRAGMA table_info(stock_data)")
                cols = [r[1].lower() for r in cur.fetchall()]
                if "kit_number" in cols:
                    cur.execute("""SELECT DISTINCT kit_number FROM stock_data
                                   WHERE kit_number IS NOT NULL AND kit_number!='None'
                                   ORDER BY kit_number""")
                    for (k,) in cur.fetchall():
                        if k: kits.add(str(k))
                if "module_number" in cols:
                    cur.execute("""SELECT DISTINCT module_number FROM stock_data
                                   WHERE module_number IS NOT NULL AND module_number!='None'
                                   ORDER BY module_number""")
                    for (m,) in cur.fetchall():
                        if m: modules.add(str(m))
            except sqlite3.Error:
                pass
            finally:
                cur.close(); conn.close()
        kit_vals = ["All"] + sorted(kits, key=lambda x: (len(x), x))
        mod_vals = ["All"] + sorted(modules, key=lambda x: (len(x), x))
        self.kit_cb['values'] = kit_vals
        self.module_cb['values'] = mod_vals
        if self.kit_filter_var.get() not in kit_vals:
            self.kit_filter_var.set("All")
        if self.module_filter_var.get() not in mod_vals:
            self.module_filter_var.set("All")

    # ---------- Refresh ----------
    def refresh(self):
        self.load_scenarios()
        self.populate_kit_module_lists()
        try:
            expiry_period = int(self.expiry_period_var.get() or "12")
        except ValueError:
            expiry_period = 12
            self.expiry_period_var.set("12")
        try:
            amc_months = int(self.amc_months_var.get() or "6")
        except ValueError:
            amc_months = 6
            self.amc_months_var.set("6")

        calc = StockAvailabilityCalculator(
            scenario_name_map=self._scenario_map(),
            management_mode=self.mgmt_mode_var.get(),
            scenario_filter=self.scenario_var.get(),
            kit_filter=self.kit_filter_var.get(),
            module_filter=self.module_filter_var.get(),
            item_search=self.item_search_var.get(),
            type_filter=self.type_var.get(),
            expiry_period=expiry_period,
            amc_months=amc_months
        )
        try:
            self.data_rows = calc.compute()
        except ValueError as e:
            custom_popup(self, self.t("error","Error"), str(e), "error")
            return
        self._populate_tree()
        self.status_var.set(self.t("loaded","Loaded {n} rows").format(n=len(self.data_rows)))

    # ---------- Current Columns ----------
    def current_columns(self):
        if self.simple_mode:
            return self.SIMPLE_COLUMNS
        return self.SIMPLE_COLUMNS + self.DETAILED_EXTRA

    # ---------- Populate Tree ----------
    def _populate_tree(self):
        cols = self.current_columns()
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            label = self.t(f"columns.{c}", self.COL_LABELS.get(c, c.title()))
            width = 140
            if c == "description": width = 340
            elif c == "comments": width = 220
            elif c == "code": width = 160
            elif c == "quantity": width = 100
            elif c == "amc": width = 80
            elif c == "expiry_date": width = 160
            elif c in ("scenario","kit_number","module_number","management_mode","type"):
                width = 120
            elif c in ("qty_in","qty_out","discrepancy","months_to_expiry"):
                width = 110
            self.tree.heading(c, text=label)
            self.tree.column(c, width=width, anchor="w")

        for idx, row in enumerate(self.data_rows):
            # Format expiry date
            disp_exp = format_expiry_display(row.get("expiry_date"))
            row_display = dict(row)
            row_display["expiry_date"] = disp_exp
            values = [row_display.get(c, "") for c in cols]

            # Tagging logic
            tags = []
            if row.get("_expired_flag"):
                tags.append("expired_light")
            dtype = (row.get("type") or "").upper()
            if dtype == "KIT":
                tags.append("kitrow")
            elif dtype == "MODULE":
                tags.append("modrow")
            if not tags:
                tags.append("alt" if idx % 2 else "norm")

            self.tree.insert("", "end", values=values, tags=tuple(tags))

    # ---------- Toggle ----------
    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        self.toggle_btn.config(text=self.t("toggle_simple","Simple") if not self.simple_mode else self.t("toggle_detailed","Detailed"))
        self._populate_tree()

    # ---------- Clear ----------
    def clear_filters(self):
        self.mgmt_mode_var.set("All")
        self.scenario_var.set("All")
        self.kit_filter_var.set("All")
        self.module_filter_var.set("All")
        self.item_search_var.set("")
        self.type_var.set("All")
        self.expiry_period_var.set("12")
        self.amc_months_var.set("6")
        self.simple_mode = True
        self.toggle_btn.config(text=self.t("toggle_detailed","Detailed"))
        self.refresh()

    # ---------- Export ----------
    def export_excel(self):
        if not self.data_rows:
            custom_popup(self, self.t("no_data","No Data"),
                         self.t("nothing_export","Nothing to export."), "warning")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files","*.xlsx")],
            title=self.t("export_dialog","Save Stock Availability"),
            initialfile=f"Stock_Availability_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "StockAvailability"
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([self.t("generated","Generated"), now_str])
            ws.append([self.t("filters","Filters Used")])
            ws.append(["Management Mode", self.mgmt_mode_var.get(),
                       "Scenario", self.scenario_var.get(),
                       "Type", self.type_var.get()])
            ws.append(["Expiry Period", self.expiry_period_var.get(),
                       "AMC Months", self.amc_months_var.get(),
                       "Mode", "Simple" if self.simple_mode else "Detailed"])
            ws.append(["Kit", self.kit_filter_var.get(),
                       "Module", self.module_filter_var.get(),
                       "Item Search", self.item_search_var.get()])
            ws.append([])

            cols = self.current_columns()
            header = [self.COL_LABELS.get(c, c.title()) for c in cols]
            ws.append(header)

            kit_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            module_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            expired_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")

            for r in self.data_rows:
                # Prepare row values with formatted expiry
                row_display = dict(r)
                row_display["expiry_date"] = format_expiry_display(r.get("expiry_date"))
                row_vals = [row_display.get(c, "") for c in cols]
                ws.append(row_vals)
                dtyp = (r.get("type") or "").upper()
                row_cells = ws[ws.max_row]
                if r.get("_expired_flag"):
                    for c in row_cells:
                        c.fill = expired_fill
                if dtyp == "KIT":
                    for c in row_cells:
                        c.fill = kit_fill
                elif dtyp == "MODULE":
                    for c in row_cells:
                        c.fill = module_fill

            # Auto width
            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            wb.save(file_path)
            custom_popup(self, self.t("success","Success"),
                         self.t("export_success","Export completed: {fp}").format(fp=file_path),
                         "info")
        except Exception as e:
            custom_popup(self, self.t("error","Error"),
                         self.t("export_failed","Export failed: {err}").format(err=str(e)),
                         "error")

# Backward compatibility alias (if desired)
class StockAvailabilityReport(StockAvailability):
    pass

__all__ = ["StockAvailability","StockAvailabilityReport"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Availability v1.1")
    class Dummy: pass
    d = Dummy()
    StockAvailability(root, d).pack(fill="both", expand=True)
    root.geometry("1400x760")
    root.mainloop()