# ============================= out_kit.py (Part 1/6) =============================
# Break Kit / Module
# Derived and extended from dispatch_kit.py to implement "Break" logic:
#  - Movement primary mode renamed "Break Kit" (internal key keeps dispatch_Kit for reuse of logic patterns)
#  - OUT Type fixed: "Internal move from in-box items"
#  - Dual logging: for every OUT transaction, a mirror IN transaction is also recorded (Qty_IN + IN_Type)
#  - Hidden columns: line_id, qty_in_hidden (mirrors user-entered qty_to_issue)
#  - Tree ordering restored using treecode via kit_items index
#  - Color highlighting: distinct styles for headers and Kit/Module data rows
#  - Quantities for kits/modules propagate to sub-levels; star/★ marking editable cells
#  - Document number generation with "BRK" abbreviation
#
# ================================================================================
import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import logging
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from popup_utils import custom_popup, custom_askyesno, custom_dialog
import os


from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang
from theme_config import AppTheme, apply_global_style, get_button_style, configure_tree_tags, create_styled_button
from popup_utils import custom_popup, custom_askyesno, custom_dialog

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")

# Canonical English value (used for database storage)
OUT_TYPE_CANONICAL = "Internal move from in-box items"

def fetch_project_details():
    conn = connect_db()
    if conn is None:
        logging.error("[BREAK] DB connection failed (fetch_project_details)")
        return "Unknown Project", "Unknown Code"
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        return (row[0] if row and row[0] else "Unknown Project",
                row[1] if row and row[1] else "Unknown Code")
    except sqlite3.Error as e:
        logging.error(f"[BREAK] fetch_project_details error: {e}")
        return "Unknown Project", "Unknown Code"
    finally:
        cur.close()
        conn.close()

def configure_db_pragmas():
    conn = connect_db()
    if conn is None:
        return
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA busy_timeout=5000;")
        conn.commit()
    except Exception:
        pass
    finally:
        try: conn.close()
        except: pass


class StockOutKit(tk.Frame):
    """
    Break Kit / Module logic with dual transaction logging.
    Retains dispatch-like structure but adapted for "break" internal movement.
    """

    def __init__(self, parent, app, role="supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        self.role = role.lower()

        try:
            configure_db_pragmas()
        except Exception:
            pass

        # UI variable references
        self.scenario_var = tk.StringVar()
        self.mode_var = tk.StringVar()
        self.Kit_var = tk.StringVar()
        self.Kit_number_var = tk.StringVar()
        self.module_var = tk.StringVar()
        self.module_number_var = tk.StringVar()
        self.out_type_var = tk.StringVar(value="")  # Will be set after lang is available
        self.search_var = tk.StringVar()
        self.status_var = tk.StringVar(value=lang.t("break_kit.ready", "Ready"))

        # Widgets placeholders
        self.scenario_cb = None
        self.mode_cb = None
        self.Kit_cb = None
        self.Kit_number_cb = None
        self.module_cb = None
        self.module_number_cb = None
        self.tree = None

        # Data state
        self.scenario_map = self.fetch_scenario_map()
        self.selected_scenario_id = None
        self.selected_scenario_name = None
        self.mode_definitions = []
        self.mode_label_to_key = {}
        self.full_items = []      # enriched rows (without headers)
        self.row_data = {}        # iid -> metadata
        self.search_min_chars = 2
        self.editing_cell = None
        self._item_index_cache = {}

        if self.parent and self.parent.winfo_exists():
            self.pack(fill="both", expand=True)
            self.after(50, self.render_ui)

    # ---------------------------------------------------------
    # Localization Helpers
    # ---------------------------------------------------------
    def _all_label(self):
        """Return localized 'All' label for dropdowns."""
        return lang.t("out_kit.all", "All")
    
    def _norm_all(self, val):
        """Normalize 'All' variants to English 'All' for internal use."""
        all_lbl = self._all_label()
        return "All" if (val is None or val == "" or val == all_lbl) else val
    
    def _extract_code_from_display(self, display_string: str) -> str:
        """
        Extract code from "CODE - Description" format.
        Handles prefixes like "● CODE - Description".
        
        Args:
            display_string: Either "CODE" or "CODE - Description"
        
        Returns:
            Just the code part, or None if empty/invalid
        """
        if not display_string:
            return None
        
        display_string = display_string.strip()
        
        if display_string == "-----":
            return None
        
        # Strip visual indicators
        prefixes = ["●", "■", "◆", "►", "[S]", "[Primary]", "[Standalone]"]
        for prefix in prefixes:
            if display_string.startswith(prefix):
                display_string = display_string[len(prefix):].strip()
                break
        
        # Extract code from "CODE - Description"
        if " - " in display_string:
            code = display_string.split(" - ", 1)[0].strip()
            return code if code else None
        
        return display_string

    def _canon_movement_type(self, display_label: str) -> str:
        """
        Convert localized movement type to canonical English for database storage.
        
        Args:
            display_label: Localized label from dropdown
        
        Returns:
            Canonical English movement type
        """
        internal_key = self.mode_label_to_key.get(display_label)
        
        if not internal_key:
            logging.warning(f"[OUT_KIT] Unknown movement type label: {display_label}")
            return display_label
        
        # Map internal keys to canonical English
        canon_map = {
            "out_kit": "Out Kit",
            "out_standalone": "Out standalone items",
            "out_module_scenario": "Out module from scenario",
            "out_module_kit": "Out module from Kit",
            "out_items_kit": "Out items from Kit",
            "out_items_module": "Out items from module"
        }
        
        canonical = canon_map.get(internal_key, internal_key)
        logging.debug(f"[OUT_KIT] Movement type: '{display_label}' → '{canonical}'")
        return canonical

    def _display_for_movement_type(self, canonical_value: str) -> str:
        """
        Convert canonical English movement type to localized display label.
        
        Args:
            canonical_value: English movement type from database
        
        Returns:
            Localized display label
        """
        reverse_canon_map = {
            "Out Kit": "out_kit",
            "Out standalone items": "out_standalone",
            "Out module from scenario": "out_module_scenario",
            "Out module from Kit": "out_module_kit",
            "Out items from Kit": "out_items_kit",
            "Out items from module": "out_items_module"
        }
        
        internal_key = reverse_canon_map.get(canonical_value, "out_kit")
        
        for label, key in self.mode_label_to_key.items():
            if key == internal_key:
                return label
        
        return canonical_value

    # -------------------- Scenario / Modes --------------------
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None: return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            rows = cur.fetchall()
            return {str(r['scenario_id']): r['name'] for r in rows}
        except sqlite3.Error as e:
            logging.error(f"[BREAK] fetch_scenario_map error: {e}")
            return {}
        finally:
            cur.close()
            conn.close()

    def build_mode_definitions(self):
        """
        Build localized movement type definitions.
        All labels are translated, but internal keys remain English.
        """
        scenario = self.selected_scenario_name or ""
        self.mode_definitions = [
            ("out_kit", lang.t("out_kit.mode_out_kit", "Out Kit")),
            ("out_standalone", lang.t("out_kit.mode_out_standalone", "Out standalone item/s from {scenario}", scenario=scenario)),
            ("out_module_scenario", lang.t("out_kit.mode_out_module_scenario", "Out module from {scenario}", scenario=scenario)),
            ("out_module_kit", lang.t("out_kit.mode_out_module_kit", "Out module from a Kit")),
            ("out_items_kit", lang.t("out_kit.mode_out_items_kit", "Out items from a Kit")),
            ("out_items_module", lang.t("out_kit.mode_out_items_module", "Out items from a module")),
        ]
        self.mode_label_to_key = {label: key for key, label in self.mode_definitions}
        logging.info(f"[OUT_KIT] Built {len(self.mode_definitions)} mode definitions")

    def current_mode_key(self):
        return self.mode_label_to_key.get(self.mode_var.get())

    # -------------------- unique_id Management Helpers --------------------
    @staticmethod
    def count_slashes(unique_id: str) -> int:
        """Count slashes in unique_id."""
        return unique_id.count("/") if unique_id else 0

    @staticmethod
    def is_inbox_item(unique_id: str) -> bool:
        """Check if unique_id is in-box format (> 6 slashes)."""
        return StockOutKit.count_slashes(unique_id) > 6

    @staticmethod
    def is_onshelf_item(unique_id: str) -> bool:
        """Check if unique_id is on-shelf format (≤ 6 slashes)."""
        slash_count = StockOutKit.count_slashes(unique_id)
        return 0 < slash_count <= 6

    @staticmethod
    def convert_inbox_to_onshelf(unique_id: str) -> str:
        """
        Convert in-box unique_id to on-shelf format.
        Removes kit_number and module_number (last 2 parts).
        
        Format:
        - in-box:   scenario/kit/module/item/std_qty/exp_date/kit_number/module_number (>6 slashes)
        - on-shelf: scenario/kit/module/item/std_qty/exp_date (≤6 slashes)
        
        Example:
            IN:  1/KMEDKCHO1--/KMEDMCHO01-/MED001/10/2025-12-31/CHOL-001/M-001
            OUT: 1/KMEDKCHO1--/KMEDMCHO01-/MED001/10/2025-12-31
        """
        if not unique_id:
            return unique_id
        
        parts = unique_id.split("/")
        
        # If already on-shelf or invalid, return as-is
        if len(parts) <= 6:
            logging.debug(f"[CONVERT] Already on-shelf or invalid: {unique_id}")
            return unique_id
        
        # Remove last 2 parts (kit_number, module_number)
        onshelf_parts = parts[:6]
        onshelf_id = "/".join(onshelf_parts)
        
        logging.debug(f"[CONVERT] in-box: {unique_id} → on-shelf: {onshelf_id}")
        return onshelf_id

    def check_if_adopted_expiry(self, unique_id_inbox):
        """
        Check if an in-box item has an adopted expiry date.
        
        Args:
            unique_id_inbox: The in-box unique_id
        
        Returns:
            tuple: (is_adopted: bool, comment: str)
        """
        conn = connect_db()
        if conn is None:
            return False, ""
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            cur.execute("""
                SELECT comments FROM stock_data WHERE unique_id = ?
            """, (unique_id_inbox,))
            
            row = cur.fetchone()
            
            if not row or not row['comments']:
                return False, ""
            
            comment = str(row['comments']).lower()
            
            # Check for adoption indicators
            adoption_keywords = ['adopted', 'candidate', 'inherited', 'propagated']
            is_adopted = any(keyword in comment for keyword in adoption_keywords)
            
            logging.debug(f"[ADOPTED_CHECK] {unique_id_inbox}: adopted={is_adopted}, comment={row['comments'][:50] if row['comments'] else 'None'}")
            
            return is_adopted, row['comments'] if is_adopted else ""
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] check_if_adopted_expiry error: {e}")
            return False, ""
        finally:
            cur.close()
            conn.close()


    def _validate_date_format(self, date_str):
        """
        Validate date format (YYYY-MM-DD).
        
        Args:
            date_str: Date string to validate
        
        Returns:
            bool: True if valid format
        """
        if not date_str:
            return True  # Empty is allowed
        
        import re
        
        # Check basic format
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return False
        
        # Try parsing as actual date
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    
    def _is_date_in_future(self, date_str):
        """
        Check if date is in the future (or today).
        
        Args:
            date_str: Date string in YYYY-MM-DD format
        
        Returns:
            bool: True if date is today or in future
        """
        if not date_str:
            return True  # Empty is allowed
        
        try:
            exp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            today = datetime.today().date()
            return exp_date >= today
        except ValueError:
            return False



    def _parse_flexible_date(self, date_input):
        """
        Parse flexible date formats and convert to YYYY-MM-DD.
        
        Supported formats:
        - YYYY-MM-DD (standard)
        - YYYY-MM (month only → last day of month)
        - YYYY (year only → December 31)
        - MM/YYYY (e.g., 12/2029 → 2029-12-31)
        - DD-MM-YYYY, DD/MM/YYYY (day first, e.g., 31-12-2029)
        - MM-DD-YYYY, MM/DD/YYYY (month first, US format)
        
        Args:
            date_input: User's date string
        
        Returns:
            tuple: (success: bool, parsed_date: str or None, error_msg: str)
        """
        if not date_input:
            return True, None, ""
        
        import re
        from calendar import monthrange
        
        date_input = date_input.strip()
        
        # Try various formats in priority order
        patterns = [
            # 1. YYYY-MM-DD (standard ISO format)
            (r'^(\d{4})-(\d{2})-(\d{2})$', 'YYYY-MM-DD', 
             lambda m: f"{m[0]}-{m[1]}-{m[2]}"),
            
            # 2. YYYY-MM (month only → last day)
            (r'^(\d{4})-(\d{2})$', 'YYYY-MM', 
             lambda m: self._last_day_of_month(int(m[0]), int(m[1]))),
            
            # 3. YYYY (year only → December 31)
            (r'^(\d{4})$', 'YYYY', 
             lambda m: f"{m[0]}-12-31"),
            
            # 4. MM/YYYY (e.g., 12/2029 → 2029-12-31)
            (r'^(\d{1,2})/(\d{4})$', 'MM/YYYY', 
             lambda m: self._last_day_of_month(int(m[1]), int(m[0]))),
            
            # 5. DD-MM-YYYY (day first, common in Europe)
            (r'^(\d{2})-(\d{2})-(\d{4})$', 'DD-MM-YYYY', 
             lambda m: f"{m[2]}-{m[1]}-{m[0]}"),
            
            # 6. DD/MM/YYYY (day first with slashes)
            # Disambiguate by checking if day > 12
            (r'^(\d{1,2})/(\d{1,2})/(\d{4})$', 'DD/MM/YYYY or MM/DD/YYYY', 
             self._parse_ambiguous_slash_date),
        ]
        
        for pattern, format_name, converter in patterns:
            match = re.match(pattern, date_input)
            if match:
                try:
                    groups = match.groups()
                    result = converter(groups)
                    
                    if result is None:
                        continue  # Ambiguous pattern, try next
                    
                    # Validate the result is a real date
                    datetime.strptime(result, '%Y-%m-%d')
                    
                    logging.debug(f"[PARSE_DATE] '{date_input}' ({format_name}) → '{result}'")
                    return True, result, ""
                    
                except (ValueError, IndexError) as e:
                    logging.debug(f"[PARSE_DATE] Pattern {format_name} matched but failed validation: {e}")
                    continue
        
        # No pattern matched
        error_msg = (
            f"Cannot parse date: '{date_input}'\n\n"
            f"Supported formats:\n"
            f"  • YYYY-MM-DD    (e.g., 2027-12-31)\n"
            f"  • YYYY-MM       (e.g., 2027-12 → 2027-12-31)\n"
            f"  • YYYY          (e.g., 2027 → 2027-12-31)\n"
            f"  • MM/YYYY       (e.g., 12/2029 → 2029-12-31)\n"
            f"  • DD-MM-YYYY    (e.g., 31-12-2027)\n"
            f"  • DD/MM/YYYY    (e.g., 31/12/2027)\n"
            f"  • MM/DD/YYYY    (e.g., 12/31/2027)"
        )
        return False, None, error_msg
    
    def _parse_ambiguous_slash_date(self, groups):
        """
        Parse DD/MM/YYYY vs MM/DD/YYYY by checking if first number > 12.
        
        Args:
            groups: (first, second, year) from regex
        
        Returns:
            str: YYYY-MM-DD or None if ambiguous
        """
        first = int(groups[0])
        second = int(groups[1])
        year = groups[2]
        
        # If first > 12, must be DD/MM/YYYY
        if first > 12:
            return f"{year}-{second:02d}-{first:02d}"
        
        # If second > 12, must be MM/DD/YYYY
        if second > 12:
            return f"{year}-{first:02d}-{second:02d}"
        
        # Both ≤ 12: ambiguous - default to DD/MM/YYYY (European)
        # You could also add a preference setting here
        logging.debug(f"[PARSE_DATE] Ambiguous date {first}/{second}/{year}, assuming DD/MM/YYYY")
        return f"{year}-{second:02d}-{first:02d}"
    
    def _last_day_of_month(self, year, month):
        """Get last day of month in YYYY-MM-DD format."""
        from calendar import monthrange
        
        if not (1 <= month <= 12):
            raise ValueError(f"Invalid month: {month}")
        
        last_day = monthrange(year, month)[1]
        return f"{year:04d}-{month:02d}-{last_day:02d}"






    def load_scenarios(self):
        vals = [f"{sid} - {nm}" for sid, nm in self.scenario_map.items()]
        self.scenario_cb['values'] = vals
        if vals:
            self.scenario_cb.current(0)
            self.on_scenario_selected()

    def on_scenario_selected(self, event=None):
        sel = self.scenario_var.get()
        if not sel:
            self.selected_scenario_id = None
            self.selected_scenario_name = None
            return
        self.selected_scenario_id = sel.split(" - ")[0]
        self.selected_scenario_name = sel.split(" - ", 1)[1] if " - " in sel else ""
        self.build_mode_definitions()
        self.mode_cb['values'] = [lbl for _, lbl in self.mode_definitions]
        if self.mode_definitions:
            self.mode_var.set(self.mode_definitions[0][1])
        self.on_mode_changed()

    def on_mode_changed(self, event=None):
        """
        Called when movement type changes.
        Enables/disables appropriate selectors based on mode.
        """
        mode_key = self.current_mode_key()
        
        logging.debug(f"[OUT_KIT] Mode changed to: {mode_key}")

        # Disable all selectors initially
        for cb in [self.Kit_cb, self.Kit_number_cb, self.module_cb, self.module_number_cb]:
            cb.config(state="disabled")

        # Clear selections
        self.Kit_var.set("")
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        
        # Clear dropdown values
        self.Kit_cb['values'] = []
        self.Kit_number_cb['values'] = []
        self.module_cb['values'] = []
        self.module_number_cb['values'] = []

        # Clear table
        self.clear_table_only()

        if not self.selected_scenario_id:
            return

        # ===== Mode-specific logic =====
        
        if mode_key == "out_kit":
            # Show primary kits only
            self.Kit_cb.config(state="readonly")
            self.Kit_cb['values'] = self.fetch_Kits(self.selected_scenario_id)
            logging.debug(f"[OUT_KIT] out_kit: Populated {len(self.Kit_cb['values'])} primary kits")
        
        elif mode_key == "out_standalone":
            # Populate with standalone items
            self.populate_standalone_items()
            logging.debug(f"[OUT_KIT] out_standalone: Populated standalone items")
        
        elif mode_key == "out_module_scenario":
            # Show primary modules
            self.module_cb.config(state="readonly")
            modules = self.fetch_all_modules(self.selected_scenario_id)
            self.module_cb['values'] = modules
            logging.debug(f"[OUT_KIT] out_module_scenario: Populated {len(modules)} primary modules")
        
        elif mode_key == "out_module_kit":
            # Enable kit selector
            self.Kit_cb.config(state="readonly")
            self.Kit_cb['values'] = self.fetch_Kits(self.selected_scenario_id)
            logging.debug(f"[OUT_KIT] out_module_kit: Populated {len(self.Kit_cb['values'])} primary kits")
        
        elif mode_key == "out_items_kit":
            # Enable kit selector
            self.Kit_cb.config(state="readonly")
            self.Kit_cb['values'] = self.fetch_Kits(self.selected_scenario_id)
            logging.debug(f"[OUT_KIT] out_items_kit: Populated {len(self.Kit_cb['values'])} primary kits")
        
        elif mode_key == "out_items_module":
            # Enable both kit and module selectors
            self.Kit_cb.config(state="readonly")
            self.Kit_cb['values'] = self.fetch_Kits(self.selected_scenario_id)
            
            self.module_cb.config(state="readonly")
            self.module_cb['values'] = self.fetch_all_modules(self.selected_scenario_id)
            
            logging.debug(f"[OUT_KIT] out_items_module: Kits and modules both enabled")


# ============================= out_kit.py (Part 2/6) =============================
    # -------------------- Index / Treecode --------------------
    def ensure_item_index(self, scenario_id):
        if (self._item_index_cache.get("scenario_id") == scenario_id and
            self._item_index_cache.get("flat_map")):
            return
        self._item_index_cache = {
            "scenario_id": scenario_id,
            "flat_map": {},
            "triple_map": {}
        }
        conn = connect_db()
        if conn is None: return
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT code, kit, module, item, treecode, level
                  FROM kit_items
                 WHERE scenario_id=?
            """,(scenario_id,))
            for r in cur.fetchall():
                code = (r["code"] or "").strip()
                kit = (r["kit"] or "").strip()
                module = (r["module"] or "").strip()
                item = (r["item"] or "").strip()
                treecode = r["treecode"]
                entry = {
                    "code": code,
                    "kit": kit,
                    "module": module,
                    "item": item,
                    "treecode": treecode,
                    "level": (r["level"] or "").lower()
                }
                for tok in (code, kit, module, item):
                    if tok and tok.upper() != "NONE":
                        self._item_index_cache["flat_map"].setdefault(tok, entry)
                self._item_index_cache["triple_map"][(kit or None,
                                                      module or None,
                                                      item or None)] = entry
        except sqlite3.Error as e:
            logging.error(f"[BREAK] ensure_item_index error: {e}")
        finally:
            cur.close()
            conn.close()

    @staticmethod
    def parse_unique_id(unique_id: str):
        kit = module = item = None
        std_qty = 1
        parts = unique_id.split("/") if unique_id else []
        if len(parts) >= 2: kit = parts[1] or None
        if len(parts) >= 3: module = parts[2] or None
        if len(parts) >= 4: item = parts[3] or None
        if len(parts) >= 5:
            try:
                v = int(parts[4])
                if v > 0: std_qty = v
            except:
                std_qty = 1
        return {"kit": kit, "module": module, "item": item, "std_qty": std_qty}

    def enrich_stock_row(self, scenario_id, unique_id, final_qty, exp_date,
                         Kit_number, module_number, line_id):
        self.ensure_item_index(scenario_id)
        parsed = self.parse_unique_id(unique_id)
        kit_code = parsed["kit"]
        module_code = parsed["module"]
        item_code = parsed["item"]
        std_qty = parsed["std_qty"]

        if item_code and item_code.upper() != "NONE":
            display_code = item_code
            forced = "Item"
            triple_key = (kit_code, module_code, item_code)
        elif module_code and module_code.upper() != "NONE":
            display_code = module_code
            forced = "Module"
            triple_key = (kit_code, module_code, None)
        else:
            display_code = kit_code
            forced = "Kit"
            triple_key = (kit_code, None, None)

        idx = self._item_index_cache
        entry = idx.get("triple_map", {}).get(triple_key) or \
                idx.get("flat_map", {}).get(display_code or "")
        treecode = entry.get("treecode") if entry else None

        description = get_item_description(display_code or "")
        detected = detect_type(display_code or "", description) or forced
        m = {"KIT":"Kit","MODULE":"Module","ITEM":"Item"}
        detected_norm = m.get(detected.upper(), forced)
        final_type = forced if forced in ("Module","Item") else ("Kit" if detected_norm not in ("Kit",) else detected_norm)

        return {
            "unique_id": unique_id,
            "code": display_code or "",
            "description": description,
            "type": final_type,
            "Kit": kit_code or "-----",
            "module": module_code or "-----",
            "current_stock": final_qty,
            "expiry_date": exp_date or "",
            "batch_no": "",
            "Kit_number": Kit_number,
            "module_number": module_number,
            "std_qty": std_qty if final_type == "Item" else None,
            "line_id": line_id,
            "treecode": treecode
        }

    # -------------------- Stock Fetch --------------------
    def fetch_stock_data_for_Kit_number(self, scenario_id, Kit_number, Kit_code=None):
        """
        Fetch stock data for a kit number.
        ✅ FILTERS: Only in-box items (> 6 slashes in unique_id).
        """
        self.ensure_item_index(scenario_id)
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_stock_data_for_Kit_number")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # ✅ Filter: Only in-box items (> 6 slashes)
            cur.execute("""
                SELECT unique_id, final_qty, exp_date, kit_number, module_number, line_id
                FROM stock_data
                WHERE kit_number=? 
                  AND unique_id LIKE ? 
                  AND final_qty > 0
                  AND LENGTH(unique_id) - LENGTH(REPLACE(unique_id, '/', '')) > 6
                ORDER BY unique_id
            """, (Kit_number, f"{scenario_id}/%"))
            
            rows = cur.fetchall()
            
            result = [
                self.enrich_stock_row(
                    scenario_id, r["unique_id"], r["final_qty"],
                    r["exp_date"], r["kit_number"], r["module_number"], r["line_id"]
                )
                for r in rows
            ]
            
            logging.info(f"[OUT_KIT] Found {len(result)} in-box items for Kit_number={Kit_number}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_stock_data_for_Kit_number error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_stock_data_for_module_number(self, scenario_id, module_number, Kit_code=None, module_code=None):
        """
        Fetch stock data for a module number.
        ✅ FILTERS: Only in-box items (> 6 slashes in unique_id).
        """
        self.ensure_item_index(scenario_id)
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_stock_data_for_module_number")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # ✅ Filter: Only in-box items (> 6 slashes)
            cur.execute("""
                SELECT unique_id, final_qty, exp_date, kit_number, module_number, line_id
                FROM stock_data
                WHERE module_number=? 
                  AND unique_id LIKE ? 
                  AND final_qty > 0
                  AND LENGTH(unique_id) - LENGTH(REPLACE(unique_id, '/', '')) > 6
                ORDER BY unique_id
            """, (module_number, f"{scenario_id}/%"))
            
            rows = cur.fetchall()
            
            result = [
                self.enrich_stock_row(
                    scenario_id, r["unique_id"], r["final_qty"],
                    r["exp_date"], r["kit_number"], r["module_number"], r["line_id"]
                )
                for r in rows
            ]
            
            logging.info(f"[OUT_KIT] Found {len(result)} in-box items for module_number={module_number}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_stock_data_for_module_number error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_standalone_stock_items(self, scenario_id):
        """
        Fetch standalone stock items (no kit_number or module_number).
        ✅ FILTERS: Only in-box items (> 6 slashes in unique_id).
        """
        self.ensure_item_index(scenario_id)
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_standalone_stock_items")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # ✅ Filter: Only in-box items (> 6 slashes)
            cur.execute("""
                SELECT unique_id, final_qty, exp_date, kit_number, module_number, line_id
                FROM stock_data
                WHERE final_qty > 0
                  AND (kit_number IS NULL OR kit_number='None')
                  AND (module_number IS NULL OR module_number='None')
                  AND unique_id LIKE ?
                  AND LENGTH(unique_id) - LENGTH(REPLACE(unique_id, '/', '')) > 6
                ORDER BY unique_id
            """, (f"{scenario_id}/%",))
            
            rows = cur.fetchall()
            
            result = [
                self.enrich_stock_row(
                    scenario_id, r["unique_id"], r["final_qty"],
                    r["exp_date"], r["kit_number"], r["module_number"], r["line_id"]
                )
                for r in rows
            ]
            
            logging.info(f"[OUT_KIT] Found {len(result)} standalone in-box items")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_standalone_stock_items error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    # -------------------- Structural fetch helpers --------------------
    def fetch_Kits(self, scenario_id):
        """
        Fetch PRIMARY kits from kit_items (level='primary').
        Only includes items with type='Kit' (language-independent).
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_Kits")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get primary kits from kit_items
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id,))
            
            kit_codes = [r['code'] for r in cur.fetchall()]
            
            if not kit_codes:
                logging.debug(f"[OUT_KIT] No primary kits found for scenario {scenario_id}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for kit_code in kit_codes:
                desc = get_item_description(kit_code)
                item_type = detect_type(kit_code, desc).upper()
                
                # Only include if type is KIT
                if item_type == "KIT":
                    display = f"{kit_code} - {desc}" if desc else kit_code
                    result.append(display)
            
            logging.info(f"[OUT_KIT] Found {len(result)} primary kits for scenario {scenario_id}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_Kits error: {e}")
            return []
        finally:
            cur.close()
            conn.close()


    def fetch_all_modules(self, scenario_id):
        """
        Fetch PRIMARY modules from kit_items (level='primary', standalone modules).
        Only includes items with type='Module' (language-independent).
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_all_modules")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get primary standalone modules
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND module IS NOT NULL 
                  AND module != ''
                  AND module != 'None'
                  AND (kit IS NULL OR kit = '' OR kit = 'None')
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id,))
            
            module_codes = [r['code'] for r in cur.fetchall()]
            
            if not module_codes:
                logging.debug(f"[OUT_KIT] No primary modules found for scenario {scenario_id}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()
                
                # Only include if type is MODULE
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)
            
            logging.info(f"[OUT_KIT] Found {len(result)} primary modules for scenario {scenario_id}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_all_modules error: {e}")
            return []
        finally:
            cur.close()
            conn.close()



    def fetch_available_Kit_numbers(self, scenario_id, Kit_code=None):
        """
        Fetch kit numbers with available stock (final_qty > 0).
        
        Args:
            scenario_id: Scenario ID
            Kit_code: Optional kit code to filter by (extracted from dropdown)
        
        Returns:
            List of kit_number strings
        """
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_available_Kit_numbers")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))
            
            where_clauses = [
                "(scenario=? OR scenario=?)",
                "kit_number IS NOT NULL",
                "kit_number != 'None'",
                "final_qty > 0"
            ]
            params = [str(scenario_id), scenario_name]
            
            # Filter by kit code if provided
            if Kit_code:
                where_clauses.append("kit=?")
                params.append(Kit_code)
                logging.debug(f"[OUT_KIT] Filtering kit numbers by kit_code={Kit_code}")
            
            sql = f"""
                SELECT DISTINCT kit_number
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY kit_number
            """
            
            logging.debug(f"[OUT_KIT] SQL: {sql}")
            logging.debug(f"[OUT_KIT] Params: {params}")
            
            cur.execute(sql, params)
            results = [r['kit_number'] for r in cur.fetchall()]
            
            logging.info(f"[OUT_KIT] Found {len(results)} kit numbers")
            return results
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_available_Kit_numbers error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_modules_for_kit(self, scenario_id, kit_code):
        """
        Fetch SECONDARY modules inside a specific kit from kit_items.
        Only includes items with type='Module'.
        
        Args:
            scenario_id: Scenario ID
            kit_code: Kit code (extracted from dropdown)
        
        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_modules_for_kit")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            # Get secondary modules inside this kit
            cur.execute("""
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND kit=?
                  AND level='secondary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """, (scenario_id, kit_code))
            
            module_codes = [r['code'] for r in cur.fetchall()]
            
            if not module_codes:
                logging.debug(f"[OUT_KIT] No modules found in kit {kit_code}")
                return []
            
            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()
                
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)
            
            logging.debug(f"[OUT_KIT] Found {len(result)} modules in kit {kit_code}")
            return result
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_modules_for_kit error: {e}")
            return []
        finally:
            cur.close()
            conn.close()



    def fetch_module_numbers(self, scenario_id, Kit_code=None, module_code=None, kit_number=None):
        """
        Fetch module numbers with available stock (final_qty > 0).
        
        Args:
            scenario_id: Scenario ID
            Kit_code: NOT USED (kept for compatibility)
            module_code: Module code to filter by
            kit_number: Kit number to filter by (the actual kit instance)
        
        Returns:
            List of module_number strings
        """
        conn = connect_db()
        if conn is None:
            logging.error("[OUT_KIT] DB connection failed in fetch_module_numbers")
            return []
        
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))
            
            where_clauses = [
                "(scenario=? OR scenario=?)",
                "module_number IS NOT NULL",
                "module_number != 'None'",
                "final_qty > 0"
            ]
            params = [str(scenario_id), scenario_name]
            
            # Filter by kit_number (the actual kit instance)
            if kit_number:
                where_clauses.append("kit_number=?")
                params.append(kit_number)
                logging.debug(f"[OUT_KIT] Filtering by kit_number={kit_number}")
            
            # Filter by module code
            if module_code:
                where_clauses.append("module=?")
                params.append(module_code)
                logging.debug(f"[OUT_KIT] Filtering by module_code={module_code}")
            
            sql = f"""
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY module_number
            """
            
            logging.debug(f"[OUT_KIT] SQL: {sql}")
            logging.debug(f"[OUT_KIT] Params: {params}")
            
            cur.execute(sql, params)
            results = [r['module_number'] for r in cur.fetchall()]
            
            logging.info(f"[OUT_KIT] Found {len(results)} module numbers")
            return results
            
        except sqlite3.Error as e:
            logging.error(f"[OUT_KIT] fetch_module_numbers error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

# ============================= out_kit.py (Part 3/6) =============================
    # -------------------- UI Rendering --------------------
    def render_ui(self):
        """
        Build the complete UI matching dispatch_kit.py style.
        ✅ Uses AppTheme colors
        ✅ Fixed OUT Type (read-only label)
        ✅ Wider Kit/Module dropdowns
        ✅ Right-click quantity editing
        ✅ Fully translated
        """
        if not self.parent:
            return
        
        # Clear existing widgets
        for w in self.parent.winfo_children():
            try:
                w.destroy()
            except:
                pass

        # Apply theme
        apply_global_style()

        # ===== TITLE =====
        title_frame = tk.Frame(self.parent, bg=AppTheme.BG_MAIN)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text=lang.t("out_kit.title", "Out Kit/Module"),
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_HUGE, "bold"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.TEXT_DARK
        ).pack(pady=(10, 0))

        # ===== INSTRUCTIONS =====
        instruct_frame = tk.Frame(
            self.parent,
            bg="#FFF9C4",
            highlightbackground="#E0D890",
            highlightthickness=1,
            bd=0
        )

        # ===== MAIN CONTAINER =====
        main = tk.Frame(self.parent, bg=AppTheme.BG_MAIN)
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # Row 0: Scenario
        tk.Label(
            main,
            text=lang.t("out_kit.scenario", "Scenario:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(
            main,
            textvariable=self.scenario_var,
            state="readonly",
            width=40,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.scenario_cb.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)

        # Row 1: Movement Type
        tk.Label(
            main,
            text=lang.t("out_kit.movement_type", "Movement Type:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        
        self.mode_var = tk.StringVar()
        self.mode_cb = ttk.Combobox(
            main,
            textvariable=self.mode_var,
            state="readonly",
            width=40,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.mode_cb.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.mode_cb.bind("<<ComboboxSelected>>", self.on_mode_changed)

        # Row 2: Kit and Kit Number
        tk.Label(
            main,
            text=lang.t("out_kit.select_kit", "Select Kit:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=2, column=0, padx=5, pady=5, sticky="w")
        
        self.Kit_var = tk.StringVar()
        self.Kit_cb = ttk.Combobox(
            main,
            textvariable=self.Kit_var,
            state="disabled",
            width=80,  # ✅ WIDER: Matches dispatch_kit.py
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.Kit_cb.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.Kit_cb.bind("<<ComboboxSelected>>", self.on_Kit_selected)

        tk.Label(
            main,
            text=lang.t("out_kit.select_kit_number", "Select Kit Number:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=2, column=2, padx=5, pady=5, sticky="w")
        
        self.Kit_number_var = tk.StringVar()
        self.Kit_number_cb = ttk.Combobox(
            main,
            textvariable=self.Kit_number_var,
            state="disabled",
            width=20,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.Kit_number_cb.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.Kit_number_cb.bind("<<ComboboxSelected>>", self.on_Kit_number_selected)

        # Row 3: Module and Module Number
        tk.Label(
            main,
            text=lang.t("out_kit.select_module", "Select Module:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=3, column=0, padx=5, pady=5, sticky="w")
        
        self.module_var = tk.StringVar()
        self.module_cb = ttk.Combobox(
            main,
            textvariable=self.module_var,
            state="disabled",
            width=80,  # ✅ WIDER: Matches dispatch_kit.py
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.module_cb.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.module_cb.bind("<<ComboboxSelected>>", self.on_module_selected)

        tk.Label(
            main,
            text=lang.t("out_kit.select_module_number", "Select Module Number:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=3, column=2, padx=5, pady=5, sticky="w")
        
        self.module_number_var = tk.StringVar()
        self.module_number_cb = ttk.Combobox(
            main,
            textvariable=self.module_number_var,
            state="disabled",
            width=20,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        )
        self.module_number_cb.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        self.module_number_cb.bind("<<ComboboxSelected>>", self.on_module_number_selected)

        # Row 4: Fixed OUT Type (read-only label) - NO dropdowns for End User/Third Party
        type_frame = tk.Frame(main, bg=AppTheme.BG_MAIN)
        type_frame.grid(row=4, column=0, columnspan=4, pady=5, sticky="w")
        
        tk.Label(
            type_frame,
            text=lang.t("out_kit.out_type", "OUT Type:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=0, column=0, padx=5, sticky="w")
        
        # ✅ Display translated OUT Type (but store canonical English)
        self.out_type_var = tk.StringVar(
            value=lang.t("out_kit.out_type_fixed", "Internal move from in-box items")
        )
        tk.Label(
            type_frame,
            textvariable=self.out_type_var,
            bg="#E0E0E0",
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL),
            relief="sunken",
            width=35,
            anchor="w",
            padx=8,
            pady=4
        ).grid(row=0, column=1, padx=5, pady=5)

        # Row 5: Search
        tk.Label(
            main,
            text=lang.t("out_kit.item_search", "Kit/Module/Item:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
        ).grid(row=5, column=0, padx=5, pady=5, sticky="w")
        
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(
            main,
            textvariable=self.search_var,
            width=40,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL),
            bg=AppTheme.ENTRY_BG,
            fg=AppTheme.ENTRY_FG
        )
        self.search_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", self.search_items)

        tk.Button(
            main,
            text=lang.t("out_kit.clear_search", "Clear Search"),
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            command=self.clear_search,
            relief="flat",
            cursor="hand2"
        ).grid(row=5, column=2, padx=5, pady=5)

        # Row 6: Search Listbox
        self.search_listbox = tk.Listbox(
            main,
            height=5,
            width=60,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL),
            bg=AppTheme.ENTRY_BG,
            fg=AppTheme.ENTRY_FG
        )
        self.search_listbox.grid(row=6, column=1, columnspan=3, padx=5, pady=5, sticky="we")

        logging.debug("[RENDER_UI] Search listbox bindings set")

        # ===== TREEVIEW =====
        # ✅ Out_kit has 12 columns (includes line_id, qty_in_hidden - hidden)
        cols = (
            "code", "description", "type", "Kit", "module",
            "current_stock", "expiry_date", "batch_no",
            "qty_to_issue", "unique_id", "line_id", "qty_in_hidden"
        )
        
        self.tree = ttk.Treeview(main, columns=cols, show="headings", height=18)

        headings = {
            "code": lang.t("out_kit.code", "Code"),
            "description": lang.t("out_kit.description", "Description"),
            "type": lang.t("out_kit.type", "Type"),
            "Kit": lang.t("out_kit.kit", "Kit"),
            "module": lang.t("out_kit.module", "Module"),
            "current_stock": lang.t("out_kit.current_stock", "Current Stock"),
            "expiry_date": lang.t("out_kit.expiry_date", "Expiry Date"),
            "batch_no": lang.t("out_kit.batch_no", "Batch Number"),
            "qty_to_issue": lang.t("out_kit.qty_to_out", "Qty to Out"),
            "unique_id": "unique_id",
            "line_id": "line_id (hidden)",
            "qty_in_hidden": "qty_in (hidden)"
        }

        widths = {
            "code": 150,
            "description": 360,
            "type": 110,
            "Kit": 120,
            "module": 120,
            "current_stock": 110,
            "expiry_date": 140,
            "batch_no": 130,
            "qty_to_issue": 140,
            "unique_id": 0,      # Hidden
            "line_id": 0,        # Hidden
            "qty_in_hidden": 0   # Hidden
        }

        aligns = {
            "code": "w",
            "description": "w",
            "type": "w",
            "Kit": "w",
            "module": "w",
            "current_stock": "e",
            "expiry_date": "w",
            "batch_no": "w",
            "qty_to_issue": "e",
            "unique_id": "w",
            "line_id": "w",
            "qty_in_hidden": "e"
        }

        hidden_cols = ("unique_id", "line_id", "qty_in_hidden")
        
        for c in cols:
            self.tree.heading(c, text=headings[c])
            is_hidden = c in hidden_cols
            self.tree.column(
                c,
                width=widths[c],
                anchor=aligns[c],
                stretch=(False if is_hidden else True),
                minwidth=0 if is_hidden else widths[c]
            )

        # Scrollbars
        vsb = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        vsb.grid(row=7, column=4, sticky="ns")
        self.tree.configure(yscrollcommand=vsb.set)
        
        hsb = ttk.Scrollbar(main, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=8, column=0, columnspan=4, sticky="ew")
        self.tree.configure(xscrollcommand=hsb.set)
        
        self.tree.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")
        self.configure_tree_tags()
        main.grid_rowconfigure(7, weight=1)
        main.grid_columnconfigure(1, weight=1)

        # Tree bindings
        self.tree.bind("<Double-1>", self.start_edit)
        self.tree.bind("<KeyPress-Return>", self.start_edit)
        self.tree.bind("<KeyPress-Tab>", self.start_edit)
        self.tree.bind("<KeyPress-Up>", self.navigate_tree)
        self.tree.bind("<KeyPress-Down>", self.navigate_tree)
        
        # ✅ Right-click for quantity editing
        self.tree.bind("<Button-3>", self.show_qty_edit_menu)  # Windows/Linux
        self.tree.bind("<Control-Button-1>", self.show_qty_edit_menu)  # Mac
        
        logging.debug("[RENDER_UI] Tree bindings set - starred cells editable, qty_to_issue via right-click")

        # ✅ Configure visual tags
        self.tree.tag_configure("Kit_header", background="#E3F6E1", font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"))
        self.tree.tag_configure("module_header", background="#E1ECFC", font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"))
        self.tree.tag_configure("Kit_module_highlight", background="#FFF9C4")
        self.tree.tag_configure("editable_row", foreground="#000000")
        self.tree.tag_configure("non_editable", foreground="#666666")
        self.tree.tag_configure("header_kit", background="#E3F6E1", font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"))
        self.tree.tag_configure("kit_data", background="#C5EDC1")
        self.tree.tag_configure("header_module", background="#E1ECFC", font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"))
        self.tree.tag_configure("module_data", background="#C9E2FA")
        self.tree.tag_configure("item_row", foreground="#222222")
        
        logging.debug("[RENDER_UI] Tree visual tags configured")

        # ===== BUTTONS =====
        btnf = tk.Frame(main, bg=AppTheme.BG_MAIN)
        btnf.grid(row=9, column=0, columnspan=4, pady=5)
        
        tk.Button(
            btnf,
            text=lang.t("out_kit.save", "Save"),
            bg=AppTheme.BTN_SUCCESS,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            command=self.save_all,
            state="normal" if self.role in ["admin", "manager", "supervisor"] else "disabled",
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=6
        ).pack(side="left", padx=5)
        
        tk.Button(
            btnf,
            text=lang.t("out_kit.clear", "Clear"),
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            command=self.clear_form,
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=6
        ).pack(side="left", padx=5)
        
        tk.Button(
            btnf,
            text=lang.t("out_kit.export", "Export"),
            bg=AppTheme.BTN_EXPORT,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            command=self.export_data,
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=6
        ).pack(side="left", padx=5)

        # ===== STATUS BAR =====
        self.status_var = tk.StringVar(value=lang.t("out_kit.ready", "Ready"))
        tk.Label(
            main,
            textvariable=self.status_var,
            relief="sunken",
            anchor="w",
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL)
        ).grid(row=10, column=0, columnspan=4, sticky="ew")

        # Load initial data
        self.load_scenarios()
        
        logging.info("[OUT_KIT] UI rendered successfully matching dispatch_kit.py style")



    def configure_tree_tags(self):
        """
        Configure visual styling tags for tree rows.
        ✅ Kit headers: Green background
        ✅ Module headers: Blue background
        ✅ Kit data rows: Light green
        ✅ Module data rows: Light blue  
        ✅ Adopted expiry: Red/pink background
        ✅ Edited expiry: Light cyan background (NEW)
        """
        try:
            # ===== Header styles =====
            self.tree.tag_configure(
                "kit_header",
                background="#C8E6C9",  # Light green
                font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
            )
            
            self.tree.tag_configure(
                "module_header",
                background="#BBDEFB",  # Light blue
                font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold")
            )
            
            # ===== Data row styles =====
            self.tree.tag_configure(
                "kit_data",
                background="#C5EDC1"  # ✅ Light green for kit rows
            )
            
            self.tree.tag_configure(
                "module_data",
                background="#C9E2FA"  # Light blue for module rows
            )
            
            # ✅ DEPRECATED: Old yellow highlight (kept for compatibility)
            self.tree.tag_configure(
                "kit_module_highlight",
                background="#FFF9C4"  # Light yellow (deprecated)
            )
            
            # ===== Special states =====
            self.tree.tag_configure(
                "adopted_expiry",
                background="#FFCDD2",  # Light red/pink for adopted expiries
                foreground="#B71C1C"   # Dark red text
            )
            
            # ✅ NEW: Light cyan for user-edited expiries
            self.tree.tag_configure(
                "expiry_edited",
                background="#B2EBF2",  # Light cyan (different from yellow)
                foreground="#006064"   # Dark cyan text
            )
            
            logging.debug("[OUT_KIT] Tree visual tags configured")
            
        except Exception as e:
            logging.warning(f"[OUT_KIT] Tag configuration failed: {e}")


    # -------------------- Row building & sorting (treecode) --------------------
    def _build_with_headers(self, rows):
        def sort_key(it):
            return (
                it.get("Kit_number") or "",
                it.get("module_number") or "",
                it.get("treecode") or "ZZZ",
                it.get("code") or ""
            )
        ordered = sorted(rows, key=sort_key)
        result = []
        seen_kits = set()
        seen_modules = set()
        for it in ordered:
            kit_code = it.get("Kit") if it.get("Kit") and it.get("Kit") != "-----" else None
            module_code = it.get("module") if it.get("module") and it.get("module") != "-----" else None
            kit_number = it.get("Kit_number")
            module_number = it.get("module_number")
            if kit_code and kit_number and (kit_code, kit_number) not in seen_kits:
                result.append({
                    "is_header": True,
                    "type": "Kit",
                    "code": kit_code,
                    "description": get_item_description(kit_code),
                    "Kit": kit_number,
                    "module": "",
                    "current_stock": "",
                    "expiry_date": "",
                    "batch_no": "",
                    "unique_id": "",
                    "Kit_number": kit_number,
                    "module_number": None,
                    "line_id": "",
                    "treecode": ""
                })
                seen_kits.add((kit_code, kit_number))
            if module_code and module_number and (kit_code, module_code, module_number, kit_number) not in seen_modules:
                result.append({
                    "is_header": True,
                    "type": "Module",
                    "code": module_code,
                    "description": get_item_description(module_code),
                    "Kit": kit_number or "",
                    "module": module_number,
                    "current_stock": "",
                    "expiry_date": "",
                    "batch_no": "",
                    "unique_id": "",
                    "Kit_number": kit_number,
                    "module_number": module_number,
                    "line_id": "",
                    "treecode": ""
                })
                seen_modules.add((kit_code, module_code, module_number, kit_number))
            result.append(it)
        return result

    # -------------------- Mode rules / quantity derivation --------------------
    def get_mode_rules(self):
        mode = self.current_mode_key()
        rules = {
            "editable_types": set(),
            "derive_modules_from_Kit": False,
            "derive_items_from_modules": False
        }
        if mode == "dispatch_Kit":
            rules.update({
                "editable_types": {"Kit"},
                "derive_modules_from_Kit": True,
                "derive_items_from_modules": True
            })
        elif mode in ("issue_module_scenario","issue_module_Kit"):
            rules.update({
                "editable_types": {"Module"},
                "derive_items_from_modules": True
            })
        elif mode in ("issue_standalone","issue_items_module","issue_items_Kit"):
            rules.update({"editable_types":{"Item"}})
        return rules

    def initialize_quantities_and_highlight(self):
        rules = self.get_mode_rules()
        mode_key = self.current_mode_key()
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"): continue
            vals = list(self.tree.item(iid, "values"))
            rtype = (vals[2] or "").lower()
            try:
                stock = int(vals[5]) if vals[5] else 0
            except:
                stock = 0
            if rtype == "kit":
                qty = 1 if mode_key == "dispatch_Kit" else 0
            elif rtype == "module":
                qty = 1 if ("module" in {t.lower() for t in rules["editable_types"]} and stock>0) else 0
            else:
                qty = 0
            vals[8] = str(qty)
            vals[11] = str(qty)
            self.tree.item(iid, values=vals)
        if rules.get("derive_modules_from_Kit"):
            self._derive_modules_from_Kits()
        if rules.get("derive_items_from_modules"):
            self._derive_items_from_modules()
        self._reapply_editable_icons(rules)

    def _derive_modules_from_Kits(self):
        kit_qty = {}
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid,{})
            if meta.get("is_header"): continue
            vals = self.tree.item(iid,"values")
            if (vals[2] or "").lower()=="kit":
                raw = vals[8]
                if raw.startswith("★"):
                    raw = raw[2:].strip()
                kit_qty[meta.get("Kit_number")] = int(raw) if raw.isdigit() else 0
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid,{})
            if meta.get("is_header"): continue
            vals = list(self.tree.item(iid,"values"))
            if (vals[2] or "").lower()=="module":
                base = kit_qty.get(meta.get("Kit_number"),0)
                try:
                    stock = int(vals[5]) if vals[5] else 0
                except: stock=0
                if base>stock: base=stock
                vals[8]=str(base)
                vals[11]=str(base)
                self.tree.item(iid, values=vals)

    def _derive_items_from_modules(self):
        module_map = {}
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid,{})
            if meta.get("is_header"): continue
            vals = self.tree.item(iid,"values")
            if (vals[2] or "").lower()=="module":
                raw = vals[8]
                if raw.startswith("★"):
                    raw = raw[2:].strip()
                module_map[meta.get("module_number")] = int(raw) if raw.isdigit() else 0
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid,{})
            if meta.get("is_header"): continue
            if meta.get("row_type") != "Item":
                continue
            vals = list(self.tree.item(iid,"values"))
            std_qty = meta.get("std_qty") or 1
            mod_qty = module_map.get(meta.get("module_number"),0)
            desired = std_qty * mod_qty
            try:
                stock = int(vals[5]) if vals[5] else 0
            except: stock=0
            if desired>stock: desired=stock
            vals[8]=str(desired)
            vals[11]=str(desired)
            self.tree.item(iid, values=vals)

    def _reapply_editable_icons(self, rules):
        editable_lower = {t.lower() for t in rules["editable_types"]}
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid,{})
            vals = list(self.tree.item(iid,"values"))
            rtype = (vals[2] or "").lower()
            tags=[]
            if meta.get("is_header"):
                if rtype=="kit":
                    tags.append("header_kit")
                elif rtype=="module":
                    tags.append("header_module")
            else:
                if rtype=="kit":
                    tags.append("kit_data")
                elif rtype=="module":
                    tags.append("module_data")
                else:
                    tags.append("item_row")
            if not meta.get("is_header") and rtype in editable_lower and meta.get("unique_id"):
                core = vals[8]
                raw = core[2:].strip() if core.startswith("★") else core.strip()
                if raw=="":
                    raw="0"
                vals[8] = f"★ {raw}"
                tags.append("editable_row")
                self.tree.item(iid, values=vals, tags=tuple(tags))
            else:
                if not meta.get("is_header"):
                    tags.append("non_editable")
                self.tree.item(iid, tags=tuple(tags))

# ============================= out_kit.py (Part 4/6) =============================
    # -------------------- Event handlers (selection) --------------------
    def on_Kit_selected(self, event=None):
        """Handle kit selection - populate kit_number dropdown."""
        Kit_display = self.Kit_var.get()
        if not Kit_display:
            return
        
        Kit_code = self._extract_code_from_display(Kit_display)
        logging.debug(f"[OUT_KIT] Kit selected: '{Kit_display}' -> Code: '{Kit_code}'")
        
        # Clear dependent dropdowns
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        self.clear_table_only()
        
        if not Kit_code:
            return
        
        # Fetch kit numbers for this kit
        kit_numbers = self.fetch_available_Kit_numbers(self.selected_scenario_id, Kit_code)
        
        if kit_numbers:
            self.Kit_number_cb.config(state="readonly")
            self.Kit_number_cb['values'] = kit_numbers
            self.status_var.set(
                lang.t("out_kit.select_kit_number_msg", 
                       "Select a kit number from dropdown ({count} available)", 
                       count=len(kit_numbers))
            )
        else:
            self.Kit_number_cb.config(state="disabled")
            self.Kit_number_cb['values'] = []
            self.status_var.set(
                lang.t("out_kit.no_kit_numbers", "No kit numbers available with stock > 0")
            )

    def on_Kit_number_selected(self, event=None):
        """Handle kit_number selection - load kit stock OR populate module dropdown."""
        Kit_number = self.Kit_number_var.get()
        if not Kit_number:
            return
        
        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display)
        mode_key = self.current_mode_key()
        
        logging.debug(f"[OUT_KIT] Kit_number selected: {Kit_number}, Kit={Kit_code}, mode={mode_key}")
        
        # Clear dependent fields
        self.module_var.set("")
        self.module_number_var.set("")
        
        if mode_key == "out_kit":
            # Load entire kit
            items = self.fetch_stock_data_for_Kit_number(
                self.selected_scenario_id, 
                Kit_number, 
                Kit_code
            )
            self.populate_rows(items, status_msg=lang.t(
                "out_kit.loaded_rows_kit",
                "Loaded {n} stock rows for Kit number {k}",
                n=len(items), k=Kit_number
            ))
        
        elif mode_key == "out_module_kit":
            # ✅ FIX: Populate MODULE dropdown first (not module_number)
            # Fetch secondary modules inside this kit
            modules = self.fetch_modules_for_kit(self.selected_scenario_id, Kit_code)
            
            if modules:
                self.module_cb.config(state="readonly")
                self.module_cb['values'] = modules
                self.status_var.set(
                    lang.t("out_kit.select_module_msg",
                           "Select a module from dropdown ({count} available)",
                           count=len(modules))
                )
                logging.debug(f"[OUT_KIT] Populated {len(modules)} modules for kit {Kit_code}")
            else:
                self.module_cb.config(state="disabled")
                self.module_cb['values'] = []
                self.status_var.set(
                    lang.t("out_kit.no_modules_in_kit", "No modules found in this kit")
                )
        
        elif mode_key == "out_items_kit":
            # Populate module_number dropdown for this kit instance
            module_numbers = self.fetch_module_numbers(
                self.selected_scenario_id,
                kit_number=Kit_number
            )
            
            if module_numbers:
                self.module_number_cb.config(state="readonly")
                self.module_number_cb['values'] = module_numbers
                self.status_var.set(
                    lang.t("out_kit.select_module_number_msg",
                           "Select a module number from dropdown ({count} available)",
                           count=len(module_numbers))
                )
            else:
                self.module_number_cb.config(state="disabled")
                self.module_number_cb['values'] = []
                self.status_var.set(
                    lang.t("out_kit.no_module_numbers", "No module numbers available")
                )

    def on_module_selected(self, event=None):
        """Handle module selection - populate module_number dropdown."""
        module_display = self.module_var.get()
        if not module_display:
            return
        
        module_code = self._extract_code_from_display(module_display)
        logging.debug(f"[OUT_KIT] Module selected: '{module_display}' -> Code: '{module_code}'")
        
        # Clear dependent fields
        self.module_number_var.set("")
        self.clear_table_only()
        
        if not module_code:
            return
        
        mode_key = self.current_mode_key()
        Kit_number = self.Kit_number_var.get() if mode_key == "out_module_kit" else None
        
        # Fetch module numbers
        module_numbers = self.fetch_module_numbers(
            self.selected_scenario_id,
            module_code=module_code,
            kit_number=Kit_number
        )
        
        if module_numbers:
            self.module_number_cb.config(state="readonly")
            self.module_number_cb['values'] = module_numbers
            self.status_var.set(
                lang.t("out_kit.select_module_number_msg",
                       "Select a module number from dropdown ({count} available)",
                       count=len(module_numbers))
            )
        else:
            self.module_number_cb.config(state="disabled")
            self.module_number_cb['values'] = []
            self.status_var.set(
                lang.t("out_kit.no_module_numbers", "No module numbers available")
            )

    def on_module_number_selected(self, event=None):
        """Handle module_number selection - load module stock."""
        module_number = self.module_number_var.get()
        if not module_number:
            return
        
        module_display = self.module_var.get()
        module_code = self._extract_code_from_display(module_display) if module_display else None
        
        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display) if Kit_display else None
        
        logging.debug(f"[OUT_KIT] Module_number selected: {module_number}, module={module_code}, Kit={Kit_code}")
        
        # Load module stock
        items = self.fetch_stock_data_for_module_number(
            self.selected_scenario_id,
            module_number,
            Kit_code=Kit_code,
            module_code=module_code
        )
        
        self.populate_rows(items, status_msg=lang.t(
            "out_kit.loaded_rows_module",
            "Loaded {n} stock rows for module number {m}",
            n=len(items), m=module_number
        ))



    def populate_standalone_items(self):
        if not self.selected_scenario_id: return
        items = self.fetch_standalone_stock_items(self.selected_scenario_id)
        for it in items: it["row_type"]=it["type"]
        self.full_items = items[:]
        self.populate_rows(self.full_items,
                           f"Loaded {len(self.full_items)} standalone rows")

    # -------------------- Table population --------------------
    def clear_table_only(self):
        if self.tree:
            self.tree.delete(*self.tree.get_children())
        self.row_data.clear()

    def populate_rows(self, items=None, status_msg=""):
        """
        Populate tree with items, checking for adopted expiries.
        ✅ Highlights rows with adopted expiry dates
        ✅ Stores adoption flag in row_data
        """
        if items is None:
            items = self.full_items
        else:
            self.full_items = items
        
        self.clear_table_only()
        
        if not items:
            self.status_var.set(lang.t("out_kit.no_data", "No data to display"))
            return
        
        # Build with headers
        rows_with_headers = self._build_with_headers(items)
        
        logging.debug(f"[POPULATE_ROWS] Inserting {len(rows_with_headers)} rows ({len(items)} items)")
        
        adopted_count = 0
        
        for item in rows_with_headers:
            is_header = item.get("is_header", False)
            
            # Prepare tree values (12 columns)
            vals = (
                item.get("code", ""),
                item.get("description", ""),
                item.get("type", ""),
                item.get("Kit", "-----"),
                item.get("module", "-----"),
                str(item.get("current_stock", "")),
                item.get("expiry_date", ""),
                item.get("batch_no", ""),
                str(item.get("current_stock", "")),  # qty_to_issue = current_stock
                item.get("unique_id", ""),
                item.get("line_id", ""),
                str(item.get("current_stock", ""))  # qty_in_hidden = current_stock
            )
            
            # Insert into tree
            iid = self.tree.insert("", "end", values=vals)
            
            # ✅ CHECK: Adopted expiry? (only for data rows with unique_id)
            is_adopted = False
            original_comment = ""
            
            if not is_header and item.get("unique_id"):
                is_adopted, original_comment = self.check_if_adopted_expiry(item["unique_id"])
                
                if is_adopted:
                    adopted_count += 1
            
            # ✅ Store metadata
            self.row_data[iid] = {
                "unique_id": item.get("unique_id", ""),
                "code": item.get("code", ""),
                "is_header": is_header,
                "row_type": item.get("type", ""),  # ✅ Store type for later use
                "Kit_number": item.get("Kit_number"),
                "module_number": item.get("module_number"),
                "treecode": item.get("treecode"),
                "std_qty": item.get("std_qty"),
                "adopted_expiry": is_adopted,
                "original_comment": original_comment,
                "expiry_edited": False  # Track if user manually edits
            }
            
            # ===== BUILD AND APPLY TAGS =====
            tags = []
            row_type = item.get("type", "").lower()
            
            if is_header:
                # ✅ HEADER ROWS: Green for Kit, Blue for Module
                if row_type == "kit":
                    tags.append("kit_header")
                    logging.debug(f"[POPULATE] Kit header: {item.get('code')}")
                elif row_type == "module":
                    tags.append("module_header")
                    logging.debug(f"[POPULATE] Module header: {item.get('code')}")
            else:
                # ✅ DATA ROWS: Apply type-specific background first
                if row_type == "kit":
                    tags.append("kit_data")  # Light green background
                elif row_type == "module":
                    tags.append("module_data")  # Light blue background
                # Items get no special background (default white)
                
                # ✅ HIGHLIGHT: Adopted expiry (RED - highest priority)
                if is_adopted:
                    tags.append("adopted_expiry")  # Overrides type background with red/pink
                # ✅ HIGHLIGHT: User-edited expiry (CYAN - secondary priority)
                elif self.row_data[iid].get("expiry_edited"):
                    tags.append("expiry_edited")  # Overrides type background with cyan
            
            # Apply all tags at once
            if tags:
                self.tree.item(iid, tags=tuple(tags))
        
        # Configure visual tags
        self.configure_tree_tags()
        
        # Status message
        if adopted_count > 0:
            msg = lang.t(
                "out_kit.loaded_with_adopted",
                "Loaded {total} rows. ⚠️ {adopted} items have ADOPTED expiry dates (highlighted in red). Double-click expiry to edit.",
                total=len(rows_with_headers),
                adopted=adopted_count
            )
        else:
            msg = status_msg or lang.t("out_kit.loaded_rows", "Loaded {count} rows", count=len(rows_with_headers))
        
        self.status_var.set(msg)
        logging.info(f"[POPULATE_ROWS] Populated {len(rows_with_headers)} rows with highlights (adopted={adopted_count})")

    # -------------------- Editing --------------------
    def start_edit(self, event):
        """
        Start editing a cell on double-click.
        ✅ Allows editing expiry_date for rows with adopted expiry
        """
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        
        if not row_id or not col_id:
            return
        
        # Check if header row
        meta = self.row_data.get(row_id, {})
        if meta.get("is_header"):
            return
        
        # Get column index (0-based)
        col_index = int(col_id.replace("#", "")) - 1
        
        # Column 6 = expiry_date
        if col_index == 6:
            # ✅ Allow editing if adopted expiry
            if meta.get("adopted_expiry"):
                self._start_edit_cell(row_id, col_index)
            else:
                self.status_var.set(
                    lang.t("out_kit.expiry_not_editable", "Expiry date is not editable (not adopted)")
                )
        else:
            self.status_var.set(
                lang.t("out_kit.column_not_editable", "This column is not editable")
            )

    def navigate_tree(self, event):
        if self.editing_cell: return
        rows = list(self.tree.get_children())
        if not rows: return
        sel = self.tree.selection()
        cur = sel[0] if sel else rows[0]
        idx = rows.index(cur)
        if event.keysym=="Up" and idx>0:
            self.tree.selection_set(rows[idx-1]); self.tree.focus(rows[idx-1])
        elif event.keysym=="Down" and idx < len(rows)-1:
            self.tree.selection_set(rows[idx+1]); self.tree.focus(rows[idx+1])

    def show_qty_edit_menu(self, event):
        """
        Show context menu for editing qty_to_issue on right-click.
        Works for ALL rows (Kit/Module/Item).
        """
        # Identify what was clicked
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        
        if not row_id or not col_id:
            return
        
        # Check if qty_to_issue column (column 9, index 8)
        col_index = int(col_id.replace("#", "")) - 1
        if col_index != 8:  # qty_to_issue is column 8 (0-indexed)
            return
        
        # Get current values
        vals = self.tree.item(row_id, "values")
        
        # Check if row has stock
        try:
            current_stock = int(vals[5]) if vals[5] else 0
        except:
            current_stock = 0
        
        if current_stock == 0:
            self.status_var.set(lang.t("out_kit.no_stock", "No stock available"))
            return
        
        # Create context menu
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(
            label=lang.t("out_kit.edit_quantity", "Edit Quantity"),
            command=lambda: self.edit_qty_to_issue_popup(row_id, current_stock)
        )
        
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def edit_qty_to_issue_popup(self, row_id, max_stock):
        """
        Open dialog to edit qty_to_issue for an item row.
        Uses native Tkinter dialog for maximum compatibility.
        
        Args:
            row_id: Tree item ID
            max_stock: Maximum allowed quantity (current stock)
        """
        vals = list(self.tree.item(row_id, "values"))
        code = vals[0]
        description = vals[1]
        current_qty_str = vals[8]  # ✅ qty_to_issue is column 8 in out_kit
        
        # Strip star if present
        if isinstance(current_qty_str, str) and current_qty_str.startswith("★"):
            current_qty_str = current_qty_str[2:].strip()
        
        try:
            current_qty = int(current_qty_str) if str(current_qty_str).isdigit() else 0
        except:
            current_qty = 0
        
        # ===== CREATE CUSTOM DIALOG =====
        dialog = tk.Toplevel(self.parent)
        dialog.title(lang.t("out_kit.edit_qty_title", "Edit Quantity"))
        dialog.geometry("450x450")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Main frame
        main_frame = tk.Frame(dialog, bg="#F0F4F8", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text=lang.t("out_kit.edit_quantity", "Edit Quantity"),
            font=("Helvetica", 14, "bold"),
            bg="#F0F4F8"
        )
        title_label.pack(pady=(0, 15))

        instruction_label = tk.Label(
            main_frame,
            text=lang.t("out_kit.edit_qty_instruction", "Enter new quantity and press ENTER or click Save"),
            font=("Helvetica", 9, "italic"),
            fg="#555",
            bg="#F0F4F8"
        )
        instruction_label.pack(pady=(0, 15))
        
        # Item info frame
        info_frame = tk.Frame(main_frame, bg="white", relief="solid", borderwidth=1)
        info_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            info_frame,
            text=f"{lang.t('out_kit.code', 'Code')}: {code}",
            font=("Helvetica", 11, "bold"),
            bg="white",
            anchor="w"
        ).pack(fill="x", padx=10, pady=5)
        
        tk.Label(
            info_frame,
            text=description,
            font=("Helvetica", 10),
            bg="white",
            anchor="w",
            wraplength=400,
            justify="left"
        ).pack(fill="x", padx=10, pady=(0, 5))
        
        # Stock info frame
        stock_frame = tk.Frame(main_frame, bg="#E8F4F8", relief="solid", borderwidth=1)
        stock_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            stock_frame,
            text=f"{lang.t('out_kit.current_stock', 'Current Stock')}: {max_stock}",
            font=("Helvetica", 10),
            bg="#E8F4F8",
            anchor="w"
        ).pack(fill="x", padx=10, pady=3)
        
        tk.Label(
            stock_frame,
            text=f"{lang.t('out_kit.current_qty_issue', 'Current Qty to Issue')}: {current_qty}",
            font=("Helvetica", 10),
            bg="#E8F4F8",
            anchor="w"
        ).pack(fill="x", padx=10, pady=3)
        
        # Entry frame
        entry_frame = tk.Frame(main_frame, bg="#F0F4F8")
        entry_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(
            entry_frame,
            text=f"{lang.t('out_kit.new_quantity', 'New Quantity')}:",
            font=("Helvetica", 11, "bold"),
            bg="#F0F4F8"
        ).pack(anchor="w")
        
        qty_var = tk.StringVar(value=str(current_qty))
        qty_entry = tk.Entry(
            entry_frame,
            textvariable=qty_var,
            font=("Helvetica", 12),
            width=15
        )
        qty_entry.pack(anchor="w", pady=5)
        qty_entry.focus()
        qty_entry.select_range(0, tk.END)
        
        # ✅ FIX: Define status_label BEFORE using it in save_quantity()
        status_label = tk.Label(
            main_frame,
            text="",
            font=("Helvetica", 9),
            fg="red",
            bg="#F0F4F8",
            wraplength=400
        )
        status_label.pack(pady=5)
        
        # Result variable
        result = {"cancelled": True, "value": None}
        
        def save_quantity():
            """Validate and save the quantity."""
            new_qty_str = qty_var.get().strip()
            
            # Validate
            if not new_qty_str.isdigit():
                status_label.config(
                    text=lang.t("out_kit.error_invalid_number", "Please enter a valid number")
                )
                return
            
            new_qty = int(new_qty_str)
            
            if new_qty < 0:
                status_label.config(
                    text=lang.t("out_kit.error_negative", "Quantity cannot be negative")
                )
                return
            
            if new_qty > max_stock:
                status_label.config(
                    text=lang.t("out_kit.error_exceeds_stock", 
                               "Exceeds available stock ({stock})").format(stock=max_stock)
                )
                return
            
            # Valid input
            result["cancelled"] = False
            result["value"] = new_qty
            dialog.destroy()
        
        def cancel():
            """Close without saving."""
            result["cancelled"] = True
            dialog.destroy()
        
        # Button frame
        btn_frame = tk.Frame(main_frame, bg="#F0F4F8")
        btn_frame.pack(side="bottom", pady=10)
        
        tk.Button(
            btn_frame,
            text=lang.t("out_kit.save", "Save"),
            font=("Helvetica", 10, "bold"),
            bg="#27AE60",
            fg="white",
            width=10,
            command=save_quantity
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text=lang.t("out_kit.cancel", "Cancel"),
            font=("Helvetica", 10),
            bg="#7F8C8D",
            fg="white",
            width=10,
            command=cancel
        ).pack(side="left", padx=5)
        
        # Bind keys
        qty_entry.bind("<Return>", lambda e: save_quantity())
        qty_entry.bind("<KP_Enter>", lambda e: save_quantity())
        dialog.bind("<Escape>", lambda e: cancel())
        
        # Wait for dialog
        dialog.wait_window()
        
        # Process result
        if result["cancelled"]:
            return
        
        new_qty = result["value"]
        
        # Update tree (column 8 is qty_to_issue)
        vals[8] = str(new_qty)
        
        # ✅ Also update hidden column 11 (qty_in_mirror for dual transaction)
        vals[11] = str(new_qty)
        
        # ✅ Apply updated values to tree
        self.tree.item(row_id, values=vals)
        
        # Update row_data
        if row_id in self.row_data:
            self.row_data[row_id]['qty_to_issue'] = new_qty
        
        self.status_var.set(
            lang.t("out_kit.qty_updated", 
                   "Quantity updated for {code}: {qty}").format(code=code, qty=new_qty)
        )
        
        logging.debug(f"[OUT_KIT] Updated {code} qty_to_issue: {current_qty} ��� {new_qty}")




    def _start_edit_cell(self, row_id, col_index):
        """
        In-place editing of tree cells.
        ✅ Column 6 (expiry_date): Editable if adopted expiry, with validation
        ✅ Column 8 (qty_to_issue): Editable based on mode rules
        ✅ Single popup for all errors, keeps editor open
        """
        meta = self.row_data.get(row_id, {})
        if meta.get("is_header") or not meta.get("unique_id"):
            return
        
        vals = self.tree.item(row_id, "values")
        rtype = (vals[2] or "").lower()
        
        # ===== EXPIRY DATE EDITING (Column 6) =====
        if col_index == 6:
            if not meta.get("adopted_expiry"):
                self.status_var.set(
                    lang.t("out_kit.expiry_not_editable", "Expiry date is not editable (not adopted)")
                )
                return
            
            bbox = self.tree.bbox(row_id, f"#{col_index+1}")
            if not bbox:
                return
            
            x, y, w, h = bbox
            old_expiry = vals[6] or ""
            
            # Destroy existing editor
            if self.editing_cell:
                try:
                    self.editing_cell.destroy()
                except:
                    pass
                self.editing_cell = None
            
            # Create entry
            entry = tk.Entry(self.tree, font=("Helvetica", 10), background="#FFFBE0")
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, old_expiry)
            entry.focus()
            entry.select_range(0, tk.END)
            self.editing_cell = entry
            
            def cleanup():
                """Safely cleanup editor widget."""
                try:
                    if entry.winfo_exists():
                        entry.destroy()
                except:
                    pass
                self.editing_cell = None
            
            def save_expiry(_=None):
                """Save edited expiry date with validation."""
                # ✅ Define vals_list at the top (from parent scope)
                vals_list = list(vals)
                
                raw_input = entry.get().strip()
                
                # ===== CASE 1: Empty input (clear expiry) =====
                if not raw_input:
                    vals_list[6] = ""
                    self.tree.item(row_id, values=vals_list)
                    
                    # Mark as edited and clear adopted flag
                    if row_id in self.row_data:
                        self.row_data[row_id]["adopted_expiry"] = False
                        self.row_data[row_id]["expiry_edited"] = True
                        
                        # Build tag list: preserve type tags + add expiry_edited
                        tags = []
                        rtype = vals_list[2].lower() if len(vals_list) > 2 else ""
                        if rtype == "kit":
                            tags.append("kit_data")
                        elif rtype == "module":
                            tags.append("module_data")
                        tags.append("expiry_edited")
                        
                        self.tree.item(row_id, tags=tuple(tags))
                        logging.debug(f"[EDIT_EXPIRY] Cleared expiry, tags: {tags}")
                    
                    self.status_var.set(lang.t("out_kit.expiry_cleared", "✓ Expiry cleared"))
                    cleanup()
                    return
                
                # ===== CASE 2: Validate date format =====
                success, parsed_date, parse_error = self._parse_flexible_date(raw_input)
                
                if not success:
                    # Clear cell + show error + keep editor open
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self.parent,
                        lang.t("dialog_titles.error", "Error"),
                        parse_error,
                        "error"
                    )
                    entry.focus()
                    return
                
                # ===== CASE 3: Validate date is in future =====
                if not self._is_date_in_future(parsed_date):
                    # Clear cell + show error + keep editor open
                    entry.delete(0, tk.END)
                    entry.configure(background="#FFCCCC")
                    custom_popup(
                        self.parent,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("out_kit.date_must_be_future", 
                               "Expiry date must be in the future.\n\n"
                               "Entered date: {date}\n"
                               "Today: {today}",
                               date=parsed_date,
                               today=datetime.today().strftime('%Y-%m-%d')),
                        "error"
                    )
                    entry.focus()
                    return
                
                # ===== CASE 4: All validations passed - SAVE =====
                vals_list[6] = parsed_date
                self.tree.item(row_id, values=vals_list)
                
                # Mark as edited and clear adopted flag
                if row_id in self.row_data:
                    # Clear adopted flag since user manually edited
                    self.row_data[row_id]["adopted_expiry"] = False
                    self.row_data[row_id]["expiry_edited"] = True
                    
                    # ✅ Build tag list: preserve type tags + add expiry_edited
                    tags = []
                    
                    # Add type-specific tag (kit_data, module_data, or item_row)
                    rtype = vals_list[2].lower() if len(vals_list) > 2 else ""
                    if rtype == "kit":
                        tags.append("kit_data")
                    elif rtype == "module":
                        tags.append("module_data")
                    
                    # Add edited expiry tag (cyan background - applied last to take priority)
                    tags.append("expiry_edited")
                    
                    # Apply combined tags
                    self.tree.item(row_id, tags=tuple(tags))
                    
                    logging.debug(f"[EDIT_EXPIRY] Applied tags: {tags} to row {row_id}")
                    logging.info(f"[EDIT_EXPIRY] User edited: {meta.get('code')}: {old_expiry} → {parsed_date}")
                    
                    self.status_var.set(
                        lang.t("out_kit.expiry_updated", 
                               "✓ Expiry updated: {code} → {date}",
                               code=meta.get('code', ''), 
                               date=parsed_date)
                    )
                
                cleanup()
            
            entry.bind("<Return>", save_expiry)
            entry.bind("<KP_Enter>", save_expiry)
            entry.bind("<Tab>", save_expiry)
            entry.bind("<FocusOut>", save_expiry)
            entry.bind("<Escape>", lambda e: cleanup())
            return
        
        # ===== QUANTITY EDITING (Column 8) =====
        elif col_index == 8:
            rules = self.get_mode_rules()
            editable_lower = {t.lower() for t in rules["editable_types"]}
            
            if rtype not in editable_lower:
                return
            
            bbox = self.tree.bbox(row_id, f"#{col_index+1}")
            if not bbox:
                return
            
            x, y, w, h = bbox
            raw_old = vals[8]
            old_clean = raw_old[2:].strip() if raw_old.startswith("★") else raw_old.strip()
            
            # Destroy existing editor
            if self.editing_cell:
                try:
                    self.editing_cell.destroy()
                except:
                    pass
                self.editing_cell = None
            
            entry = tk.Entry(self.tree, font=("Helvetica", 10), background="#FFFBE0")
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, old_clean if old_clean else "")
            entry.focus()
            entry.select_range(0, tk.END)
            self.editing_cell = entry
            
            def save_qty(_=None):
                val = entry.get().strip()
                
                try:
                    stock = int(vals[5]) if vals[5] else 0
                except:
                    stock = 0
                
                if rtype in ("kit", "module"):
                    if val not in ("0", "1"):
                        val = "1" if stock > 0 else "0"
                    if stock == 0 and val == "1":
                        val = "0"
                else:
                    if not val.isdigit():
                        val = "0"
                    else:
                        iv = int(val)
                        if iv < 0:
                            iv = 0
                        if iv > stock:
                            iv = stock
                        val = str(iv)
                
                vals_list = list(vals)
                vals_list[8] = f"★ {val}"
                vals_list[11] = val
                self.tree.item(row_id, values=vals_list)
                
                try:
                    entry.destroy()
                except:
                    pass
                self.editing_cell = None
                
                if rtype == "kit" and rules.get("derive_modules_from_Kit"):
                    self._derive_modules_from_Kits()
                    if rules.get("derive_items_from_modules"):
                        self._derive_items_from_modules()
                    self._reapply_editable_icons(rules)
                elif rtype == "module" and rules.get("derive_items_from_modules"):
                    self._derive_items_from_modules()
                    self._reapply_editable_icons(rules)
            
            entry.bind("<Return>", save_qty)
            entry.bind("<KP_Enter>", save_qty)
            entry.bind("<Tab>", save_qty)
            entry.bind("<FocusOut>", save_qty)
            entry.bind("<Escape>", lambda e: (entry.destroy() if entry.winfo_exists() else None, setattr(self, "editing_cell", None)))
            return
        
        else:
            self.status_var.set(
                lang.t("out_kit.column_not_editable", "This column is not editable")
            )


    def _validate_date_format(self, date_str):
        """
        Validate date format (YYYY-MM-DD).
        
        Args:
            date_str: Date string to validate
        
        Returns:
            bool: True if valid format
        """
        if not date_str:
            return True  # Empty is allowed
        
        import re
        
        # Check basic format
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return False
        
        # Try parsing as actual date
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return True
        except ValueError:
            return False            

    # -------------------- Search --------------------
    def search_items(self, event=None):
        q = (self.search_var.get() or "").strip().lower()
        if q == "":
            self.populate_rows(self.full_items,
                               f"Showing {len(self.full_items)} rows (reset)")
            return
        if len(q) < self.search_min_chars:
            self.status_var.set(f"Type at least {self.search_min_chars} chars to search...")
            return
        filtered=[]
        for it in self.full_items:
            if q in (it['code'] or "").lower() or q in (it['description'] or "").lower():
                filtered.append(it)
        self.populate_rows(filtered, f"Found {len(filtered)} matching rows")

# ============================= out_kit.py (Part 5/6) =============================
    # -------------------- Transaction helpers --------------------
    def _insert_transaction_out(self, cur, *, unique_id, code, description,
                                expiry_date, batch_number, scenario,
                                kit, module, qty_out, out_type,
                                ts_date, ts_time, movement_type, document_number,
                                comment=None):
        """
        Insert OUT transaction record.
        ✅ Uses Kit/Module columns (not kit_number/module_number)
        ✅ Includes comments parameter for expiry remarks
        """
        try:
            cur.execute("""
                INSERT INTO stock_transactions
                (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
                 Scenario, Kit, Module,
                 Qty_Out, Out_Type, Movement_Type, document_number, comments)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                ts_date, ts_time, unique_id, code, description,
                expiry_date, batch_number, scenario,
                kit or "",  # ✅ Use kit (code), not kit_number (instance)
                module or "",  # ✅ Use module (code), not module_number (instance)
                qty_out, out_type, movement_type, document_number,
                comment
            ))
            logging.debug(f"[OUT_TRANSACTION] Logged OUT: {code} qty={qty_out} kit={kit} module={module}")
        except sqlite3.Error as e:
            logging.error(f"[OUT_TRANSACTION] Failed: {e}")
            raise

    def _insert_transaction_in_mirror(self, cur, *, unique_id, code, description,
                                      expiry_date, batch_number, scenario,
                                      kit, module, qty_in,
                                      out_type_as_in_type, ts_date, ts_time,
                                      movement_type, document_number,
                                      comment=None):
        """
        Insert mirror IN transaction record.
        ✅ Uses Kit/Module columns (not kit_number/module_number)
        ✅ Includes comments parameter for expiry remarks
        """
        try:
            cur.execute("""
                INSERT INTO stock_transactions
                (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
                 Scenario, Kit, Module,
                 Qty_IN, IN_Type, Movement_Type, document_number, comments)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                ts_date, ts_time, unique_id, code, description,
                expiry_date, batch_number, scenario,
                kit or "",  # ✅ Use kit (code), not kit_number (instance)
                module or "",  # ✅ Use module (code), not module_number (instance)
                qty_in, out_type_as_in_type, movement_type, document_number,
                comment
            ))
            logging.debug(f"[IN_TRANSACTION] Logged IN: {code} qty={qty_in} kit={kit} module={module}")
        except sqlite3.Error as e:
            logging.error(f"[IN_TRANSACTION] Failed: {e}")
            raise

    # -------------------- Document Number --------------------
    def generate_document_number(self, out_type_text: str) -> str:
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").upper()
        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/BRK"
        conn = connect_db()
        serial = 1
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("""
                    SELECT document_number FROM stock_transactions
                     WHERE document_number LIKE ?
                     ORDER BY document_number DESC LIMIT 1
                """,(prefix+"/%",))
                r = cur.fetchone()
                if r and r[0]:
                    tail = r[0].rsplit('/',1)[-1]
                    if tail.isdigit():
                        serial = int(tail)+1
            finally:
                cur.close()
                conn.close()
        doc = f"{prefix}/{serial:04d}"
        self.current_document_number = doc
        return doc

    #----------------------Adopted Expiries WARNING pop-up--------

    def _show_adopted_expiry_warning(self):
        """
        Show warning dialog if any items have adopted expiries.
        Returns True to proceed with save, False to go back for review.
        
        Returns:
            bool: True = proceed with save, False = cancel and review
        """
        # Count items with adopted expiries
        adopted_count = 0
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("adopted_expiry") and not meta.get("is_header"):
                adopted_count += 1
        
        # If no adopted expiries, proceed directly
        if adopted_count == 0:
            logging.debug("[SAVE] No adopted expiries found, proceeding with save")
            return True
        
        # ✅ Show warning dialog
        logging.info(f"[SAVE] Found {adopted_count} items with adopted expiries, showing warning")
        
        # Create custom dialog
        dialog = tk.Toplevel(self.parent)
        dialog.title(lang.t("out_kit.warning_title", "Warning - Adopted Expiries"))
        dialog.geometry("550x550")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Result variable
        result = {"proceed": False}
        
        # Main frame
        main_frame = tk.Frame(dialog, bg="#FFF3E0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Warning icon and title
        title_frame = tk.Frame(main_frame, bg="#FFF3E0")
        title_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            title_frame,
            text="⚠️",
            font=("Helvetica", 32),
            bg="#FFF3E0",
            fg="#FF6F00"
        ).pack(side="left", padx=(0, 10))
        
        tk.Label(
            title_frame,
            text=lang.t("out_kit.adopted_expiry_warning_title", "Adopted Expiry Dates Detected"),
            font=("Helvetica", 14, "bold"),
            bg="#FFF3E0",
            fg="#E65100"
        ).pack(side="left")
        
        # Warning message frame
        msg_frame = tk.Frame(main_frame, bg="white", relief="solid", borderwidth=1)
        msg_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Count label
        count_label = tk.Label(
            msg_frame,
            text=lang.t(
                "out_kit.adopted_count",
                "{count} item(s) have expiry dates adopted from kit/module.",
                count=adopted_count
            ),
            font=("Helvetica", 11, "bold"),
            bg="white",
            fg="#D84315",
            wraplength=480,
            justify="left"
        )
        count_label.pack(fill="x", padx=15, pady=(15, 10))
        
        # Warning text
        warning_text = lang.t(
            "out_kit.adopted_expiry_warning",
            "This will cause issues in stock management for on-shelf items.\n\n"
            "It is STRONGLY RECOMMENDED to check the expiries of all items manually "
            "and enter verified expiries for all the RED highlighted rows before saving.\n\n"
            "Items with adopted expiries are marked with a red/pink background in the table."
        )
        
        warning_label = tk.Label(
            msg_frame,
            text=warning_text,
            font=("Helvetica", 10),
            bg="white",
            fg="#424242",
            wraplength=480,
            justify="left"
        )
        warning_label.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Recommendation box
        rec_frame = tk.Frame(main_frame, bg="#E3F2FD", relief="solid", borderwidth=1)
        rec_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            rec_frame,
            text="💡 " + lang.t("out_kit.recommendation", "Recommendation"),
            font=("Helvetica", 10, "bold"),
            bg="#E3F2FD",
            fg="#1565C0",
            anchor="w"
        ).pack(fill="x", padx=10, pady=(8, 4))
        
        tk.Label(
            rec_frame,
            text=lang.t(
                "out_kit.recommendation_text",
                "Click 'Review' to go back and verify expiry dates.\n"
                "Double-click any RED highlighted row to edit the expiry date."
            ),
            font=("Helvetica", 9),
            bg="#E3F2FD",
            fg="#424242",
            anchor="w",
            justify="left"
        ).pack(fill="x", padx=10, pady=(0, 8))
        
        # Button frame
        btn_frame = tk.Frame(main_frame, bg="#FFF3E0")
        btn_frame.pack(side="bottom", pady=(10, 0))
        
        def on_review():
            """User wants to go back and review."""
            result["proceed"] = False
            dialog.destroy()
        
        def on_save():
            """User wants to proceed with save."""
            result["proceed"] = True
            dialog.destroy()
        
        # Review button (recommended action)
        review_btn = tk.Button(
            btn_frame,
            text=lang.t("out_kit.review_button", "📋 Review Expiries"),
            font=("Helvetica", 11, "bold"),
            bg="#2196F3",
            fg="white",
            width=18,
            command=on_review,
            relief="flat",
            cursor="hand2",
            padx=15,
            pady=8
        )
        review_btn.pack(side="left", padx=5)
        
        # Save anyway button (warning action)
        save_btn = tk.Button(
            btn_frame,
            text=lang.t("out_kit.save_anyway_button", "⚠️ Save Anyway"),
            font=("Helvetica", 11),
            bg="#FF9800",
            fg="white",
            width=18,
            command=on_save,
            relief="flat",
            cursor="hand2",
            padx=15,
            pady=8
        )
        save_btn.pack(side="left", padx=5)
        
        # Bind keys
        dialog.bind("<Escape>", lambda e: on_review())
        dialog.bind("<Return>", lambda e: on_review())  # Default = Review
        
        # Wait for dialog
        dialog.wait_window()
        
        logging.info(f"[SAVE] User chose: {'PROCEED' if result['proceed'] else 'REVIEW'}")
        return result["proceed"]    




    # -------------------- Save (Break logic) --------------------
    def save_all(self):
        """
        Save/issue all items with dual transaction logging (OUT + mirror IN).
        ✅ Converts in-box items (>6 slashes) to on-shelf items (≤6 slashes)
        ✅ Dual transaction: OUT (reduce in-box) + IN (add to on-shelf)
        ✅ Full localization with custom popups
        ✅ Stock validation (qty_out <= current_stock)
        ✅ Proper error handling with retry logic
        ✅ Comments in stock_transactions with expiry remarks
        """
        logging.info("[OUT_KIT] save_all called")
        
        # Role validation
        if self.role not in ["admin", "manager", "supervisor"]:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("out_kit.no_permission", "Only admin or manager roles can save changes."),
                "error"
            )
            return
        
        # Check for adopted expiries and show warning
        if not self._show_adopted_expiry_warning():
            # User chose to review - don't save, just return
            self.status_var.set(
                lang.t("out_kit.save_cancelled_review", "Save cancelled - please review expiry dates")
            )
            logging.info("[SAVE] User cancelled save to review adopted expiries")
            return
        # ===== Continue with existing save logic =====
        logging.info("[SAVE] Starting save process...")

        # Fixed OUT type
        out_type = OUT_TYPE_CANONICAL
        
        # Collect rows to process
        rows = []
        
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            
            # Unpack all 12 columns
            if len(vals) < 12:
                logging.warning(f"[OUT_KIT] Row {iid} has only {len(vals)} columns, skipping")
                continue
            
            (code, desc, tfield, kit_col, module_col,
             current_stock, exp_date, batch_no, qty_to_issue,
             unique_id, line_id, qty_in_hidden) = vals
            
            meta = self.row_data.get(iid, {})
            
            # Skip header rows
            if meta.get("is_header"):
                continue
            
            # Parse qty_to_issue (remove star if present)
            raw_q = str(qty_to_issue).replace("★", "").strip()
            if not raw_q or not raw_q.isdigit():
                continue
            
            q_out = int(raw_q)
            if q_out <= 0:
                continue
            
            # ✅ VALIDATION: Must be in-box item
            if not self.is_inbox_item(unique_id):
                logging.warning(f"[OUT_KIT] Skipping non-in-box item: {unique_id}")
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.warning", "Warning"),
                    lang.t("out_kit.not_inbox",
                           "Item {code} is not from in-box (unique_id has ≤6 slashes). Skipping.",
                           code=code),
                    "warning"
                )
                continue
            
            # Parse current stock
            try:
                stock_str = str(current_stock).replace("★", "").strip()
                stock_int = int(stock_str) if stock_str.isdigit() else 0
            except:
                stock_int = 0
            
            # ✅ VALIDATION: qty_out cannot exceed stock
            if q_out > stock_int:
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("out_kit.qty_exceeds_stock",
                           "Item {code}: Quantity to out ({qty}) exceeds available stock ({stock}).",
                           code=code, qty=q_out, stock=stock_int),
                    "warning"
                )
                return
            
            # ✅ Convert in-box unique_id to on-shelf unique_id
            unique_id_onshelf = self.convert_inbox_to_onshelf(unique_id)
            
            # Parse qty_in (mirror quantity)
            try:
                q_in = int(qty_in_hidden) if str(qty_in_hidden).isdigit() else q_out
            except:
                q_in = q_out

            # ✅ CHECK: Adopted expiry?
            is_adopted = meta.get("adopted_expiry", False)
            expiry_edited = meta.get("expiry_edited", False)
            original_comment = meta.get("original_comment", "")    
            
            # Collect row data
            rows.append({
                "code": code,
                "desc": desc,
                "type": tfield,
                "stock": stock_int,
                "qty_out": q_out,
                "qty_in": q_in,
                "exp_date": exp_date if exp_date else None,
                "exp_date_adopted": is_adopted,
                "expiry_edited": expiry_edited,
                "original_comment": original_comment,
                "batch_no": batch_no if batch_no else None,
                "unique_id_inbox": unique_id,
                "unique_id_onshelf": unique_id_onshelf,
                "line_id": line_id if line_id else None,
                "kit": kit_col if kit_col != "-----" else None,
                "module": module_col if module_col != "-----" else None,
                "kit_number": meta.get("Kit_number"),
                "module_number": meta.get("module_number")
            })
        
        # Check if any items to process
        if not rows:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("out_kit.no_items_to_process", "No quantities entered to process."),
                "error"
            )
            return
        
        # Get scenario and movement type
        scenario_name = self.scenario_map.get(self.selected_scenario_id, "")
        movement_label = self.mode_var.get()
        movement_canonical = self._canon_movement_type(movement_label)
        
        # Generate document number
        doc_number = self.generate_document_number(out_type)
        
        self.status_var.set(
            lang.t("out_kit.processing", "Processing... Document Number: {doc}", doc=doc_number)
        )
        
        # Process with retry logic
        import time
        max_attempts = 4
        
        for attempt in range(1, max_attempts + 1):
            conn = connect_db()
            if conn is None:
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("out_kit.db_connection_failed", "Database connection failed."),
                    "error"
                )
                return
            
            try:
                conn.execute("PRAGMA busy_timeout=5000;")
                cur = conn.cursor()
                now_date = datetime.today().strftime('%Y-%m-%d')
                now_time = datetime.now().strftime('%H:%M:%S')
                
                # Process each row
                for r in rows:
                    # ===== 1️⃣ OUT TRANSACTION (in-box) =====
                    
                    # Verify in-box stock availability
                    cur.execute("""
                        SELECT final_qty FROM stock_data WHERE unique_id = ?
                    """, (r["unique_id_inbox"],))
                    
                    db_row = cur.fetchone()
                    if not db_row or db_row[0] is None or db_row[0] < r["qty_out"]:
                        raise ValueError(
                            lang.t("out_kit.insufficient_stock",
                                   "Insufficient stock or concurrency issue for {code}",
                                   code=r["code"])
                        )
                    
                    # Update in-box stock_data: Increase qty_out
                    cur.execute("""
                        UPDATE stock_data
                        SET qty_out = qty_out + ?,
                            updated_at = ?
                        WHERE unique_id = ?
                          AND (qty_in - qty_out) >= ?
                    """, (r["qty_out"], f"{now_date} {now_time}", r["unique_id_inbox"], r["qty_out"]))
                    
                    if cur.rowcount == 0:
                        raise ValueError(
                            lang.t("out_kit.concurrent_change",
                                   "Concurrent change or insufficient stock for {code}",
                                   code=r["code"])
                        )
                    
                    # ✅ Build comment for OUT transaction
                    out_comment = f"OUT from in-box {r['kit_number']}/{r['module_number']}"
                    if r['exp_date_adopted'] and not r['expiry_edited']:
                        out_comment += f" | ⚠️ Adopted expiry: {r['exp_date']} | Original: {r['original_comment'][:50] if r['original_comment'] else 'N/A'}"
                    elif r['exp_date_adopted'] and r['expiry_edited']:
                        out_comment += f" | ✓ Expiry verified/updated by user to {r['exp_date']}"
                    
                    # Log OUT transaction (in-box)
                    self._insert_transaction_out(
                        cur,
                        unique_id=r["unique_id_inbox"],
                        code=r["code"],
                        description=r["desc"],
                        expiry_date=r["exp_date"],
                        batch_number=r["batch_no"],
                        scenario=scenario_name,
                        kit=r["kit"],  
                        module=r["module"], 
                        qty_out=r["qty_out"],
                        out_type=out_type,
                        ts_date=now_date,
                        ts_time=now_time,
                        movement_type=movement_canonical,
                        document_number=doc_number,
                        comment=out_comment  # ✅ PASS COMMENT
                    )
                    
                    logging.info(f"[OUT_KIT] OUT processed: {r['code']} qty={r['qty_out']} from in-box")
                    
                    # ===== 2️⃣ IN TRANSACTION (on-shelf) =====
                    
                    if r["qty_in"] > 0:
                        # Check if on-shelf unique_id already exists
                        cur.execute("""
                            SELECT line_id, qty_in, qty_out
                            FROM stock_data
                            WHERE unique_id = ?
                        """, (r["unique_id_onshelf"],))
                        
                        onshelf_existing = cur.fetchone()
                        
                        if onshelf_existing:
                            # Update existing on-shelf record
                            cur.execute("""
                                UPDATE stock_data
                                SET qty_in = qty_in + ?,
                                    updated_at = ?
                                WHERE unique_id = ?
                            """, (r["qty_in"], f"{now_date} {now_time}", r["unique_id_onshelf"]))
                            
                            logging.info(f"[OUT_KIT] Updated existing on-shelf: {r['unique_id_onshelf']}")
                        
                        else:
                            # ✅ Create new on-shelf record - determine expiry, comments
                            if r["exp_date_adopted"] and not r["expiry_edited"]:
                                # ✅ ADOPTED + NOT EDITED: Keep expiry but warn
                                onshelf_exp_date = r["exp_date"]
                                onshelf_comment = f"⚠️ ADOPTED EXPIRY - Verify before use | Original: {r['original_comment'][:80] if r['original_comment'] else 'Unknown'}"
                                in_comment = f"IN to on-shelf | ⚠️ Adopted expiry {r['exp_date']} needs verification"
                                logging.warning(f"[OUT_KIT] Preserved adopted expiry (not edited): {r['code']} = {r['exp_date']}")
                            
                            elif r["exp_date_adopted"] and r["expiry_edited"]:
                                # ✅ ADOPTED + EDITED: Use new expiry, mark as verified
                                onshelf_exp_date = r["exp_date"]
                                onshelf_comment = f"Expiry verified/updated during break from {r['kit_number']}/{r['module_number']}"
                                in_comment = f"IN to on-shelf | Expiry verified by user: {r['exp_date']}"
                                logging.info(f"[OUT_KIT] User verified expiry: {r['code']} = {r['exp_date']}")
                            
                            else:
                                # ✅ NOT ADOPTED: Normal expiry
                                onshelf_exp_date = r["exp_date"]
                                onshelf_comment = f"Moved from in-box {r['kit_number']}/{r['module_number']}"
                                in_comment = f"IN to on-shelf from {r['kit_number']}/{r['module_number']}"
                            
                            cur.execute("""
                                INSERT INTO stock_data
                                (unique_id, scenario, kit, module, item, std_qty,
                                 qty_in, qty_out, exp_date, 
                                 kit_number, module_number, management_mode, comments, updated_at)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, NULL, NULL, 'on-shelf', ?, ?)
                            """, (
                                r["unique_id_onshelf"],
                                scenario_name,
                                r["kit"],
                                r["module"],
                                r["code"],
                                r.get("std_qty"),
                                r["qty_in"],
                                onshelf_exp_date,
                                onshelf_comment,  # ✅ Smart comment
                                f"{now_date} {now_time}"
                            ))
                            
                            logging.info(f"[OUT_KIT] Created on-shelf: {r['unique_id_onshelf']} (adopted={r['exp_date_adopted']}, edited={r['expiry_edited']})")
                        
                        # ✅ Build comment for IN transaction (use in_comment if defined, else fallback)
                        if not onshelf_existing:
                            # Use the in_comment we built above
                            pass  # in_comment already defined
                        else:
                            # For existing records, build a simple comment
                            in_comment = f"IN to on-shelf (updated existing) from {r['kit_number']}/{r['module_number']}"
                        
                        # Log IN transaction (on-shelf)
                        self._insert_transaction_in_mirror(
                            cur,
                            unique_id=r["unique_id_onshelf"],
                            code=r["code"],
                            description=r["desc"],
                            expiry_date=onshelf_exp_date if not onshelf_existing else r["exp_date"],
                            batch_number=r["batch_no"],
                            scenario=scenario_name,
                            kit=r["kit"],
                            module=r["module"],
                            qty_in=r["qty_in"],
                            out_type_as_in_type="Internal move to on-shelf",
                            ts_date=now_date,
                            ts_time=now_time,
                            movement_type=movement_canonical,
                            document_number=doc_number,
                            comment=in_comment  # ✅ PASS SMART COMMENT
                        )
                        
                        logging.info(f"[OUT_KIT] IN processed: {r['code']} qty={r['qty_in']} to on-shelf")
                
                # Commit transaction
                conn.commit()
                
                # Success message
                total_transactions = len(rows) * 2
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.success", "Success"),
                    lang.t("out_kit.break_complete",
                           "Break complete. Logged {count} transactions (OUT from in-box + IN to on-shelf).",
                           count=total_transactions),
                    "info"
                )
                
                self.status_var.set(
                    lang.t("out_kit.break_complete_doc",
                           "Break complete. Document Number: {doc}",
                           doc=doc_number)
                )
                
                # Ask for export
                if custom_askyesno(
                    self.parent,
                    lang.t("dialog_titles.confirm", "Confirm"),
                    lang.t("out_kit.ask_export", "Export the break operation to Excel?")
                ) == "yes":
                    self.export_data(rows)
                
                # Clear form
                self.clear_form()
                return
                
            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower() and attempt < max_attempts:
                    logging.warning(f"[OUT_KIT] Database locked attempt {attempt}/{max_attempts}, retrying...")
                    try:
                        conn.rollback()
                    except:
                        pass
                    time.sleep(0.8 * attempt)
                    continue
                else:
                    try:
                        conn.rollback()
                    except:
                        pass
                    logging.error(f"[OUT_KIT] Break failed: {e}")
                    custom_popup(
                        self.parent,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t("out_kit.break_failed", "Break failed: {error}", error=str(e)),
                        "error"
                    )
                    return
                    
            except Exception as e:
                try:
                    conn.rollback()
                except:
                    pass
                logging.error(f"[OUT_KIT] Break failed: {e}")
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("out_kit.break_failed", "Break failed: {error}", error=str(e)),
                    "error"
                )
                return
                
            finally:
                try:
                    cur.close()
                except:
                    pass
                try:
                    conn.close()
                except:
                    pass
        
        # Max attempts reached
        custom_popup(
            self.parent,
            lang.t("dialog_titles.error", "Error"),
            lang.t("out_kit.break_failed_locked", "Break failed: database remained locked."),
            "error"
        )

    # -------------------- Utility / Clear / Export --------------------
    def clear_search(self):
        self.search_var.set("")
        self.populate_rows(self.full_items,
                           f"Showing {len(self.full_items)} rows (reset)")

    def clear_form(self):
        self.clear_table_only()
        self.scenario_var.set("")
        self.mode_var.set("")
        self.Kit_var.set("")
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        self.out_type_var.set(lang.t("out_kit.out_type_fixed", "Internal move from in-box items"))
        self.status_var.set(lang.t("break_kit.ready","Ready"))
        self.scenario_map = self.fetch_scenario_map()
        self.load_scenarios()

# ============================= out_kit.py (Part 6/6) =============================
    def export_data(self, rows_processed=None):
        if self.parent is None or not self.parent.winfo_exists():
            return
        try:
            export_rows = []
            for iid in self.tree.get_children():
                vals = self.tree.item(iid,"values")
                if not vals or len(vals) < 12: continue
                (code, desc, tfield, kit_col, module_col,
                 current_stock, exp_date, batch_no, qty_to_issue,
                 unique_id, line_id, qty_in_hidden) = vals
                raw_q = qty_to_issue[2:].strip() if qty_to_issue.startswith("★") else qty_to_issue
                qty_out = int(raw_q) if raw_q.isdigit() else 0
                qty_in = int(qty_in_hidden) if qty_in_hidden.isdigit() else qty_out
                meta = self.row_data.get(iid,{})
                export_rows.append({
                    "code": code,
                    "description": desc,
                    "type": tfield,
                    "kit_number": meta.get("Kit_number") or kit_col or "",
                    "module_number": meta.get("module_number") or module_col or "",
                    "current_stock": int(current_stock) if str(current_stock).isdigit() else 0,
                    "expiry_date": exp_date or "",
                    "batch_number": batch_no or "",
                    "qty_out": qty_out,
                    "qty_in": qty_in
                })
            if rows_processed:
                # Optionally refine; left as-is (already correct).
                pass
            if not any(r["qty_out"]>0 for r in export_rows):
                custom_popup(self.parent,"Error","No break quantities to export.","error")
                return

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            movement_type_raw = self.mode_var.get() or "Break Kit"
            scenario_name = self.selected_scenario_name or "N/A"
            doc_number = getattr(self, "current_document_number", None)
            out_type_raw = OUT_TYPE_CANONICAL

            def sanitize(s):
                import re
                s = re.sub(r'[^A-Za-z0-9]+','_', s)
                s = re.sub(r'_+','_', s)
                return s.strip('_') or "Unknown"

            movement_slug = sanitize(movement_type_raw)
            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)
            file_name = f"Break_{movement_slug}_{current_time.replace(':','-')}.xlsx"
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files","*.xlsx")],
                initialfile=file_name,
                initialdir=default_dir
            )
            if not path:
                self.status_var.set("Export cancelled")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws_title_base = "Break Kit-Module"
            ws.title = ws_title_base[:31]

            if doc_number:
                ws['A1'] = f"Date: {current_time}{' '*8}Document Number: {doc_number}"
            else:
                ws['A1'] = f"Date: {current_time}"
            ws['A1'].font = Font(name="Calibri", size=11)
            project_name, project_code = fetch_project_details()
            ws['A2'] = f"{ws_title_base} – Movement: {movement_type_raw}"
            ws['A2'].font = Font(name="Tahoma", size=14, bold=True)
            ws['A2'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A2:J2')
            ws['A3'] = f"{project_name} - {project_code}"
            ws['A3'].font = Font(name="Tahoma", size=14, bold=True)
            ws['A3'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A3:J3')
            ws['A4'] = f"OUT Type: {out_type_raw}"
            ws['A4'].font = Font(name="Tahoma", size=12, bold=True)
            ws['A4'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A4:J4')
            ws['A5'] = f"Scenario: {scenario_name}"
            ws['A5'].font = Font(name="Tahoma", size=12, bold=True)
            ws['A5'].alignment = Alignment(horizontal="right")
            ws.merge_cells('A5:J5')
            ws.append([])

            headers = ["Code","Description","Type","Kit Number","Module Number",
                       "Current Stock","Expiry Date","Batch Number",
                       "Qty Broken (Out)","Qty In (Mirror)"]
            ws.append(headers)
            for c in range(1,len(headers)+1):
                ws.cell(row=7,column=c).font = Font(name="Tahoma", size=11, bold=True)

            from openpyxl.styles import PatternFill
            for row_idx, r in enumerate(export_rows, start=8):
                ws.append([
                    r["code"], r["description"], r["type"], r["kit_number"],
                    r["module_number"], r["current_stock"], r["expiry_date"],
                    r["batch_number"], r["qty_out"], r["qty_in"]
                ])
                rtype = (r["type"] or "").lower()
                for col in range(1,len(headers)+1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(name="Calibri", size=11, bold=(rtype in ("kit","module")))
                    if rtype=="kit":
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif rtype=="module":
                        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Autofit
            for col in ws.columns:
                max_len=0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        ln = len(str(cell.value))
                        if ln>max_len: max_len=ln
                ws.column_dimensions[letter].width = min(max_len+2, 50)

            wb.save(path)
            custom_popup(self.parent,"Success",f"Exported to {path}","info")
            self.status_var.set(f"Export successful: {path}")
        except Exception as e:
            logging.error(f"[BREAK] Export failed: {e}")
            custom_popup(self.parent,"Error",f"Export failed: {e}","error")


# -------------------- Main Harness --------------------
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Break Kit/Module")
    app = type("App", (), {})()
    app.role = "admin"
    StockOutKit(root, app, role="admin")
    root.geometry("1400x850")
    root.mainloop()
                                                   