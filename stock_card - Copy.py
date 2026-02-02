import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import sqlite3
from db import connect_db
from language_manager import lang
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.page import PrintPageSetup
from popup_utils import custom_popup, custom_askyesno, custom_dialog
import os


def get_active_designation(code: str) -> str:
    """
    Fetch the active designation for a given code based on lang.lang_code.
    Uses designation_[lang_code], falls back to designation_en, then designation.
    """
    conn = connect_db()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT designation, designation_en, designation_fr, designation_sp FROM items_list WHERE code = ?",
            (code,))
        row = cursor.fetchone()
        if not row:
            return lang.t("stock_card.no_description", "No Description")
        lang_code = lang.lang_code.lower()
        mapping = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp"
        }
        active_col = mapping.get(lang_code, "designation_en")
        if row[active_col]:
            return row[active_col]
        if row["designation_en"]:
            return row["designation_en"]
        return row["designation"] if row["designation"] else lang.t("stock_card.no_description", "No Description")
    finally:
        cursor.close()
        conn.close()


def get_active_item_type(code: str) -> str:
    """
    Fetch the item type from items_list for a given code.
    """
    conn = connect_db()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT type FROM items_list WHERE code = ?", (code,))
        row = cursor.fetchone()
        return row['type'].upper() if row and row['type'] else lang.t("stock_card.unknown", "Unknown")
    finally:
        cursor.close()
        conn.close()


def fetch_kit_numbers(scenario_name: str | None = None) -> list:
    """
    Fetch distinct kit_number values from stock_data.
    Optionally filter by scenario (scenario_name, not ID) if provided and not 'All Scenarios'.
    """
    conn = connect_db()
    if conn is None:
        return []
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        base = """
            SELECT DISTINCT kit_number
              FROM stock_data
             WHERE kit_number IS NOT NULL
               AND kit_number != 'None'
        """
        params = []
        if scenario_name and scenario_name != lang.t("stock_card.all_scenarios", "All Scenarios"):
            base += " AND scenario = ?"
            params.append(scenario_name)
        base += " ORDER BY kit_number"
        cur.execute(base, params)
        return [r["kit_number"] for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()


def fetch_module_numbers(scenario_name: str | None = None,
                         kit_number: str | None = None) -> list:
    """
    Fetch distinct module_number values from stock_data.
    Optionally filter by scenario and kit_number (in that order).
    """
    conn = connect_db()
    if conn is None:
        return []
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        base = """
            SELECT DISTINCT module_number
              FROM stock_data
             WHERE module_number IS NOT NULL
               AND module_number != 'None'
        """
        params = []
        if scenario_name and scenario_name != lang.t("stock_card.all_scenarios", "All Scenarios"):
            base += " AND scenario = ?"
            params.append(scenario_name)
        if kit_number and kit_number != lang.t("stock_card.all_kits", "All Kits"):
            base += " AND kit_number = ?"
            params.append(kit_number)
        base += " ORDER BY module_number"
        cur.execute(base, params)
        return [r["module_number"] for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()


class StockCard(tk.Frame):
    def __init__(self, parent: tk.Widget, root: tk.Tk, role: str = "supervisor") -> None:
        super().__init__(parent)
        self.root = root
        self.role = role.lower()

        self.scenario_map = self.fetch_scenario_map()
        self.kit_numbers = fetch_kit_numbers()
        self.module_numbers = fetch_module_numbers()

        # One-line item info shown on the same line as the search bar
        self.item_info_var = tk.StringVar(value="")

        self.status_var = tk.StringVar(value=lang.t("stock_card.ready", "Ready"))
        self.pack(fill="both", expand=True)
        self.render_ui()

    def fetch_scenario_map(self) -> dict:
        """
        Fetch scenario_id to name mapping from scenarios table.
        """
        conn = connect_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT scenario_id, name FROM scenarios")
            return {str(row['scenario_id']): row['name'] for row in cursor.fetchall()}
        finally:
            cursor.close()
            conn.close()

    def fetch_search_results(self, query: str) -> list:
        """
        Fetch items matching the search query, filtered by scenario, kit_number, and module_number.
        """
        conn = connect_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        lang_code = lang.lang_code.lower()
        mapping = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp"
        }
        active_designation = mapping.get(lang_code, "designation_en")
        query_sql = f"""
            SELECT DISTINCT s.code
            FROM stock_transactions s
            LEFT JOIN items_list i ON s.code = i.code
            WHERE (
                LOWER(s.code) LIKE ?
                OR LOWER(i.{active_designation}) LIKE ?
                OR (i.{active_designation} IS NULL AND LOWER(i.designation_en) LIKE ?)
                OR (i.{active_designation} IS NULL AND i.designation_en IS NULL AND LOWER(i.designation) LIKE ?)
            )
        """
        params = (f"%{query.lower()}%", f"%{query.lower()}%", f"%{query.lower()}%", f"%{query.lower()}%")
        if self.scenario_var.get() != lang.t("stock_card.all_scenarios", "All Scenarios"):
            query_sql += " AND s.Scenario = ?"
            params = params + (self.scenario_var.get(),)
        if self.kit_var.get() != lang.t("stock_card.all_kits", "All Kits"):
            query_sql += " AND s.Kit = ?"
            params = params + (self.kit_var.get(),)
        if self.module_var.get() != lang.t("stock_card.all_modules", "All Modules"):
            query_sql += " AND s.Module = ?"
            params = params + (self.module_var.get(),)
        query_sql += " ORDER BY s.code"
        try:
            cursor.execute(query_sql, params)
            results = cursor.fetchall()
            return [{'code': row['code'], 'description': get_active_designation(row['code'])} for row in results]
        except Exception as e:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_card.search_error", "Search failed: {error}").format(error=str(e)),
                         "error")
            return []
        finally:
            cursor.close()
            conn.close()

    # ---------------- Localization helpers ----------------
    def _to_display_in_type(self, canonical: str) -> str:
        if not canonical:
            return ""
        return lang.enum_to_display("stock_in.in_types_map", canonical, fallback=canonical)

    def _to_display_out_type(self, canonical: str) -> str:
        if not canonical:
            return ""
        fallback = canonical.replace("_", " ").title()
        return lang.enum_to_display("stock_out.out_types_map", canonical, fallback=fallback)

    def _to_display_remarks(self, text: str) -> str:
        if text is None:
            return ""
        return lang.enum_to_display("stock_transactions.remarks_map", text, fallback=text)

    def _to_display_comments(self, text: str) -> str:
        if text is None:
            return ""
        return lang.enum_to_display("stock_transactions.comments_map", text, fallback=text)

    def _combine_remarks_and_comments(self, remarks: str | None, comments: str | None) -> str:
        r = self._to_display_remarks(remarks)
        c = self._to_display_comments(comments)

        r = (r or "").strip()
        c = (c or "").strip()

        if r and c:
            return f"{r}, {c}"
        return r or c or ""

    # ---------------- Document window ----------------
    def _get_stock_transactions_columns(self) -> list[str]:
        conn = connect_db()
        if conn is None:
            return []
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_transactions)")
            return [r[1] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()

    def _auto_fit_tree_columns(self, tree: ttk.Treeview, max_width: int = 420, padding: int = 24) -> None:
        cols = tree["columns"]
        for col in cols:
            header_text = tree.heading(col, "text") or str(col)
            max_len = len(str(header_text))

            for iid in tree.get_children():
                val = tree.set(iid, col)
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))

            width = min(max_width, max(80, (max_len * 7) + padding))
            tree.column(col, width=width, stretch=True)

    def open_document_window(self, document_number: str) -> None:
        doc = (document_number or "").strip()
        if not doc:
            custom_popup(self, lang.t("dialog_titles.info", "Info"),
                         lang.t("stock_card.no_document_number", "No document number for this row."),
                         "info")
            return

        active_code = (self.code_entry.get() or "").strip()

        win = tk.Toplevel(self)
        win.title(lang.t("stock_card.document_details_title", "Document Details") + f" - {doc}")
        win.configure(bg="#F5F5F5")
        win.geometry("1200x700")
        win.minsize(900, 500)
        try:
            win.transient(self.winfo_toplevel())
        except Exception:
            pass

        header = tk.Frame(win, bg="#F5F5F5")
        header.pack(fill="x", padx=10, pady=(10, 5))
        tk.Label(
            header,
            text=lang.t("stock_card.document_details_title", "Document Details"),
            font=("Helvetica", 16, "bold"),
            bg="#F5F5F5",
            fg="#2C3E50",
            anchor="w"
        ).pack(side="left", fill="x", expand=True)

        right_header = tk.Frame(header, bg="#F5F5F5")
        right_header.pack(side="right")

        # We'll fill these after loading the first row (Date/Time)
        doc_lbl = tk.Label(
            right_header,
            text=f"{lang.t('stock_card.document_number_label', 'Document Number')}: {doc}",
            font=("Helvetica", 11),
            bg="#F5F5F5",
            fg="#2C3E50",
            anchor="e"
        )
        doc_lbl.pack(anchor="e")

        dt_lbl = tk.Label(
            right_header,
            text="",
            font=("Helvetica", 10),
            bg="#F5F5F5",
            fg="#2C3E50",
            anchor="e"
        )
        dt_lbl.pack(anchor="e")

        tree_frame = tk.Frame(win, bg="#F5F5F5")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        all_columns = self._get_stock_transactions_columns()
        if not all_columns:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_card.document_fetch_error", "Failed to load document rows: {error}")
                         .format(error="schema not found"),
                         "error")
            win.destroy()
            return

        # Hide: unique_id, document_number, Date, Time
        hidden_cols = {"unique_id", "document_number", "date", "time"}
        visible_columns = [c for c in all_columns if c.lower() not in hidden_cols]

        tree = ttk.Treeview(tree_frame, columns=visible_columns, show="headings")
        tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        tree.configure(xscrollcommand=hsb.set)

        tree.tag_configure("active_code", background="#FFF3CD")  # highlight for selected code

        for col in visible_columns:
            key = f"stock_transactions.{col.lower()}"
            tree.heading(col, text=lang.t(key, col))
            tree.column(col, width=120, stretch=True)

        conn = connect_db()
        if conn is None:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_card.document_fetch_error", "Failed to load document rows: {error}")
                         .format(error="Database connection failed"),
                         "error")
            win.destroy()
            return

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT * FROM stock_transactions WHERE document_number = ? ORDER BY Date, Time",
                (doc,)
            )
            rows = cur.fetchall()

            # Show Date/Time once in header (from first row if exists)
            if rows:
                try:
                    d = rows[0]["Date"]
                except Exception:
                    d = ""
                try:
                    t = rows[0]["Time"]
                except Exception:
                    t = ""
                if d or t:
                    dt_lbl.config(text=f"{lang.t('stock_card.document_datetime', 'Date/Time')}: {d} {t}".strip())

            for r in rows:
                values = []
                for c in visible_columns:
                    try:
                        values.append(r[c])
                    except Exception:
                        values.append("")

                try:
                    row_code = (r["code"] or "").strip()
                except Exception:
                    row_code = ""

                tags = ("active_code",) if (active_code and row_code == active_code) else ()
                tree.insert("", "end", values=values, tags=tags)

            self._auto_fit_tree_columns(tree)

        except Exception as e:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_card.document_fetch_error", "Failed to load document rows: {error}")
                         .format(error=str(e)),
                         "error")
            try:
                win.destroy()
            except Exception:
                pass
            return
        finally:
            try:
                cur.close()
                conn.close()
            except Exception:
                pass

        footer = tk.Frame(win, bg="#F5F5F5")
        footer.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(
            footer,
            text=lang.t("stock_card.close", "Close"),
            bg="#7F8C8D",
            fg="white",
            command=win.destroy
        ).pack(side="right")

    # ---------------- Data fetch ----------------
    def fetch_transactions_for_code(self, code: str) -> list:
        conn = connect_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            query = """
                SELECT Date, Time, IN_Type, Out_Type, End_User, Third_Party,
                       Qty_IN, Qty_Out, Expiry_date, Remarks, comments, document_number
                FROM stock_transactions
                WHERE code = ?
            """
            params = [code]
            if self.scenario_var.get() != lang.t("stock_card.all_scenarios", "All Scenarios"):
                query += " AND Scenario = ?"
                params.append(self.scenario_var.get())
            if self.kit_var.get() != lang.t("stock_card.all_kits", "All Kits"):
                query += " AND Kit = ?"
                params.append(self.kit_var.get())
            if self.module_var.get() != lang.t("stock_card.all_modules", "All Modules"):
                query += " AND Module = ?"
                params.append(self.module_var.get())
            query += " ORDER BY Date, Time"
            cursor.execute(query, params)
            rows = cursor.fetchall()

            processed_rows = []
            running_balance = 0
            for row in rows:
                qty_in = row['Qty_IN'] or 0
                qty_out = row['Qty_Out'] or 0
                running_balance += qty_in - qty_out

                origin_destination = ""
                if row['IN_Type']:
                    in_disp = self._to_display_in_type(row['IN_Type'])
                    origin_destination = in_disp
                    if row['Third_Party']:
                        origin_destination += f" ({row['Third_Party']})"
                elif row['Out_Type']:
                    out_disp = self._to_display_out_type(row['Out_Type'])
                    origin_destination = out_disp
                    if row['End_User']:
                        origin_destination += f" ({row['End_User']})"

                remarks_disp = self._combine_remarks_and_comments(row['Remarks'], row['comments'])

                processed_rows.append({
                    'date': row['Date'],
                    'time': row['Time'],
                    'origin_destination': origin_destination,
                    'qty_in': qty_in,
                    'qty_out': qty_out,
                    'final_stock': running_balance,
                    'expiry_date': row['Expiry_date'] or "",
                    'remarks': remarks_disp,
                    'document_number': row['document_number'] or "",
                    'mismatch': False
                })

            stock_query = "SELECT SUM(final_qty) AS current_final_stock FROM stock_data WHERE item = ?"
            stock_params = [code]
            if self.scenario_var.get() != lang.t("stock_card.all_scenarios", "All Scenarios"):
                stock_query += " AND scenario = ?"
                stock_params.append(self.scenario_var.get())
            if self.kit_var.get() != lang.t("stock_card.all_kits", "All Kits"):
                stock_query += " AND kit = ?"
                stock_params.append(self.kit_var.get())
            if self.module_var.get() != lang.t("stock_card.all_modules", "All Modules"):
                stock_query += " AND module = ?"
                stock_params.append(self.module_var.get())
            cursor.execute(stock_query, stock_params)
            current_final_stock = cursor.fetchone()['current_final_stock'] or 0

            if processed_rows:
                processed_rows[-1]['mismatch'] = processed_rows[-1]['final_stock'] != current_final_stock

            return processed_rows
        except Exception as e:
            custom_popup(self, lang.t("dialog_titles.error", "Error"),
                         lang.t("stock_card.transaction_error", "Transaction fetch failed: {error}").format(error=str(e)),
                         "error")
            return []
        finally:
            cursor.close()
            conn.close()

    def fetch_project_details(self) -> tuple:
        conn = connect_db()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
            row = cursor.fetchone()
            return row[0] if row and row[0] else lang.t("stock_card.unknown_project", "Unknown Project"), \
                   row[1] if row and row[1] else lang.t("stock_card.unknown_code", "Unknown Code")
        finally:
            cursor.close()
            conn.close()

    def render_ui(self) -> None:
        for w in self.winfo_children():
            w.destroy()

        tk.Label(self, text=lang.t("stock_card.title", "Stock Card"),
                 font=("Helvetica", 20, "bold"), bg="#F5F5F5").pack(pady=10, fill="x")

        filter_frame = tk.Frame(self, bg="#F5F5F5")
        filter_frame.pack(pady=5, fill="x")

        tk.Label(filter_frame, text=lang.t("stock_card.scenario", "Scenario:"), bg="#F5F5F5").grid(row=0, column=0, padx=5, sticky="w")
        self.scenario_var = tk.StringVar(value=lang.t("stock_card.all_scenarios", "All Scenarios"))
        scenarios = list(self.scenario_map.values()) + [lang.t("stock_card.all_scenarios", "All Scenarios")]
        self.scenario_cb = ttk.Combobox(filter_frame, textvariable=self.scenario_var,
                                        values=scenarios, state="readonly", width=30)
        self.scenario_cb.grid(row=0, column=1, padx=5, pady=5)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_filter_selected)

        tk.Label(filter_frame, text=lang.t("stock_card.kit_number", "Kit Number:"), bg="#F5F5F5").grid(row=0, column=2, padx=5, sticky="w")
        self.kit_var = tk.StringVar(value=lang.t("stock_card.all_kits", "All Kits"))
        kits = self.kit_numbers + [lang.t("stock_card.all_kits", "All Kits")]
        self.kit_cb = ttk.Combobox(filter_frame, textvariable=self.kit_var,
                                   values=kits, state="readonly", width=30)
        self.kit_cb.grid(row=0, column=3, padx=5, pady=5)
        self.kit_cb.bind("<<ComboboxSelected>>", self.on_filter_selected)

        tk.Label(filter_frame, text=lang.t("stock_card.module_number", "Module Number:"), bg="#F5F5F5").grid(row=0, column=4, padx=5, sticky="w")
        self.module_var = tk.StringVar(value=lang.t("stock_card.all_modules", "All Modules"))
        modules = self.module_numbers + [lang.t("stock_card.all_modules", "All Modules")]
        self.module_cb = ttk.Combobox(filter_frame, textvariable=self.module_var,
                                      values=modules, state="readonly", width=30)
        self.module_cb.grid(row=0, column=5, padx=5, pady=5)
        self.module_cb.bind("<<ComboboxSelected>>", self.on_filter_selected)

        btn_frame = tk.Frame(self, bg="#F5F5F5")
        btn_frame.pack(pady=5, fill="x")
        tk.Button(btn_frame, text=lang.t("stock_card.clear", "Clear All"), bg="#7F8C8D", fg="white",
                  command=self.clear_form).pack(side="left", padx=5)
        tk.Button(btn_frame, text=lang.t("stock_card.export", "Export"), bg="#2980B9", fg="white",
                  command=self.export_data).pack(side="left", padx=5)

        search_frame = tk.Frame(self, bg="#F5F5F5")
        search_frame.pack(pady=10, fill="x")
        search_frame.grid_columnconfigure(0, weight=0)
        search_frame.grid_columnconfigure(1, weight=0)
        search_frame.grid_columnconfigure(2, weight=0)
        search_frame.grid_columnconfigure(3, weight=1)
        search_frame.grid_columnconfigure(4, weight=1)

        tk.Label(search_frame, text=lang.t("stock_card.item_code", "Item Code"), bg="#F5F5F5") \
            .grid(row=0, column=0, padx=5, sticky="w")
        self.code_entry = tk.Entry(search_frame)
        self.code_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.code_entry.bind("<KeyRelease>", self.search_items)
        self.code_entry.bind("<Return>", self.select_first_result)

        tk.Button(search_frame, text=lang.t("stock_card.clear_search", "Clear Search"),
                  bg="#7F8C8D", fg="white", command=self.clear_search) \
            .grid(row=0, column=2, padx=5, pady=5, sticky="w")

        self.search_listbox = tk.Listbox(search_frame, height=5)
        self.search_listbox.grid(row=0, column=3, padx=5, pady=5, sticky="we")
        self.search_listbox.bind("<<ListboxSelect>>", self.fill_code_from_search)

        info_lbl = tk.Label(search_frame, textvariable=self.item_info_var,
                            font=("Helvetica", 12, "bold"), fg="#2C3E50", bg="#F5F5F5", anchor="e", justify="right")
        info_lbl.grid(row=0, column=4, padx=8, pady=5, sticky="e")

        tree_frame = tk.Frame(self)
        tree_frame.pack(expand=True, fill="both", pady=10)

        self.cols = ("date", "time", "origin_destination", "qty_in", "qty_out", "final_stock",
                     "expiry_date", "remarks", "document_number")
        self.tree = ttk.Treeview(tree_frame, columns=self.cols, show="headings", height=18)
        self.tree.tag_configure("mismatch", foreground="red")

        self.headers = {
            "date": lang.t("stock_card.date", "Date"),
            "time": lang.t("stock_card.time", "Time"),
            "origin_destination": lang.t("stock_card.origin_destination", "Origin/Destination"),
            "qty_in": lang.t("stock_card.qty_in", "IN"),
            "qty_out": lang.t("stock_card.qty_out", "OUT"),
            "final_stock": lang.t("stock_card.final_stock", "Final Stock"),
            "expiry_date": lang.t("stock_card.expiry_date", "Expiry Date"),
            "remarks": lang.t("stock_card.remarks", "Remarks"),
            "document_number": lang.t("stock_transactions.document_number", "Document Number")
        }
        self.widths = {
            "date": 103,
            "time": 100,
            "origin_destination": 220,
            "qty_in": 80,
            "qty_out": 80,
            "final_stock": 100,
            "expiry_date": 110,
            "remarks": 300,
            "document_number": 150
        }

        for c in self.cols:
            self.tree.heading(c, text=self.headers.get(c, c))
            self.tree.column(c, width=self.widths.get(c, 100), stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<ButtonRelease-1>", self.on_tree_click)

        tk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w", bg="#F5F5F5").pack(fill="x", pady=(5, 0))

    def on_tree_click(self, event: tk.Event = None) -> None:
        try:
            region = self.tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            col_id = self.tree.identify_column(event.x)
            row_id = self.tree.identify_row(event.y)
            if not row_id or not col_id:
                return

            idx = int(col_id.replace("#", "")) - 1
            if idx < 0 or idx >= len(self.cols):
                return
            if self.cols[idx] != "document_number":
                return

            values = self.tree.item(row_id, "values") or []
            if idx >= len(values):
                return
            doc = values[idx]
            if doc:
                self.open_document_window(str(doc))
        except Exception:
            return

    def on_filter_selected(self, event: tk.Event = None) -> None:
        self.tree.delete(*self.tree.get_children())
        code = self.code_entry.get().strip()

        if code:
            desc = get_active_designation(code)
            self.item_info_var.set(f"{code} — {desc}")
            self.status_var.set(lang.t("stock_card.loading", "Loading..."))
            self.update_idletasks()
            data = self.fetch_transactions_for_code(code)
            if not data:
                self.status_var.set(lang.t("stock_card.no_items", "No items found for code {code}").format(code=code))
            else:
                for row in data:
                    self.insert_tree_row(row)
                self.status_var.set(lang.t("stock_card.loaded", "Loaded {count} records for code {code}")
                                    .format(count=len(self.tree.get_children()), code=code))
        else:
            self.item_info_var.set("")
            self.status_var.set(lang.t("stock_card.ready", "Ready"))

        self.search_items()

    def search_items(self, event: tk.Event = None) -> None:
        query = self.code_entry.get().strip()
        self.search_listbox.delete(0, tk.END)
        if not query:
            self.tree.delete(*self.tree.get_children())
            self.status_var.set(lang.t("stock_card.ready", "Ready"))
            return
        results = self.fetch_search_results(query)
        for item in results:
            text = f"{item['code']} - {item['description'] or lang.t('stock_card.no_description', 'No Description')}"
            self.search_listbox.insert(tk.END, text)
        self.status_var.set(lang.t("stock_card.found_items", "Found {count} items").format(count=self.search_listbox.size()))

    def clear_search(self) -> None:
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.item_info_var.set("")
        self.status_var.set(lang.t("stock_card.ready", "Ready"))

    def select_first_result(self, event: tk.Event = None) -> None:
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self.fill_code_from_search()

    def fill_code_from_search(self, event: tk.Event = None) -> None:
        sel = self.search_listbox.curselection()
        if not sel:
            return
        code = self.search_listbox.get(sel[0]).split(" - ")[0]
        self.code_entry.delete(0, tk.END)
        self.code_entry.insert(0, code)
        self.search_listbox.delete(0, tk.END)

        desc = get_active_designation(code)
        self.item_info_var.set(f"{code} — {desc}")

        self.tree.delete(*self.tree.get_children())
        self.status_var.set(lang.t("stock_card.loading", "Loading..."))
        self.update_idletasks()
        data = self.fetch_transactions_for_code(code)
        if not data:
            self.status_var.set(lang.t("stock_card.no_items", "No items found for code {code}").format(code=code))
        else:
            for row in data:
                self.insert_tree_row(row)
            self.status_var.set(lang.t("stock_card.loaded", "Loaded {count} records for code {code}")
                                .format(count=len(self.tree.get_children()), code=code))

    def insert_tree_row(self, row: dict) -> None:
        tags = ("mismatch",) if row['mismatch'] else ()
        self.tree.insert("", "end", values=(
            row['date'],
            row['time'],
            row['origin_destination'],
            row['qty_in'],
            row['qty_out'],
            row['final_stock'],
            row['expiry_date'],
            row['remarks'],
            row.get('document_number', "")
        ), tags=tags)

    def clear_form(self) -> None:
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.scenario_var.set(lang.t("stock_card.all_scenarios", "All Scenarios"))
        self.kit_var.set(lang.t("stock_card.all_kits", "All Kits"))
        self.module_var.set(lang.t("stock_card.all_modules", "All Modules"))
        self.item_info_var.set("")
        self.status_var.set(lang.t("stock_card.ready", "Ready"))

    def export_data(self) -> None:
        try:
            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)

            localized_label = lang.t("stock_card.stock_card", "Stock Card")

            import re
            safe_label = re.sub(r"[^\w\-]+", "_", localized_label).strip("_")

            current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            code = self.code_entry.get().strip()
            description = get_active_designation(code) if code else lang.t("stock_card.no_description", "No Description")

            base_name = f"IsEPREP_{safe_label}"
            file_name = f"{base_name}_{code}_{current_time}.xlsx" if code else f"{base_name}_{current_time}.xlsx"

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                title=lang.t("stock_card.save_excel", "Save Excel"),
                initialfile=file_name,
                initialdir=default_dir
            )
            if not file_path:
                self.status_var.set(lang.t("stock_card.export_cancelled", "Export cancelled"))
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = localized_label

            project_name, project_code = self.fetch_project_details()

            ws["A1"] = localized_label
            ws["A1"].font = Font(name="Tahoma", size=14)
            ws["A1"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A1:I1")

            ws["A2"] = f"{project_name} - {project_code}"
            ws["A2"].font = Font(name="Tahoma", size=14)
            ws["A2"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A2:I2")

            ws["A3"] = f"{lang.t('stock_card.code', 'Code')}: {code} - {description}" if code else ""
            ws["A3"].font = Font(name="Tahoma")
            ws["A3"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A3:I3")

            ws.append([])

            headers = [self.tree.heading(col)["text"] for col in self.cols]
            ws.append(headers)

            kit_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            module_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

            rows_data = []
            for iid in self.tree.get_children():
                rows_data.append(self.tree.item(iid)["values"])

            for row_idx, values in enumerate(rows_data, start=6):
                ws.append(values)

                code_val = self.code_entry.get().strip()
                item_type = get_active_item_type(code_val)

                if item_type == "KIT":
                    for col_cells in ws[f"A{row_idx}:I{row_idx}"]:
                        for cell in col_cells:
                            cell.fill = kit_fill
                elif item_type == "MODULE":
                    for col_cells in ws[f"A{row_idx}:I{row_idx}"]:
                        for cell in col_cells:
                            cell.fill = module_fill

            ws.column_dimensions["A"].width = 103.4 / 7
            ws.column_dimensions["B"].width = 100 / 7
            ws.column_dimensions["C"].width = 220 / 7
            ws.column_dimensions["D"].width = 80 / 7
            ws.column_dimensions["E"].width = 80 / 7
            ws.column_dimensions["F"].width = 100 / 7
            ws.column_dimensions["G"].width = 110 / 7
            ws.column_dimensions["H"].width = 300 / 7
            ws.column_dimensions["I"].width = 150 / 7

            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1

            wb.save(file_path)
            wb.close()

            custom_popup(
                self,
                lang.t("dialog_titles.success", "Success"),
                lang.t("stock_card.exported", "Exported to {path}").format(path=file_path),
                "success"
            )
            self.status_var.set(lang.t("stock_card.export_success", "Export successful: {path}").format(path=file_path))

        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("stock_card.export_failed", "Export failed: {error}").format(error=str(e)),
                "error"
            )
            self.status_var.set(lang.t("stock_card.export_error", "Export error: {error}").format(error=str(e)))


if __name__ == "__main__":
    root = tk.Tk()
    app = tk.Tk()  # Dummy app for testing
    app.role = "admin"
    StockCard(root, app, role="admin")
    root.mainloop()