import tkinter as tk
from tkinter import ttk, Toplevel, StringVar, Listbox, Scrollbar, filedialog
import sqlite3
from db import connect_db
from language_manager import lang
import pandas as pd
from popup_utils import custom_popup, custom_askyesno, custom_dialog, show_toast
import openpyxl
from openpyxl.styles import PatternFill

# ============ COLORS ============
COLORS = {
    "scenario": "#0077CC",
    "KIT": "#228B22",
    "MODULE": "#FF8C00",
    "ITEM": "#000000"
}

# ============ ROLE RESTRICTION ============
# Users whose role (canonical or symbol) is in this set are READ-ONLY.
# (manager -> "~", supervisor -> "$")
RESTRICTED_MODIFY = {"manager", "supervisor", "~", "$"}


class KitsComposition(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.role = (app.role.lower() if (app and getattr(app, "role", None)) else "admin")
        self.pack(fill="both", expand=True)

        self.conn = connect_db()
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        self.cursor.execute("PRAGMA foreign_keys = ON")

        self.node_treecode = {}
        self.selected_scenario = None
        self.selected_scenario_id = None
        self.last_menu_position = None

        self.menu = None  # context menu reference

        self.ensure_db_structure()
        self._build_ui()
        self.after(3000, self.auto_refresh_scenarios)
        self.load_scenarios()

    # ---------------- Permission helpers ----------------
    def _can_modify(self):
        return (self.role or "").strip().lower() not in RESTRICTED_MODIFY

    def _deny(self):
        custom_popup(
            self,
            lang.t("dialog_titles.restricted", "Restricted"),
            lang.t("kits.read_only_mode", "Read-only: Your role cannot modify this content."),
            "warning"
        )

    # ---------------- Internal helpers ----------------
    def _normalize_iid(self, iid: str) -> str:
        return iid.split("__", 1)[0] if "__" in iid else iid

    def _unique_iid(self, base: str) -> str:
        if not self.tree.exists(base):
            return base
        idx = 2
        while self.tree.exists(f"{base}__{idx}"):
            idx += 1
        return f"{base}__{idx}"

    def _auto_resize_main_column(self, event):
        total_width = event.width
        fixed_width_qty = 80
        fixed_width_level = 100
        new_width = max(total_width - fixed_width_qty - fixed_width_level - 5, 200)
        self.tree.column("#0", width=new_width)

    # ---------------- UI Build ----------------
    def _build_ui(self):
        main_frame = tk.Frame(self, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=0)
        main_frame.grid_columnconfigure(1, weight=1)

        self.left_frame = tk.Frame(main_frame, bg="#f0f0f0", width=180)
        self.left_frame.grid(row=0, column=0, sticky="ns")
        self.left_frame.pack_propagate(False)

        tk.Label(
            self.left_frame,
            text=lang.t("kits.scenarios", "Scenarios"),
            font=("Arial", 12, "bold"),
            bg="#f0f0f0"
        ).pack(pady=5)

        self.scenario_listbox = tk.Listbox(
            self.left_frame,
            font=("Arial", 10),
            selectbackground="#0077CC",
            selectforeground="white"
        )
        self.scenario_listbox.pack(fill=tk.Y, expand=True, padx=5, pady=5)
        self.scenario_listbox.bind("<<ListboxSelect>>", self.load_selected_scenario)

        button_frame = tk.Frame(self.left_frame, bg="#f0f0f0")
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(
            button_frame,
            text=lang.t("kits.export_excel", "Export to Excel"),
            command=self.export_to_excel,
            font=("Arial", 10),
            bg="#0077CC",
            fg="white",
            relief="flat",
            pady=2
        ).pack(fill=tk.X, pady=2)
        tk.Button(
            button_frame,
            text=lang.t("kits.import_excel", "Import from Excel"),
            command=self.import_from_excel,
            font=("Arial", 10),
            bg="#0077CC",
            fg="white",
            relief="flat",
            pady=2
        ).pack(fill=tk.X, pady=2)

        self.right_frame = tk.Frame(main_frame, bg="#ffffff")
        self.right_frame.grid(row=0, column=1, sticky="nsew")

        tree_frame = tk.Frame(self.right_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            rowheight=25,
            font=("Arial", 10),
            indent=20,
            background="#ffffff",
            fieldbackground="#ffffff"
        )
        style.configure(
            "Treeview.Heading",
            font=("Arial", 11, "bold"),
            background="#d3d3d3",
            foreground="#333333"
        )
        style.map("Treeview",
                  background=[("selected", "#0077CC")],
                  foreground=[("selected", "white")])

        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)

        self.tree = ttk.Treeview(
            tree_frame,
            columns=("std_qty", "level"),
            show="tree headings",
            xscrollcommand=x_scroll.set,
            yscrollcommand=y_scroll.set
        )
        x_scroll.config(command=self.tree.xview)
        y_scroll.config(command=self.tree.yview)
        self.tree.grid(row=0, column=0, sticky="nsew")
        x_scroll.grid(row=1, column=0, sticky="ew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree.heading("#0", text=lang.t("kits.hierarchy", "Hierarchy"), anchor="w")
        self.tree.heading("std_qty", text=lang.t("kits.std_qty", "Standard Quantity"), anchor="center")
        self.tree.heading("level", text=lang.t("kits.level", "Level"), anchor="center")
        self.tree.column("#0", width=500, stretch=True)
        self.tree.column("std_qty", width=80, anchor="center", stretch=False)
        self.tree.column("level", width=100, anchor="center", stretch=False)
        self.tree.bind("<Configure>", self._auto_resize_main_column)

        # Event bindings
        self.tree.bind("<Button-3>", self._debug_event)
        self.tree.bind("<Control-Button-1>", self._debug_event)
        self.tree.bind("<Double-1>", self._debug_event)
        self.tree.bind("<Button-1>", self.select_node)

        self.status_var = tk.StringVar(value=lang.t("kits.ready", "Ready"))
        status_bar = tk.Label(
            self.right_frame,
            textvariable=self.status_var,
            bg="#f0f0f0",
            font=("Arial", 9),
            anchor="w",
            padx=10
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=2)

    # ---------------- Event Debug / Context ----------------
    def _debug_event(self, event):
        node = self.tree.identify_row(event.y)
        if not node:
            self.status_var.set(lang.t("kits.no_node_selected", "No node selected"))
            return
        if event.type == "4":  # ButtonPress
            if event.num == 3 or (event.num == 1 and event.state & 0x4):
                self.last_menu_position = (event.x_root, event.y_root)
                self.show_context_menu(event)
        elif event.type == "3":  # DoubleButton?
            self.toggle_node(event)

    # ---------------- DB Structure ----------------
    def ensure_db_structure(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS scenarios (
                scenario_id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                activity_type TEXT,
                target_population INTEGER,
                responsible_person TEXT,
                stock_location TEXT
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS items_list (
                code TEXT PRIMARY KEY,
                type TEXT,
                designation TEXT,
                designation_en TEXT,
                designation_fr TEXT,
                designation_sp TEXT
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS kit_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scenario_id INTEGER,
                scenario TEXT NOT NULL,
                kit TEXT,
                module TEXT,
                item TEXT,
                code TEXT NOT NULL,
                std_qty INTEGER NOT NULL,
                level TEXT,
                treecode TEXT,
                FOREIGN KEY (scenario_id) REFERENCES scenarios(scenario_id),
                FOREIGN KEY (code) REFERENCES items_list(code),
                CHECK (std_qty > 0)
            )
        """)
        self.conn.commit()
        self.cursor.execute("SELECT COUNT(*) AS cnt FROM scenarios")
        if self.cursor.fetchone()["cnt"] == 0:
            self.cursor.execute("""
                INSERT INTO scenarios (name, activity_type, target_population, responsible_person, stock_location)
                VALUES (?, ?, ?, ?, ?)
            """, (
                lang.t("kits.default_scenario", "Default Scenario"),
                lang.t("kits.general", "General"),
                0,
                lang.t("kits.admin", "Admin"),
                "Default Location"
            ))
            self.conn.commit()

    # ---------------- Scenario Loading ----------------
    def load_scenarios(self):
        self.scenario_listbox.delete(0, tk.END)
        self.cursor.execute("SELECT scenario_id, name FROM scenarios ORDER BY scenario_id")
        for row in self.cursor.fetchall():
            self.scenario_listbox.insert(tk.END, f"{row['scenario_id']} - {row['name']}")

    def auto_refresh_scenarios(self):
        current_count = self.scenario_listbox.size()
        self.cursor.execute("SELECT COUNT(*) AS cnt FROM scenarios")
        new_count = self.cursor.fetchone()["cnt"]
        if new_count != current_count:
            self.load_scenarios()
        self.after(3000, self.auto_refresh_scenarios)

    def load_selected_scenario(self, event):
        selection = self.scenario_listbox.curselection()
        if not selection:
            self.status_var.set(lang.t("kits.no_scenario_selected", "No scenario selected"))
            return
        scenario_line = self.scenario_listbox.get(selection[0])
        try:
            scenario_id = int(scenario_line.split(" - ")[0])
            scenario_name = scenario_line.split(" - ")[1]
        except (IndexError, ValueError):
            self.status_var.set(lang.t("kits.invalid_scenario_format", "Invalid scenario format"))
            return

        self.selected_scenario = scenario_name
        self.selected_scenario_id = scenario_id
        self.node_treecode.clear()
        self.tree.delete(*self.tree.get_children())

        root_iid = f"scenario_{scenario_id}"
        self.tree.insert("", "end", iid=root_iid, text=scenario_line, values=("", ""), tags=("scenario",))
        self.load_hierarchy(scenario_id, scenario_name)
        self.colorize_tree()
        self.status_var.set(f"{lang.t('kits.loaded_scenario', 'Loaded scenario')}: {scenario_name}")

    # ---------------- Tree / Hierarchy ----------------
    def _parse_treecode(self, treecode: str):
        if not treecode or len(treecode) != 11 or not treecode.isdigit():
            return None
        return {
            "SS": treecode[0:2],
            "PPP": treecode[2:5],
            "MMM": treecode[5:8],
            "III": treecode[8:11]
        }

    def load_hierarchy(self, scenario_id, scenario_name):
        kit_instance_map = {}
        module_instance_map = {}
        primary_module_map = {}

        self.cursor.execute("""
            SELECT ki.*, il.type
              FROM kit_items ki
              JOIN items_list il ON ki.code = il.code
             WHERE ki.scenario_id = ?
             ORDER BY ki.treecode
        """, (scenario_id,))
        rows = self.cursor.fetchall()

        # Primary
        for row in rows:
            if row["level"] != "primary":
                continue
            code = row["code"]
            tc = row["treecode"] or ""
            segs = self._parse_treecode(tc)
            ppp = segs["PPP"] if segs else "000"
            qty = row["std_qty"]
            itype = (row["type"] or "UNKNOWN").upper()
            desc = self._get_designation(code)
            text = f"{code} - {desc}" if desc else code
            tag = itype.lower()
            parent = f"scenario_{scenario_id}"
            try:
                if itype == "KIT":
                    base = f"kit_{code}_{ppp}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    kit_instance_map[(code, ppp)] = iid
                    self.node_treecode[iid] = tc
                elif itype == "MODULE":
                    base = f"module_{code}_{ppp}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    primary_module_map[(code, ppp)] = iid
                    self.node_treecode[iid] = tc
                else:
                    base = f"item_{code}_{ppp}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    self.node_treecode[iid] = tc
            except tk.TclError:
                continue

        # Secondary
        for row in rows:
            if row["level"] != "secondary":
                continue
            tc = row["treecode"] or ""
            segs = self._parse_treecode(tc)
            ppp = segs["PPP"] if segs else "000"
            mmm = segs["MMM"] if segs else "000"
            code = row["code"]
            kit = row["kit"]
            module = row["module"]
            item = row["item"]
            qty = row["std_qty"]
            itype = (row["type"] or "UNKNOWN").upper()
            desc = self._get_designation(code)
            text = f"{code} - {desc}" if desc else code
            tag = itype.lower()
            try:
                if module and not item:
                    parent = kit_instance_map.get((kit, ppp), f"scenario_{scenario_id}")
                    base = f"module_{kit or 'none'}_{module}_{ppp}_{mmm}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    module_instance_map[(kit, module, ppp, mmm)] = iid
                    self.node_treecode[iid] = tc
                elif item and not module:
                    parent = kit_instance_map.get((kit, ppp), f"scenario_{scenario_id}")
                    base = f"item_{kit or 'none'}_{item}_{ppp}_{mmm}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    self.node_treecode[iid] = tc
                elif module and item:
                    parent = primary_module_map.get((module, ppp), f"scenario_{scenario_id}")
                    base = f"item_none_{module}_{item}_{ppp}_{mmm}"
                    iid = self._unique_iid(base)
                    self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                    self.node_treecode[iid] = tc
            except tk.TclError:
                continue

        # Tertiary
        for row in rows:
            if row["level"] != "tertiary":
                continue
            tc = row["treecode"] or ""
            segs = self._parse_treecode(tc)
            ppp = segs["PPP"] if segs else "000"
            mmm = segs["MMM"] if segs else "000"
            iii = segs["III"] if segs else "000"
            code = row["code"]
            kit = row["kit"]
            module = row["module"]
            item = row["item"]
            qty = row["std_qty"]
            itype = (row["type"] or "UNKNOWN").upper()
            desc = self._get_designation(code)
            text = f"{code} - {desc}" if desc else code
            tag = itype.lower()
            parent = module_instance_map.get((kit, module, ppp, mmm), f"scenario_{scenario_id}")
            base = f"item_{kit or 'none'}_{module or 'none'}_{item}_{ppp}_{mmm}_{iii}"
            iid = self._unique_iid(base)
            try:
                self.tree.insert(parent, "end", iid=iid, text=text, values=(qty, row["level"]), tags=(tag,))
                self.node_treecode[iid] = tc
            except tk.TclError:
                continue

    def _get_designation(self, code):
        lang_code = lang.lang_code
        cols = {"fr": "designation_fr", "es": "designation_sp", "en": "designation_en"}
        col = cols.get(lang_code, "designation_en")
        self.cursor.execute(f"SELECT {col} AS d FROM items_list WHERE code=?", (code,))
        r = self.cursor.fetchone()
        if r and r["d"]:
            return r["d"]
        if lang_code != "en":
            self.cursor.execute("SELECT designation_en AS d FROM items_list WHERE code=?", (code,))
            r = self.cursor.fetchone()
            if r and r["d"]:
                return r["d"]
        for c in cols.values():
            self.cursor.execute(f"SELECT {c} AS d FROM items_list WHERE code=?", (code,))
            r = self.cursor.fetchone()
            if r and r["d"]:
                return r["d"]
        return None

    # ---------------- Selection ----------------
    def select_node(self, event):
        node = self.tree.identify_row(event.y)
        if node:
            self.tree.focus_set()
            self.tree.focus(node)
            self.tree.selection_set(node)
            self.status_var.set(f"{lang.t('kits.node_selected','Node selected')}: {self.tree.item(node,'text')}")
        else:
            self.status_var.set(lang.t("kits.no_node_selected","No node selected"))
        return "break"

    # ---------------- Context Menu ----------------
    def show_context_menu(self, event):
        node = self.tree.identify_row(event.y)
        if not node:
            return
        self.tree.focus_set()
        self.tree.focus(node)
        self.tree.selection_set(node)
        tags = self.tree.item(node, "tags")
        if not tags:
            return
        if self.menu:
            self.menu.destroy()
        self.menu = tk.Menu(self.tree, tearoff=0, font=("Arial", 10), bg="#f0f0f0", fg="#333333")
        itype = tags[0].upper()
        read_only = not self._can_modify()

        # Helper to add a command respecting read-only
        def add_cmd(label, cmd):
            if read_only:
                self.menu.add_command(label=label, command=self._deny)
            else:
                self.menu.add_command(label=label, command=cmd)

        if itype == "SCENARIO":
            add_cmd(lang.t("kits.add_kit","Add Kit"), lambda: self.add_node("KIT"))
            add_cmd(lang.t("kits.add_module","Add Module"), lambda: self.add_node("MODULE"))
            add_cmd(lang.t("kits.add_item","Add Item"), lambda: self.add_node("ITEM"))
            self.menu.add_separator()
            self.menu.add_command(label=lang.t("kits.view_details","View Details"), command=self.view_details)
            self.menu.add_separator()
            self.menu.add_command(label=lang.t("kits.export_excel","Export to Excel"), command=self.export_to_excel)
            # Import modifies DB; block when read-only
            if read_only:
                self.menu.add_command(label=lang.t("kits.import_excel","Import from Excel"), command=self._deny)
            else:
                self.menu.add_command(label=lang.t("kits.import_excel","Import from Excel"), command=self.import_from_excel)
        elif itype == "KIT":
            add_cmd(lang.t("kits.add_module","Add Module"), lambda: self.add_node("MODULE"))
            add_cmd(lang.t("kits.add_item","Add Item"), lambda: self.add_node("ITEM"))
            self.menu.add_separator()
            add_cmd(lang.t("kits.duplicate","Duplicate Kit Instance"), self.duplicate_selected_kit)
            add_cmd(lang.t("kits.edit_qty","Edit Quantity"), self.edit_quantity)
            add_cmd(lang.t("kits.delete","Delete"), self.delete_node)
            self.menu.add_command(label=lang.t("kits.view_details","View Details"), command=self.view_details)
        elif itype == "MODULE":
            add_cmd(lang.t("kits.add_item","Add Item"), lambda: self.add_node("ITEM"))
            self.menu.add_separator()
            add_cmd(lang.t("kits.duplicate","Duplicate Module Instance"), self.duplicate_selected_module)
            add_cmd(lang.t("kits.edit_qty","Edit Quantity"), self.edit_quantity)
            add_cmd(lang.t("kits.delete","Delete"), self.delete_node)
            self.menu.add_command(label=lang.t("kits.view_details","View Details"), command=self.view_details)
        elif itype == "ITEM":
            add_cmd(lang.t("kits.edit_qty","Edit Quantity"), self.edit_quantity)
            add_cmd(lang.t("kits.delete","Delete"), self.delete_node)
            self.menu.add_command(label=lang.t("kits.view_details","View Details"), command=self.view_details)
        self.menu.post(event.x_root, event.y_root)

    def toggle_node(self, event):
        node = self.tree.identify_row(event.y)
        if not node:
            return
        self.tree.focus(node)
        self.tree.selection_set(node)
        self.tree.item(node, open=not self.tree.item(node, "open"))

    # ---------------- View Details ----------------
    def view_details(self):
        sel = self.tree.selection()
        if not sel:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.no_node_selected","No node selected"), "warning")
            return
        node = sel[0]
        vals = self.tree.item(node, "values")
        qty = vals[0] if vals else "N/A"
        level = vals[1] if vals and len(vals) > 1 else "N/A"
        treecode = self.node_treecode.get(node, "N/A")
        custom_popup(
            self,
            lang.t("kits.view_details","View Details"),
            f"{lang.t('kits.details_for','Details for')} {self.tree.item(node,'text')}\n"
            f"{lang.t('kits.std_qty','Standard Quantity')}: {qty}\n"
            f"{lang.t('kits.level','Level')}: {level}\n"
            f"TreeCode: {treecode}",
            "info"
        )

    # ---------------- Add Node ----------------
    def add_node(self, level):
        if not self._can_modify():
            self._deny()
            return
        parent_sel = self.tree.selection()
        if not parent_sel:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.no_node_selected","No node selected"), "error")
            return
        parent = parent_sel[0]
        parent_tags = self.tree.item(parent, "tags")
        parent_type = parent_tags[0].upper() if parent_tags else None

        if parent_type == "ITEM":
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.invalid_parent","Cannot add children to an ITEM"), "error")
            return
        if parent_type == "MODULE" and level.upper() != "ITEM":
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.module_item_only","Modules can only have items"), "error")
            return

        code = self.search_dialog(level)
        if not code:
            return

        if level.upper() in ("KIT", "MODULE"):
            self._process_node_addition(level, parent, parent_tags, code, 1)
            return

        dlg = Toplevel(self)
        dlg.title(lang.t("kits.std_qty","Standard Quantity"))
        self._center_toplevel(dlg, width=260, height=150)
        tk.Label(dlg, text=lang.t("kits.enter_qty","Enter Quantity"), font=("Arial", 10), pady=10).pack()
        qty_var = tk.StringVar(value="1")
        entry = tk.Entry(dlg, textvariable=qty_var, font=("Arial", 10), width=10)
        entry.pack(pady=5)
        entry.focus()

        def submit():
            try:
                qv = int(qty_var.get())
                if qv <= 0:
                    raise ValueError
                dlg.destroy()
                self._process_node_addition(level, parent, parent_tags, code, qv)
            except ValueError:
                custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.invalid_qty","Invalid quantity"), "error")
                entry.focus_force()

        tk.Button(dlg, text=lang.t("kits.ok","OK"), command=submit, bg="#0077CC", fg="white", relief="flat").pack(pady=8)
        entry.bind("<Return>", lambda e: submit())
        dlg.grab_set()
        dlg.transient(self)
        dlg.wait_window()

    def _process_node_addition(self, level, parent, parent_tags, code, qty):
        if not self._can_modify():
            self._deny()
            return
        code_norm = code.strip()
        self.cursor.execute("SELECT code, type FROM items_list WHERE code=?", (code_norm,))
        row = self.cursor.fetchone()
        if not row:
            custom_popup(self, lang.t("kits.error","Error"), f"{lang.t('kits.invalid_code','Invalid code')}: {code_norm}", "error")
            return
        item_type = (row["type"] or "").upper()
        if level.upper() != item_type:
            custom_popup(self, lang.t("kits.error","Error"), f"{code_norm}: {item_type} != {level.upper()}", "error")
            return
        if not self.selected_scenario_id:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.no_scenario_selected","No scenario selected"), "error")
            return

        parent_type = parent_tags[0].upper() if parent_tags else None

        if parent_type == "SCENARIO":  
            db_level = "primary"
        elif parent_type == "KIT":
            db_level = "secondary"
        elif parent_type == "MODULE":
            # ✅ FIX: Check if parent module is at primary level
            parent_level = self.tree.item(parent, "values")[1] if self.tree.item(parent, "values") else None
    
            if parent_level == "primary":
                # Primary-level module → children are secondary
                db_level = "secondary"
            else:
                # Secondary-level module → children are tertiary
                db_level = "tertiary"
        else:
            custom_popup(self, lang. t("kits.error","Error"), lang.t("kits.invalid_parent","Invalid parent"), "error")
            return

        if item_type in ("KIT", "MODULE"):
            qty = 1

        kit = module = item = None
        ppp_override = None
        mmm_override = None
        parent_base = self._normalize_iid(parent)
        parts = parent_base.split("_")

        if parent_type == "KIT":
            if len(parts) >= 3:
                ppp_override = parts[2]
        elif parent_type == "MODULE":
            if len(parts) == 3:
                ppp_override = parts[2]
            elif len(parts) >= 5:
                ppp_override = parts[3]
                mmm_override = parts[4]

        if db_level == "primary":
            if item_type == "KIT":
                kit = code_norm
            elif item_type == "MODULE":
                module = code_norm
            else:
                item = code_norm
        elif db_level == "secondary":
            if parts[0] == "kit":
                kit = parts[1]
            if item_type == "MODULE":
                module = code_norm
            else:
                item = code_norm
        elif db_level == "tertiary":
            if item_type != "ITEM":
                custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.item_required","Item required"), "error")
                return
            if parts[0] == "module":
                kit = parts[1] if parts[1] != 'none' else None
                module = parts[2]
            item = code_norm

        treecode = self._generate_treecode(db_level, kit, module, item, ppp_override, mmm_override)
        if treecode is None:
            custom_popup(self, lang.t("kits.error","Error"), "Treecode allocation failed (limit?)", "error")
            return

        try:
            self.cursor.execute("""
                INSERT INTO kit_items
                    (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (self.selected_scenario_id, self.selected_scenario, kit, module, item, code_norm, qty, db_level, treecode))
            self.conn.commit()
            self.status_var.set(f"{lang.t('kits.node_added','Node added')}: {code_norm}")
        except Exception as e:
            custom_popup(self, lang.t("kits.error","Error"), f"{lang.t('kits.db_error','Database error')}: {e}", "error")
            return

        self.refresh_current_tree(preserve_view=True)

    # -------- Treecode Generation (unchanged logic, only inside class) --------
    def _generate_treecode(self, level_db, kit, module, item, ppp_override=None, mmm_override=None):
        ss = f"{self.selected_scenario_id:02d}"

        if level_db == "primary":
            self.cursor.execute("""
                SELECT COUNT(*) AS cnt FROM kit_items
                 WHERE scenario_id=? AND level='primary'
            """, (self.selected_scenario_id,))
            c = self.cursor.fetchone()["cnt"]
            if c >= 999:
                return None
            ppp = f"{c+1:03d}"
            return f"{ss}{ppp}000000"

        if level_db == "secondary":
            if kit:
                if ppp_override:
                    ppp = ppp_override
                else:
                    self.cursor.execute("""
                        SELECT treecode FROM kit_items
                         WHERE scenario_id=? AND kit=? AND level='primary'
                         ORDER BY treecode LIMIT 1
                    """, (self.selected_scenario_id, kit))
                    krow = self.cursor.fetchone()
                    if not krow:
                        return None
                    ppp = krow["treecode"][2:5]
            else:
                ppp = ppp_override or "001"

            self.cursor.execute("""
                SELECT COUNT(*) AS cnt FROM kit_items
                 WHERE scenario_id=?
                   AND level='secondary'
                   AND substr(treecode,3,3)=?
                   AND (? IS NULL OR kit=?)
            """, (self.selected_scenario_id, ppp, kit, kit))
            c = self.cursor.fetchone()["cnt"]
            if c >= 999:
                return None
            mmm = f"{c+1:03d}"
            return f"{ss}{ppp}{mmm}000"

        if level_db == "tertiary":
            # ✅ FIX:   Check BOTH primary and secondary levels for the parent module
            if ppp_override and mmm_override: 
                # Try secondary first (module under kit)
                self.cursor. execute("""
                    SELECT treecode FROM kit_items
                    WHERE scenario_id=? AND module=?  AND level='secondary'
                    AND substr(treecode,3,3)=? AND substr(treecode,6,3)=?
                    AND (? IS NULL OR kit=?)
                    ORDER BY treecode LIMIT 1
                """, (self.selected_scenario_id, module, ppp_override, mmm_override, kit, kit))
                sec_row = self.cursor.fetchone()
        
                if not sec_row: 
                    # ✅ Try primary level (standalone module in scenario)
                    self.cursor. execute("""
                        SELECT treecode FROM kit_items
                        WHERE scenario_id=? AND code=? AND level='primary'
                        AND substr(treecode,3,3)=?
                        ORDER BY treecode LIMIT 1
                    """, (self.selected_scenario_id, module, ppp_override))
                    sec_row = self. cursor.fetchone()
        
                if not sec_row: 
                    return None
                ppp = ppp_override
                mmm = mmm_override
            else: 
                # Try secondary first (module under kit)
                self.cursor.execute("""
                    SELECT treecode FROM kit_items
                    WHERE scenario_id=? AND module=? AND level='secondary'
                    AND (? IS NULL OR kit=?)
                    ORDER BY treecode LIMIT 1
                """, (self. selected_scenario_id, module, kit, kit))
                sec_row = self.cursor.fetchone()
        
                if not sec_row:
                    # ✅ Try primary level (standalone module in scenario)
                    self.cursor.execute("""
                        SELECT treecode FROM kit_items
                        WHERE scenario_id=?  AND code=? AND level='primary'
                        ORDER BY treecode LIMIT 1
                    """, (self.selected_scenario_id, module))
                    sec_row = self. cursor.fetchone()
        
                if not sec_row: 
                    return None
        
                segs = self._parse_treecode(sec_row["treecode"])
                if not segs:
                    return None
                ppp = segs["PPP"]
                mmm = segs["MMM"]

            prefix = f"{ss}{ppp}{mmm}"
            self.cursor.execute("""
                SELECT COUNT(*) AS cnt FROM kit_items
                WHERE scenario_id=? AND module=? AND level='tertiary'
                AND substr(treecode,3,3)=? AND substr(treecode,6,3)=?
                AND (? IS NULL OR kit=?)
            """, (self.selected_scenario_id, module, ppp, mmm, kit, kit))
            c = self.cursor.fetchone()["cnt"]
            if c >= 999:
                return None
            iii = f"{c+1:03d}"
            return f"{prefix}{iii}"

    # ---------------- Edit Quantity ----------------
    def edit_quantity(self):
        if not self._can_modify():
            self._deny()
            return
        sel = self.tree.selection()
        if not sel:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.no_node_selected","No node selected"), "error")
            return
        node = sel[0]
        tags = self.tree.item(node, "tags")
        if not tags:
            return
        itype = tags[0].upper()
        if itype in ("KIT", "MODULE"):
            custom_popup(self, lang.t("kits.info","Info"), lang.t("kits.fixed_qty","Quantity fixed at 1. Duplicate for more instances."), "info")
            return
        current_qty = self.tree.item(node, "values")[0]

        dlg = Toplevel(self)
        dlg.title(lang.t("kits.edit_qty","Edit Quantity"))
        self._center_toplevel(dlg, width=260, height=150)
        tk.Label(dlg, text=lang.t("kits.enter_new_qty","Enter new quantity"), pady=10).pack()
        qty_var = tk.StringVar(value=str(current_qty))
        entry = tk.Entry(dlg, textvariable=qty_var, width=10)
        entry.pack(pady=5)
        entry.focus()

        def save():
            try:
                new_qty = int(qty_var.get())
                if new_qty <= 0:
                    raise ValueError
                tc = self.node_treecode.get(node)
                if not tc:
                    custom_popup(self, lang.t("kits.error","Error"), "Missing treecode", "error")
                    return
                self.cursor.execute("""
                    UPDATE kit_items
                       SET std_qty=?
                     WHERE scenario_id=? AND treecode=?
                """, (new_qty, self.selected_scenario_id, tc))
                self.conn.commit()
                dlg.destroy()
                self.refresh_current_tree(preserve_view=True)
            except ValueError:
                custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.invalid_qty","Invalid quantity"), "error")
                entry.focus_force()
            except Exception as e:
                custom_popup(self, lang.t("kits.error","Error"), f"{lang.t('kits.db_error','Database error')}: {e}", "error")

        tk.Button(dlg, text=lang.t("kits.ok","OK"), command=save, bg="#0077CC", fg="white", relief="flat").pack(pady=8)
        entry.bind("<Return>", lambda e: save())
        dlg.grab_set()
        dlg.wait_window()

    # ---------------- Delete Node ----------------
    def delete_node(self):
        if not self._can_modify():
            self._deny()
            return
        sel = self.tree.selection()
        if not sel:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.no_node_selected","No node selected"), "error")
            return
        node = sel[0]
        tags = self.tree.item(node, "tags")
        if not tags or "scenario" in tags:
            custom_popup(self, lang.t("kits.error","Error"), lang.t("kits.cannot_delete_scenario","Cannot delete scenario"), "warning")
            return
        tc = self.node_treecode.get(node)
        if not tc:
            custom_popup(self, lang.t("kits.error","Error"), "Internal: missing treecode", "error")
            return
        segs = self._parse_treecode(tc)
        itype = tags[0].upper()
        parent = self.tree.parent(node)
        siblings = list(self.tree.get_children(parent)) if parent else []
        try:
            pos = siblings.index(node)
        except ValueError:
            pos = 0

        confirm = custom_askyesno(self, lang.t("kits.delete","Delete"),
                                  lang.t("kits.delete_confirm","Are you sure you want to delete this node and its contents?"))
        if confirm != "yes":
            return

        try:
            if itype == "KIT":
                ppp = segs["PPP"]
                base = self._normalize_iid(node).split("_")
                kit_code = base[1]
                self.cursor.execute("""
                    DELETE FROM kit_items
                     WHERE scenario_id=? AND kit=? AND substr(treecode,3,3)=?
                """, (self.selected_scenario_id, kit_code, ppp))
            elif itype == "MODULE":
                level = self.tree.item(node, "values")[1]
                base = self._normalize_iid(node).split("_")
                ppp = segs["PPP"]
                if level == "primary":
                    module_code = base[1]
                    self.cursor.execute("""
                        DELETE FROM kit_items
                         WHERE scenario_id=? AND module=? AND substr(treecode,3,3)=?
                    """, (self.selected_scenario_id, module_code, ppp))
                else:
                    module_code = base[2]
                    kit_code = base[1] if base[1] != 'none' else None
                    mmm = segs["MMM"]
                    self.cursor.execute("""
                        DELETE FROM kit_items
                         WHERE scenario_id=? AND module=? AND (kit=? OR (? IS NULL AND kit IS NULL))
                           AND substr(treecode,3,3)=? AND substr(treecode,6,3)=?
                    """, (self.selected_scenario_id, module_code, kit_code, kit_code, ppp, mmm))
            else:
                self.cursor.execute("""
                    DELETE FROM kit_items
                     WHERE scenario_id=? AND treecode=?
                """, (self.selected_scenario_id, tc))

            self.conn.commit()
            new_focus = None
            if parent and self.tree.exists(parent):
                if node in siblings:
                    siblings.remove(node)
                if siblings:
                    new_focus = siblings[min(pos, len(siblings)-1)]
                else:
                    new_focus = parent
            self.refresh_current_tree(preserve_view=True, force_focus_treecode=None)
            if new_focus and self.tree.exists(new_focus):
                self.tree.selection_set(new_focus)
                self.tree.focus(new_focus)
                self.tree.see(new_focus)
            show_toast(self, lang.t("kits.node_deleted","Node deleted"), "success")
        except Exception as e:
            custom_popup(self, lang.t("kits.error","Error"), f"{lang.t('kits.db_error','Database error')}: {e}", "error")

    # ---------------- Search Dialog ----------------
    def search_dialog(self, level):
        dialog = Toplevel(self)
        dialog.title(f"{lang.t('kits.search','Search')} {level}")
        self._center_toplevel(dialog, width=520, height=320)
        dialog.grab_set()
        dialog.transient(self)

        tk.Label(dialog, text=f"{lang.t('kits.search','Search')} {level}:", font=("Arial", 10), pady=10).pack()
        search_var = StringVar()
        entry = tk.Entry(dialog, textvariable=search_var, font=("Arial", 10))
        entry.pack(fill=tk.X, padx=10, pady=5)
        listbox = Listbox(dialog, font=("Arial", 10), selectbackground="#0077CC", selectforeground="white")
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scrollbar = Scrollbar(listbox, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)

        results = []
        lang_code = lang.lang_code
        designation_cols = {"fr": "designation_fr", "es": "designation_sp", "en": "designation_en"}
        designation_col = designation_cols.get(lang_code, "designation_en")
        type_filter = level.upper()

        def update_list(*_):
            query = search_var.get().strip()
            results.clear()
            listbox.delete(0, tk.END)
            try:
                self.cursor.execute(f"""
                    SELECT code, {designation_col} AS designation FROM items_list
                     WHERE UPPER(type)=? AND (code LIKE ? OR {designation_col} LIKE ?)
                     ORDER BY code LIMIT 30
                """, (type_filter, f"%{query}%", f"%{query}%"))
                rows = self.cursor.fetchall()
                if not rows and lang_code != "en":
                    self.cursor.execute("""
                        SELECT code, designation_en AS designation FROM items_list
                         WHERE UPPER(type)=? AND (code LIKE ? OR designation_en LIKE ?)
                         ORDER BY code LIMIT 30
                    """, (type_filter, f"%{query}%", f"%{query}%"))
                    rows = self.cursor.fetchall()
                if not rows:
                    for col in designation_cols.values():
                        if col == designation_col:
                            continue
                        self.cursor.execute(f"""
                            SELECT code, {col} AS designation FROM items_list
                             WHERE UPPER(type)=? AND (code LIKE ? OR {col} LIKE ?)
                             ORDER BY code LIMIT 30
                        """, (type_filter, f"%{query}%", f"%{query}%"))
                        alt = self.cursor.fetchall()
                        if alt:
                            rows = alt
                            break
                if not rows:
                    listbox.insert(tk.END, lang.t("kits.no_items_found","No items found"))
                else:
                    for r in rows:
                        cd = r["code"].strip()
                        listbox.insert(tk.END, f"{cd} - {r['designation'] or 'No Description'}")
                        results.append(cd)
            except Exception as e:
                listbox.insert(tk.END, f"DB error: {e}")

        search_var.trace("w", update_list)
        update_list()

        chosen = {"code": None}

        def on_select(_=None):
            idxs = listbox.curselection()
            if idxs and results:
                chosen["code"] = results[idxs[0]]
            dialog.destroy()

        listbox.bind("<Double-Button-1>", on_select)
        entry.bind("<Return>", on_select)
        entry.focus()
        dialog.wait_window()
        return chosen["code"]

    # ---------------- Export ----------------
    def export_to_excel(self):
        if not self.selected_scenario_id:
            custom_popup(self, lang.t("kits.warning","Warning"), lang.t("kits.no_scenario_selected","No scenario selected"), "warning")
            return
        try:
            self.cursor.execute("""
                SELECT scenario, kit, module, item, code, std_qty, level, treecode
                  FROM kit_items
                 WHERE scenario_id=?
                 ORDER BY treecode
            """, (self.selected_scenario_id,))
            rows = self.cursor.fetchall()
            if not rows:
                custom_popup(self, lang.t("kits.info","Info"), lang.t("kits.no_data_to_export","No data to export"), "info")
                return
            data = []
            for r in rows:
                data.append({
                    "Scenario": r["scenario"],
                    "Kit": r["kit"] or "",
                    "Module": r["module"] or "",
                    "Item": r["item"] or "",
                    "Code": r["code"],
                    "Designation": self._get_designation(r["code"]) or "No Description",
                    "Standard Quantity": r["std_qty"],
                    "Level": r["level"],
                    "TreeCode": r["treecode"]
                })
            df = pd.DataFrame(data)
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files","*.xlsx")],
                title=lang.t("kits.save_excel","Save Excel"),
                initialfile=f"{self.selected_scenario}_kits.xlsx"
            )
            if not file_path:
                self.status_var.set(lang.t("kits.export_cancelled","Export cancelled"))
                return
            df.to_excel(file_path, index=False, engine="openpyxl")
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            kit_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            module_fill = PatternFill(start_color="F7F6CD", end_color="F7F6CD", fill_type="solid")
            headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
            level_col = headers.get("Level", 0)
            code_col = headers.get("Code", 0)
            for r in range(2, ws.max_row + 1):
                level_val = ws.cell(row=r, column=level_col).value
                code_val = ws.cell(row=r, column=code_col).value
                self.cursor.execute("SELECT type FROM items_list WHERE code=?", (code_val,))
                trow = self.cursor.fetchone()
                ttype = trow["type"].upper() if trow and trow["type"] else None
                if level_val == "primary" and ttype == "KIT":
                    for c in ws[r]: c.fill = kit_fill
                elif level_val == "secondary" and ttype == "MODULE":
                    for c in ws[r]: c.fill = module_fill
            wb.save(file_path)
            wb.close()
            custom_popup(self, lang.t("kits.success","Success"),
                         f"{lang.t('kits.export_success','Export successful')}: {file_path}", "success")
        except Exception as e:
            custom_popup(self, lang.t("kits.error","Error"),
                         f"{lang.t('kits.export_error','Export error')}: {e}", "error")

    # ---------------- Import (Restricted if read-only) ----------------
    def import_from_excel(self):
        if not self._can_modify():
            self._deny()
            return
        if not self.selected_scenario_id:
            custom_popup(self, lang.t("kits.warning","Warning"),
                         lang.t("kits.no_scenario_selected","No scenario selected"), "warning")
            return
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files","*.xlsx *.xls")],
                title=lang.t("kits.open_excel","Open Excel")
            )
            if not file_path:
                self.status_var.set(lang.t("kits.import_cancelled","Import cancelled"))
                return
            df = pd.read_excel(file_path, engine="openpyxl")
            df.columns = [c.lower() for c in df.columns]
            required = {'kit','module','item','code','standard quantity'}
            if not required.issubset(df.columns):
                custom_popup(self, lang.t("kits.error","Error"),
                             lang.t("kits.invalid_excel_format","Invalid Excel format"), "error")
                return

            imported = 0
            skipped = []
            codes = df['code'].dropna().unique()
            type_map = {}
            if len(codes):
                placeholders = ",".join(["?"] * len(codes))
                self.cursor.execute(
                    f"SELECT code, type FROM items_list WHERE code IN ({placeholders})",
                    list(codes)
                )
                type_map = {r['code']: (r['type'] or '').upper() for r in self.cursor.fetchall()}

            for idx, r in df.iterrows():
                code = str(r['code']) if pd.notna(r['code']) else None
                if not code:
                    skipped.append(f"Row {idx+2}: missing code")
                    continue
                if code not in type_map:
                    skipped.append(f"Row {idx+2}: unknown code {code}")
                    continue
                try:
                    qty = int(r['standard quantity']) if pd.notna(r['standard quantity']) else None
                except Exception:
                    qty = None
                if not qty or qty <= 0:
                    skipped.append(f"Row {idx+2}: invalid qty")
                    continue
                kit = str(r['kit']) if pd.notna(r['kit']) else None
                module = str(r['module']) if pd.notna(r['module']) else None
                item = str(r['item']) if pd.notna(r['item']) else None
                present = [x for x in [kit, module, item] if x]

                if len(present) == 1:
                    level = "primary"
                    kit = code
                    module = item = None
                elif len(present) == 2:
                    level = "secondary"
                    module = code
                    item = None
                elif len(present) == 3:
                    level = "tertiary"
                    item = code
                else:
                    skipped.append(f"Row {idx+2}: hierarchy error")
                    continue

                if type_map[code] in ("KIT","MODULE"):
                    qty = 1

                treecode = self._generate_treecode(level, kit, module, item)
                if not treecode:
                    skipped.append(f"Row {idx+2}: allocation fail")
                    continue

                self.cursor.execute("""
                    INSERT INTO kit_items
                        (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (self.selected_scenario_id, self.selected_scenario,
                      kit, module, item, code, qty, level, treecode))
                imported += 1

            self.conn.commit()
            self.refresh_current_tree(preserve_view=True)
            msg = f"{lang.t('kits.import_success','Import successful')}: {imported}"
            if skipped:
                msg += f"\n{lang.t('kits.skipped_rows','Skipped rows')} ({len(skipped)}):\n" + "\n".join(skipped[:10])
                if len(skipped) > 10:
                    msg += f"\n... {len(skipped)-10} more"
                custom_popup(self, lang.t("kits.warning","Warning"), msg, "warning")
            else:
                custom_popup(self, lang.t("kits.success","Success"), msg, "success")
            self.status_var.set(msg)
        except Exception as e:
            custom_popup(self, lang.t("kits.error","Error"),
                         f"{lang.t('kits.import_error','Import error')}: {e}", "error")

    # ---------------- View State ----------------
    def _save_view_state(self):
        expanded_tcs = set()
        for top in self.tree.get_children():
            self._collect_expanded(top, expanded_tcs)
        sel_tc = None
        sel = self.tree.selection()
        if sel:
            sel_tc = self.node_treecode.get(sel[0])
        y0, _ = self.tree.yview()
        return {"expanded": expanded_tcs, "selected": sel_tc, "y": y0}

    def _restore_view_state(self, state):
        if not state:
            return
        expanded = state.get("expanded", set())
        selected = state.get("selected")
        y = state.get("y", 0.0)
        for iid, tc in self.node_treecode.items():
            if tc in expanded:
                self.tree.item(iid, open=True)
        if selected:
            new_iid = self.find_iid_by_treecode(selected)
            if new_iid:
                self.tree.selection_set(new_iid)
                self.tree.focus(new_iid)
                self.tree.see(new_iid)
        self.after_idle(lambda: self.tree.yview_moveto(y))

    def refresh_current_tree(self, parent_iid=None, focused_iid=None, preserve_view=False, force_focus_treecode=None):
        if not self.selected_scenario_id:
            return
        state = self._save_view_state() if preserve_view else None
        self.node_treecode.clear()
        self.tree.delete(*self.tree.get_children())
        root_iid = f"scenario_{self.selected_scenario_id}"
        self.tree.insert("", "end", iid=root_iid,
                         text=f"{self.selected_scenario_id} - {self.selected_scenario}",
                         values=("", ""), tags=("scenario",))
        self.load_hierarchy(self.selected_scenario_id, self.selected_scenario)
        self.colorize_tree()
        if force_focus_treecode:
            new_iid = self.find_iid_by_treecode(force_focus_treecode)
            if new_iid:
                self.tree.selection_set(new_iid)
                self.tree.focus(new_iid)
                self.tree.see(new_iid)
        elif preserve_view:
            self._restore_view_state(state)

    def _collect_expanded(self, iid, store: set):
        if self.tree.item(iid, "open"):
            tc = self.node_treecode.get(iid)
            if tc:
                store.add(tc)
        for c in self.tree.get_children(iid):
            self._collect_expanded(c, store)

    def find_iid_by_treecode(self, treecode: str):
        for iid, tc in self.node_treecode.items():
            if tc == treecode:
                return iid
        return None

    # ---------------- Colors ----------------
    def colorize_tree(self):
        for iid in self.tree.get_children():
            self.apply_color_recursive(iid)

    def apply_color_recursive(self, iid):
        tags = self.tree.item(iid, "tags")
        if tags:
            self.tree.tag_configure(tags[0], foreground=COLORS.get(tags[0].upper(), "#000000"))
        for child in self.tree.get_children(iid):
            self.apply_color_recursive(child)

    # ---------------- ID helpers ----------------
    def _next_primary_ppp(self):
        self.cursor.execute("""
            SELECT treecode FROM kit_items
             WHERE scenario_id=? AND level='primary'
        """, (self.selected_scenario_id,))
        used = {self._parse_treecode(r["treecode"])["PPP"] for r in self.cursor.fetchall()
                if self._parse_treecode(r["treecode"])}
        for i in range(1, 1000):
            p = f"{i:03d}"
            if p not in used:
                return p
        return None

    def _next_mmm_in_ppp(self, ppp):
        self.cursor.execute("""
            SELECT treecode FROM kit_items
             WHERE scenario_id=? AND substr(treecode,3,3)=?
               AND (level='secondary' OR level='tertiary')
        """, (self.selected_scenario_id, ppp))
        used = {self._parse_treecode(r["treecode"])["MMM"] for r in self.cursor.fetchall()
                if self._parse_treecode(r["treecode"])}
        for i in range(1, 1000):
            m = f"{i:03d}"
            if m not in used:
                return m
        return None

    def _iter_new_iii(self, ppp, mmm):
        self.cursor.execute("""
            SELECT treecode FROM kit_items
             WHERE scenario_id=? AND substr(treecode,3,3)=? AND substr(treecode,6,3)=?
               AND level='tertiary'
        """, (self.selected_scenario_id, ppp, mmm))
        used = {self._parse_treecode(r["treecode"])["III"] for r in self.cursor.fetchall()
                if self._parse_treecode(r["treecode"])}
        for i in range(1, 1000):
            iii = f"{i:03d}"
            if iii not in used:
                yield iii

    # ---------------- Duplication (Restricted) ----------------
    def duplicate_selected_kit(self):
        if not self._can_modify():
            self._deny()
            return
        sel = self.tree.selection()
        if not sel:
            custom_popup(self, "Error", "Select a KIT to duplicate.", "error")
            return
        node = sel[0]
        tags = self.tree.item(node, "tags") or []
        if not tags or tags[0].upper() != "KIT":
            custom_popup(self, "Error", "Selected node is not a KIT.", "error")
            return
        if not self.selected_scenario_id:
            custom_popup(self, "Error", "No scenario selected.", "error")
            return
        base = self._normalize_iid(node).split("_")
        kit_code = base[1]
        new_ppp = self._next_primary_ppp()
        if not new_ppp:
            custom_popup(self, "Error", "No free primary slot available.", "error")
            return
        ss = f"{self.selected_scenario_id:02d}"
        primary_tc = f"{ss}{new_ppp}000000"
        try:
            self.cursor.execute("BEGIN")
            self.cursor.execute("""
                INSERT INTO kit_items
                    (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                VALUES (?, ?, ?, ?, ?, ?, 1, 'primary', ?)
            """, (self.selected_scenario_id, self.selected_scenario, kit_code, None, None, kit_code, primary_tc))
            self.cursor.execute("""
                SELECT DISTINCT module FROM kit_items
                 WHERE scenario_id=? AND kit=? AND module IS NOT NULL
                   AND level='secondary'
            """, (self.selected_scenario_id, kit_code))
            modules = [r['module'] for r in self.cursor.fetchall()]
            for module_code in modules:
                new_mmm = self._next_mmm_in_ppp(new_ppp)
                if not new_mmm:
                    raise ValueError("No free MMM for module")
                module_tc = f"{ss}{new_ppp}{new_mmm}000"
                self.cursor.execute("""
                    INSERT INTO kit_items
                        (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                    VALUES (?, ?, ?, ?, ?, ?, 1, 'secondary', ?)
                """, (self.selected_scenario_id, self.selected_scenario,
                      kit_code, module_code, None, module_code, module_tc))
                self.cursor.execute("""
                    SELECT DISTINCT item, code, std_qty
                      FROM kit_items
                     WHERE scenario_id=? AND kit=? AND module=? AND level='tertiary'
                """, (self.selected_scenario_id, kit_code, module_code))
                items = self.cursor.fetchall()
                iii_gen = self._iter_new_iii(new_ppp, new_mmm)
                for it in items:
                    iii = next(iii_gen)
                    item_tc = f"{ss}{new_ppp}{new_mmm}{iii}"
                    self.cursor.execute("""
                        INSERT INTO kit_items
                            (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'tertiary', ?)
                    """, (self.selected_scenario_id, self.selected_scenario,
                          kit_code, module_code, it['item'], it['code'], it['std_qty'], item_tc))
            self.cursor.execute("""
                SELECT DISTINCT item, code, std_qty
                  FROM kit_items
                 WHERE scenario_id=? AND kit=? AND module IS NULL
                   AND item IS NOT NULL AND level='secondary'
            """, (self.selected_scenario_id, kit_code))
            direct_items = self.cursor.fetchall()
            for di in direct_items:
                new_mmm = self._next_mmm_in_ppp(new_ppp)
                if not new_mmm:
                    raise ValueError("No free MMM for direct item")
                sec_tc = f"{ss}{new_ppp}{new_mmm}000"
                self.cursor.execute("""
                    INSERT INTO kit_items
                        (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'secondary', ?)
                """, (self.selected_scenario_id, self.selected_scenario,
                      kit_code, None, di['item'], di['code'], di['std_qty'], sec_tc))
            self.conn.commit()
            custom_popup(self, "Success", f"Kit '{kit_code}' duplicated (PPP={new_ppp}).", "success")
            self.refresh_current_tree(preserve_view=True)
        except Exception as e:
            self.conn.rollback()
            custom_popup(self, "Error", f"Kit duplication failed: {e}", "error")

    def duplicate_selected_module(self):
        if not self._can_modify():
            self._deny()
            return
        sel = self.tree.selection()
        if not sel:
            custom_popup(self, "Error", "Select a MODULE to duplicate.", "error")
            return
        node = sel[0]
        tags = self.tree.item(node, "tags")
        if not tags or tags[0].upper() != "MODULE":
            custom_popup(self, "Error", "Selected node is not a MODULE.", "error")
            return
        if not self.selected_scenario_id:
            custom_popup(self, "Error", "No scenario selected.", "error")
            return
        tc = self.node_treecode.get(node)
        if not tc:
            custom_popup(self, "Error", "Missing treecode.", "error")
            return
        segs = self._parse_treecode(tc)
        if not segs:
            custom_popup(self, "Error", "Invalid treecode format.", "error")
            return
        level = self.tree.item(node, "values")[1]
        base_parts = self._normalize_iid(node).split("_")
        try:
            self.cursor.execute("BEGIN")
            if level == "primary":
                module_code = base_parts[1]
                new_ppp = self._next_primary_ppp()
                if not new_ppp:
                    raise ValueError("No PPP available")
                ss = f"{self.selected_scenario_id:02d}"
                new_mod_tc = f"{ss}{new_ppp}000000"
                self.cursor.execute("""
                    INSERT INTO kit_items
                        (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                    VALUES (?, ?, ?, ?, ?, ?, 1, 'primary', ?)
                """, (self.selected_scenario_id, self.selected_scenario,
                      None, module_code, None, module_code, new_mod_tc))
                self.cursor.execute("""
                    SELECT DISTINCT item, code, std_qty
                      FROM kit_items
                     WHERE scenario_id=? AND module=? AND kit IS NULL AND level='secondary'
                """, (self.selected_scenario_id, module_code))
                sec_items = self.cursor.fetchall()
                for it in sec_items:
                    new_mmm = self._next_mmm_in_ppp(new_ppp)
                    if not new_mmm:
                        raise ValueError("No MMM for items")
                    item_tc = f"{ss}{new_ppp}{new_mmm}000"
                    self.cursor.execute("""
                        INSERT INTO kit_items
                            (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'secondary', ?)
                    """, (self.selected_scenario_id, self.selected_scenario,
                          None, module_code, it['item'], it['code'], it['std_qty'], item_tc))
                custom_popup(self, "Success", f"Module '{module_code}' duplicated (PPP={new_ppp}).", "success")
            else:
                kit_code = base_parts[1] if base_parts[1] != 'none' else None
                module_code = base_parts[2]
                ppp = segs["PPP"]
                new_mmm = self._next_mmm_in_ppp(ppp)
                if not new_mmm:
                    raise ValueError("No MMM available")
                ss = f"{self.selected_scenario_id:02d}"
                module_tc = f"{ss}{ppp}{new_mmm}000"
                self.cursor.execute("""
                    INSERT INTO kit_items
                        (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                    VALUES (?, ?, ?, ?, ?, ?, 1, 'secondary', ?)
                """, (self.selected_scenario_id, self.selected_scenario,
                      kit_code, module_code, None, module_code, module_tc))
                self.cursor.execute("""
                    SELECT DISTINCT item, code, std_qty
                      FROM kit_items
                     WHERE scenario_id=? AND kit=? AND module=? AND level='tertiary'
                """, (self.selected_scenario_id, kit_code, module_code))
                items = self.cursor.fetchall()
                iii_gen = self._iter_new_iii(ppp, new_mmm)
                for it in items:
                    iii = next(iii_gen)
                    item_tc = f"{ss}{ppp}{new_mmm}{iii}"
                    self.cursor.execute("""
                        INSERT INTO kit_items
                            (scenario_id, scenario, kit, module, item, code, std_qty, level, treecode)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'tertiary', ?)
                    """, (self.selected_scenario_id, self.selected_scenario,
                          kit_code, module_code, it['item'], it['code'], it['std_qty'], item_tc))
                custom_popup(self, "Success", f"Module '{module_code}' duplicated (PPP={ppp}, MMM={new_mmm}).", "success")
            self.conn.commit()
            self.refresh_current_tree(preserve_view=True)
        except Exception as e:
            self.conn.rollback()
            custom_popup(self, "Error", f"Module duplication failed: {e}", "error")

    # ---------------- Positioning Helper ----------------
    def _center_toplevel(self, win: Toplevel, width=300, height=180):
        win.update_idletasks()
        parent = self.winfo_toplevel()
        try:
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            x = px + (pw // 2) - (width // 2)
            y = py + (ph // 2) - (height // 2)
        except Exception:
            sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
            x = (sw // 2) - (width // 2)
            y = (sh // 2) - (height // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")

    # ---------------- Cleanup ----------------
    def destroy(self):
        try:
            self.conn.close()
        except Exception:
            pass
        super().destroy()


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Kits Composition (Test)")
    class Dummy: pass
    d = Dummy()
    d.role = "$"  # test read-only
    KitsComposition(root, d)
    root.geometry("1180x820")
    root.mainloop()