"""
order.py  v1.8

FIXED in v1.8:
  * FIXED: Column widths now truly preserved - only reconfigure columns when necessary
  * Separated data population from column configuration
  * Columns only reset on mode change or initial load
  * Double-click cells to edit (Back Orders, Loan Balance, Planned Give, Donations Receive, Qty To Order, Remarks)
  * Right-click for Stock Card
  * Auto-calculation on edit

FORMULA FOR QTY_NEEDED:
  qty_needed = standard_qty - current_stock + qty_expiring - back_orders
               - loan_balance + planned_dons_give - dons_receive
  (clamped to minimum 0)
"""

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import math
import time
from datetime import date, datetime
from calendar import monthrange
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    from tkcalendar import DateEntry

    TKCAL_AVAILABLE = True
except Exception:
    TKCAL_AVAILABLE = False

from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang
from popup_utils import custom_popup
from theme_config import AppTheme, configure_tree_tags, enable_column_auto_resize

KIT_FILL_COLOR = "228B22"
MODULE_FILL_COLOR = "ADD8E6"

LOAN_OUT_TYPES = {"Loan", "Return of Borrowing"}
LOAN_IN_TYPES = {"In Borrowing", "In Return of Loan"}

INT_ENTRY_MIN = 0
INT_ENTRY_MAX = 24


def add_months(base: date, months: int) -> date:
    if months <= 0:
        return base
    y = base.year + (base.month - 1 + months) // 12
    m = (base.month - 1 + months) % 12 + 1
    d = base.day
    last = monthrange(y, m)[1]
    if d > last:
        d = last
    return date(y, m, d)


def fetch_project_settings():
    conn = connect_db()
    if conn is None:
        return 0, 0, 0
    cur = conn.cursor()
    lead = cover = buf = 0
    try:
        cur.execute("PRAGMA table_info(project_details)")
        cols = {r[1].lower(): r[1] for r in cur.fetchall()}
        needed = []
        if "lead_time_months" in cols:
            needed.append("lead_time_months")
        if "cover_period_months" in cols:
            needed.append("cover_period_months")
        if "buffer_months" in cols:
            needed.append("buffer_months")
        if not needed:
            return 0, 0, 0
        cur.execute(f"SELECT {', '.join(needed)} FROM project_details LIMIT 1")
        row = cur.fetchone()
        if row:
            idx = 0
            if "lead_time_months" in needed:
                lead = int(row[idx] or 0)
                idx += 1
            if "cover_period_months" in needed:
                cover = int(row[idx] or 0)
                idx += 1
            if "buffer_months" in needed:
                buf = int(row[idx] or 0)
    except Exception:
        pass
    finally:
        cur.close()
        conn.close()
    return tuple(max(0, min(24, x)) for x in (lead, cover, buf))


# ---------- Data Aggregator ----------
class OrderData:
    def __init__(
        self, kit_filter, module_filter, type_filter, item_search, lead, cover, buffer
    ):
        self.kit_filter = kit_filter
        self.module_filter = module_filter
        self.type_filter = type_filter
        self.item_search = (item_search or "").strip()
        self.lead = lead
        self.cover = cover
        self.buffer = buffer
        self.horizon_months = (lead or 0) + (cover or 0) + (buffer or 0)
        self.horizon_end = add_months(date.today(), self.horizon_months)

    def fetch(self):
        std_map = self._fetch_standard_qty()
        stock_map = self._fetch_current_stock()
        exp_map = self._fetch_expiring_qty()
        loan_map = self._fetch_loan_balance()
        commercial = self._fetch_commercial_data(
            set().union(std_map, stock_map, exp_map, loan_map)
        )

        codes = set(std_map) | set(stock_map) | set(exp_map) | set(loan_map)
        rows = []
        for code in sorted(codes):
            desc = get_item_description(code)
            dtype = detect_type(code, desc)
            if self.type_filter and self.type_filter.lower() != "all":
                if dtype.lower() != self.type_filter.lower():
                    continue
            if self.item_search:
                if dtype.lower() != "item":
                    continue
                if (
                    self.item_search.lower() not in code.lower()
                    and self.item_search.lower() not in desc.lower()
                ):
                    continue
            cdat = commercial.get(code, {})
            row = {
                "code": code,
                "description": desc,
                "type": dtype,
                "standard_qty": std_map.get(code, 0),
                "current_stock": stock_map.get(code, 0),
                "qty_expiring": exp_map.get(code, 0),
                "back_orders": 0,
                "loan_balance": loan_map.get(code, 0),
                "planned_dons_give": 0,
                "dons_receive": 0,
                "pack_size": cdat.get("pack_size", 0),
                "qty_needed": 0,
                "qty_to_order": "",
                "qty_to_order_rounded": 0,
                "price_per_pack": cdat.get("price", 0.0),
                "weight_per_pack": cdat.get("weight", 0.0),
                "volume_per_pack_dm3": cdat.get("volume", 0.0),
                "amount": 0.0,
                "weight_kg": 0.0,
                "volume_m3": 0.0,
                "account_code": cdat.get("account", ""),
                "remarks": "",
            }
            rows.append(row)
        return rows

    def _fetch_standard_qty(self):
        res = defaultdict(int)
        conn = connect_db()
        if not conn:
            return res
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(std_qty_helper)")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if "code" not in cols or "std_qty" not in cols:
                return res
            filters = []
            params = []
            if self.kit_filter and self.kit_filter.lower() != "all" and "kit" in cols:
                filters.append("kit=?")
                params.append(self.kit_filter)
            if (
                self.module_filter
                and self.module_filter.lower() != "all"
                and "module" in cols
            ):
                filters.append("module=?")
                params.append(self.module_filter)
            where = "WHERE " + " AND ".join(filters) if filters else ""
            cur.execute(f"SELECT code,std_qty FROM std_qty_helper {where}", params)
            for c, q in cur.fetchall():
                if c:
                    try:
                        res[c] += int(q or 0)
                    except:
                        pass
        finally:
            cur.close()
            conn.close()
        return res

    def _fetch_current_stock(self):
        res = defaultdict(int)
        conn = connect_db()
        if not conn:
            return res
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_data)")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if "final_qty" not in cols:
                return res
            has_code = "code" in cols
            code_col = "code" if has_code else "unique_id"
            filters = ["final_qty IS NOT NULL"]
            params = []
            if (
                self.kit_filter
                and self.kit_filter.lower() != "all"
                and "kit_number" in cols
            ):
                filters.append("kit_number=?")
                params.append(self.kit_filter)
            if (
                self.module_filter
                and self.module_filter.lower() != "all"
                and "module_number" in cols
            ):
                filters.append("module_number=?")
                params.append(self.module_filter)
            where = "WHERE " + " AND ".join(filters)
            cur.execute(f"SELECT {code_col}, final_qty FROM stock_data {where}", params)
            for c, q in cur.fetchall():
                code = c if has_code else self._extract_code_from_unique_id(c)
                if code:
                    try:
                        res[code] += int(q or 0)
                    except:
                        pass
        finally:
            cur.close()
            conn.close()
        return res

    def _fetch_expiring_qty(self):
        res = defaultdict(int)
        if self.horizon_months <= 0:
            return res
        conn = connect_db()
        if not conn:
            return res
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_data)")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if not {"final_qty", "exp_date"} <= set(cols):
                return res
            has_code = "code" in cols
            code_col = "code" if has_code else "unique_id"
            horizon_end = self.horizon_end.strftime("%Y-%m-%d")
            filters = [
                "final_qty IS NOT NULL",
                "exp_date IS NOT NULL",
                "exp_date!=''",
                "exp_date <= ?",
            ]
            params = [horizon_end]
            if (
                self.kit_filter
                and self.kit_filter.lower() != "all"
                and "kit_number" in cols
            ):
                filters.append("kit_number=?")
                params.append(self.kit_filter)
            if (
                self.module_filter
                and self.module_filter.lower() != "all"
                and "module_number" in cols
            ):
                filters.append("module_number=?")
                params.append(self.module_filter)
            where = "WHERE " + " AND ".join(filters)
            cur.execute(f"SELECT {code_col}, final_qty FROM stock_data {where}", params)
            for c, q in cur.fetchall():
                code = c if has_code else self._extract_code_from_unique_id(c)
                if code:
                    try:
                        res[code] += int(q or 0)
                    except:
                        pass
        finally:
            cur.close()
            conn.close()
        return res

    def _fetch_loan_balance(self):
        res = defaultdict(int)
        conn = connect_db()
        if not conn:
            return res
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_transactions)")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if not {"qty_in", "qty_out", "in_type", "out_type", "code"} <= set(cols):
                return res
            filters = []
            params = []
            if self.kit_filter and self.kit_filter.lower() != "all" and "kit" in cols:
                filters.append("Kit=?")
                params.append(self.kit_filter)
            if (
                self.module_filter
                and self.module_filter.lower() != "all"
                and "module" in cols
            ):
                filters.append("Module=?")
                params.append(self.module_filter)

            loan_condition = f"(Out_Type IN ({','.join('?'*len(LOAN_OUT_TYPES))}) OR IN_Type IN ({','.join('?'*len(LOAN_IN_TYPES))}))"
            params_all = params + list(LOAN_OUT_TYPES) + list(LOAN_IN_TYPES)
            if filters:
                where = "WHERE " + " AND ".join(filters) + f" AND {loan_condition}"
            else:
                where = "WHERE " + loan_condition

            cur.execute(
                f"""
                SELECT code, Qty_IN, IN_Type, Qty_Out, Out_Type
                FROM stock_transactions
                {where}
            """,
                params_all,
            )
            for code, q_in, in_type, q_out, out_type in cur.fetchall():
                if not code:
                    continue
                if out_type in LOAN_OUT_TYPES and q_out:
                    try:
                        res[code] += int(q_out)
                    except:
                        pass
                if in_type in LOAN_IN_TYPES and q_in:
                    try:
                        res[code] -= int(q_in)
                    except:
                        pass
        except Exception:
            pass
        finally:
            cur.close()
            conn.close()
        return res

    def _fetch_commercial_data(self, codes):
        res = {}
        if not codes:
            return res
        conn = connect_db()
        if not conn:
            return res
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(items_list)")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if "code" not in cols:
                return res
            field_map = {
                "pack": "pack_size",
                "price_per_pack_euros": "price",
                "weight_per_pack_kg": "weight",
                "volume_per_pack_dm3": "volume",
                "account_code": "account",
            }
            sel = ["code"] + [c for c in field_map if c in cols]
            placeholders = ",".join("?" for _ in codes)
            cur.execute(
                f"SELECT {', '.join(sel)} FROM items_list WHERE code IN ({placeholders})",
                list(codes),
            )
            idx = {name: i for i, name in enumerate(sel)}
            for row in cur.fetchall():
                code = row[idx["code"]]
                data = {}
                for src, dst in field_map.items():
                    if src in idx:
                        val = row[idx[src]]
                        try:
                            if dst in ("pack_size",):
                                data[dst] = int(val or 0)
                            elif dst in ("price", "weight", "volume"):
                                data[dst] = float(val or 0)
                            else:
                                data[dst] = val or ""
                        except:
                            data[dst] = 0 if dst != "account" else ""
                res[code] = data
        finally:
            cur.close()
            conn.close()
        return res

    @staticmethod
    def _extract_code_from_unique_id(u):
        try:
            p = u.split("/")
            if len(p) >= 4 and p[3] != "None":
                return p[3]
            if len(p) >= 3 and p[2] != "None":
                return p[2]
            if len(p) >= 2 and p[1] != "None":
                return p[1]
            return u
        except:
            return u


# ---------- UI ----------
class OrderNeeds(tk.Frame):
    SIMPLE_COLS = [
        "code",
        "description",
        "standard_qty",
        "current_stock",
        "qty_needed",
        "qty_to_order_rounded",
        "amount",
        "weight_kg",
        "volume_m3",
    ]
    DETAIL_COLS = [
        "code",
        "description",
        "type",
        "standard_qty",
        "current_stock",
        "qty_expiring",
        "back_orders",
        "loan_balance",
        "planned_dons_give",
        "dons_receive",
        "pack_size",
        "qty_needed",
        "qty_to_order",
        "qty_to_order_rounded",
        "price_per_pack",
        "weight_per_pack",
        "volume_per_pack_dm3",
        "amount",
        "weight_kg",
        "volume_m3",
        "account_code",
        "remarks",
    ]
    EDITABLE_COLS = {
        "back_orders",
        "loan_balance",
        "planned_dons_give",
        "dons_receive",
        "qty_to_order",
        "remarks",
    }

    def __init__(self, parent, app, *args, **kwargs):
        super().__init__(parent, bg=AppTheme.BG_MAIN, *args, **kwargs)
        self.app = app
        self.rows = []
        self.simple_mode = False

        self.kit_var = tk.StringVar(value="All")
        self.module_var = tk.StringVar(value="All")
        self.type_var = tk.StringVar(value="All")
        self.item_search_var = tk.StringVar()
        self.lead_var = tk.StringVar()
        self.cover_var = tk.StringVar()
        self.buffer_var = tk.StringVar()

        self.total_amount_var = tk.StringVar(value="0.00")
        self.total_weight_var = tk.StringVar(value="0.00")
        self.total_volume_var = tk.StringVar(value="0.000")
        self.missing_price_var = tk.StringVar(value="")

        self.tree = None
        self.edit_entry = None
        self.status_var = tk.StringVar(value=lang.t("order_needs.ready", "Ready"))
        self._selected_item_values = None

        # Double-click tracking
        self._last_click_time = 0
        self._last_click_item = None
        self._last_click_column = None

        # FIXED: Track current columns to avoid unnecessary reconfiguration
        self._current_tree_columns = None

        self._build_ui()
        self._load_initial_params()
        self.populate_dropdowns()
        self.refresh()

    def _build_ui(self):
        header = tk.Frame(self, bg=AppTheme.BG_MAIN)
        header.pack(fill="x", padx=12, pady=(12, 6))

        title_frame = tk.Frame(header, bg=AppTheme.BG_MAIN)
        title_frame.pack(side="left", fill="x", expand=True)

        tk.Label(
            title_frame,
            text=lang.t("menu.reports.order_needs", "Order / Needs"),
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_HUGE, "bold"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(side="left")

        info_btn = tk.Button(
            title_frame,
            text="ℹ️",
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            relief="flat",
            font=(AppTheme.FONT_FAMILY, 12),
            cursor="hand2",
            command=self._show_formula_info,
        )
        info_btn.pack(side="left", padx=10)

        self.toggle_btn = tk.Button(
            header,
            text=lang.t("order_needs.detailed", "Detailed"),
            bg=AppTheme.BTN_TOGGLE,
            fg=AppTheme.TEXT_WHITE,
            relief="flat",
            padx=14,
            pady=6,
            command=self.toggle_mode,
        )
        self.toggle_btn.pack(side="right", padx=(6, 0))
        tk.Button(
            header,
            text=lang.t("generic.export", "Export"),
            bg=AppTheme.BTN_EXPORT,
            fg=AppTheme.TEXT_WHITE,
            relief="flat",
            padx=14,
            pady=6,
            command=self.export_excel,
        ).pack(side="right", padx=(6, 0))
        tk.Button(
            header,
            text=lang.t("generic.clear", "Clear"),
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            relief="flat",
            padx=14,
            pady=6,
            command=self.clear_all,
        ).pack(side="right", padx=(6, 0))
        tk.Button(
            header,
            text=lang.t("generic.refresh", "Refresh"),
            bg=AppTheme.BTN_REFRESH,
            fg=AppTheme.TEXT_WHITE,
            relief="flat",
            padx=14,
            pady=6,
            command=self.refresh,
        ).pack(side="right", padx=(6, 0))

        # Filters
        filt = tk.Frame(self, bg=AppTheme.BG_MAIN)
        filt.pack(fill="x", padx=12, pady=(0, 6))
        r1 = tk.Frame(filt, bg=AppTheme.BG_MAIN)
        r1.pack(fill="x", pady=2)
        tk.Label(
            r1, text=lang.t("generic.kit_number", "Kit Number"), bg=AppTheme.BG_MAIN
        ).grid(row=0, column=0, sticky="w", padx=(0, 4))
        self.kit_cb = ttk.Combobox(
            r1, textvariable=self.kit_var, state="readonly", width=18
        )
        self.kit_cb.grid(row=0, column=1, padx=(0, 12))
        self.kit_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        tk.Label(
            r1,
            text=lang.t("generic.module_number", "Module Number"),
            bg=AppTheme.BG_MAIN,
        ).grid(row=0, column=2, sticky="w", padx=(0, 4))
        self.module_cb = ttk.Combobox(
            r1, textvariable=self.module_var, state="readonly", width=18
        )
        self.module_cb.grid(row=0, column=3, padx=(0, 12))
        self.module_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        tk.Label(r1, text=lang.t("generic.type", "Type"), bg=AppTheme.BG_MAIN).grid(
            row=0, column=4, sticky="w", padx=(0, 4)
        )
        self.type_cb = ttk.Combobox(
            r1,
            textvariable=self.type_var,
            state="readonly",
            width=12,
            values=["All", "Kit", "Module", "Item"],
        )
        self.type_cb.grid(row=0, column=5, padx=(0, 12))
        self.type_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        tk.Label(
            r1,
            text=lang.t("order_needs.item_search", "Item Search"),
            bg=AppTheme.BG_MAIN,
        ).grid(row=0, column=6, sticky="w", padx=(0, 4))
        self.item_entry = tk.Entry(r1, textvariable=self.item_search_var, width=20)
        self.item_entry.grid(row=0, column=7, padx=(0, 12))
        self.item_entry.bind("<Return>", lambda e: self.refresh())
        self.item_entry.bind(
            "<Escape>", lambda e: self._clear_field(self.item_search_var)
        )

        # Parameters
        r2 = tk.Frame(filt, bg=AppTheme.BG_MAIN)
        r2.pack(fill="x", pady=2)
        tk.Label(
            r2,
            text=lang.t("order_needs.lead_time", "Lead Time (months)"),
            bg=AppTheme.BG_MAIN,
        ).grid(row=0, column=0, sticky="w", padx=(0, 4))
        self.lead_entry = tk.Entry(r2, textvariable=self.lead_var, width=6)
        self.lead_entry.grid(row=0, column=1, padx=(0, 12))
        self.lead_entry.bind("<FocusOut>", lambda e: self._param_focus_refresh())
        self.lead_entry.bind("<Return>", lambda e: self._param_focus_refresh())

        tk.Label(
            r2,
            text=lang.t("order_needs.cover_period", "Cover Period (months)"),
            bg=AppTheme.BG_MAIN,
        ).grid(row=0, column=2, sticky="w", padx=(0, 4))
        self.cover_entry = tk.Entry(r2, textvariable=self.cover_var, width=6)
        self.cover_entry.grid(row=0, column=3, padx=(0, 12))
        self.cover_entry.bind("<FocusOut>", lambda e: self._param_focus_refresh())
        self.cover_entry.bind("<Return>", lambda e: self._param_focus_refresh())

        tk.Label(
            r2,
            text=lang.t("order_needs.buffer", "Security Stock (buffer, months)"),
            bg=AppTheme.BG_MAIN,
        ).grid(row=0, column=4, sticky="w", padx=(0, 4))
        self.buffer_entry = tk.Entry(r2, textvariable=self.buffer_var, width=6)
        self.buffer_entry.grid(row=0, column=5, padx=(0, 12))
        self.buffer_entry.bind("<FocusOut>", lambda e: self._param_focus_refresh())
        self.buffer_entry.bind("<Return>", lambda e: self._param_focus_refresh())

        # Instruction label
        instruction_frame = tk.Frame(self, bg=AppTheme.BG_MAIN)
        instruction_frame.pack(fill="x", padx=12, pady=(0, 2))
        tk.Label(
            instruction_frame,
            text="💡 Double-click ✎ cells to edit | Double-click column headers to auto-fit | Right-click rows for Stock Card",
            font=(AppTheme.FONT_FAMILY, 9, "bold"),
            bg=AppTheme.BG_MAIN,
            fg="#2563EB",
        ).pack(side="left")

        # Totals Row
        info = tk.Frame(self, bg=AppTheme.BG_MAIN)
        info.pack(fill="x", padx=12, pady=(2, 4))
        tk.Label(
            info,
            text=lang.t("order_needs.total_amount", "Total Amount (€):"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 4))
        tk.Label(
            info,
            textvariable=self.total_amount_var,
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=1, padx=(0, 14), sticky="w")

        tk.Label(
            info,
            text=lang.t("order_needs.total_weight", "Total Weight (kg):"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
        ).grid(row=0, column=2, sticky="w", padx=(0, 4))
        tk.Label(
            info,
            textvariable=self.total_weight_var,
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=3, padx=(0, 14), sticky="w")

        tk.Label(
            info,
            text=lang.t("order_needs.total_volume", "Total Volume (m3):"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
        ).grid(row=0, column=4, sticky="w", padx=(0, 4))
        tk.Label(
            info,
            textvariable=self.total_volume_var,
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=5, padx=(0, 14), sticky="w")

        tk.Label(
            info,
            textvariable=self.missing_price_var,
            bg=AppTheme.BG_MAIN,
            fg="#B45309",
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_SMALL, "italic"),
        ).grid(row=0, column=6, sticky="w")

        tk.Label(
            self,
            textvariable=self.status_var,
            anchor="w",
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
            relief="sunken",
        ).pack(fill="x", padx=12, pady=(0, 8))

        # Table
        frame = tk.Frame(self, bg=AppTheme.COLOR_BORDER, bd=1, relief="solid")
        frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.tree = ttk.Treeview(frame, columns=(), show="headings", height=24)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        hsb.pack(fill="x", padx=12, pady=(0, 12))
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Style configuration
        style = ttk.Style()
        style.configure(
            "Treeview",
            background=AppTheme.BG_PANEL,
            fieldbackground=AppTheme.BG_PANEL,
            foreground=AppTheme.COLOR_PRIMARY,
            rowheight=24,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL),
        )
        style.configure(
            "Treeview.Heading",
            background="#E5E8EB",
            foreground=AppTheme.COLOR_PRIMARY,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_HEADING, "bold"),
        )
        self.tree.tag_configure("alt", background=AppTheme.ROW_ALT)
        self.tree.tag_configure(
            "kitrow", background=AppTheme.KIT_COLOR, foreground=AppTheme.TEXT_WHITE
        )
        self.tree.tag_configure("modrow", background=AppTheme.MODULE_COLOR)
        self.tree.tag_configure(
            "remarks_italic",
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "italic"),
        )

        # Event bindings
        self.tree.bind("<ButtonRelease-1>", self._on_click_detect_double)
        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Motion>", self._on_tree_motion)

        # Context menu
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(
            label=lang.t("stock_summary.show_stock_card", "Show Stock Card"),
            command=self._show_stock_card_from_menu,
        )

        self.bind_all("<Escape>", self._global_esc)

    def _on_click_detect_double(self, event):
        """Detect double-click using time tracking"""
        current_time = time.time()
        item_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)

        # Check if it's a double-click (within 500ms on same cell)
        if (
            current_time - self._last_click_time < 0.5
            and item_id == self._last_click_item
            and col_id == self._last_click_column
            and item_id
        ):
            # It's a double-click!
            self._handle_double_click(event, item_id, col_id)
            # Reset tracking
            self._last_click_time = 0
            self._last_click_item = None
            self._last_click_column = None
        else:
            # Single click - update tracking
            self._last_click_time = current_time
            self._last_click_item = item_id
            self._last_click_column = col_id

    def _handle_double_click(self, event, row_id, col_id):
        """Handle the actual double-click editing"""
        # Destroy any existing entry
        if self.edit_entry:
            try:
                self.edit_entry.destroy()
            except:
                pass
            self.edit_entry = None

        if not row_id or not col_id:
            return

        # Get column name
        col_index = int(col_id.replace("#", "")) - 1
        cols = self._current_columns()

        if col_index < 0 or col_index >= len(cols):
            return

        col_name = cols[col_index]

        # Check if editable
        if col_name not in self.EDITABLE_COLS:
            return

        # Get bbox
        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return

        x, y, w, h = bbox

        # Get current value
        current_value = self.tree.set(row_id, col_name)

        # Create entry
        entry = tk.Entry(
            self.tree,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL),
            bg="#FFFFCC",  # Light yellow
            fg="black",
            relief="solid",
            bd=2,
        )

        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, str(current_value))
        entry.focus_set()
        entry.select_range(0, tk.END)
        entry.icursor(tk.END)

        self.edit_entry = entry
        self.status_var.set(
            f"✎ Editing {col_name}... Press Enter to save, Escape to cancel"
        )

        # Save handler
        def save_edit(evt=None):
            new_value = entry.get()
            try:
                entry.destroy()
            except:
                pass
            self.edit_entry = None
            self._apply_edit(row_id, col_name, new_value)

        # Cancel handler
        def cancel_edit(evt=None):
            try:
                entry.destroy()
            except:
                pass
            self.edit_entry = None
            self.status_var.set("Edit cancelled")

        entry.bind("<Return>", save_edit)
        entry.bind("<KP_Enter>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", save_edit)

    def _on_tree_motion(self, event):
        """Change cursor when hovering over editable columns"""
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            col_id = self.tree.identify_column(event.x)
            if col_id:
                col_index = int(col_id.replace("#", "")) - 1
                cols = self._current_columns()
                if 0 <= col_index < len(cols):
                    col_name = cols[col_index]
                    if col_name in self.EDITABLE_COLS:
                        self.tree.config(cursor="hand2")
                        return
        self.tree.config(cursor="")

    def _show_formula_info(self):
        """Show formula information popup"""
        formula_text = """
╔═══════════════════════════════════════════════════════════════╗
║                    ORDER CALCULATION FORMULA                    ║
╚═══════════════════════════════════════════════════════════════╝

📊 STEP 1: Calculate Quantity Needed
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
qty_needed = standard_qty - current_stock + qty_expiring 
             - back_orders - loan_balance + planned_dons_give 
             - dons_receive

(If result is negative, it's clamped to 0)

📦 STEP 2: Determine Quantity to Order
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
qty_to_order = qty_needed (default, but you can override)

📐 STEP 3: Round Up to Full Packs
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
qty_to_order_rounded = CEIL(qty_to_order / pack_size) × pack_size

💰 STEP 4: Calculate Costs & Logistics
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
number_of_packs = qty_to_order_rounded / pack_size
amount (€) = number_of_packs × price_per_pack
weight (kg) = number_of_packs × weight_per_pack
volume (m³) = (number_of_packs × volume_per_pack_dm³) / 1000

✏️ EDITABLE COLUMNS (Double-click to edit):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━���━━━━━━━━
• back_orders       • loan_balance
• planned_dons_give • dons_receive  
• qty_to_order      • remarks

All calculations update automatically after editing! ⚡
        """

        popup = tk.Toplevel(self)
        popup.title("Order Calculation Formula")
        popup.geometry("700x600")
        popup.configure(bg=AppTheme.BG_MAIN)

        tk.Label(
            popup,
            text="📐 Order Calculation Formula",
            font=(AppTheme.FONT_FAMILY, 16, "bold"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(pady=10)

        text_frame = tk.Frame(popup, bg="white", relief="sunken", bd=2)
        text_frame.pack(fill="both", expand=True, padx=20, pady=10)

        text_widget = tk.Text(
            text_frame,
            wrap="word",
            font=("Courier New", 10),
            bg="white",
            fg=AppTheme.COLOR_PRIMARY,
            padx=10,
            pady=10,
        )
        text_widget.pack(fill="both", expand=True)
        text_widget.insert("1.0", formula_text)
        text_widget.config(state="disabled")

        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side="right", fill="y")
        text_widget.config(yscrollcommand=scrollbar.set)

        tk.Button(
            popup,
            text="Close",
            command=popup.destroy,
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, 10, "bold"),
            padx=20,
            pady=5,
        ).pack(pady=10)

    def _on_right_click(self, event):
        """Show context menu on right-click"""
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        values = self.tree.item(item_id, "values")
        if not values or len(values) < 1:
            return

        self.tree.selection_set(item_id)
        self._selected_item_values = values

        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    def _show_stock_card_from_menu(self):
        """Open stock card from context menu"""
        if hasattr(self, "_selected_item_values"):
            self._open_stock_card(self._selected_item_values)

    def _open_stock_card(self, values):
        """Open stock card for selected item"""
        try:
            item_code = str(values[0]).strip()

            if not item_code:
                return

            from stock_card import StockCard

            win = tk.Toplevel(self)
            win.title(lang.t("stock_card.title", "Stock Card") + f" - {item_code}")
            win.geometry("1400x800")
            win.configure(bg="#F5F5F5")

            card = StockCard(
                win, win, role=getattr(self.app, "role", "user") if self.app else "user"
            )
            card.pack(fill="both", expand=True)

            def load_data():
                try:
                    if hasattr(card, "code_entry"):
                        card.code_entry.delete(0, tk.END)
                        card.code_entry.insert(0, item_code)

                    if hasattr(card, "search_items"):
                        card.search_items()

                        def auto_select():
                            try:
                                if (
                                    hasattr(card, "search_listbox")
                                    and card.search_listbox.size() > 0
                                ):
                                    card.search_listbox.selection_clear(0, tk.END)
                                    card.search_listbox.selection_set(0)
                                    card.search_listbox.event_generate(
                                        "<<ListboxSelect>>"
                                    )
                            except:
                                pass

                        win.after(300, auto_select)

                except Exception as e:
                    print(f"[ERROR] Load data failed: {e}")

            win.after(100, load_data)

        except ImportError:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                "stock_card.py module not found.",
                "error",
            )
        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                f"Error opening stock card: {str(e)}",
                "error",
            )

    def _configure_columns(self):
        """Configure tree columns - only call when columns change"""
        cols = self._current_columns()

        # Only reconfigure if columns actually changed
        if self._current_tree_columns == tuple(cols):
            return  # Columns haven't changed, skip reconfiguration

        self._current_tree_columns = tuple(cols)
        self.tree["columns"] = cols

        headers = {
            "code": lang.t("generic.code", "Code"),
            "description": lang.t("generic.description", "Description"),
            "type": lang.t("generic.type", "Type"),
            "standard_qty": lang.t("order_needs.standard_qty", "Standard Qty"),
            "current_stock": lang.t("order_needs.current_stock", "Current Stock"),
            "qty_expiring": lang.t("order_needs.qty_expiring", "Qty Expiring"),
            "back_orders": "✎ " + lang.t("order_needs.back_orders", "Back Orders"),
            "loan_balance": "✎ " + lang.t("order_needs.loan_balance", "Loan Balance"),
            "planned_dons_give": "✎ "
            + lang.t("order_needs.planned_dons_give", "Planned Give"),
            "dons_receive": "✎ " + lang.t("order_needs.dons_receive", "Donations Rcv"),
            "pack_size": lang.t("order_needs.pack_size", "Pack Size"),
            "qty_needed": lang.t("order_needs.qty_needed", "Qty Needed"),
            "qty_to_order": "✎ " + lang.t("order_needs.qty_to_order", "Qty To Order"),
            "qty_to_order_rounded": lang.t(
                "order_needs.qty_to_order_rounded", "Order Rounded"
            ),
            "price_per_pack": lang.t("order_needs.price_per_pack", "Price/Pack (€)"),
            "weight_per_pack": lang.t(
                "order_needs.weight_per_pack", "Weight/Pack (kg)"
            ),
            "volume_per_pack_dm3": lang.t(
                "order_needs.volume_per_pack", "Volume/Pack (dm³)"
            ),
            "amount": lang.t("order_needs.amount", "Amount (€)"),
            "weight_kg": lang.t("order_needs.weight", "Weight (kg)"),
            "volume_m3": lang.t("order_needs.volume_m3", "Volume (m³)"),
            "account_code": lang.t("order_needs.account_code", "Account Code"),
            "remarks": "✎ " + lang.t("order_needs.remarks", "Remarks"),
        }
        widths = {
            "code": 170,
            "description": 340,
            "type": 90,
            "standard_qty": 110,
            "current_stock": 110,
            "qty_expiring": 110,
            "back_orders": 120,
            "loan_balance": 140,
            "planned_dons_give": 140,
            "dons_receive": 140,
            "pack_size": 90,
            "qty_needed": 110,
            "qty_to_order": 130,
            "qty_to_order_rounded": 140,
            "price_per_pack": 110,
            "weight_per_pack": 130,
            "volume_per_pack_dm3": 140,
            "amount": 110,
            "weight_kg": 110,
            "volume_m3": 110,
            "account_code": 130,
            "remarks": 260,
        }
        for c in cols:
            self.tree.heading(c, text=headers.get(c, c))
            self.tree.column(c, width=widths.get(c, 120), anchor="w", stretch=True)

    def _populate_tree_data(self):
        """Populate tree data WITHOUT reconfiguring columns"""
        # Clear existing data
        self.tree.delete(*self.tree.get_children())

        cols = self._current_columns()

        # Insert new data
        for idx, r in enumerate(self.rows):
            vals = [self._format_cell(r, c) for c in cols]
            tag = "alt" if idx % 2 else ""
            dtype = r.get("type", "").upper()
            if dtype == "KIT":
                tag = "kitrow"
            elif dtype == "MODULE":
                tag = "modrow"
            if r.get("remarks") and "remarks" in cols:
                self.tree.insert("", "end", values=vals, tags=(tag, "remarks_italic"))
            else:
                self.tree.insert("", "end", values=vals, tags=(tag,))

    def _format_cell(self, row, col):
        v = row.get(col, "")
        if col in ("amount", "price_per_pack"):
            return f"{float(v):.2f}"
        if col in ("weight_kg", "weight_per_pack", "volume_per_pack_dm3"):
            return f"{float(v):.3f}"
        if col == "volume_m3":
            return f"{float(v):.4f}"
        if col == "qty_to_order" and v == "":
            return ""
        return v

    def refresh(self):
        od = OrderData(
            kit_filter=self.kit_var.get(),
            module_filter=self.module_var.get(),
            type_filter=self.type_var.get(),
            item_search=self.item_search_var.get(),
            lead=self._safe_int(self.lead_var.get()),
            cover=self._safe_int(self.cover_var.get()),
            buffer=self._safe_int(self.buffer_var.get()),
        )
        self.rows = od.fetch()
        for r in self.rows:
            self._recompute_row(r)

        # Configure columns first (only if changed)
        self._configure_columns()
        # Then populate data
        self._populate_tree_data()
        self._recompute_totals()

        # Enable auto-resize only on first load
        if not hasattr(self, "_initial_resize_done"):
            enable_column_auto_resize(self.tree)
            self._initial_resize_done = True

        self.status_var.set(
            lang.t(
                "order_needs.loaded",
                "Loaded {n} rows - Double-click editable columns (✎) to modify",
            ).format(n=len(self.rows))
        )

    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        self.toggle_btn.config(
            text=(
                lang.t("order_needs.simple", "Simple")
                if self.simple_mode
                else lang.t("order_needs.detailed", "Detailed")
            )
        )
        # Mode change requires column reconfiguration
        self._configure_columns()
        self._populate_tree_data()
        self._recompute_totals()

    def _apply_edit(self, iid, col_name, new_val):
        """Apply edit and auto-recalculate while preserving column widths"""
        code_index = self._current_columns().index("code")
        code = self.tree.item(iid, "values")[code_index]
        row = next((r for r in self.rows if r["code"] == code), None)
        if not row:
            return

        # Validate and apply
        if col_name in {
            "back_orders",
            "loan_balance",
            "planned_dons_give",
            "dons_receive",
        }:
            if new_val.strip() == "":
                row[col_name] = 0
            elif re.match(r"^-?\d+$", new_val.strip()):
                row[col_name] = int(new_val.strip())
            else:
                custom_popup(
                    self,
                    lang.t("generic.error", "Error"),
                    lang.t("order_needs.invalid_int", "Enter whole integer."),
                    "error",
                )
                return
        elif col_name == "qty_to_order":
            if new_val.strip() == "":
                row[col_name] = ""
            elif re.match(r"^-?\d+$", new_val.strip()):
                row[col_name] = int(new_val.strip())
            else:
                custom_popup(
                    self,
                    lang.t("generic.error", "Error"),
                    lang.t("order_needs.invalid_int", "Enter whole integer."),
                    "error",
                )
                return
        elif col_name == "remarks":
            row[col_name] = new_val

        # Recalculate
        self._recompute_row(row)
        # FIXED: Only repopulate data, don't reconfigure columns
        self._populate_tree_data()
        self._recompute_totals()

        self.status_var.set(
            f"✓ Updated {col_name} for {code} - All calculations refreshed!"
        )

    def _recompute_row(self, r):
        """Calculate order quantities"""
        std = r.get("standard_qty", 0)
        cur = r.get("current_stock", 0)
        exp = r.get("qty_expiring", 0)
        back = r.get("back_orders", 0)
        loan = r.get("loan_balance", 0)
        give = r.get("planned_dons_give", 0)
        rec = r.get("dons_receive", 0)

        qty_needed = std - cur + exp - back - loan + give - rec
        if qty_needed < 0:
            qty_needed = 0
        r["qty_needed"] = qty_needed

        q_to = r.get("qty_to_order", "")
        if q_to == "":
            q_to = qty_needed
        r["qty_to_order"] = q_to

        pack = r.get("pack_size", 0)
        if pack and pack > 0:
            rrounded = math.ceil(q_to / pack) * pack
        else:
            rrounded = q_to
        r["qty_to_order_rounded"] = rrounded

        pack_div = pack if pack and pack > 0 else None
        price = float(r.get("price_per_pack") or 0)
        wt = float(r.get("weight_per_pack") or 0)
        vol_dm3 = float(r.get("volume_per_pack_dm3") or 0)

        if pack_div:
            packs = rrounded / pack_div
            amount = packs * price
            weight = packs * wt
            volume_m3 = (packs * vol_dm3) / 1000
        else:
            amount = 0.0
            weight = 0.0
            volume_m3 = 0.0

        r["amount"] = amount
        r["weight_kg"] = weight
        r["volume_m3"] = volume_m3

    def _recompute_totals(self):
        total_amount = sum(r.get("amount", 0) for r in self.rows)
        total_weight = sum(r.get("weight_kg", 0) for r in self.rows)
        total_volume = sum(r.get("volume_m3", 0) for r in self.rows)
        missing = sum(1 for r in self.rows if (r.get("price_per_pack") or 0) == 0)
        self.total_amount_var.set(f"{total_amount:,.2f}")
        self.total_weight_var.set(f"{total_weight:,.2f}")
        self.total_volume_var.set(f"{total_volume:,.3f}")
        if missing:
            self.missing_price_var.set(
                lang.t(
                    "order_needs.missing_prices",
                    fallback="{n} items have missing price (0€ used)",
                ).format(n=missing)
            )
        else:
            self.missing_price_var.set("")

    def _param_focus_refresh(self):
        self._validate_int_field(self.lead_var)
        self._validate_int_field(self.cover_var)
        self._validate_int_field(self.buffer_var)
        self.refresh()

    def _load_initial_params(self):
        lead, cover, buffer = fetch_project_settings()
        self.lead_var.set(str(lead))
        self.cover_var.set(str(cover))
        self.buffer_var.set(str(buffer))

    def populate_dropdowns(self):
        self.kit_cb["values"] = ["All"] + self._distinct("stock_data", "kit_number")
        self.module_cb["values"] = ["All"] + self._distinct(
            "stock_data", "module_number"
        )

    def _distinct(self, table, column):
        conn = connect_db()
        if not conn:
            return []
        cur = conn.cursor()
        try:
            cur.execute(f"PRAGMA table_info({table})")
            cols = {r[1].lower(): r[1] for r in cur.fetchall()}
            if column.lower() not in cols:
                return []
            cur.execute(
                f"""
                SELECT DISTINCT {column} FROM {table}
                WHERE {column} IS NOT NULL AND {column}!='' AND {column}!='None'
                ORDER BY {column}
            """
            )
            return [r[0] for r in cur.fetchall()]
        except:
            return []
        finally:
            cur.close()
            conn.close()

    def clear_all(self):
        self.kit_var.set("All")
        self.module_var.set("All")
        self.type_var.set("All")
        self.item_search_var.set("")
        self._load_initial_params()
        self.refresh()

    def _current_columns(self):
        return self.SIMPLE_COLS if self.simple_mode else self.DETAIL_COLS

    def _validate_int_field(self, var):
        val = var.get().strip()
        if val == "" or not val.isdigit():
            var.set("0")
            return
        iv = int(val)
        if iv < INT_ENTRY_MIN:
            iv = INT_ENTRY_MIN
        if iv > INT_ENTRY_MAX:
            iv = INT_ENTRY_MAX
        var.set(str(iv))

    def _clear_field(self, var):
        var.set("")
        self.refresh()

    def _global_esc(self, event):
        if isinstance(event.widget, tk.Entry) and event.widget in (
            self.item_entry,
            self.lead_entry,
            self.cover_entry,
            self.buffer_entry,
        ):
            return
        if self.edit_entry:
            try:
                self.edit_entry.destroy()
            except:
                pass
            self.edit_entry = None
            self.status_var.set("Edit cancelled")

    @staticmethod
    def _safe_int(txt, default=0):
        try:
            return int(txt)
        except:
            return default

    def export_excel(self):
        if not self.rows:
            custom_popup(
                self,
                lang.t("generic.info", "Info"),
                lang.t("order_needs.no_data_export", "Nothing to export."),
                "warning",
            )
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title=lang.t("order_needs.export_title", "Save Order/Needs Report"),
            initialfile=f"OrderNeeds_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OrderNeeds"
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([lang.t("generic.generated", "Generated"), now])
            ws.append([lang.t("generic.filters_used", "Filters Used")])
            ws.append(
                [
                    "Kit",
                    self.kit_var.get(),
                    "Module",
                    self.module_var.get(),
                    "Type",
                    self.type_var.get(),
                ]
            )
            ws.append(
                [
                    "Item Search",
                    self.item_search_var.get(),
                    "Lead",
                    self.lead_var.get(),
                    "Cover",
                    self.cover_var.get(),
                ]
            )
            ws.append(
                [
                    "Buffer",
                    self.buffer_var.get(),
                    "Mode",
                    "Simple" if self.simple_mode else "Detailed",
                ]
            )
            ws.append([])
            ws.append(
                [
                    "Total Amount (€)",
                    self.total_amount_var.get(),
                    "Total Weight (kg)",
                    self.total_weight_var.get(),
                    "Total Volume (m3)",
                    self.total_volume_var.get(),
                    "Missing Price Rows",
                    self.missing_price_var.get(),
                ]
            )
            ws.append([])

            cols = self._current_columns()
            ws.append([c.replace("_", " ").title().replace("✎ ", "") for c in cols])

            kit_fill = PatternFill(
                start_color=KIT_FILL_COLOR, end_color=KIT_FILL_COLOR, fill_type="solid"
            )
            module_fill = PatternFill(
                start_color=MODULE_FILL_COLOR,
                end_color=MODULE_FILL_COLOR,
                fill_type="solid",
            )

            for r in self.rows:
                row_out = []
                for c in cols:
                    val = r.get(c, "")
                    if c in (
                        "amount",
                        "price_per_pack",
                        "weight_kg",
                        "weight_per_pack",
                        "volume_per_pack_dm3",
                        "volume_m3",
                    ):
                        val = float(val)
                    row_out.append(val)
                ws.append(row_out)
                dtype = r.get("type", "").upper()
                if dtype == "KIT":
                    for cell in ws[ws.max_row]:
                        cell.fill = kit_fill
                elif dtype == "MODULE":
                    for cell in ws[ws.max_row]:
                        cell.fill = module_fill
                if "remarks" in cols:
                    try:
                        idx = cols.index("remarks")
                        ws.cell(row=ws.max_row, column=idx + 1).font = Font(italic=True)
                    except:
                        pass

            for col in ws.columns:
                length = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > length:
                        length = len(val)
                ws.column_dimensions[letter].width = min(length + 2, 60)
            ws.freeze_panes = "A10"
            wb.save(path)
            custom_popup(
                self,
                lang.t("generic.success", "Success"),
                lang.t("order_needs.export_success", "Export completed: {f}").format(
                    f=path
                ),
                "info",
            )
        except Exception as e:
            custom_popup(
                self,
                lang.t("generic.error", "Error"),
                lang.t("order_needs.export_fail", "Export failed: {err}").format(
                    err=str(e)
                ),
                "error",
            )


__all__ = ["OrderNeeds"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Order / Needs v1.8 - TRULY Fixed Column Widths!")
    OrderNeeds(root, None).pack(fill="both", expand=True)
    root.geometry("1850x880")
    root.mainloop()
