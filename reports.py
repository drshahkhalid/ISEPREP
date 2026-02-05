"""
reports.py
Stock Statement Report
Real-time search + manual expiry period override + translation support.

NEW FEATURES:
  - Double-click row to view code details from std_qty_helper table
  - Summary box showing Total Amount (EUR), Total Weight (kg), Total Volume (m³) - TOP RIGHT
  - LANGUAGE-AWARE descriptions from items_list
  - AUTO-APPLY FILTERS when any parameter changes (no need to click Load/Refresh)

Adjustments in this version:
  - Heading is now translation-enabled (reports.heading_stock_statement).
  - "Expiry Horizon" renamed everywhere in UI to "Expiry Period".
  - Added new translation keys (reports.expiry_period, reports.expiry_period_recommended, etc.)
  - Backward compatibility: if older translation files only have expiry_horizon keys, code provides fallbacks.
  - All functions from previous version retained (no removals).

Features:
  - Refreshes snapshot tables (std_list_combined, std_qty_helper) each Load.
  - Manual Expiry Period override (months); blank = recommended (lead+cover+buffer).
  - Real-time search (filters the populated table only, no extra DB hits).
  - Translation aware (language_manager.lang.t).
  - Aggregations handle kit/module/item codes from stock_data (no direct 'code' column).
  - Compatible with very old SQLite (no CTE/window functions).

Missing Dependencies:
  - db.connect_db()
  - language_manager.lang
  - popup_utils.custom_popup (falls back to simple Tk messagebox if absent)
"""

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import datetime
from calendar import monthrange
import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from db import connect_db
from language_manager import lang

try:
    from popup_utils import custom_popup
except ImportError:
    from tkinter import messagebox

    def custom_popup(parent, title, msg, kind="info"):
        if kind == "error":
            messagebox.showerror(title, msg, parent=parent)
        elif kind == "warning":
            messagebox.showwarning(title, msg, parent=parent)
        else:
            messagebox.showinfo(title, msg, parent=parent)


# ------------------------------------------------------------------
# Language-aware designation helper
# ------------------------------------------------------------------
def _get_designation_field():
    """Return the appropriate designation field name based on current language"""
    lang_code = getattr(lang, "lang_code", "en").lower()
    if lang_code == "fr":
        return "designation_fr"
    elif lang_code in ("es", "sp"):
        return "designation_sp"
    else:
        return "designation_en"


def _build_designation_sql():
    """Build SQL COALESCE expression for language-aware designation with fallback"""
    lang_code = getattr(lang, "lang_code", "en").lower()

    if lang_code == "fr":
        return "COALESCE(il.designation_fr, il.designation_en, il.designation)"
    elif lang_code in ("es", "sp"):
        return "COALESCE(il.designation_sp, il.designation_en, il.designation)"
    else:
        return "COALESCE(il.designation_en, il.designation)"


# ------------------------------------------------------------------
# DB Helpers
# ------------------------------------------------------------------
def _fetchall(sql, params=()):
    conn = connect_db()
    if conn is None:
        return []
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        return cur.fetchall()
    except sqlite3.Error:
        return []
    finally:
        cur.close()
        conn.close()


def _fetchone(sql, params=()):
    rows = _fetchall(sql, params)
    return rows[0] if rows else None


# ------------------------------------------------------------------
# Snapshot Refresh (Option B real-time) - LANGUAGE AWARE
# ------------------------------------------------------------------
def _refresh_snapshots_internal():
    """Refresh snapshot tables with language-aware descriptions"""
    conn = connect_db()
    if conn is None:
        raise RuntimeError("DB connection failed for snapshot refresh")
    cur = conn.cursor()

    # Build language-aware designation expression
    designation_expr = _build_designation_sql()

    try:
        # Refresh std_list_combined
        cur.execute("DELETE FROM std_list_combined")
        sql_std_list = f"""
        INSERT INTO std_list_combined (code, description, type, std_qty_collective)
        SELECT
            all_codes.code,
            {designation_expr},
            il.type,
            SUM(all_codes.qty_component) AS std_qty_collective
        FROM (
            SELECT code, quantity AS qty_component FROM compositions
            UNION ALL
            SELECT code, std_qty AS qty_component FROM kit_items
        ) AS all_codes
        LEFT JOIN items_list il ON il.code = all_codes.code
        GROUP BY all_codes.code
        """
        cur.execute(sql_std_list)

        # Refresh std_qty_helper
        cur.execute("DELETE FROM std_qty_helper")
        sql_helper = f"""
        INSERT INTO std_qty_helper (id, code, description, type, scenario_id, scenario, kit, module, std_qty)
        SELECT
          c.std_id,
          c.code,
          {designation_expr.replace('il.', 'il.')},
          il.type,
          c.scenario_id,
          s.name,
          NULL,
          NULL,
          c.quantity
        FROM compositions c
        LEFT JOIN items_list il ON il.code = c.code
        LEFT JOIN scenarios s ON s.scenario_id = c.scenario_id
        UNION ALL
        SELECT
          CAST(k.id AS TEXT),
          k.code,
          {designation_expr.replace('il.', 'il.')},
          il.type,
          k.scenario_id,
          s.name,
          k.kit,
          k.module,
          k.std_qty
        FROM kit_items k
        LEFT JOIN items_list il ON il.code = k.code
        LEFT JOIN scenarios s ON s.scenario_id = k.scenario_id
        """
        cur.execute(sql_helper)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


# ------------------------------------------------------------------
# Period Logic
# ------------------------------------------------------------------
def compute_horizon_months():
    cols = {
        r["name"].lower(): True for r in _fetchall("PRAGMA table_info(project_details)")
    }
    needed = ["lead_time_months", "cover_period_months", "buffer_months"]
    if not all(n in cols for n in needed):
        return 0
    row = _fetchone(
        """
        SELECT
          COALESCE(lead_time_months,0) lt,
          COALESCE(cover_period_months,0) cp,
          COALESCE(buffer_months,0) bf
        FROM project_details
        LIMIT 1
    """
    )
    if not row:
        return 0
    return int(row["lt"]) + int(row["cp"]) + int(row["bf"])


def compute_cutoff_date(months):
    today = datetime.date.today()
    year = today.year
    month = today.month + months
    while month > 12:
        month -= 12
        year += 1
    last_day = monthrange(year, month)[1]
    return datetime.date(year, month, last_day).isoformat()


# ------------------------------------------------------------------
# Aggregations
# ------------------------------------------------------------------
def aggregate_std_qty(filters):
    where = []
    params = []
    if filters["scenario"]:
        where.append("scenario = ?")
        params.append(filters["scenario"])
    if filters["kit_code"]:
        where.append("(kit = ? OR (type='KIT' AND code=?))")
        params.extend([filters["kit_code"], filters["kit_code"]])
    if filters["module_code"]:
        where.append("(module = ? OR (type='MODULE' AND code=?))")
        params.extend([filters["module_code"], filters["module_code"]])
    if filters["item_code"]:
        where.append("code = ?")
        params.append(filters["item_code"])
    if filters["type_filter"] and filters["type_filter"].upper() in (
        "KIT",
        "MODULE",
        "ITEM",
    ):
        where.append("UPPER(type)=?")
        params.append(filters["type_filter"].upper())

    sql = "SELECT code, SUM(std_qty) AS total_std FROM std_qty_helper WHERE 1=1"
    if where:
        sql += " AND " + " AND ".join(where)
    sql += " GROUP BY code"
    rows = _fetchall(sql, tuple(params))
    out = {}
    for r in rows:
        out[r["code"]] = r["total_std"] if r["total_std"] is not None else 0
    return out


def aggregate_stock(filters, cutoff_iso):
    scenario = filters["scenario"]
    kit_code = filters["kit_code"]
    module_code = filters["module_code"]
    item_code = filters["item_code"]
    type_filter = (filters["type_filter"] or "All").upper()

    where_base = []
    params = []
    if scenario:
        where_base.append("scenario = ?")
        params.append(scenario)

    sql = "SELECT kit, module, item, final_qty, exp_date, scenario FROM stock_data"
    if where_base:
        sql += " WHERE " + " AND ".join(where_base)

    rows = _fetchall(sql, tuple(params))

    type_map = {}
    for r in _fetchall("SELECT code, type FROM items_list"):
        type_map[r["code"]] = (r["type"] or "").upper()

    out = {}
    for r in rows:
        final_qty = r["final_qty"] if r["final_qty"] is not None else 0
        exp_date = r["exp_date"]

        if r["kit"] and r["kit"].strip().lower() != "none":
            c = r["kit"]
            if not kit_code or kit_code == c:
                _try_accumulate(
                    out, c, final_qty, exp_date, cutoff_iso, type_filter, type_map
                )

        if r["module"] and r["module"].strip().lower() != "none":
            c = r["module"]
            if not module_code or module_code == c:
                _try_accumulate(
                    out, c, final_qty, exp_date, cutoff_iso, type_filter, type_map
                )

        if r["item"] and r["item"].strip().lower() != "none":
            c = r["item"]
            if not item_code or item_code == c:
                _try_accumulate(
                    out, c, final_qty, exp_date, cutoff_iso, type_filter, type_map
                )

    return out


def _try_accumulate(out, code, qty, exp_date, cutoff_iso, type_filter, type_map):
    if type_filter in ("KIT", "MODULE", "ITEM"):
        if type_map.get(code, "") != type_filter:
            return
    entry = out.setdefault(code, {"current_stock": 0, "expiring": 0})
    entry["current_stock"] += qty
    if (
        exp_date
        and cutoff_iso
        and re.match(r"^\d{4}-\d{2}-\d{2}$", exp_date)
        and exp_date <= cutoff_iso
    ):
        entry["expiring"] += qty


def load_item_metadata(codes):
    """Load item metadata with language-aware designation"""
    if not codes:
        return {}
    meta = {}
    conn = connect_db()
    if conn is None:
        return meta
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Build language-aware designation
    designation_expr = _build_designation_sql()

    try:
        clist = list(codes)
        chunk = 200
        for i in range(0, len(clist), chunk):
            part = clist[i : i + chunk]
            placeholders = ",".join(["?"] * len(part))
            sql = f"""
                SELECT code, {designation_expr} AS designation, type,
                       pack, price_per_pack_euros, unit_price_euros,
                       weight_per_pack_kg, volume_per_pack_dm3,
                       shelf_life_months, remarks, account_code
                FROM items_list il
                WHERE code IN ({placeholders})
            """
            cur.execute(sql, part)
            for r in cur.fetchall():
                meta[r["code"]] = r
    finally:
        cur.close()
        conn.close()
    return meta


# ------------------------------------------------------------------
# Role helper
# ------------------------------------------------------------------
def get_role_from_args(args, kwargs):
    if "role" in kwargs and isinstance(kwargs["role"], str):
        return kwargs["role"]
    if len(args) >= 2:
        for candidate in reversed(args[1:]):
            if isinstance(candidate, str):
                return candidate
    return None


# ------------------------------------------------------------------
# Reports Frame
# ------------------------------------------------------------------
class ReportsFrame(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, bg="#F0F4F8")
        self.parent = parent
        self.role = get_role_from_args((parent,) + args, kwargs) or "user"
        self.pack(fill="both", expand=True)
        self._all_rows = []  # cached result set for client-side search

        # ========== NEW: Auto-refresh control flags ==========
        self._initializing = True  # Prevent auto-refresh during UI build
        self._auto_refresh_enabled = False  # Enable after first manual load
        self._pending_refresh = None  # Timer ID for debounced refresh

        self._build_ui()
        self.populate_scenarios()

        # ========== NEW: Perform initial load automatically ==========
        self.after(100, self._initial_load)

        self.status_var.set(
            lang.t("reports.ready", "Ready (role={role})", role=self.role)
        )

    def _initial_load(self):
        """Perform initial data load and enable auto-refresh"""
        self._initializing = False
        self.load_report()
        self._auto_refresh_enabled = True
        self.refresh_button.focus_set()

    # ========== NEW: Auto-refresh trigger with debouncing ==========
    def _schedule_auto_refresh(self, delay_ms=800):
        """Schedule an auto-refresh with debouncing to avoid too many refreshes"""
        if not self._auto_refresh_enabled or self._initializing:
            return

        # Cancel pending refresh if exists
        if self._pending_refresh:
            self.after_cancel(self._pending_refresh)

        # Schedule new refresh
        self._pending_refresh = self.after(delay_ms, self._execute_auto_refresh)

    def _execute_auto_refresh(self):
        """Execute the auto-refresh"""
        self._pending_refresh = None
        self.load_report()

    # ---------------- UI ----------------
    def _build_ui(self):
        # Header container for title and summary box
        header_container = tk.Frame(self, bg="#F0F4F8")
        header_container.pack(fill="x", pady=10, padx=10)

        # Translation-enabled heading (LEFT)
        tk.Label(
            header_container,
            text=lang.t("reports.heading_stock_statement", "Stock Statement"),
            font=("Helvetica", 20, "bold"),
            bg="#F0F4F8",
        ).pack(side="left", anchor="w")

        # Summary Box (TOP RIGHT)
        summary_frame = tk.Frame(header_container, bg="#FFFFFF", relief="ridge", bd=2)
        summary_frame.pack(side="right", anchor="e")

        summary_title = tk.Label(
            summary_frame,
            text=lang.t("reports.summary_title", "Summary Totals"),
            font=("Helvetica", 10, "bold"),
            bg="#FFFFFF",
            fg="#2C3E50",
        )
        summary_title.grid(
            row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(4, 2)
        )

        # Total Amount
        self.amount_label = tk.Label(
            summary_frame,
            text=lang.t("reports.total_amount", "Total Amount (€):"),
            font=("Helvetica", 8),
            bg="#FFFFFF",
            fg="#34495E",
        )
        self.amount_label.grid(row=1, column=0, sticky="w", padx=(6, 2), pady=1)
        self.total_amount_var = tk.StringVar(value="0.00")
        amount_entry = tk.Entry(
            summary_frame,
            textvariable=self.total_amount_var,
            font=("Helvetica", 8, "bold"),
            state="readonly",
            readonlybackground="#F8F9FA",
            fg="#27AE60",
            bd=1,
            relief="flat",
            width=15,
        )
        amount_entry.grid(row=1, column=1, sticky="w", padx=2, pady=1)

        # Total Weight
        self.weight_label = tk.Label(
            summary_frame,
            text=lang.t("reports.total_weight", "Total Weight (kg):"),
            font=("Helvetica", 8),
            bg="#FFFFFF",
            fg="#34495E",
        )
        self.weight_label.grid(row=2, column=0, sticky="w", padx=(6, 2), pady=1)
        self.total_weight_var = tk.StringVar(value="0.00")
        weight_entry = tk.Entry(
            summary_frame,
            textvariable=self.total_weight_var,
            font=("Helvetica", 8, "bold"),
            state="readonly",
            readonlybackground="#F8F9FA",
            fg="#3498DB",
            bd=1,
            relief="flat",
            width=15,
        )
        weight_entry.grid(row=2, column=1, sticky="w", padx=2, pady=1)

        # Total Volume
        self.volume_label = tk.Label(
            summary_frame,
            text=lang.t("reports.total_volume", "Total Volume (m³):"),
            font=("Helvetica", 8),
            bg="#FFFFFF",
            fg="#34495E",
        )
        self.volume_label.grid(row=3, column=0, sticky="w", padx=(6, 2), pady=(1, 4))
        self.total_volume_var = tk.StringVar(value="0.000")
        volume_entry = tk.Entry(
            summary_frame,
            textvariable=self.total_volume_var,
            font=("Helvetica", 8, "bold"),
            state="readonly",
            readonlybackground="#F8F9FA",
            fg="#E67E22",
            bd=1,
            relief="flat",
            width=15,
        )
        volume_entry.grid(row=3, column=1, sticky="w", padx=2, pady=(1, 4))

        self.summary_frame_ref = summary_frame

        # Filters frame
        filters_frame = tk.Frame(self, bg="#F0F4F8")
        filters_frame.pack(fill="x", padx=10, pady=4)

        # Scenario
        tk.Label(
            filters_frame, text=lang.t("reports.scenario", "Scenario:"), bg="#F0F4F8"
        ).grid(row=0, column=0, sticky="w", padx=4, pady=2)
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(
            filters_frame, textvariable=self.scenario_var, width=30, state="readonly"
        )
        self.scenario_cb.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        # ========== NEW: Auto-refresh on scenario change ==========
        self.scenario_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._schedule_auto_refresh()
        )

        # Type
        tk.Label(
            filters_frame, text=lang.t("reports.type", "Type:"), bg="#F0F4F8"
        ).grid(row=0, column=2, sticky="w", padx=4, pady=2)
        self.type_var = tk.StringVar(value=lang.t("reports.type_all", "All"))
        self.type_cb = ttk.Combobox(
            filters_frame,
            textvariable=self.type_var,
            values=[lang.t("reports.type_all", "All"), "KIT", "MODULE", "ITEM"],
            width=10,
            state="readonly",
        )
        self.type_cb.grid(row=0, column=3, padx=4, pady=2, sticky="w")
        self.type_cb.current(0)
        # ========== NEW: Auto-refresh on type change ==========
        self.type_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._schedule_auto_refresh()
        )

        # Expiry Period override
        tk.Label(
            filters_frame,
            text=lang.t("reports.expiry_period", "Expiry Period (months):"),
            bg="#F0F4F8",
        ).grid(row=0, column=4, sticky="w", padx=4, pady=2)
        self.horizon_override_var = tk.StringVar()
        # ========== NEW: Auto-refresh on period change (debounced) ==========
        self.horizon_override_var.trace_add(
            "write", lambda *args: self._schedule_auto_refresh(1200)
        )
        self.horizon_entry = tk.Entry(
            filters_frame, textvariable=self.horizon_override_var, width=8
        )
        self.horizon_entry.grid(row=0, column=5, padx=4, pady=2, sticky="w")

        # Kit / Module / Item selectors
        tk.Label(filters_frame, text=lang.t("reports.kit", "Kit:"), bg="#F0F4F8").grid(
            row=1, column=0, sticky="w", padx=4, pady=2
        )
        self.kit_var = tk.StringVar()
        # ========== NEW: Auto-refresh on kit change ==========
        self.kit_var.trace_add("write", lambda *args: self._schedule_auto_refresh())
        self.kit_entry = tk.Entry(
            filters_frame, textvariable=self.kit_var, width=20, state="readonly"
        )
        self.kit_entry.grid(row=1, column=1, padx=4, pady=2, sticky="w")
        tk.Button(
            filters_frame,
            text=lang.t("reports.search", "Search"),
            command=lambda: self.search_code("KIT"),
        ).grid(row=1, column=1, padx=(150, 4), pady=2, sticky="w")

        tk.Label(
            filters_frame, text=lang.t("reports.module", "Module:"), bg="#F0F4F8"
        ).grid(row=1, column=2, sticky="w", padx=4, pady=2)
        self.module_var = tk.StringVar()
        # ========== NEW: Auto-refresh on module change ==========
        self.module_var.trace_add("write", lambda *args: self._schedule_auto_refresh())
        self.module_entry = tk.Entry(
            filters_frame, textvariable=self.module_var, width=20, state="readonly"
        )
        self.module_entry.grid(row=1, column=3, padx=4, pady=2, sticky="w")
        tk.Button(
            filters_frame,
            text=lang.t("reports.search", "Search"),
            command=lambda: self.search_code("MODULE"),
        ).grid(row=1, column=3, padx=(150, 4), pady=2, sticky="w")

        tk.Label(
            filters_frame, text=lang.t("reports.item", "Item:"), bg="#F0F4F8"
        ).grid(row=1, column=4, sticky="w", padx=4, pady=2)
        self.item_var = tk.StringVar()
        # ========== NEW: Auto-refresh on item change ==========
        self.item_var.trace_add("write", lambda *args: self._schedule_auto_refresh())
        self.item_entry = tk.Entry(
            filters_frame, textvariable=self.item_var, width=18, state="readonly"
        )
        self.item_entry.grid(row=1, column=5, padx=4, pady=2, sticky="w")
        tk.Button(
            filters_frame,
            text=lang.t("reports.search", "Search"),
            command=lambda: self.search_code("ITEM"),
        ).grid(row=1, column=5, padx=(130, 4), pady=2, sticky="w")

        # Recommended period label
        self.recommended_label = tk.Label(
            self, text="", font=("Helvetica", 8), fg="#555555", bg="#F0F4F8"
        )
        self.recommended_label.pack(anchor="w", padx=12, pady=(0, 2))

        # Real-time table search
        search_frame = tk.Frame(self, bg="#F0F4F8")
        search_frame.pack(fill="x", padx=10, pady=(2, 4))
        tk.Label(
            search_frame, text=lang.t("reports.search", "Search:"), bg="#F0F4F8"
        ).pack(side="left")
        self.table_search_var = tk.StringVar()
        self.table_search_var.trace_add("write", self._on_table_search_change)
        self.table_search_entry = tk.Entry(
            search_frame, textvariable=self.table_search_var, width=40
        )
        self.table_search_entry.pack(side="left", padx=6)

        # Buttons
        btn_frame = tk.Frame(self, bg="#F0F4F8")
        btn_frame.pack(fill="x", padx=10, pady=4)
        self.refresh_button = tk.Button(
            btn_frame,
            text=lang.t("reports.load_refresh", "Load / Refresh"),
            bg="#27AE60",
            fg="white",
            command=self.load_report,
        )
        self.refresh_button.pack(side="left", padx=4)
        tk.Button(
            btn_frame,
            text=lang.t("reports.clear_filters", "Clear Filters"),
            bg="#7F8C8D",
            fg="white",
            command=self.clear_filters,
        ).pack(side="left", padx=4)
        tk.Button(
            btn_frame,
            text=lang.t("reports.export", "Export"),
            bg="#2980B9",
            fg="white",
            command=self.export_excel,
        ).pack(side="left", padx=4)

        self.status_var = tk.StringVar(
            value=lang.t("reports.ready", "Ready (role={role})", role=self.role)
        )
        tk.Label(self, textvariable=self.status_var, anchor="w", bg="#E0E4E8").pack(
            fill="x", padx=10, pady=(0, 4)
        )

        # Table
        cols = [
            "code",
            "description",
            "type",
            "standard_qty",
            "current_stock",
            "qty_expiring",
            "over_stock",
            "missing_qty",
            "pack",
            "price_per_pack",
            "unit_price",
            "weight_per_pack",
            "volume_per_pack",
            "shelf_life",
            "remarks",
            "account_code",
        ]
        headers = {
            "code": lang.t("reports.code", "Code"),
            "description": lang.t("reports.description", "Description"),
            "type": lang.t("reports.type_header", "Type"),
            "standard_qty": lang.t("reports.standard_qty", "Standard Quantity"),
            "current_stock": lang.t("reports.current_stock", "Current Stock"),
            "qty_expiring": lang.t("reports.qty_expiring", "Qty Expiring (Period)"),
            "over_stock": lang.t("reports.over_stock", "Over Stock"),
            "missing_qty": lang.t("reports.missing_qty", "Missing Quantity"),
            "pack": lang.t("reports.pack", "Pack"),
            "price_per_pack": lang.t("reports.price_per_pack", "Price/Pack (EUR)"),
            "unit_price": lang.t("reports.unit_price", "Unit Price (EUR)"),
            "weight_per_pack": lang.t("reports.weight_per_pack", "Weight/Pack (kg)"),
            "volume_per_pack": lang.t("reports.volume_per_pack", "Volume/Pack (dm³)"),
            "shelf_life": lang.t("reports.shelf_life", "Shelf Life (months)"),
            "remarks": lang.t("reports.remarks", "Remarks"),
            "account_code": lang.t("reports.account_code", "Account Code"),
        }
        widths = {
            "code": 150,
            "description": 300,
            "type": 70,
            "standard_qty": 120,
            "current_stock": 110,
            "qty_expiring": 140,
            "over_stock": 90,
            "missing_qty": 110,
            "pack": 120,
            "price_per_pack": 120,
            "unit_price": 110,
            "weight_per_pack": 130,
            "volume_per_pack": 140,
            "shelf_life": 130,
            "remarks": 200,
            "account_code": 110,
        }
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        for c in cols:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths.get(c, 100), anchor="w")

        self.tree.bind("<Double-1>", self._on_row_double_click)

        vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 4))
        vsb.place(relx=1.0, rely=0.28, relheight=0.66, anchor="ne")
        hsb.pack(fill="x", padx=10, pady=(0, 6))

        self.tree.tag_configure("missing", background="#FFCCCC")
        self.tree.tag_configure("overstock", background="#CCFFCC")

    # ========== Double-click handler ==========
    def _on_row_double_click(self, event):
        """Open details popup showing std_qty_helper data for clicked code"""
        selection = self.tree.selection()
        if not selection:
            return

        item = selection[0]
        values = self.tree.item(item, "values")
        if not values:
            return

        code = values[0]
        self._show_code_details(code)

    def _show_code_details(self, code):
        """Show popup with std_qty_helper details for the given code (LANGUAGE-AWARE)"""
        sql = """
            SELECT code, description, type, scenario, kit, module, std_qty
            FROM std_qty_helper
            WHERE code = ?
            ORDER BY scenario, kit, module
        """
        rows = _fetchall(sql, (code,))

        if not rows:
            custom_popup(
                self,
                lang.t("reports.info", "Info"),
                lang.t(
                    "reports.no_details", "No details found for code: {code}", code=code
                ),
                "info",
            )
            return

        popup = tk.Toplevel(self)
        popup.title(
            lang.t("reports.code_details_title", "Code Details: {code}", code=code)
        )
        popup.geometry("1100x500")
        popup.transient(self)
        popup.grab_set()
        popup.configure(bg="#F0F4F8")

        header_frame = tk.Frame(popup, bg="#FFFFFF", relief="flat", bd=1)
        header_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(
            header_frame,
            text=lang.t(
                "reports.code_details_header",
                "Standard Quantity Details for Code: {code}",
                code=code,
            ),
            font=("Helvetica", 12, "bold"),
            bg="#FFFFFF",
            fg="#2C3E50",
        ).pack(pady=8)

        details_container = tk.Frame(popup, bg="#F0F4F8")
        details_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["code", "description", "type", "scenario", "kit", "module", "std_qty"]
        headers = {
            "code": lang.t("reports.code", "Code"),
            "description": lang.t("reports.description", "Description"),
            "type": lang.t("reports.type_header", "Type"),
            "scenario": lang.t("reports.scenario", "Scenario"),
            "kit": lang.t("reports.kit", "Kit"),
            "module": lang.t("reports.module", "Module"),
            "std_qty": lang.t("reports.standard_qty", "Standard Qty"),
        }
        widths = {
            "code": 120,
            "description": 280,
            "type": 80,
            "scenario": 180,
            "kit": 120,
            "module": 120,
            "std_qty": 100,
        }

        details_tree = ttk.Treeview(
            details_container, columns=cols, show="headings", height=16
        )
        for c in cols:
            details_tree.heading(c, text=headers[c])
            details_tree.column(c, width=widths.get(c, 100), anchor="w")

        vsb = ttk.Scrollbar(
            details_container, orient="vertical", command=details_tree.yview
        )
        hsb = ttk.Scrollbar(
            details_container, orient="horizontal", command=details_tree.xview
        )
        details_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        details_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        details_container.grid_rowconfigure(0, weight=1)
        details_container.grid_columnconfigure(0, weight=1)

        for r in rows:
            details_tree.insert(
                "",
                "end",
                values=(
                    r["code"] or "",
                    r["description"] or "",
                    r["type"] or "",
                    r["scenario"] or "",
                    r["kit"] or "",
                    r["module"] or "",
                    r["std_qty"] or 0,
                ),
            )

        btn_frame = tk.Frame(popup, bg="#F0F4F8")
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        tk.Button(
            btn_frame,
            text=lang.t("reports.close", "Close"),
            font=("Helvetica", 10),
            bg="#7F8C8D",
            fg="white",
            command=popup.destroy,
        ).pack(side="right", padx=4)

        popup.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() // 2) - (popup.winfo_width() // 2)
        y = (
            self.winfo_rooty()
            + (self.winfo_height() // 2)
            - (popup.winfo_height() // 2)
        )
        popup.geometry(f"+{max(0, x)}+{max(0, y)}")

    # ========== Update summary totals ==========
    def _update_summary_totals(self):
        """Calculate and update summary totals based on missing quantities"""
        total_amount = 0.0
        total_weight = 0.0
        total_volume = 0.0

        for row in self._all_rows:
            missing_qty = row.get("missing_qty", 0) or 0
            if missing_qty <= 0:
                continue

            unit_price = row.get("unit_price", 0) or 0
            weight_per_pack = row.get("weight_per_pack", 0) or 0
            volume_per_pack = row.get("volume_per_pack", 0) or 0

            try:
                unit_price = float(unit_price) if unit_price else 0.0
                weight_per_pack = float(weight_per_pack) if weight_per_pack else 0.0
                volume_per_pack = float(volume_per_pack) if volume_per_pack else 0.0
            except (ValueError, TypeError):
                unit_price = 0.0
                weight_per_pack = 0.0
                volume_per_pack = 0.0

            total_amount += missing_qty * unit_price
            total_weight += missing_qty * weight_per_pack
            total_volume += missing_qty * (volume_per_pack / 1000.0)

        self.total_amount_var.set(f"{total_amount:,.2f}")
        self.total_weight_var.set(f"{total_weight:,.2f}")
        self.total_volume_var.set(f"{total_volume:,.3f}")

    # ---------------- Scenario list ----------------
    def populate_scenarios(self):
        rows = _fetchall("SELECT name FROM scenarios ORDER BY name")
        values = [r["name"] for r in rows]
        self.scenario_cb["values"] = [""] + values
        if values:
            self.scenario_cb.current(0)
        self._update_recommended_label()

    # ---------------- Search item popups (LANGUAGE-AWARE) ----------------
    def search_code(self, type_filter):
        d = tk.Toplevel(self)
        d.title(f"{lang.t('reports.search','Search')} {type_filter}")
        d.geometry("520x320")
        d.transient(self)
        d.grab_set()
        tk.Label(
            d,
            text=f"{lang.t('reports.search','Search')} {type_filter}:",
            font=("Helvetica", 10, "bold"),
        ).pack(pady=4)
        sv = tk.StringVar()
        entry = tk.Entry(d, textvariable=sv)
        entry.pack(fill="x", padx=8, pady=4)
        lb = tk.Listbox(d, height=10)
        lb.pack(fill="both", expand=True, padx=8, pady=4)

        def load_list(*_):
            q = sv.get().strip().lower()
            designation_expr = _build_designation_sql()
            sql = f"""
                SELECT code, {designation_expr} AS designation
                FROM items_list il
                WHERE UPPER(type)=? AND (
                      UPPER(code) LIKE UPPER(?) OR
                      UPPER(COALESCE({designation_expr},'')) LIKE UPPER(?)
                ) ORDER BY code LIMIT 50
            """
            rows = _fetchall(sql, (type_filter.upper(), f"%{q}%", f"%{q}%"))
            lb.delete(0, tk.END)
            if not rows:
                lb.insert(tk.END, lang.t("reports.no_results", "No results"))
                return
            for r in rows:
                lb.insert(tk.END, f"{r['code']} - {r['designation'] or ''}")

        def choose(_=None):
            sel = lb.curselection()
            if not sel:
                d.destroy()
                return
            line = lb.get(sel[0])
            if line == lang.t("reports.no_results", "No results"):
                d.destroy()
                return
            code = line.split(" - ")[0]
            if type_filter.upper() == "KIT":
                self.kit_var.set(code)
            elif type_filter.upper() == "MODULE":
                self.module_var.set(code)
            else:
                self.item_var.set(code)
            d.destroy()

        sv.trace("w", load_list)
        lb.bind("<Double-1>", choose)
        entry.bind("<Return>", choose)
        load_list()
        entry.focus()
        d.wait_window()

    # ---------------- Filters ----------------
    def gather_filters(self):
        tsel = self.type_var.get()
        type_filter = "All" if tsel in ("", lang.t("reports.type_all", "All")) else tsel
        return {
            "scenario": self.scenario_var.get() or None,
            "kit_code": self.kit_var.get() or None,
            "module_code": self.module_var.get() or None,
            "item_code": self.item_var.get() or None,
            "type_filter": type_filter,
        }

    def clear_filters(self):
        # ========== NEW: Temporarily disable auto-refresh during clear ==========
        was_enabled = self._auto_refresh_enabled
        self._auto_refresh_enabled = False

        self.scenario_var.set("")
        self.kit_var.set("")
        self.module_var.set("")
        self.item_var.set("")
        self.type_var.set("All")
        self.horizon_override_var.set("")
        self.table_search_var.set("")
        self.tree.delete(*self.tree.get_children())
        self._all_rows.clear()
        self._update_recommended_label()
        self._update_summary_totals()
        self.status_var.set(
            lang.t(
                "reports.filters_cleared",
                "Filters cleared. Ready (role={role}).",
                role=self.role,
            )
        )

        # Re-enable and trigger refresh
        self._auto_refresh_enabled = was_enabled
        if was_enabled:
            self._schedule_auto_refresh(200)

    # ---------------- Recommended period label ----------------
    def _update_recommended_label(self):
        rec = compute_horizon_months()
        self.recommended_label.config(
            text=lang.t(
                "reports.expiry_period_recommended",
                lang.t(
                    "reports.expiry_horizon_recommended",
                    "Recommended: {months} month(s) (lead+cover+buffer)",
                    months=rec,
                ),
                months=rec,
            )
        )

    # ---------------- Load Report ----------------
    def load_report(self):
        self.tree.delete(*self.tree.get_children())
        self._all_rows.clear()

        # Refresh snapshots
        try:
            self.status_var.set(
                lang.t("reports.refreshing_snapshots", "Refreshing snapshots...")
            )
            self.update_idletasks()
            _refresh_snapshots_internal()
        except Exception as e:
            self.status_var.set(
                lang.t(
                    "reports.snapshot_refresh_failed",
                    "Snapshot refresh failed: {err}",
                    err=str(e),
                )
            )
            custom_popup(
                self,
                lang.t("reports.error", "Error"),
                lang.t(
                    "reports.snapshot_refresh_failed",
                    "Snapshot refresh failed: {err}",
                    err=str(e),
                ),
                "error",
            )
            return

        filters = self.gather_filters()

        # Determine period
        recommended = compute_horizon_months()
        raw_override = self.horizon_override_var.get().strip()
        if raw_override == "":
            horizon_months = recommended
        else:
            try:
                val = int(raw_override)
                horizon_months = val if val >= 0 else recommended
            except Exception:
                horizon_months = recommended
        cutoff_iso = compute_cutoff_date(horizon_months) if horizon_months > 0 else None

        self.status_var.set(lang.t("reports.computing", "Computing aggregates..."))
        self.update_idletasks()
        std_map = aggregate_std_qty(filters)
        stock_map = aggregate_stock(filters, cutoff_iso)
        all_codes = set(std_map.keys()) | set(stock_map.keys())
        meta = load_item_metadata(all_codes)

        for code in sorted(all_codes):
            std_qty = std_map.get(code, 0) or 0
            stock_entry = stock_map.get(code, {"current_stock": 0, "expiring": 0})
            current_stock = stock_entry["current_stock"] or 0
            qty_expiring = stock_entry["expiring"] or 0
            over_stock = max(0, current_stock - std_qty)
            missing_qty = max(0, (std_qty - current_stock) + min(qty_expiring, std_qty))

            m = meta.get(code)
            row = {
                "code": code,
                "description": m["designation"] if m else "",
                "type": (m["type"] if m else ""),
                "standard_qty": std_qty,
                "current_stock": current_stock,
                "qty_expiring": qty_expiring,
                "over_stock": over_stock,
                "missing_qty": missing_qty,
                "pack": m["pack"] if m else "",
                "price_per_pack": (
                    m["price_per_pack_euros"]
                    if m and m["price_per_pack_euros"] is not None
                    else ""
                ),
                "unit_price": (
                    m["unit_price_euros"]
                    if m and m["unit_price_euros"] is not None
                    else ""
                ),
                "weight_per_pack": (
                    m["weight_per_pack_kg"]
                    if m and m["weight_per_pack_kg"] is not None
                    else ""
                ),
                "volume_per_pack": (
                    m["volume_per_pack_dm3"]
                    if m and m["volume_per_pack_dm3"] is not None
                    else ""
                ),
                "shelf_life": (
                    m["shelf_life_months"]
                    if m and m["shelf_life_months"] is not None
                    else ""
                ),
                "remarks": (m["remarks"] if m else ""),
                "account_code": (m["account_code"] if m else ""),
            }
            self._all_rows.append(row)

        self._render_rows(self._all_rows)
        self._update_summary_totals()

        horizon_text = f"{horizon_months}" if horizon_months is not None else "0"
        self.status_var.set(
            lang.t(
                "reports.loaded_rows_status_period",
                lang.t(
                    "reports.loaded_rows_status",
                    "Loaded {rows} rows. Period: {h} month(s) (cutoff {cutoff}) role={role}",
                    rows=len(self._all_rows),
                    h=horizon_text,
                    cutoff=cutoff_iso or "N/A",
                    role=self.role,
                ),
                rows=len(self._all_rows),
                h=horizon_text,
                cutoff=cutoff_iso or "N/A",
                role=self.role,
            )
        )
        self._update_recommended_label()

    # ---------------- Rendering & filtering ----------------
    def _render_rows(self, rows):
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            tags = []
            if r["missing_qty"] > 0:
                tags.append("missing")
            elif r["over_stock"] > 0:
                tags.append("overstock")
            self.tree.insert(
                "",
                "end",
                values=(
                    r["code"],
                    r["description"],
                    r["type"],
                    r["standard_qty"],
                    r["current_stock"],
                    r["qty_expiring"],
                    r["over_stock"],
                    r["missing_qty"],
                    r["pack"],
                    r["price_per_pack"],
                    r["unit_price"],
                    r["weight_per_pack"],
                    r["volume_per_pack"],
                    r["shelf_life"],
                    r["remarks"],
                    r["account_code"],
                ),
                tags=tuple(tags),
            )

    def _on_table_search_change(self, *_):
        query = self.table_search_var.get().strip().lower()
        if not query:
            self._render_rows(self._all_rows)
            return
        filtered = []
        for r in self._all_rows:
            if (
                query in (r["code"] or "").lower()
                or query in (r["description"] or "").lower()
                or query in (r["type"] or "").lower()
            ):
                filtered.append(r)
        self._render_rows(filtered)
        horizon_display = (
            self.horizon_override_var.get().strip() or compute_horizon_months()
        )
        self.status_var.set(
            lang.t(
                "reports.loaded_rows_status_period",
                "Loaded {rows} rows. Period: {h} month(s) role={role}",
                rows=len(filtered),
                h=horizon_display,
                role=self.role,
                cutoff="",
            )
        )

    # ---------------- Export ----------------
    def export_excel(self):
        rows = self.tree.get_children("")
        if not rows:
            custom_popup(
                self,
                lang.t("reports.info", "Info"),
                lang.t("reports.nothing_to_export", "Nothing to export."),
                "info",
            )
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="stock_statement.xlsx",
        )
        if not file_path:
            self.status_var.set(lang.t("reports.export_cancelled", "Export cancelled"))
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock_Statement"
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws["A1"] = (
                f"{lang.t('reports.generated','Generated')}: {now} (role={self.role})"
            )
            ws["A1"].font = Font(size=10)
            ws.append([])
            headers = [self.tree.heading(col)["text"] for col in self.tree["columns"]]
            ws.append(headers)

            missing_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            over_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )

            for iid in rows:
                vals = self.tree.item(iid, "values")
                ws.append(list(vals))
                tagset = set(self.tree.item(iid, "tags") or [])
                if "missing" in tagset:
                    for cell in ws[ws.max_row]:
                        cell.fill = missing_fill
                elif "overstock" in tagset:
                    for cell in ws[ws.max_row]:
                        cell.fill = over_fill

            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    v = str(cell.value) if cell.value is not None else ""
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[letter].width = min(max_len + 2, 55)
            ws.freeze_panes = "A3"
            wb.save(file_path)
            custom_popup(
                self,
                lang.t("reports.success", "Success"),
                lang.t("reports.export_success", "Export successful."),
                "success",
            )
            self.status_var.set(lang.t("reports.export_success", "Export successful."))
        except Exception as e:
            custom_popup(
                self,
                lang.t("reports.error", "Error"),
                lang.t("reports.export_failed", "Export failed: {err}", err=str(e)),
                "error",
            )
            self.status_var.set(
                lang.t("reports.export_failed", "Export failed: {err}", err=str(e))
            )


# Backward compatibility alias
class Reports(ReportsFrame):
    pass


def create_reports(parent, *args, **kwargs):
    return ReportsFrame(parent, *args, **kwargs)


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Statement")
    root.geometry("1480x780")
    ReportsFrame(root, role="admin")
    root.mainloop()
