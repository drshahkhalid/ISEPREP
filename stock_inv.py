import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import time
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from db import connect_db
from language_manager import lang
from stock_data import parse_expiry
from manage_items import get_item_description, detect_type
from popup_utils import custom_popup, custom_askyesno, custom_dialog

# Optional custom popups
try:
    from popup_utils import custom_popup, custom_askyesno
except ImportError:

    def custom_popup(parent, title, message, kind="info"):
        if kind == "error":
            messagebox.showerror(title, message)
        elif kind == "warning":
            messagebox.showwarning(title, message)
        else:
            messagebox.showinfo(title, message)

    def custom_askyesno(parent, title, message):
        return "yes" if messagebox.askyesno(title, message) else "no"


_SCENARIO_CACHE = {}


# ======================== HELPER FUNCTIONS FROM STOCK_SUMMARY ========================


def normalize_type_text(text: str) -> str:
    """
    Normalize item type text for comparison.
    Handles translations: Module/Módulo, Kit, Item
    Returns uppercase normalized text: KIT, MODULE, or ITEM
    """
    if not text:
        return ""

    # Remove accents and convert to uppercase
    import unicodedata

    normalized = unicodedata.normalize("NFD", text)
    without_accents = "".join(
        char for char in normalized if unicodedata.category(char) != "Mn"
    )
    upper = without_accents.upper()

    # Map variations to standard types
    if upper in ("KIT",):
        return "KIT"
    elif upper in ("MODULE", "MODULO"):  # Handles both Module and Módulo
        return "MODULE"
    elif upper in ("ITEM",):
        return "ITEM"

    return upper


def load_scenario_maps():
    """Load scenario ID to name mappings"""
    conn = connect_db()
    if conn is None:
        return {}, set()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT scenario_id, name FROM scenarios")
        rows = cur.fetchall()
        id_to_name = {}
        name_set = set()
        for r in rows:
            sid = r["scenario_id"]
            nm = r["name"]
            if nm:
                name_set.add(nm)
            id_to_name[str(sid)] = nm
        return id_to_name, name_set
    finally:
        cur.close()
        conn.close()


def normalize_scenario(raw_val, id_to_name, name_set):
    """Normalize scenario value to name"""
    if raw_val is None:
        return ""
    if raw_val in name_set:
        return raw_val
    txt = str(raw_val)
    if txt in id_to_name:
        return id_to_name[txt] or txt
    return txt


def load_std_quantities_by_scenario(scenario_name=None):
    """
    Load standard quantities with treecodes and management type info
    Returns: dict with structure {scenario_name: {key: {"code": code, "std_qty": qty, "mgmt_type": type, ...}}}
    Key is: code for on-shelf, treecode for in-box
    """
    result = {}

    # Determine scenarios to load
    if not scenario_name:
        scenario_rows = _fetchall("SELECT scenario_id, name FROM scenarios")
        scenarios = [(r["scenario_id"], r["name"]) for r in scenario_rows]
    else:
        scenario_row = _fetchone(
            "SELECT scenario_id, name FROM scenarios WHERE name = ?", (scenario_name,)
        )
        if not scenario_row:
            return {}
        scenarios = [(scenario_row["scenario_id"], scenario_row["name"])]

    for scenario_id, scenario_name in scenarios:
        if scenario_name not in result:
            result[scenario_name] = {}

        # 1. Load ON-SHELF items from compositions
        onshelf_rows = _fetchall(
            """
            SELECT code, quantity
            FROM compositions 
            WHERE scenario_id = ?
            ORDER BY code
            """,
            (scenario_id,),
        )

        for r in onshelf_rows:
            code = r["code"]
            qty = r["quantity"] or 0

            # Key is code for on-shelf items
            result[scenario_name][code] = {
                "code": code,
                "std_qty": qty,
                "mgmt_type": "on-shelf",
                "kit_code": None,
                "module_code": None,
                "treecode": None,
                "scenario_id": scenario_id,
            }

        # 2. Load IN-BOX items from kit_items
        inbox_rows = _fetchall(
            """
            SELECT code, kit, module, std_qty, treecode
            FROM kit_items
            WHERE scenario_id = ?
            ORDER BY treecode, kit, module
            """,
            (scenario_id,),
        )

        for r in inbox_rows:
            code = r["code"]
            treecode = r["treecode"] or code
            kit = r["kit"] or ""
            module = r["module"] or ""
            qty = r["std_qty"] or 0

            # Key is treecode for in-box items
            result[scenario_name][treecode] = {
                "code": code,
                "std_qty": qty,
                "mgmt_type": "in-box",
                "kit_code": kit,
                "module_code": module,
                "treecode": treecode,
                "scenario_id": scenario_id,
            }

    return result


def aggregate_stock_by_key(scenario_name, mgmt_mode):
    """
    ✅ FIXED: Aggregate stock data by key + EXPIRY DATE + COMMENTS
    Uses ENGLISH constants for DB queries
    """
    result = {}

    id_to_name, name_set = load_scenario_maps()

    conn = connect_db()
    if conn is None:
        return result

    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        # Build WHERE clause
        where_parts = ["final_qty > 0"]
        params = []

        if scenario_name:
            where_parts.append(
                "(scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name=? LIMIT 1))"
            )
            params.extend([scenario_name, scenario_name])

        where_clause = " AND ".join(where_parts)

        # ✅ NEW: Determine which management modes to query (use English)
        query_onshelf = False
        query_inbox = False

        # Map translated mgmt_mode to English for DB query
        if not mgmt_mode or mgmt_mode == "" or mgmt_mode == "All":
            query_onshelf = True
            query_inbox = True
        elif mgmt_mode == lang.t("stock_inv.management_on_shelf", "On-Shelf"):
            query_onshelf = True
        elif mgmt_mode == lang.t("stock_inv.management_in_box", "In-Box"):
            query_inbox = True
        else:
            # Fallback: try to detect if it's a translated value
            query_onshelf = True
            query_inbox = True

        # ✅ Query ON-SHELF items (match by item code + EXPIRY + COMMENTS)
        if query_onshelf:
            onshelf_sql = f"""
                SELECT
                    CAST(scenario AS TEXT) AS raw_scenario,
                    item AS code,
                    exp_date,
                    SUM(final_qty) AS current_stock,
                    GROUP_CONCAT(DISTINCT kit_number) AS kit_numbers,
                    GROUP_CONCAT(DISTINCT module_number) AS module_numbers,
                    GROUP_CONCAT(DISTINCT comments) AS comments
                FROM stock_data
                WHERE {where_clause}
                  AND LOWER(management_mode) IN ('on_shelf','on-shelf','onshelf')
                  AND item IS NOT NULL
                  AND item <> ''
                GROUP BY raw_scenario, item, exp_date
                ORDER BY item, exp_date
            """

            onshelf_rows = cur.execute(onshelf_sql, tuple(params)).fetchall()

            for r in onshelf_rows:
                norm_scen = normalize_scenario(r["raw_scenario"], id_to_name, name_set)
                key = (norm_scen, r["code"], r["exp_date"] or "")

                result[key] = {
                    "current_stock": r["current_stock"] or 0,
                    "exp_date": r["exp_date"] or "",
                    "kit_number": "",
                    "module_number": "",
                    "mgmt_type": "on-shelf",  # Always English
                    "comments": r["comments"] or "",
                }

        # ✅ Query IN-BOX items (match by treecode + EXPIRY + COMMENTS)
        if query_inbox:
            inbox_sql = f"""
                SELECT
                    CAST(scenario AS TEXT) AS raw_scenario,
                    treecode,
                    exp_date,
                    SUM(final_qty) AS current_stock,
                    MIN(kit_number) AS kit_number,
                    MIN(module_number) AS module_number,
                    GROUP_CONCAT(DISTINCT comments) AS comments
                FROM stock_data
                WHERE {where_clause}
                  AND LOWER(management_mode) IN ('in_box','in-box','inbox')
                  AND treecode IS NOT NULL
                  AND treecode <> ''
                GROUP BY raw_scenario, treecode, exp_date
                ORDER BY treecode, exp_date
            """

            inbox_rows = cur.execute(inbox_sql, tuple(params)).fetchall()

            for r in inbox_rows:
                norm_scen = normalize_scenario(r["raw_scenario"], id_to_name, name_set)
                key = (norm_scen, r["treecode"], r["exp_date"] or "")

                result[key] = {
                    "current_stock": r["current_stock"] or 0,
                    "exp_date": r["exp_date"] or "",
                    "kit_number": r["kit_number"] or "",
                    "module_number": r["module_number"] or "",
                    "mgmt_type": "in-box",  # Always English
                    "comments": r["comments"] or "",
                }

    finally:
        cur.close()
        conn.close()

    return result


def _fetchall(sql, params=()):
    """Helper to fetch all rows"""
    conn = connect_db()
    if conn is None:
        return []
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        return cur.fetchall()
    except sqlite3.Error:
        return []
    finally:
        cur.close()
        conn.close()


def _fetchone(sql, params=()):
    """Helper to fetch one row"""
    rows = _fetchall(sql, params)
    return rows[0] if rows else None


def scenario_id_to_name(scenario_id: str) -> str:
    if not scenario_id:
        return lang.t("stock_inv.unknown", "Unknown")
    if scenario_id in _SCENARIO_CACHE:
        return _SCENARIO_CACHE[scenario_id]
    conn = connect_db()
    if conn is None:
        return scenario_id
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM scenarios WHERE scenario_id = ?", (scenario_id,))
        row = cur.fetchone()
        name = row["name"] if row and row["name"] else scenario_id
        _SCENARIO_CACHE[scenario_id] = name
        return name
    finally:
        cur.close()
        conn.close()


def get_active_designation(code):
    if not code:
        return lang.t("stock_inv.no_description", "No Description")
    conn = connect_db()
    if conn is None:
        return code
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute(
            """SELECT designation, designation_en, designation_fr, designation_sp
                          FROM items_list WHERE code = ?""",
            (code,),
        )
        row = cursor.fetchone()
        if not row:
            return code
        lang_code = lang.lang_code.lower()
        col_map = {
            "en": "designation_en",
            "fr": "designation_fr",
            "es": "designation_sp",
            "sp": "designation_sp",
        }
        preferred = col_map.get(lang_code, "designation_en")
        if row[preferred]:
            return row[preferred]
        if row["designation_en"]:
            return row["designation_en"]
        return row["designation"] or code
    finally:
        cursor.close()
        conn.close()


def check_expiry_required(code):
    if not code:
        return False
    conn = connect_db()
    if conn is None:
        return False
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT remarks FROM items_list WHERE code = ?", (code,))
        row = cursor.fetchone()
        return bool(row and row["remarks"] and "exp" in row["remarks"].lower())
    finally:
        cursor.close()
        conn.close()


def get_std_qty(code, scenario_name):
    conn = connect_db()
    if conn is None:
        return 0
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        if scenario_name == lang.t("stock_inv.all_scenarios", "All Scenarios"):
            cursor.execute(
                "SELECT SUM(quantity) AS total_qty FROM compositions WHERE code = ?",
                (code,),
            )
            row = cursor.fetchone()
            return row["total_qty"] if row and row["total_qty"] is not None else 0
        else:
            cursor.execute(
                """
                SELECT quantity FROM compositions
                 WHERE code = ? AND scenario_id = (
                     SELECT scenario_id FROM scenarios WHERE name = ?
                 )""",
                (code, scenario_name),
            )
            row = cursor.fetchone()
            return row["quantity"] if row and row["quantity"] is not None else 0
    finally:
        cursor.close()
        conn.close()


def get_treecode(scenario_id, kit_code, module_code, item_code):
    """
    Fetch treecode from kit_items for the given scenario, kit, module, item.
    """
    conn = connect_db()
    if conn is None:
        return None
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT treecode FROM kit_items
            WHERE scenario_id = ? AND kit = ? AND module = ? AND item = ?
            LIMIT 1
        """,
            (scenario_id, kit_code or None, module_code or None, item_code),
        )
        row = cur.fetchone()
        return row["treecode"] if row else None
    finally:
        cur.close()
        conn.close()


def get_item_type(code: str) -> str:
    if not code:
        return lang.t("stock_inv.item", "Item")
    conn = connect_db()
    if conn is None:
        return lang.t("stock_inv.item", "Item")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT type, designation_en FROM items_list WHERE code = ?", (code,)
        )
        row = cur.fetchone()
        if row and row["type"]:
            t = row["type"].strip().upper()
            if t in ("KIT", "MODULE", "ITEM"):
                return lang.t(f"stock_inv.{t.lower()}", t.title())
        desc = row["designation_en"] if row else get_item_description(code)
        dt = detect_type(code, desc) if desc else "ITEM"
        return {
            "KIT": lang.t("stock_inv.kit", "Kit"),
            "MODULE": lang.t("stock_inv.module", "Module"),
            "ITEM": lang.t("stock_inv.item", "Item"),
        }.get(dt.upper(), lang.t("stock_inv.item", "Item"))
    finally:
        cur.close()
        conn.close()


def validate_expiry_for_save(code, expiry_date):
    """
    For items requiring expiry (remarks contains 'exp'), an expiry date must exist and be future.
    """
    if not check_expiry_required(code):
        return True
    parsed = parse_expiry(expiry_date) if expiry_date else None
    if not parsed:
        return False
    # parsed may return datetime or date; convert to date
    if hasattr(parsed, "date"):
        parsed_date = parsed.date()
    else:
        parsed_date = parsed
    return parsed_date > datetime.now().date()


def parse_inventory_unique_id(unique_id: str) -> dict:
    parts = unique_id.split("/") if unique_id else []
    out = {
        "scenario_id": None,
        "scenario_name": None,
        "kit_code": None,
        "module_code": None,
        "item_code": None,
        "std_qty": None,
        "exp_date": None,
        "kit_number": None,
        "module_number": None,
        "treecode": None,
    }
    if not parts:
        return out
    if len(parts) >= 1:
        out["scenario_id"] = parts[0]
        out["scenario_name"] = scenario_id_to_name(parts[0])
    if len(parts) >= 2 and parts[1] not in ("None", ""):
        out["kit_code"] = parts[1]
    if len(parts) >= 3 and parts[2] not in ("None", ""):
        out["module_code"] = parts[2]
    if len(parts) >= 4 and parts[3] not in ("None", ""):
        out["item_code"] = parts[3]
    if len(parts) >= 5:
        try:
            out["std_qty"] = int(parts[4])
        except:
            out["std_qty"] = None
    if len(parts) >= 6 and parts[5] not in ("None", ""):
        out["exp_date"] = parts[5]
    if len(parts) >= 7 and parts[6] not in ("None", ""):
        out["kit_number"] = parts[6]
    if len(parts) >= 8 and parts[7] not in ("None", ""):
        out["module_number"] = parts[7]
    if len(parts) >= 9 and parts[8] not in ("None", ""):
        out["treecode"] = parts[8]
    return out


def construct_unique_id(
    scenario_id: str,
    kit_code: str,
    module_code: str,
    item_code: str,
    std_qty: int,
    exp_date: str,
    kit_number: str = None,
    module_number: str = None,
    force_box_format: bool = False,
    treecode: str = None,
) -> str:
    base = [
        scenario_id or "0",
        kit_code or "None",
        module_code or "None",
        item_code or "None",
        str(std_qty if std_qty is not None else 0),
        exp_date or "None",
    ]
    if kit_number or module_number or force_box_format:
        base.append(kit_number or "None")
        base.append(module_number or "None")
        if treecode:
            base.append(treecode)
    return "/".join(base)


class StockInventory(tk.Frame):
    def __init__(self, parent, app, role="supervisor"):
        super().__init__(parent)
        self.app = app
        self.role = role.lower()
        self.scenario_map = self.fetch_scenario_map()
        self.ctx_menu = None
        self.ctx_row = None
        self.base_physical_inputs = {}
        self.user_row_states = {}
        self.temp_row_counter = 0
        self.pack(fill="both", expand=True)
        self.render_ui()

    # ---------- DB fetchers ----------
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None:
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios")
            return {str(r["scenario_id"]): r["name"] for r in cur.fetchall()}
        finally:
            cur.close()
            conn.close()

    def fetch_kit_numbers(self, scenario_name=None):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            sql = """SELECT DISTINCT kit_number
                       FROM stock_data
                      WHERE kit_number IS NOT NULL
                        AND kit_number != 'None'"""
            params = []
            if scenario_name and scenario_name != lang.t(
                "stock_inv.all_scenarios", "All Scenarios"
            ):
                sql += " AND scenario = ?"
                params.append(scenario_name)
            sql += " ORDER BY kit_number"
            cur.execute(sql, params)
            return [r["kit_number"] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()

    def fetch_module_numbers(self, scenario_name=None, kit_number=None):
        conn = connect_db()
        if conn is None:
            return []
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            sql = """SELECT DISTINCT module_number
                       FROM stock_data
                      WHERE module_number IS NOT NULL
                        AND module_number != 'None'"""
            params = []
            if scenario_name and scenario_name != lang.t(
                "stock_inv.all_scenarios", "All Scenarios"
            ):
                sql += " AND scenario = ?"
                params.append(scenario_name)
            if kit_number and kit_number not in (
                lang.t("stock_inv.all_kits", "All Kits"),
                lang.t("stock_inv.stand_alone_items", "Stand alone items"),
            ):
                sql += " AND kit_number = ?"
                params.append(kit_number)
            sql += " ORDER BY module_number"
            cur.execute(sql, params)
            return [r["module_number"] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()

    def fetch_project_details(self):
        try:
            conn = connect_db()
            if conn is None:
                return (
                    lang.t("stock_inv.unknown", "Unknown"),
                    lang.t("stock_inv.unknown", "Unknown"),
                )
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            try:
                cur.execute(
                    "SELECT project_name, project_code FROM project_details ORDER BY id DESC LIMIT 1"
                )
                row = cur.fetchone()
                return (
                    (row["project_name"], row["project_code"])
                    if row
                    else (
                        lang.t("stock_inv.unknown", "Unknown"),
                        lang.t("stock_inv.unknown", "Unknown"),
                    )
                )
            finally:
                cur.close()
                conn.close()
        except Exception as e:
            # Log or handle the error, return defaults
            print(f"Error fetching project details: {e}")
            return (
                lang.t("stock_inv.unknown", "Unknown"),
                lang.t("stock_inv.unknown", "Unknown"),
            )

    # ---------- Dropdown refresh ----------
    def refresh_kit_dropdown(self):
        scenario_val = self.scenario_var.get()
        kits = self.fetch_kit_numbers(scenario_val)
        all_label = lang.t("stock_inv.all_kits", "All Kits")
        standalone_label = lang.t("stock_inv.stand_alone_items", "Stand alone items")
        kits = [k for k in kits if k not in (all_label, standalone_label)]
        kits_sorted = sorted(kits) + [standalone_label, all_label]
        current = self.kit_number_var.get()
        self.kit_number_cb["values"] = kits_sorted
        if current not in kits_sorted:
            self.kit_number_var.set(all_label)

    def refresh_module_dropdown(self):
        scenario_val = self.scenario_var.get()
        kit_val = self.kit_number_var.get()
        modules = self.fetch_module_numbers(
            scenario_val,
            (
                kit_val
                if kit_val
                not in (
                    lang.t("stock_inv.all_kits", "All Kits"),
                    lang.t("stock_inv.stand_alone_items", "Stand alone items"),
                )
                else None
            ),
        )
        all_label = lang.t("stock_inv.all_modules", "All Modules")
        if all_label not in modules:
            modules.append(all_label)
        modules_sorted = [m for m in modules if m != all_label] + [all_label]
        self.module_number_cb["values"] = modules_sorted
        if self.module_number_var.get() not in modules_sorted:
            self.module_number_var.set(all_label)

    # ---------- Search base ----------
    def fetch_search_results(self, query):
        """
        NEW: Search in standard quantities (compositions + kit_items)
        Returns items matching query, with stock status
        """
        if not query or len(query) < 2:
            return []

        scenario_filter = self.scenario_var.get()
        mgmt_mode = self.mgmt_mode_var.get()

        # ✅ FIX: Convert "All Scenarios" to None
        all_scenarios_text = lang.t("stock_inv.all_scenarios", "All Scenarios")
        if scenario_filter == all_scenarios_text or not scenario_filter:
            scenario_filter = None

        # Get all items from std sources
        std_data_by_scenario = load_std_quantities_by_scenario(scenario_filter)

        # Get stock data
        stock_map = aggregate_stock_by_key(scenario_filter, mgmt_mode)

        # Search through items
        query_lower = query.lower()
        results = []

        for scenario_name, std_data in std_data_by_scenario.items():
            for key, std_info in std_data.items():
                code = std_info["code"]
                description = get_active_designation(code)

                # Search in code and description
                if query_lower in code.lower() or query_lower in description.lower():
                    # Check stock status
                    stock_key = (scenario_name, key)
                    stock_entry = stock_map.get(stock_key, {"current_stock": 0})

                    results.append(
                        {
                            "code": code,
                            "description": description,
                            "scenario": scenario_name,
                            "has_stock": stock_entry["current_stock"] > 0,
                            "std_qty": std_info["std_qty"],
                            "mgmt_type": std_info["mgmt_type"],
                        }
                    )

        return results

    def on_search_keyrelease(self, event=None):
        query = self.code_entry.get().strip()
        self.search_listbox.delete(0, tk.END)

        # ✅ NEW: If search is cleared in Complete Inventory mode, reload tree
        if len(query) == 0:
            inv_type = self.inv_type_var.get()
            if inv_type == lang.t("stock_inv.complete_inventory", "Complete Inventory"):
                # Reload tree to show all items again
                self.rebuild_tree_preserving_state()
            return

        if len(query) < 2:
            return

        results = self.fetch_search_results(query)

        for r in results:
            # Show stock status indicator
            indicator = "✓" if r["has_stock"] else "○"
            display = f"{indicator} {r['code']} - {r['description']} [{r['mgmt_type']}]"
            self.search_listbox.insert(tk.END, display)

        self.status_var.set(
            lang.t(
                "stock_inv.found_items", "Found {count} items matching '{query}'"
            ).format(count=len(results), query=query)
        )

    def on_search_select(self, event=None):
        """Handle search selection - load the item from std sources"""
        sel = self.search_listbox.curselection()
        if not sel:
            return

        line = self.search_listbox.get(sel[0])

        # Extract code (handle both ✓ and ○ prefixes)
        code = line.split(" - ")[0].replace("✓ ", "").replace("○ ", "").strip()

        # Load the item using fetch_inventory_items (filtered by code)
        scenario_filter = self.scenario_var.get()
        mgmt_mode = self.mgmt_mode_var.get()

        # ✅ FIX: Convert "All Scenarios" to None
        all_scenarios_text = lang.t("stock_inv.all_scenarios", "All Scenarios")
        if scenario_filter == all_scenarios_text or not scenario_filter:
            scenario_filter = None

        std_data_by_scenario = load_std_quantities_by_scenario(scenario_filter)
        stock_map = aggregate_stock_by_key(scenario_filter, mgmt_mode)

        # Find matching items
        added = 0
        for scenario_name, std_data in std_data_by_scenario.items():
            for key, std_info in std_data.items():
                if std_info["code"] == code:
                    scenario_id = std_info["scenario_id"]
                    mgmt_type = std_info["mgmt_type"]

                    # Get stock data
                    stock_key = (scenario_name, key)
                    stock_entry = stock_map.get(
                        stock_key,
                        {
                            "current_stock": 0,
                            "earliest_expiry": "",
                            "kit_number": "",
                            "module_number": "",
                        },
                    )

                    # Generate proper unique_id
                    if mgmt_type == "on-shelf":
                        unique_id = construct_unique_id(
                            scenario_id=str(scenario_id),
                            kit_code=None,
                            module_code=None,
                            item_code=code,
                            std_qty=std_info["std_qty"],
                            exp_date=stock_entry.get("earliest_expiry", ""),
                            force_box_format=False,
                        )
                    else:
                        unique_id = construct_unique_id(
                            scenario_id=str(scenario_id),
                            kit_code=std_info.get("kit_code"),
                            module_code=std_info.get("module_code"),
                            item_code=code,
                            std_qty=std_info["std_qty"],
                            exp_date=stock_entry.get("earliest_expiry", ""),
                            kit_number=stock_entry.get("kit_number"),
                            module_number=stock_entry.get("module_number"),
                            force_box_format=True,
                            treecode=std_info.get("treecode"),
                        )

                    # Check if already in tree
                    if unique_id not in self.user_row_states:
                        item = {
                            "unique_id": unique_id,
                            "code": code,
                            "description": get_active_designation(code),
                            "type": get_item_type(code),
                            "scenario": scenario_name,
                            "kit_number": stock_entry.get("kit_number", "-----"),
                            "module_number": stock_entry.get("module_number", "-----"),
                            "current_stock": stock_entry["current_stock"],
                            "exp_date": stock_entry.get("earliest_expiry", ""),
                            "std_qty": std_info["std_qty"],
                        }

                        self.insert_tree_row(item, physical_qty="")
                        added += 1

        self._highlight_missing_required_expiry()

        self.status_var.set(
            lang.t(
                "stock_inv.added_items", "Added {count} items for code {code}"
            ).format(count=added, code=code)
        )

        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)

    # ---------- Items in stock with filters ----------

    def fetch_inventory_items(self):
        """
        ✅ FIXED: Load ALL items from standard quantities (compositions + kit_items)
        Uses ENGLISH constants for backend filtering, translations only for display
        """
        scenario_filter = self.scenario_var.get()
        mgmt_mode = self.mgmt_mode_var.get()
        kit_filter = self.kit_number_var.get()
        module_filter = self.module_number_var.get()

        # ✅ FIX: Convert "All Scenarios" to None
        all_scenarios_text = lang.t("stock_inv.all_scenarios", "All Scenarios")
        if scenario_filter == all_scenarios_text or not scenario_filter:
            scenario_filter = None

        # Get standard quantities for all items
        std_data_by_scenario = load_std_quantities_by_scenario(scenario_filter)

        # ✅ Get current stock data (now grouped by expiry!)
        stock_map = aggregate_stock_by_key(scenario_filter, mgmt_mode)

        items = []

        # ✅ Create a set to track which (scenario, key) combinations we've seen
        seen_combinations = set()

        # ✅ Get filter labels
        all_kits_label = lang.t("stock_inv.all_kits", "All Kits")
        standalone_label = lang.t("stock_inv.stand_alone_items", "Stand alone items")
        all_modules_label = lang.t("stock_inv.all_modules", "All Modules")

        # ✅ NEW: Map translated mgmt_mode back to English for DB comparison
        mgmt_mode_english = None
        if mgmt_mode == lang.t("stock_inv.management_on_shelf", "On-Shelf"):
            mgmt_mode_english = "on-shelf"
        elif mgmt_mode == lang.t("stock_inv.management_in_box", "In-Box"):
            mgmt_mode_english = "in-box"
        # If "All" or empty, leave as None to show all

        # First pass: Add rows from EXISTING stock (with actual expiries)
        for stock_key, stock_entry in stock_map.items():
            scenario_name, key, exp_date = stock_key

            # Find matching std data
            std_data = std_data_by_scenario.get(scenario_name, {})
            std_info = std_data.get(key)

            if not std_info:
                continue

            code = std_info["code"]
            mgmt_type = std_info[
                "mgmt_type"
            ]  # This is ALWAYS "on-shelf" or "in-box" from DB
            scenario_id = std_info["scenario_id"]

            # ✅ FIXED: Compare using English constants, NOT translated values
            if mgmt_mode_english:
                if mgmt_type != mgmt_mode_english:
                    continue

            # Get kit/module numbers from stock
            kit_num = stock_entry.get("kit_number", "")
            mod_num = stock_entry.get("module_number", "")

            # ✅ STANDALONE DETECTION: Items with no kit/module numbers
            is_standalone = (not kit_num or kit_num == "-----") and (
                not mod_num or mod_num == "-----"
            )

            # ✅ FIXED: Apply kit number filter
            if kit_filter and kit_filter != all_kits_label:
                if kit_filter == standalone_label:
                    if not is_standalone:
                        continue
                else:
                    if kit_num != kit_filter:
                        continue

            # ✅ FIXED: Apply module number filter
            if module_filter and module_filter != all_modules_label:
                if mod_num != module_filter:
                    continue

            # ✅ Generate unique_id with ACTUAL expiry date
            if mgmt_type == "on-shelf":
                unique_id = construct_unique_id(
                    scenario_id=str(scenario_id),
                    kit_code=None,
                    module_code=None,
                    item_code=code,
                    std_qty=std_info["std_qty"],
                    exp_date=exp_date,
                    force_box_format=False,
                )
            else:
                treecode = std_info.get("treecode")
                unique_id = construct_unique_id(
                    scenario_id=str(scenario_id),
                    kit_code=std_info.get("kit_code"),
                    module_code=std_info.get("module_code"),
                    item_code=code,
                    std_qty=std_info["std_qty"],
                    exp_date=exp_date,
                    kit_number=stock_entry.get("kit_number"),
                    module_number=stock_entry.get("module_number"),
                    force_box_format=True,
                    treecode=treecode,
                )

            items.append(
                {
                    "unique_id": unique_id,
                    "code": code,
                    "description": get_active_designation(code),
                    "type": get_item_type(code),
                    "mgmt_type": mgmt_type,  # Keep as English internally
                    "scenario": scenario_name,
                    "kit_number": kit_num if kit_num else "-----",
                    "module_number": mod_num if mod_num else "-----",
                    "current_stock": stock_entry["current_stock"],
                    "exp_date": exp_date,
                    "std_qty": std_info["std_qty"],
                    "treecode": std_info.get("treecode", code),
                    "comments": stock_entry.get("comments", ""),
                }
            )

            # Mark this combination as seen
            seen_combinations.add((scenario_name, key))

        # Second pass: Add rows for items WITHOUT stock (current_stock = 0)
        # Only in Complete Inventory mode
        if self.inv_type_var.get() == lang.t(
            "stock_inv.complete_inventory", "Complete Inventory"
        ):
            for scenario_name, std_data in std_data_by_scenario.items():
                for key, std_info in std_data.items():
                    # Skip if we already added this item (has stock)
                    if (scenario_name, key) in seen_combinations:
                        continue

                    code = std_info["code"]
                    mgmt_type = std_info["mgmt_type"]  # English: "on-shelf" or "in-box"
                    scenario_id = std_info["scenario_id"]

                    # ✅ FIXED: Compare using English constants
                    if mgmt_mode_english:
                        if mgmt_type != mgmt_mode_english:
                            continue

                    # ✅ FIXED: For items without stock, apply kit/module filters
                    if kit_filter and kit_filter != all_kits_label:
                        if kit_filter == standalone_label:
                            if mgmt_type != "in-box":
                                continue
                        else:
                            continue

                    if module_filter and module_filter != all_modules_label:
                        continue

                    # ✅ Generate unique_id WITHOUT expiry (blank)
                    if mgmt_type == "on-shelf":
                        unique_id = construct_unique_id(
                            scenario_id=str(scenario_id),
                            kit_code=None,
                            module_code=None,
                            item_code=code,
                            std_qty=std_info["std_qty"],
                            exp_date="",
                            force_box_format=False,
                        )
                    else:
                        treecode = std_info.get("treecode")
                        unique_id = construct_unique_id(
                            scenario_id=str(scenario_id),
                            kit_code=std_info.get("kit_code"),
                            module_code=std_info.get("module_code"),
                            item_code=code,
                            std_qty=std_info["std_qty"],
                            exp_date="",
                            kit_number="",
                            module_number="",
                            force_box_format=True,
                            treecode=treecode,
                        )

                    items.append(
                        {
                            "unique_id": unique_id,
                            "code": code,
                            "description": get_active_designation(code),
                            "type": get_item_type(code),
                            "mgmt_type": mgmt_type,  # Keep as English
                            "scenario": scenario_name,
                            "kit_number": "-----",
                            "module_number": "-----",
                            "current_stock": 0,
                            "exp_date": "",
                            "std_qty": std_info["std_qty"],
                            "treecode": std_info.get("treecode", code),
                        }
                    )

        # ✅ FIXED SORTING
        def sort_key(item):
            scenario = item["scenario"]
            mgmt_type = item["mgmt_type"]

            if mgmt_type == "in-box":
                group_key = item.get("treecode", item["code"])
            else:
                group_key = item["code"]

            exp_date = item["exp_date"] or "9999-12-31"

            return (scenario, group_key, exp_date)

        items.sort(key=sort_key)

        return items

    # ---------- State preservation helpers ----------
    def capture_current_rows(self):
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            uid = vals[0]
            # ✅ FIXED: Index shifted by 1 (management_type at index 4)
            base_ph = self.base_physical_inputs.get(
                iid,
                (
                    int(vals[10]) if str(vals[10]).isdigit() else 0
                ),  # ✅ physical_qty is now at index 10
            )
            self.user_row_states[uid] = {
                "unique_id": uid,
                "code": vals[1],
                "description": vals[2],
                "type": vals[3],
                "management_type": vals[4],  # ✅ NEW
                "scenario": vals[5],
                "kit_number": vals[6],
                "module_number": vals[7],
                "current_stock": vals[8],
                "exp_date": vals[9],
                "physical_qty": vals[10],
                "updated_exp_date": vals[11],
                "discrepancy": vals[12],
                "remarks": vals[13],
                "std_qty": vals[14],
                "base_physical": base_ph,
                "is_custom": uid.startswith("temp::")
                or (int(vals[8]) if str(vals[8]).isdigit() else 0)
                == 0,  # ✅ current_stock is now at index 8
            }

    def rebuild_tree_preserving_state(self):
        self.capture_current_rows()
        self.tree.delete(*self.tree.get_children())

        # ✅ NEW: Load ALL items from standard quantities
        items = self.fetch_inventory_items()

        # Check if we are in "Complete Inventory" mode
        is_complete_inventory = self.inv_type_var.get() == lang.t(
            "stock_inv.complete_inventory", "Complete Inventory"
        )

        for item in items:
            # Get preserved state if exists
            state = self.user_row_states.get(item["unique_id"])

            if state:
                phys = state["physical_qty"]
                upd = state["updated_exp_date"]
                disc = state["discrepancy"]
                remarks = state.get("remarks", "")
                base_ph = state.get("base_physical", 0)
            else:
                phys = ""
                upd = ""
                disc = ""
                remarks = ""
                base_ph = 0

            # ✅ FIXED: Only force 0 if no previous state exists AND in Complete Inventory mode
            if is_complete_inventory and not state:
                # New item in Complete Inventory - default to 0
                phys_to_insert = "0"
                base_ph_to_insert = 0
            else:
                # Preserve existing data (user may have already entered values)
                phys_to_insert = phys
                base_ph_to_insert = base_ph

            iid = self._insert_from_state(
                item, phys_to_insert, upd, disc, remarks, base_ph_to_insert
            )

            # ✅ REMOVED: Don't update stored state - preserve user input!

        # Recompute quantities
        self.recompute_all_physical_quantities()

        # Update status with count breakdown
        total = len(items)
        with_stock = sum(1 for item in items if item["current_stock"] > 0)
        missing = total - with_stock

        self.status_var.set(
            lang.t(
                "stock_inv.loaded_with_breakdown",
                "Loaded {total} items ({with_stock} with stock, {missing} missing)",
            ).format(total=total, with_stock=with_stock, missing=missing)
        )

    def _insert_from_state(
        self, item_dict, physical_qty, updated_exp, discrepancy, remarks, base_physical
    ):
        """Insert row with auto-filled updated_exp for existing rows."""
        phys_str = physical_qty if physical_qty and str(physical_qty).isdigit() else ""
        disc_str = discrepancy if discrepancy not in ("", None) else ""

        current_stock = item_dict.get("current_stock", 0)
        current_exp = item_dict.get("exp_date", "")

        if isinstance(current_stock, str):
            current_stock = int(current_stock) if current_stock.isdigit() else 0

        if current_stock > 0 and current_exp:
            auto_updated_exp = current_exp
        else:
            auto_updated_exp = updated_exp

        if current_stock > 0 and not remarks:
            remarks = item_dict.get("comments", "")

        # ✅ mgmt_type is ALWAYS English internally ("on-shelf" or "in-box")
        mgmt_type = item_dict.get("mgmt_type", "on-shelf")

        # ✅ Translate ONLY for display in the tree
        mgmt_label = (
            lang.t("stock_inv.on_shelf", "On-Shelf")
            if mgmt_type == "on-shelf"
            else lang.t("stock_inv.in_box", "In-Box")
        )

        iid = self.tree.insert(
            "",
            "end",
            values=(
                item_dict["unique_id"],  # 0
                item_dict["code"],  # 1
                item_dict["description"],  # 2
                item_dict["type"],  # 3
                mgmt_label,  # 4 ✅ TRANSLATED for display
                item_dict["scenario"],  # 5
                item_dict.get("kit_number", "-----"),  # 6
                item_dict.get("module_number", "-----"),  # 7
                item_dict["current_stock"],  # 8
                current_exp,  # 9
                phys_str,  # 10
                auto_updated_exp,  # 11
                disc_str,  # 12
                remarks,  # 13
                item_dict["std_qty"],  # 14
            ),
        )

        # ✅ Type normalization for highlighting (handles Module/Módulo)
        t = normalize_type_text(item_dict["type"] or "")
        if t == "KIT":
            self.tree.item(iid, tags=("kit_row",))
        elif t == "MODULE":
            self.tree.item(iid, tags=("module_row",))
        else:
            self.tree.item(iid, tags=())

        self.base_physical_inputs[iid] = base_physical
        return iid

    # ---------- In-box multiplier logic ----------

    def recompute_all_physical_quantities(self):
        """
        Universal quantity calculation engine.
        """
        # --- Phase 1: Calculate all final quantities ---
        kit_factors = {}
        module_factors = {}

        all_iids = self.tree.get_children()

        # First pass: Identify all KIT/MODULE rows and calculate their factors (0 or 1)
        for iid in all_iids:
            vals = self.tree.item(iid, "values")
            if not vals:
                continue

            unique_id = vals[0]
            is_physical = unique_id and unique_id.count("/") == 7
            if not is_physical:
                continue

            # ✅ FIXED: Normalize type text (handles Module/Módulo)
            row_type = normalize_type_text(vals[3] or "")
            base_input = self.base_physical_inputs.get(iid, 0)

            if row_type == "KIT":
                kit_number = vals[6]  # ✅ Updated index
                if kit_number and kit_number != "-----":
                    final_qty = 1 if base_input > 0 else 0
                    kit_factors[kit_number] = final_qty
                    self.tree.set(iid, "physical_qty", str(final_qty))

            elif row_type == "MODULE":
                module_number = vals[7]  # ✅ Updated index
                kit_number = vals[6]  # ✅ Updated index
                if module_number and module_number != "-----":
                    parent_kit_factor = kit_factors.get(kit_number, 1)
                    base_factor = 1 if base_input > 0 else 0
                    final_qty = base_factor * parent_kit_factor
                    module_factors[module_number] = final_qty
                    self.tree.set(iid, "physical_qty", str(final_qty))

                    # Enhanced popup for MODULEs zeroed out by a KIT
                    if base_input > 0 and final_qty == 0 and parent_kit_factor == 0:
                        custom_popup(
                            self,
                            lang.t("dialog_titles.info", "Info"),
                            lang.t(
                                "stock_inv.kit_zero_module_zero",
                                "Quantity for {code} {desc} will remain 0 as the kit number {kit} has 0 quantity.",
                            ).format(code=vals[1], desc=vals[2], kit=kit_number),
                            "info",
                        )
                        # State Synchronization: Reset base input to prevent repeated popups
                        self.base_physical_inputs[iid] = 0

        # Second pass: Calculate quantities for all other rows (ITEMS and non-physical)
        for iid in all_iids:
            vals = self.tree.item(iid, "values")
            if not vals:
                continue

            unique_id = vals[0]
            is_physical = unique_id and unique_id.count("/") == 7
            base_input = self.base_physical_inputs.get(iid, 0)

            # ✅ FIXED: Normalize type text (handles Module/Módulo)
            row_type = normalize_type_text(vals[3] or "")

            final_qty = base_input  # Default for non-physical rows

            if is_physical and row_type == "ITEM":
                kit_number = vals[6]  # ✅ Updated index
                module_number = vals[7]  # ✅ Updated index

                parent_kit_factor = kit_factors.get(kit_number, 1)
                parent_module_factor = module_factors.get(module_number, 1)

                final_qty = base_input * parent_kit_factor * parent_module_factor
                self.tree.set(iid, "physical_qty", str(final_qty))

                if base_input > 0 and final_qty == 0:
                    reason = ""
                    if parent_module_factor == 0:
                        reason = (
                            lang.t("stock_inv.module_number", "module number")
                            + f" {module_number}"
                        )
                    elif parent_kit_factor == 0:
                        reason = (
                            lang.t("stock_inv.kit_number", "kit number")
                            + f" {kit_number}"
                        )

                    if reason:
                        custom_popup(
                            self,
                            lang.t("dialog_titles.info", "Info"),
                            lang.t(
                                "stock_inv.item_zero_reason",
                                "Quantity for {code} {desc} will remain 0 as the {reason} has 0 quantity.",
                            ).format(code=vals[1], desc=vals[2], reason=reason),
                            "info",
                        )
                        # State Synchronization: Reset base input
                        self.base_physical_inputs[iid] = 0

            elif not is_physical:
                self.tree.set(iid, "physical_qty", str(final_qty))

        # --- Phase 2: Update discrepancies for all rows ---
        for iid in all_iids:
            vals = self.tree.item(iid, "values")
            if not vals:
                continue

            current_stock = (
                int(vals[8]) if str(vals[8]).isdigit() else 0
            )  # ✅ Updated index
            physical_qty = (
                int(vals[10]) if str(vals[10]).isdigit() else 0
            )  # ✅ Updated index
            discrepancy = physical_qty - current_stock

            self.tree.set(
                iid, "discrepancy", "" if discrepancy == 0 else str(discrepancy)
            )
            self._update_state_from_row(iid)

        # --- Phase 3: Re-apply all visual highlights ---
        self._highlight_missing_required_expiry()

    # ---------- Document number generation ----------
    def generate_document_number(self):
        project_name, project_code = self.fetch_project_details()
        project_code = (project_code or "PRJ").strip().upper()
        inv_type_label = self.inv_type_var.get()
        abbr = (
            "CINV"
            if inv_type_label
            == lang.t("stock_inv.complete_inventory", "Complete Inventory")
            else "PINV"
        )
        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"
        conn = connect_db()
        serial = 1
        if conn:
            cur = conn.cursor()
            try:
                cur.execute(
                    """
                    SELECT document_number FROM stock_transactions
                    WHERE document_number LIKE ?
                    ORDER BY document_number DESC LIMIT 1
                """,
                    (prefix + "/%",),
                )
                row = cur.fetchone()
                if row and row[0]:
                    last_serial = row[0].rsplit("/", 1)[-1]
                    if last_serial.isdigit():
                        serial = int(last_serial) + 1
            finally:
                cur.close()
                conn.close()
        return f"{prefix}/{serial:04d}"

    # --------------Adopted Expiry warning----------

    def _show_adopted_expiry_warning(self):
        """
        Show warning dialog if any items have adopted expiries in remarks/comments AND physical quantity > 0.
        Returns True to proceed with save, False to go back for review.

        Returns:
            bool: True = proceed with save, False = cancel and review
        """
        # Count items with adopted expiries AND physical qty > 0
        adopted_count = 0
        adopted_items = []

        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals:
                continue

            code = vals[1]
            remarks = (vals[13] or "").lower()  # Remarks column (index 13)
            physical_qty = vals[10]  # Physical quantity (index 10)

            # ✅ Parse physical quantity
            phys_qty_int = 0
            if physical_qty and str(physical_qty).isdigit():
                phys_qty_int = int(physical_qty)

            # ✅ Check if "adopted_exp" or "adopted expiry" appears in remarks AND physical qty > 0
            if phys_qty_int > 0 and (
                "adopted_exp" in remarks or "adopted expiry" in remarks
            ):
                adopted_count += 1
                adopted_items.append(code)

        # If no adopted expiries with qty > 0, proceed directly
        if adopted_count == 0:
            return True

        # ✅ Show warning dialog
        # Create custom dialog
        dialog = tk.Toplevel(self)
        dialog.title(lang.t("stock_inv.warning_title", "Warning - Adopted Expiries"))
        dialog.geometry("550x550")
        dialog.transient(self)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # Result variable
        result = {"proceed": False}

        # Main frame
        main_frame = tk.Frame(dialog, bg="#FFF3E0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Warning icon and title
        title_frame = tk.Frame(main_frame, bg="#FFF3E0")
        title_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            title_frame, text="⚠️", font=("Helvetica", 32), bg="#FFF3E0", fg="#FF6F00"
        ).pack(side="left", padx=(0, 10))

        tk.Label(
            title_frame,
            text=lang.t(
                "stock_inv.adopted_expiry_warning_title",
                "Adopted Expiry Dates Detected",
            ),
            font=("Helvetica", 14, "bold"),
            bg="#FFF3E0",
            fg="#E65100",
        ).pack(side="left")

        # Warning message frame
        msg_frame = tk.Frame(main_frame, bg="white", relief="solid", borderwidth=1)
        msg_frame.pack(fill="both", expand=True, pady=(0, 15))

        # Count label
        count_label = tk.Label(
            msg_frame,
            text=lang.t(
                "stock_inv.adopted_count",
                "{count} item(s) have expiry dates adopted from kit/module or other sources.",
            ).format(count=adopted_count),
            font=("Helvetica", 11, "bold"),
            bg="white",
            fg="#D84315",
            wraplength=480,
            justify="left",
        )
        count_label.pack(fill="x", padx=15, pady=(15, 10))

        # Warning text
        warning_text = lang.t(
            "stock_inv.adopted_expiry_warning",
            "This may cause issues in stock management.\n\n"
            "It is STRONGLY RECOMMENDED to verify the expiry dates manually "
            "and enter actual expiry dates before saving.\n\n"
            "Items with 'Adopted_Exp' in remarks should be reviewed carefully.",
        )

        warning_label = tk.Label(
            msg_frame,
            text=warning_text,
            font=("Helvetica", 10),
            bg="white",
            fg="#424242",
            wraplength=480,
            justify="left",
        )
        warning_label.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        # Recommendation box
        rec_frame = tk.Frame(main_frame, bg="#E3F2FD", relief="solid", borderwidth=1)
        rec_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            rec_frame,
            text="💡 " + lang.t("stock_inv.recommendation", "Recommendation"),
            font=("Helvetica", 10, "bold"),
            bg="#E3F2FD",
            fg="#1565C0",
            anchor="w",
        ).pack(fill="x", padx=10, pady=(8, 4))

        tk.Label(
            rec_frame,
            text=lang.t(
                "stock_inv.recommendation_text",
                "Click 'Review' to go back and verify expiry dates.\n"
                "Double-click any row to edit the expiry date and remarks.",
            ),
            font=("Helvetica", 9),
            bg="#E3F2FD",
            fg="#424242",
            anchor="w",
            justify="left",
        ).pack(fill="x", padx=10, pady=(0, 8))

        # Button frame
        btn_frame = tk.Frame(main_frame, bg="#FFF3E0")
        btn_frame.pack(side="bottom", pady=(10, 0))

        def on_review():
            """User wants to go back and review."""
            result["proceed"] = False
            dialog.destroy()

        def on_save():
            """User wants to proceed with save."""
            result["proceed"] = True
            dialog.destroy()

        # Review button (recommended action)
        review_btn = tk.Button(
            btn_frame,
            text=lang.t("stock_inv.review_button", "📋 Review Expiries"),
            font=("Helvetica", 11, "bold"),
            bg="#2196F3",
            fg="white",
            width=18,
            command=on_review,
            relief="flat",
            cursor="hand2",
            padx=15,
            pady=8,
        )
        review_btn.pack(side="left", padx=5)

        # Save anyway button (warning action)
        save_btn = tk.Button(
            btn_frame,
            text=lang.t("stock_inv.save_anyway_button", "⚠️ Save Anyway"),
            font=("Helvetica", 11),
            bg="#FF9800",
            fg="white",
            width=18,
            command=on_save,
            relief="flat",
            cursor="hand2",
            padx=15,
            pady=8,
        )
        save_btn.pack(side="left", padx=5)

        # Bind keys
        dialog.bind("<Escape>", lambda e: on_review())
        dialog.bind("<Return>", lambda e: on_review())  # Default = Review

        # Wait for dialog
        dialog.wait_window()

        return result["proceed"]

    # ---------- Export ----------
    def export_to_excel(self, rows_to_export=None, document_number=None):
        """
        Export inventory to Excel WITHOUT user prompt.
        Auto-generates filename and saves to default directory.
        """
        try:
            # Create default directory if it doesn't exist
            default_dir = "D:/ISEPREP"
            os.makedirs(default_dir, exist_ok=True)

            # Auto-generate filename
            inv_type = self.inv_type_var.get().replace(" ", "_")
            mgmt_mode = self.mgmt_mode_var.get().replace(" ", "_")
            current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_name = (
                f"IsEPREP_Stock-Inventory_{inv_type}_{mgmt_mode}_{current_time}.xlsx"
            )

            # Auto-generate full path (NO user prompt)
            path = os.path.join(default_dir, file_name)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = lang.t("stock_inv.stock_inventory", "Stock Inventory")

            # Get project details
            project_name, project_code = self.fetch_project_details()
            inv_type_label = self.inv_type_var.get()
            mgmt_mode_label = self.mgmt_mode_var.get()
            scenario = self.scenario_var.get()
            kit_filter = self.kit_number_var.get()
            module_filter = self.module_number_var.get()
            current_date = datetime.now().strftime("%Y-%m-%d")

            # ✅ FIXED: Headers now span 13 columns (A to M) instead of 12
            # Header row 1: Title
            ws["A1"] = lang.t("stock_inv.stock_inventory", "Stock Inventory")
            ws["A1"].font = Font(name="Tahoma", size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A1:M1")  # ✅ Changed from L1 to M1

            # Header row 2: Project info
            ws["A2"] = f"{project_name} - {project_code}"
            ws["A2"].font = Font(name="Tahoma", size=14)
            ws["A2"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A2:M2")  # ✅ Changed from L2 to M2

            # Header row 3: Filters
            ws["A3"] = (
                f"{lang.t('stock_inv.inventory_type', 'Inventory Type')}: {inv_type_label}, "
                f"{lang.t('stock_inv.management_mode', 'Management Mode')}: {mgmt_mode_label}, "
                f"{lang.t('stock_inv.scenario', 'Scenario')}: {scenario}, "
                f"{lang.t('stock_inv.kit_number', 'Kit Number')}: {kit_filter}, "
                f"{lang.t('stock_inv.module_number', 'Module Number')}: {module_filter}"
            )
            ws["A3"].font = Font(name="Tahoma")
            ws["A3"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A3:M3")  # ✅ Changed from L3 to M3

            # Header row 4: Date
            ws["A4"] = (
                f"{lang.t('stock_inv.inventory_date', 'Inventory Date')}: {current_date}"
            )
            ws["A4"].font = Font(name="Tahoma")
            ws["A4"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A4:M4")  # ✅ Changed from L4 to M4

            # Header row 5: Document number (if provided)
            if document_number:
                ws["A5"] = (
                    f"{lang.t('stock_inv.document_number', 'Document Number')}: {document_number}"
                )
                ws["A5"].font = Font(name="Tahoma", bold=True)
                ws["A5"].alignment = Alignment(horizontal="right")
                ws.merge_cells("A5:M5")  # ✅ Changed from L5 to M5
                ws.append([])  # Blank row after document number

            # Column headers (13 columns total)
            headers = [
                lang.t("stock_inv.code", "Code"),
                lang.t("stock_inv.description", "Description"),
                lang.t("stock_inv.type", "Type"),
                lang.t("stock_inv.management_type", "Management"),  # ✅ Column 4
                lang.t("stock_inv.scenario", "Scenario"),
                lang.t("stock_inv.kit_number", "Kit Number"),
                lang.t("stock_inv.module_number", "Module Number"),
                lang.t("stock_inv.current_stock", "Current Stock"),
                lang.t("stock_inv.expiry_date", "Expiry Date"),
                lang.t("stock_inv.physical_qty", "Physical Quantity"),
                lang.t("stock_inv.updated_exp_date", "Updated Expiry Date"),
                lang.t("stock_inv.discrepancy", "Discrepancy"),
                lang.t("stock_inv.remarks", "Remarks"),  # ✅ Column 13 (M)
            ]
            ws.append(headers)

            # Define fill colors for highlighting
            kit_fill = PatternFill(
                start_color="D8F5D0", end_color="D8F5D0", fill_type="solid"
            )
            module_fill = PatternFill(
                start_color="D5ECFF", end_color="D5ECFF", fill_type="solid"
            )
            exp_warn_fill = PatternFill(
                start_color="FFD8D8", end_color="FFD8D8", fill_type="solid"
            )

            # Get rows to export (either provided or from tree)
            rows_data = rows_to_export or [
                {
                    "code": vals[1],
                    "description": vals[2],
                    "type": vals[3],
                    "management_type": vals[4],  # ✅ NEW
                    "scenario": vals[5],
                    "kit_number": vals[6],
                    "module_number": vals[7],
                    "current_stock": vals[8],
                    "exp_date": vals[9],
                    "physical_qty": vals[10],
                    "updated_exp_date": vals[11],
                    "discrepancy": vals[12],
                    "remarks": vals[13],
                }
                for item in self.tree.get_children()
                if (vals := self.tree.item(item)["values"])
            ]

            # Write data rows with color coding
            row_start = ws.max_row + 1
            for idx, row in enumerate(rows_data, start=row_start):
                ws.append(
                    [
                        row["code"],
                        row["description"],
                        row["type"],
                        row["management_type"],  # ✅ NEW
                        row["scenario"],
                        row["kit_number"],
                        row["module_number"],
                        row["current_stock"],
                        row["exp_date"],
                        row["physical_qty"],
                        row["updated_exp_date"],
                        row["discrepancy"],
                        row["remarks"],
                    ]
                )

                # ✅ FIXED: Normalize type for color coding (handles Módulo/Module)
                t = normalize_type_text(row["type"] or "")
                fill = None
                if t == "KIT":
                    fill = kit_fill
                elif t == "MODULE":
                    fill = module_fill

                # Highlight missing required expiry
                if check_expiry_required(row["code"]) and not (
                    row["updated_exp_date"] or row["exp_date"]
                ):
                    fill = exp_warn_fill

                if fill:
                    for c in ws[f"A{idx}:M{idx}"]:  # ✅ Changed from L{idx} to M{idx}
                        for cell in c:
                            cell.fill = fill

            # ✅ Set column widths (13 columns now)
            widths = [100, 300, 100, 100, 120, 120, 130, 110, 110, 120, 130, 110, 200]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w / 7

            # Set page layout for printing
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1

            # Save workbook
            wb.save(path)
            wb.close()

            # Show success message with file location
            custom_popup(
                self,
                lang.t("dialog_titles.success", "Success"),
                lang.t("stock_inv.exported", "Exported to {path}").format(path=path),
                "info",
            )
            self.status_var.set(
                lang.t("stock_inv.export_success", "Export successful: {path}").format(
                    path=path
                )
            )

        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("stock_inv.export_failed", "Export failed: {error}").format(
                    error=str(e)
                ),
                "error",
            )
            self.status_var.set(
                lang.t("stock_inv.export_error", "Export error: {error}").format(
                    error=str(e)
                )
            )

    # ---------- UI ----------
    def render_ui(self):
        for w in self.winfo_children():
            w.destroy()

        tk.Label(
            self,
            text=lang.t("stock_inv.title", "Stock Inventory Adjustment"),
            font=("Helvetica", 20, "bold"),
            bg="#F5F5F5",
        ).pack(pady=10, anchor="w")

        filter_frame = tk.Frame(self, bg="#F5F5F5")
        filter_frame.pack(pady=5, fill="x")

        tk.Label(
            filter_frame, text=lang.t("stock_inv.scenario", "Scenario:"), bg="#F5F5F5"
        ).grid(row=0, column=0, padx=5, sticky="w")
        self.scenario_var = tk.StringVar(
            value=lang.t("stock_inv.all_scenarios", "All Scenarios")
        )
        scenarios = list(self.scenario_map.values())
        all_scen = lang.t("stock_inv.all_scenarios", "All Scenarios")
        if all_scen not in scenarios:
            scenarios.append(all_scen)
        self.scenario_cb = ttk.Combobox(
            filter_frame,
            textvariable=self.scenario_var,
            values=scenarios,
            state="readonly",
            width=20,
        )
        self.scenario_cb.grid(row=0, column=1, padx=5, pady=5)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)

        tk.Label(
            filter_frame,
            text=lang.t("stock_inv.management_mode", "Management Mode:"),
            bg="#F5F5F5",
        ).grid(row=0, column=2, padx=5, sticky="w")
        self.mgmt_mode_var = tk.StringVar(
            value=lang.t("stock_inv.management_all", "All")
        )
        self.mgmt_mode_cb = ttk.Combobox(
            filter_frame,
            textvariable=self.mgmt_mode_var,
            values=[
                lang.t("stock_inv.management_all", "All"),
                lang.t("stock_inv.management_on_shelf", "On-Shelf"),
                lang.t("stock_inv.management_in_box", "In-Box"),
            ],
            state="readonly",
            width=14,
        )
        self.mgmt_mode_cb.grid(row=0, column=3, padx=5, pady=5)
        self.mgmt_mode_cb.bind("<<ComboboxSelected>>", self.on_management_mode_change)

        tk.Label(
            filter_frame,
            text=lang.t("stock_inv.kit_number", "Kit Number:"),
            bg="#F5F5F5",
        ).grid(row=0, column=4, padx=5, sticky="w")
        self.kit_number_var = tk.StringVar(
            value=lang.t("stock_inv.all_kits", "All Kits")
        )
        self.kit_number_cb = ttk.Combobox(
            filter_frame,
            textvariable=self.kit_number_var,
            values=[lang.t("stock_inv.all_kits", "All Kits")],
            state="readonly",
            width=14,
        )
        self.kit_number_cb.grid(row=0, column=5, padx=5, pady=5)
        self.kit_number_cb.bind("<<ComboboxSelected>>", self.on_filter_change)

        tk.Label(
            filter_frame,
            text=lang.t("stock_inv.module_number", "Module Number:"),
            bg="#F5F5F5",
        ).grid(row=0, column=6, padx=5, sticky="w")
        self.module_number_var = tk.StringVar(
            value=lang.t("stock_inv.all_modules", "All Modules")
        )
        self.module_number_cb = ttk.Combobox(
            filter_frame,
            textvariable=self.module_number_var,
            values=[lang.t("stock_inv.all_modules", "All Modules")],
            state="readonly",
            width=14,
        )
        self.module_number_cb.grid(row=0, column=7, padx=5, pady=5)
        self.module_number_cb.bind("<<ComboboxSelected>>", self.on_filter_change)

        # Inventory Type
        type_frame = tk.Frame(self, bg="#F5F5F5")
        type_frame.pack(pady=5, fill="x")
        tk.Label(
            type_frame,
            text=lang.t("stock_inv.inventory_type", "Inventory Type:"),
            bg="#F5F5F5",
        ).grid(row=0, column=0, padx=5, sticky="w")
        self.inv_type_var = tk.StringVar(
            value=lang.t("stock_inv.complete_inventory", "Complete Inventory")
        )
        self.inv_type_cb = ttk.Combobox(
            type_frame,
            textvariable=self.inv_type_var,
            state="readonly",
            values=[
                lang.t("stock_inv.complete_inventory", "Complete Inventory"),
                lang.t("stock_inv.partial_inventory", "Partial Inventory"),
            ],
            width=28,
        )
        self.inv_type_cb.grid(row=0, column=1, padx=5, pady=5)
        self.inv_type_cb.bind("<<ComboboxSelected>>", self.on_inv_type_selected)

        # Search (partial only)
        self.search_frame = tk.Frame(self, bg="#F5F5F5")
        tk.Label(
            self.search_frame,
            text=lang.t("stock_inv.search", "Search Code or Description:"),
            bg="#F5F5F5",
        ).grid(row=0, column=0, padx=5, sticky="w")
        self.code_entry = tk.Entry(self.search_frame, width=55)
        self.code_entry.grid(row=0, column=1, padx=5, pady=5)
        self.code_entry.bind("<KeyRelease>", self.on_search_keyrelease)
        self.code_entry.bind("<Return>", self.select_first_search_result)
        self.search_listbox = tk.Listbox(self.search_frame, height=5, width=50)
        self.search_listbox.grid(row=0, column=2, padx=5, pady=5)
        self.search_listbox.bind("<<ListboxSelect>>", self.on_search_select)

        # ✅ NEW: Clear Search Button
        tk.Button(
            self.search_frame,
            text=lang.t("stock_inv.clear_search", "Clear Search"),
            command=self.clear_search,
            bg="#95A5A6",
            fg="white",
        ).grid(row=0, column=3, padx=5, pady=5)

        # ✅ NEW: Real-time Tree Filter
        filter_row_frame = tk.Frame(self, bg="#F5F5F5")
        filter_row_frame.pack(pady=5, fill="x")
        tk.Label(
            filter_row_frame,
            text=lang.t("stock_inv.filter_items", "Filter Displayed Items:"),
            bg="#F5F5F5",
        ).pack(side="left", padx=5)
        self.filter_entry = tk.Entry(filter_row_frame, width=40)
        self.filter_entry.pack(side="left", padx=5)
        self.filter_entry.bind("<KeyRelease>", self.on_filter_keyrelease)
        tk.Button(
            filter_row_frame,
            text=lang.t("stock_inv.clear_filter", "Clear Filter"),
            command=self.clear_tree_filter,
            bg="#95A5A6",
            fg="white",
        ).pack(side="left", padx=5)

        # Tree
        tree_frame = tk.Frame(self)
        tree_frame.pack(expand=True, fill="both", pady=10)

        # ✅ UPDATED: Add management_type column
        self.cols = (
            "unique_id",
            "code",
            "description",
            "type",
            "management_type",  # ✅ NEW COLUMN
            "scenario",
            "kit_number",
            "module_number",
            "current_stock",
            "exp_date",
            "physical_qty",
            "updated_exp_date",
            "discrepancy",
            "remarks",
            "std_qty",
        )

        self.display_cols = (
            "code",
            "description",
            "type",
            "management_type",  # ✅ NEW COLUMN
            "scenario",
            "kit_number",
            "module_number",
            "current_stock",
            "exp_date",
            "physical_qty",
            "updated_exp_date",
            "discrepancy",
            "remarks",
        )

        self.tree = ttk.Treeview(
            tree_frame,
            columns=self.cols,
            show="headings",
            height=18,
            displaycolumns=self.display_cols,
        )
        self.tree.tag_configure("light_red", background="#FF9999")
        self.tree.tag_configure("kit_row", background="#D8F5D0")
        self.tree.tag_configure("module_row", background="#D5ECFF")

        headers = {
            "unique_id": lang.t("stock_inv.unique_id", "Unique ID"),
            "code": lang.t("stock_inv.code", "Code"),
            "description": lang.t("stock_inv.description", "Description"),
            "type": lang.t("stock_inv.type", "Type"),
            "management_type": lang.t(
                "stock_inv.management_type", "Management"
            ),  # ✅ NEW
            "scenario": lang.t("stock_inv.scenario", "Scenario"),
            "kit_number": lang.t("stock_inv.kit_number", "Kit Number"),
            "module_number": lang.t("stock_inv.module_number", "Module Number"),
            "current_stock": lang.t("stock_inv.current_stock", "Current Stock"),
            "exp_date": lang.t("stock_inv.expiry_date", "Expiry Date"),
            "physical_qty": lang.t("stock_inv.physical_qty", "Physical Quantity"),
            "updated_exp_date": lang.t(
                "stock_inv.updated_exp_date", "Updated Expiry Date"
            ),
            "discrepancy": lang.t("stock_inv.discrepancy", "Discrepancy"),
            "remarks": lang.t("stock_inv.remarks", "Remarks"),
            "std_qty": lang.t("stock_inv.std_qty", "Standard Qty"),
        }

        widths = {
            "unique_id": 0,
            "code": 110,
            "description": 360,
            "type": 80,
            "management_type": 100,  # ✅ NEW
            "scenario": 110,
            "kit_number": 110,
            "module_number": 120,
            "current_stock": 90,
            "exp_date": 100,
            "physical_qty": 110,
            "updated_exp_date": 120,
            "discrepancy": 90,
            "remarks": 180,
            "std_qty": 80,
        }

        for c in self.cols:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths.get(c, 100), anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        btn_frame = tk.Frame(self, bg="#F5F5F5")
        btn_frame.pack(pady=5, fill="x")
        actions = [
            (
                "Save Adjustments",
                "#27AE60",
                self.save_all,
                "normal" if self.role in ["admin", "manager"] else "disabled",
            ),
            ("Clear All", "#7F8C8D", self.clear_form, "normal"),
            (
                "Export to Excel",
                "#3B82F6",
                self.export_to_excel,
                "normal" if self.role in ["admin", "manager"] else "disabled",
            ),
        ]
        for txt, bg, cmd, st in actions:
            tk.Button(
                btn_frame,
                text=lang.t(f"stock_inv.{txt.lower().replace(' ', '_')}", txt),
                bg=bg,
                fg="white",
                command=cmd,
                state=st,
            ).pack(side="left", padx=5)

        batch_frame = tk.Frame(self, bg="#F5F5F5")
        batch_frame.pack(pady=5, fill="x")
        self.batch_info_var = tk.StringVar()
        tk.Label(
            batch_frame,
            text=lang.t("stock_inv.batch_info", "Batch Information:"),
            bg="#F5F5F5",
        ).grid(row=0, column=0, padx=5, sticky="w")
        tk.Label(
            batch_frame,
            textvariable=self.batch_info_var,
            bg="#F5F5F5",
            wraplength=600,
            anchor="w",
        ).grid(row=0, column=1, padx=5, sticky="w")

        self.status_var = tk.StringVar(value="")
        tk.Label(self, textvariable=self.status_var, bg="#F5F5F5", anchor="w").pack(
            pady=5, anchor="w"
        )

        self.tree.bind("<Double-1>", self.start_edit)
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.search_listbox.bind("<<ListboxSelect>>", self.on_search_select)
        self.tree.bind("<Tab>", self.on_tab_press)

        self.refresh_kit_dropdown()
        self.refresh_module_dropdown()
        self.on_inv_type_selected()

    def _is_valid_future_updated_exp(self, updated_exp_str: str) -> bool:
        """
        Return True only if updated_exp_str is a valid parsable future date.
        Empty / None returns False.
        """
        if not updated_exp_str:
            return False
        parsed = parse_expiry(updated_exp_str)
        if not parsed:
            return False
        d = parsed.date() if hasattr(parsed, "date") else parsed
        return d > datetime.now().date()

    def _highlight_missing_required_expiry(self):
        """
        Highlight rows (light_red) ONLY when:
        - Item requires expiry (remarks contains 'exp')
        - Physical quantity > 0
        - AND (updated_exp_date is missing OR invalid / not future)
        """
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals:
                continue
            code = vals[1]
            phys = vals[10]  # ✅ Updated index
            updated = vals[11]  # ✅ Updated index
            t = (vals[3] or "").upper()
            phys_int = int(phys) if phys.isdigit() else 0
            requires = check_expiry_required(code)

            if requires and phys_int > 0:
                if self._is_valid_future_updated_exp(updated):
                    # Valid updated future expiry -> normal coloring
                    if t == "KIT":
                        self.tree.item(iid, tags=("kit_row",))
                    elif t == "MODULE":
                        self.tree.item(iid, tags=("module_row",))
                    else:
                        self.tree.item(iid, tags=())
                else:
                    # Missing or invalid updated future expiry -> highlight
                    self.tree.item(iid, tags=("light_red",))
            else:
                # Not requiring OR phys not positive -> normal coloring
                if t == "KIT":
                    self.tree.item(iid, tags=("kit_row",))
                elif t == "MODULE":
                    self.tree.item(iid, tags=("module_row",))
                else:
                    self.tree.item(iid, tags=())

    def _show_expiry_change_instructions(self, code, description):
        """
        Show detailed instructions popup for how to re-date existing stock.
        This is shown when user tries to edit updated_exp_date for an existing row.
        """
        message = lang.t(
            "stock_inv.expiry_change_blocked",
            "Cannot change expiry date for existing stock.\n\n"
            "To re-date stock for {code} - {desc}:\n\n"
            "1. Set 'Physical Quantity' to the amount you want to KEEP with the current expiry date\n"
            "   (or 0 to move all stock to new expiry)\n\n"
            "2. Right-click this row and select 'Add New Row'\n\n"
            "3. In the new row:\n"
            "   - Enter the quantity for the NEW expiry date\n"
            "   - Fill in 'Updated Expiry Date' with the new date\n\n"
            "4. Click 'Save Adjustments'\n\n"
            "Example:\n"
            "- Current: 100 units, Expiry Dec-26\n"
            "- Want: 50 units Dec-26, 50 units Dec-29\n"
            "  → Row 1 (existing): Physical = 50 (keeps Dec-26)\n"
            "  → Row 2 (new): Physical = 50, Updated Expiry = Dec-29",
        ).format(code=code, desc=description)

        # Create custom dialog
        dialog = tk.Toplevel(self)
        dialog.title(
            lang.t("stock_inv.expiry_change_title", "Cannot Change Expiry Date")
        )
        dialog.geometry("650x720")
        dialog.transient(self)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # Main frame
        main_frame = tk.Frame(dialog, bg="#FFF3E0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Warning icon and title
        title_frame = tk.Frame(main_frame, bg="#FFF3E0")
        title_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            title_frame, text="🚫", font=("Helvetica", 32), bg="#FFF3E0", fg="#E65100"
        ).pack(side="left", padx=(0, 10))

        tk.Label(
            title_frame,
            text=lang.t("stock_inv.expiry_change_title", "Cannot Change Expiry Date"),
            font=("Helvetica", 14, "bold"),
            bg="#FFF3E0",
            fg="#D84315",
        ).pack(side="left")

        # Message frame with scrollbar
        msg_frame = tk.Frame(main_frame, bg="white", relief="solid", borderwidth=1)
        msg_frame.pack(fill="both", expand=True, pady=(0, 15))

        # Text widget for scrollable content
        text_widget = tk.Text(
            msg_frame,
            wrap="word",
            font=("Courier", 10),
            bg="white",
            fg="#424242",
            padx=15,
            pady=15,
            relief="flat",
        )
        text_widget.pack(fill="both", expand=True)
        text_widget.insert("1.0", message)
        text_widget.config(state="disabled")

        # Info box
        info_frame = tk.Frame(main_frame, bg="#E3F2FD", relief="solid", borderwidth=1)
        info_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            info_frame,
            text="ℹ️ " + lang.t("stock_inv.why_blocked", "Why is this blocked?"),
            font=("Helvetica", 10, "bold"),
            bg="#E3F2FD",
            fg="#1565C0",
            anchor="w",
        ).pack(fill="x", padx=10, pady=(8, 4))

        tk.Label(
            info_frame,
            text=lang.t(
                "stock_inv.why_blocked_text",
                "Existing stock already has a recorded expiry date in the system.\n"
                "To maintain accurate stock tracking, you must create separate records\n"
                "for different expiry dates.",
            ),
            font=("Helvetica", 9),
            bg="#E3F2FD",
            fg="#424242",
            anchor="w",
            justify="left",
        ).pack(fill="x", padx=10, pady=(0, 8))

        # Button frame
        btn_frame = tk.Frame(main_frame, bg="#FFF3E0")
        btn_frame.pack(side="bottom", pady=(10, 0))

        def on_close():
            dialog.destroy()

        # OK button
        ok_btn = tk.Button(
            btn_frame,
            text=lang.t("stock_inv.understood", "✓ I Understand"),
            font=("Helvetica", 11, "bold"),
            bg="#4CAF50",
            fg="white",
            width=20,
            command=on_close,
            relief="flat",
            cursor="hand2",
            padx=15,
            pady=8,
        )
        ok_btn.pack(pady=5)

        # Bind keys
        dialog.bind("<Escape>", lambda e: on_close())
        dialog.bind("<Return>", lambda e: on_close())

        # Wait for dialog
        dialog.wait_window()

    # ---------- Tab navigation ----------
    def on_tab_press(self, event):
        row_id = self.tree.focus()
        if not row_id:
            return "break"
        cols_order = ["physical_qty", "updated_exp_date", "remarks"]
        # Determine current column from focus not reliable; start next cycle
        # Always move to next editable column
        current_values = self.tree.item(row_id, "values")
        # Find which one is currently being edited by scanning focus widget - simplify: always start with first blank or cycle
        # Just start editing first editable if none editing
        self._begin_inline_edit(row_id, cols_order[0])
        return "break"

    # ---------- Context menu ----------
    def show_context_menu(self, event):
        """Show context menu with validation for updated expiry."""
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return

        self.ctx_row = row_id
        vals = self.tree.item(row_id, "values")
        current_stock = int(vals[7]) if str(vals[7]).isdigit() else 0
        current_exp = vals[8]

        if self.ctx_menu:
            self.ctx_menu.destroy()

        self.ctx_menu = tk.Menu(self, tearoff=0)

        # Physical Qty - always editable
        self.ctx_menu.add_command(
            label=lang.t("stock_inv.edit_physical", "Edit Physical Qty"),
            command=lambda: self._begin_inline_edit(row_id, "physical_qty"),
        )

        # ✅ Updated Expiry - conditional
        if current_stock > 0 and current_exp:
            # Existing row - show instructions instead
            self.ctx_menu.add_command(
                label=lang.t("stock_inv.edit_expiry", "Edit Updated Expiry"),
                command=lambda: self._show_expiry_change_instructions(vals[1], vals[2]),
            )
        else:
            # New row or no current expiry - allow editing
            self.ctx_menu.add_command(
                label=lang.t("stock_inv.edit_expiry", "Edit Updated Expiry"),
                command=lambda: self._begin_inline_edit(row_id, "updated_exp_date"),
            )

        # Remarks - always editable
        self.ctx_menu.add_command(
            label=lang.t("stock_inv.edit_remarks", "Edit Remarks"),
            command=lambda: self._begin_inline_edit(row_id, "remarks"),
        )

        self.ctx_menu.add_separator()

        # ✅ Add new row (re-date same item)
        self.ctx_menu.add_command(
            label=lang.t("stock_inv.add_new_row_redate", "Add New Row (Re-date)"),
            command=lambda: self.add_new_row_below(row_id),
        )

        self.ctx_menu.add_separator()

        # Clear physical qty
        self.ctx_menu.add_command(
            label=lang.t("stock_inv.clear_physical", "Clear Physical Qty"),
            command=lambda: self._clear_physical(row_id),
        )

        # Remove row
        self.ctx_menu.add_command(
            label=lang.t("stock_inv.remove_row", "Remove Row"),
            command=lambda: (self.tree.delete(row_id), self._remove_state(row_id)),
        )

        self.ctx_menu.tk_popup(event.x_root, event.y_root)

    # -------------------------For Item search--------------------------
    def on_filter_keyrelease(self, event=None):
        """Filter tree rows in real-time based on code/description."""
        filter_text = self.filter_entry.get().strip().lower()

        if not filter_text:
            # Show all rows if filter is empty
            for iid in self.tree.get_children():
                self.tree.reattach(iid, "", "end")
            return

        # Hide rows that don't match filter
        for iid in self.tree.get_children():
            vals = self.tree.item(iid, "values")
            if not vals:
                continue

            code = (vals[1] or "").lower()
            description = (vals[2] or "").lower()

            # Check if filter text is in code or description
            if filter_text in code or filter_text in description:
                # Show row (reattach to tree)
                try:
                    self.tree.reattach(iid, "", "end")
                except:
                    pass  # Already attached
            else:
                # Hide row (detach from tree)
                try:
                    self.tree.detach(iid)
                except:
                    pass  # Already detached

        # Update status
        visible_count = len(
            [iid for iid in self.tree.get_children() if self.tree.parent(iid) == ""]
        )
        self.status_var.set(
            lang.t(
                "stock_inv.filtered_items", "Showing {count} items matching filter"
            ).format(count=visible_count)
        )

    def clear_tree_filter(self):
        """Clear the tree filter and show all rows."""
        self.filter_entry.delete(0, tk.END)

        # ✅ FIXED: Reload tree instead of just showing rows
        # This ensures all items are loaded, not just unhiding filtered ones
        self.rebuild_tree_preserving_state()

        self.status_var.set(
            lang.t("stock_inv.showing_all", "Showing all {count} items").format(
                count=len(self.tree.get_children())
            )
        )

    # ---------- State utility ----------
    def _remove_state(self, row_id):
        vals = self.tree.item(row_id, "values")
        if vals:
            uid = vals[0]
            self.user_row_states.pop(uid, None)
        self.base_physical_inputs.pop(row_id, None)

    def _update_state_from_row(self, row_id):
        vals = self.tree.item(row_id, "values")
        if not vals:
            return
        uid = vals[0]
        st = self.user_row_states.get(uid)
        if not st:
            return
        st["physical_qty"] = vals[9]
        st["updated_exp_date"] = vals[10]
        st["discrepancy"] = vals[11]
        st["remarks"] = vals[12]
        st["base_physical"] = self.base_physical_inputs.get(
            row_id, st.get("base_physical", 0)
        )

    # ---------- Insert row ----------
    def insert_tree_row(self, item, physical_qty=""):
        """Insert tree row with auto-filled updated_exp for existing rows."""
        current_stock = item.get("current_stock", 0)
        current_exp = item.get("exp_date", "")

        if isinstance(current_stock, str):
            current_stock = int(current_stock) if current_stock.isdigit() else 0

        # ✅ Auto-fill updated_exp_date for existing rows
        if current_stock > 0 and current_exp:
            auto_updated_exp = current_exp
        else:
            auto_updated_exp = ""  # New row - blank for user to fill

        iid = self.tree.insert(
            "",
            "end",
            values=(
                item["unique_id"],
                item["code"],
                item["description"],
                item["type"],
                item["scenario"],
                item.get("kit_number", "-----"),
                item.get("module_number", "-----"),
                current_stock,
                current_exp,
                "",  # physical qty blank
                auto_updated_exp,  # ✅ Auto-filled for existing, blank for new
                "",  # discrepancy blank
                "",  # remarks blank
                item["std_qty"],
            ),
        )
        t = item["type"].upper()
        if t == "KIT":
            self.tree.item(iid, tags=("kit_row",))
        elif t == "MODULE":
            self.tree.item(iid, tags=("module_row",))
        else:
            self.tree.item(iid, tags=())
        self.base_physical_inputs[iid] = 0
        self.user_row_states[item["unique_id"]] = {
            "unique_id": item["unique_id"],
            "code": item["code"],
            "description": item["description"],
            "type": item["type"],
            "scenario": item["scenario"],
            "kit_number": item.get("kit_number", "-----"),
            "module_number": item.get("module_number", "-----"),
            "current_stock": current_stock,
            "exp_date": current_exp,
            "physical_qty": "",
            "updated_exp_date": auto_updated_exp,  # ✅ Auto-filled
            "discrepancy": "",
            "remarks": "",
            "std_qty": item["std_qty"],
            "base_physical": 0,
            "is_custom": current_stock == 0,
        }
        return True

    def add_new_row_below(self, row_id):
        """
        Add a new row below the selected row with current stock = 0.
        Used for RE-DATING stock (same item, different expiry).
        Preserves parent row format (6-segment On-Shelf or 8-segment In-Box).
        """
        vals = self.tree.item(row_id, "values")
        if not vals:
            return

        # ✅ Parse the parent row's unique_id to get ALL original data
        parent_unique_id = vals[0]
        parsed = parse_inventory_unique_id(parent_unique_id)
        code = vals[1]

        # ✅ Determine if parent is In-Box format (8 segments) or On-Shelf (6 segments)
        is_in_box = len(parent_unique_id.split("/")) >= 8

        # Generate temporary unique ID for new row
        self.temp_row_counter += 1
        temp_uid = f"temp::{code}::{self.temp_row_counter}"

        # ✅ NEW RULE: New rows always have current_stock = 0 and blank updated_exp
        new_item = {
            "unique_id": temp_uid,
            "code": code,
            "description": vals[2],
            "type": vals[3],
            "scenario": vals[4],
            "kit_number": vals[5],
            "module_number": vals[6],
            "current_stock": "0",  # ✅ Always 0 for new rows
            "exp_date": "",  # ✅ Blank expiry - new batch
            "std_qty": parsed.get("std_qty", 0),
        }

        # Insert below the selected row
        idx = self.tree.index(row_id)
        new_iid = self.tree.insert(
            "",
            idx + 1,
            values=(
                new_item["unique_id"],
                new_item["code"],
                new_item["description"],
                new_item["type"],
                new_item["scenario"],
                new_item["kit_number"],
                new_item["module_number"],
                new_item["current_stock"],  # 0
                new_item["exp_date"],  # blank
                "",  # physical_qty - user will enter
                "",  # ✅ updated_exp_date - BLANK for new row (user will fill)
                "",  # discrepancy - will be calculated
                "",  # remarks
                new_item["std_qty"],
            ),
        )

        # Apply row styling
        t = (vals[3] or "").upper()
        if t == "KIT":
            self.tree.item(new_iid, tags=("kit_row",))
        elif t == "MODULE":
            self.tree.item(new_iid, tags=("module_row",))

        # Initialize base physical input to 0
        self.base_physical_inputs[new_iid] = 0

        # ✅ Save state WITH parent row metadata (kit_code, module_code, treecode, is_in_box)
        self.user_row_states[temp_uid] = {
            "unique_id": temp_uid,
            "code": code,
            "description": vals[2],
            "type": vals[3],
            "scenario": vals[4],
            "kit_number": vals[5],
            "module_number": vals[6],
            "current_stock": "0",
            "exp_date": "",
            "physical_qty": "",
            "updated_exp_date": "",
            "discrepancy": "",
            "remarks": "",
            "std_qty": new_item["std_qty"],
            "base_physical": 0,
            "is_custom": True,
            # ✅ CRITICAL: Store parent row metadata
            "parent_unique_id": parent_unique_id,
            "scenario_id": parsed.get("scenario_id"),
            "kit_code": parsed.get("kit_code"),
            "module_code": parsed.get("module_code"),
            "treecode": parsed.get("treecode"),
            "is_in_box": is_in_box,  # ✅ Flag to preserve format (6 or 8 segments)
        }

        # Show success message with format info
        if is_in_box:
            self.status_var.set(
                lang.t(
                    "stock_inv.row_added_inbox",
                    "New row added for {code} (In-Box: {kit}/{module}). Fill in Physical Qty and Updated Expiry.",
                ).format(code=code, kit=vals[5], module=vals[6])
            )
        else:
            self.status_var.set(
                lang.t(
                    "stock_inv.row_added",
                    "New row added for {code} (On-Shelf). Fill in Physical Qty and Updated Expiry.",
                ).format(code=code)
            )

        # Highlight if expiry is required
        self._highlight_missing_required_expiry()

    def _clear_physical(self, row_id):
        vals = self.tree.item(row_id, "values")
        if not vals:
            return
        uid = vals[0]
        self.base_physical_inputs[row_id] = 0
        self.tree.set(row_id, "physical_qty", "")
        self.tree.set(row_id, "discrepancy", "")
        st = self.user_row_states.get(uid)
        if st:
            st["physical_qty"] = ""
            st["discrepancy"] = ""
            st["base_physical"] = 0
        t = (vals[3] or "").upper()
        if t == "KIT":
            self.tree.item(row_id, tags=("kit_row",))
        elif t == "MODULE":
            self.tree.item(row_id, tags=("module_row",))
        if self.mgmt_mode_var.get() == lang.t("stock_inv.management_in_box", "In-Box"):
            self.recompute_all_physical_quantities()
        self._highlight_missing_required_expiry()

        # ---------- Inline edit ----------

    def _begin_inline_edit(self, row_id, col_key):
        """
        Begin inline editing for a specific cell.
        ✅ BLOCKS editing of updated_exp_date for existing rows (current_stock > 0)
        """
        vals = self.tree.item(row_id, "values")
        if not vals:
            return

        code = vals[1]
        description = vals[2]
        current_stock = vals[8]  # Index 8 for current_stock

        # Parse current stock
        current_stock_int = 0
        if current_stock and str(current_stock).isdigit():
            current_stock_int = int(current_stock)

        # ✅ BLOCK editing updated_exp_date for existing rows
        if col_key == "updated_exp_date" and current_stock_int > 0:
            self._show_expiry_change_instructions(code, description)
            return

        # Get column index
        col_index = {
            "physical_qty": "#10",
            "updated_exp_date": "#11",
            "remarks": "#13",
        }.get(col_key)

        if not col_index:
            return

        # Get current value
        current_value = vals[int(col_index.replace("#", ""))]

        # Get column position and width
        x = self.tree.bbox(row_id, col_index)[0]
        y = self.tree.bbox(row_id, col_index)[1]
        width = self.tree.column(col_index, "width")
        height = self.tree.bbox(row_id, col_index)[3]

        # Create entry widget
        entry = tk.Entry(self.tree, width=width)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, current_value)
        entry.select_range(0, tk.END)
        entry.focus()

        def save_edit(evt=None):
            new_value = entry.get().strip()
            entry.destroy()

            # Update tree
            self.tree.set(row_id, col_key, new_value)

            # Update state
            if col_key == "physical_qty":
                # Update base physical input
                if new_value.isdigit():
                    self.base_physical_inputs[row_id] = int(new_value)
                else:
                    self.base_physical_inputs[row_id] = 0
                # Recompute all quantities
                self.recompute_all_physical_quantities()
            else:
                # For remarks or updated_exp_date, just update state
                self._update_state_from_row(row_id)

        # Bind events
        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def start_edit(self, event):
        """Handle double-click editing with validation for updated expiry."""
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return

        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(self.display_cols):
            return

        col_key = self.display_cols[col_idx]

        # Only allow editing specific columns
        if col_key not in ("physical_qty", "updated_exp_date", "remarks"):
            return

        # ✅ BLOCK editing updated_exp_date for existing rows
        if col_key == "updated_exp_date":
            vals = self.tree.item(row_id, "values")
            current_stock = int(vals[7]) if str(vals[7]).isdigit() else 0
            current_exp = vals[8]  # Current expiry date

            # If row has existing stock and expiry, block editing
            if current_stock > 0 and current_exp:
                self._show_expiry_change_instructions(vals[1], vals[2])
                return  # Don't allow editing

        # Allow editing
        self._begin_inline_edit(row_id, col_key)

    # ---------- Filter / mode events ----------
    def on_scenario_selected(self, event=None):
        self.refresh_kit_dropdown()
        self.refresh_module_dropdown()
        self.rebuild_tree_preserving_state()

    def on_management_mode_change(self, event=None):
        mgmt_mode = self.mgmt_mode_var.get()
        on_shelf = mgmt_mode == lang.t("stock_inv.management_on_shelf", "On-Shelf")
        if on_shelf:
            self.kit_number_var.set(lang.t("stock_inv.all_kits", "All Kits"))
            self.module_number_var.set(lang.t("stock_inv.all_modules", "All Modules"))
            self.kit_number_cb.config(state="disabled")
            self.module_number_cb.config(state="disabled")
        else:
            self.kit_number_cb.config(state="readonly")
            self.module_number_cb.config(state="readonly")
        self.rebuild_tree_preserving_state()

    def on_filter_change(self, event=None):
        """Handle kit/module number filter changes with state preservation."""
        widget = event.widget if event else None
        mgmt_mode = self.mgmt_mode_var.get()
        on_shelf = mgmt_mode == lang.t("stock_inv.management_on_shelf", "On-Shelf")

        # ✅ ALWAYS capture current state before any changes
        self.capture_current_rows()

        # Refresh module dropdown if kit number changed
        if not on_shelf and widget == getattr(self, "kit_number_cb", None):
            self.refresh_module_dropdown()

        # ✅ ALWAYS rebuild tree when filters change (not just in Complete Inventory)
        self.rebuild_tree_preserving_state()

    def on_inv_type_selected(self, event=None):
        """Handle inventory type changes."""
        inv_type = self.inv_type_var.get()

        # ✅ ALWAYS clear data when inventory type changes (both directions)
        self.user_row_states.clear()
        self.base_physical_inputs.clear()

        if inv_type == lang.t("stock_inv.complete_inventory", "Complete Inventory"):
            # Complete Inventory mode: Load all items
            self.search_frame.pack_forget()
            self.rebuild_tree_preserving_state()
            self.batch_info_var.set(
                lang.t(
                    "stock_inv.complete_info",
                    "Complete inventory: ALL items loaded (including those with 0 stock). Items with current stock are highlighted.",
                )
            )
        else:
            # Partial Inventory mode: Clear tree and show search
            self.tree.delete(*self.tree.get_children())
            self.search_frame.pack(pady=5, fill="x")
            self.status_var.set(
                lang.t(
                    "stock_inv.partial_info",
                    "Search for items from standard quantities. ✓=Has stock, ○=Missing stock",
                )
            )
            self.batch_info_var.set(
                lang.t(
                    "stock_inv.partial_batch_info",
                    "Enter 2+ characters to search. Items are loaded from compositions and kit_items tables.",
                )
            )

    # ---------- Search interactions ----------

    def select_first_search_result(self, event=None):
        """Select first search result when Enter is pressed."""
        if self.search_listbox.size() > 0:
            self.search_listbox.selection_set(0)
            self.on_search_select()

    def clear_search(self):
        """Clear search field and reload tree in Complete Inventory mode."""
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)

        # Reload tree if in Complete Inventory mode (preserving user data)
        inv_type = self.inv_type_var.get()
        if inv_type == lang.t("stock_inv.complete_inventory", "Complete Inventory"):
            self.rebuild_tree_preserving_state()

        self.status_var.set(lang.t("stock_inv.search_cleared", "Search cleared"))

    def save_all(self):
        """
        Save inventory adjustments following the 4-scenario principle.

        Scenarios:
        1. Existing row, discrepancy > 0 (Surplus) → qty_in += discrepancy
        2. Existing row, discrepancy < 0 (Shortage) → qty_out += abs(discrepancy)
        3. New row (current_stock=0), physical > 0 → Create new batch with updated_exp
        4. Existing row, discrepancy = 0 (Match) → No database change

        NO expiry changes allowed for existing rows (blocked in UI).
        """

        # ✅ Check for adopted expiries before saving
        if not self._show_adopted_expiry_warning():
            # User chose to review - cancel save
            self.status_var.set(
                lang.t(
                    "stock_inv.save_cancelled",
                    "Save cancelled - Please review expiry dates",
                )
            )
            return

        # Role validation
        if self.role.lower() not in ["admin", "manager"]:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "stock_inv.restricted",
                    "Only admin or manager roles can save changes.",
                ),
                "error",
            )
            return

        # Check if there are rows to save
        rows = self.tree.get_children()
        if not rows:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("stock_inv.no_rows", "No rows to save."),
                "error",
            )
            return

        # Validate: NEW rows (current_stock=0) with physical qty MUST have updated_exp
        blocking = []
        for rid in rows:
            vals = self.tree.item(rid, "values")
            code = vals[1]
            current_stock = int(vals[7]) if str(vals[7]).isdigit() else 0
            phys_str = vals[9]
            updated_exp = vals[10]
            phys_int = int(phys_str) if phys_str.isdigit() else 0

            # Only validate NEW rows with physical quantity
            if current_stock == 0 and phys_int > 0 and check_expiry_required(code):
                if not self._is_valid_future_updated_exp(updated_exp):
                    blocking.append(code)

        if blocking:
            self._highlight_missing_required_expiry()
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "stock_inv.invalid_expiry_new_rows",
                    "Valid future expiry dates required for new items: {items}",
                ).format(items=", ".join(sorted(set(blocking)))),
                "error",
            )
            return

        # Generate document number
        doc_number = self.generate_document_number()

        # Connect to database
        conn = connect_db()
        if conn is None:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("stock_inv.db_error", "Database connection failed"),
                "error",
            )
            return
        cur = conn.cursor()

        errors = []
        max_retries = 4

        def attempt(sql, params):
            """Execute SQL with retry logic for database locks."""
            for attempt_i in range(max_retries):
                try:
                    cur.execute(sql, params)
                    return True
                except sqlite3.OperationalError as e:
                    if "locked" in str(e).lower() and attempt_i < max_retries - 1:
                        time.sleep(0.35 * (attempt_i + 1))
                        continue
                    errors.append(str(e))
                    return False
                except Exception as e:
                    errors.append(str(e))
                    return False

        # Inventory type abbreviation
        inv_type_label = self.inv_type_var.get()
        inv_abbr = (
            "Complete INV"
            if inv_type_label
            == lang.t("stock_inv.complete_inventory", "Complete Inventory")
            else "Partial INV"
        )

        def log_transaction(
            uid,
            exp,
            qty_in=None,
            qty_out=None,
            discrepancy=None,
            remarks=None,
            code=None,
            scen=None,
            kit=None,
            mod=None,
        ):
            """Log transaction to stock_transactions table."""
            if not qty_in and not qty_out:
                return

            attempt(
                """
                INSERT INTO stock_transactions
                (Date, Time, document_number, unique_id, code, Description, Expiry_date, Batch_Number,
                Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
                Third_Party, End_User, Discrepancy, Remarks, Movement_Type)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    datetime.today().strftime("%Y-%m-%d"),
                    datetime.now().strftime("%H:%M:%S"),
                    doc_number,
                    uid,
                    code,
                    get_active_designation(code),
                    exp or None,
                    None,
                    scen,
                    kit,
                    mod,
                    qty_in if qty_in and qty_in > 0 else None,
                    inv_abbr if qty_in and qty_in > 0 else None,
                    qty_out if qty_out and qty_out > 0 else None,
                    inv_abbr if qty_out and qty_out > 0 else None,
                    None,
                    None,
                    discrepancy if discrepancy else None,
                    remarks if remarks else None,
                    "stock_inv",
                ),
            )

        # Process each row

        processed_count = 0
        for rid in rows:
            vals = self.tree.item(rid, "values")
            unique_id = vals[0]
            code = vals[1]
            current_stock = int(vals[7]) if str(vals[7]).isdigit() else 0
            current_exp = vals[8]
            physical_str = vals[9]
            updated_exp = vals[10]
            remarks_val = vals[12]

            # Skip if no physical quantity entered
            if not physical_str or not physical_str.isdigit():
                continue

            physical = int(physical_str)
            discrepancy = physical - current_stock

            # ✅ FIX: Handle temp vs real unique_ids
            if unique_id.startswith("temp::"):
                # For temp rows, get ALL data from stored state
                state = self.user_row_states.get(unique_id, {})

                # Get scenario from row values (vals[4])
                scenario_name = vals[4]

                # ✅ Get scenario_id from stored state (from parent row)
                scenario_id = state.get("scenario_id")

                if not scenario_id:
                    # Fallback: Look up scenario_id from scenario_name
                    for sid, sname in self.scenario_map.items():
                        if sname == scenario_name:
                            scenario_id = sid
                            break

                if not scenario_id:
                    errors.append(
                        f"Could not find scenario_id for scenario '{scenario_name}'"
                    )
                    continue

                # ✅ Get kit/module codes from stored state (from parent row)
                kit_code = state.get("kit_code")
                module_code = state.get("module_code")
                item_code = code  # Use actual code from vals[1]

                # Get kit/module numbers from row
                kit_number = vals[5] if vals[5] != "-----" else None
                module_number = vals[6] if vals[6] != "-----" else None

                # ✅ Get treecode from stored state (from parent row)
                treecode_stored = state.get("treecode")

                std_qty_val = int(vals[13]) if str(vals[13]).isdigit() else 0

            else:
                # Regular row - parse unique_id normally
                parsed = parse_inventory_unique_id(unique_id)
                scenario_id = parsed["scenario_id"]
                scenario_name = parsed["scenario_name"]
                kit_code = parsed["kit_code"]
                module_code = parsed["module_code"]
                item_code = parsed["item_code"]
                kit_number = parsed["kit_number"]
                module_number = parsed["module_number"]
                treecode_stored = parsed.get("treecode")
                std_qty_val = parsed["std_qty"] or (
                    int(vals[13]) if str(vals[13]).isdigit() else 0
                )

            # ✅ FIX: For logging purposes, use kit_number and module_number (NOT codes)
            kit_for_log = kit_number if kit_number else None
            mod_for_log = module_number if module_number else None

            # ============================================================
            # SCENARIO LOGIC - 4 Scenarios Based on Agreed Principles
            # ============================================================

            # SCENARIO 1 & 2: Existing row (current_stock > 0)
            if current_stock > 0:
                # Check if unique_id exists in stock_data
                cur.execute(
                    "SELECT qty_in, qty_out FROM stock_data WHERE unique_id = ?",
                    (unique_id,),
                )
                existing = cur.fetchone()

                if not existing:
                    errors.append(
                        f"Row {code} has current_stock > 0 but no stock_data entry found!"
                    )
                    continue

                old_qty_in = existing[0] or 0
                old_qty_out = existing[1] or 0

                # SCENARIO 1: Surplus (discrepancy > 0)
                if discrepancy > 0:
                    new_qty_in = old_qty_in + discrepancy
                    if not attempt(
                        "UPDATE stock_data SET qty_in = ? WHERE unique_id = ?",
                        (new_qty_in, unique_id),
                    ):
                        errors.append(f"Failed to update qty_in for {code}")
                        continue

                    log_transaction(
                        unique_id,
                        current_exp,
                        qty_in=discrepancy,
                        discrepancy=discrepancy,
                        code=code,
                        scen=scenario_id,
                        kit=kit_for_log,  # ✅ Now uses kit_number
                        mod=mod_for_log,  # ✅ Now uses module_number
                        remarks=remarks_val or f"Surplus: +{discrepancy} units",
                    )
                    processed_count += 1

                # SCENARIO 2: Shortage (discrepancy < 0)
                elif discrepancy < 0:
                    new_qty_out = old_qty_out + abs(discrepancy)
                    if not attempt(
                        "UPDATE stock_data SET qty_out = ? WHERE unique_id = ?",
                        (new_qty_out, unique_id),
                    ):
                        errors.append(f"Failed to update qty_out for {code}")
                        continue

                    log_transaction(
                        unique_id,
                        current_exp,
                        qty_out=abs(discrepancy),
                        discrepancy=discrepancy,
                        code=code,
                        scen=scenario_id,
                        kit=kit_for_log,  # ✅ Now uses kit_number
                        mod=mod_for_log,  # ✅ Now uses module_number
                        remarks=remarks_val or f"Shortage: {discrepancy} units",
                    )
                    processed_count += 1

                # SCENARIO 4: Match (discrepancy == 0) - No action needed
                else:
                    # Just log if remarks exist
                    if remarks_val:
                        log_transaction(
                            unique_id,
                            current_exp,
                            discrepancy=0,
                            code=code,
                            scen=scenario_id,
                            kit=kit_for_log,  # ✅ Now uses kit_number
                            mod=mod_for_log,  # ✅ Now uses module_number
                            remarks=remarks_val,
                        )
                    processed_count += 1

            # SCENARIO 3: New row (current_stock = 0) - Create new batch
            elif current_stock == 0 and physical > 0:
                # Must have updated_exp for new rows
                if not updated_exp:
                    errors.append(f"New row for {code} missing Updated Expiry Date!")
                    continue

                # Parse expiry date
                parsed_exp = parse_expiry(updated_exp)
                exp_iso = parsed_exp.strftime("%Y-%m-%d") if parsed_exp else updated_exp

                # Determine if in-box format
                had_box = kit_number or module_number or len(unique_id.split("/")) >= 8

                # ✅ Use stored treecode, or fetch if not available
                treecode = treecode_stored
                if had_box and not treecode:
                    treecode = get_treecode(
                        scenario_id, kit_code, module_code, item_code
                    )

                # ✅ Construct proper unique_id with ALL original metadata
                new_uid = construct_unique_id(
                    scenario_id=scenario_id,
                    kit_code=kit_code,  # ✅ From parent row
                    module_code=module_code,  # ✅ From parent row
                    item_code=item_code,
                    std_qty=std_qty_val,
                    exp_date=exp_iso,  # ✅ ONLY this changes
                    kit_number=kit_number,
                    module_number=module_number,
                    force_box_format=had_box,
                    treecode=treecode,  # ✅ From parent row
                )

                # Check if batch already exists
                cur.execute(
                    "SELECT qty_in, qty_out FROM stock_data WHERE unique_id = ?",
                    (new_uid,),
                )
                existing_batch = cur.fetchone()

                if existing_batch:
                    # Batch exists - add to it
                    new_qty_in = (existing_batch[0] or 0) + physical
                    if not attempt(
                        "UPDATE stock_data SET qty_in = ? WHERE unique_id = ?",
                        (new_qty_in, new_uid),
                    ):
                        errors.append(f"Failed to update existing batch for {code}")
                        continue
                else:
                    # ✅ Create new batch with ALL original metadata including treecode
                    if not attempt(
                        """
                        INSERT INTO stock_data 
                        (unique_id, scenario, kit_number, module_number, kit, module, item, 
                        std_qty, qty_in, qty_out, exp_date, discrepancy, treecode)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            new_uid,
                            scenario_name,
                            kit_number,
                            module_number,
                            kit_code,  # ✅ From parent
                            module_code,  # ✅ From parent
                            item_code,
                            std_qty_val,
                            physical,
                            0,
                            exp_iso,
                            0,
                            treecode,  # ✅ From parent
                        ),
                    ):
                        errors.append(f"Failed to create new batch for {code}")
                        continue

                # Log transaction
                log_transaction(
                    new_uid,
                    exp_iso,
                    qty_in=physical,
                    discrepancy=physical,
                    code=item_code,
                    scen=scenario_id,
                    kit=kit_for_log,  # ✅ Now uses kit_number
                    mod=mod_for_log,  # ✅ Now uses module_number
                    remarks=remarks_val
                    or f"New batch: {physical} units, exp: {exp_iso}",
                )
                processed_count += 1

        # Commit all changes
        try:
            conn.commit()

            # ✅ AUTO-EXPORT TO EXCEL
            rows_to_export = [
                {
                    "code": vals[1],
                    "description": vals[2],
                    "type": vals[3],
                    "scenario": vals[4],
                    "kit_number": vals[5],
                    "module_number": vals[6],
                    "current_stock": vals[7],
                    "exp_date": vals[8],
                    "physical_qty": vals[9],
                    "updated_exp_date": vals[10],
                    "discrepancy": vals[11],
                    "remarks": vals[12],
                }
                for item in self.tree.get_children()
                if (vals := self.tree.item(item)["values"])
            ]

            # Export with document number
            self.export_to_excel(
                rows_to_export=rows_to_export, document_number=doc_number
            )

            # Success message
            custom_popup(
                self,
                lang.t("dialog_titles.success", "Success"),
                lang.t(
                    "stock_inv.save_success",
                    "Inventory saved successfully!\nDocument: {doc}\nProcessed: {count} rows",
                ).format(doc=doc_number, count=processed_count),
                "info",
            )

            # Clear form after successful save
            self.clear_form()

        except Exception as e:
            conn.rollback()
            errors.append(f"Commit failed: {str(e)}")
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("stock_inv.save_failed", "Failed to save: {error}").format(
                    error=str(e)
                ),
                "error",
            )
        finally:
            cur.close()
            conn.close()

        # Show any errors that occurred
        if errors:
            error_msg = "\n".join(errors[:5])  # Show first 5 errors
            custom_popup(
                self,
                lang.t("dialog_titles.warning", "Warning"),
                lang.t(
                    "stock_inv.partial_errors", "Completed with errors:\n{errors}"
                ).format(errors=error_msg),
                "warning",
            )

    # ---------- Clear form ----------
    def clear_form(self):
        self.tree.delete(*self.tree.get_children())
        self.code_entry.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
        self.inv_type_var.set(
            lang.t("stock_inv.complete_inventory", "Complete Inventory")
        )
        self.scenario_var.set(lang.t("stock_inv.all_scenarios", "All Scenarios"))
        self.mgmt_mode_var.set(lang.t("stock_inv.management_all", "All"))
        self.kit_number_var.set(lang.t("stock_inv.all_kits", "All Kits"))
        self.module_number_var.set(lang.t("stock_inv.all_modules", "All Modules"))
        self.refresh_kit_dropdown()
        self.refresh_module_dropdown()
        self.search_frame.pack_forget()
        self.batch_info_var.set("")
        self.status_var.set("")
        self.user_row_states.clear()
        self.base_physical_inputs.clear()
        self.on_inv_type_selected()


# ---------- Standalone run ----------
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Inventory Adjustment")
    app = type("App", (), {})()
    app.role = "admin"
    StockInventory(root, app, role="admin")
    root.geometry("1550x900")
    root.mainloop()
