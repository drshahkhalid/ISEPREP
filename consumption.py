"""
consumption.py  v2.1

Combined Reception (IN) + Consumption (OUT) Monthly Analysis

v2.1 Changes:
  * Renamed 'Dataset' label to 'Stock Action'.
  * Added horizontal scrollbar to results table (Treeview now has both vertical & horizontal).
  * Added global key bindings:
       - Enter (Return): triggers refresh (anywhere in the module window).
       - Escape (Esc): clears all filters (same as Clear button).
  * Added key bindings on main editable filter widgets so pressing Enter immediately refreshes.
  * Added optional horizontal scrolling for graph popup canvas (if resized smaller than content).
  * Ensured movement type dropdowns update after scenario changes; bindings preserved.
  * Minor UI spacing/layout refinements for clarity.

See earlier version notes (v2.0) for full feature list (Reception + Consumption integration, two-line graph, etc.).
"""

import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
from datetime import date, datetime, timedelta
from calendar import monthrange
import re
from collections import defaultdict, OrderedDict
import os
import time

# Optional calendar
try:
    from tkcalendar import DateEntry
    TKCAL_AVAILABLE = True
except Exception:
    TKCAL_AVAILABLE = False

# Optional Pillow for image export
PIL_AVAILABLE = False
try:
    from PIL import Image, ImageGrab
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from db import connect_db
from language_manager import lang
from popup_utils import custom_popup
from manage_items import get_item_description, detect_type

# ---------------- Configuration (Image Export) ----------------
USE_SCREENSHOT_CAPTURE = True
SCREENSHOT_RETRIES = 4
SCREENSHOT_DELAY_MS = 140
SCREENSHOT_TOLERANCE_RATIO = 0.80

# ---------------- Theme Constants ----------------
BG_MAIN        = "#F0F4F8"
BG_PANEL       = "#FFFFFF"
COLOR_PRIMARY  = "#2C3E50"
COLOR_BORDER   = "#D0D7DE"
ROW_ALT_COLOR  = "#F7FAFC"
ROW_NORM_COLOR = "#FFFFFF"
BTN_EXPORT     = "#2980B9"
BTN_REFRESH    = "#2563EB"
BTN_CLEAR      = "#7F8C8D"
BTN_TOGGLE     = "#8E44AD"
BTN_GRAPH      = "#0B7285"

KIT_FILL_COLOR    = "228B22"
MODULE_FILL_COLOR = "ADD8E6"

LINE_COLOR_IN  = "#2E8B57"  # green
LINE_COLOR_OUT = "#C0392B"  # red

OUT_TYPES_LIST = [
    "Issue to End User",
    "Expired Items",
    "Damaged Items",
    "Cold Chain Break",
    "Batch Recall",
    "Theft",
    "Other Losses",
    "Out Donation",
    "Loan",
    "Return of Borrowing",
    "Quarantine",
    "All"
]
IN_TYPES_LIST = [
    "In MSF",
    "In Local Purchase",
    "In from Quarantine",
    "In Donation",
    "Return from End User",
    "In Supply Non-MSF",
    "In Borrowing",
    "In Return of Loan",
    "In Correction of Previous Transaction",
    "All"
]

OUT_MOVEMENT_TEMPLATES = [
    "Dispatch Kit",
    "Issue standalone item/s from {scenario}",
    "Issue module from {scenario}",
    "Issue module from a kit",
    "Issue items from a kit",
    "Issue items from a module",
    "ALL"
]
IN_MOVEMENT_TEMPLATES = [
    "Receive Kit",
    "Add standalone item/s in {scenario}",
    "Add module to {scenario}",
    "Add module to a kit",
    "Add items to a kit",
    "Add items to a module",
    "ALL"
]

DATE_REGEXES = [
    ("%Y-%m-%d", re.compile(r"^\d{4}-\d{2}-\d{2}$")),
    ("%d/%m/%Y", re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")),
    ("%d-%m-%Y", re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$")),
    ("%d %b %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$")),
    ("%d %B %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]+\s+\d{4}$")),
    ("%Y/%m/%d", re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")),
]
MONTH_ONLY_REGEXES = [
    ("%Y-%m", re.compile(r"^\d{4}-\d{1,2}$")),
    ("%m/%Y", re.compile(r"^\d{1,2}/\d{4}$")),
    ("%b-%Y", re.compile(r"^[A-Za-z]{3}-\d{4}$")),
    ("%B-%Y", re.compile(r"^[A-Za-z]+-\d{4}$")),
]

def parse_user_date(text: str, role: str):
    if not text:
        return None
    raw = text.strip()
    for fmt, rx in DATE_REGEXES:
        if rx.match(raw):
            try: return datetime.strptime(raw, fmt).date()
            except Exception: pass
    for fmt, rx in MONTH_ONLY_REGEXES:
        if rx.match(raw):
            try:
                dt = datetime.strptime(raw, fmt)
                y, m = dt.year, dt.month
                return date(y, m, 1) if role == "from" else date(y, m, monthrange(y, m)[1])
            except Exception:
                pass
    if re.match(r"^\d{4}$", raw):
        y = int(raw)
        return date(y,1,1) if role=="from" else date(y,12,31)
    return None

def month_iter(start: date, end: date):
    y, m = start.year, start.month
    while (y < end.year) or (y == end.year and m <= end.month):
        yield y, m
        if m == 12: y += 1; m = 1
        else: m += 1

def ym_label(y, m):
    return date(y, m, 1).strftime("%b-%Y")

def fetch_project_details():
    conn = connect_db()
    if conn is None:
        return ("Unknown Project","Unknown Code")
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        if not row: return ("Unknown Project","Unknown Code")
        return (row[0] or "Unknown Project", row[1] or "Unknown Code")
    except sqlite3.Error:
        return ("Unknown Project","Unknown Code")
    finally:
        cur.close(); conn.close()

class CombinedCalculator:
    def __init__(self, **kwargs):
        # Accept all kwargs for flexibility
        for k,v in kwargs.items():
            setattr(self, k, v)
        if self.date_from and self.date_to and self.date_from > self.date_to:
            self.date_from, self.date_to = self.date_to, self.date_from

    def _load_mgmt_map(self, conn):
        mapping = {}
        need = (self.management_mode.lower() != "all" or
                self.kit.lower() != "all" or
                self.module.lower() != "all")
        if not need: return mapping
        cur = conn.cursor()
        try:
            cur.execute("PRAGMA table_info(stock_data)")
            cols = [r[1].lower() for r in cur.fetchall()]
            if "unique_id" not in cols: return mapping
            fields = ["unique_id"]
            if "management_mode" in cols: fields.append("management_mode")
            if "kit_number" in cols: fields.append("kit_number")
            if "module_number" in cols: fields.append("module_number")
            cur.execute(f"SELECT {', '.join(fields)} FROM stock_data")
            for row in cur.fetchall():
                mapping[row[0]] = dict(zip(fields, row))
        except sqlite3.Error:
            pass
        finally: cur.close()
        return mapping

    def compute(self):
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed")
        try:
            cur = conn.cursor()
            mgmt_map = self._load_mgmt_map(conn)
            where = []
            params = []
            if self.date_from:
                where.append("Date >= ?")
                params.append(self.date_from.strftime("%Y-%m-%d"))
            if self.date_to:
                where.append("Date <= ?")
                params.append(self.date_to.strftime("%Y-%m-%d"))
            if self.document_number:
                where.append("document_number LIKE ?")
                params.append(f"%{self.document_number}%")
            if self.scenario.lower() != "all":
                where.append("Scenario = ?")
                params.append(self.scenario)

            sql = f"""
                SELECT Date, code, Scenario, Kit, Module,
                       Qty_IN, IN_Type, Qty_Out, Out_Type,
                       Movement_Type, document_number, unique_id
                FROM stock_transactions
                WHERE {' AND '.join(where) if where else '1=1'}
            """
            cur.execute(sql, params)
            rows = cur.fetchall()

            if not self.date_from or not self.date_to:
                md, Md = None, None
                for r in rows:
                    try:
                        d = datetime.strptime(r[0], "%Y-%m-%d").date()
                        if not md or d < md: md = d
                        if not Md or d > Md: Md = d
                    except Exception:
                        continue
                if not self.date_from: self.date_from = md or date.today()
                if not self.date_to: self.date_to = Md or date.today()

            months_seq = list(month_iter(self.date_from, self.date_to))

            per_code = {}
            meta_map = {}
            for (dt_text, code, scen, kit_num, module_num,
                 qty_in, in_type, qty_out, out_type, movement_type,
                 doc_no, unique_id) in rows:
                if not code: continue

                # Management filtering
                if self.management_mode.lower() != "all":
                    mm = (mgmt_map.get(unique_id, {}).get("management_mode") or "").lower()
                    if mm != self.management_mode.lower():
                        continue
                if self.kit.lower() != "all":
                    kit_sd = (mgmt_map.get(unique_id, {}).get("kit_number") or "").lower()
                    if kit_sd:
                        if kit_sd != self.kit.lower(): continue
                    else:
                        if not kit_num or kit_num.lower() != self.kit.lower(): continue
                if self.module.lower() != "all":
                    mod_sd = (mgmt_map.get(unique_id, {}).get("module_number") or "").lower()
                    if mod_sd:
                        if mod_sd != self.module.lower(): continue
                    else:
                        if not module_num or module_num.lower() != self.module.lower(): continue

                dtype = detect_type(code, "")
                if self.type_filter.lower() != "all" and dtype.lower() != self.type_filter.lower():
                    continue
                if self.item_search:
                    if dtype.lower() != "item":
                        continue
                    desc_search = get_item_description(code)
                    if (self.item_search.lower() not in code.lower() and
                        self.item_search.lower() not in desc_search.lower()):
                        continue

                # Filter directions
                dataset_mode = self.dataset_mode
                # IN filters
                if dataset_mode in ("All","Reception"):
                    if self.in_type.lower() != "all":
                        if (in_type or "").strip() != self.in_type:
                            continue
                    if self.in_movement.upper() != "ALL":
                        if (movement_type or "").strip() != self.in_movement:
                            continue
                # OUT filters
                if dataset_mode in ("All","Consumption"):
                    if self.out_type.lower() != "all":
                        if (out_type or "").strip() != self.out_type:
                            continue
                    if self.out_movement.upper() != "ALL":
                        if (movement_type or "").strip() != self.out_movement:
                            continue

                try:
                    d_obj = datetime.strptime(dt_text, "%Y-%m-%d").date()
                except Exception:
                    continue
                ym = (d_obj.year, d_obj.month)
                if ym not in months_seq:
                    continue

                entry = per_code.setdefault(code, {
                    "per_month_in": defaultdict(int),
                    "per_month_out": defaultdict(int),
                    "total_in": 0,
                    "total_out": 0
                })
                if qty_in and qty_in > 0:
                    entry["per_month_in"][ym] += qty_in
                    entry["total_in"] += qty_in
                if qty_out and qty_out > 0:
                    entry["per_month_out"][ym] += qty_out
                    entry["total_out"] += qty_out

                if code not in meta_map:
                    meta_map[code] = {
                        "scenario": scen or "",
                        "kit_number": kit_num or "",
                        "module_number": module_num or "",
                        "movement_type": movement_type or "",
                        "document_number": doc_no or "",
                        "in_type": in_type or "",
                        "out_type": out_type or "",
                        "type": dtype
                    }

            result = []
            for code, agg in per_code.items():
                desc = get_item_description(code)
                meta = meta_map.get(code, {})
                result.append({
                    "code": code,
                    "description": desc,
                    "scenario": meta.get("scenario",""),
                    "kit_number": meta.get("kit_number",""),
                    "module_number": meta.get("module_number",""),
                    "movement_type": meta.get("movement_type",""),
                    "document_number": meta.get("document_number",""),
                    "in_type": meta.get("in_type",""),
                    "out_type": meta.get("out_type",""),
                    "type": meta.get("type",""),
                    "per_month_in": agg["per_month_in"],
                    "per_month_out": agg["per_month_out"],
                    "total_in": agg["total_in"],
                    "total_out": agg["total_out"]
                })

            result.sort(key=lambda r: (r["type"], r["code"]))
            return result, months_seq
        finally:
            try: conn.close()
            except Exception: pass

class Consumption(tk.Frame):
    def __init__(self, parent, app, *args, **kwargs):
        super().__init__(parent, bg=BG_MAIN, *args, **kwargs)
        self.app = app
        self.rows = []
        self.months_seq = []
        self.simple_mode = False
        self.graph_visible = False
        self.graph_window = None
        self.chart_canvas = None
        self.chart_pad = 40

        self.project_name, self.project_code = fetch_project_details()
        self.dataset_mode = tk.StringVar(value="All")

        self._build_ui()
        self.populate_dropdowns()
        self._set_default_dates()
        self._setup_global_keybindings()
        self.refresh()

    def t(self, key, fallback=None, **kwargs):
        return lang.t(f"consumption.{key}", fallback=fallback if fallback else key, **kwargs)

    # ---------- UI ----------
    def _build_ui(self):
        header = tk.Frame(self, bg=BG_MAIN)
        header.pack(fill="x", padx=12, pady=(12,4))
        tk.Label(header, text=self.t("title","Stock Actions (Reception & Consumption)"),
                 font=("Helvetica",20,"bold"),
                 bg=BG_MAIN, fg=COLOR_PRIMARY, anchor="w").pack(side="left", fill="x", expand=True)

        self.toggle_btn = tk.Button(header, text=self.t("toggle_detailed","Detailed"),
                                    bg=BTN_TOGGLE, fg="#FFFFFF", relief="flat",
                                    padx=14, pady=6, command=self.toggle_mode)
        self.toggle_btn.pack(side="right")

        self.graph_btn = tk.Button(header, text=self.t("toggle_graph_show","Show Graph"),
                                   bg=BTN_GRAPH, fg="#FFFFFF", relief="flat",
                                   padx=14, pady=6, command=self.toggle_graph_popup)
        self.graph_btn.pack(side="right", padx=(0,6))

        filters = tk.Frame(self, bg=BG_MAIN)
        filters.pack(fill="x", padx=12, pady=(0,10))

        # Row 1 - Stock Action (Dataset)
        r1 = tk.Frame(filters, bg=BG_MAIN); r1.pack(fill="x", pady=2)
        tk.Label(r1, text=self.t("stock_action","Stock Action"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.dataset_cb = ttk.Combobox(r1, textvariable=self.dataset_mode,
                                       state="readonly", width=16,
                                       values=["All","Reception","Consumption"])
        self.dataset_cb.grid(row=0, column=1, padx=(0,12))
        self.dataset_cb.bind("<<ComboboxSelected>>", lambda e: self._update_filter_states())

        tk.Label(r1, text=self.t("management_mode","Management Mode"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.mgmt_var = tk.StringVar(value="All")
        ttk.Combobox(r1, textvariable=self.mgmt_var, state="readonly", width=14,
                     values=["All","on-shelf","in-box"]).grid(row=0, column=3, padx=(0,12))

        tk.Label(r1, text=self.t("scenario","Scenario"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.scenario_var = tk.StringVar(value="All")
        self.scenario_cb = ttk.Combobox(r1, textvariable=self.scenario_var, state="readonly", width=20)
        self.scenario_cb.grid(row=0, column=5, padx=(0,12))
        self.scenario_cb.bind("<<ComboboxSelected>>", lambda e: (self._refresh_movement_lists(), self.refresh()))

        # Row 2
        r2 = tk.Frame(filters, bg=BG_MAIN); r2.pack(fill="x", pady=2)
        tk.Label(r2, text=self.t("kit_number","Kit"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.kit_var = tk.StringVar(value="All")
        self.kit_cb = ttk.Combobox(r2, textvariable=self.kit_var, state="readonly", width=16)
        self.kit_cb.grid(row=0, column=1, padx=(0,14))

        tk.Label(r2, text=self.t("module_number","Module"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.module_var = tk.StringVar(value="All")
        self.module_cb = ttk.Combobox(r2, textvariable=self.module_var, state="readonly", width=16)
        self.module_cb.grid(row=0, column=3, padx=(0,14))

        tk.Label(r2, text=self.t("type","Type"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.type_var = tk.StringVar(value="All")
        ttk.Combobox(r2, textvariable=self.type_var, state="readonly", width=12,
                     values=["All","Kit","Module","Item"]).grid(row=0, column=5, padx=(0,14))

        tk.Label(r2, text=self.t("item_search","Item Search"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.item_search_var = tk.StringVar()
        item_entry = tk.Entry(r2, textvariable=self.item_search_var, width=18)
        item_entry.grid(row=0, column=7, padx=(0,14))
        item_entry.bind("<Return>", lambda e: self.refresh())

        # Row 3 (Dates and Document)
        r3 = tk.Frame(filters, bg=BG_MAIN); r3.pack(fill="x", pady=2)
        tk.Label(r3, text=self.t("from_date","From Date"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.from_var = tk.StringVar()
        self.from_entry = self._make_date_widget(r3, self.from_var)
        self.from_entry.grid(row=0, column=1, padx=(0,14))

        tk.Label(r3, text=self.t("to_date","To Date"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.to_var = tk.StringVar()
        self.to_entry = self._make_date_widget(r3, self.to_var)
        self.to_entry.grid(row=0, column=3, padx=(0,14))

        tk.Label(r3, text=self.t("document_number","Document Number"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.doc_var = tk.StringVar()
        doc_entry = tk.Entry(r3, textvariable=self.doc_var, width=16)
        doc_entry.grid(row=0, column=5, padx=(0,14))
        doc_entry.bind("<Return>", lambda e: self.refresh())

        # Row 4 (IN/OUT filters)
        r4 = tk.Frame(filters, bg=BG_MAIN); r4.pack(fill="x", pady=2)
        tk.Label(r4, text=self.t("in_type","IN Type"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.in_type_var = tk.StringVar(value="All")
        self.in_type_cb = ttk.Combobox(r4, textvariable=self.in_type_var, state="readonly",
                                       width=26, values=IN_TYPES_LIST)
        self.in_type_cb.grid(row=0, column=1, padx=(0,14))

        tk.Label(r4, text=self.t("in_movement_type","IN Movement Type"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.in_move_var = tk.StringVar(value="ALL")
        self.in_move_cb = ttk.Combobox(r4, textvariable=self.in_move_var, state="readonly", width=34)
        self.in_move_cb.grid(row=0, column=3, padx=(0,14))

        tk.Label(r4, text=self.t("out_type","Out Type"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.out_type_var = tk.StringVar(value="All")
        self.out_type_cb = ttk.Combobox(r4, textvariable=self.out_type_var, state="readonly",
                                        width=26, values=OUT_TYPES_LIST)
        self.out_type_cb.grid(row=0, column=5, padx=(0,14))

        tk.Label(r4, text=self.t("out_movement_type","Out Movement Type"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.out_move_var = tk.StringVar(value="ALL")
        self.out_move_cb = ttk.Combobox(r4, textvariable=self.out_move_var, state="readonly", width=34)
        self.out_move_cb.grid(row=0, column=7, padx=(0,14))

        # Buttons
        btn_row = tk.Frame(filters, bg=BG_MAIN)
        btn_row.pack(fill="x", pady=(6,4))
        tk.Button(btn_row, text=self.t("refresh","Refresh"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.refresh).pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("clear","Clear"),
                  bg=BTN_CLEAR, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.clear_filters).pack(side="left", padx=(0,6))
        tk.Button(btn_row, text=self.t("export","Export"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.export_excel).pack(side="left", padx=(0,6))

        self.status_var = tk.StringVar(value=self.t("ready","Ready"))
        tk.Label(self, textvariable=self.status_var, anchor="w",
                 bg=BG_MAIN, fg=COLOR_PRIMARY, relief="sunken")\
            .pack(fill="x", padx=12, pady=(0,8))

        # Table frame with BOTH scrollbars
        table_frame_outer = tk.Frame(self, bg=COLOR_BORDER, bd=1, relief="solid")
        table_frame_outer.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self.tree = ttk.Treeview(table_frame_outer, columns=(), show="headings", height=22)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(table_frame_outer, orient="vertical", command=self.tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(table_frame_outer, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame_outer.rowconfigure(0, weight=1)
        table_frame_outer.columnconfigure(0, weight=1)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("Treeview", background=BG_PANEL, fieldbackground=BG_PANEL,
                        foreground=COLOR_PRIMARY, rowheight=24, font=("Helvetica",10))
        style.configure("Treeview.Heading", background="#E5E8EB", foreground=COLOR_PRIMARY,
                        font=("Helvetica",11,"bold"))
        self.tree.tag_configure("norm", background=ROW_NORM_COLOR)
        self.tree.tag_configure("alt", background=ROW_ALT_COLOR)
        self.tree.tag_configure("kitrow", background="#228B22", foreground="#FFFFFF")
        self.tree.tag_configure("modrow", background="#ADD8E6")

    def _make_date_widget(self, parent, var):
        if TKCAL_AVAILABLE:
            w = DateEntry(parent, textvariable=var, width=12,
                          date_pattern="yyyy-mm-dd", showweeknumbers=False,
                          background="#2563EB", foreground="white", borderwidth=1)
            w.bind("<Return>", lambda e: self.refresh())
            return w
        else:
            w = tk.Entry(parent, textvariable=var, width=12)
            w.bind("<Return>", lambda e: self.refresh())
            return w

    # Global keybindings
    def _setup_global_keybindings(self):
        root = self.winfo_toplevel()
        root.bind("<Return>", lambda e: self.refresh())
        root.bind("<Escape>", lambda e: self.clear_filters())

    def _refresh_movement_lists(self):
        scen = self.scenario_var.get()
        def fmt(templates):
            out=[]
            for t in templates:
                if "{scenario}" in t:
                    out.append(t.format(scenario=scen if scen!="All" else "Scenario"))
                else:
                    out.append(t)
            return out
        self.in_move_cb['values'] = fmt(IN_MOVEMENT_TEMPLATES)
        if self.in_move_var.get() not in self.in_move_cb['values']:
            self.in_move_var.set("ALL")
        self.out_move_cb['values'] = fmt(OUT_MOVEMENT_TEMPLATES)
        if self.out_move_var.get() not in self.out_move_cb['values']:
            self.out_move_var.set("ALL")

    def _update_filter_states(self):
        mode = self.dataset_mode.get()
        enable_in = (mode in ("All","Reception"))
        enable_out = (mode in ("All","Consumption"))
        self._set_combo_state(self.in_type_cb, enable_in)
        self._set_combo_state(self.in_move_cb, enable_in)
        self._set_combo_state(self.out_type_cb, enable_out)
        self._set_combo_state(self.out_move_cb, enable_out)

    def _set_combo_state(self, cb, enable):
        try:
            cb.config(state="readonly" if enable else "disabled")
        except Exception:
            pass

    def _set_default_dates(self):
        today = date.today()
        one_year_before = today - timedelta(days=365)
        self.from_var.set(one_year_before.strftime("%Y-%m-%d"))
        self.to_var.set(today.strftime("%Y-%m-%d"))

    def populate_dropdowns(self):
        conn = connect_db()
        scenarios = []
        kits = []
        modules = []
        if conn:
            cur = conn.cursor()
            try:
                cur.execute("SELECT name FROM scenarios ORDER BY name")
                scenarios = [r[0] for r in cur.fetchall()]
                cur.execute("PRAGMA table_info(stock_data)")
                sd_cols = [r[1].lower() for r in cur.fetchall()]
                if "kit_number" in sd_cols:
                    cur.execute("""SELECT DISTINCT kit_number FROM stock_data
                                   WHERE kit_number IS NOT NULL AND kit_number!='None'
                                   ORDER BY kit_number""")
                    kits = [r[0] for r in cur.fetchall()]
                if "module_number" in sd_cols:
                    cur.execute("""SELECT DISTINCT module_number FROM stock_data
                                   WHERE module_number IS NOT NULL AND module_number!='None'
                                   ORDER BY module_number""")
                    modules = [r[0] for r in cur.fetchall()]
            except sqlite3.Error:
                pass
            finally:
                cur.close(); conn.close()

        self.scenario_cb['values'] = ["All"] + scenarios
        if self.scenario_var.get() not in self.scenario_cb['values']:
            self.scenario_var.set("All")

        self.kit_cb['values'] = ["All"] + kits
        if self.kit_var.get() not in self.kit_cb['values']:
            self.kit_var.set("All")

        self.module_cb['values'] = ["All"] + modules
        if self.module_var.get() not in self.module_cb['values']:
            self.module_var.set("All")

        self._refresh_movement_lists()
        self._update_filter_states()

    def refresh(self):
        from_dt = parse_user_date(self.from_var.get().strip(), "from") if self.from_var.get().strip() else None
        to_dt = parse_user_date(self.to_var.get().strip(), "to") if self.to_var.get().strip() else None
        if to_dt and to_dt > date.today():
            to_dt = date.today()
        calc = CombinedCalculator(
            dataset_mode=self.dataset_mode.get(),
            management_mode=self.mgmt_var.get(),
            scenario=self.scenario_var.get(),
            kit=self.kit_var.get(),
            module=self.module_var.get(),
            type_filter=self.type_var.get(),
            item_search=self.item_search_var.get(),
            out_type=self.out_type_var.get(),
            out_movement=self.out_move_var.get(),
            in_type=self.in_type_var.get(),
            in_movement=self.in_move_var.get(),
            document_number=self.doc_var.get(),
            date_from=from_dt,
            date_to=to_dt
        )
        try:
            self.rows, self.months_seq = calc.compute()
        except ValueError as e:
            custom_popup(self, self.t("error","Error"), str(e), "error")
            return
        self._populate_tree()
        self.status_var.set(self.t("loaded","Loaded {n} rows").format(n=len(self.rows)))
        if self.graph_visible:
            self.draw_or_update_graph()

    def clear_filters(self):
        self.dataset_mode.set("All")
        self.mgmt_var.set("All")
        self.scenario_var.set("All")
        self.kit_var.set("All")
        self.module_var.set("All")
        self.type_var.set("All")
        self.item_search_var.set("")
        self.out_type_var.set("All")
        self.out_move_var.set("ALL")
        self.in_type_var.set("All")
        self.in_move_var.set("ALL")
        self.doc_var.set("")
        self._set_default_dates()
        self._refresh_movement_lists()
        self._update_filter_states()
        self.refresh()

    # Columns
    def _current_columns(self):
        mode = self.dataset_mode.get()
        month_labels = []
        if mode == "Reception":
            month_labels = [f"{ym_label(y,m)} IN" for (y,m) in self.months_seq]
        elif mode == "Consumption":
            month_labels = [f"{ym_label(y,m)} OUT" for (y,m) in self.months_seq]
        else:
            for (y,m) in self.months_seq:
                base = ym_label(y,m)
                month_labels.append(f"{base} IN")
                month_labels.append(f"{base} OUT")
        totals = []
        if mode in ("All","Reception"):
            totals.append("total_in")
        if mode in ("All","Consumption"):
            totals.append("total_out")

        if self.simple_mode:
            return ["code","description"] + month_labels + totals
        else:
            return ["scenario","kit_number","module_number","code","description"] + month_labels + totals + ["in_type","out_type","movement_type","document_number"]

    def _populate_tree(self):
        cols = self._current_columns()
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        heading_map = {c: c.replace("_"," ").title() for c in cols}
        heading_map.update({
            "code":"Code","description":"Description","kit_number":"Kit Number",
            "module_number":"Module Number","movement_type":"Movement Type",
            "document_number":"Document Number","in_type":"IN Type","out_type":"Out Type",
            "total_in": self.t("total_in","Total IN"),
            "total_out": self.t("total_out","Total OUT"),
            "scenario":"Scenario"
        })
        for c in cols:
            width = 130
            if c == "description": width = 340
            elif c == "code": width = 160
            elif c in ("scenario","kit_number","module_number","movement_type","document_number","in_type","out_type"):
                width = 150
            elif c.startswith(tuple(["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"])):
                width = 110
            elif c in ("total_in","total_out"):
                width = 120
            self.tree.heading(c, text=heading_map.get(c,c))
            self.tree.column(c, width=width, minwidth=80, stretch=False, anchor="w")

        for idx, row in enumerate(self.rows):
            dtype = (row.get("type") or "").upper()
            values = []
            for c in cols:
                if c in ("code","description","scenario","kit_number","module_number","movement_type","document_number","in_type","out_type"):
                    values.append(row.get(c,""))
                elif c == "total_in":
                    values.append(row.get("total_in",0))
                elif c == "total_out":
                    values.append(row.get("total_out",0))
                elif re.match(r"^[A-Za-z]{3}-\d{4} IN$", c):
                    base = c[:-3]
                    dt = datetime.strptime(base, "%b-%Y")
                    values.append(row["per_month_in"].get((dt.year, dt.month),0))
                elif re.match(r"^[A-Za-z]{3}-\d{4} OUT$", c):
                    base = c[:-4]
                    dt = datetime.strptime(base, "%b-%Y")
                    values.append(row["per_month_out"].get((dt.year, dt.month),0))
                else:
                    values.append(row.get(c,""))
            tag = "alt" if idx % 2 else "norm"
            if dtype == "KIT": tag = "kitrow"
            elif dtype == "MODULE": tag = "modrow"
            self.tree.insert("", "end", values=values, tags=(tag,))

    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        self.toggle_btn.config(
            text=self.t("toggle_simple","Simple") if self.simple_mode else self.t("toggle_detailed","Detailed")
        )
        self._populate_tree()
        if self.graph_visible:
            self.draw_or_update_graph()

    # Graph popup
    def toggle_graph_popup(self):
        if self.graph_visible: self.close_graph_popup()
        else: self.open_graph_popup()

    def open_graph_popup(self):
        if self.graph_window and tk.Toplevel.winfo_exists(self.graph_window):
            self.graph_window.focus(); return
        self.graph_window = tk.Toplevel(self)
        self.graph_window.title(self._dynamic_chart_title())
        self.graph_window.configure(bg=BG_MAIN)
        self.graph_window.geometry("840x560")
        self.graph_window.protocol("WM_DELETE_WINDOW", self.close_graph_popup)

        ctrl = tk.Frame(self.graph_window, bg=BG_MAIN)
        ctrl.pack(fill="x", pady=(8,4), padx=8)
        tk.Button(ctrl, text=self.t("chart_refresh","Refresh Chart"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  command=self.draw_or_update_graph).pack(side="left", padx=(0,6))
        tk.Button(ctrl, text=self.t("chart_export","Export Chart"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  command=self.export_chart_ps).pack(side="left", padx=(0,6))
        tk.Button(ctrl, text=self.t("chart_export_image","Export Chart (Image)"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  command=self.export_chart_image).pack(side="left", padx=(0,6))

        # Scrollable canvas area (horizontal if needed)
        canvas_container = tk.Frame(self.graph_window, bg=BG_MAIN)
        canvas_container.pack(fill="both", expand=True, padx=8, pady=(0,8))
        hsb = ttk.Scrollbar(canvas_container, orient="horizontal")
        hsb.pack(side="bottom", fill="x")
        self.chart_canvas = tk.Canvas(canvas_container, bg="#FFFFFF",
                                      highlightthickness=1,
                                      highlightbackground=COLOR_BORDER,
                                      xscrollcommand=hsb.set)
        self.chart_canvas.pack(side="left", fill="both", expand=True)
        hsb.config(command=self.chart_canvas.xview)

        self.chart_canvas.bind("<Configure>", lambda e: self._redraw_chart())
        self.graph_visible = True
        self.graph_btn.config(text=self.t("toggle_graph_hide","Hide Graph"))
        self.draw_or_update_graph()

    def close_graph_popup(self):
        if self.graph_window and tk.Toplevel.winfo_exists(self.graph_window):
            try: self.graph_window.destroy()
            except Exception: pass
        self.graph_window = None
        self.chart_canvas = None
        self.graph_visible = False
        self.graph_btn.config(text=self.t("toggle_graph_show","Show Graph"))

    def _dynamic_chart_title(self):
        proj = f"{self.project_name} - {self.project_code}"
        template = lang.t("consumption.chart_title_dynamic",
                          fallback="{project} Consumption per month (number of items)")
        mode = self.dataset_mode.get()
        if mode == "Reception": suffix = " (reception)"
        elif mode == "Consumption": suffix = " (consumption)"
        else: suffix = " (reception & consumption)"
        return template.format(project=proj) + suffix

    def _aggregate_month_lines(self):
        months = self.months_seq
        mode = self.dataset_mode.get()
        in_line = OrderedDict()
        out_line = OrderedDict()
        for ym in months:
            in_line[ym] = 0
            out_line[ym] = 0
        for row in self.rows:
            if mode in ("All","Reception"):
                for ym, q in row["per_month_in"].items():
                    if ym in in_line: in_line[ym] += q
            if mode in ("All","Consumption"):
                for ym, q in row["per_month_out"].items():
                    if ym in out_line: out_line[ym] += q
        return in_line, out_line

    def draw_or_update_graph(self):
        if not self.graph_visible or not self.chart_canvas:
            return
        if self.graph_window and tk.Toplevel.winfo_exists(self.graph_window):
            self.graph_window.title(self._dynamic_chart_title())
        self._redraw_chart()

    def _redraw_chart(self):
        if not self.chart_canvas: return
        self.chart_canvas.delete("all")
        w = self.chart_canvas.winfo_width()
        h = self.chart_canvas.winfo_height()
        if w < 240 or h < 220:
            return
        if not self.months_seq:
            self.chart_canvas.create_text(w/2,h/2,text=self.t("no_data","No Data"),
                                          font=("Helvetica",12,"bold"), fill="#444")
            return
        mode = self.dataset_mode.get()
        in_line, out_line = self._aggregate_month_lines()
        has_in = (mode in ("All","Reception"))
        has_out = (mode in ("All","Consumption"))
        if (has_in and all(v==0 for v in in_line.values())) and \
           (has_out and all(v==0 for v in out_line.values())):
            self.chart_canvas.create_text(w/2,h/2,text=self.t("no_data","No Data"),
                                          font=("Helvetica",12,"bold"), fill="#444")
            return

        pad = self.chart_pad
        # If many months, extend virtual width
        months_count = len(self.months_seq)
        extra_width = max(0, (months_count - 14) * 50)
        virtual_width = w + extra_width
        left = pad
        right = virtual_width - pad - 20
        top = pad + 30
        bottom = h - pad - 30

        max_val = 1
        if has_in: max_val = max(max_val, max(in_line.values()) or 0)
        if has_out: max_val = max(max_val, max(out_line.values()) or 0)
        if max_val == 0: max_val = 1

        # Axes
        self.chart_canvas.create_line(left, bottom, right, bottom, width=2)
        self.chart_canvas.create_line(left, bottom, left, top, width=2)

        steps = 5
        for i in range(steps+1):
            y_val = bottom - (bottom - top)*i/steps
            self.chart_canvas.create_line(left, y_val, right, y_val,
                                          fill="#E5E8EB",
                                          dash=() if i in (0,steps) else (2,4))
            val_label = int(round(max_val*i/steps))
            self.chart_canvas.create_text(left-8, y_val, text=str(val_label),
                                          anchor="e", font=("Helvetica",8), fill="#333")

        # X coords
        n = len(self.months_seq)
        if n == 1:
            x_coords = [(left + right)/2]
        else:
            x_coords = [ left + i*(right-left)/(n-1) for i in range(n) ]

        def build_points(line):
            pts=[]
            for (x,(y,m)) in zip(x_coords, self.months_seq):
                val = line[(y,m)]
                y_pix = bottom - (val/max_val)*(bottom-top)
                pts.append((x,y_pix,val, ym_label(y,m)))
            return pts

        in_pts = build_points(in_line) if has_in else []
        out_pts = build_points(out_line) if has_out else []

        def draw_series(pts, color):
            for i in range(1,len(pts)):
                self.chart_canvas.create_line(pts[i-1][0], pts[i-1][1],
                                              pts[i][0], pts[i][1],
                                              fill=color, width=2, smooth=True)
            for (x,y_pix,val,label) in pts:
                self.chart_canvas.create_oval(x-4,y_pix-4,x+4,y_pix+4,
                                              fill=color, outline=color)
                self.chart_canvas.create_text(x, y_pix-10, text=str(val),
                                              font=("Helvetica",8,"bold"), fill=color)

        if has_in: draw_series(in_pts, LINE_COLOR_IN)
        if has_out: draw_series(out_pts, LINE_COLOR_OUT)

        rotate = 45 if n > 10 else 0
        label_pts = in_pts if in_pts else out_pts
        for (x,y_pix,val,label) in label_pts:
            self.chart_canvas.create_text(x, bottom+ (20 if rotate else 12),
                                          text=label, angle=rotate,
                                          font=("Helvetica",8), fill="#333")

        self.chart_canvas.create_text((left+right)/2, top-40,
                                      text=self._dynamic_chart_title(),
                                      font=("Helvetica",13,"bold"), fill="#111827")

        # Legend
        legend_items = []
        if has_in: legend_items.append((LINE_COLOR_IN, self.t("legend_reception","Reception")))
        if has_out: legend_items.append((LINE_COLOR_OUT, self.t("legend_consumption","Consumption")))
        lx = left + 10
        ly = top - 12
        for color,label in legend_items:
            self.chart_canvas.create_rectangle(lx, ly-8, lx+16, ly+8, fill=color, outline=color)
            self.chart_canvas.create_text(lx+24, ly, text=label, anchor="w",
                                          font=("Helvetica",9))
            lx += 140

        # Configure scroll region
        self.chart_canvas.config(scrollregion=(0,0,virtual_width, h))

    # Exports
    def export_chart_ps(self):
        if not self.chart_canvas:
            custom_popup(self, self.t("error","Error"), self.t("no_data","No Data"), "error"); return
        file_path = filedialog.asksaveasfilename(defaultextension=".ps",
                                                 filetypes=[("PostScript","*.ps")],
                                                 title=self.t("chart_export","Export Chart"))
        if not file_path: return
        try:
            self.chart_canvas.postscript(colormode='color', file=file_path)
            custom_popup(self, self.t("success","Success"), f"Chart exported: {file_path}", "info")
        except Exception as e:
            custom_popup(self, self.t("error","Error"), f"Export failed: {e}", "error")

    def export_chart_image(self):
        if not self.chart_canvas:
            custom_popup(self, self.t("error","Error"), self.t("no_data","No Data"), "error"); return
        if not PIL_AVAILABLE:
            custom_popup(self, self.t("chart_export_image","Export Chart (Image)"),
                         "Pillow not installed.\nInstall:\n  pip install Pillow", "warning")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG","*.png"),("JPEG","*.jpg;*.jpeg"),("All Files","*.*")],
                                                 title=self.t("chart_export_image","Export Chart (Image)"))
        if not file_path: return
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".png",".jpg",".jpeg"):
            file_path += ".png"; ext = ".png"
        expected_w = self.chart_canvas.winfo_width()
        expected_h = self.chart_canvas.winfo_height()

        def capture_once():
            self.graph_window.lift()
            self.graph_window.attributes("-topmost", True)
            self.graph_window.update_idletasks()
            self.graph_window.after(40)
            self.graph_window.update()
            x0 = self.chart_canvas.winfo_rootx()
            y0 = self.chart_canvas.winfo_rooty()
            x1 = x0 + expected_w
            y1 = y0 + expected_h
            img = ImageGrab.grab(bbox=(x0,y0,x1,y1))
            self.graph_window.after(50, lambda: self.graph_window.attributes("-topmost", False))
            return img

        success_img = None
        for attempt in range(1, SCREENSHOT_RETRIES+1):
            try:
                self.graph_window.update_idletasks()
                time.sleep(SCREENSHOT_DELAY_MS/1000.0)
                img = capture_once()
                iw, ih = img.size
                if iw >= expected_w*SCREENSHOT_TOLERANCE_RATIO and ih >= expected_h*SCREENSHOT_TOLERANCE_RATIO:
                    success_img = img
                    break
            except Exception:
                continue
        if not success_img:
            custom_popup(self, self.t("error","Error"),
                         "Screenshot capture failed. Use PostScript export instead.",
                         "error")
            return
        try:
            if ext in (".jpg",".jpeg") and success_img.mode != "RGB":
                success_img = success_img.convert("RGB")
            args = {"quality":92} if ext in (".jpg",".jpeg") else {}
            success_img.save(file_path, **args)
            custom_popup(self, self.t("success","Success"),
                         self.t("chart_export_success","Chart image saved: {f}").format(f=file_path),
                         "info")
        except Exception as e:
            custom_popup(self, self.t("error","Error"),
                         self.t("chart_export_fail","Image export failed: {err}").format(err=str(e)),
                         "error")

    def export_excel(self):
        if not self.rows:
            custom_popup(self, self.t("no_data","No Data"),
                         self.t("nothing_export","Nothing to export."),"warning")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Files","*.xlsx")],
                                                 title=self.t("export_dialog","Save Combined Report"),
                                                 initialfile=f"Combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not file_path: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Combined"
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append(["Generated", now_str])
            ws.append(["Stock Action", self.dataset_mode.get()])
            ws.append(["Mgmt Mode", self.mgmt_var.get(), "Scenario", self.scenario_var.get(), "Type", self.type_var.get()])
            ws.append(["Kit", self.kit_var.get(), "Module", self.module_var.get(), "Item Search", self.item_search_var.get()])
            ws.append(["IN Type", self.in_type_var.get(), "IN Movement", self.in_move_var.get(),
                       "Out Type", self.out_type_var.get(), "Out Movement", self.out_move_var.get()])
            ws.append(["Document", self.doc_var.get(), "From", self.from_var.get(), "To", self.to_var.get()])
            ws.append([])
            cols = self._current_columns()
            ws.append([c.replace("_"," ").title() for c in cols])

            kit_fill = PatternFill(start_color=KIT_FILL_COLOR, end_color=KIT_FILL_COLOR, fill_type="solid")
            module_fill = PatternFill(start_color=MODULE_FILL_COLOR, end_color=MODULE_FILL_COLOR, fill_type="solid")

            for r in self.rows:
                row_vals=[]
                for c in cols:
                    if c in ("code","description","scenario","kit_number","module_number",
                             "movement_type","document_number","in_type","out_type"):
                        row_vals.append(r.get(c,""))
                    elif c == "total_in":
                        row_vals.append(r.get("total_in",0))
                    elif c == "total_out":
                        row_vals.append(r.get("total_out",0))
                    elif re.match(r"^[A-Za-z]{3}-\d{4} IN$", c):
                        base = c[:-3]
                        dt = datetime.strptime(base, "%b-%Y")
                        row_vals.append(r["per_month_in"].get((dt.year, dt.month), 0))
                    elif re.match(r"^[A-Za-z]{3}-\d{4} OUT$", c):
                        base = c[:-4]
                        dt = datetime.strptime(base, "%b-%Y")
                        row_vals.append(r["per_month_out"].get((dt.year, dt.month), 0))
                    else:
                        row_vals.append(r.get(c,""))
                ws.append(row_vals)
                dtype = (r.get("type") or "").upper()
                if dtype == "KIT":
                    for cell in ws[ws.max_row]: cell.fill = kit_fill
                elif dtype == "MODULE":
                    for cell in ws[ws.max_row]: cell.fill = module_fill

            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(v))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            wb.save(file_path)
            custom_popup(self, self.t("success","Success"),
                         self.t("export_success","Export completed: {f}").format(f=file_path),
                         "info")
        except Exception as e:
            custom_popup(self, self.t("error","Error"),
                         self.t("export_fail","Export failed: {err}").format(err=str(e)),
                         "error")

__all__ = ["Consumption"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Actions (Reception & Consumption) v2.1")
    class Dummy: pass
    d = Dummy()
    Consumption(root, d).pack(fill="both", expand=True)
    root.geometry("1650x880")
    root.mainloop()