"""
donations.py  v1.0

Donations Monitor

Purpose:
  Track all IN and OUT donation movements per Date, Code and Third Party.
  Aggregates In-Donations (IN_Type='In Donation') and Out-Donations (Out_Type='Out Donation')
  so you can quickly see net donated vs distributed quantities per third party and item.

Grouping Key:
  (Date, Code, Third_Party)  -- daily granularity to show chronological donation flow.

Filters (Drop-down / text):
  - Scenario        (scenarios.name)
  - Kit Number      (stock_transactions.Kit)
  - Module Number   (stock_transactions.Module)
  - Third Party     (third_parties.name)
  - Type            (Kit, Module, Item, All)   (determined via detect_type)
  - Item Search     (search code + translated description; applies only when Type=Item or All)
  - Date Range      (From / To) multi-format parsing; To capped at today
  - Document Number (substring search)
  - Simple / Detailed mode toggle

Columns:
  Simple Mode:
     date, code, description, in_donations, out_donations, remarks (aggregated)
  Detailed Mode:
     date, scenarios, kits, modules, type, code, description, third_party,
     in_donations, out_donations, expiry_dates, documents, remarks

  * scenarios / kits / modules / documents / remarks / expiry_dates:
        comma‑separated unique sorted lists aggregated per group
  * expiry_dates taken from stock_transactions.Expiry_date (fallback to None if blank)

Keyboard:
  - Enter in search/date/document fields => Refresh
  - Esc in a field => clears that field & refreshes
  - Esc globally (when focus not inside an Entry) => clear all filters

Excel Export:
  - Includes filter summary + mode
  - Color rows for KIT (green) & MODULE (light blue)
  - Preserves all displayed columns in current mode

Dependencies:
  - db.connect_db
  - manage_items.get_item_description, manage_items.detect_type
  - language_manager.lang
  - popup_utils.custom_popup
  - tkcalendar (optional) for date picking

NOTE:
  If you later decide to add balance (in - out) or running totals, you can extend
  the aggregation loop easily—left as-is per current specification.
"""

import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import re
from datetime import date, datetime
from calendar import monthrange
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Optional calendar
try:
    from tkcalendar import DateEntry
    TKCAL_AVAILABLE = True
except Exception:
    TKCAL_AVAILABLE = False

from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang
from popup_utils import custom_popup

# ---------------- THEME ----------------
BG_MAIN        = "#F0F4F8"
BG_PANEL       = "#FFFFFF"
COLOR_PRIMARY  = "#2C3E50"
COLOR_BORDER   = "#D0D7DE"
ROW_ALT_COLOR  = "#F7FAFC"
ROW_NORM_COLOR = "#FFFFFF"
BTN_EXPORT     = "#2980B9"
BTN_REFRESH    = "#2563EB"
BTN_CLEAR      = "#7F8C8D"
BTN_TOGGLE     = "#8E44AD"

KIT_FILL_COLOR    = "228B22"
MODULE_FILL_COLOR = "ADD8E6"

# ---------------- DATE PARSING ----------------
DATE_REGEXES = [
    ("%Y-%m-%d", re.compile(r"^\d{4}-\d{2}-\d{2}$")),
    ("%d/%m/%Y", re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")),
    ("%d-%m-%Y", re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$")),
    ("%Y/%m/%d", re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")),
    ("%d %b %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$")),
    ("%d %B %Y", re.compile(r"^\d{1,2}\s+[A-Za-z]+\s+\d{4}$")),
]
MONTH_ONLY_REGEXES = [
    ("%Y-%m", re.compile(r"^\d{4}-\d{1,2}$")),
    ("%m/%Y", re.compile(r"^\d{1,2}/\d{4}$")),
    ("%b-%Y", re.compile(r"^[A-Za-z]{3}-\d{4}$")),
    ("%B-%Y", re.compile(r"^[A-Za-z]+-\d{4}$")),
]

def parse_user_date(text: str, bound: str):
    if not text:
        return None
    raw = text.strip()
    for fmt, rx in DATE_REGEXES:
        if rx.match(raw):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                pass
    for fmt, rx in MONTH_ONLY_REGEXES:
        if rx.match(raw):
            try:
                tmp = datetime.strptime(raw, fmt)
                y, m = tmp.year, tmp.month
                return date(y, m, 1) if bound == "from" else date(y, m, monthrange(y, m)[1])
            except Exception:
                pass
    if re.match(r"^\d{4}$", raw):
        y = int(raw)
        return date(y, 1, 1) if bound == "from" else date(y, 12, 31)
    return None

# ---------------- FILTER DATA LOADERS ----------------
def fetch_scenarios():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM scenarios ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()

def fetch_kit_numbers():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT Kit FROM stock_transactions
            WHERE Kit IS NOT NULL AND Kit!='None' AND Kit!=''
            ORDER BY Kit
        """)
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()

def fetch_module_numbers():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT Module FROM stock_transactions
            WHERE Module IS NOT NULL AND Module!='None' AND Module!=''
            ORDER BY Module
        """)
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()

def fetch_third_parties():
    conn = connect_db()
    if conn is None: return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM third_parties ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error:
        return []
    finally:
        cur.close(); conn.close()

# ---------------- AGGREGATION ----------------
def aggregate_donations(filters):
    scenario = filters.get("scenario")
    kit_number = filters.get("kit")
    module_number = filters.get("module")
    type_filter = filters.get("type")
    item_search = filters.get("item_search")
    doc_search = filters.get("doc_number")
    third_party_filter = filters.get("third_party")
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")

    conn = connect_db()
    if conn is None:
        return []

    cur = conn.cursor()
    clauses = []
    params = []

    if scenario and scenario.lower() != "all":
        clauses.append("Scenario = ?"); params.append(scenario)
    if kit_number and kit_number.lower() != "all":
        clauses.append("Kit = ?"); params.append(kit_number)
    if module_number and module_number.lower() != "all":
        clauses.append("Module = ?"); params.append(module_number)
    if third_party_filter and third_party_filter.lower() != "all":
        clauses.append("Third_Party = ?"); params.append(third_party_filter)
    if doc_search:
        clauses.append("document_number LIKE ?"); params.append(f"%{doc_search}%")
    if date_from:
        clauses.append("Date >= ?"); params.append(date_from.strftime("%Y-%m-%d"))
    if date_to:
        clauses.append("Date <= ?"); params.append(date_to.strftime("%Y-%m-%d"))

    donation_clause = "(IN_Type='In Donation' OR Out_Type='Out Donation')"
    base_where = " AND ".join(clauses) if clauses else "1=1"
    sql = f"""
        SELECT Date, code, Scenario, Kit, Module,
               Qty_IN, IN_Type, Qty_Out, Out_Type,
               document_number, Third_Party, Remarks, Expiry_date
        FROM stock_transactions
        WHERE {base_where} AND {donation_clause}
    """
    try:
        cur.execute(sql, params)
        data = cur.fetchall()
    except sqlite3.Error:
        cur.close(); conn.close()
        return []

    # Group by (Date, Code, Third_Party)
    groups = {}
    scen_map = defaultdict(set)
    kit_map = defaultdict(set)
    mod_map = defaultdict(set)
    doc_map = defaultdict(set)
    remarks_map = defaultdict(set)
    expiry_map = defaultdict(set)

    for (dt_str, code, scen, kit, module,
         qty_in, in_type, qty_out, out_type,
         doc, third_party, remarks, exp_date) in data:
        if not code:
            continue
        desc = get_item_description(code)
        dtype = detect_type(code, desc)

        if type_filter and type_filter.lower() != "all":
            if dtype.lower() != type_filter.lower():
                continue
        if item_search:
            if dtype.lower() != "item":
                continue
            if (item_search.lower() not in code.lower()
                and item_search.lower() not in desc.lower()):
                continue

        third_party = third_party or ""
        key = (dt_str, code, third_party)
        rec = groups.setdefault(key, {
            "date": dt_str,
            "code": code,
            "third_party": third_party,
            "description": desc,
            "type": dtype,
            "in_donations": 0,
            "out_donations": 0
        })

        if in_type == "In Donation" and qty_in:
            try: rec["in_donations"] += int(qty_in)
            except: pass
        if out_type == "Out Donation" and qty_out:
            try: rec["out_donations"] += int(qty_out)
            except: pass

        if scen: scen_map[key].add(scen)
        if kit: kit_map[key].add(kit)
        if module: mod_map[key].add(module)
        if doc: doc_map[key].add(doc)
        if remarks: remarks_map[key].add(remarks)
        if exp_date: expiry_map[key].add(exp_date)

    rows = []
    for key, rec in groups.items():
        dt_str, code, tp = key
        rec["scenarios"] = ", ".join(sorted(scen_map[key])) if scen_map[key] else ""
        rec["kits"] = ", ".join(sorted(kit_map[key])) if kit_map[key] else ""
        rec["modules"] = ", ".join(sorted(mod_map[key])) if mod_map[key] else ""
        rec["documents"] = ", ".join(sorted(doc_map[key])) if doc_map[key] else ""
        rec["remarks"] = ", ".join(sorted(remarks_map[key])) if remarks_map[key] else ""
        rec["expiry_dates"] = ", ".join(sorted(expiry_map[key])) if expiry_map[key] else ""
        rows.append(rec)

    rows.sort(key=lambda r: (r["date"], r["type"], r["code"], r["third_party"]))
    cur.close(); conn.close()
    return rows

# ---------------- UI CLASS ----------------
class Donations(tk.Frame):
    SIMPLE_COLS = ["date","code","description","in_donations","out_donations","remarks"]
    DETAIL_COLS = [
        "date","scenarios","kits","modules","type",
        "code","description","third_party",
        "in_donations","out_donations","expiry_dates",
        "documents","remarks"
    ]

    def __init__(self, parent, app, *args, **kwargs):
        super().__init__(parent, bg=BG_MAIN, *args, **kwargs)
        self.app = app
        self.rows = []
        self.simple_mode = False

        # Filters
        self.scenario_var = tk.StringVar(value="All")
        self.kit_var = tk.StringVar(value="All")
        self.module_var = tk.StringVar(value="All")
        self.third_party_var = tk.StringVar(value="All")
        self.type_var = tk.StringVar(value="All")
        self.item_search_var = tk.StringVar()
        self.doc_var = tk.StringVar()
        self.from_var = tk.StringVar()
        self.to_var = tk.StringVar()

        self.status_var = tk.StringVar(value=lang.t("donations.ready","Ready"))
        self.tree = None

        self._build_ui()
        self.populate_filters()
        self._set_default_dates()
        self.refresh()

    # ---------- UI BUILD ----------
    def _build_ui(self):
        header = tk.Frame(self, bg=BG_MAIN)
        header.pack(fill="x", padx=12, pady=(12,6))

        tk.Label(header,
                 text=lang.t("donations.title","Donations"),
                 font=("Helvetica",20,"bold"),
                 bg=BG_MAIN, fg=COLOR_PRIMARY).pack(side="left")

        self.toggle_btn = tk.Button(header,
                                    text=lang.t("donations.detailed","Detailed"),
                                    bg=BTN_TOGGLE, fg="#FFFFFF",
                                    relief="flat", padx=14, pady=6,
                                    command=self.toggle_mode)
        self.toggle_btn.pack(side="right", padx=(6,0))

        tk.Button(header,
                  text=lang.t("generic.export","Export"),
                  bg=BTN_EXPORT, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.export_excel)\
            .pack(side="right", padx=(6,0))
        tk.Button(header,
                  text=lang.t("generic.clear","Clear"),
                  bg=BTN_CLEAR, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.clear_filters)\
            .pack(side="right", padx=(6,0))
        tk.Button(header,
                  text=lang.t("generic.refresh","Refresh"),
                  bg=BTN_REFRESH, fg="#FFFFFF", relief="flat",
                  padx=14, pady=6, command=self.refresh)\
            .pack(side="right", padx=(6,0))

        filters = tk.Frame(self, bg=BG_MAIN)
        filters.pack(fill="x", padx=12, pady=(0,8))

        # Row 1
        r1 = tk.Frame(filters, bg=BG_MAIN)
        r1.pack(fill="x", pady=2)

        tk.Label(r1, text=lang.t("generic.scenario","Scenario"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.scenario_cb = ttk.Combobox(r1, textvariable=self.scenario_var,
                                        state="readonly", width=20)
        self.scenario_cb.grid(row=0, column=1, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.kit_number","Kit Number"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.kit_cb = ttk.Combobox(r1, textvariable=self.kit_var,
                                   state="readonly", width=16)
        self.kit_cb.grid(row=0, column=3, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.module_number","Module Number"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.module_cb = ttk.Combobox(r1, textvariable=self.module_var,
                                      state="readonly", width=16)
        self.module_cb.grid(row=0, column=5, padx=(0,12))

        tk.Label(r1, text=lang.t("generic.type","Type"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.type_cb = ttk.Combobox(r1, textvariable=self.type_var,
                                    state="readonly", width=12,
                                    values=["All","Kit","Module","Item"])
        self.type_cb.grid(row=0, column=7, padx=(0,12))

        tk.Label(r1, text=lang.t("donations.third_party","Third Party"), bg=BG_MAIN)\
            .grid(row=0, column=8, sticky="w", padx=(0,4))
        self.third_party_cb = ttk.Combobox(r1, textvariable=self.third_party_var,
                                           state="readonly", width=18)
        self.third_party_cb.grid(row=0, column=9, padx=(0,4))

        # Row 2
        r2 = tk.Frame(filters, bg=BG_MAIN)
        r2.pack(fill="x", pady=2)

        tk.Label(r2, text=lang.t("donations.item_search","Item Search"), bg=BG_MAIN)\
            .grid(row=0, column=0, sticky="w", padx=(0,4))
        self.item_entry = tk.Entry(r2, textvariable=self.item_search_var, width=20)
        self.item_entry.grid(row=0, column=1, padx=(0,12))
        self.item_entry.bind("<Return>", lambda e: self.refresh())
        self.item_entry.bind("<Escape>", lambda e: self._clear_field(self.item_search_var))

        tk.Label(r2, text=lang.t("generic.document_number","Document Number"), bg=BG_MAIN)\
            .grid(row=0, column=2, sticky="w", padx=(0,4))
        self.doc_entry = tk.Entry(r2, textvariable=self.doc_var, width=20)
        self.doc_entry.grid(row=0, column=3, padx=(0,12))
        self.doc_entry.bind("<Return>", lambda e: self.refresh())
        self.doc_entry.bind("<Escape>", lambda e: self._clear_field(self.doc_var))

        tk.Label(r2, text=lang.t("generic.from_date","From Date"), bg=BG_MAIN)\
            .grid(row=0, column=4, sticky="w", padx=(0,4))
        self.from_entry = self._date_widget(r2, self.from_var)
        self.from_entry.grid(row=0, column=5, padx=(0,12))

        tk.Label(r2, text=lang.t("generic.to_date","To Date"), bg=BG_MAIN)\
            .grid(row=0, column=6, sticky="w", padx=(0,4))
        self.to_entry = self._date_widget(r2, self.to_var)
        self.to_entry.grid(row=0, column=7, padx=(0,12))

        # Status Bar
        tk.Label(self, textvariable=self.status_var, anchor="w",
                 bg=BG_MAIN, fg=COLOR_PRIMARY, relief="sunken")\
            .pack(fill="x", padx=12, pady=(0,8))

        # Table Container
        table_frame = tk.Frame(self, bg=COLOR_BORDER, bd=1, relief="solid")
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self.tree = ttk.Treeview(table_frame, columns=(), show="headings", height=22)
        self.tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        hsb.pack(fill="x", padx=12, pady=(0,12))
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        style = ttk.Style()
        try: style.theme_use("clam")
        except: pass
        style.configure("Treeview",
                        background=BG_PANEL, fieldbackground=BG_PANEL,
                        foreground=COLOR_PRIMARY, rowheight=24,
                        font=("Helvetica",10))
        style.configure("Treeview.Heading",
                        background="#E5E8EB", foreground=COLOR_PRIMARY,
                        font=("Helvetica",11,"bold"))
        self.tree.tag_configure("alt", background=ROW_ALT_COLOR)
        self.tree.tag_configure("kitrow", background="#228B22", foreground="#FFFFFF")
        self.tree.tag_configure("modrow", background="#ADD8E6")

        # Global ESC for clearing all (when not inside entry)
        self.bind_all("<Escape>", self._esc_global)

    def _esc_global(self, event):
        if isinstance(event.widget, tk.Entry):
            return
        self.clear_filters()

    def _date_widget(self, parent, var):
        if TKCAL_AVAILABLE:
            w = DateEntry(parent, textvariable=var, width=12,
                          date_pattern="yyyy-mm-dd",
                          showweeknumbers=False,
                          background="#2563EB", foreground="white", borderwidth=1)
            w.bind("<Return>", lambda e: self.refresh())
            w.bind("<Escape>", lambda e: self._clear_field(var))
            return w
        else:
            e = tk.Entry(parent, textvariable=var, width=12)
            e.bind("<Return>", lambda ev: self.refresh())
            e.bind("<Escape>", lambda ev: self._clear_field(var))
            return e

    def _clear_field(self, tk_var):
        tk_var.set("")
        self.refresh()

    def _set_default_dates(self):
        today = date.today()
        try:
            past = date(today.year - 1, today.month, today.day)
        except:
            past = today
        self.from_var.set(past.strftime("%Y-%m-%d"))
        self.to_var.set(today.strftime("%Y-%m-%d"))

    # ---------- Filters ----------
    def populate_filters(self):
        scenarios = fetch_scenarios()
        kits = fetch_kit_numbers()
        modules = fetch_module_numbers()
        tps = fetch_third_parties()

        self.scenario_cb['values'] = ["All"] + scenarios
        self.kit_cb['values'] = ["All"] + kits
        self.module_cb['values'] = ["All"] + modules
        self.third_party_cb['values'] = ["All"] + tps

        for var, cb in [
            (self.scenario_var, self.scenario_cb),
            (self.kit_var, self.kit_cb),
            (self.module_var, self.module_cb),
            (self.third_party_var, self.third_party_cb)
        ]:
            if var.get() not in cb['values']:
                var.set("All")

    # ---------- Mode ----------
    def toggle_mode(self):
        self.simple_mode = not self.simple_mode
        self.toggle_btn.config(text=lang.t("donations.simple","Simple") if self.simple_mode
                               else lang.t("donations.detailed","Detailed"))
        self._populate_tree()

    # ---------- Refresh ----------
    def refresh(self):
        filters = {
            "scenario": self.scenario_var.get(),
            "kit": self.kit_var.get(),
            "module": self.module_var.get(),
            "third_party": self.third_party_var.get(),
            "type": self.type_var.get(),
            "item_search": self.item_search_var.get().strip(),
            "doc_number": self.doc_var.get().strip(),
            "date_from": parse_user_date(self.from_var.get().strip(), "from") if self.from_var.get().strip() else None,
            "date_to": parse_user_date(self.to_var.get().strip(), "to") if self.to_var.get().strip() else None
        }
        if filters["date_to"] and filters["date_to"] > date.today():
            filters["date_to"] = date.today()
        self.rows = aggregate_donations(filters)
        self._populate_tree()
        self.status_var.set(lang.t("donations.loaded","Loaded {n} rows").format(n=len(self.rows)))

    def clear_filters(self):
        self.scenario_var.set("All")
        self.kit_var.set("All")
        self.module_var.set("All")
        self.third_party_var.set("All")
        self.type_var.set("All")
        self.item_search_var.set("")
        self.doc_var.set("")
        self._set_default_dates()
        self.refresh()

    # ---------- Columns ----------
    def _current_columns(self):
        return self.SIMPLE_COLS if self.simple_mode else self.DETAIL_COLS

    def _populate_tree(self):
        cols = self._current_columns()
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols

        headings = {
            "date": lang.t("generic.date","Date"),
            "scenarios": lang.t("generic.scenario","Scenario(s)"),
            "kits": lang.t("generic.kit_number","Kit Number(s)"),
            "modules": lang.t("generic.module_number","Module Number(s)"),
            "type": lang.t("generic.type","Type"),
            "code": lang.t("generic.code","Code"),
            "description": lang.t("generic.description","Description"),
            "third_party": lang.t("donations.third_party","Third Party"),
            "in_donations": lang.t("donations.in_donations","In-Donations"),
            "out_donations": lang.t("donations.out_donations","Out-Donations"),
            "expiry_dates": lang.t("donations.expiry_dates","Expiry Date(s)"),
            "documents": lang.t("generic.document_number","Document Number(s)"),
            "remarks": lang.t("donations.remarks","Remarks")
        }

        for c in cols:
            width = 120
            if c == "description":
                width = 340
            elif c == "code":
                width = 160
            elif c in ("scenarios","kits","modules","documents","remarks","expiry_dates"):
                width = 220
            elif c == "third_party":
                width = 180
            elif c in ("in_donations","out_donations"):
                width = 130
            self.tree.heading(c, text=headings.get(c, c))
            self.tree.column(c, width=width, anchor="w", stretch=True)

        for idx, r in enumerate(self.rows):
            if self.simple_mode:
                values = [
                    r["date"], r["code"], r["description"],
                    r["in_donations"], r["out_donations"], r["remarks"]
                ]
            else:
                values = [
                    r["date"], r["scenarios"], r["kits"], r["modules"], r["type"],
                    r["code"], r["description"], r["third_party"],
                    r["in_donations"], r["out_donations"], r["expiry_dates"],
                    r["documents"], r["remarks"]
                ]
            tag = "alt" if idx % 2 else ""
            if r["type"].upper() == "KIT":
                tag = "kitrow"
            elif r["type"].upper() == "MODULE":
                tag = "modrow"
            self.tree.insert("", "end", values=values, tags=(tag,))

    # ---------- Export ----------
    def export_excel(self):
        if not self.rows:
            custom_popup(self, lang.t("generic.info","Info"),
                         lang.t("donations.no_data_export","Nothing to export."), "warning")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files","*.xlsx")],
            title=lang.t("donations.export_title","Save Donations Report"),
            initialfile=f"Donations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Donations"

            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([lang.t("generic.generated","Generated"), now_str])
            ws.append([lang.t("generic.filters_used","Filters Used")])
            ws.append(["Scenario", self.scenario_var.get(),
                       "Kit", self.kit_var.get(),
                       "Module", self.module_var.get()])
            ws.append(["Third Party", self.third_party_var.get(),
                       "Type", self.type_var.get(),
                       "Document", self.doc_var.get()])
            ws.append(["From", self.from_var.get(),
                       "To", self.to_var.get(),
                       "Mode", "Simple" if self.simple_mode else "Detailed"])
            ws.append([])

            cols = self._current_columns()
            headers = [c.replace("_"," ").title() for c in cols]
            ws.append(headers)

            kit_fill = PatternFill(start_color=KIT_FILL_COLOR, end_color=KIT_FILL_COLOR, fill_type="solid")
            module_fill = PatternFill(start_color=MODULE_FILL_COLOR, end_color=MODULE_FILL_COLOR, fill_type="solid")

            for r in self.rows:
                if self.simple_mode:
                    line = [
                        r["date"], r["code"], r["description"],
                        r["in_donations"], r["out_donations"], r["remarks"]
                    ]
                else:
                    line = [
                        r["date"], r["scenarios"], r["kits"], r["modules"], r["type"],
                        r["code"], r["description"], r["third_party"],
                        r["in_donations"], r["out_donations"], r["expiry_dates"],
                        r["documents"], r["remarks"]
                    ]
                ws.append(line)
                dtype = r["type"].upper()
                if dtype == "KIT":
                    for c in ws[ws.max_row]: c.fill = kit_fill
                elif dtype == "MODULE":
                    for c in ws[ws.max_row]: c.fill = module_fill

            # Autosize
            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(v))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            wb.save(path)
            custom_popup(self, lang.t("generic.success","Success"),
                         lang.t("donations.export_success","Export completed: {f}").format(f=path),
                         "info")
        except Exception as e:
            custom_popup(self, lang.t("generic.error","Error"),
                         lang.t("donations.export_fail","Export failed: {err}").format(err=str(e)),
                         "error")

# ------------- Module Export -------------
__all__ = ["Donations"]

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Donations Monitor v1.0")
    Donations(root, None).pack(fill="both", expand=True)
    root.geometry("1580x820")
    root.mainloop()