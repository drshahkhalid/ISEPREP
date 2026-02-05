"""
stock_summary.py -
Stock Summary Dashboard

FEATURES:
  ✅ Group headers in wide description column for proper display
  ✅ Language-aware descriptions (EN/FR/ES from items_list)
  ✅ Stock details popup on double-click, stock card via right-click
  ✅ Right-click context menu with dual options
  ✅ Theme integration with colorful metrics
  ✅ Auto-refresh on filter changes
  ✅ Excel export with formatting
"""

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import datetime
from calendar import monthrange
import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from db import connect_db
from language_manager import lang

try:
    from popup_utils import custom_popup
except ImportError:
    from tkinter import messagebox

    def custom_popup(parent, title, msg, kind="info"):
        if kind == "error":
            messagebox.showerror(title, msg, parent=parent)
        elif kind == "warning":
            messagebox.showwarning(title, msg, parent=parent)
        else:
            messagebox.showinfo(title, msg, parent=parent)


from theme_config import AppTheme, enable_column_auto_resize


# ----------------------------- DB Helpers -----------------------------
def _fetchall(sql, params=()):
    conn = connect_db()
    if conn is None:
        return []
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        return cur.fetchall()
    except sqlite3.Error:
        return []
    finally:
        cur.close()
        conn.close()


def _fetchone(sql, params=()):
    rows = _fetchall(sql, params)
    return rows[0] if rows else None


# ------------------------- Expiry Period Logic ------------------------
def compute_recommended_months():
    cols = {
        r["name"].lower(): True for r in _fetchall("PRAGMA table_info(project_details)")
    }
    for need in ("lead_time_months", "cover_period_months", "buffer_months"):
        if need not in cols:
            return 0
    row = _fetchone(
        """
        SELECT
          COALESCE(lead_time_months,0) lt,
          COALESCE(cover_period_months,0) cp,
          COALESCE(buffer_months,0) bf
        FROM project_details
        ORDER BY rowid ASC
        LIMIT 1
    """
    )
    if not row:
        return 0
    return int(row["lt"]) + int(row["cp"]) + int(row["bf"])


def cutoff_date_from_months(months: int | None):
    if not months or months <= 0:
        return None
    today = datetime.date.today()
    year = today.year
    month = today.month + months
    while month > 12:
        month -= 12
        year += 1
    last_day = monthrange(year, month)[1]
    return datetime.date(year, month, last_day).isoformat()


# --------------------------- Scenario Mapping -------------------------
def load_scenario_maps():
    rows = _fetchall("SELECT scenario_id, name FROM scenarios")
    id_to_name = {}
    name_set = set()
    for r in rows:
        sid = r["scenario_id"]
        nm = r["name"]
        if nm:
            name_set.add(nm)
        id_to_name[str(sid)] = nm
    return id_to_name, name_set


def normalize_scenario(raw_val, id_to_name, name_set):
    if raw_val is None:
        return ""
    if raw_val in name_set:
        return raw_val
    txt = str(raw_val)
    if txt in id_to_name:
        return id_to_name[txt] or txt
    return txt


# Replace load_std_quantities_by_scenario() function:
def load_std_quantities_by_scenario(scenario_name):
    """
    Load standard quantities with treecodes
    Returns: dict with structure {scenario_name: {treecode: {"code": code, "std_qty": qty, "type": "on-shelf"/"in-box", ...}}}
    """
    result = {}

    # Check if loading all scenarios
    all_text = lang.t("stock_summary.all_scenarios", "All")
    load_all = not scenario_name or scenario_name == all_text

    # Get scenario_id(s) and names
    if load_all:
        scenario_rows = _fetchall("SELECT scenario_id, name FROM scenarios")
        scenarios = [(r["scenario_id"], r["name"]) for r in scenario_rows]
    else:
        scenario_row = _fetchone(
            "SELECT scenario_id, name FROM scenarios WHERE name = ?", (scenario_name,)
        )
        if not scenario_row:
            return {}
        scenarios = [(scenario_row["scenario_id"], scenario_row["name"])]

    for scenario_id, scenario_name in scenarios:
        if scenario_name not in result:
            result[scenario_name] = {}

        # 1. Load ON-SHELF items from compositions
        # NOTE: compositions table doesn't have treecode column
        on_shelf_rows = _fetchall(
            """
            SELECT unique_id_2, code, quantity
            FROM compositions 
            WHERE scenario_id = ?
            ORDER BY code
            """,
            (scenario_id,),
        )

        for r in on_shelf_rows:
            code = r["code"]
            treecode = code  # Use code as treecode for on-shelf items
            qty = r["quantity"] or 0

            result[scenario_name][treecode] = {
                "code": code,
                "std_qty": qty,
                "mgmt_type": "on-shelf",
                "kit_code": "",
                "module_code": "",
            }

        # 2. Load IN-BOX items from kit_items
        in_box_rows = _fetchall(
            """
            SELECT code, kit, module, std_qty, treecode
            FROM kit_items
            WHERE scenario_id = ?
            ORDER BY treecode, kit, module
            """,
            (scenario_id,),
        )

        for r in in_box_rows:
            code = r["code"]
            treecode = r["treecode"] or code
            kit = r["kit"] or ""
            module = r["module"] or ""
            qty = r["std_qty"] or 0

            # Use treecode as key
            result[scenario_name][treecode] = {
                "code": code,
                "std_qty": qty,
                "mgmt_type": "in-box",
                "kit_code": kit,
                "module_code": module,
            }

    return result


def load_item_metadata(codes):
    """Load item metadata with language-aware designation (using lang.lang_code)"""
    if not codes:
        return {}

    # CORRECT: Use lang.lang_code (discovered from debug output)
    lang_code = getattr(lang, "lang_code", "en").lower()

    # Build SQL COALESCE expression with fallback chain
    if lang_code == "fr":
        designation_expr = "COALESCE(designation_fr, designation_en, designation, code)"
    elif lang_code in ("es", "sp"):
        designation_expr = "COALESCE(designation_sp, designation_en, designation, code)"
    else:
        designation_expr = "COALESCE(designation_en, designation, code)"

    meta = {}
    conn = connect_db()
    if conn is None:
        return meta
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        chunk = 200
        clist = list(codes)
        for i in range(0, len(clist), chunk):
            part = clist[i : i + chunk]
            ph = ",".join(["?"] * len(part))

            cur.execute(
                f"""
                SELECT 
                    code, 
                    {designation_expr} as description,
                    type, 
                    remarks
                FROM items_list
                WHERE code IN ({ph})
            """,
                part,
            )
            for r in cur.fetchall():
                meta[r["code"]] = {
                    "description": r["description"] or r["code"],
                    "type": r["type"] or "",
                    "remarks": r["remarks"] or "",
                }
    finally:
        cur.close()
        conn.close()
    return meta


def aggregate_stock_by_treecode(filters, cutoff_iso, id_to_name, name_set):
    """
    Aggregate stock by treecode (in-box) and by code (on-shelf)
    Returns: dict with key = (scenario, key) where key is code for on-shelf, treecode for in-box
    """
    where_parts = ["1=1"]
    params = []

    scen = filters["scenario"]
    all_text = lang.t("stock_summary.all_scenarios", "All")
    if scen and scen != all_text:
        where_parts.append(
            "(scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name=? LIMIT 1))"
        )
        params.extend([scen, scen])

    mm = filters["management_mode"].lower()
    mm_filter = ""
    if mm == "on-shelf":
        mm_filter = "LOWER(management_mode) IN ('on_shelf','on-shelf','onshelf')"
    elif mm == "in-box":
        mm_filter = "LOWER(management_mode) IN ('in_box','in-box','inbox')"

    if filters["kit_number"]:
        where_parts.append("kit_number = ?")
        params.append(filters["kit_number"])

    if filters["module_number"]:
        where_parts.append("module_number = ?")
        params.append(filters["module_number"])

    if filters["item_code"]:
        where_parts.append("item = ?")
        params.append(filters["item_code"])

    where_clause = " AND ".join(where_parts)

    result = {}

    # 1. ON-SHELF items (no treecode, match by item code)
    if mm in ("", "all", "on-shelf"):
        onshelf_where = where_clause
        if mm_filter and mm == "on-shelf":
            onshelf_where += f" AND {mm_filter}"
        else:
            onshelf_where += (
                " AND LOWER(management_mode) IN ('on_shelf','on-shelf','onshelf')"
            )

        onshelf_sql = f"""
            SELECT
               CAST(scenario AS TEXT) AS raw_scenario,
               item AS code,
               SUM(final_qty) AS current_stock,
               MIN(exp_date) AS earliest_expiry,
               GROUP_CONCAT(DISTINCT management_mode) AS management_modes,
               GROUP_CONCAT(comments, '; ') AS comments
            FROM stock_data
            WHERE {onshelf_where}
              AND item IS NOT NULL
              AND item <> ''
              AND final_qty IS NOT NULL
            GROUP BY raw_scenario, item
        """

        onshelf_rows = _fetchall(onshelf_sql, tuple(params))

        for r in onshelf_rows:
            norm_scen = normalize_scenario(r["raw_scenario"], id_to_name, name_set)
            code = r["code"]
            key = (norm_scen, code)  # Use code directly for on-shelf

            result[key] = {
                "current_stock": r["current_stock"] or 0,
                "earliest_expiry": r["earliest_expiry"],
                "kit_code": "",
                "module_code": "",
                "kit_number": "",
                "module_number": "",
                "management_modes": (
                    set(r["management_modes"].split(","))
                    if r["management_modes"]
                    else set()
                ),
                "comments": r["comments"] or "",
                "expiring_qty": 0,
            }

    # 2. IN-BOX items (with treecode)
    if mm in ("", "all", "in-box"):
        inbox_where = where_clause
        if mm_filter and mm == "in-box":
            inbox_where += f" AND {mm_filter}"
        else:
            inbox_where += " AND LOWER(management_mode) IN ('in_box','in-box','inbox')"

        inbox_sql = f"""
            SELECT
               CAST(scenario AS TEXT) AS raw_scenario,
               treecode,
               COALESCE(kit, '') AS kit_code,
               COALESCE(module, '') AS module_code,
               COALESCE(kit_number, '') AS kit_number,
               COALESCE(module_number, '') AS module_number,
               SUM(final_qty) AS current_stock,
               MIN(exp_date) AS earliest_expiry,
               GROUP_CONCAT(DISTINCT management_mode) AS management_modes,
               GROUP_CONCAT(comments, '; ') AS comments
            FROM stock_data
            WHERE {inbox_where}
              AND treecode IS NOT NULL
              AND treecode <> ''
              AND final_qty IS NOT NULL
            GROUP BY raw_scenario, treecode
        """

        inbox_rows = _fetchall(inbox_sql, tuple(params))

        for r in inbox_rows:
            norm_scen = normalize_scenario(r["raw_scenario"], id_to_name, name_set)
            treecode = r["treecode"]
            key = (norm_scen, treecode)  # Use treecode for in-box

            result[key] = {
                "current_stock": r["current_stock"] or 0,
                "earliest_expiry": r["earliest_expiry"],
                "kit_code": r["kit_code"],
                "module_code": r["module_code"],
                "kit_number": r["kit_number"],
                "module_number": r["module_number"],
                "management_modes": (
                    set(r["management_modes"].split(","))
                    if r["management_modes"]
                    else set()
                ),
                "comments": r["comments"] or "",
                "expiring_qty": 0,
            }

    # 3. Expiring quantities
    if cutoff_iso:
        # On-shelf expiring
        if mm in ("", "all", "on-shelf"):
            exp_onshelf_where = where_clause
            if mm_filter and mm == "on-shelf":
                exp_onshelf_where += f" AND {mm_filter}"
            else:
                exp_onshelf_where += (
                    " AND LOWER(management_mode) IN ('on_shelf','on-shelf','onshelf')"
                )

            exp_onshelf_sql = f"""
                SELECT
                  CAST(scenario AS TEXT) AS raw_scenario,
                  item AS code,
                  SUM(final_qty) AS expiring_sum
                FROM stock_data
                WHERE {exp_onshelf_where}
                  AND item IS NOT NULL
                  AND item <> ''
                  AND exp_date IS NOT NULL
                  AND exp_date <= ?
                GROUP BY raw_scenario, item
            """
            exp_params = params + [cutoff_iso]
            exp_onshelf_rows = _fetchall(exp_onshelf_sql, tuple(exp_params))

            for er in exp_onshelf_rows:
                norm_scen = normalize_scenario(er["raw_scenario"], id_to_name, name_set)
                key = (norm_scen, er["code"])
                if key in result:
                    result[key]["expiring_qty"] = er["expiring_sum"] or 0

        # In-box expiring
        if mm in ("", "all", "in-box"):
            exp_inbox_where = where_clause
            if mm_filter and mm == "in-box":
                exp_inbox_where += f" AND {mm_filter}"
            else:
                exp_inbox_where += (
                    " AND LOWER(management_mode) IN ('in_box','in-box','inbox')"
                )

            exp_inbox_sql = f"""
                SELECT
                  CAST(scenario AS TEXT) AS raw_scenario,
                  treecode,
                  SUM(final_qty) AS expiring_sum
                FROM stock_data
                WHERE {exp_inbox_where}
                  AND treecode IS NOT NULL
                  AND treecode <> ''
                  AND exp_date IS NOT NULL
                  AND exp_date <= ?
                GROUP BY raw_scenario, treecode
            """
            exp_params = params + [cutoff_iso]
            exp_inbox_rows = _fetchall(exp_inbox_sql, tuple(exp_params))

            for er in exp_inbox_rows:
                norm_scen = normalize_scenario(er["raw_scenario"], id_to_name, name_set)
                key = (norm_scen, er["treecode"])
                if key in result:
                    result[key]["expiring_qty"] = er["expiring_sum"] or 0

    return result


# ------------------------- Distinct Values ---------------------------
def distinct_kit_numbers(scenario=None):
    if scenario:
        rows = _fetchall(
            """
            SELECT DISTINCT kit_number FROM stock_data
            WHERE kit_number IS NOT NULL
              AND (scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name=? LIMIT 1))
            ORDER BY kit_number
        """,
            (scenario, scenario),
        )
    else:
        rows = _fetchall(
            "SELECT DISTINCT kit_number FROM stock_data WHERE kit_number IS NOT NULL ORDER BY kit_number"
        )
    return [r["kit_number"] for r in rows if r["kit_number"]]


def distinct_module_numbers(scenario=None, kit_number=None):
    where = ["module_number IS NOT NULL"]
    params = []
    if scenario:
        where.append(
            "(scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name=? LIMIT 1))"
        )
        params.extend([scenario, scenario])
    if kit_number:
        where.append("kit_number = ?")
        params.append(kit_number)
    sql = (
        "SELECT DISTINCT module_number FROM stock_data WHERE "
        + " AND ".join(where)
        + " ORDER BY module_number"
    )
    rows = _fetchall(sql, tuple(params))
    return [r["module_number"] for r in rows if r["module_number"]]


# ---------------------------- UI Window ------------------------------
class StockSummaryWindow(tk.Toplevel):
    def __init__(self, parent=None, role="user"):
        super().__init__(parent)
        self.role = role or "user"
        self.title(lang.t("stock_summary.title", "Stock Summary"))
        self.geometry("1600x900")
        self.configure(bg=AppTheme.BG_MAIN)
        self._all_rows = []
        self._build_ui()
        self.populate_scenarios()
        self.populate_kit_module_lists()
        self.status_var.set(
            lang.t("reports.ready", "Ready (role={role})", role=self.role)
        )
        self.after(100, self.load_data)

    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self, bg=AppTheme.BG_MAIN)
        header_frame.pack(fill="x", padx=12, pady=(10, 5))

        tk.Label(
            header_frame,
            text=lang.t("stock_summary.header", "Stock Summary Dashboard"),
            font=(AppTheme.FONT_FAMILY, 20, "bold"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(side="left")

        # Metrics Panel
        metrics_frame = tk.Frame(self, bg=AppTheme.BG_MAIN)
        metrics_frame.pack(fill="x", padx=12, pady=(0, 5))

        self.metric_labels = {}
        metrics_config = [
            ("total_items", lang.t("stock_summary.total_items", "Total"), "#3498DB"),
            ("with_stock", lang.t("stock_summary.with_stock", "In Stock"), "#27AE60"),
            ("missing", lang.t("stock_summary.missing", "Missing"), "#E74C3C"),
            ("overstock", lang.t("stock_summary.overstock", "Overstock"), "#F39C12"),
            ("expiring", lang.t("stock_summary.expiring", "Expiring"), "#E67E22"),
            ("coverage", lang.t("stock_summary.coverage", "Coverage"), "#9B59B6"),
        ]

        for key, label, color in metrics_config:
            card = tk.Frame(metrics_frame, bg=color, relief="raised", bd=1)
            card.pack(side="left", padx=3)

            tk.Label(
                card,
                text=label,
                bg=color,
                fg="white",
                font=(AppTheme.FONT_FAMILY, 8, "bold"),
            ).pack(padx=6, pady=(2, 0))

            val_lbl = tk.Label(
                card,
                text="0",
                bg=color,
                fg="white",
                font=(AppTheme.FONT_FAMILY, 14, "bold"),
            )
            val_lbl.pack(padx=6, pady=(0, 2))
            self.metric_labels[key] = val_lbl

        # Filters
        f = tk.Frame(self, bg=AppTheme.BG_MAIN)
        f.pack(fill="x", padx=10, pady=2)

        tk.Label(
            f,
            text=lang.t("reports.scenario", "Scenario:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=0, sticky="w", padx=4, pady=2)
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(
            f, textvariable=self.scenario_var, width=28, state="readonly"
        )
        self.scenario_cb.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        self.scenario_cb.bind("<<ComboboxSelected>>", self._on_scenario_change)

        tk.Label(
            f,
            text=lang.t("stock_summary.management_mode", "Management Mode:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=2, sticky="w", padx=4, pady=2)
        self.management_var = tk.StringVar(
            value=lang.t("stock_summary.management_all", "All")
        )
        self.management_cb = ttk.Combobox(
            f,
            textvariable=self.management_var,
            values=[
                lang.t("stock_summary.management_all", "All"),
                lang.t("stock_summary.management_on_shelf", "On-Shelf"),
                lang.t("stock_summary.management_in_box", "In-Box"),
            ],
            width=14,
            state="readonly",
        )
        self.management_cb.grid(row=0, column=3, padx=4, pady=2, sticky="w")
        self.management_cb.current(0)
        self.management_cb.bind("<<ComboboxSelected>>", self._on_filter_change)

        tk.Label(
            f,
            text=lang.t("reports.type", "Type:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=4, sticky="w", padx=4, pady=2)
        self.type_var = tk.StringVar(value=lang.t("reports.type_all", "All"))
        self.type_cb = ttk.Combobox(
            f,
            textvariable=self.type_var,
            values=[lang.t("reports.type_all", "All"), "KIT", "MODULE", "ITEM"],
            width=10,
            state="readonly",
        )
        self.type_cb.grid(row=0, column=5, padx=4, pady=2, sticky="w")
        self.type_cb.current(0)
        self.type_cb.bind("<<ComboboxSelected>>", self._on_filter_change)

        tk.Label(
            f,
            text=lang.t("reports.expiry_period", "Expiry Period (months):"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=0, column=6, sticky="w", padx=4, pady=2)
        self.period_override_var = tk.StringVar()
        self.period_entry = tk.Entry(f, textvariable=self.period_override_var, width=7)
        self.period_entry.grid(row=0, column=7, padx=4, pady=2, sticky="w")
        self.period_entry.bind("<KeyRelease>", self._on_filter_change)

        tk.Label(
            f,
            text=lang.t("stock_summary.kit_number", "Kit Number:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=1, column=0, sticky="w", padx=4, pady=2)
        self.kit_number_var = tk.StringVar()
        self.kit_number_cb = ttk.Combobox(
            f, textvariable=self.kit_number_var, width=20, state="readonly"
        )
        self.kit_number_cb.grid(row=1, column=1, padx=4, pady=2, sticky="w")
        self.kit_number_cb.bind("<<ComboboxSelected>>", self._on_kit_change)

        tk.Label(
            f,
            text=lang.t("stock_summary.module_number", "Module Number:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=1, column=2, sticky="w", padx=4, pady=2)
        self.module_number_var = tk.StringVar()
        self.module_number_cb = ttk.Combobox(
            f, textvariable=self.module_number_var, width=20, state="readonly"
        )
        self.module_number_cb.grid(row=1, column=3, padx=4, pady=2, sticky="w")
        self.module_number_cb.bind("<<ComboboxSelected>>", self._on_filter_change)

        tk.Label(
            f,
            text=lang.t("reports.item", "Item Code:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).grid(row=1, column=4, sticky="w", padx=4, pady=2)
        self.item_var = tk.StringVar()
        self.item_entry = tk.Entry(f, textvariable=self.item_var, width=18)
        self.item_entry.grid(row=1, column=5, padx=4, pady=2, sticky="w")
        self.item_entry.bind("<KeyRelease>", self._on_filter_change)

        self.recommended_label = tk.Label(
            self,
            text="",
            font=(AppTheme.FONT_FAMILY, 8),
            fg=AppTheme.COLOR_SECONDARY,
            bg=AppTheme.BG_MAIN,
        )
        self.recommended_label.pack(anchor="w", padx=14, pady=(0, 2))

        # Search
        sframe = tk.Frame(self, bg=AppTheme.BG_MAIN)
        sframe.pack(fill="x", padx=10, pady=(2, 4))
        tk.Label(
            sframe,
            text=lang.t("reports.search", "Search:"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search_change)
        self.search_entry = tk.Entry(sframe, textvariable=self.search_var, width=40)
        self.search_entry.pack(side="left", padx=6)

        # Buttons
        bframe = tk.Frame(self, bg=AppTheme.BG_MAIN)
        bframe.pack(fill="x", padx=10, pady=2)

        tk.Button(
            bframe,
            text=lang.t("reports.load_refresh", "Manual Refresh"),
            bg=AppTheme.BTN_SUCCESS,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            relief="flat",
            command=self.load_data,
        ).pack(side="left", padx=4)

        tk.Button(
            bframe,
            text=lang.t("reports.clear_filters", "Clear Filters"),
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            relief="flat",
            command=self.clear_filters,
        ).pack(side="left", padx=4)

        tk.Button(
            bframe,
            text=lang.t("reports.export", "Export to Excel"),
            bg=AppTheme.BTN_EXPORT,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, AppTheme.FONT_SIZE_NORMAL, "bold"),
            relief="flat",
            command=self.export_excel,
        ).pack(side="left", padx=4)

        # Status
        self.status_var = tk.StringVar(value=lang.t("reports.ready", "Ready"))
        tk.Label(
            self,
            textvariable=self.status_var,
            anchor="w",
            bg=AppTheme.BG_PANEL,
            fg=AppTheme.COLOR_PRIMARY,
            relief="sunken",
        ).pack(fill="x", padx=10, pady=(0, 4))

        self.columns = [
            "scenario",
            "kit_number",
            "module_number",
            "management_modes",
            "code",
            "description",
            "type",
            "standard_qty",
            "current_stock",
            "coverage_pct",
            "qty_expiring",
            "earliest_expiry",
            "over_stock",
            "missing_qty",
            "remarks",
        ]

        headers = {
            "scenario": lang.t("stock_summary.col_scenario", "Scenario"),
            "kit_number": lang.t("stock_summary.col_kit_number", "Kit #"),
            "module_number": lang.t("stock_summary.col_module_number", "Module #"),
            "management_modes": lang.t("stock_summary.col_management", "Mgmt Mode"),
            "code": lang.t("stock_summary.col_code", "Item Code"),
            "description": lang.t("stock_summary.col_description", "Description"),
            "type": lang.t("stock_summary.col_type", "Type"),
            "standard_qty": lang.t("stock_summary.col_std_qty", "Std Qty"),
            "current_stock": lang.t("stock_summary.col_current_stock", "Current Stock"),
            "coverage_pct": lang.t("stock_summary.col_coverage", "Coverage %"),
            "qty_expiring": lang.t("stock_summary.col_qty_expiring", "Expiring Qty"),
            "earliest_expiry": lang.t(
                "stock_summary.col_earliest_expiry", "Next Expiry"
            ),
            "over_stock": lang.t("stock_summary.col_over_stock", "Over Stock"),
            "missing_qty": lang.t("stock_summary.col_missing", "Shortage"),
            "remarks": lang.t("stock_summary.col_remarks", "Remarks"),
        }

        widths = {
            "scenario": 130,
            "kit_number": 100,
            "module_number": 100,
            "management_modes": 100,
            "code": 120,
            "description": 500,
            "type": 60,
            "standard_qty": 80,
            "current_stock": 100,
            "coverage_pct": 90,
            "qty_expiring": 100,
            "earliest_expiry": 100,
            "over_stock": 80,
            "missing_qty": 80,
            "remarks": 300,
        }

        # Create container frame for treeview with scrollbars
        tree_container = tk.Frame(self, bg=AppTheme.BG_MAIN)
        tree_container.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        # Create frame for tree and scrollbars
        tree_frame = tk.Frame(
            tree_container, bg=AppTheme.COLOR_BORDER, bd=2, relief="solid"
        )
        tree_frame.pack(fill="both", expand=True)

        # Create treeview
        self.tree = ttk.Treeview(
            tree_frame, columns=self.columns, show="headings", height=25
        )

        # Configure columns
        for c in self.columns:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths.get(c, 100), anchor="w")

        # Create VISIBLE scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Pack with scrollbars ALWAYS visible
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Configure grid weights
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure tags
        self.tree.tag_configure("missing", background="#FFCCCC")
        self.tree.tag_configure("overstock", background="#CCFFCC")
        self.tree.tag_configure(
            "group",
            background="#E8E8E8",
            foreground="#8B0000",
            font=(AppTheme.FONT_FAMILY, 10, "bold"),
        )

        # Event bindings - WORKING solution for double-click
        self.tree.bind("<Button-3>", self._on_right_click)

        # Double-click workaround (works on all systems)
        self._last_click_time = 0
        self._last_click_item = None

        def on_single_click(event):
            import time

            current_time = time.time()
            item_id = self.tree.identify_row(event.y)

            # Check if it's a double-click (within 500ms)
            if (
                current_time - self._last_click_time < 0.5
                and item_id == self._last_click_item
                and item_id
            ):
                # It's a double-click!
                self._on_double_click(event)
                self._last_click_time = 0  # Reset
                self._last_click_item = None
            else:
                # Single click
                self._last_click_time = current_time
                self._last_click_item = item_id

        self.tree.bind("<ButtonRelease-1>", on_single_click)

        # Context menu
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(
            label=lang.t("stock_summary.show_stock_card", "Show Stock Card"),
            command=self._show_stock_card_from_menu,
        )
        self.context_menu.add_separator()
        self.context_menu.add_command(
            label=lang.t("stock_summary.stock_details", "Stock Details"),
            command=self._show_stock_details,
        )

    def _on_double_click(self, event):
        """Open stock details on double-click"""
        try:
            item_id = self.tree.identify_row(event.y)

            if not item_id:
                print("[DEBUG] No item identified")
                return

            values = self.tree.item(item_id, "values")
            print(f"[DEBUG] Values: {values}")

            if not values or len(values) < 6:
                print("[DEBUG] Not enough values")
                return

            # Check if it's a group header
            desc_value = str(values[5])
            if desc_value.startswith("━"):
                print("[DEBUG] Group header detected, skipping")
                return

            # Get item code
            item_code = str(values[4]).strip()
            if not item_code:
                print("[DEBUG] No item code found")
                return

            print(f"[DEBUG] Opening stock details for: {item_code}")

            # Store values and call stock details
            self._selected_item_values = values
            self._show_stock_details()

        except Exception as e:
            print(f"[ERROR] Double-click failed: {e}")
            import traceback

            traceback.print_exc()

    def _on_right_click(self, event):
        """Show context menu on right-click"""
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        values = self.tree.item(item_id, "values")
        if not values or len(values) < 6:  # CHANGED from 7 to 6
            return

        # FIX: Check description column (index 5, not 6!)
        desc_value = str(values[5])  # CHANGED from values[6] to values[5]
        if desc_value.startswith("━"):
            return

        self.tree.selection_set(item_id)
        self._selected_item_values = values

        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    def _show_stock_card_from_menu(self):
        """Open stock card from context menu"""
        if hasattr(self, "_selected_item_values"):
            self._open_stock_card(self._selected_item_values)

    def _show_stock_details(self):
        """Show detailed stock breakdown with all expiry dates"""
        if not hasattr(self, "_selected_item_values"):
            return

        values = self._selected_item_values
        scenario = str(values[0]).strip()
        item_code = str(values[4]).strip()
        mgmt_mode = str(values[3]).strip()
        treecode_val = None

        # Get treecode from the row data
        for r in self._all_rows:
            if r["code"] == item_code and r["scenario"] == scenario:
                treecode_val = r.get("treecode")
                break

        if not item_code:
            return

        # Query stock_data for detailed breakdown
        id_to_name, name_set = load_scenario_maps()

        # Build query based on management mode
        if "on-shelf" in mgmt_mode.lower() or "on_shelf" in mgmt_mode.lower():
            # On-shelf: match by scenario + item
            sql = """
                SELECT 
                    unique_id,
                    scenario,
                    item AS code,
                    final_qty,
                    exp_date,
                    management_mode,
                    kit_number,
                    module_number,
                    comments
                FROM stock_data
                WHERE (scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name = ?))
                  AND item = ?
                  AND LOWER(management_mode) IN ('on_shelf', 'on-shelf', 'onshelf')
                  AND final_qty > 0
                ORDER BY exp_date
            """
            rows = _fetchall(sql, (scenario, scenario, item_code))
        else:
            # In-box: match by treecode
            if treecode_val:
                sql = """
                    SELECT 
                        unique_id,
                        scenario,
                        item AS code,
                        final_qty,
                        exp_date,
                        management_mode,
                        kit_number,
                        module_number,
                        kit,
                        module,
                        comments
                    FROM stock_data
                    WHERE (scenario = ? OR scenario = (SELECT CAST(scenario_id AS TEXT) FROM scenarios WHERE name = ?))
                      AND treecode = ?
                      AND final_qty > 0
                    ORDER BY exp_date, kit_number, module_number
                """
                rows = _fetchall(sql, (scenario, scenario, treecode_val))
            else:
                rows = []

        # Create popup window
        popup = tk.Toplevel(self)
        popup.title(
            lang.t("stock_summary.stock_details_title", "Stock Details")
            + f" - {item_code}"
        )
        popup.geometry("1000x600")
        popup.configure(bg=AppTheme.BG_MAIN)

        # Header
        header_frame = tk.Frame(popup, bg=AppTheme.BG_MAIN)
        header_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(
            header_frame,
            text=lang.t(
                "stock_summary.stock_breakdown",
                "Stock Breakdown: {code}",
                code=item_code,
            ),
            font=(AppTheme.FONT_FAMILY, 16, "bold"),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(side="left")

        tk.Label(
            header_frame,
            text=lang.t(
                "stock_summary.scenario_mode",
                "Scenario: {scenario} | Mode: {mode}",
                scenario=scenario,
                mode=mgmt_mode,
            ),
            font=(AppTheme.FONT_FAMILY, 10),
            bg=AppTheme.BG_MAIN,
            fg=AppTheme.COLOR_SECONDARY,
        ).pack(side="left", padx=20)

        # Summary
        summary_frame = tk.Frame(popup, bg=AppTheme.BG_PANEL, relief="raised", bd=2)
        summary_frame.pack(fill="x", padx=10, pady=5)

        total_qty = sum(r["final_qty"] for r in rows)
        tk.Label(
            summary_frame,
            text=lang.t(
                "stock_summary.total_lines_qty",
                "Total Lines: {lines} | Total Quantity: {qty}",
                lines=len(rows),
                qty=total_qty,
            ),
            font=(AppTheme.FONT_FAMILY, 12, "bold"),
            bg=AppTheme.BG_PANEL,
            fg=AppTheme.COLOR_PRIMARY,
        ).pack(pady=5)

        # Create treeview for details
        tree_frame = tk.Frame(popup)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Columns
        columns = ["kit_number", "module_number", "quantity", "exp_date", "comments"]
        detail_tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", height=20
        )

        # Configure columns with translations
        detail_tree.heading(
            "kit_number", text=lang.t("stock_summary.col_kit_num", "Kit #")
        )
        detail_tree.heading(
            "module_number", text=lang.t("stock_summary.col_module_num", "Module #")
        )
        detail_tree.heading(
            "quantity", text=lang.t("stock_summary.col_quantity", "Quantity")
        )
        detail_tree.heading(
            "exp_date", text=lang.t("stock_summary.col_exp_date", "Expiry Date")
        )
        detail_tree.heading(
            "comments", text=lang.t("stock_summary.col_comments", "Comments")
        )

        detail_tree.column("kit_number", width=120)
        detail_tree.column("module_number", width=120)
        detail_tree.column("quantity", width=100)
        detail_tree.column("exp_date", width=120)
        detail_tree.column("comments", width=400)

        # Scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=detail_tree.yview)
        detail_tree.configure(yscrollcommand=vsb.set)

        detail_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Populate data
        if len(rows) == 0:
            # No data found
            tk.Label(
                tree_frame,
                text=lang.t(
                    "stock_summary.no_data_found", "No stock data found for this item"
                ),
                font=(AppTheme.FONT_FAMILY, 12),
                bg=AppTheme.BG_MAIN,
                fg=AppTheme.COLOR_SECONDARY,
            ).pack(pady=50)
        else:
            for r in rows:
                detail_tree.insert(
                    "",
                    "end",
                    values=(
                        r["kit_number"] if r["kit_number"] else "",
                        r["module_number"] if r["module_number"] else "",
                        r["final_qty"],
                        r["exp_date"] if r["exp_date"] else "",
                        r["comments"] if r["comments"] else "",
                    ),
                )

        # Close button
        tk.Button(
            popup,
            text="Close",
            command=popup.destroy,
            bg=AppTheme.BTN_NEUTRAL,
            fg=AppTheme.TEXT_WHITE,
            font=(AppTheme.FONT_FAMILY, 10, "bold"),
        ).pack(pady=10)

    def _open_stock_card(self, values):
        """Open stock card for selected item"""
        try:
            item_code = str(values[4]).strip()
            print(f"[DEBUG] _open_stock_card called with code: {item_code}")

            if not item_code:
                print("[DEBUG] Empty item code")
                return

            from stock_card import StockCard

            win = tk.Toplevel(self)
            win.title(lang.t("stock_card.title", "Stock Card") + f" - {item_code}")
            win.geometry("1400x800")
            win.configure(bg="#F5F5F5")

            # Create StockCard
            card = StockCard(win, win, role=self.role)
            card.pack(fill="both", expand=True)

            # Simple approach: just set the code and trigger load
            def load_data():
                try:
                    if hasattr(card, "code_entry"):
                        card.code_entry.delete(0, tk.END)
                        card.code_entry.insert(0, item_code)

                    if hasattr(card, "search_items"):
                        card.search_items()

                        # Auto-select first result after a delay
                        def auto_select():
                            try:
                                if (
                                    hasattr(card, "search_listbox")
                                    and card.search_listbox.size() > 0
                                ):
                                    card.search_listbox.selection_clear(0, tk.END)
                                    card.search_listbox.selection_set(0)
                                    card.search_listbox.event_generate(
                                        "<<ListboxSelect>>"
                                    )
                            except:
                                pass

                        win.after(300, auto_select)

                except Exception as e:
                    print(f"[ERROR] Load data failed: {e}")

            win.after(100, load_data)

        except ImportError as e:
            print(f"[ERROR] Import error: {e}")
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                "stock_card.py module not found.",
                "error",
            )
        except Exception as e:
            print(f"[ERROR] Stock card error: {e}")
            import traceback

            traceback.print_exc()
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                f"Error opening stock card: {str(e)}",
                "error",
            )

        # Auto-refresh handlers

    def _on_filter_change(self, event=None):
        if hasattr(self, "_refresh_timer"):
            self.after_cancel(self._refresh_timer)
        self._refresh_timer = self.after(500, self.load_data)

    # Scenario & Distinct lists

    def populate_scenarios(self):
        rows = _fetchall("SELECT name FROM scenarios ORDER BY name")
        vals = [r["name"] for r in rows]
        # Add "All" option at the beginning
        all_text = lang.t("stock_summary.all_scenarios", "All")
        self.scenario_cb["values"] = [all_text] + vals
        if vals:
            self.scenario_cb.current(0)  # Select "All" by default
        self._update_recommended_label()

    def populate_kit_module_lists(self):
        """Populate kit numbers based on selected scenario"""
        scen = self.scenario_var.get() or None
        # Don't filter by "All" scenario
        all_text = lang.t("stock_summary.all_scenarios", "All")
        if scen == all_text or not scen:
            scen = None

        kits = distinct_kit_numbers(scen)
        self.kit_number_cb["values"] = [""] + kits
        self.kit_number_var.set("")
        self.populate_module_numbers()

    def populate_module_numbers(self):
        """Populate module numbers based on selected scenario and kit"""
        scen = self.scenario_var.get() or None
        # Don't filter by "All" scenario
        all_text = lang.t("stock_summary.all_scenarios", "All")
        if scen == all_text or not scen:
            scen = None

        kitn = self.kit_number_var.get() or None
        # Don't filter by empty kit number
        if kitn == "":
            kitn = None

        modules = distinct_module_numbers(scen, kitn)
        self.module_number_cb["values"] = [""] + modules  # ✅ FIXED!
        self.module_number_var.set("")

    def _on_scenario_change(self, event=None):
        """Handle scenario change - refresh kit/module lists and data"""
        # Clear kit and module selections
        self.kit_number_var.set("")
        self.module_number_var.set("")
        # Refresh kit and module lists when scenario changes
        self.populate_kit_module_lists()
        # Trigger data refresh
        self._on_filter_change()

    def _on_kit_change(self, event=None):
        """Handle kit number change - refresh module list and data"""
        # Clear module selection when kit changes
        self.module_number_var.set("")
        # Repopulate modules for the selected kit
        self.populate_module_numbers()
        # Trigger data refresh
        self._on_filter_change()

    def gather_filters(self):
        msel = self.management_var.get()
        if msel in ("", lang.t("stock_summary.management_all", "All")):
            management_mode = "All"
        elif msel == lang.t("stock_summary.management_on_shelf", "On-Shelf"):
            management_mode = "On-Shelf"
        elif msel == lang.t("stock_summary.management_in_box", "In-Box"):
            management_mode = "In-Box"
        else:
            management_mode = msel

        tsel = self.type_var.get()
        if tsel in ("", lang.t("reports.type_all", "All")):
            type_filter = "All"
        else:
            type_filter = tsel

        # Handle "All Scenarios" option
        scenario = self.scenario_var.get() or None
        all_text = lang.t("stock_summary.all_scenarios", "All Scenarios")
        if scenario == all_text:
            scenario = None  # Load all scenarios

        return {
            "scenario": scenario,
            "management_mode": management_mode,
            "kit_number": self.kit_number_var.get() or None,
            "module_number": self.module_number_var.get() or None,
            "item_code": self.item_var.get().strip() or None,
            "type_filter": type_filter,
        }

    def clear_filters(self):
        self.scenario_var.set("")
        self.management_var.set(lang.t("stock_summary.management_all", "All"))
        self.kit_number_var.set("")
        self.module_number_var.set("")
        self.item_var.set("")
        self.type_var.set(lang.t("reports.type_all", "All"))
        self.period_override_var.set("")
        self.search_var.set("")
        self.tree.delete(*self.tree.get_children())
        self._all_rows.clear()
        self._update_recommended_label()
        self._update_metrics({})
        self.status_var.set(
            lang.t(
                "reports.filters_cleared",
                "Filters cleared. Ready (role={role}).",
                role=self.role,
            )
        )
        self.after(100, self.load_data)

    def _update_recommended_label(self):
        rec = compute_recommended_months()
        self.recommended_label.config(
            text=lang.t(
                "stock_summary.recommended",
                "Recommended expiry period: {months} month(s) (lead+cover+buffer)",
                months=rec,
            )
        )

    # Replace load_data() method:
    def load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._all_rows.clear()
        filters = self.gather_filters()

        all_text = lang.t("stock_summary.all_scenarios", "All")
        loading_all = not filters["scenario"] or filters["scenario"] == all_text

        if loading_all:
            self.status_var.set(
                lang.t("stock_summary.loading_all", "Loading all scenarios...")
            )

        recommended = compute_recommended_months()
        raw = self.period_override_var.get().strip()
        if raw == "":
            months = recommended
        else:
            try:
                v = int(raw)
                months = v if v >= 0 else recommended
            except:
                months = recommended

        cutoff_iso = cutoff_date_from_months(months)
        exp_header = lang.t(
            "stock_summary.qty_expiring_header",
            "Expiring in next {months} months",
            months=months,
        )
        self.tree.heading("qty_expiring", text=exp_header)

        self.status_var.set(lang.t("reports.computing", "Computing..."))
        self.update_idletasks()

        id_to_name, name_set = load_scenario_maps()

        std_data_by_scenario = load_std_quantities_by_scenario(
            None if loading_all else filters["scenario"]
        )

        # NEW: Use treecode-based aggregation
        stock_map = aggregate_stock_by_treecode(
            filters, cutoff_iso, id_to_name, name_set
        )

        if not std_data_by_scenario and not stock_map:
            self.status_var.set(
                lang.t("stock_summary.no_data", "No data for selected filters.")
            )
            self._update_metrics({})
            return

        # Collect all codes for metadata lookup
        all_codes = set()
        for scenario_data in std_data_by_scenario.values():
            for item_data in scenario_data.values():
                all_codes.add(item_data["code"])

        meta = load_item_metadata(all_codes)
        type_filter = filters["type_filter"].upper()

        # Build rows - WITH FILTER SUPPORT!
        for scenario_name, std_data in sorted(std_data_by_scenario.items()):
            # Sort by treecode for proper display order
            for treecode in sorted(std_data.keys()):
                std_info = std_data[treecode]
                code = std_info["code"]
                std_qty = std_info["std_qty"]
                mgmt_type = std_info["mgmt_type"]
                kit_code = std_info["kit_code"]
                module_code = std_info["module_code"]

                # Get stock data by treecode
                stock_key = (scenario_name, treecode)
                stock_entry = stock_map.get(
                    stock_key,
                    {
                        "current_stock": 0,
                        "expiring_qty": 0,
                        "earliest_expiry": None,
                        "kit_code": "",
                        "module_code": "",
                        "kit_number": "",
                        "module_number": "",
                        "management_modes": set(),
                        "comments": "",
                    },
                )

                # ✅ FILTER FIX: Skip items that don't match kit/module filters
                if filters["kit_number"]:
                    # If kit filter is active, only show items that have stock with this kit
                    if (
                        not stock_entry["kit_number"]
                        or stock_entry["kit_number"] != filters["kit_number"]
                    ):
                        # No stock with this kit number - skip this item
                        continue

                if filters["module_number"]:
                    # If module filter is active, only show items that have stock with this module
                    if (
                        not stock_entry["module_number"]
                        or stock_entry["module_number"] != filters["module_number"]
                    ):
                        # No stock with this module number - skip this item
                        continue

                m = meta.get(code, {"description": code, "type": "", "remarks": ""})
                ctype = m["type"]

                # Type filter
                if (
                    type_filter in ("KIT", "MODULE", "ITEM")
                    and ctype.upper() != type_filter
                ):
                    continue

                current_stock = stock_entry["current_stock"]
                expiring_qty = stock_entry["expiring_qty"]
                over_stock = max(0, current_stock - std_qty)
                missing_qty = max(
                    0, std_qty - current_stock + min(expiring_qty, std_qty)
                )

                coverage_pct = ""
                if std_qty > 0:
                    available_stock = current_stock - expiring_qty
                    coverage_pct = round((available_stock * 100.0) / std_qty, 1)

                earliest = stock_entry.get("earliest_expiry") or ""
                if earliest and not re.match(r"^\d{4}-\d{2}-\d{2}$", earliest):
                    earliest = ""

                modes = (
                    ", ".join(sorted(list(stock_entry["management_modes"])))
                    if stock_entry["management_modes"]
                    else mgmt_type.title()
                )

                # Use kit/module info from stock_entry if available, otherwise from std_info
                display_kit_code = stock_entry.get("kit_code") or kit_code
                display_module_code = stock_entry.get("module_code") or module_code

                self._all_rows.append(
                    {
                        "scenario": scenario_name,
                        "kit_code": display_kit_code,
                        "module_code": display_module_code,
                        "kit_number": stock_entry.get("kit_number", ""),
                        "module_number": stock_entry.get("module_number", ""),
                        "management_modes": modes,
                        "treecode": treecode,
                        "code": code,
                        "description": m["description"],
                        "type": ctype,
                        "standard_qty": std_qty,
                        "current_stock": current_stock,
                        "coverage_pct": coverage_pct,
                        "qty_expiring": expiring_qty,
                        "earliest_expiry": earliest,
                        "over_stock": over_stock,
                        "missing_qty": missing_qty,
                        "remarks": stock_entry.get("comments", "") or m["remarks"],
                    }
                )

        self._render_rows(self._all_rows)
        enable_column_auto_resize(self.tree)

        metrics = self._compute_metrics(self._all_rows)
        self._update_metrics(metrics)

        self.status_var.set(
            lang.t(
                "stock_summary.loaded_status",
                "Loaded {rows} items | Expiry: {m} months (cutoff {cutoff}) | Role: {role}",
                rows=len(self._all_rows),
                m=months,
                cutoff=cutoff_iso or "N/A",
                role=self.role,
            )
        )
        self._update_recommended_label()

    def _render_rows(self, rows):
        """Render rows with proper grouping headers - only for kit/module, with spacing"""
        self.tree.delete(*self.tree.get_children())
        last_group = (None, None, None)  # scenario, kit_code, module_code

        for r in rows:
            # Get kit/module CODES for grouping (not numbers)
            kit_code = r.get("kit_code", "") or ""
            module_code = r.get("module_code", "") or ""

            # Current group tuple (ignore mgmt_mode)
            current_group = (r["scenario"], kit_code, module_code)

            # Check if we need spacing or headers
            if current_group != last_group:
                # Add blank row when scenario or kit changes
                if (
                    current_group[0] != last_group[0]  # Scenario changed
                    or current_group[1] != last_group[1]
                ):  # Kit changed

                    # Insert blank row for spacing
                    self.tree.insert(
                        "",
                        "end",
                        values=(
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ),
                        tags=(),
                    )

                # Build header only for kit/module changes (not for on-shelf/in-box)
                parts = []

                # Scenario changed - show scenario name
                if current_group[0] != last_group[0]:
                    parts.append(f"{current_group[0]}")

                # Kit/Module changed (for In-Box items only)
                if kit_code:  # Only if there's a kit
                    # Kit changed
                    if kit_code != last_group[1]:
                        parts.append(f"Kit: {kit_code}")
                        # Show first module of the kit
                        if module_code:
                            parts.append(f"Module: {module_code}")
                    # Only module changed (same kit)
                    elif module_code and module_code != last_group[2]:
                        parts.append(f"Module: {module_code}")

                # Insert header if there's something to show
                if parts:
                    header_text = "━━ " + " → ".join(parts) + " ━━"

                    self.tree.insert(
                        "",
                        "end",
                        values=(
                            r["scenario"],
                            r["kit_number"],
                            r["module_number"],
                            "",
                            "",
                            header_text,
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ),
                        tags=("group",),
                    )

                last_group = current_group

            # Insert the actual data row
            tags = []
            if r["missing_qty"] > 0:
                tags.append("missing")
            elif r["over_stock"] > 0:
                tags.append("overstock")

            self.tree.insert(
                "",
                "end",
                values=(
                    r["scenario"],
                    r["kit_number"],
                    r["module_number"],
                    r["management_modes"],
                    r["code"],
                    r["description"],
                    r["type"],
                    r["standard_qty"],
                    r["current_stock"],
                    r["coverage_pct"],
                    r["qty_expiring"],
                    r["earliest_expiry"],
                    r["over_stock"],
                    r["missing_qty"],
                    r["remarks"],
                ),
                tags=tuple(tags),
            )

    def _compute_metrics(self, rows):
        total_items = 0
        with_stock = 0
        missing = 0
        overstock = 0
        expiring_count = 0  # COUNT of rows with expiring qty > 0
        total_coverage = 0
        items_with_std_qty = 0

        for r in rows:
            total_items += 1

            # FIXED: Count only items with available stock (after removing expiring)
            current = r.get("current_stock", 0) or 0
            expiring = r.get("qty_expiring", 0) or 0
            available_stock = current - expiring
            if available_stock > 0:
                with_stock += 1

            if r["missing_qty"] > 0:
                missing += 1
            if r["over_stock"] > 0:
                overstock += 1

            # COUNT rows with expiring quantities (not sum)
            if expiring > 0:
                expiring_count += 1  # Just count this row

            # Calculate coverage
            std_qty = r.get("standard_qty", 0) or 0
            if std_qty > 0:
                items_with_std_qty += 1
                coverage_pct = min((available_stock / std_qty) * 100, 100)
                total_coverage += coverage_pct

        # Calculate average coverage
        avg_coverage = (
            round(total_coverage / items_with_std_qty) if items_with_std_qty > 0 else 0
        )

        return {
            "total_items": total_items,
            "with_stock": with_stock,  # Now counts only available stock (current - expiring)
            "missing": missing,
            "overstock": overstock,
            "expiring": expiring_count,
            "coverage": avg_coverage,
        }

    def _update_metrics(self, m):
        if not m:
            for k in self.metric_labels:
                self.metric_labels[k].config(text="0")
            return

        self.metric_labels["total_items"].config(text=str(m.get("total_items", 0)))
        self.metric_labels["with_stock"].config(text=str(m.get("with_stock", 0)))
        self.metric_labels["missing"].config(text=str(m.get("missing", 0)))
        self.metric_labels["overstock"].config(text=str(m.get("overstock", 0)))
        self.metric_labels["expiring"].config(text=str(m.get("expiring", 0)))
        self.metric_labels["coverage"].config(text=f"{m.get('coverage', 0)}%")

    def _on_search_change(self, *_):
        q = self.search_var.get().strip().lower()
        if not q:
            self._render_rows(self._all_rows)
            metrics = self._compute_metrics(self._all_rows)
            self._update_metrics(metrics)
            return
        filtered = []
        for r in self._all_rows:
            fields = [
                r["scenario"],
                r["kit_number"],
                r["module_number"],
                r["management_modes"],
                r["treecode"],
                r["code"],
                r["description"],
                r["type"],
                r["remarks"],
            ]
            combined = " ".join([str(x).lower() for x in fields if x])
            if q in combined:
                filtered.append(r)
        self._render_rows(filtered)
        metrics = self._compute_metrics(filtered)
        self._update_metrics(metrics)

    def export_excel(self):
        items = self.tree.get_children("")
        if not items:
            custom_popup(
                self,
                lang.t("reports.info", "Info"),
                lang.t("reports.nothing_to_export", "Nothing to export."),
                "info",
            )
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"stock_summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            self.status_var.set(lang.t("reports.export_cancelled", "Export cancelled"))
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock_Summary"
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws["A1"] = (
                f"{lang.t('reports.generated','Generated')}: {now} (role={self.role})"
            )
            ws["A1"].font = Font(size=10, bold=True)
            ws.append([])

            headers = [self.tree.heading(c)["text"] for c in self.columns]
            ws.append(headers)
            header_row = ws[ws.max_row]
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="D9E1EC", end_color="D9E1EC", fill_type="solid"
                )

            missing_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            over_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            group_fill = PatternFill(
                start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
            )

            for iid in items:
                vals = self.tree.item(iid, "values")
                tags = set(self.tree.item(iid, "tags") or [])
                ws.append(list(vals))
                rcells = ws[ws.max_row]

                if "group" in tags:
                    for c in rcells:
                        c.fill = group_fill
                        c.font = Font(bold=True)
                elif "missing" in tags:
                    for c in rcells:
                        c.fill = missing_fill
                elif "overstock" in tags:
                    for c in rcells:
                        c.fill = over_fill

            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[letter].width = min(max_len + 2, 55)

            ws.freeze_panes = "A4"
            wb.save(path)

            custom_popup(
                self,
                lang.t("reports.success", "Success"),
                lang.t("reports.export_success", "Export successful!"),
                "success",
            )
            self.status_var.set(lang.t("reports.export_success", "Export successful!"))
        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", "Error"),
                lang.t("reports.export_failed", "Export failed: {err}", err=str(e)),
                "error",
            )
            self.status_var.set(
                lang.t("reports.export_failed", "Export failed: {err}", err=str(e))
            )


# Launcher
def open_stock_summary(parent=None, role="user"):
    win = StockSummaryWindow(parent=parent, role=role)
    win.focus()
    return win


# Standalone Run
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stock Summary")
    root.geometry("1600x900")
    root.withdraw()
    open_stock_summary(root, role="admin")
    root.mainloop()
