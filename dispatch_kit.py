import tkinter as tk
from tkinter import ttk, filedialog
import sqlite3
import logging
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from popup_utils import custom_popup, custom_askyesno, custom_dialog
from db import connect_db
from manage_items import get_item_description, detect_type
from language_manager import lang

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


# ------------------------- DB HELPERS -------------------------
def fetch_end_users():
    conn = connect_db()
    if conn is None:
        logging.error("[DISPATCH] DB connection failed (fetch_end_users)")
        return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM end_users ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error as e:
        logging.error(f"[DISPATCH] fetch_end_users error: {e}")
        return []
    finally:
        cur.close()
        conn.close()


def fetch_third_parties():
    conn = connect_db()
    if conn is None:
        logging.error("[DISPATCH] DB connection failed (fetch_third_parties)")
        return []
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM third_parties ORDER BY name")
        return [r[0] for r in cur.fetchall()]
    except sqlite3.Error as e:
        logging.error(f"[DISPATCH] fetch_third_parties error: {e}")
        return []
    finally:
        cur.close()
        conn.close()


def fetch_project_details():
    conn = connect_db()
    if conn is None:
        logging.error("[DISPATCH] DB connection failed (fetch_project_details)")
        return lang.t("dispatch_kit.unknown_project", "Unknown Project"), lang.t(
            "dispatch_kit.unknown_code", "Unknown Code"
        )
    cur = conn.cursor()
    try:
        cur.execute("SELECT project_name, project_code FROM project_details LIMIT 1")
        row = cur.fetchone()
        return (
            (
                row[0]
                if row and row[0]
                else lang.t("dispatch_kit.unknown_project", "Unknown Project")
            ),
            (
                row[1]
                if row and row[1]
                else lang.t("dispatch_kit.unknown_code", "Unknown Code")
            ),
        )
    except sqlite3.Error as e:
        logging.error(f"[DISPATCH] fetch_project_details error: {e}")
        return lang.t("dispatch_kit.unknown_project", "Unknown Project"), lang.t(
            "dispatch_kit.unknown_code", "Unknown Code"
        )
    finally:
        cur.close()
        conn.close()


def log_transaction(
    unique_id,
    code,
    description,
    expiry_date,
    batch_number,
    scenario,
    Kit,
    module,
    qty_out,
    out_type,
    third_party,
    end_user,
    remarks,
    movement_type,
):
    conn = connect_db()
    if conn is None:
        raise ValueError("DB connection failed")
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO stock_transactions
            (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
             Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
             Third_Party, End_User, Remarks, Movement_Type)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
            (
                datetime.today().strftime("%Y-%m-%d"),
                datetime.now().strftime("%H:%M:%S"),
                unique_id,
                code,
                description,
                expiry_date,
                batch_number,
                scenario,
                Kit,
                module,
                None,
                None,
                qty_out,
                out_type,
                third_party,
                end_user,
                remarks,
                movement_type,
            ),
        )
        conn.commit()
    except sqlite3.Error as e:
        conn.rollback()
        logging.error(f"[DISPATCH] Transaction log error: {e}")
        raise
    finally:
        cur.close()
        conn.close()


def configure_db_pragmas():
    conn = connect_db()
    if conn is None:
        return
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA busy_timeout=5000;")
        conn.commit()
    except Exception:
        pass
    finally:
        try:
            conn.close()
        except:
            pass


# =============================================================
#                        MAIN CLASS
# =============================================================
class StockDispatchKit(tk.Frame):
    """
    Dispatch (Issue Out) Module
    Modes:
      - dispatch_kit
      - issue_standalone
      - issue_module_scenario
      - issue_module_kit
      - issue_items_kit
      - issue_items_module
    """

    def __init__(self, parent, app, role="supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        self.role = role.lower()
        try:
            configure_db_pragmas()
        except Exception:
            pass

        self.scenario_map = self.fetch_scenario_map()
        self.selected_scenario_id = None
        self.selected_scenario_name = None

        # UI variable references
        self.scenario_var = None
        self.scenario_cb = None
        self.mode_var = None
        self.mode_cb = None
        self.Kit_var = None
        self.Kit_cb = None
        self.Kit_number_var = None
        self.Kit_number_cb = None
        self.module_var = None
        self.module_cb = None
        self.module_number_var = None
        self.module_number_cb = None
        self.trans_type_var = None
        self.trans_type_cb = None
        self.end_user_var = None
        self.end_user_cb = None
        self.third_party_var = None
        self.third_party_cb = None
        self.remarks_entry = None
        self.search_var = None
        self.search_entry = None
        self.search_listbox = None
        self.tree = None
        self.status_var = None
        self.editing_cell = None

        # Data caches
        self.row_data = {}
        self.full_items = []
        self.search_min_chars = 2

        if self.parent and self.parent.winfo_exists():
            self.pack(fill="both", expand=True)
            self.after(50, self.render_ui)

    # ---------------------------------------------------------
    # Helpers for "All" label and out-type mapping
    # ---------------------------------------------------------
    def _all_label(self):
        return lang.t("dispatch_kit.all", "All")

    def _norm_all(self, val):
        all_lbl = self._all_label()
        return "All" if (val is None or val == "" or val == all_lbl) else val

    def _out_type_options(self):
        """
        Returns list of (value, display_label) tuples.
        Value stays English (used in DB / calculations); display is translated.
        """
        opts = [
            (
                "Issue to End User",
                lang.t("dispatch_kit.out_issue_end_user", "Issue to End User"),
            ),
            ("Expired Items", lang.t("dispatch_kit.out_expired", "Expired Items")),
            ("Damaged Items", lang.t("dispatch_kit.out_damaged", "Damaged Items")),
            (
                "Cold Chain Break",
                lang.t("dispatch_kit.out_cold_chain", "Cold Chain Break"),
            ),
            ("Batch Recall", lang.t("dispatch_kit.out_batch_recall", "Batch Recall")),
            ("Theft", lang.t("dispatch_kit.out_theft", "Theft")),
            ("Other Losses", lang.t("dispatch_kit.out_other_losses", "Other Losses")),
            ("Out Donation", lang.t("dispatch_kit.out_donation", "Out Donation")),
            ("Loan", lang.t("dispatch_kit.out_loan", "Loan")),
            (
                "Return of Borrowing",
                lang.t("dispatch_kit.out_return_borrowing", "Return of Borrowing"),
            ),
            ("Quarantine", lang.t("dispatch_kit.out_quarantine", "Quarantine")),
        ]
        return opts

    def _display_for_out_type(self, value):
        for v, lbl in self._out_type_options():
            if v == value:
                return lbl
        return value

    def _value_for_out_type(self, display):
        for v, lbl in self._out_type_options():
            if lbl == display or v == display:
                return v
        return display

    # --------------Canonical Movement Type Mapping -----------------
    def _canon_movement_type(self, display_label: str) -> str:
        """
        Convert any localized movement type display label to canonical English.

        Args:
            display_label:  The label shown in the dropdown (could be FR/ES/EN)

        Returns:
            Canonical English movement type name for database storage
        """
        # Get the internal key from the display label
        internal_key = self.mode_label_to_key.get(display_label)

        if not internal_key:
            # Fallback:  if not found, return as-is (shouldn't happen)
            logging.warning(
                f"[DISPATCH] Unknown movement type label:   {display_label}"
            )
            return display_label

        # Map internal keys to canonical English display names
        canon_map = {
            "dispatch_kit": "Dispatch Kit",
            "issue_standalone": "Issue standalone items",
            "issue_module_scenario": "Issue module from scenario",
            "issue_module_kit": "Issue module from Kit",
            "issue_items_kit": "Issue items from Kit",
            "issue_items_module": "Issue items from module",
        }

        canonical = canon_map.get(internal_key, internal_key)

        logging.debug(
            f"[DISPATCH] Movement type:  '{display_label}' → internal:  '{internal_key}' → canonical: '{canonical}'"
        )

        return canonical

    def _display_for_movement_type(self, canonical_value: str) -> str:
        """
        Convert canonical English movement type to current language display label.
        Used when reading from database for display.

        Args:
            canonical_value: English movement type from database

        Returns:
            Localized display label for current language
        """
        # Reverse map:  canonical English → internal key
        reverse_canon_map = {
            "Dispatch Kit": "dispatch_kit",
            "Issue standalone items": "issue_standalone",
            "Issue module from scenario": "issue_module_scenario",
            "Issue module from Kit": "issue_module_kit",
            "Issue items from Kit": "issue_items_kit",
            "Issue items from module": "issue_items_module",
        }

        internal_key = reverse_canon_map.get(canonical_value, "dispatch_kit")

        # Find the localized label for this key
        for label, key in self.mode_label_to_key.items():
            if key == internal_key:
                return label

        # Fallback to canonical if not found
        return canonical_value

    # ------------Helper to display code- discreption-------#
    def _extract_code_from_display(self, display_string: str) -> str:
        """
        Extract code from "CODE - Description" format.
        Handles prefixes like "● CODE - Description" or "[S] CODE - Description"

        Args:
            display_string: Either "CODE" or "CODE - Description" or "● CODE - Description"

        Returns:
            Just the code part, or None if empty/invalid
        """
        if not display_string:
            return None

        # Remove whitespace
        display_string = display_string.strip()

        # Handle "-----" placeholder
        if display_string == "-----":
            return None

        # ✅ Strip visual indicators/prefixes (excluding ⭐ which is for editable cells)
        prefixes = ["●", "■", "◆", "►", "[S]", "[Primary]", "[Standalone]"]
        for prefix in prefixes:
            if display_string.startswith(prefix):
                display_string = display_string[len(prefix) :].strip()
                break

        # Check if it contains " - " separator
        if " - " in display_string:
            code = display_string.split(" - ", 1)[0].strip()
            return code if code else None

        # Already just a code
        return display_string

    # -----------------FEFO Helper----------------
    def _distribute_qty_by_fefo(self, item_rows, required_qty):
        """
        Distribute required quantity across multiple rows using FEFO (First Expiry First Out).

        Args:
            item_rows: List of dicts with keys: iid, expiry_date, current_stock, etc.
            required_qty: Total quantity needed (typically std_qty)

        Returns:
            Dict mapping iid ��� quantity_to_issue
        """
        if not item_rows or required_qty <= 0:
            return {}

        # Sort by expiry date (earliest first), then by iid for stability
        def parse_exp_date(exp_str):
            """Convert expiry date string to comparable format."""
            if not exp_str or exp_str == "":
                return "9999-12-31"  # Put items without expiry at end
            try:
                # Try parsing as YYYY-MM-DD
                if len(exp_str) == 10 and exp_str[4] == "-":
                    return exp_str
                # Try parsing as DD-Mon-YYYY (e.g., "31-Dec-2025")
                dt = datetime.strptime(exp_str, "%d-%b-%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return "9999-12-31"

        sorted_rows = sorted(
            item_rows,
            key=lambda r: (parse_exp_date(r.get("expiry_date", "")), r.get("iid", "")),
        )

        result = {}
        remaining = required_qty

        for row in sorted_rows:
            iid = row.get("iid")
            try:
                available = int(row.get("current_stock", 0))
            except Exception:
                available = 0

            if remaining <= 0:
                # No more quantity needed
                result[iid] = 0
            elif available >= remaining:
                # This row can fulfill remaining requirement
                result[iid] = remaining
                remaining = 0
            else:
                # Take all from this row and continue
                result[iid] = available
                remaining -= available

        logging.debug(
            f"[DISPATCH][FEFO] Distributed {required_qty} qty across {len(item_rows)} rows: {result}"
        )
        return result

    # ---------------------------------------------------------
    # Index / Parsing / Enrichment
    # ---------------------------------------------------------
    def ensure_item_index(self, scenario_id):
        if (
            hasattr(self, "_item_index_cache")
            and self._item_index_cache.get("scenario_id") == scenario_id
        ):
            return
        self._item_index_cache = {
            "scenario_id": scenario_id,
            "flat_map": {},
            "triple_map": {},
        }

        conn = connect_db()
        if conn is None:
            logging.warning("[DISPATCH] Cannot build item index (no DB connection).")
            return
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT code, Kit, module, item, treecode, level
                  FROM Kit_items
                 WHERE scenario_id = ?
            """,
                (scenario_id,),
            )
            for row in cur.fetchall():
                code = (row["code"] or "").strip()
                Kit = (row["Kit"] or "").strip()
                module = (row["module"] or "").strip()
                item = (row["item"] or "").strip()
                treecode = row["treecode"]
                level = (row["level"] or "").strip().lower()
                entry = {
                    "code": code,
                    "Kit": Kit,
                    "module": module,
                    "item": item,
                    "treecode": treecode,
                    "level": level,
                }
                for tok in (code, Kit, module, item):
                    if tok and tok.upper() != "NONE":
                        self._item_index_cache["flat_map"].setdefault(tok, entry)
                self._item_index_cache["triple_map"][
                    (Kit or None, module or None, item or None)
                ] = entry
            logging.debug(
                f"[DISPATCH] item_index built:{len(self._item_index_cache['flat_map'])} tokens"
            )
        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] ensure_item_index error: {e}")
        finally:
            cur.close()
            conn.close()

    @staticmethod
    def parse_unique_id(unique_id: str):
        """
        Expected pattern:
          scenario / Kit / module / item / std_qty / exp_date / Kit_number / module_number
        std_qty parsed from index 4 if present & >0 else 1.
        """
        Kit_code = module_code = item_code = None
        std_qty = 1
        if not unique_id:
            return {
                "Kit_code": None,
                "module_code": None,
                "item_code": None,
                "std_qty": 1,
            }
        parts = unique_id.split("/")
        if len(parts) >= 2:
            Kit_code = parts[1] or None
        if len(parts) >= 3:
            module_code = parts[2] or None
        if len(parts) >= 4:
            item_code = parts[3] or None
        if len(parts) >= 5:
            try:
                pv = int(parts[4])
                if pv > 0:
                    std_qty = pv
            except:
                std_qty = 1
        return {
            "Kit_code": Kit_code,
            "module_code": module_code,
            "item_code": item_code,
            "std_qty": std_qty,
        }

    @staticmethod
    def extract_std_qty_from_unique_id(unique_id: str) -> int:
        """
        Extract std_qty from unique_id.
        Format: scenario/kit/module/item/std_qty/exp_date/kit_number/module_number

        Args:
            unique_id: The unique identifier string

        Returns:
            std_qty as integer (defaults to 1 if not found or invalid)
        """
        if not unique_id:
            return 1

        parts = unique_id.split("/")
        if len(parts) < 5:
            return 1

        try:
            std_qty = int(parts[4])
            return std_qty if std_qty > 0 else 1
        except (ValueError, IndexError):
            return 1

    def enrich_stock_row(
        self, scenario_id, unique_id, final_qty, exp_date, Kit_number, module_number
    ):
        """
        Returns a dict with a normalized 'type' field: one of 'Kit', 'Module', 'Item'.
        Forced hierarchy: if module/item segments exist they override a broader detect_type result.
        """
        self.ensure_item_index(scenario_id)
        parsed = self.parse_unique_id(unique_id)
        Kit_code = parsed["Kit_code"]
        module_code = parsed["module_code"]
        item_code = parsed["item_code"]
        std_qty = parsed["std_qty"]

        # Determine display code & forced_type from unique_id structure
        if item_code and item_code.upper() != "NONE":
            display_code = item_code
            forced_type = "Item"
        elif module_code and module_code.upper() != "NONE":
            display_code = module_code
            forced_type = "Module"
        else:
            display_code = Kit_code
            forced_type = "Kit"

        # Lookup treecode (unchanged)
        treecode = None
        triple_key = (
            Kit_code if Kit_code else None,
            module_code if module_code else None,
            item_code if item_code else None,
        )
        idx = getattr(self, "_item_index_cache", {})
        triple_map = idx.get("triple_map", {})
        flat_map = idx.get("flat_map", {})
        entry = triple_map.get(triple_key) or flat_map.get(display_code or "")
        if entry:
            treecode = entry.get("treecode")

        description = get_item_description(display_code or "")
        detected = detect_type(display_code or "", description) or forced_type

        # Normalize detected to canonical Title case if valid
        detected_upper = detected.upper()
        valid_upper = {"KIT": "Kit", "MODULE": "Module", "ITEM": "Item"}
        if detected_upper in valid_upper:
            detected_norm = valid_upper[detected_upper]
        else:
            detected_norm = forced_type  # fallback

        # If forced_type narrows (Module/Item), honor it over detected
        if forced_type in ("Module", "Item"):
            final_type = forced_type
        else:
            # forced_type == 'Kit'; keep detected_norm only if it is not narrower
            final_type = "Kit" if detected_norm not in ("Kit",) else detected_norm

        return {
            "unique_id": unique_id,
            "code": display_code or "",
            "description": description,
            "type": final_type,  # 'Kit', 'Module', or 'Item'
            "Kit": Kit_code or "-----",
            "module": module_code or "-----",
            "current_stock": final_qty,
            "expiry_date": exp_date or "",
            "batch_no": "",
            "treecode": treecode,
            "Kit_number": Kit_number,
            "module_number": module_number,
            "std_qty": std_qty if final_type == "Item" else None,
        }

    # ---------------------------------------------------------
    # Scenario / Modes
    # ---------------------------------------------------------
    def fetch_scenario_map(self):
        conn = connect_db()
        if conn is None:
            logging.error("[DISPATCH] DB connection failed in fetch_scenario_map")
            return {}
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            rows = cur.fetchall()
            mapping = {str(r["scenario_id"]): r["name"] for r in rows}
            logging.info(f"[DISPATCH] Loaded {len(mapping)} scenarios")
            return mapping
        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] Error loading scenarios: {e}")
            return {}
        finally:
            cur.close()
            conn.close()

    def build_mode_definitions(self):
        scenario = self.selected_scenario_name or ""
        self.mode_definitions = [
            ("dispatch_kit", lang.t("dispatch_kit.mode_dispatch_kit", "Dispatch Kit")),
            (
                "issue_standalone",
                lang.t(
                    "dispatch_kit.mode_issue_standalone",
                    "Issue standalone item/s from {scenario}",
                    scenario=scenario,
                ),
            ),
            (
                "issue_module_scenario",
                lang.t(
                    "dispatch_kit.mode_issue_module_scenario",
                    "Issue module from {scenario}",
                    scenario=scenario,
                ),
            ),
            (
                "issue_module_kit",
                lang.t("dispatch_kit.mode_issue_module_kit", "Issue module from a Kit"),
            ),
            (
                "issue_items_kit",
                lang.t("dispatch_kit.mode_issue_items_kit", "Issue items from a Kit"),
            ),
            (
                "issue_items_module",
                lang.t(
                    "dispatch_kit.mode_issue_items_module", "Issue items from a module"
                ),
            ),
        ]
        self.mode_label_to_key = {label: key for key, label in self.mode_definitions}

    def current_mode_key(self):
        return self.mode_label_to_key.get(self.mode_var.get())

    def load_scenarios(self):
        values = [f"{sid} - {name}" for sid, name in self.scenario_map.items()]
        self.scenario_cb["values"] = values
        if values:
            self.scenario_cb.current(0)
            self.on_scenario_selected()
        else:
            if self.status_var:
                self.status_var.set(
                    lang.t(
                        "dispatch_kit.no_scenarios", "No scenarios found (check DB)."
                    )
                )

    def on_scenario_selected(self, event=None):
        sel = self.scenario_var.get()
        if not sel:
            self.selected_scenario_id = None
            self.selected_scenario_name = None
            return
        self.selected_scenario_id = sel.split(" - ")[0]
        self.selected_scenario_name = sel.split(" - ", 1)[1] if " - " in sel else ""

        self.build_mode_definitions()
        self.mode_cb["values"] = [lbl for _, lbl in self.mode_definitions]
        if self.mode_definitions:
            self.mode_var.set(self.mode_definitions[0][1])
        self.on_mode_changed()

    def on_mode_changed(self, event=None):
        """
        Called when movement type changes.
        Enables/disables appropriate selectors based on mode.
        """
        mode_key = self.current_mode_key()

        # Disable all selectors initially
        for cb in [
            self.Kit_cb,
            self.Kit_number_cb,
            self.module_cb,
            self.module_number_cb,
        ]:
            cb.config(state="disabled")

        # Clear selections
        self.Kit_var.set("")
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")

        # Clear dropdown values
        self.Kit_cb["values"] = []
        self.Kit_number_cb["values"] = []
        self.module_cb["values"] = []
        self.module_number_cb["values"] = []

        # Clear table
        self.full_items = []
        self.clear_table_only()

        # Clear search
        self.search_var.set("")
        self.search_listbox.delete(0, tk.END)

        if not self.selected_scenario_id:
            return

        # ===== Mode-specific logic =====

        if mode_key == "dispatch_kit":
            # Show primary kits only
            self.Kit_cb.config(state="readonly")
            self.Kit_cb["values"] = self.fetch_kits(self.selected_scenario_id)
            logging.debug(
                f"[MODE_CHANGED] dispatch_kit: Populated {len(self.Kit_cb['values'])} primary kits"
            )

        elif mode_key == "issue_standalone":
            # ✅ Populate search with primary items
            items = self.fetch_primary_items(self.selected_scenario_id)
            for item_display in items:
                self.search_listbox.insert(tk.END, item_display)
            logging.debug(
                f"[MODE_CHANGED] issue_standalone: Populated {len(items)} primary items in search"
            )
            if items:
                self.status_var.set(
                    lang.t(
                        "dispatch_kit.found_items",
                        "Found {count} items",
                        count=len(items),
                    )
                )

        elif mode_key == "issue_module_scenario":
            # ✅ Populate search with primary modules
            modules = self.fetch_all_modules(self.selected_scenario_id)
            for module_display in modules:
                self.search_listbox.insert(tk.END, module_display)

            # Also activate module_number dropdown
            self.module_cb.config(state="readonly")
            self.module_cb["values"] = modules
            logging.debug(
                f"[MODE_CHANGED] issue_module_scenario: Populated {len(modules)} primary modules"
            )
            if modules:
                self.status_var.set(
                    lang.t(
                        "dispatch_kit.found_modules",
                        "Found {count} modules",
                        count=len(modules),
                    )
                )

        elif mode_key == "issue_module_kit":
            # Enable kit selector (primary kits)
            # Module selector will populate when kit is selected
            self.Kit_cb.config(state="readonly")
            self.Kit_cb["values"] = self.fetch_kits(self.selected_scenario_id)
            logging.debug(
                f"[MODE_CHANGED] issue_module_kit: Populated {len(self.Kit_cb['values'])} primary kits"
            )

        elif mode_key == "issue_items_kit":
            # Enable kit selector (primary kits)
            # Items will be found via search after kit is selected
            self.Kit_cb.config(state="readonly")
            self.Kit_cb["values"] = self.fetch_kits(self.selected_scenario_id)
            logging.debug(
                f"[MODE_CHANGED] issue_items_kit: Populated {len(self.Kit_cb['values'])} primary kits"
            )

        elif mode_key == "issue_items_module":
            # ✅ Enable BOTH kit and module dropdowns
            # User can choose either:
            # - Select kit first (module inside kit) OR
            # - Select module directly (standalone primary module)

            self.Kit_cb.config(state="readonly")
            self.Kit_cb["values"] = self.fetch_kits(self.selected_scenario_id)

            self.module_cb.config(state="readonly")
            # ✅ Show ALL modules (both primary standalone AND modules inside kits)
            all_modules = self.fetch_all_modules_combined(self.selected_scenario_id)
            self.module_cb["values"] = all_modules

            logging.debug(
                f"[MODE_CHANGED] issue_items_module: Kits and modules both enabled"
            )

    # ---------------------------------------------------------
    # Structural Helpers
    # ---------------------------------------------------------
    def fetch_kits(self, scenario_id):
        """
        Fetch PRIMARY kits from kit_items (level='primary').
        Only includes items with type='Kit' (language-independent).

        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # Get primary kits from kit_items (level='primary')
            cur.execute(
                """
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """,
                (scenario_id,),
            )

            kit_codes = [r["code"] for r in cur.fetchall()]

            if not kit_codes:
                logging.debug(
                    f"[FETCH_KITS] No primary kits found for scenario {scenario_id}"
                )
                return []

            # Get descriptions and filter by type
            result = []
            for kit_code in kit_codes:
                desc = get_item_description(kit_code)
                item_type = detect_type(kit_code, desc).upper()

                # Only include if type is KIT
                if item_type == "KIT":
                    display = f"{kit_code} - {desc}" if desc else kit_code
                    result.append(display)

            logging.debug(
                f"[FETCH_KITS] Found {len(result)} primary kits for scenario {scenario_id}"
            )
            return result

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_kits error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_all_modules(self, scenario_id):
        """
        Fetch PRIMARY modules from kit_items (level='primary', module column filled, no parent kit).
        Only includes items with type='Module' (language-independent).

        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # ✅ Get primary modules (level='primary' AND module column is filled, kit is empty)
            cur.execute(
                """
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND module IS NOT NULL 
                  AND module != ''
                  AND module != 'None'
                  AND (kit IS NULL OR kit = '' OR kit = 'None')
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """,
                (scenario_id,),
            )

            module_codes = [r["code"] for r in cur.fetchall()]

            if not module_codes:
                logging.debug(
                    f"[FETCH_MODULES] No primary modules found for scenario {scenario_id}"
                )
                return []

            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()

                # Only include if type is MODULE
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)

            logging.debug(
                f"[FETCH_MODULES] Found {len(result)} primary modules for scenario {scenario_id}"
            )
            return result

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_all_modules error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_modules_for_kit(self, scenario_id, kit_code):
        """
        Fetch SECONDARY modules inside a specific kit from kit_items.
        Only includes items with type='Module'.

        Args:
            scenario_id: Scenario ID
            kit_code: Kit code (extracted from dropdown)

        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # Get secondary modules inside this kit (level='secondary' and kit=kit_code)
            cur.execute(
                """
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND kit=?
                  AND level='secondary'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """,
                (scenario_id, kit_code),
            )

            module_codes = [r["code"] for r in cur.fetchall()]

            if not module_codes:
                logging.debug(
                    f"[FETCH_MODULES_FOR_KIT] No modules found in kit {kit_code}"
                )
                return []

            # Get descriptions and filter by type
            result = []
            for module_code in module_codes:
                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()

                # Only include if type is MODULE
                if item_type == "MODULE":
                    display = f"{module_code} - {desc}" if desc else module_code
                    result.append(display)

            logging.debug(
                f"[FETCH_MODULES_FOR_KIT] Found {len(result)} modules in kit {kit_code}"
            )
            return result

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_modules_for_kit error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_primary_items(self, scenario_id):
        """
        Fetch PRIMARY items from kit_items (level='primary', item column filled, no parent kit or module).
        Only includes items with type='Item'.

        Returns:
            List of formatted strings: "CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # ✅ Get primary items (level='primary' AND item column is filled, kit/module are empty)
            cur.execute(
                """
                SELECT DISTINCT code
                FROM kit_items
                WHERE scenario_id=? 
                  AND level='primary'
                  AND item IS NOT NULL 
                  AND item != ''
                  AND item != 'None'
                  AND (kit IS NULL OR kit = '' OR kit = 'None')
                  AND (module IS NULL OR module = '' OR module = 'None')
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY code
            """,
                (scenario_id,),
            )

            item_codes = [r["code"] for r in cur.fetchall()]

            if not item_codes:
                logging.debug(
                    f"[FETCH_PRIMARY_ITEMS] No primary items found for scenario {scenario_id}"
                )
                return []

            # Get descriptions and filter by type
            result = []
            for item_code in item_codes:
                desc = get_item_description(item_code)
                item_type = detect_type(item_code, desc).upper()

                # Only include if type is ITEM
                if item_type == "ITEM":
                    display = f"{item_code} - {desc}" if desc else item_code
                    result.append(display)

            logging.debug(
                f"[FETCH_PRIMARY_ITEMS] Found {len(result)} primary items for scenario {scenario_id}"
            )
            return result

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_primary_items error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_available_kit_numbers(self, scenario_id, kit_code=None):
        """
        Fetch kit numbers with available stock (final_qty > 0).

        Args:
            scenario_id: Scenario ID
            kit_code: Optional kit code to filter by (extracted from dropdown)

        Returns:
            List of kit_number strings
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))

            where_clauses = [
                "(scenario=? OR scenario=?)",
                "kit_number IS NOT NULL",
                "kit_number != 'None'",
                "final_qty > 0",
            ]
            params = [str(scenario_id), scenario_name]

            # ✅ Filter by kit code if provided
            if kit_code:
                where_clauses.append("kit=?")
                params.append(kit_code)
                logging.debug(f"[FETCH_KIT_NUMBERS] Filtering by kit_code={kit_code}")

            sql = f"""
                SELECT DISTINCT kit_number
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY kit_number
            """

            logging.debug(f"[FETCH_KIT_NUMBERS] SQL: {sql}")
            logging.debug(f"[FETCH_KIT_NUMBERS] Params: {params}")

            cur.execute(sql, params)
            results = [r["kit_number"] for r in cur.fetchall()]

            logging.debug(
                f"[FETCH_KIT_NUMBERS] Found {len(results)} kit numbers: {results}"
            )
            return results

        except sqlite3.Error as e:
            logging.error(f"[FETCH_KIT_NUMBERS] Error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_module_numbers(
        self, scenario_id, kit_code=None, module_code=None, kit_number=None
    ):
        """
        Fetch module numbers with available stock (final_qty > 0).

        Args:
            scenario_id: Scenario ID
            kit_code: NOT USED (kept for compatibility)
            module_code: Module code to filter by (extracted from dropdown)
            kit_number: Kit number to filter by (the actual kit instance)

        Returns:
            List of module_number strings
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))

            where_clauses = [
                "(scenario=? OR scenario=?)",
                "module_number IS NOT NULL",
                "module_number != 'None'",
                "final_qty > 0",
            ]
            params = [str(scenario_id), scenario_name]

            # ✅ Filter by kit_number (the actual kit instance)
            if kit_number:
                where_clauses.append("kit_number=?")
                params.append(kit_number)
                logging.debug(
                    f"[FETCH_MODULE_NUMBERS] Filtering by kit_number={kit_number}"
                )

            # ✅ Filter by module code
            if module_code:
                where_clauses.append("module=?")
                params.append(module_code)
                logging.debug(
                    f"[FETCH_MODULE_NUMBERS] Filtering by module_code={module_code}"
                )

            sql = f"""
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY module_number
            """

            logging.debug(f"[FETCH_MODULE_NUMBERS] SQL: {sql}")
            logging.debug(f"[FETCH_MODULE_NUMBERS] Params: {params}")

            cur.execute(sql, params)
            results = [r["module_number"] for r in cur.fetchall()]

            logging.debug(
                f"[FETCH_MODULE_NUMBERS] Found {len(results)} module numbers: {results}"
            )
            return results

        except sqlite3.Error as e:
            logging.error(f"[FETCH_MODULE_NUMBERS] Error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_module_numbers_standalone(self, scenario_id, module_code):
        """
        Fetch module_numbers for a PRIMARY standalone module (not inside a kit).

        Args:
            scenario_id: Scenario ID
            module_code: Code of the standalone module

        Returns:
            List of module_number strings
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # Query stock_data for module_numbers where:
            # - module = module_code
            # - kit is NULL or empty (standalone)
            cur.execute(
                """
                SELECT DISTINCT module_number
                FROM stock_data
                WHERE (scenario=? OR scenario=?)
                AND module_number IS NOT NULL
                AND module_number != 'None'
                AND final_qty > 0
                AND module=?
                AND (kit IS NULL OR kit = '' OR kit = 'None')
                ORDER BY module_number
            """,
                (scenario_id, self.scenario_map.get(scenario_id, ""), module_code),
            )

            module_numbers = [r["module_number"] for r in cur.fetchall()]

            logging.debug(
                f"[FETCH_MODULE_NUMBERS_STANDALONE] Found {len(module_numbers)} for module={module_code}"
            )
            return module_numbers

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_module_numbers_standalone error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_all_modules_combined(self, scenario_id):
        """
        Fetch ALL modules for a scenario:
        - Primary standalone modules (marked with ●)
        - Secondary modules (inside kits)

        Returns:
            List of formatted strings: "CODE - Description" or "● CODE - Description"
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # ✅ Get all modules with level information
            cur.execute(
                """
                SELECT DISTINCT code, level, kit
                FROM kit_items
                WHERE scenario_id=? 
                  AND module IS NOT NULL 
                  AND module != ''
                  AND module != 'None'
                  AND code IS NOT NULL 
                  AND code != ''
                ORDER BY 
                  CASE 
                    WHEN level='primary' AND (kit IS NULL OR kit='' OR kit='None') THEN 0
                    ELSE 1
                  END,
                code
            """,
                (scenario_id,),
            )

            rows = cur.fetchall()

            result = []
            for row in rows:
                module_code = row["code"]
                level = row["level"]
                kit = row["kit"]

                desc = get_item_description(module_code)
                item_type = detect_type(module_code, desc).upper()

                if item_type == "MODULE":
                    # ✅ Check if it's a standalone primary module
                    is_standalone = level == "primary" and (
                        kit is None or kit == "" or kit == "None"
                    )

                    # ✅ Add bullet for standalone modules
                    if is_standalone:
                        display = (
                            f"● {module_code} - {desc}" if desc else f"● {module_code}"
                        )
                    else:
                        display = f"{module_code} - {desc}" if desc else module_code

                    result.append(display)

            logging.debug(f"[FETCH_ALL_MODULES_COMBINED] Found {len(result)} modules")
            return result

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_all_modules_combined error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    # ---------------------------------------------------------
    # Stock Fetching
    # ---------------------------------------------------------
    def fetch_stock_data_for_kit_number(self, scenario_id, Kit_number, Kit_code=None):
        """
        Fetch stock data for a specific kit number.

        Args:
            scenario_id: Scenario ID
            Kit_number: The kit instance number (e.g., "CHOL-001")
            Kit_code: Optional kit code for additional filtering (extracted from dropdown)

        Returns:
            List of stock items with final_qty > 0
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))

            where_clauses = [
                "(scenario=? OR scenario=?)",
                "kit_number=?",
                "final_qty > 0",
            ]
            params = [str(scenario_id), scenario_name, Kit_number]

            # ✅ Optional: filter by kit code
            if Kit_code:
                where_clauses.append("kit=?")
                params.append(Kit_code)
                logging.debug(
                    f"[FETCH_KIT_STOCK] Additional filter by Kit_code={Kit_code}"
                )

            sql = f"""
                SELECT *
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY unique_id
            """

            logging.debug(f"[FETCH_KIT_STOCK] SQL: {sql}")
            logging.debug(f"[FETCH_KIT_STOCK] Params: {params}")

            cur.execute(sql, params)
            rows = cur.fetchall()

            items = []
            for row in rows:
                item = dict(row)
                # Enrich with additional details
                enriched = self.enrich_stock_row(
                    scenario_id,
                    item["unique_id"],
                    item["final_qty"],
                    item.get("exp_date"),
                    item.get("kit_number"),
                    item.get("module_number"),
                )
                items.append(enriched)

            logging.debug(
                f"[FETCH_KIT_STOCK] Found {len(items)} items for Kit_number={Kit_number}"
            )
            return items

        except sqlite3.Error as e:
            logging.error(f"[FETCH_KIT_STOCK] Error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fetch_stock_data_for_module_number(
        self, scenario_id, module_number, Kit_code=None, module_code=None
    ):
        """
        Fetch stock data for a specific module number.

        Args:
            scenario_id: Scenario ID
            module_number: The module instance number (e.g., "M-001")
            Kit_code: Optional kit code for filtering (extracted from dropdown)
            module_code: Optional module code for filtering (extracted from dropdown)

        Returns:
            List of stock items with final_qty > 0
        """
        conn = connect_db()
        if conn is None:
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            scenario_name = self.scenario_map.get(str(scenario_id), str(scenario_id))

            where_clauses = [
                "(scenario=? OR scenario=?)",
                "module_number=?",
                "final_qty > 0",
            ]
            params = [str(scenario_id), scenario_name, module_number]

            # ✅ Optional: filter by kit code
            if Kit_code:
                where_clauses.append("kit=?")
                params.append(Kit_code)
                logging.debug(
                    f"[FETCH_MODULE_STOCK] Additional filter by Kit_code={Kit_code}"
                )

            # ✅ Optional: filter by module code
            if module_code:
                where_clauses.append("module=?")
                params.append(module_code)
                logging.debug(
                    f"[FETCH_MODULE_STOCK] Additional filter by module_code={module_code}"
                )

            sql = f"""
                SELECT *
                FROM stock_data
                WHERE {' AND '.join(where_clauses)}
                ORDER BY unique_id
            """

            logging.debug(f"[FETCH_MODULE_STOCK] SQL: {sql}")
            logging.debug(f"[FETCH_MODULE_STOCK] Params: {params}")

            cur.execute(sql, params)
            rows = cur.fetchall()

            items = []
            for row in rows:
                item = dict(row)
                # Enrich with additional details
                enriched = self.enrich_stock_row(
                    scenario_id,
                    item["unique_id"],
                    item["final_qty"],
                    item.get("exp_date"),
                    item.get("kit_number"),
                    item.get("module_number"),
                )
                items.append(enriched)

            logging.debug(
                f"[FETCH_MODULE_STOCK] Found {len(items)} items for module_number={module_number}"
            )
            return items

        except sqlite3.Error as e:
            logging.error(f"[FETCH_MODULE_STOCK] Error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def fill_from_search(self, event=None):
        """
        Load stock data when user clicks/selects an item from search listbox.
        Works for issue_standalone and issue_module_scenario modes.
        """
        mode_key = self.current_mode_key()

        # Get selected item from listbox
        selection = self.search_listbox.curselection()
        if not selection:
            logging.debug("[FILL_FROM_SEARCH] No selection")
            return

        selected_display = self.search_listbox.get(selection[0])
        selected_code = self._extract_code_from_display(selected_display)

        if not selected_code:
            logging.warning(
                f"[FILL_FROM_SEARCH] Could not extract code from '{selected_display}'"
            )
            return

        logging.debug(
            f"[FILL_FROM_SEARCH] Selected: '{selected_display}' -> Code: '{selected_code}', mode: {mode_key}"
        )

        # ===== Mode-specific logic =====

        if mode_key == "issue_standalone":
            # Fetch all stock rows for this item code
            items = self.fetch_standalone_stock_items(
                self.selected_scenario_id, selected_code
            )

            if items:
                self.populate_rows(
                    items,
                    lang.t(
                        "dispatch_kit.loaded_item_stock",
                        "Loaded {count} stock rows for item {code}",
                        count=len(items),
                        code=selected_code,
                    ),
                )
            else:
                self.clear_table_only()
                self.status_var.set(
                    lang.t(
                        "dispatch_kit.no_stock",
                        "No stock found for {code}",
                        code=selected_code,
                    )
                )

        elif mode_key == "issue_module_scenario":
            # Fetch all stock rows for this module code
            items = self.fetch_standalone_stock_items(
                self.selected_scenario_id, selected_code
            )

            if items:
                self.populate_rows(
                    items,
                    lang.t(
                        "dispatch_kit.loaded_module_stock",
                        "Loaded {count} stock rows for module {code}",
                        count=len(items),
                        code=selected_code,
                    ),
                )
            else:
                self.clear_table_only()
                self.status_var.set(
                    lang.t(
                        "dispatch_kit.no_stock",
                        "No stock found for {code}",
                        code=selected_code,
                    )
                )
        else:
            logging.debug(
                f"[FILL_FROM_SEARCH] Mode {mode_key} doesn't support listbox selection"
            )

    def fetch_standalone_stock_items(self, scenario_id, item_code):
        """
        Fetch all stock data for a standalone item or module (by code).
        Used for issue_standalone and issue_module_scenario modes.

        Args:
            scenario_id: Scenario ID
            item_code: The code of the item or module to fetch

        Returns:
            List of enriched stock item dicts
        """
        conn = connect_db()
        if conn is None:
            logging.error("[DISPATCH] DB connection failed")
            return []

        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        try:
            # ✅ Query stock_data for rows matching this code
            # Check kit, module, OR item columns
            logging.debug(
                f"[FETCH_STANDALONE_STOCK] Fetching stock for code={item_code}, scenario={scenario_id}"
            )

            cur.execute(
                """
                SELECT unique_id, final_qty, exp_date, kit_number, module_number
                FROM stock_data
                WHERE (scenario=? OR scenario=?) 
                  AND final_qty > 0
                  AND (kit=? OR module=? OR item=?)
                ORDER BY exp_date, unique_id
            """,
                (
                    scenario_id,
                    self.scenario_map.get(scenario_id, ""),
                    item_code,
                    item_code,
                    item_code,
                ),
            )

            rows = cur.fetchall()

            if not rows:
                logging.debug(
                    f"[FETCH_STANDALONE_STOCK] No stock found for code={item_code}"
                )
                return []

            logging.debug(
                f"[FETCH_STANDALONE_STOCK] Found {len(rows)} rows in stock_data for code={item_code}"
            )

            # Enrich each row
            items = []
            for row in rows:
                enriched = self.enrich_stock_row(
                    scenario_id,
                    row["unique_id"],
                    row["final_qty"],
                    row["exp_date"],
                    row["kit_number"],
                    row["module_number"],
                )
                items.append(enriched)
                logging.debug(
                    f"[FETCH_STANDALONE_STOCK] Enriched: {enriched['code']} qty={enriched['current_stock']}"
                )

            logging.debug(
                f"[FETCH_STANDALONE_STOCK] Returning {len(items)} enriched items"
            )
            return items

        except sqlite3.Error as e:
            logging.error(f"[DISPATCH] fetch_standalone_stock_items error: {e}")
            return []
        finally:
            cur.close()
            conn.close()

    def on_qty_to_issue_changed(self, iid: str):
        """
        Called when qty_to_issue changes for an item.
        Recalculates qty_required for all subsequent siblings with same code.

        Args:
            iid: Tree item ID that was changed
        """
        if iid not in self.row_data:
            return

        item_data = self.row_data[iid]
        code = item_data.get("code", "")
        parent_iid = self.tree.parent(iid)

        if not parent_iid:
            return

        # Get all children of parent
        siblings = list(self.tree.get_children(parent_iid))

        # Find current item index
        try:
            current_index = siblings.index(iid)
        except ValueError:
            return

        # Update qty_required for current item and all subsequent siblings with same code
        for sibling_iid in siblings[current_index:]:
            if sibling_iid not in self.row_data:
                continue

            sibling_data = self.row_data[sibling_iid]

            # Only update siblings with same item code
            if sibling_data.get("code") == code:
                # Recalculate qty_required
                qty_required = self.calculate_qty_required(sibling_iid)

                # Update tree display
                self.tree.set(sibling_iid, "qty_required", qty_required)

                # Update row_data
                self.row_data[sibling_iid]["qty_required"] = qty_required

                logging.debug(
                    f"[QTY_REQUIRED] Updated {sibling_iid}: qty_required={qty_required}"
                )

    def show_qty_edit_menu(self, event):
        """
        Show context menu for editing qty_to_issue on right-click.
        Only shows for ITEM rows (Kit/Module use direct editing).
        """
        # Identify what was clicked
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)

        if not row_id or not col_id:
            return

        # Check if qty_to_issue column (column 10, index 9)
        col_index = int(col_id.replace("#", "")) - 1
        if col_index != 9:  # qty_to_issue is column 9 (0-indexed)
            return

        # Check if row is an Item (not Kit/Module)
        vals = self.tree.item(row_id, "values")
        row_type = (vals[2] or "").upper()

        if row_type != "ITEM":
            return  # Only items can be edited via right-click

        # Check if row has stock
        meta = self.row_data.get(row_id, {})
        if meta.get("is_header"):
            return

        try:
            current_stock = int(vals[5]) if vals[5] else 0
        except:
            current_stock = 0

        if current_stock == 0:
            self.status_var.set(lang.t("dispatch_kit.no_stock", "No stock available"))
            return

        # Create context menu
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(
            label=lang.t("dispatch_kit.edit_quantity", "Edit Quantity"),
            command=lambda: self.edit_qty_to_issue_popup(row_id, current_stock),
        )

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def calculate_qty_required(self, iid: str) -> int:
        """
        Calculate quantity required for a specific tree item.
        Formula: std_qty - sum(qty_to_issue of previous siblings with same code)

        This implements FEFO by showing remaining requirement after earlier batches.

        Args:
            iid: Tree item ID

        Returns:
            Quantity still required (0 if fully allocated)
        """
        if iid not in self.row_data:
            return 0

        item_data = self.row_data[iid]
        unique_id = item_data.get("unique_id", "")
        code = item_data.get("code", "")
        item_type = item_data.get("type", "").upper()

        # Only calculate for items (not kits or modules)
        if item_type != "ITEM":
            return 0

        # Extract std_qty from unique_id
        std_qty = self.extract_std_qty_from_unique_id(unique_id)

        if std_qty <= 0:
            return 0

        # Find parent
        parent_iid = self.tree.parent(iid)
        if not parent_iid:
            # No parent - return full std_qty
            return std_qty

        # Get all children of parent (in order)
        siblings = self.tree.get_children(parent_iid)

        # Sum quantities already allocated to EARLIER siblings with same code
        allocated_qty = 0
        for sibling_iid in siblings:
            # Stop when we reach current item (FEFO: only count earlier batches)
            if sibling_iid == iid:
                break

            if sibling_iid not in self.row_data:
                continue

            sibling_data = self.row_data[sibling_iid]

            # Only count siblings with same item code
            if sibling_data.get("code") == code:
                try:
                    qty_to_issue = int(sibling_data.get("qty_to_issue", 0))
                    allocated_qty += qty_to_issue
                except (ValueError, TypeError):
                    pass

        # Calculate remaining requirement
        remaining = std_qty - allocated_qty

        return max(0, remaining)  # Never negative

    # ---------------------------------------------------------
    # UI
    # ---------------------------------------------------------
    def render_ui(self):
        if not self.parent:
            return
        for w in self.parent.winfo_children():
            try:
                w.destroy()
            except:
                pass

        title_frame = tk.Frame(self.parent, bg="#F0F4F8")
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text=lang.t("dispatch_kit.title", "Dispatch Kit-Module"),
            font=("Helvetica", 20, "bold"),
            bg="#F0F4F8",
        ).pack(pady=(10, 0))

        # ✅ UPDATED: Instructions mention right-click for qty editing
        instruct_frame = tk.Frame(
            self.parent,
            bg="#FFF9C4",
            highlightbackground="#E0D890",
            highlightthickness=1,
            bd=0,
        )
        instruct_frame.pack(fill="x", padx=10, pady=(6, 4))
        tk.Label(
            instruct_frame,
            text=lang.t(
                "dispatch_kit.instructions",
                "Cells marked with ★ are editable. Right-click 'Qty to Issue' to edit quantity. "
                "For Kits and modules quantity entered can be either 1 or 0.",
            ),
            fg="#444",
            bg="#FFF9C4",
            font=("Helvetica", 10, "italic"),
        ).pack(padx=8, pady=4, anchor="w")

        main = tk.Frame(self.parent, bg="#F0F4F8")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(
            main, text=lang.t("dispatch_kit.scenario", "Scenario:"), bg="#F0F4F8"
        ).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(
            main, textvariable=self.scenario_var, state="readonly", width=40
        )
        self.scenario_cb.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)

        tk.Label(
            main,
            text=lang.t("dispatch_kit.movement_type", "Movement Type:"),
            bg="#F0F4F8",
        ).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.mode_var = tk.StringVar()
        self.mode_cb = ttk.Combobox(
            main, textvariable=self.mode_var, state="readonly", width=40
        )
        self.mode_cb.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        self.mode_cb.bind("<<ComboboxSelected>>", self.on_mode_changed)

        # Kit selector
        self.Kit_var = tk.StringVar()
        self.Kit_cb = ttk.Combobox(
            main, textvariable=self.Kit_var, state="disabled", width=80
        )
        tk.Label(
            main, text=lang.t("dispatch_kit.select_kit", "Select Kit:"), bg="#F0F4F8"
        ).grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.Kit_cb.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.Kit_cb.bind("<<ComboboxSelected>>", self.on_kit_selected)

        # Kit number selector
        self.Kit_number_var = tk.StringVar()
        self.Kit_number_cb = ttk.Combobox(
            main, textvariable=self.Kit_number_var, state="disabled", width=20
        )
        tk.Label(
            main,
            text=lang.t("dispatch_kit.select_kit_number", "Select Kit Number:"),
            bg="#F0F4F8",
        ).grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.Kit_number_cb.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.Kit_number_cb.bind("<<ComboboxSelected>>", self.on_kit_number_selected)

        # Module selector
        self.module_var = tk.StringVar()
        self.module_cb = ttk.Combobox(
            main, textvariable=self.module_var, state="disabled", width=80
        )
        tk.Label(
            main,
            text=lang.t("dispatch_kit.select_module", "Select Module:"),
            bg="#F0F4F8",
        ).grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.module_cb.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.module_cb.bind("<<ComboboxSelected>>", self.on_module_selected)

        # Module number selector
        self.module_number_var = tk.StringVar()
        self.module_number_cb = ttk.Combobox(
            main, textvariable=self.module_number_var, state="disabled", width=20
        )
        tk.Label(
            main,
            text=lang.t("dispatch_kit.select_module_number", "Select Module Number:"),
            bg="#F0F4F8",
        ).grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.module_number_cb.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        self.module_number_cb.bind(
            "<<ComboboxSelected>>", self.on_module_number_selected
        )

        type_frame = tk.Frame(main, bg="#F0F4F8")
        type_frame.grid(row=4, column=0, columnspan=4, pady=5, sticky="w")
        tk.Label(
            type_frame, text=lang.t("dispatch_kit.out_type", "OUT Type:"), bg="#F0F4F8"
        ).grid(row=0, column=0, padx=5, sticky="w")
        self.trans_type_var = tk.StringVar()
        out_type_values = [lbl for _, lbl in self._out_type_options()]
        self.trans_type_cb = ttk.Combobox(
            type_frame,
            textvariable=self.trans_type_var,
            values=out_type_values,
            state="readonly",
            width=30,
        )
        self.trans_type_cb.grid(row=0, column=1, padx=5, pady=5)
        self.trans_type_cb.bind("<<ComboboxSelected>>", self.on_out_type_selected)

        tk.Label(
            type_frame, text=lang.t("dispatch_kit.end_user", "End User:"), bg="#F0F4F8"
        ).grid(row=0, column=2, padx=5, sticky="w")
        self.end_user_var = tk.StringVar()
        self.end_user_cb = ttk.Combobox(
            type_frame, textvariable=self.end_user_var, state="disabled", width=30
        )
        self.end_user_cb["values"] = fetch_end_users()
        self.end_user_cb.grid(row=0, column=3, padx=5, pady=5)

        tk.Label(
            type_frame,
            text=lang.t("dispatch_kit.third_party", "Third Party:"),
            bg="#F0F4F8",
        ).grid(row=0, column=4, padx=5, sticky="w")
        self.third_party_var = tk.StringVar()
        self.third_party_cb = ttk.Combobox(
            type_frame, textvariable=self.third_party_var, state="disabled", width=30
        )
        self.third_party_cb["values"] = fetch_third_parties()
        self.third_party_cb.grid(row=0, column=5, padx=5, pady=5)

        tk.Label(
            type_frame, text=lang.t("dispatch_kit.remarks", "Remarks:"), bg="#F0F4F8"
        ).grid(row=0, column=6, padx=5, sticky="w")
        self.remarks_entry = tk.Entry(type_frame, width=40, state="disabled")
        self.remarks_entry.grid(row=0, column=7, padx=5, pady=5)

        tk.Label(
            main,
            text=lang.t("dispatch_kit.item_search", "Kit/Module/Item:"),
            bg="#F0F4F8",
        ).grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(main, textvariable=self.search_var, width=40)
        self.search_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", self.search_items)

        tk.Button(
            main,
            text=lang.t("dispatch_kit.clear_search", "Clear Search"),
            bg="#7F8C8D",
            fg="white",
            command=self.clear_search,
        ).grid(row=5, column=2, padx=5, pady=5)

        self.search_listbox = tk.Listbox(main, height=5, width=60)
        self.search_listbox.grid(
            row=6, column=1, columnspan=3, padx=5, pady=5, sticky="we"
        )

        self.search_listbox.bind("<<ListboxSelect>>", self.fill_from_search)
        self.search_listbox.bind("<Return>", self.fill_from_search)
        self.search_listbox.bind("<space>", self.fill_from_search)

        logging.debug("[RENDER_UI] Search listbox bindings set")

        # ✅ UPDATED: Added qty_required column between batch_no and qty_to_issue
        cols = (
            "code",
            "description",
            "type",
            "Kit",
            "module",
            "current_stock",
            "expiry_date",
            "batch_no",
            "qty_required",
            "qty_to_issue",
            "unique_id",
        )
        self.tree = ttk.Treeview(main, columns=cols, show="headings", height=18)

        # ✅ UPDATED: Added qty_required heading
        headings = {
            "code": lang.t("dispatch_kit.code", "Code"),
            "description": lang.t("dispatch_kit.description", "Description"),
            "type": lang.t("dispatch_kit.type", "Type"),
            "Kit": lang.t("dispatch_kit.kit", "Kit"),
            "module": lang.t("dispatch_kit.module", "Module"),
            "current_stock": lang.t("dispatch_kit.current_stock", "Current Stock"),
            "expiry_date": lang.t("dispatch_kit.expiry_date", "Expiry Date"),
            "batch_no": lang.t("dispatch_kit.batch_no", "Batch Number"),
            "qty_required": lang.t(
                "dispatch_kit.qty_required", "Qty Required"
            ),  # ✅ NEW
            "qty_to_issue": lang.t("dispatch_kit.qty_to_issue", "Quantity to Issue"),
            "unique_id": "Unique ID",
        }

        # ✅ UPDATED: Added qty_required width
        widths = {
            "code": 160,
            "description": 380,
            "type": 120,
            "Kit": 120,
            "module": 120,
            "current_stock": 110,
            "expiry_date": 150,
            "batch_no": 140,
            "qty_required": 110,  # ✅ NEW
            "qty_to_issue": 140,
            "unique_id": 0,
        }

        # ✅ UPDATED: Added qty_required alignment
        aligns = {
            "code": "w",
            "description": "w",
            "type": "w",
            "Kit": "w",
            "module": "w",
            "current_stock": "e",
            "expiry_date": "w",
            "batch_no": "w",
            "qty_required": "e",  # ✅ NEW - right-aligned for numbers
            "qty_to_issue": "e",
            "unique_id": "w",
        }

        for c in cols:
            self.tree.heading(c, text=headings[c])
            self.tree.column(
                c,
                width=widths[c],
                anchor=aligns[c],
                stretch=(False if c == "unique_id" else True),
                minwidth=0 if c == "unique_id" else widths[c],
            )

        vsb = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        vsb.grid(row=7, column=4, sticky="ns")
        self.tree.configure(yscrollcommand=vsb.set)
        hsb = ttk.Scrollbar(main, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=8, column=0, columnspan=4, sticky="ew")
        self.tree.configure(xscrollcommand=hsb.set)
        self.tree.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")
        main.grid_rowconfigure(7, weight=1)
        main.grid_columnconfigure(1, weight=1)

        # ✅ Existing bindings (starred cells remain directly editable)
        self.tree.bind("<Double-1>", self.start_edit)
        self.tree.bind("<KeyPress-Return>", self.start_edit)
        self.tree.bind("<KeyPress-Tab>", self.start_edit)
        self.tree.bind("<KeyPress-Up>", self.navigate_tree)
        self.tree.bind("<KeyPress-Down>", self.navigate_tree)

        # ✅ NEW: Right-click context menu for qty_to_issue column
        self.tree.bind(
            "<Button-3>", self.show_qty_edit_menu
        )  # Right-click (Windows/Linux)
        self.tree.bind(
            "<Control-Button-1>", self.show_qty_edit_menu
        )  # Ctrl+Click (Mac)

        logging.debug(
            "[RENDER_UI] Tree bindings set - starred cells editable, qty_to_issue via right-click"
        )

        # ✅ Configure visual tags for tree styling
        self.tree.tag_configure(
            "Kit_header", background="#E3F6E1", font=("Helvetica", 10, "bold")
        )
        self.tree.tag_configure(
            "module_header", background="#E1ECFC", font=("Helvetica", 10, "bold")
        )
        self.tree.tag_configure("Kit_module_highlight", background="#FFF9C4")
        self.tree.tag_configure("editable_row", foreground="#000000")
        self.tree.tag_configure("non_editable", foreground="#666666")

        logging.debug("[RENDER_UI] Tree visual tags configured")

        btnf = tk.Frame(main, bg="#F0F4F8")
        btnf.grid(row=9, column=0, columnspan=4, pady=5)
        tk.Button(
            btnf,
            text=lang.t("dispatch_kit.save", "Save"),
            bg="#27AE60",
            fg="white",
            command=self.save_all,
            state="normal" if self.role in ["admin", "manager"] else "disabled",
        ).pack(side="left", padx=5)
        tk.Button(
            btnf,
            text=lang.t("dispatch_kit.clear", "Clear"),
            bg="#7F8C8D",
            fg="white",
            command=self.clear_form,
        ).pack(side="left", padx=5)
        tk.Button(
            btnf,
            text=lang.t("dispatch_kit.export", "Export"),
            bg="#2980B9",
            fg="white",
            command=self.export_data,
        ).pack(side="left", padx=5)

        self.status_var = tk.StringVar(value=lang.t("dispatch_kit.ready", "Ready"))
        tk.Label(
            main,
            textvariable=self.status_var,
            relief="sunken",
            anchor="w",
            bg="#F0F4F8",
        ).grid(row=10, column=0, columnspan=4, sticky="ew")

        self.load_scenarios()

    # ---------------------------------------------------------
    # Headers + Display Assembly
    # ---------------------------------------------------------
    def _build_with_headers(self, rows):
        """
        Build hierarchical view with kit/module headers.
        ✅ FIXED: Kit rows, then modules (each with items sorted by CODE then EXPIRY).

        Structure:
        - Kit header
        - Kit row (type='Kit')
        - For each module:
        - Module header
        - Module row (type='Module')
        - Item rows (type='Item', sorted by code alphabetically, then expiry FEFO)
        """
        # Separate rows by type and group
        kit_rows = {}  # {kit_number: [kit_rows]}
        module_groups = {}  # {(kit_num, mod_num): [module_rows + item_rows]}

        for idx, item in enumerate(rows):
            item_type = (item.get("type") or "").upper()
            kit_num = item.get("Kit_number") or ""
            mod_num = item.get("module_number") or ""

            if item_type == "KIT":
                # Kit rows go in separate dict
                if kit_num not in kit_rows:
                    kit_rows[kit_num] = []
                kit_rows[kit_num].append((idx, item))

            else:
                # Module and Item rows grouped together by (kit, module)
                key = (kit_num, mod_num)
                if key not in module_groups:
                    module_groups[key] = []
                module_groups[key].append((idx, item, item_type))

        # Sort function for items within a module group
        def sort_within_module(triple):
            """
            Sort by:
            1. Type: Module (0) before Item (1)
            2. Item code (alphabetically) - ✅ NEW
            3. Expiry date (earliest first) for Items
            4. Original index for stability
            """
            idx, item, item_type = triple

            # Type priority
            type_priority = 0 if item_type == "MODULE" else 1

            # ✅ Item code for alphabetical grouping
            item_code = item.get("code", "").upper()

            # Expiry date
            exp_str = item.get("expiry_date", "")
            if not exp_str or exp_str == "":
                exp_date = "9999-12-31"
            else:
                try:
                    if len(exp_str) == 10 and exp_str[4] == "-":
                        exp_date = exp_str
                    else:
                        dt = datetime.strptime(exp_str, "%d-%b-%Y")
                        exp_date = dt.strftime("%Y-%m-%d")
                except:
                    exp_date = "9999-12-31"

            # ✅ UPDATED: Sort by type, then code, then expiry, then index
            return (type_priority, item_code, exp_date, idx)

        # Sort each module group (Module rows first, then Items by code+expiry)
        for key in module_groups:
            module_groups[key].sort(key=sort_within_module)

        # Sort groups by (kit_number, module_number)
        sorted_module_groups = sorted(
            module_groups.items(), key=lambda x: (x[0][0], x[0][1])
        )

        # Build final result
        result = []
        seen_kit = set()
        seen_module = set()
        current_kit = None

        for (kit_num, mod_num), item_triples in sorted_module_groups:
            if not item_triples:
                continue

            # Get first item to extract codes
            first_item = item_triples[0][1]
            kit_code = (
                first_item.get("Kit")
                if first_item.get("Kit") and first_item.get("Kit") != "-----"
                else None
            )
            module_code = (
                first_item.get("module")
                if first_item.get("module") and first_item.get("module") != "-----"
                else None
            )

            # Add Kit header + Kit rows (only once per kit)
            if kit_code and kit_num and kit_num != current_kit:
                current_kit = kit_num

                # Add Kit header
                if (kit_code, kit_num) not in seen_kit:
                    result.append(
                        {
                            "is_header": True,
                            "header_level": "Kit",
                            "code": kit_code,
                            "description": get_item_description(kit_code),
                            "type": "Kit",
                            "Kit": kit_num,
                            "module": "",
                            "current_stock": "",
                            "expiry_date": "",
                            "batch_no": "",
                            "qty_required": "",
                            "unique_id": "",
                            "Kit_number": kit_num,
                            "module_number": None,
                            "treecode": first_item.get("treecode"),
                            "std_qty": None,
                        }
                    )
                    seen_kit.add((kit_code, kit_num))

                # Add Kit rows (type='Kit') right after header
                if kit_num in kit_rows:
                    for idx, kit_item in kit_rows[kit_num]:
                        result.append(kit_item)

            # Add Module header
            if (
                module_code
                and mod_num
                and (kit_code, module_code, mod_num, kit_num) not in seen_module
            ):
                result.append(
                    {
                        "is_header": True,
                        "header_level": "module",
                        "code": module_code,
                        "description": get_item_description(module_code),
                        "type": "Module",
                        "Kit": kit_num or "",
                        "module": mod_num,
                        "current_stock": "",
                        "expiry_date": "",
                        "batch_no": "",
                        "qty_required": "",
                        "unique_id": "",
                        "Kit_number": kit_num,
                        "module_number": mod_num,
                        "treecode": first_item.get("treecode"),
                        "std_qty": None,
                    }
                )
                seen_module.add((kit_code, module_code, mod_num, kit_num))

            # Add Module row + Item rows (sorted: Module first, Items by CODE then EXPIRY)
            for idx, item, item_type in item_triples:
                result.append(item)

        logging.debug(
            f"[BUILD_HEADERS] Built {len(result)} rows (Kits→Modules→Items, code+expiry-sorted)"
        )
        return result

    # ---------------------------------------------------------
    # Mode Rules & Quantity Logic
    # ---------------------------------------------------------
    def get_mode_rules(self):
        mode = self.current_mode_key()
        rules = {
            "editable_types": set(),
            "derive_modules_from_kit": False,
            "derive_items_from_modules": False,
        }
        if mode == "dispatch_kit":
            rules.update(
                {
                    "editable_types": {"Kit"},
                    "derive_modules_from_kit": True,
                    "derive_items_from_modules": True,
                }
            )
        elif mode in ("issue_module_scenario", "issue_module_kit"):
            rules.update(
                {"editable_types": {"Module"}, "derive_items_from_modules": True}
            )
        elif mode in ("issue_standalone", "issue_items_module", "issue_items_kit"):
            rules.update({"editable_types": {"Item"}})
        return rules

    def initialize_quantities_and_highlight(self):
        """
        Initialize quantities and apply visual highlighting to rows.
        ✅ FIXED: Stars added to ALL editable rows (Kit/Module/Item), not just Kit/Module.
        """
        rules = self.get_mode_rules()
        mode_key = self.current_mode_key()
        editable_types_lower = {t.lower() for t in rules.get("editable_types", [])}

        # ✅ Configure tags FIRST
        try:
            self.tree.tag_configure(
                "Kit_header", background="#E3F6E1", font=("Helvetica", 10, "bold")
            )
            self.tree.tag_configure(
                "module_header", background="#E1ECFC", font=("Helvetica", 10, "bold")
            )
            self.tree.tag_configure("Kit_module_highlight", background="#FFF9C4")
            self.tree.tag_configure("editable_row", foreground="#000000")
            self.tree.tag_configure("non_editable", foreground="#666666")
        except Exception as e:
            logging.warning(f"[INIT_QTY] Tag configuration failed: {e}")

        # ✅ STEP 1: Update Kit/Module qty_to_issue (items already have it from populate_rows)
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"):
                continue

            vals = list(self.tree.item(iid, "values"))
            if len(vals) < 10:
                logging.warning(
                    f"[INIT_QTY] Row {iid} has only {len(vals)} columns, skipping"
                )
                continue

            row_type_lower = (vals[2] or "").lower()

            # Only update Kit/Module qty_to_issue (items already set by populate_rows)
            if row_type_lower in ("kit", "module"):
                try:
                    stock = int(vals[5]) if vals[5] else 0
                except Exception:
                    stock = 0

                if row_type_lower == "kit":
                    if mode_key == "dispatch_kit":
                        qty = 1
                    else:
                        qty = 1 if ("kit" in editable_types_lower and stock > 0) else 0
                elif row_type_lower == "module":
                    qty = 1 if ("module" in editable_types_lower and stock > 0) else 0
                else:
                    qty = 0

                vals[9] = str(qty)
                self.tree.item(iid, values=vals)
                logging.debug(
                    f"[INIT_QTY] Updated {row_type_lower} {iid} qty_to_issue={qty}"
                )

        # ✅ STEP 2: Force refresh after quantity updates
        self.tree.update_idletasks()

        # ✅ STEP 3: Derive quantities if needed
        if rules.get("derive_modules_from_kit") and hasattr(
            self, "_derive_modules_from_kits"
        ):
            self._derive_modules_from_kits()
        if rules.get("derive_items_from_modules"):
            self._derive_items_from_modules()

        # ✅ STEP 4: Add stars and apply tags (CRITICAL SECTION)
        stars_added = 0

        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})

            # Handle header rows
            if meta.get("is_header"):
                vals = self.tree.item(iid, "values")
                rt = (vals[2] or "").lower() if len(vals) > 2 else ""
                if rt == "kit":
                    self.tree.item(iid, tags=("Kit_header", "Kit_module_highlight"))
                elif rt == "module":
                    self.tree.item(iid, tags=("module_header", "Kit_module_highlight"))
                logging.debug(f"[INIT_QTY] Header row {iid}: type={rt}")
                continue

            # Handle data rows
            vals = list(self.tree.item(iid, "values"))
            if len(vals) < 10:
                continue

            rt_low = (vals[2] or "").lower()
            tags = []
            stars_added_this_row = False

            # Highlight kits/modules
            if rt_low in ("kit", "module"):
                tags.append("Kit_module_highlight")

            # Check if editable
            is_editable = rt_low in editable_types_lower and meta.get("unique_id")

            if is_editable:
                # ✅ Add star to batch_no (column 7) if allowed
                if rules.get("can_edit_batch"):
                    batch_val = str(vals[7]) if vals[7] else ""
                    if not batch_val.startswith("★"):
                        old_batch = batch_val
                        vals[7] = f"★ {batch_val}"
                        stars_added_this_row = True
                        logging.debug(
                            f"[INIT_QTY] Row {iid} batch: '{old_batch}' → '{vals[7]}'"
                        )

                # ✅ FIXED: Add star to qty_to_issue (column 9) for ALL editable rows (not just Kit/Module)
                qty_val = str(vals[9]) if vals[9] else "0"
                if qty_val and qty_val != "0" and not qty_val.startswith("★"):
                    old_qty = qty_val
                    vals[9] = f"★ {qty_val}"
                    stars_added_this_row = True
                    stars_added += 1
                    logging.debug(
                        f"[INIT_QTY] Row {iid} ({rt_low}) qty: '{old_qty}' → '{vals[9]}'"
                    )

                tags.append("editable_row")
            else:
                tags.append("non_editable")

            # ✅ Update tree (CRITICAL: update both values AND tags together)
            if stars_added_this_row:
                self.tree.item(iid, values=tuple(vals), tags=tuple(tags))
            else:
                self.tree.item(iid, tags=tuple(tags))

        logging.debug(f"[INIT_QTY] Added stars to {stars_added} rows")

        # ✅ STEP 5: AGGRESSIVE MULTI-LEVEL REFRESH
        try:
            self.tree.update_idletasks()
            self.update_idletasks()
            self.tree.update()
            self.update()
            self.after(5, lambda: self.tree.update_idletasks())
            self.after(10, lambda: self.tree.update())
            self.after(20, lambda: self._force_tree_redraw())
            logging.debug("[INIT_QTY] Multi-level refresh completed")
        except Exception as e:
            logging.error(f"[INIT_QTY] Refresh failed: {e}")

        logging.info(
            f"[INIT_QTY] Quantities initialized, {stars_added} stars added to editable cells"
        )

    def _force_tree_redraw(self):
        """Helper to force tree redraw - called with delay."""
        try:
            if self.tree and self.tree.winfo_exists():
                # Get first visible item
                visible = self.tree.get_children()
                if visible:
                    # Force tree to redraw by briefly selecting/deselecting
                    current_selection = self.tree.selection()
                    self.tree.selection_set(visible[0])
                    self.tree.update()
                    if current_selection:
                        self.tree.selection_set(current_selection)
                    else:
                        self.tree.selection_remove(visible[0])
                    self.tree.update()
                logging.debug("[INIT_QTY] Forced tree redraw completed")
        except Exception as e:
            logging.debug(f"[INIT_QTY] Force redraw failed (non-critical): {e}")

    def _derive_modules_from_kits(self):
        """Derive module quantities from kit quantities."""
        kit_quantities = {}

        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"):
                continue

            vals = self.tree.item(iid, "values")
            if (vals[2] or "").lower() == "kit":
                # ✅ FIXED: Column 9 is qty_to_issue now (was 8)
                raw = vals[9]
                if raw.startswith("★"):
                    raw = raw[2:].strip()
                kit_qty = int(raw) if raw.isdigit() else 0
                kit_quantities[meta.get("Kit_number")] = kit_qty

        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"):
                continue

            vals = list(self.tree.item(iid, "values"))
            if (vals[2] or "").lower() == "module":
                base_qty = kit_quantities.get(meta.get("Kit_number"), 0)
                try:
                    stock = int(vals[5]) if vals[5] else 0
                except Exception:
                    stock = 0
                if base_qty > stock:
                    base_qty = stock

                # ✅ FIXED: Column 9 is qty_to_issue now (was 8)
                vals[9] = str(base_qty)
                self.tree.item(iid, values=vals)

    def _reapply_editable_icons(self, rules):
        editable_lower = {t.lower() for t in rules["editable_types"]}

        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            vals = list(self.tree.item(iid, "values"))
            row_type = vals[2] or ""
            rt_low = row_type.lower()

            if meta.get("is_header"):
                if rt_low in ("kit", "module"):
                    base_tag = "Kit_header" if rt_low == "kit" else "module_header"
                    self.tree.item(iid, tags=(base_tag, "Kit_module_highlight"))
                continue

            tags = []
            if rt_low in ("kit", "module"):
                tags.append("Kit_module_highlight")

            if rt_low in editable_lower and meta.get("unique_id"):
                cell_val = vals[9]  # qty_to_issue column
                core = (
                    cell_val[2:].strip()
                    if cell_val.startswith("★")
                    else cell_val.strip()
                )
                if core == "":
                    core = "0"
                vals[8] = f"★ {core}"
                tags.append("editable_row")
                self.tree.item(iid, values=vals, tags=tuple(tags))
            else:
                tags.append("non_editable")
                self.tree.item(iid, tags=tuple(tags))

    def _derive_items_from_modules(self):
        """
        Derive item quantities from module quantities using FEFO logic.
        Groups items by code and distributes quantities starting from earliest expiry.
        """
        # Collect module quantities
        module_quantities = {}
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"):
                continue
            vals = self.tree.item(iid, "values")
            if (vals[2] or "").lower() == "module":
                # ✅ FIXED: Column 9 is qty_to_issue now (was 8)
                raw = vals[9]
                if raw.startswith("★"):
                    raw = raw[2:].strip()
                module_qty = int(raw) if raw.isdigit() else 0

                # Key by (kit_number, module_number) for uniqueness
                key = (meta.get("Kit_number"), meta.get("module_number"))
                module_quantities[key] = module_qty

        # Group items by (kit_number, module_number, item_code)
        item_groups = {}
        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})
            if meta.get("is_header"):
                continue
            vals = list(self.tree.item(iid, "values"))
            if (vals[2] or "").lower() != "item":
                continue

            item_code = vals[0]
            kit_number = meta.get("Kit_number")
            module_number = meta.get("module_number")

            # Get std_qty for this item
            std_qty = meta.get("std_qty")
            if std_qty is None:
                std_qty = 1  # Default

            try:
                current_stock = int(vals[5]) if vals[5] else 0
            except Exception:
                current_stock = 0

            # Group key
            group_key = (kit_number, module_number, item_code)

            if group_key not in item_groups:
                item_groups[group_key] = []

            item_groups[group_key].append(
                {
                    "iid": iid,
                    "expiry_date": vals[6] or "",
                    "current_stock": current_stock,
                    "std_qty": std_qty,
                }
            )

        # Apply FEFO distribution for each item group
        for group_key, item_rows in item_groups.items():
            kit_number, module_number, item_code = group_key

            # Get module quantity for this group
            module_key = (kit_number, module_number)
            module_qty = module_quantities.get(module_key, 0)

            if module_qty == 0:
                # Module qty is 0, set all items to 0
                for row in item_rows:
                    vals = list(self.tree.item(row["iid"], "values"))
                    # ✅ FIXED: Column 9 is qty_to_issue now (was 8)
                    vals[9] = "0"
                    self.tree.item(row["iid"], values=vals)
                continue

            # Calculate required quantity: module_qty * std_qty
            std_qty = item_rows[0]["std_qty"]  # All rows should have same std_qty
            required_qty = module_qty * std_qty

            # Distribute using FEFO
            fefo_distribution = self._distribute_qty_by_fefo(item_rows, required_qty)

            # Apply quantities to tree
            for row in item_rows:
                iid = row["iid"]
                qty = fefo_distribution.get(iid, 0)

                vals = list(self.tree.item(iid, "values"))
                # ✅ FIXED: Column 9 is qty_to_issue now (was 8)
                vals[9] = str(qty)
                self.tree.item(iid, values=vals)

    def _auto_fill_fefo_quantities(self):
        """
        Auto-fill qty_to_issue for items using FEFO logic.
        Distributes quantity required across available stock by earliest expiry.
        """
        # Group items by parent and code
        parent_map = {}  # {parent_iid: {code: [iid_list]}}

        def collect_items(iid):
            """Recursively collect item nodes."""
            if iid not in self.row_data:
                return

            item_type = self.row_data[iid].get("type", "").upper()

            if item_type == "ITEM":
                # Leaf node
                parent_iid = self.tree.parent(iid)
                if parent_iid:
                    code = self.row_data[iid].get("code", "")
                    if parent_iid not in parent_map:
                        parent_map[parent_iid] = {}
                    if code not in parent_map[parent_iid]:
                        parent_map[parent_iid][code] = []
                    parent_map[parent_iid][code].append(iid)
            else:
                # Branch node - recurse
                for child_iid in self.tree.get_children(iid):
                    collect_items(child_iid)

        # Collect all items
        for root_iid in self.tree.get_children():
            collect_items(root_iid)

        # Process each parent's items
        for parent_iid, code_map in parent_map.items():
            for code, iid_list in code_map.items():
                if not iid_list:
                    continue

                # Get std_qty from first item
                first_item = self.row_data.get(iid_list[0], {})
                std_qty = self.extract_std_qty_from_unique_id(
                    first_item.get("unique_id", "")
                )

                # Build stock rows for FEFO distribution
                stock_rows = []
                for iid in iid_list:
                    item = self.row_data.get(iid, {})
                    stock_rows.append(
                        {
                            "iid": iid,
                            "expiry_date": item.get("expiry_date", ""),
                            "current_stock": item.get("current_stock", 0),
                        }
                    )

                # Use existing FEFO helper
                distribution = self._distribute_qty_by_fefo(stock_rows, std_qty)

                # Apply distribution
                for iid, qty in distribution.items():
                    self.tree.set(iid, "qty_to_issue", qty)
                    if iid in self.row_data:
                        self.row_data[iid]["qty_to_issue"] = qty

                logging.debug(
                    f"[FEFO] Distributed {std_qty} for {code}: {distribution}"
                )

    def _calculate_all_qty_required(self):
        """
        Calculate and set qty_required for all item rows in the tree.
        Should be called after qty_to_issue is populated.
        """

        def process_node(iid):
            """Recursively process nodes."""
            if iid not in self.row_data:
                return

            item_type = self.row_data[iid].get("type", "").upper()

            if item_type == "ITEM":
                # Calculate qty_required
                qty_required = self.calculate_qty_required(iid)

                # Update tree
                self.tree.set(iid, "qty_required", qty_required)

                # Update row_data
                self.row_data[iid]["qty_required"] = qty_required

            # Process children
            for child_iid in self.tree.get_children(iid):
                process_node(child_iid)

        # Process all root nodes
        for root_iid in self.tree.get_children():
            process_node(root_iid)

        logging.debug("[QTY_REQUIRED] Calculated qty_required for all items")

    # ---------------------------------------------------------
    # Out Type Dependents
    # ---------------------------------------------------------
    def on_out_type_selected(self, event=None):
        out_type_display = self.trans_type_var.get()
        out_type = self._value_for_out_type(out_type_display)

        third_party_required = {"Out Donation", "Loan", "Return of Borrowing"}
        end_user_required = {"Issue to End User"}
        remarks_required = {
            "Expired Items",
            "Damaged Items",
            "Cold Chain Break",
            "Batch Recall",
            "Theft",
            "Other Losses",
            "Quarantine",
        }

        self.end_user_cb.config(state="disabled")
        self.third_party_cb.config(state="disabled")
        self.remarks_entry.config(state="disabled")

        if out_type not in end_user_required:
            self.end_user_var.set("")
        if out_type not in third_party_required:
            self.third_party_var.set("")
        if out_type not in remarks_required:
            self.remarks_entry.delete(0, tk.END)

        if out_type in end_user_required:
            self.end_user_cb.config(state="readonly")
        if out_type in third_party_required:
            self.third_party_cb.config(state="readonly")
        if out_type in remarks_required:
            self.remarks_entry.config(state="normal")

    # ---------------------------------------------------------
    # Event Handlers / Loading
    # ---------------------------------------------------------
    def on_kit_selected(self, event=None):
        """Handle kit selection - update dependent dropdowns based on mode."""
        mode_key = self.current_mode_key()

        # ✅ Extract code from dropdown display
        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display) if Kit_display else None

        logging.debug(
            f"[KIT_SELECTED] Display: '{Kit_display}' -> Code: '{Kit_code}', mode: {mode_key}"
        )

        # Reset dependent dropdowns
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")

        # Clear table
        self.clear_table_only()

        # ===== Mode-specific logic =====

        if mode_key == "dispatch_kit":
            # Activate kit_number dropdown
            if Kit_code:
                Kit_numbers = self.fetch_available_kit_numbers(
                    self.selected_scenario_id, Kit_code
                )
                self.Kit_number_cb["values"] = Kit_numbers
                self.Kit_number_cb.config(
                    state="readonly" if Kit_numbers else "disabled"
                )
                logging.debug(f"[KIT_SELECTED] Kit numbers available: {Kit_numbers}")
            else:
                self.Kit_number_cb["values"] = []
                self.Kit_number_cb.config(state="disabled")

        elif mode_key == "issue_module_kit":
            # ✅ CRITICAL: Activate BOTH kit_number and module dropdowns
            if Kit_code:
                # 1. Populate kit_number dropdown
                Kit_numbers = self.fetch_available_kit_numbers(
                    self.selected_scenario_id, Kit_code
                )
                self.Kit_number_cb["values"] = Kit_numbers
                self.Kit_number_cb.config(
                    state="readonly" if Kit_numbers else "disabled"
                )
                logging.debug(f"[KIT_SELECTED] Kit numbers available: {Kit_numbers}")

                # 2. Populate module dropdown with secondary modules in this kit
                modules = self.fetch_modules_for_kit(
                    self.selected_scenario_id, Kit_code
                )
                self.module_cb["values"] = modules
                self.module_cb.config(state="readonly" if modules else "disabled")
                logging.debug(f"[KIT_SELECTED] Modules in kit: {modules}")
            else:
                self.Kit_number_cb["values"] = []
                self.Kit_number_cb.config(state="disabled")
                self.module_cb["values"] = []
                self.module_cb.config(state="disabled")

        elif mode_key == "issue_items_kit":
            # Activate kit_number dropdown only
            if Kit_code:
                Kit_numbers = self.fetch_available_kit_numbers(
                    self.selected_scenario_id, Kit_code
                )
                self.Kit_number_cb["values"] = Kit_numbers
                self.Kit_number_cb.config(
                    state="readonly" if Kit_numbers else "disabled"
                )
                logging.debug(f"[KIT_SELECTED] Kit numbers available: {Kit_numbers}")
            else:
                self.Kit_number_cb["values"] = []
                self.Kit_number_cb.config(state="disabled")

        elif mode_key == "issue_items_module":
            # Activate BOTH kit_number and module dropdowns
            if Kit_code:
                # 1. Populate module dropdown
                modules = self.fetch_modules_for_kit(
                    self.selected_scenario_id, Kit_code
                )
                self.module_cb["values"] = modules
                self.module_cb.config(state="readonly" if modules else "disabled")
                logging.debug(f"[KIT_SELECTED] Modules in kit: {modules}")

                # 2. Populate kit_number dropdown
                Kit_numbers = self.fetch_available_kit_numbers(
                    self.selected_scenario_id, Kit_code
                )
                self.Kit_number_cb["values"] = Kit_numbers
                self.Kit_number_cb.config(
                    state="readonly" if Kit_numbers else "disabled"
                )
                logging.debug(f"[KIT_SELECTED] Kit numbers available: {Kit_numbers}")
            else:
                self.module_cb["values"] = []
                self.module_cb.config(state="disabled")
                self.Kit_number_cb["values"] = []
                self.Kit_number_cb.config(state="disabled")

    def on_kit_number_selected(self, event=None):
        """Handle kit number selection - load stock data or update module numbers."""
        mode_key = self.current_mode_key()
        Kit_number = (
            self.Kit_number_var.get().strip() if self.Kit_number_var.get() else None
        )

        # ✅ Extract kit code from dropdown
        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display) if Kit_display else None

        logging.debug(
            f"[KIT_NUMBER_SELECTED] Kit_number={Kit_number}, Kit_code={Kit_code}, mode={mode_key}"
        )

        if not Kit_number:
            self.clear_table_only()
            return

        # ===== Mode-specific logic =====

        if mode_key == "issue_module_kit":
            # Don't load data yet - wait for module_number selection
            # Just log that we're ready
            logging.debug(f"[KIT_NUMBER_SELECTED] Ready - waiting for module selection")
            return

        elif mode_key == "issue_items_module":
            # ✅ Extract module code from dropdown
            module_display = self.module_var.get()
            module_code = (
                self._extract_code_from_display(module_display)
                if module_display
                else None
            )

            if module_code and Kit_number:
                # ✅ Fetch module numbers filtered by Kit_number AND module_code
                module_numbers = self.fetch_module_numbers(
                    self.selected_scenario_id,
                    kit_code=None,  # Not used
                    module_code=module_code,
                    kit_number=Kit_number,
                )
                self.module_number_cb["values"] = module_numbers
                self.module_number_cb.config(
                    state="readonly" if module_numbers else "disabled"
                )
                logging.debug(
                    f"[KIT_NUMBER_SELECTED] Module numbers available: {module_numbers}"
                )
            else:
                self.module_number_cb["values"] = []
                self.module_number_cb.config(state="disabled")
            return

        # For dispatch_kit and issue_items_kit: load stock data
        elif mode_key in ["dispatch_kit", "issue_items_kit"]:
            if Kit_code and Kit_number:
                items = self.fetch_stock_data_for_kit_number(
                    self.selected_scenario_id, Kit_number, Kit_code
                )
                self.populate_rows(
                    items,
                    lang.t(
                        "dispatch_kit.loaded_kit_stock",
                        "Loaded {count} items for kit {Kit_number}",
                        count=len(items),
                        Kit_number=Kit_number,
                    ),
                )
            else:
                self.clear_table_only()

    def on_module_selected(self, event=None):
        """
        Handle module selection - update module_number dropdown based on mode.
        Supports both:
        - Modules inside kits (kit-based flow)
        - Standalone primary modules (direct selection)
        """
        mode_key = self.current_mode_key()

        # ✅ Extract codes from dropdowns
        module_display = self.module_var.get()
        module_code = (
            self._extract_code_from_display(module_display) if module_display else None
        )

        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display) if Kit_display else None

        # ✅ Get kit_number if available (critical for filtering!)
        Kit_number = (
            self.Kit_number_var.get().strip() if self.Kit_number_var.get() else None
        )

        logging.debug(
            f"[MODULE_SELECTED] module_code={module_code}, Kit_code={Kit_code}, Kit_number={Kit_number}, mode={mode_key}"
        )

        # Reset module number
        self.module_number_var.set("")

        # ===== For issue_module_kit mode =====
        if mode_key == "issue_module_kit":
            # ✅ CRITICAL: Filter module_numbers by kit_number AND module_code
            if module_code:
                # Fetch module numbers filtered by kit_number (if selected) and module_code
                module_numbers = self.fetch_module_numbers(
                    self.selected_scenario_id,
                    kit_code=None,  # Not used in the query
                    module_code=module_code,
                    kit_number=Kit_number,  # ✅ This is the key filter!
                )
                self.module_number_cb["values"] = module_numbers
                self.module_number_cb.config(
                    state="readonly" if module_numbers else "disabled"
                )

                if Kit_number:
                    logging.debug(
                        f"[MODULE_SELECTED] Module numbers for kit_number={Kit_number}, module={module_code}: {module_numbers}"
                    )
                else:
                    logging.debug(
                        f"[MODULE_SELECTED] Module numbers for module={module_code} (all kits): {module_numbers}"
                    )
            else:
                self.module_number_cb["values"] = []
                self.module_number_cb.config(state="disabled")

        # ===== For issue_items_module mode =====
        elif mode_key == "issue_items_module":
            if not module_code:
                # No module selected yet
                self.module_number_cb["values"] = []
                self.module_number_cb.config(state="disabled")
                logging.debug(f"[MODULE_SELECTED] No module selected yet")

            elif Kit_code and Kit_number:
                # ✅ Flow A: Module inside a kit (existing behavior)
                # Fetch module numbers filtered by Kit_number AND module_code
                module_numbers = self.fetch_module_numbers(
                    self.selected_scenario_id,
                    kit_code=None,  # Not used
                    module_code=module_code,
                    kit_number=Kit_number,
                )
                self.module_number_cb["values"] = module_numbers
                self.module_number_cb.config(
                    state="readonly" if module_numbers else "disabled"
                )
                logging.debug(
                    f"[MODULE_SELECTED] Kit-based: {len(module_numbers)} module numbers for kit_number={Kit_number}, module={module_code}"
                )

            elif module_code and not Kit_code:
                # ✅ Flow B: Standalone primary module (NEW!)
                # User selected module directly without selecting a kit first
                module_numbers = self.fetch_module_numbers_standalone(
                    self.selected_scenario_id, module_code
                )
                self.module_number_cb["values"] = module_numbers
                self.module_number_cb.config(
                    state="readonly" if module_numbers else "disabled"
                )
                logging.debug(
                    f"[MODULE_SELECTED] Standalone: {len(module_numbers)} module numbers for primary module={module_code}"
                )

            else:
                # Kit selected but no kit_number yet, or other incomplete state
                self.module_number_cb["values"] = []
                self.module_number_cb.config(state="disabled")
                logging.debug(
                    f"[MODULE_SELECTED] Waiting for kit_number selection or standalone module"
                )

        # Clear table until module number is selected
        self.clear_table_only()

    def on_module_number_selected(self, event=None):
        """Handle module number selection - works for both kit-based and standalone."""
        mode_key = self.current_mode_key()
        module_number = (
            self.module_number_var.get().strip()
            if self.module_number_var.get()
            else None
        )

        if mode_key not in [
            "issue_items_module",
            "issue_module_kit",
            "issue_module_scenario",
        ]:
            return

        if not module_number:
            self.clear_table_only()
            return

        # Extract codes
        Kit_display = self.Kit_var.get()
        Kit_code = self._extract_code_from_display(Kit_display) if Kit_display else None

        module_display = self.module_var.get()
        module_code = (
            self._extract_code_from_display(module_display) if module_display else None
        )

        logging.debug(
            f"[MODULE_NUMBER_SELECTED] module_number={module_number}, kit={Kit_code}, module={module_code}"
        )

        # Load stock data (works for both kit-based and standalone)
        items = self.fetch_stock_data_for_module_number(
            self.selected_scenario_id,
            module_number,
            Kit_code,  # Will be None for standalone
            module_code,
        )

        self.populate_rows(
            items,
            lang.t(
                "dispatch_kit.loaded_module_stock",
                "Loaded {count} items for module {module_number}",
                count=len(items),
                module_number=module_number,
            ),
        )

    def populate_standalone_items(self):
        if not self.selected_scenario_id:
            return
        items = self.fetch_standalone_stock_items(self.selected_scenario_id)
        self.full_items = items[:]
        self.populate_rows(
            self.full_items,
            lang.t(
                "dispatch_kit.loaded_standalone", "Loaded {n} standalone item rows"
            ).format(n=len(self.full_items)),
        )

    # ---------------------------------------------------------
    # Table Helpers
    # ---------------------------------------------------------
    def clear_table_only(self):
        if self.tree:
            self.tree.delete(*self.tree.get_children())
        self.row_data.clear()

    def populate_rows(self, items=None, status_msg="", update_cache=True):
        """
        Populate the tree with items, applying FEFO logic for qty_to_issue.
        ✅ FIXED: No global sorting - hierarchy preserved, FEFO applied per module.
        """
        if items is None:
            items = []

        # Cache items for search restore
        if update_cache:
            self.full_items = items.copy()
            logging.debug(
                f"[POPULATE_ROWS] Cached {len(items)} items to self.full_items"
            )

        # Clear existing rows
        self.clear_table_only()
        self.row_data = {}

        if not items:
            if status_msg:
                self.status_var.set(status_msg)
            else:
                self.status_var.set(
                    lang.t("dispatch_kit.no_items", "No items to display")
                )
            return

        # ✅ NO GLOBAL SORTING - preserve order from database
        logging.debug(
            f"[POPULATE_ROWS] Processing {len(items)} items without global sorting"
        )

        # Group items by (kit_number, module_number, code) for FEFO calculation
        fefo_groups = {}
        for item in items:
            kit_num = item.get("Kit_number", "")
            mod_num = item.get("module_number", "")
            code = item.get("code", "")
            key = (kit_num, mod_num, code)

            if key not in fefo_groups:
                fefo_groups[key] = []
            fefo_groups[key].append(item)

        # Calculate qty_to_issue and qty_required using FEFO within each group
        for (kit_num, mod_num, code), group_items in fefo_groups.items():
            # Get std_qty from first item
            std_qty = group_items[0].get("std_qty")
            if std_qty is None or std_qty <= 0:
                # Not an item, skip FEFO calculation
                for item in group_items:
                    item["qty_to_issue"] = 0
                    item["qty_required"] = ""
                continue

            # Sort group by expiry date (FEFO within this module/code)
            def expiry_key(it):
                """Sort by expiry date only - preserve original order when dates equal."""
                exp_str = it.get("expiry_date", "")
                if not exp_str:
                    return "9999-12-31"
                try:
                    if len(exp_str) == 10 and exp_str[4] == "-":
                        return exp_str
                    dt = datetime.strptime(exp_str, "%d-%b-%Y")
                    return dt.strftime("%Y-%m-%d")
                except:
                    return "9999-12-31"

            group_items.sort(key=expiry_key)  # ✅ Only sorts by expiry date

            # Apply FEFO: distribute std_qty across items with earliest expiry first
            remaining = std_qty
            for item in group_items:
                current_stock = item.get("current_stock", 0)

                # Calculate qty_required (how much is still needed)
                qty_required = max(0, remaining)
                item["qty_required"] = qty_required

                # Calculate qty_to_issue (how much from this batch)
                if remaining <= 0:
                    qty_to_issue = 0
                elif current_stock >= remaining:
                    qty_to_issue = remaining
                    remaining = 0
                else:
                    qty_to_issue = current_stock
                    remaining -= current_stock

                item["qty_to_issue"] = qty_to_issue

                logging.debug(
                    f"[QTY_REQUIRED] {code}: std_qty={std_qty}, "
                    f"qty_required={qty_required}, qty_to_issue={qty_to_issue}"
                )

        # Build hierarchical structure with headers (preserves order, adds headers)
        structured = self._build_with_headers(items)

        # Insert into tree
        for item in structured:
            code = item.get("code", "")
            desc = item.get("description", "")
            item_type = item.get("type", "")
            kit_val = item.get("Kit", "")
            module_val = item.get("module", "")
            current_stock = item.get("current_stock", "")
            expiry_date = item.get("expiry_date", "")
            batch_no = item.get("batch_no", "")
            unique_id = item.get("unique_id", "")

            # ✅ Format qty_required (safe handling - FIXED duplicate try-except)
            qty_required = item.get("qty_required", "")
            if qty_required == "" or qty_required is None:
                qty_required_display = ""
            else:
                try:
                    qty_required_display = str(int(qty_required))
                except (ValueError, TypeError):
                    qty_required_display = ""

            # Format qty_to_issue (NO STAR YET - added in initialize_quantities_and_highlight)
            qty_to_issue = item.get("qty_to_issue", 0)
            try:
                qty_to_issue_display = str(int(qty_to_issue))
            except (ValueError, TypeError):
                qty_to_issue_display = "0"

            # Insert row with 11 columns (NO STARS YET)
            iid = self.tree.insert(
                "",
                "end",
                values=(
                    code,
                    desc,
                    item_type,
                    kit_val,
                    module_val,
                    current_stock,
                    expiry_date,
                    batch_no,
                    qty_required_display,
                    qty_to_issue_display,  # ✅ NO STAR - will be added by initialize_quantities_and_highlight
                    unique_id,
                ),
            )

            # Store metadata
            self.row_data[iid] = {
                "unique_id": unique_id,
                "code": code,
                "description": desc,
                "type": item_type,
                "Kit_number": item.get("Kit_number"),
                "module_number": item.get("module_number"),
                "current_stock": current_stock,
                "qty_to_issue": qty_to_issue,
                "qty_required": qty_required,
                "is_header": item.get("is_header", False),
                "std_qty": item.get("std_qty"),
                "row_type": item_type,  # ✅ Added for initialize_quantities_and_highlight
            }

        # ✅ Initialize quantities and highlighting (THIS ADDS THE STARS)
        self.initialize_quantities_and_highlight()

        # Set status
        if status_msg:
            self.status_var.set(status_msg)
        else:
            self.status_var.set(
                lang.t("dispatch_kit.items_loaded", "Loaded {count} items").format(
                    count=len(items)
                )
            )

        logging.info(
            f"[POPULATE_ROWS] Populated {len(items)} items with hierarchy preserved"
        )

    def _extract_std_qty_from_unique_id(self, unique_id: str) -> int:
        """
        Extract std_qty from unique_id.
        Format: scenario/kit/module/item/std_qty/exp_date/kit_number/module_number

        Args:
            unique_id: The unique identifier string

        Returns:
            std_qty as integer (defaults to 1 if not found or invalid)
        """
        if not unique_id:
            return 1

        parts = unique_id.split("/")
        if len(parts) < 5:
            return 1

        try:
            std_qty = int(parts[4])
            return std_qty if std_qty > 0 else 1
        except (ValueError, IndexError):
            return 1

    def get_selected_unique_ids(self):
        uids = []
        for iid in self.tree.selection():
            hidden_val = self.tree.set(iid, "unique_id")
            if hidden_val:
                uids.append(hidden_val)
            else:
                meta = self.row_data.get(iid, {})
                if meta.get("unique_id"):
                    uids.append(meta["unique_id"])
        return uids

    def _debug_log_items(self, label, items, limit=12):
        logging.info(f"[DISPATCH][DEBUG] {label}: total={len(items)}")
        for i, it in enumerate(items[:limit]):
            logging.info(
                f"[DISPATCH][DEBUG] {label} row{i}: "
                f"unique_id={it.get('unique_id')} code={it.get('code')} "
                f"type={it.get('type')} Kit={it.get('Kit')} module={it.get('module')} "
                f"Kit_no={it.get('Kit_number')} module_no={it.get('module_number')} std_qty={it.get('std_qty')}"
            )

    # ---------------------------------------------------------
    # Editing
    # ---------------------------------------------------------
    def _flatten_rows(self):
        out = []

        def dive(iids):
            for r in iids:
                out.append(r)
                dive(self.tree.get_children(r))

        dive(self.tree.get_children())
        return out

    def navigate_tree(self, event):
        """
        Navigate tree with arrow keys and start editing with Enter/Tab.
        """
        if self.editing_cell:
            return

        rows = self._flatten_rows()
        if not rows:
            return

        sel = self.tree.selection()
        if not sel:
            self.tree.selection_set(rows[0])
            self.tree.focus(rows[0])
            self.start_edit_cell(rows[0], 9)
            return

        cur = sel[0]
        try:
            idx = rows.index(cur)
        except ValueError:
            return

        if event.keysym == "Up" and idx > 0:
            self.tree.selection_set(rows[idx - 1])
            self.tree.focus(rows[idx - 1])
            self.tree.see(rows[idx - 1])
        elif event.keysym == "Down" and idx < len(rows) - 1:
            self.tree.selection_set(rows[idx + 1])
            self.tree.focus(rows[idx + 1])
            self.tree.see(rows[idx + 1])
        elif event.keysym in ("Return", "Tab"):
            self.start_edit_cell(cur, 9)

    def start_edit(self, event):
        """
        Handle editing trigger from double-click or keyboard.
        Only allows editing qty_to_issue column (column 9, index 9).
        """
        if event.type == tk.EventType.KeyPress:
            sel = self.tree.selection()
            if not sel:
                return
            # For keyboard, edit currently selected row's qty_to_issue column
            self.start_edit_cell(sel[0], 9)
            return

        # For mouse double-click
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)

        if not row_id or not col_id:
            return

        col_index = int(col_id.replace("#", "")) - 1

        # Only allow editing qty_to_issue column (index 9)
        if col_index != 9:
            return

        self.start_edit_cell(row_id, 9)

    def start_edit_cell(self, row_id, col_index):
        """
        Start inline editing for a cell.
        Only works for qty_to_issue column (index 9) on editable rows.
        """
        # Only allow editing qty_to_issue column
        if col_index != 9:
            logging.debug(
                f"[EDIT] Column {col_index} not editable (only col 9 allowed)"
            )
            return

        rules = self.get_mode_rules()
        editable_lower = {t.lower() for t in rules["editable_types"]}
        meta = self.row_data.get(row_id, {})

        # Don't edit headers
        if meta.get("is_header"):
            logging.debug(f"[EDIT] Row {row_id} is header, not editable")
            return

        # Don't edit rows without unique_id
        if not meta.get("unique_id"):
            logging.debug(f"[EDIT] Row {row_id} has no unique_id, not editable")
            return

        vals = self.tree.item(row_id, "values")
        rt_low = (vals[2] or "").lower()

        # Check if row type is editable in current mode
        if rt_low not in editable_lower:
            logging.debug(
                f"[EDIT] Row type '{rt_low}' not editable in current mode. Editable types: {editable_lower}"
            )
            return

        # Get cell position
        bbox = self.tree.bbox(row_id, f"#{col_index + 1}")
        if not bbox:
            logging.debug(f"[EDIT] Cell not visible")
            return

        x, y, w, h = bbox

        # Get current value (strip star if present)
        raw_old = self.tree.set(row_id, "qty_to_issue")
        old_clean = raw_old[2:].strip() if raw_old.startswith("★") else raw_old.strip()

        logging.debug(
            f"[EDIT] Starting edit for {vals[0]} (type={rt_low}), current value={old_clean}"
        )

        # Close any existing editor
        if self.editing_cell:
            try:
                self.editing_cell.destroy()
            except Exception:
                pass

        # Create entry widget
        entry = tk.Entry(self.tree, font=("Helvetica", 10), background="#FFFBE0")
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, old_clean if old_clean else "")
        entry.focus()
        entry.select_range(0, tk.END)
        self.editing_cell = entry

        def set_status(msg):
            if self.status_var:
                self.status_var.set(msg)

        def save(_=None):
            val = entry.get().strip()
            try:
                stock = int(vals[5]) if vals[5] else 0
            except Exception:
                stock = 0

            # Validate based on row type
            if rt_low in ("kit", "module"):
                # Kit/Module: only allow 0 or 1
                if val not in ("0", "1"):
                    set_status(
                        lang.t(
                            "dispatch_kit.msg_qty_binary",
                            "Only 0 or 1 allowed – auto-corrected.",
                        )
                    )
                    val = (
                        old_clean
                        if old_clean in ("0", "1")
                        else ("1" if stock > 0 else "0")
                    )
                if stock == 0 and val == "1":
                    val = "0"
            else:  # item
                # Item: allow any non-negative integer up to stock
                if not val.isdigit():
                    set_status(
                        lang.t(
                            "dispatch_kit.msg_invalid_number",
                            "Invalid number – set to 0.",
                        )
                    )
                    val = "0"
                else:
                    iv = int(val)
                    if iv < 0:
                        iv = 0
                        set_status(
                            lang.t(
                                "dispatch_kit.msg_negative",
                                "Negative not allowed – set to 0.",
                            )
                        )
                    if iv > stock:
                        iv = stock
                        set_status(
                            lang.t(
                                "dispatch_kit.msg_exceeded_stock",
                                "Exceeded stock – capped.",
                            )
                        )
                    val = str(iv)

            # Update tree with star for editable types
            self.tree.set(row_id, "qty_to_issue", f"★ {val}")
            entry.destroy()
            self.editing_cell = None

            logging.debug(f"[EDIT] Saved {vals[0]}: {old_clean} → {val}")

            # Trigger cascading updates if needed
            if rt_low == "kit" and rules.get("derive_modules_from_kit"):
                self._derive_modules_from_kits()
                if rules.get("derive_items_from_modules"):
                    self._derive_items_from_modules()
                self._reapply_editable_icons(rules)
            elif rt_low == "module" and rules.get("derive_items_from_modules"):
                self._derive_items_from_modules()
                self._reapply_editable_icons(rules)

        # Bind save triggers
        entry.bind("<Return>", save)
        entry.bind("<KP_Enter>", save)  # Numpad Enter
        entry.bind("<Tab>", save)
        entry.bind("<FocusOut>", save)
        entry.bind(
            "<Escape>", lambda _: (entry.destroy(), setattr(self, "editing_cell", None))
        )

    # ---------------------------------------------------------
    # Search
    # ---------------------------------------------------------
    def search_items(self, event=None):
        """
        Dynamic search - filters items as user types.
        If search is empty, restore original full_items.
        """
        query = self.search_var.get().strip()

        # If search is empty, restore full items
        if not query:
            logging.debug("[SEARCH_ITEMS] Empty query - restoring full items")
            if hasattr(self, "full_items") and self.full_items:
                # ✅ update_cache=False because we're restoring, not loading new data
                self.populate_rows(
                    self.full_items,
                    lang.t("dispatch_kit.restored", "Restored all items"),
                    update_cache=False,
                )
            return

        # If no full_items cache, nothing to search
        if not hasattr(self, "full_items") or not self.full_items:
            logging.debug("[SEARCH_ITEMS] No full_items cache available")
            return

        # Filter items by query (search in code and description)
        query_lower = query.lower()
        filtered = []

        for item in self.full_items:
            code = item.get("code", "").lower()
            desc = item.get("description", "").lower()

            if query_lower in code or query_lower in desc:
                filtered.append(item)

        logging.debug(
            f"[SEARCH_ITEMS] Query='{query}' matched {len(filtered)}/{len(self.full_items)} items"
        )

        # Repopulate with filtered items
        if filtered:
            # ✅ update_cache=False to preserve original full_items
            self.populate_rows(
                filtered,
                lang.t(
                    "dispatch_kit.found_items",
                    "Found {count} items",
                    count=len(filtered),
                ),
                update_cache=False,
            )
        else:
            self.clear_table_only()
            self.status_var.set(
                lang.t(
                    "dispatch_kit.no_matches", "No items match '{query}'", query=query
                )
            )

    # ---------------------------------------------------------
    # Save (Issue)
    # ---------------------------------------------------------
    def save_all(self):
        """
        Save/issue all items that have qty_to_issue > 0.
        ✅ Complete version with proper 11-column unpacking and validation.
        """
        logging.info("[DISPATCH] save_all called")

        # Role validation
        if self.role not in ["admin", "manager", "supervisor"]:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "dispatch_kit.no_permission",
                    "Only admin or manager roles can save changes.",
                ),
                "error",
            )
            return

        # Get and validate out_type
        out_type_display = (self.trans_type_var.get() or "").strip()
        out_type = self._value_for_out_type(out_type_display)
        if not out_type:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("dispatch_kit.no_out_type", "OUT Type is mandatory."),
                "error",
            )
            return

        # Define required fields by out_type
        third_party_required = {"Out Donation", "Loan", "Return of Borrowing"}
        end_user_required = {"Issue to End User"}
        remarks_required = {
            "Expired Items",
            "Damaged Items",
            "Cold Chain Break",
            "Batch Recall",
            "Theft",
            "Other Losses",
            "Quarantine",
        }

        # Get form values
        end_user = (self.end_user_var.get() or "").strip()
        third_party = (self.third_party_var.get() or "").strip()
        remarks = (self.remarks_entry.get() or "").strip()

        # Validate required fields based on out_type
        if out_type in end_user_required and not end_user:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "dispatch_kit.err_end_user_required",
                    "End User is required for this Out Type.",
                ),
                "error",
            )
            return

        if out_type in third_party_required and not third_party:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "dispatch_kit.err_third_party_required",
                    "Third Party is required for this Out Type.",
                ),
                "error",
            )
            return

        if out_type in remarks_required and len(remarks) < 3:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "dispatch_kit.err_remarks_required",
                    "Remarks are required (min 3 chars) for this Out Type.",
                ),
                "error",
            )
            return

        # Collect rows to issue
        rows_to_issue = []

        for iid in self.tree.get_children():
            meta = self.row_data.get(iid, {})

            # Skip header rows
            if meta.get("is_header"):
                continue

            vals = list(self.tree.item(iid, "values"))
            if not vals:
                continue

            # ✅ CRITICAL: Tree has 11 columns!
            # Columns: Code, Description, Type, Kit, Module, Stock, Expiry, Batch, Qty_Required, Qty_to_Issue, unique_id
            if len(vals) < 11:
                logging.warning(
                    f"[SAVE_ALL] Row {iid} has {len(vals)} columns (expected 11), skipping"
                )
                continue

            # ✅ UNPACK 11 VALUES (added qty_required between batch_no and qty_to_issue)
            (
                code,
                desc,
                type_field,
                kit_col,
                module_col,
                current_stock,
                exp_date,
                batch_no,
                qty_required,
                qty_to_issue,
                tree_unique_id,
            ) = vals

            # Get unique_id from metadata (more reliable than tree column)
            unique_id = meta.get("unique_id") or tree_unique_id
            if not unique_id:
                logging.warning(f"[SAVE_ALL] Row {iid} has no unique_id, skipping")
                continue

            # ✅ Parse qty_to_issue (remove star if present)
            raw_q = str(qty_to_issue).replace("★", "").strip()
            if not raw_q or not raw_q.isdigit():
                continue

            q_int = int(raw_q)
            if q_int <= 0:
                continue

            # ✅ Parse current_stock (remove star if present)
            try:
                stock_str = str(current_stock).replace("★", "").strip()
                stock_int = int(stock_str) if stock_str.isdigit() else 0
            except Exception:
                stock_int = 0

            # Collect row data
            rows_to_issue.append(
                {
                    "iid": iid,
                    "code": code,
                    "description": desc,
                    "type": type_field,
                    "kit": kit_col if kit_col != "-----" else None,
                    "module": module_col if module_col != "-----" else None,
                    "current_stock": stock_int,
                    "qty_to_issue": q_int,
                    "expiry_date": exp_date if exp_date else None,
                    "batch_no": batch_no if batch_no else None,
                    "unique_id": unique_id,
                    "metadata": meta,
                }
            )

        # Check if any items to issue
        if not rows_to_issue:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("dispatch_kit.no_issue_qty", "No quantities entered to issue."),
                "error",
            )
            return

        # Validate stock availability
        over_issue = [
            r["code"]
            for r in rows_to_issue
            if r["qty_to_issue"] > r["current_stock"] and r["current_stock"] > 0
        ]
        if over_issue:
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t(
                    "dispatch_kit.over_issue",
                    "Cannot issue more than stock for: {list}",
                ).format(list=", ".join(over_issue)),
                "error",
            )
            return

        # Get scenario and movement type
        scenario_name = self.scenario_map.get(self.selected_scenario_id, "")
        movement_label = self.mode_var.get()  # Get display label from dropdown
        movement_type_canonical = self._canon_movement_type(
            movement_label
        )  # ✅ Convert to English

        # Generate document number
        doc_number = self.generate_document_number(out_type)
        self.status_var.set(
            lang.t(
                "dispatch_kit.pending_dispatch",
                "Pending dispatch... Document Number: {doc}",
            ).format(doc=doc_number)
        )

        # Process transactions with retry logic
        import time

        max_attempts = 4

        for attempt in range(1, max_attempts + 1):
            conn = connect_db()
            if conn is None:
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("dispatch_kit.db_error", "Database connection failed"),
                    "error",
                )
                return

            try:
                conn.execute("PRAGMA busy_timeout=5000;")
                cur = conn.cursor()
                now_date = datetime.today().strftime("%Y-%m-%d")
                now_time = datetime.now().strftime("%H:%M:%S")

                # Process each row
                for row in rows_to_issue:
                    # Verify stock availability (concurrency check)
                    cur.execute(
                        """
                        SELECT final_qty FROM stock_data WHERE unique_id = ?
                    """,
                        (row["unique_id"],),
                    )

                    db_row = cur.fetchone()
                    if (
                        not db_row
                        or db_row[0] is None
                        or db_row[0] < row["qty_to_issue"]
                    ):
                        raise ValueError(
                            f"Insufficient stock or concurrency issue for {row['code']}"
                        )

                    # Update stock_data
                    cur.execute(
                        """
                        UPDATE stock_data
                        SET qty_out = qty_out + ?,
                            updated_at = ?
                        WHERE unique_id = ?
                        AND (qty_in - qty_out) >= ?
                    """,
                        (
                            row["qty_to_issue"],
                            f"{now_date} {now_time}",
                            row["unique_id"],
                            row["qty_to_issue"],
                        ),
                    )

                    if cur.rowcount == 0:
                        raise ValueError(
                            f"Concurrent change or insufficient stock for {row['code']}"
                        )

                    # Get kit_number and module_number from metadata
                    rd = row["metadata"]
                    kit_number = (
                        rd.get("Kit_number")
                        or rd.get("kit_number")
                        or row["kit"]
                        or None
                    )
                    module_number = rd.get("module_number") or row["module"] or None

                    # Insert transaction record
                    self._insert_transaction_issue(
                        cur,
                        unique_id=row["unique_id"],
                        code=row["code"],
                        description=row["description"],
                        expiry_date=row["expiry_date"],
                        batch_number=row["batch_no"],
                        scenario=scenario_name,
                        kit_number=kit_number,
                        module_number=module_number,
                        qty_out=row["qty_to_issue"],
                        out_type=out_type,
                        third_party=third_party if third_party else None,
                        end_user=end_user if end_user else None,
                        remarks=remarks if remarks else None,
                        movement_type=movement_type_canonical,
                        ts_date=now_date,
                        ts_time=now_time,
                        document_number=doc_number,
                    )

                # Commit transaction
                conn.commit()

                # Success message
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.success", "Success"),
                    lang.t("dispatch_kit.issue_success", "Stock issued successfully."),
                    "info",
                )

                self.status_var.set(
                    lang.t(
                        "dispatch_kit.issue_complete",
                        "Issue complete. Document Number: {doc}",
                    ).format(doc=doc_number)
                )

                # Ask for Excel export
                if (
                    custom_askyesno(
                        self.parent,
                        lang.t("dialog_titles.confirm", "Confirm"),
                        lang.t(
                            "dispatch_kit.ask_export",
                            "Do you want to export the issuance to Excel?",
                        ),
                    )
                    == "yes"
                ):
                    # Prepare export data
                    export_tuples = [
                        (
                            r["iid"],
                            r["code"],
                            r["description"],
                            r["current_stock"],
                            r["qty_to_issue"],
                            r["expiry_date"],
                            r["batch_no"],
                        )
                        for r in rows_to_issue
                    ]
                    self.export_data(export_tuples)

                # Clear form
                self.clear_form()
                return

            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower():
                    logging.warning(
                        f"[DISPATCH] Database locked attempt {attempt}/{max_attempts}; retrying..."
                    )
                    try:
                        conn.rollback()
                    except:
                        pass
                    time.sleep(0.8 * attempt)
                else:
                    try:
                        conn.rollback()
                    except:
                        pass
                    logging.error(f"[DISPATCH] Issue failed: {e}")
                    custom_popup(
                        self.parent,
                        lang.t("dialog_titles.error", "Error"),
                        lang.t(
                            "dispatch_kit.issue_failed", "Issue failed: {err}"
                        ).format(err=e),
                        "error",
                    )
                    return

            except Exception as e:
                try:
                    conn.rollback()
                except:
                    pass
                logging.error(f"[DISPATCH] Issue failed: {e}")
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t("dispatch_kit.issue_failed", "Issue failed: {err}").format(
                        err=e
                    ),
                    "error",
                )
                return

            finally:
                try:
                    cur.close()
                except:
                    pass
                try:
                    conn.close()
                except:
                    pass

        # Max attempts reached
        custom_popup(
            self.parent,
            lang.t("dialog_titles.error", "Error"),
            lang.t(
                "dispatch_kit.issue_failed_locked",
                "Issue failed: database remained locked.",
            ),
            "error",
        )

    # ---------------------------------------------------------
    # Helper: insert transaction using existing cursor
    # ---------------------------------------------------------
    def _insert_transaction_issue(
        self,
        cur,
        *,
        unique_id,
        code,
        description,
        expiry_date,
        batch_number,
        scenario,
        kit_number,
        module_number,
        qty_out,
        out_type,
        third_party,
        end_user,
        remarks,
        movement_type,
        ts_date,
        ts_time,
        document_number,
    ):
        cur.execute(
            """
            INSERT INTO stock_transactions
            (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number,
             Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type,
             Third_Party, End_User, Remarks, Movement_Type, document_number)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
            (
                ts_date,
                ts_time,
                unique_id,
                code,
                description,
                expiry_date,
                batch_number,
                scenario,
                kit_number,
                module_number,
                None,
                None,
                qty_out,
                out_type,
                third_party,
                end_user,
                remarks,
                movement_type,
                document_number,
            ),
        )

    # ---------------------------------------------------------
    # Generate Document Number
    # ---------------------------------------------------------
    def generate_document_number(self, out_type_text: str) -> str:
        project_name, project_code = fetch_project_details()
        project_code = (project_code or "PRJ").strip().upper()

        base_map = {
            "Issue to End User": "OEU",
            "Expired Items": "OEXP",
            "Damaged Items": "ODMG",
            "Cold Chain Break": "OCCB",
            "Batch Recall": "OBRC",
            "Theft": "OTHF",
            "Other Losses": "OLS",
            "Out Donation": "ODN",
            "Loan": "OLOAN",
            "Return of Borrowing": "OROB",
            "Quarantine": "OQRT",
        }

        raw = (out_type_text or "").strip()
        import re

        norm_raw = re.sub(r"[^a-z0-9]+", "", raw.lower())
        abbr = None
        for k, v in base_map.items():
            if re.sub(r"[^a-z0-9]+", "", k.lower()) == norm_raw:
                abbr = v
                break

        if not abbr:
            stop = {"OF", "FROM", "THE", "AND", "DE", "DU", "DES", "LA", "LE", "LES"}
            parts = []
            for token in re.split(r"\s+", raw.upper()):
                if not token or token in stop:
                    continue
                if token == "MSF":
                    parts.append("MSF")
                else:
                    parts.append(token[0])
            if not parts:
                abbr = (raw[:4].upper() or "DOC").replace(" ", "")
            else:
                abbr = "".join(parts)
            if len(abbr) > 8:
                abbr = abbr[:8]

        now = datetime.now()
        prefix = f"{now.year:04d}/{now.month:02d}/{project_code}/{abbr}"

        conn = connect_db()
        serial_num = 1
        if conn is not None:
            cur = conn.cursor()
            try:
                cur.execute(
                    """
                    SELECT document_number
                      FROM stock_transactions
                     WHERE document_number LIKE ?
                     ORDER BY document_number DESC
                     LIMIT 1
                """,
                    (prefix + "/%",),
                )
                row = cur.fetchone()
                if row and row[0]:
                    last_serial = row[0].rsplit("/", 1)[-1]
                    if last_serial.isdigit():
                        serial_num = int(last_serial) + 1
            except Exception as e:
                logging.error(f"[DISPATCH] Error fetching last document_number: {e}")
            finally:
                cur.close()
                conn.close()

        document_number = f"{prefix}/{serial_num:04d}"
        self.current_document_number = document_number
        return document_number

    # ---------------------------------------------------------
    # Utility / Clear / Export
    # ---------------------------------------------------------
    def clear_search(self):
        """Clear search and restore the original full tree."""
        self.search_var.set("")

        # Restore the original full tree if it exists
        if hasattr(self, "full_items") and self.full_items:
            logging.debug(
                f"[CLEAR_SEARCH] Restoring {len(self.full_items)} original items"
            )
            # Trigger search_items which will restore with update_cache=False
            self.search_items()
        else:
            logging.debug("[CLEAR_SEARCH] No full_items cache to restore")
            self.status_var.set(lang.t("dispatch_kit.ready", "Ready"))

    def clear_form(self):
        self.clear_table_only()
        self.scenario_var.set("")
        self.mode_var.set("")
        self.Kit_var.set("")
        self.Kit_number_var.set("")
        self.module_var.set("")
        self.module_number_var.set("")
        self.trans_type_var.set("")
        self.end_user_var.set("")
        self.third_party_var.set("")
        self.remarks_entry.config(state="normal")
        self.remarks_entry.delete(0, tk.END)
        self.remarks_entry.config(state="disabled")
        self.status_var.set(lang.t("dispatch_kit.ready", "Ready"))
        self.scenario_map = self.fetch_scenario_map()
        self.load_scenarios()

    def edit_qty_to_issue_popup(self, row_id, max_stock):
        """
        Open dialog to edit qty_to_issue for an item row.
        Uses native Tkinter dialog for maximum compatibility.

        Args:
            row_id: Tree item ID
            max_stock: Maximum allowed quantity (current stock)
        """
        vals = list(self.tree.item(row_id, "values"))
        code = vals[0]
        description = vals[1]
        current_qty_str = vals[9]  # qty_to_issue column

        # Strip star if present
        if current_qty_str.startswith("★"):
            current_qty_str = current_qty_str[2:].strip()

        try:
            current_qty = int(current_qty_str) if current_qty_str.isdigit() else 0
        except:
            current_qty = 0

        # ✅ Get qty_required for context
        qty_required_str = vals[8]
        if qty_required_str != "N/A":
            try:
                qty_required = int(qty_required_str)
            except:
                qty_required = None
        else:
            qty_required = None

        # ===== CREATE CUSTOM DIALOG =====
        dialog = tk.Toplevel(self.parent)
        dialog.title(lang.t("dispatch_kit.edit_qty_title", "Edit Quantity"))
        dialog.geometry("450x450")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # Main frame
        main_frame = tk.Frame(dialog, bg="#F0F4F8", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_label = tk.Label(
            main_frame,
            text=lang.t("dispatch_kit.edit_quantity", "Edit Quantity"),
            font=("Helvetica", 14, "bold"),
            bg="#F0F4F8",
        )
        title_label.pack(pady=(0, 15))

        instruction_label = tk.Label(
            main_frame,
            text=lang.t(
                "dispatch_kit.edit_qty_instruction",
                "Enter new quantity and press ENTER or click Save",
            ),
            font=("Helvetica", 9, "italic"),
            fg="#555",
            bg="#F0F4F8",
        )
        instruction_label.pack(pady=(0, 15))

        # Item info frame
        info_frame = tk.Frame(main_frame, bg="white", relief="solid", borderwidth=1)
        info_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            info_frame,
            text=f"{lang.t('dispatch_kit.code', 'Code')}: {code}",
            font=("Helvetica", 11, "bold"),
            bg="white",
            anchor="w",
        ).pack(fill="x", padx=10, pady=5)

        tk.Label(
            info_frame,
            text=description,
            font=("Helvetica", 10),
            bg="white",
            anchor="w",
            wraplength=400,
            justify="left",
        ).pack(fill="x", padx=10, pady=(0, 5))

        # Stock info frame
        stock_frame = tk.Frame(main_frame, bg="#E8F4F8", relief="solid", borderwidth=1)
        stock_frame.pack(fill="x", pady=(0, 15))

        tk.Label(
            stock_frame,
            text=f"{lang.t('dispatch_kit.current_stock', 'Current Stock')}: {max_stock}",
            font=("Helvetica", 10),
            bg="#E8F4F8",
            anchor="w",
        ).pack(fill="x", padx=10, pady=3)

        if qty_required is not None:
            tk.Label(
                stock_frame,
                text=f"{lang.t('dispatch_kit.qty_required', 'Qty Required')}: {qty_required}",
                font=("Helvetica", 10),
                bg="#E8F4F8",
                anchor="w",
            ).pack(fill="x", padx=10, pady=3)

        tk.Label(
            stock_frame,
            text=f"{lang.t('dispatch_kit.current_qty_issue', 'Current Qty to Issue')}: {current_qty}",
            font=("Helvetica", 10),
            bg="#E8F4F8",
            anchor="w",
        ).pack(fill="x", padx=10, pady=3)

        # Entry frame
        entry_frame = tk.Frame(main_frame, bg="#F0F4F8")
        entry_frame.pack(fill="x", pady=(0, 10))

        tk.Label(
            entry_frame,
            text=f"{lang.t('dispatch_kit.new_quantity', 'New Quantity')}:",
            font=("Helvetica", 11, "bold"),
            bg="#F0F4F8",
        ).pack(anchor="w")

        qty_var = tk.StringVar(value=str(current_qty))
        qty_entry = tk.Entry(
            entry_frame, textvariable=qty_var, font=("Helvetica", 12), width=15
        )
        qty_entry.pack(anchor="w", pady=5)
        qty_entry.focus()
        qty_entry.select_range(0, tk.END)

        # Status/error label
        status_label = tk.Label(
            main_frame,
            text="",
            font=("Helvetica", 9),
            fg="red",
            bg="#F0F4F8",
            wraplength=400,
        )
        status_label.pack(pady=5)

        # Result variable
        result = {"cancelled": True, "value": None}

        def save_quantity():
            new_qty_str = qty_var.get().strip()

            # Validate
            if not new_qty_str.isdigit():
                status_label.config(
                    text=lang.t(
                        "dispatch_kit.error_invalid_number",
                        "Please enter a valid number",
                    )
                )
                return

            new_qty = int(new_qty_str)

            if new_qty < 0:
                status_label.config(
                    text=lang.t(
                        "dispatch_kit.error_negative", "Quantity cannot be negative"
                    )
                )
                return

            if new_qty > max_stock:
                status_label.config(
                    text=lang.t(
                        "dispatch_kit.error_exceeds_stock",
                        "Exceeds available stock ({stock})",
                    ).format(stock=max_stock)
                )
                return

            # Valid input
            result["cancelled"] = False
            result["value"] = new_qty
            dialog.destroy()

        def cancel():
            result["cancelled"] = True
            dialog.destroy()

        # Button frame
        btn_frame = tk.Frame(main_frame, bg="#F0F4F8")
        btn_frame.pack(side="bottom", pady=10)

        tk.Button(
            btn_frame,
            text=lang.t("dispatch_kit.save", "Save"),
            font=("Helvetica", 10, "bold"),
            bg="#27AE60",
            fg="white",
            width=10,
            command=save_quantity,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text=lang.t("dispatch_kit.cancel", "Cancel"),
            font=("Helvetica", 10),
            bg="#7F8C8D",
            fg="white",
            width=10,
            command=cancel,
        ).pack(side="left", padx=5)

        # Bind keys
        qty_entry.bind("<Return>", lambda e: save_quantity())
        qty_entry.bind("<KP_Enter>", lambda e: save_quantity())
        dialog.bind("<Escape>", lambda e: cancel())

        # Wait for dialog
        dialog.wait_window()

        # Process result
        if result["cancelled"]:
            return

        new_qty = result["value"]

        # Update tree (NO star for items)
        vals[9] = str(new_qty)
        self.tree.item(row_id, values=vals)

        # Update row_data
        if row_id in self.row_data:
            self.row_data[row_id]["qty_to_issue"] = new_qty

        self.status_var.set(
            lang.t(
                "dispatch_kit.qty_updated", "Quantity updated for {code}: {qty}"
            ).format(code=code, qty=new_qty)
        )

        logging.debug(
            f"[EDIT_QTY] Updated {code} qty_to_issue: {current_qty} → {new_qty}"
        )

    def export_data(self, rows_to_issue=None):
        logging.info("[DISPATCH] export_data called")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in export_data")
            return
        try:
            export_rows = []
            all_iids = []

            def collect_iids(item=""):
                for child in self.tree.get_children(item):
                    all_iids.append(child)
                    collect_iids(child)

            collect_iids()

            for iid in all_iids:
                vals = self.tree.item(iid, "values")
                if not vals or len(vals) < 10:
                    logging.warning(f"[DISPATCH] Skipping invalid row {iid}: {vals}")
                    continue
                (
                    code,
                    desc,
                    tfield,
                    kit_col,
                    module_col,
                    current_stock,
                    exp_date,
                    batch_no,
                    qty_required,
                    qty_to_issue,
                    _uid,
                ) = vals
                raw_q = (
                    qty_to_issue[2:].strip()
                    if qty_to_issue.startswith("★")
                    else qty_to_issue
                )
                qty = int(raw_q) if raw_q.isdigit() else 0
                rd = self.row_data.get(iid, {})
                kit_no = rd.get("Kit_number") or rd.get("kit_number") or kit_col or ""
                mod_no = rd.get("module_number") or module_col or ""
                export_rows.append(
                    {
                        "iid": iid,
                        "code": code,
                        "description": desc,
                        "type": tfield or "Item",
                        "kit_number": kit_no,
                        "module_number": mod_no,
                        "current_stock": int(current_stock or 0),
                        "expiry_date": exp_date or "",
                        "batch_number": batch_no or "",
                        "qty_issued": qty,
                    }
                )

            if rows_to_issue is not None:
                for row in rows_to_issue:
                    if len(row) != 7:
                        logging.warning(
                            f"[DISPATCH] Skipping invalid provided row: {row}"
                        )
                        continue
                    rid, code, desc, stock, qty, exp_date, batch_no = row
                    for er in export_rows:
                        if er["iid"] == rid:
                            er["current_stock"] = stock
                            er["qty_issued"] = qty
                            er["expiry_date"] = exp_date or er["expiry_date"]
                            er["batch_number"] = batch_no or er["batch_number"]
                            break

            has_issued = any(r["qty_issued"] > 0 for r in export_rows)
            if not has_issued:
                custom_popup(
                    self.parent,
                    lang.t("dialog_titles.error", "Error"),
                    lang.t(
                        "dispatch_kit.no_issue_qty", "No quantities entered to issue."
                    ),
                    "error",
                )
                return

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            out_type_raw = self.trans_type_var.get() or lang.t(
                "dispatch_kit.unknown", "Unknown"
            )
            movement_type_raw = self.mode_var.get() or lang.t(
                "dispatch_kit.unknown", "Unknown"
            )
            scenario_name = self.selected_scenario_name or "N/A"
            doc_number = getattr(self, "current_document_number", None)

            import re

            def sanitize(s: str) -> str:
                s = re.sub(r"[^A-Za-z0-9]+", "_", s)
                s = re.sub(r"_+", "_", s)
                return s.strip("_") or "Unknown"

            out_type_slug = sanitize(out_type_raw)
            movement_type_slug = sanitize(movement_type_raw)

            default_dir = "D:/ISEPREP"
            if not os.path.exists(default_dir):
                os.makedirs(default_dir)

            file_name = f"Dispatch_{movement_type_slug}_{out_type_slug}_{current_time.replace(':', '-')}.xlsx"

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=file_name,
                initialdir=default_dir,
            )
            if not path:
                self.status_var.set(
                    lang.t("dispatch_kit.export_cancelled", "Export cancelled")
                )
                return

            wb = openpyxl.Workbook()
            ws = wb.active

            if doc_number:
                ws["A1"] = lang.t(
                    "dispatch_kit.date_doc",
                    "Date: {date}        Document Number: {doc}",
                ).format(date=current_time, doc=doc_number)
            else:
                ws["A1"] = lang.t("dispatch_kit.date_only", "Date: {date}").format(
                    date=current_time
                )
            ws["A1"].font = Font(name="Calibri", size=11)
            ws["A1"].alignment = Alignment(horizontal="left")

            project_name, project_code = fetch_project_details()

            ws_title_base = lang.t("dispatch_kit.sheet_title_base", "Dispatch")
            ws_title = f"{ws_title_base[:15]}-{movement_type_slug[:12]}"
            ws.title = ws_title

            ws["A2"] = (
                f"{ws_title_base} – {lang.t('dispatch_kit.movement', 'Movement')}: {movement_type_raw}"
            )
            ws["A2"].font = Font(name="Tahoma", size=14, bold=True)
            ws["A2"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A2:I2")

            ws["A3"] = f"{project_name} - {project_code}"
            ws["A3"].font = Font(name="Tahoma", size=14, bold=True)
            ws["A3"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A3:I3")

            ws["A4"] = f"{lang.t('dispatch_kit.out_type', 'OUT Type')}: {out_type_raw}"
            ws["A4"].font = Font(name="Tahoma", size=12, bold=True)
            ws["A4"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A4:I4")

            ws["A5"] = f"{lang.t('dispatch_kit.scenario', 'Scenario')}: {scenario_name}"
            ws["A5"].font = Font(name="Tahoma", size=12, bold=True)
            ws["A5"].alignment = Alignment(horizontal="right")
            ws.merge_cells("A5:I5")

            ws.append([])
            ws["A6"].font = Font(name="Tahoma", size=11, bold=True)

            headers = [
                lang.t("dispatch_kit.code", "Code"),
                lang.t("dispatch_kit.description", "Description"),
                lang.t("dispatch_kit.type", "Type"),
                lang.t("dispatch_kit.kit_number", "Kit Number"),
                lang.t("dispatch_kit.module_number", "Module Number"),
                lang.t("dispatch_kit.current_stock", "Current Stock"),
                lang.t("dispatch_kit.expiry_date", "Expiry Date"),
                lang.t("dispatch_kit.batch_no", "Batch Number"),
                lang.t("dispatch_kit.qty_to_issue_short", "Qty Issued"),
            ]
            ws.append(headers)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=7, column=col)
                cell.font = Font(name="Tahoma", size=11, bold=True)

            from openpyxl.styles import PatternFill

            for row_idx, row in enumerate(export_rows, start=8):
                ws.append(
                    [
                        row["code"],
                        row["description"],
                        row["type"],
                        row["kit_number"],
                        row["module_number"],
                        row["current_stock"],
                        row["expiry_date"],
                        row["batch_number"],
                        row["qty_issued"],
                    ]
                )
                row_type = row["type"].lower() if row["type"] else ""
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(
                        name="Calibri", size=11, bold=(row_type in ("kit", "module"))
                    )
                    if row_type == "kit":
                        cell.fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
                    elif row_type == "module":
                        cell.fill = PatternFill(
                            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
                        )

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        l = len(str(cell.value)) if cell.value is not None else 0
                        if l > max_length:
                            max_length = l
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.print_title_rows = "1:7"
            ws.oddFooter.center.text = "&P of &N"
            ws.evenFooter.center.text = "&P of &N"

            wb.save(path)
            custom_popup(
                self.parent,
                lang.t("dialog_titles.success", "Success"),
                lang.t("dispatch_kit.export_success", "Exported to {path}").format(
                    path=path
                ),
                "info",
            )
            self.status_var.set(
                lang.t("dispatch_kit.export_success", "Exported to {path}").format(
                    path=path
                )
            )
            logging.info(f"[DISPATCH] Exported file: {path}")
        except Exception as e:
            logging.error(f"[DISPATCH] Export failed: {e}")
            custom_popup(
                self.parent,
                lang.t("dialog_titles.error", "Error"),
                lang.t("dispatch_kit.export_failed", "Export failed: {err}").format(
                    err=str(e)
                ),
                "error",
            )


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Dispatch")
    app = type("App", (), {})()
    app.role = "admin"
    StockDispatchKit(root, app, role="admin")
    root.mainloop()
