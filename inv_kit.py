import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sqlite3
from db import connect_db
from popup_utils import custom_popup, custom_askyesno, custom_dialog
from manage_items import get_item_description
from language_manager import lang
import logging
from dateutil.parser import parse
import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_expiry(date_str: str) -> str:
    try:
        dt = parse(date_str, fuzzy=True)
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        return None

class StockData:
    @staticmethod
    def add_or_update(unique_id, qty_in=0, qty_out=0, exp_date=None):
        conn = connect_db()
        if conn is None:
            raise ValueError("Database connection failed")
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT qty_in, qty_out FROM stock_data WHERE unique_id = ?", (unique_id,))
            row = cursor.fetchone()
            if row:
                new_qty_in = row[0] + qty_in
                new_qty_out = row[1] + qty_out
                cursor.execute("UPDATE stock_data SET qty_in = ?, qty_out = ?, final_qty = ?, exp_date = ? WHERE unique_id = ?",
                               (new_qty_in, new_qty_out, new_qty_in - new_qty_out, exp_date, unique_id))
            else:
                final_qty = qty_in - qty_out
                cursor.execute("""
                    INSERT INTO stock_data (unique_id, qty_in, qty_out, final_qty, exp_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (unique_id, qty_in, qty_out, final_qty, exp_date))
            conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Error in StockData.add_or_update: {e}")
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()

def log_transaction(unique_id, code, description, expiry_date, batch_number, scenario, kit, module, qty_in, in_type, qty_out, out_type, movement_type):
    conn = connect_db()
    if conn is None:
        raise ValueError("Database connection failed")
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO stock_transactions (Date, Time, unique_id, code, Description, Expiry_date, Batch_Number, Scenario, Kit, Module, Qty_IN, IN_Type, Qty_Out, Out_Type, Movement_Type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.today().strftime('%Y-%m-%d'),
            datetime.now().time().strftime('%H:%M:%S'),
            unique_id, code, description, expiry_date, batch_number, scenario, kit, module,
            qty_in, in_type, qty_out, out_type, movement_type
        ))
        conn.commit()
    except sqlite3.Error as e:
        logging.error(f"Error logging transaction: {e}")
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

class InventoryKit(tk.Frame):
    def __init__(self, parent, app, role: str = "supervisor"):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        self.role = role.lower()
        self.scenario_map = self.fetch_scenario_map()
        self.selected_kit_id = None
        self.tree = None
        self.scenario_var = None
        self.kit_var = None
        self.status_var = None
        logging.info("InventoryKit initialized")
        if self.parent is not None and self.parent.winfo_exists():
            self.pack(fill="both", expand=True)
            self.after(100, self.initialize_ui)
        else:
            logging.error("Parent window is None or does not exist at initialization")

    def initialize_ui(self):
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in initialize_ui")
            return
        try:
            logging.info("Starting UI initialization")
            self.render_ui()
            logging.info("UI initialization completed")
        except tk.TclError as e:
            logging.error(f"Error rendering UI: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to render UI: {str(e)}"), parent=self.parent)

    def fetch_scenario_map(self):
        logging.info("Fetching scenario map")
        conn = connect_db()
        if conn is None:
            logging.error("Database connection failed")
            return {}
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT scenario_id, name FROM scenarios")
            scenario_map = {str(row['scenario_id']): row['name'] for row in cursor.fetchall()}
            logging.info(f"Fetched scenario map: {scenario_map}")
            return scenario_map
        except sqlite3.Error as e:
            logging.error(f"Error fetching scenario map: {str(e)}")
            return {}
        finally:
            cursor.close()
            conn.close()

    def fetch_scenarios(self):
        logging.info("Fetching scenarios")
        conn = connect_db()
        if conn is None:
            logging.error("Database connection failed")
            return []
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT scenario_id, name FROM scenarios ORDER BY name")
            scenarios = [{"id": row['scenario_id'], "name": row['name']} for row in cursor.fetchall()]
            logging.info(f"Fetched {len(scenarios)} scenarios")
            return scenarios
        except sqlite3.Error as e:
            logging.error(f"Error fetching scenarios: {str(e)}")
            return []
        finally:
            cursor.close()
            conn.close()

    def fetch_kits_for_scenario(self, scenario_id):
        logging.info(f"Fetching kits for scenario: {scenario_id}")
        conn = connect_db()
        if conn is None:
            logging.error("Database connection failed")
            return []
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT DISTINCT kit AS kit_code 
                FROM kit_items 
                WHERE scenario_id = ? AND kit IS NOT NULL AND level = 'primary'
            """, (scenario_id,))
            kits = [row['kit_code'] for row in cursor.fetchall()]
            kits_with_desc = []
            for kit in kits:
                description = get_item_description(kit) or lang.t("kits.no_description", "No Description")
                kits_with_desc.append({
                    'code': kit,
                    'description': description
                })
            logging.info(f"Fetched {len(kits)} kits for scenario {scenario_id}")
            return kits_with_desc
        except sqlite3.Error as e:
            logging.error(f"Error fetching kits: {str(e)}")
            return []
        finally:
            cursor.close()
            conn.close()

    def fetch_kit_inventory(self, kit_code, scenario_id):
        logging.info(f"Fetching inventory for kit: {kit_code} in scenario: {scenario_id}")
        conn = connect_db()
        if conn is None:
            logging.error("Database connection failed")
            return []
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT ki.item, ki.module, ki.code, ki.treecode, ki.level, ki.std_qty, 
                       sd.unique_id, sd.kit_number, sd.module_number, sd.final_qty, sd.exp_date
                FROM kit_items ki
                LEFT JOIN stock_data sd 
                ON ki.code = sd.item 
                AND ki.kit = sd.kit 
                AND ki.module = sd.module 
                AND ki.scenario_id = sd.scenario
                WHERE ki.scenario_id = ? AND ki.kit = ? 
                AND ki.treecode >= ? AND ki.treecode < ?
                AND (sd.final_qty > 0 OR sd.final_qty IS NULL)
            """, (scenario_id, kit_code, '01001000000', '01002000000'))
            inventory = []
            for row in cursor.fetchall():
                description = get_item_description(row['code']) or lang.t("kits.no_description", "No Description")
                type_field = "Kit" if row['level'] == 'primary' else "Module" if row['level'] == 'secondary' else "Item"
                inventory.append({
                    'unique_id': row['unique_id'] or f"{scenario_id}/{kit_code}/{row['module'] or 'None'}/{row['code']}/{row['std_qty']}/None",
                    'kit_number': row['kit_number'] or kit_code,
                    'module_number': row['module'] or '',
                    'item': row['code'],
                    'description': description,
                    'final_qty': row['final_qty'] or 0,
                    'exp_date': row['exp_date'] or '',
                    'type': type_field
                })
            logging.info(f"Fetched {len(inventory)} inventory items for kit {kit_code}")
            return inventory
        except sqlite3.Error as e:
            logging.error(f"Error fetching kit inventory: {str(e)}")
            return []
        finally:
            cursor.close()
            conn.close()

    def render_ui(self):
        logging.info("Entering render_ui")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in render_ui")
            return
        try:
            for w in self.parent.winfo_children():
                w.destroy()
                logging.debug(f"Destroyed widget: {w}")

            tk.Label(self.parent, text=lang.t("inv_kit.title", "Inventory Kits/Modules"),
                     font=("Helvetica", 20, "bold"), bg="#F5F5F5").pack(pady=10)
            logging.debug("Created title label")

            main_frame = tk.Frame(self.parent, bg="#F5F5F5")
            main_frame.pack(fill="both", expand=True, padx=10, pady=10)
            logging.debug("Created main frame")

            tk.Label(main_frame, text=lang.t("inv_kit.scenario", "Scenario:"), bg="#F5F5F5").grid(row=0, column=0, padx=5, pady=5, sticky="w")
            self.scenario_var = tk.StringVar()
            self.scenario_cb = ttk.Combobox(main_frame, textvariable=self.scenario_var, state="readonly", width=40)
            self.scenario_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")
            self.scenario_cb.bind("<<ComboboxSelected>>", self.on_scenario_selected)
            logging.debug("Created scenario combobox")

            tk.Label(main_frame, text=lang.t("inv_kit.kit", "Kit:"), bg="#F5F5F5").grid(row=1, column=0, padx=5, pady=5, sticky="w")
            self.kit_var = tk.StringVar()
            self.kit_cb = ttk.Combobox(main_frame, textvariable=self.kit_var, state="disabled", width=40)
            self.kit_cb.grid(row=1, column=1, padx=5, pady=5, sticky="w")
            self.kit_cb.bind("<<ComboboxSelected>>", self.on_kit_selected)
            logging.debug("Created kit combobox")

            cols = ("unique_id", "type", "kit_number", "module_number", "item", "description", "final_qty", "exp_date")
            self.tree = ttk.Treeview(main_frame, columns=cols, show="headings", height=20)
            
            headers = {
                "unique_id": lang.t("inv_kit.unique_id", "Unique ID"),
                "type": lang.t("inv_kit.type", "Type"),
                "kit_number": lang.t("inv_kit.kit_number", "Kit Number"),
                "module_number": lang.t("inv_kit.module_number", "Module Number"),
                "item": lang.t("inv_kit.item", "Item"),
                "description": lang.t("inv_kit.description", "Description"),
                "final_qty": lang.t("inv_kit.final_qty", "Final Qty"),
                "exp_date": lang.t("inv_kit.exp_date", "Expiry Date")
            }
            
            widths = {
                "unique_id": 160,
                "type": 130,
                "kit_number": 130,
                "module_number": 130,
                "item": 160,
                "description": 400,
                "final_qty": 130,
                "exp_date": 160
            }
            
            for col in cols:
                self.tree.heading(col, text=headers[col])
                self.tree.column(col, width=widths.get(col, 130), anchor="center", stretch=True)
            
            self.tree.grid(row=2, column=0, columnspan=2, pady=10, sticky="nsew")
            self.tree.bind("<Double-1>", self.start_edit)
            logging.debug("Created Treeview")

            btn_frame = tk.Frame(main_frame, bg="#F5F5F5")
            btn_frame.grid(row=3, column=0, columnspan=2, pady=5)
            tk.Button(btn_frame, text=lang.t("inv_kit.save", "Save"), bg="#27AE60", fg="white",
                      command=self.save_all,
                      state="normal" if self.role in ["admin", "manager"] else "disabled").pack(side="left", padx=5)
            tk.Button(btn_frame, text=lang.t("inv_kit.clear", "Clear"), bg="#7F8C8D", fg="white",
                      command=self.clear_form).pack(side="left", padx=5)
            tk.Button(btn_frame, text=lang.t("inv_kit.export", "Export"), bg="#2980B9", fg="white",
                      command=self.export_data).pack(side="left", padx=5)
            logging.debug("Created button frame")

            self.status_var = tk.StringVar()
            tk.Label(main_frame, textvariable=self.status_var, bg="#F5F5F5").grid(row=4, column=0, columnspan=2, pady=5)
            logging.debug("Created status label")

            self.load_scenarios()
            if self.parent is not None and self.parent.winfo_exists():
                self.parent.update()
            logging.info("Exiting render_ui")
        except tk.TclError as e:
            logging.error(f"Error in render_ui: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to render UI: {str(e)}"), parent=self.parent)

    def load_scenarios(self):
        logging.info("Entering load_scenarios")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in load_scenarios")
            return
        try:
            scenarios = self.fetch_scenarios()
            scenario_names = [f"{s['id']} - {s['name']}" for s in scenarios]
            self.scenario_cb['values'] = scenario_names
            logging.debug(f"Set {len(scenario_names)} scenario names in combobox")
            if scenario_names:
                self.scenario_cb.current(0)
                self.on_scenario_selected()
            logging.info("Exiting load_scenarios")
        except tk.TclError as e:
            logging.error(f"Error in load_scenarios: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to load scenarios: {str(e)}"), parent=self.parent)

    def on_scenario_selected(self, event=None):
        logging.info("Entering on_scenario_selected")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in on_scenario_selected")
            return
        try:
            scenario_str = self.scenario_var.get()
            if not scenario_str:
                logging.debug("No scenario selected")
                return
            scenario_id = scenario_str.split(" - ")[0]
            kits = self.fetch_kits_for_scenario(scenario_id)
            kit_display = [f"{k['code']} - {k['description']}" for k in kits]
            self.kit_cb['values'] = kit_display
            self.kit_cb['state'] = 'readonly'
            self.kit_var.set("")
            if self.tree is not None:
                self.tree.delete(*self.tree.get_children())
            self.status_var.set("")
            logging.debug(f"Set {len(kit_display)} kits in combobox")
            logging.info("Exiting on_scenario_selected")
        except tk.TclError as e:
            logging.error(f"Error loading kits: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.load_kits", f"Failed to load kits: {str(e)}"), parent=self.parent)

    def on_kit_selected(self, event=None):
        logging.info("Entering on_kit_selected")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in on_kit_selected")
            return
        try:
            kit_str = self.kit_var.get()
            if not kit_str:
                logging.debug("No kit selected")
                return
            kit_code = kit_str.split(" - ")[0]
            scenario_id = self.scenario_var.get().split(" - ")[0]
            inventory = self.fetch_kit_inventory(kit_code, scenario_id)
            if self.tree is not None:
                self.tree.delete(*self.tree.get_children())
            for inv in inventory:
                self.tree.insert("", "end", values=(
                    inv['unique_id'], inv['type'], inv['kit_number'], inv['module_number'],
                    inv['item'], inv['description'], inv['final_qty'], inv['exp_date']
                ))
            self.status_var.set(lang.t("inv_kit.loaded_records", f"Loaded {len(self.tree.get_children())} records"))
            logging.info(f"Populated {len(self.tree.get_children())} rows in Treeview")
            logging.info("Exiting on_kit_selected")
        except tk.TclError as e:
            logging.error(f"Error loading inventory: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.load_inventory", f"Failed to load inventory: {str(e)}"), parent=self.parent)

    def start_edit(self, event):
        logging.info("Entering start_edit")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in start_edit")
            return
        try:
            region = self.tree.identify("region", event.x, event.y)
            if region != "cell":
                logging.debug("Not a cell region in start_edit")
                return
            row_id = self.tree.identify_row(event.y)
            col = self.tree.identify_column(event.x)
            col_name = self.tree.column(col, "id")
            if col_name != "final_qty":
                logging.debug("Invalid column for editing")
                return
            x, y, width, height = self.tree.bbox(row_id, col)
            value = self.tree.item(row_id, "values")[self.tree["columns"].index(col_name)]
            entry = tk.Entry(self.tree)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, value)
            entry.focus()
            logging.debug(f"Created edit entry for row {row_id}, column {col_name}")
            def save_edit(event=None):
                try:
                    new_val = entry.get().strip()
                    new_qty = int(new_val)
                    if new_qty < 0:
                        messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.invalid_qty", "Quantity cannot be negative"), parent=self.parent)
                        entry.destroy()
                        return
                    values = list(self.tree.item(row_id, "values"))
                    col_idx = self.tree["columns"].index(col_name)
                    values[col_idx] = new_qty
                    self.tree.item(row_id, values=values)
                    entry.destroy()
                    logging.debug(f"Saved edit for row {row_id}, column {col_name}")
                except ValueError as e:
                    messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.invalid_qty", str(e)), parent=self.parent)
                    entry.destroy()
                except tk.TclError as e:
                    logging.error(f"Error in save_edit: {e}")
                    entry.destroy()
            entry.bind("<Return>", save_edit)
            entry.bind("<FocusOut>", save_edit)
            logging.info("Exiting start_edit")
        except tk.TclError as e:
            logging.error(f"Error in start_edit: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to edit cell: {str(e)}"), parent=self.parent)

    def save_all(self):
        logging.info("Entering save_all")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in save_all")
            return
        try:
            if self.role.lower() not in ["admin", "manager"]:
                messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.restricted", "Only admin or manager roles can save changes."), parent=self.parent)
                return
            rows = self.tree.get_children()
            if not rows:
                messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.no_rows", "No rows to save."), parent=self.parent)
                return
            conn = connect_db()
            if conn is None:
                logging.error("Database connection failed")
                messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.db_error", "Database connection failed"), parent=self.parent)
                return
            cursor = conn.cursor()
            try:
                for row in rows:
                    vals = self.tree.item(row, "values")
                    unique_id, type_field, kit_number, module_number, item, description, final_qty, exp_date = vals
                    qty_change = int(final_qty)
                    StockData.add_or_update(unique_id, qty_in=qty_change, qty_out=0, exp_date=exp_date or None)
                    log_transaction(
                        unique_id=unique_id,
                        code=item,
                        description=description,
                        expiry_date=parse_expiry(exp_date) if exp_date else None,
                        batch_number=None,
                        scenario=self.scenario_var.get().split(" - ")[0],
                        kit=kit_number if type_field == "Kit" else None,
                        module=module_number if type_field == "Module" else None,
                        qty_in=qty_change,
                        in_type="inventory_adjustment",
                        qty_out=None,
                        out_type=None,
                        movement_type="inventory_adjustment"
                    )
                conn.commit()
                messagebox.showinfo(lang.t("dialog_titles.info", "Success"), lang.t("inv_kit.save_success", "Inventory saved successfully."), parent=self.parent)
                self.clear_form()
                logging.info("Save completed successfully")
            except Exception as e:
                conn.rollback()
                logging.error(f"Failed to save inventory: {e}")
                messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.save_failed", f"Failed to save inventory: {str(e)}"), parent=self.parent)
            finally:
                cursor.close()
                conn.close()
        except tk.TclError as e:
            logging.error(f"Error in save_all: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to save: {str(e)}"), parent=self.parent)
        logging.info("Exiting save_all")

    def clear_form(self):
        logging.info("Entering clear_form")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in clear_form")
            return
        try:
            self.scenario_var.set("")
            self.kit_var.set("")
            self.kit_cb['state'] = 'disabled'
            if self.tree is not None:
                self.tree.delete(*self.tree.get_children())
            self.status_var.set("")
            self.load_scenarios()
            logging.info("Exiting clear_form")
        except tk.TclError as e:
            logging.error(f"Error in clear_form: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("error.ui_render", f"Failed to clear form: {str(e)}"), parent=self.parent)

    def export_data(self):
        logging.info("Entering export_data")
        if self.parent is None or not self.parent.winfo_exists():
            logging.error("Parent window is None or does not exist in export_data")
            return
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not file_path:
                logging.debug("No file path selected for export")
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = lang.t("inv_kit.title", "Inventory Kits/Modules")
            headers = [self.tree.heading(col)['text'] for col in self.tree["columns"]]
            ws.append(headers)
            for item in self.tree.get_children():
                ws.append(self.tree.item(item)["values"])
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(file_path)
            messagebox.showinfo(lang.t("dialog_titles.info", "Success"), lang.t("inv_kit.export_success", f"Exported to {file_path}"), parent=self.parent)
            logging.info(f"Exported data to {file_path}")
        except Exception as e:
            logging.error(f"Export failed: {e}")
            messagebox.showerror(lang.t("dialog_titles.error", "Error"), lang.t("inv_kit.export_failed", f"Export failed: {str(e)}"), parent=self.parent)
        logging.info("Exiting export_data")