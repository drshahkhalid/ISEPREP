import tkinter as tk
from tkinter import ttk, filedialog
import openpyxl
import sqlite3
import re
from db import connect_db
from language_manager import lang
from datetime import datetime
from popup_utils import custom_popup, custom_askyesno, custom_dialog

# Roles (canonical or symbol) that are NOT allowed to edit
RESTRICTED_EDIT = {"manager", "supervisor", "~", "$"}

def validate_code(code):
    """Validate code:  at least 9 characters (uppercase letters, digits, hyphens). Strip spaces."""
    if not code:
        return None
    code = code.replace(" ", "").upper()
    if len(code) >= 9 and re.match(r'^[A-Z0-9-]*$', code):
        return code
    return None

def generate_unique_id_2(scenario_id, code, quantity):
    """Generate unique_id_2: scenario_id/None/None/code/std_qty."""
    cleaned_code = validate_code(code) or code.replace(" ", "").upper()
    return f"{scenario_id}/None/None/{cleaned_code}/{int(quantity or 0)}"

def scenario_to_column_letter(index):
    """Map scenario index (0-based) to unique_id_2a ..  unique_id_2o."""
    return f"unique_id_2{chr(ord('a') + index)}"

def get_description_with_priority(row):
    """
    Get description with priority:  active language → en → fr → sp
    row is a sqlite3.Row object
    """
    lang_code = lang.lang_code. lower()
    
    # Get available keys from Row object
    row_keys = row.keys()
    
    # Try active language first
    if lang_code == "fr" and "designation_fr" in row_keys and row["designation_fr"]:
        return row["designation_fr"]
    elif lang_code in ("es", "sp") and "designation_sp" in row_keys and row["designation_sp"]:
        return row["designation_sp"]
    elif lang_code == "en" and "designation_en" in row_keys and row["designation_en"]: 
        return row["designation_en"]
    
    # Fallback priority: en → fr → sp
    if "designation_en" in row_keys and row["designation_en"]:
        return row["designation_en"]
    if "designation_fr" in row_keys and row["designation_fr"]:
        return row["designation_fr"]
    if "designation_sp" in row_keys and row["designation_sp"]:
        return row["designation_sp"]
    
    # Last resort
    if "designation" in row_keys and row["designation"]:
        return row["designation"]
    return ""

class StandardList(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.role = getattr(app, "role", "supervisor")
        role_l = (self.role or "").lower()
        # Read-only if role or symbol is restricted
        self.read_only = role_l in RESTRICTED_EDIT or self.role in RESTRICTED_EDIT
        self.search_term = tk.StringVar()
        self.scenarios = []
        self.show_all = False
        self. kit_codes = []  # Reserved
        self.tree_order = []  # Store code order
        self.pack(fill="both", expand=True)
        self.render_ui()

    def render_ui(self):
        """Set up the UI with Treeview, buttons, and styles."""
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", rowheight=25, font=('Helvetica', 10))
        style.configure("Treeview.Heading", font=('Helvetica', 10, 'bold'), background="#2980B9", foreground="white")
        style.map("Treeview", background=[('selected', '#3498DB')])
        style.configure("Accent.TButton", background="#2980B9", foreground="white", font=('Helvetica', 10))
        style.configure("Invalid.TLabel", foreground="red")

        main_frame = ttk.Frame(self, padding=10)
        main_frame. pack(fill="both", expand=True)

        ttk.Label(main_frame, text=lang.t("standard_list.title", fallback="Standard List"), 
                  font=("Helvetica", 20, "bold")).pack(pady=(0, 10))

        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill="x", pady=5)

        ttk.Label(top_frame, text=lang.t("standard_list.search", fallback="Search Code or Description")).pack(side="left", padx=(0, 5))
        search_entry = ttk.Entry(top_frame, textvariable=self.search_term, width=30)
        search_entry.pack(side="left")
        search_entry.bind("<KeyRelease>", lambda e: self.load_data())

        ttk.Button(top_frame, text=lang.t("standard_list.refresh", fallback="Refresh"), 
                   command=self.load_data).pack(side="left", padx=5)

        self.toggle_btn = ttk. Button(
            top_frame,
            text=lang.t("standard_list.show_all_items", fallback="Show All Items"),
            command=self.toggle_mode
        )
        self.toggle_btn.pack(side="left", padx=5)

        self.last_update_label = ttk. Label(top_frame, text="", foreground="#555")
        self.last_update_label. pack(side="right")

        self.status_var = tk.StringVar(value=lang.t("standard_list.ready", fallback="Ready"))
        ttk.Label(main_frame, textvariable=self.status_var, relief="sunken", anchor="w").pack(fill="x", pady=(5, 0))

        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(expand=True, fill="both", pady=10)
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(tree_frame, show="headings", height=20)
        self.tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.tag_configure('oddrow', background='#F5F5F5')
        self.tree.tag_configure('evenrow', background='#E8ECEF')
        self.tree.tag_configure('invalid', foreground='red')

        # Editing only if not read-only
        if not self.read_only:
            self.tree.bind("<Double-1>", self.on_double_click)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=5)

        # Export always allowed
        ttk.Button(btn_frame, text=lang.t("standard_list.export_excel", fallback="Export to Excel"),
                   style="Accent.TButton", command=self.export_excel).pack(side="left", padx=5)

        # Import conditional
        ttk.Button(
            btn_frame,
            text=lang.t("standard_list.import_excel", fallback="Import from Excel"),
            style="Accent.TButton",
            command=(self.import_excel if not self. read_only else self._deny)
        ).pack(side="left", padx=5)

        # Clear all only if not read-only
        if not self.read_only:
            ttk.Button(btn_frame, text=lang.t("standard_list.clear_all", fallback="Clear All"),
                       style="Accent.TButton", command=self.clear_all).pack(side="left", padx=5)

        if self.read_only:
            self.status_var.set(lang.t("standard_list.read_only_mode", fallback="Read-only mode (no modifications allowed)"))

        self.load_data()

    def _deny(self):
        custom_popup(
            self,
            lang.t("dialog_titles.restricted", fallback="Restricted"),
            lang.t("standard_list. read_only_alert", fallback="You do not have permission to modify the standard list. "),
            "warning"
        )

    def toggle_mode(self):
        """Toggle between showing all items and compositions only."""
        self.show_all = not self.show_all
        self.toggle_btn.config(
            text=lang.t(
                "standard_list.show_compositions_only" if self.show_all else "standard_list. show_all_items",
                fallback="Show Standard List Only" if self.show_all else "Show All Items"
            )
        )
        self.load_data()

    def get_scenarios(self):
        """Fetch scenarios from database, limit to 15."""
        try:
            conn = connect_db()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT scenario_id, name FROM scenarios ORDER BY scenario_id")
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
            scenarios = [{"scenario_id": row["scenario_id"], "name": row["name"]} for row in rows]
            if len(scenarios) > 15:
                custom_popup(
                    self,
                    lang.t("dialog_titles.warning", fallback="Warning"),
                    lang.t("standard_list.warning_15_scenarios", fallback="Only first 15 scenarios supported."),
                    "warning"
                )
                return scenarios[:15]
            return scenarios
        except Exception as e: 
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list.db_error_scenarios", fallback="Failed to fetch scenarios:  {error}").format(error=str(e)),
                "error"
            )
            return []

    def load_data(self):
        """Load and display data in Treeview, preserving order."""
        self.status_var.set(lang.t("standard_list.loading", fallback="Loading..."))
        self.app.update_idletasks()

        try:
            self.scenarios = self.get_scenarios()
            if not self.scenarios:
                self.status_var.set(lang.t("standard_list.no_scenarios_found", fallback="No scenarios found"))
                return

            columns = [
                lang.t("standard_list.code", fallback="Code"),
                lang.t("standard_list. description", fallback="Description"),
                lang.t("standard_list.type", fallback="Type")
            ] + [s["name"] for s in self. scenarios] + [lang.t("standard_list. remarks", fallback="Remarks")]
            
            self.tree["columns"] = columns

            code_label = lang.t("standard_list.code", fallback="Code")
            desc_label = lang.t("standard_list.description", fallback="Description")
            
            for col in columns:
                width, anchor = (135, "w") if col == code_label else (380, "w") if col == desc_label else (120, "center")
                self.tree.heading(col, text=col)
                self.tree.column(col, width=width, anchor=anchor, stretch=True)

            conn = connect_db()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Fetch all designation columns to allow priority selection
            query = f"""
                SELECT DISTINCT i.code, i.type, 
                       i.designation_en, i.designation_fr, i.designation_sp, i.designation,
                       c.scenario_id, c.quantity, c.remarks, c.updated_at
                FROM items_list i
                {"LEFT JOIN" if self.show_all else "JOIN"} compositions c ON i.code = c. code
                ORDER BY i.code ASC
            """
            cursor. execute(query)
            rows = cursor.fetchall()

            cursor.execute("SELECT MAX(updated_at) AS last_update FROM compositions")
            last_update_row = cursor.fetchone()
            last_update = last_update_row["last_update"] if last_update_row else None
            last_update_text = (last_update if isinstance(last_update, str)
                                else last_update.strftime('%Y-%m-%d %H:%M') if last_update else 'N/A')
            self.last_update_label.config(
                text=f"{lang.t('standard_list. last_update', fallback='Standard List last updated on')}: {last_update_text}"
            )
            cursor.close()
            conn.close()

            data_map = {}
            invalid_codes = set()

            for row in rows:
                cleaned_code = validate_code(row["code"])
                code = cleaned_code if cleaned_code else row["code"]. replace(" ", "").upper()
                if not cleaned_code:
                    invalid_codes.add(row["code"])
                
                if code not in data_map:
                    # Get description with priority
                    description = get_description_with_priority(row)
                    
                    data_map[code] = {
                        "desc": description,
                        "type": row["type"] or "",
                        "remarks": row["remarks"] or "",
                        "qtys": {s["scenario_id"]: 0 for s in self.scenarios}
                    }
                
                if row["scenario_id"]:
                    data_map[code]["qtys"][row["scenario_id"]] = row["quantity"] or 0
                    if row["remarks"]:
                        data_map[code]["remarks"] = row["remarks"]

            if invalid_codes:
                custom_popup(
                    self,
                    lang.t("dialog_titles.warning", fallback="Warning"),
                    lang.t("standard_list.invalid_code_warning", 
                           fallback="Found {count} invalid codes (e.g., {examples}). Codes must be at least 9 characters with letters, digits, or hyphens.").format(
                        count=len(invalid_codes),
                        examples=', '.join(list(invalid_codes)[:3])
                    ),
                    "warning"
                )

            self.tree.delete(*self.tree.get_children())
            search = self.search_term.get().lower()
            new_order = []

            # Existing order first
            for code in self.tree_order:
                if code not in data_map:
                    continue
                item = data_map[code]
                if search and search not in code.lower() and search not in item["desc"]. lower():
                    continue
                values = [code, item["desc"], item["type"]] + [
                    int(item["qtys"]. get(s["scenario_id"], 0)) for s in self.scenarios
                ] + [item["remarks"]]
                tag = 'oddrow' if len(new_order) % 2 else 'evenrow'
                if not validate_code(code):
                    tag = ('invalid',)
                self.tree.insert("", "end", iid=code, values=values, tags=tag)
                new_order.append(code)

            # New items
            for code in data_map:
                if code in new_order:
                    continue
                item = data_map[code]
                if search and search not in code.lower() and search not in item["desc"].lower():
                    continue
                values = [code, item["desc"], item["type"]] + [
                    int(item["qtys"].get(s["scenario_id"], 0)) for s in self.scenarios
                ] + [item["remarks"]]
                tag = 'oddrow' if len(new_order) % 2 else 'evenrow'
                if not validate_code(code):
                    tag = ('invalid',)
                self.tree.insert("", "end", iid=code, values=values, tags=tag)
                new_order.append(code)

            self.tree_order = new_order

            loaded_records = len(self.tree. get_children())
            self.status_var.set(lang.t("standard_list.loaded_records", fallback="Loaded {count} record(s)").format(count=loaded_records))

            if loaded_records == 0:
                if not self.show_all:
                    self.show_all = True
                    self.toggle_btn.config(
                        text=lang.t("standard_list.show_compositions_only", fallback="Show Standard List Only")
                    )
                    self.load_data()
                    return
                else:
                    custom_popup(
                        self,
                        lang.t("dialog_titles.info", fallback="Info"),
                        lang.t("standard_list.no_data_info", fallback="No data loaded. Check database, code formats, or search term. "),
                        "info"
                    )
        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list.error_loading_data", fallback="Error loading data") + f": {str(e)}",
                "error"
            )
            self.status_var.set(lang.t("standard_list. error_loading_data", fallback="Error loading data"))

    def on_double_click(self, event):
        """Handle double-click to edit scenario quantities or remarks."""
        if self.read_only:
            self._deny()
            return
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        col_index = int(col_id[1:]) - 1
        col_name = self.tree["columns"][col_index]
        remarks_label = lang.t("standard_list.remarks", fallback="Remarks")
        if col_name not in [s["name"] for s in self. scenarios] + [remarks_label]:
            return

        x, y, width, height = self.tree.bbox(row_id, col_id)
        value = self.tree.item(row_id, "values")[col_index]

        entry = ttk.Entry(self.tree)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus()

        def save_edit(event=None):
            new_value = entry.get()
            entry.destroy()
            values = list(self.tree.item(row_id, "values"))
            values[col_index] = new_value
            self.tree.item(row_id, values=values)
            self.save_changes(row_id, col_name, new_value)

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)

    def save_changes(self, code, column_name, new_value):
        """Save changes to quantities or remarks in the database."""
        if self.read_only:
            self._deny()
            return
        try:
            cleaned_code = validate_code(code) or code. replace(" ", "").upper()
            conn = connect_db()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            remarks_label = lang.t("standard_list.remarks", fallback="Remarks")
            if column_name == remarks_label:
                cursor.execute("UPDATE compositions SET remarks=?, updated_at=CURRENT_TIMESTAMP WHERE code=?", 
                               (new_value, cleaned_code))
            else:
                scenario = next((s for s in self.scenarios if s["name"] == column_name), None)
                if not scenario:
                    return
                scenario_id, scenario_name = scenario["scenario_id"], scenario["name"]
                try:
                    qty = float(new_value) if new_value else 0
                except ValueError:
                    qty = 0

                cursor.execute("SELECT unique_id_2a, unique_id_2b, unique_id_2c, unique_id_2d, unique_id_2e, "
                               "unique_id_2f, unique_id_2g, unique_id_2h, unique_id_2i, unique_id_2j, "
                               "unique_id_2k, unique_id_2l, unique_id_2m, unique_id_2n, unique_id_2o "
                               "FROM compositions WHERE code = ? AND scenario_id = ?", (cleaned_code, scenario_id))
                existing_row = cursor.fetchone()

                col_letter = None
                if existing_row:
                    for i in 'abcdefghijklmno':
                        if existing_row[f'unique_id_2{i}']:
                            col_letter = f"unique_id_2{i}"
                            break
                else:
                    cursor.execute("SELECT unique_id_2a, unique_id_2b, unique_id_2c, unique_id_2d, unique_id_2e, "
                                   "unique_id_2f, unique_id_2g, unique_id_2h, unique_id_2i, unique_id_2j, "
                                   "unique_id_2k, unique_id_2l, unique_id_2m, unique_id_2n, unique_id_2o "
                                   "FROM compositions WHERE code = ? ", (cleaned_code,))
                    rows = cursor.fetchall()
                    used_columns = {f'unique_id_2{i}' for row in rows for i in 'abcdefghijklmno' if row[f'unique_id_2{i}']}
                    for i in 'abcdefghijklmno':
                        if f'unique_id_2{i}' not in used_columns: 
                            col_letter = f"unique_id_2{i}"
                            break

                if not col_letter:
                    custom_popup(
                        self,
                        lang.t("dialog_titles.error", fallback="Error"),
                        lang.t("standard_list.no_available_column", fallback="No available unique_id_2 column.  Maximum scenarios reached."),
                        "error"
                    )
                    return

                unique_id_2 = generate_unique_id_2(scenario_id, cleaned_code, qty)
                if qty > 0:
                    cursor.execute(f"""
                        INSERT INTO compositions (code, scenario_id, quantity, unique_id_2, {col_letter})
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(code, scenario_id) DO UPDATE SET
                            quantity=excluded.quantity, unique_id_2=excluded.unique_id_2, {col_letter}=excluded.{col_letter},
                            updated_at=CURRENT_TIMESTAMP
                    """, (cleaned_code, scenario_id, qty, unique_id_2, unique_id_2))
                else:
                    cursor.execute("DELETE FROM compositions WHERE code = ? AND scenario_id = ?", 
                                   (cleaned_code, scenario_id))

            conn.commit()
            self.status_var.set(lang.t("standard_list.changes_saved", fallback="Changes saved"))
            self.load_data()
        except Exception as e: 
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list.save_changes_failed", fallback="Failed to save changes: {error}").format(error=str(e)),
                "error"
            )
            try:
                conn.rollback()
            except Exception:
                pass
        finally:
            try:
                cursor.close()
                conn.close()
            except Exception:
                pass

    def export_excel(self):
        """Export Treeview data to Excel (allowed even in read-only)."""
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                      filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Standard List"
            ws.append(self.tree["columns"])
            for item in self.tree.get_children():
                ws.append(self.tree.item(item)["values"])

            for col in ws. columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_length

            wb.save(file_path)
            custom_popup(
                self,
                lang.t("dialog_titles.success", fallback="Success"),
                lang.t("standard_list. export_success", fallback="Exported to {path}").format(path=file_path),
                "success"
            )
        except Exception as e:
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list. export_failed", fallback="Export failed: {error}").format(error=str(e)),
                "error"
            )

    def import_excel(self):
        """Import data from Excel and update compositions in the database (blocked if read-only)."""
        if self.read_only:
            self._deny()
            return
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [str(cell.value).lower() for cell in ws[1] if cell.value]
            if "code" not in headers:
                custom_popup(
                    self,
                    lang.t("dialog_titles.error", fallback="Error"),
                    lang.t("standard_list.excel_missing_code", fallback="Excel must have a 'Code' column (case-insensitive)."),
                    "error"
                )
                return

            code_idx = headers.index("code")
            scenario_names = {s["name"]. lower() for s in self.scenarios}
            scenario_idxs = {header:  idx for idx, header in enumerate(headers) if header in scenario_names}
            remarks_idx = headers.index("remarks") if "remarks" in headers else None

            excel_codes = set()
            code_quantities = {}
            conn = connect_db()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[code_idx]:
                    continue
                raw_code = str(row[code_idx])
                cleaned_code = validate_code(raw_code) or raw_code.replace(" ", "").upper()
                if cleaned_code. startswith('K'):
                    continue
                excel_codes.add(cleaned_code)

                if cleaned_code not in code_quantities:
                    code_quantities[cleaned_code] = {
                        'remarks': None, 
                        'quantities': {s["name"]: 0 for s in self.scenarios}
                    }

                remarks = str(row[remarks_idx]) if remarks_idx is not None and row[remarks_idx] else None
                if remarks is not None:
                    code_quantities[cleaned_code]['remarks'] = remarks

                for scen_name, idx in scenario_idxs.items():
                    val = row[idx]
                    qty = 0
                    if val: 
                        try:
                            qty = float(val)
                        except ValueError:
                            pass
                    original_scen_name = next(s["name"] for s in self.scenarios if s["name"]. lower() == scen_name)
                    code_quantities[cleaned_code]['quantities'][original_scen_name] = qty

            for cleaned_code, data in code_quantities.items():
                remarks = data['remarks']
                if remarks is not None:
                    cursor.execute("UPDATE compositions SET remarks=?, updated_at=CURRENT_TIMESTAMP WHERE code=?", 
                                   (remarks, cleaned_code))

                for scen_name, qty in data['quantities'].items():
                    scenario = next(s for s in self.scenarios if s["name"] == scen_name)
                    scenario_id = scenario["scenario_id"]
                    scenario_index = next(i for i, s in enumerate(self.scenarios) if s["scenario_id"] == scenario_id)
                    col_letter = scenario_to_column_letter(scenario_index)

                    unique_id_2 = generate_unique_id_2(scenario_id, cleaned_code, qty)
                    if qty > 0:
                        cursor.execute(f"""
                            INSERT INTO compositions (code, scenario_id, quantity, unique_id_2, {col_letter})
                            VALUES (?, ?, ?, ?, ?)
                            ON CONFLICT(code, scenario_id) DO UPDATE SET
                                quantity=excluded.quantity, unique_id_2=excluded.unique_id_2, {col_letter}=excluded.{col_letter},
                                updated_at=CURRENT_TIMESTAMP
                        """, (cleaned_code, scenario_id, qty, unique_id_2, unique_id_2))
                    else:
                        cursor.execute("DELETE FROM compositions WHERE code = ?  AND scenario_id = ?", 
                                       (cleaned_code, scenario_id))

            conn.commit()
            cursor.close()
            conn.close()

            # Check for missing items in items_list
            conn = connect_db()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT code FROM items_list")
            existing_codes = {validate_code(row[0]) or row[0]. replace(" ", "").upper() for row in cursor.fetchall()}
            cursor.close()
            conn.close()

            missing = excel_codes - existing_codes
            if missing:
                custom_popup(
                    self,
                    lang.t("dialog_titles.warning", fallback="Warning"),
                    lang.t("standard_list.missing_items_warning", 
                           fallback="The following codes are missing in the items_list: {codes}.  Data was still updated.").format(
                        codes=', '.join(sorted(missing)[: 10]) + ('...' if len(missing) > 10 else '')
                    ),
                    "warning"
                )

            self.load_data()
            custom_popup(
                self,
                lang.t("dialog_titles.success", fallback="Success"),
                lang.t("standard_list.import_success", fallback="Import completed. "),
                "success"
            )
        except Exception as e: 
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list.import_failed", fallback="Import failed: {error}").format(error=str(e)),
                "error"
            )
        finally:
            if 'wb' in locals():
                wb.close()

    def clear_all(self):
        """Clear all compositions from database (blocked if read-only)."""
        if self.read_only:
            self._deny()
            return
        
        result = custom_askyesno(
            self,
            lang.t("dialog_titles.confirm", fallback="Confirm"),
            lang.t("standard_list.clear_all_confirm", fallback="This will delete all compositions. Continue?")
        )
        
        if result != "yes":
            return
            
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM compositions")
            conn.commit()
            cursor.close()
            conn.close()
            self.load_data()
            custom_popup(
                self,
                lang.t("dialog_titles.success", fallback="Success"),
                lang.t("standard_list.cleared", fallback="All compositions deleted"),
                "success"
            )
        except Exception as e: 
            custom_popup(
                self,
                lang.t("dialog_titles.error", fallback="Error"),
                lang.t("standard_list.clear_failed", fallback="Failed to clear:  {error}").format(error=str(e)),
                "error"
            )